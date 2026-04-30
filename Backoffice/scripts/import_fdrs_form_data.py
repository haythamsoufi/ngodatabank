#!/usr/bin/env python3
"""
Import or sync FDRS-prepared data into the form_data table.

Data can come from:
  - A CSV or Excel file (Ready to import shape), or
  - An FDRS API that returns Ready-to-import JSON, or
  - data-api.ifrc.org + Indicator Bank API (no files needed; set fdrs_import_local_config.py).

With no arguments, the script runs the full pipeline from APIs (no files or flags):
data-api.ifrc.org (fdrsdata, indicator codebook, entities/ns) for FDRS Data + disagg;
databank API (assigned-forms, form-items, indicator-bank) for ready-to-import merge.
Databank API = current app (BASE_URL) unless databank_base_url is overridden.
Same-app: assignments / form items / indicator bank are read in-process from the DB (no HTTP /api/v1/*).
Remote databank: set databank_base_url and FDRS_DATABANK_API_KEY (or DATABANK_API_KEY / MOBILE_APP_API_KEY) for Bearer auth.

Expects records with: assignment_entity_status_id, item_id (or form_item_id),
value, disagg_data, data_not_available, not_applicable, prefilled_value,
imputed_value, submitted_at. Inserts new rows or updates existing ones keyed by
(assignment_entity_status_id, form_item_id).

Usage:
    # From file (e.g. export from Power BI / Excel)
    python scripts/import_fdrs_form_data.py --input path/to/ready_to_import.csv [options]
    python scripts/import_fdrs_form_data.py --input path/to/ready_to_import.xlsx --batch-size 500 --dry-run

    # From FDRS API (API returns JSON array of ready-to-import records)
    python scripts/import_fdrs_form_data.py --fdrs-api-url "https://..." [--fdrs-api-key KEY] [options]

Options:
    --input                  Path to CSV or Excel file (use this OR --fdrs-api-url OR --fdrs-from-data-api)
    --fdrs-api-url           URL that returns Ready-to-import JSON (same shape as file)
    --fdrs-api-key           Optional API key for --fdrs-api-url
    --fdrs-from-data-api     Fetch FDRS from data-api.ifrc.org (fdrsdata, datareport, entities, indicator)
    --fdrs-data-api-base     Base URL for FDRS data API (default: https://data-api.ifrc.org)
    --fdrs-data-api-key      API key for data-api.ifrc.org (required if using --fdrs-from-data-api)
    --fdrs-imputed-url       Optional URL for imputed values (e.g. Power Platform export)
    --fdrs-years             Comma-separated years for FDRS fetch (default: 2010-2024)
    --fdrs-reported-states   Comma-separated IFRC row State codes to treat as importable reported values
                             (0=Not filled, 100=Saved, 200=Reopened, 300=Submitted, 400=Validated, 500=Published).
                             Default when omitted: 100,200,300,400,500 (all except Not filled) or FDRS_REPORTED_IMPORT_STATES env.
    --indicator-mapping      CSV or Excel with fdrs_code, indicator_bank_id, item_id (for --fdrs-from-data-api). Use Excel after editing --export-mapping output.
    --export-mapping FILE   Export current KPI->item_id mapping to Excel and exit; edit then pass back via --indicator-mapping.
    --snapshot-excel FILE   Export raw FDRS snapshot with original KPI rows + import status (default for --fdrs-from-data-api: fdrs_raw_snapshot.xlsx).
    --dry-run                Preview changes without writing to the database
    --batch-size             Commit every N rows (default: 1000)
    --template-id            Restrict form_item_id to items of this template (default: 21)

    Debugging (when using --fdrs-from-data-api):
    - Why was a BaseKPI or KPI_code not imported?
      Use --snapshot-excel to get fdrs_raw_snapshot.xlsx. Each row has original KPI_code, BaseKPI,
      import_status (imported | filtered_out), and import_filter_reason (e.g. no_indicator_bank_match,
      no_form_item_for_template, no_assignment_entity_status_id, main_value_empty_or_zero,
      null_or_empty_value, ends_with_IP). Filter by BaseKPI or KPI_code in Excel to see reasons.
    - Single-row console debug:
      FDRS_DEBUG_ROW=ISO3,YEAR,ITEM_ID  e.g. FDRS_DEBUG_ROW=AFG,2024,917  prints the transformed
      row and disagg for that (iso3, year, form_item_id).
    - Per-KPI summary in console:
      --debug-kpi BASE_KPI  e.g. --debug-kpi KPI_DonBlood  prints counts by import_status and
      import_filter_reason for that BaseKPI (after snapshot is built).
    - What values are in disagg_data (direct + indirect):
      With --preview-excel, the script prints up to 5 sample disagg_data contents (direct keys/values and indirect).
      With FDRS_DEBUG_ROW=ISO3,YEAR,ITEM_ID it prints the full disagg_data (direct + indirect) for that row.
      The preview Excel also has a disagg_data column with the full JSON per row.

---
What's excluded / left out (fdrs-from-data-api; see fdrs_data_fetcher.py for pipeline exclusions)
-------------------------------------------------------------------------------------------------
  - (iso3, year) with no assignment: no assignment_entity_status_id for that period/country → row not built.
  - BaseKPI with no indicator_bank match: fdrs_kpi_code not in indicator bank → row not built.
  - Indicator with no form_item for template: indicator not on template (e.g. 21) → row not built.
  - Rows with same (assignment_entity_status_id, form_item_id): one row per (iso3, year, BaseKPI); later processing could use Tot_M/Tot_F etc. elsewhere if needed.
"""

import argparse
import csv
import json
import logging
import os
import re
import sys

logger = logging.getLogger(__name__)
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, Callable, Dict, List, Optional, Tuple

import urllib.request
import urllib.parse

# Add Backoffice to path
script_dir = os.path.dirname(os.path.abspath(__file__))
backoffice_dir = os.path.dirname(script_dir)
if backoffice_dir not in sys.path:
    sys.path.insert(0, backoffice_dir)
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

if "FLASK_CONFIG" not in os.environ:
    os.environ["FLASK_CONFIG"] = "development"

# Default APIs (no flags or local config needed; override with CLI if desired)
DEFAULT_FDRS_DATA_API_BASE = "https://data-api.ifrc.org"
# Never embed real keys in source; set FDRS_DATA_API_KEY (or pass --fdrs-data-api-key).
DEFAULT_FDRS_DATA_API_KEY = os.environ.get("FDRS_DATA_API_KEY")
# Databank API = current app (BASE_URL). Key: explicit, then env FDRS_DATABANK_API_KEY/DATABANK_API_KEY, then app API_KEY.
DEFAULT_DATABANK_API_KEY = (
    os.environ.get("FDRS_DATABANK_API_KEY") or os.environ.get("DATABANK_API_KEY") or os.environ.get("API_KEY")
)
DEFAULT_INDICATOR_BANK_API_BASE = "https://ifrc-indicatorbank.azurewebsites.net"

# Expected column names (input file may use item_id or form_item_id)
COL_ASSIGNMENT = "assignment_entity_status_id"
COL_PUBLIC = "public_submission_id"
COL_ITEM = "item_id"
COL_ITEM_ALT = "form_item_id"
COL_VALUE = "value"
COL_DISAGG = "disagg_data"
COL_DATA_NA = "data_not_available"
COL_NA = "not_applicable"
COL_PREFILLED = "prefilled_value"
COL_IMPUTED = "imputed_value"
COL_SUBMITTED = "submitted_at"

ALL_COLUMNS = (
    COL_ASSIGNMENT,
    COL_PUBLIC,
    COL_ITEM,
    COL_VALUE,
    COL_DISAGG,
    COL_DATA_NA,
    COL_NA,
    COL_PREFILLED,
    COL_IMPUTED,
    COL_SUBMITTED,
)

# Debug metadata columns for preview Excel only (ignored by row_to_payload / DB)
DEBUG_EXCEL_COLUMNS = ("_debug_year", "_debug_iso3", "_debug_kpi_code")
PREVIEW_EXCEL_COLUMNS = DEBUG_EXCEL_COLUMNS + ALL_COLUMNS

# Columns for KPI mapping Excel (export for manual edit; later load with --indicator-mapping)
MAPPING_EXCEL_COLUMNS = ("fdrs_code", "indicator_bank_id", "item_id", "indicator_name")

# Question KPIs: no indicator bank link; map directly to form_item_id. Single-choice answer: Male, Female, Other.
FDRS_QUESTION_KPI_TO_ITEM = {"KPI_pr_sex": 924, "KPI_sg_sex": 934}

SNAPSHOT_EXCEL_COLUMNS = (
    "year",
    "ISO3",
    "DonCode",
    "KPI_code",
    "BaseKPI",
    "Value",
    "ValueStatus",
    "State",
    "in_fdrs_data_stage",
    "fdrs_data_filter_reason",
    "import_status",
    "import_filter_reason",
    "in_disagg_data",
    "disagg_role",
    "disagg_data_json",
)


def _normalize_headers(row: Dict[str, str]) -> Dict[str, str]:
    """Normalize keys to strip whitespace and match expected names."""
    return {k.strip(): v for k, v in row.items()}


def _parse_submitted_at(raw: Optional[str]) -> Optional[datetime]:
    """Parse submitted_at from Power Query style (e.g. 17/09/2025  14:56:03)."""
    if not raw or not str(raw).strip():
        return None
    raw = str(raw).strip()
    # DD/MM/YYYY HH:MM:SS or DD/MM/YYYY
    for fmt in ("%d/%m/%Y  %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def _parse_json_field(raw: Optional[str]) -> Optional[Any]:
    """Parse a JSON field (disagg_data, prefilled_value, imputed_value)."""
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None
    s = str(raw).strip()
    if s.lower() in ("null", "none", ""):
        return None
    try:
        return json.loads(s)
    except json.JSONDecodeError:
        return None


def _disagg_data_for_db(val: Any) -> Optional[Any]:
    """Return value to store in disagg_data column; use None (DB NULL) when empty or 'null'."""
    if val is None:
        return None
    if isinstance(val, str):
        if not val.strip() or val.strip().lower() == "null":
            return None
        return val  # will be parsed elsewhere; caller typically passes dict from _parse_json_field
    if isinstance(val, dict) and not val:
        return None
    return val


def _parse_bool(raw: Optional[str]) -> Optional[bool]:
    """Parse boolean (data_not_available, not_applicable)."""
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None
    s = str(raw).strip().lower()
    if s in ("1", "true", "yes", "on"):
        return True
    if s in ("0", "false", "no", "off"):
        return False
    return None


_THOUSANDS_GROUPING_RE = re.compile(r"^\s*[-+]?\d{1,3}(,\d{3})+(\.\d+)?\s*$")


def _normalize_numeric_string_for_parse(s: str) -> str:
    """
    Normalize numeric strings for parsing:
    - If it looks like thousands grouping (e.g. 1,234 or 1,234.56), remove commas.
    - If it contains both ',' and '.', assume ',' are thousand separators and remove them.
    Otherwise return as-is (do not guess locale decimals).
    """
    t = (s or "").strip()
    if not t:
        return t
    if _THOUSANDS_GROUPING_RE.match(t) or ("," in t and "." in t):
        return t.replace(",", "")
    return t


def _to_whole_number_int(val: Any) -> Any:
    """
    Convert numeric-ish values to whole-number ints (no decimals).
    Uses ROUND_HALF_UP for non-integers. Leaves non-numeric values unchanged.
    """
    if val is None:
        return None
    # bool is a subclass of int; keep booleans as-is.
    if isinstance(val, bool):
        return val
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        try:
            d = Decimal(str(val))
            if d.is_finite():
                return int(d.to_integral_value(rounding=ROUND_HALF_UP))
        except Exception as e:
            logger.debug("_parse_value_int fallback: %s", e)
            return int(Decimal(val).to_integral_value(rounding=ROUND_HALF_UP))  # best effort
        return val
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return val
        s = _normalize_numeric_string_for_parse(s)
        try:
            d = Decimal(s)
        except InvalidOperation:
            return val
        if not d.is_finite():
            return val
        return int(d.to_integral_value(rounding=ROUND_HALF_UP))
    return val


def _normalize_json_numbers_to_ints(obj: Any) -> Any:
    """Recursively normalize floats/decimal-like strings inside JSON payloads to whole-number ints."""
    if obj is None:
        return None
    if isinstance(obj, (bool, int)):
        return obj
    if isinstance(obj, (float, str)):
        return _to_whole_number_int(obj)
    if isinstance(obj, list):
        return [_normalize_json_numbers_to_ints(x) for x in obj]
    if isinstance(obj, dict):
        return {k: _normalize_json_numbers_to_ints(v) for k, v in obj.items()}
    return obj


def _excel_cell_value(val: Any) -> Any:
    """Prefer int for whole numbers so Excel does not show values like 388.0 in cells."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return val
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        if val != val or val in (float("inf"), float("-inf")):
            return val
        if val == int(val):
            return int(val)
        return val
    if isinstance(val, Decimal):
        if not val.is_finite():
            return val
        try:
            as_int = int(val.to_integral_value(rounding=ROUND_HALF_UP))
            if Decimal(as_int) == val:
                return as_int
        except Exception as e:
            logger.debug("_excel_cell_value decimal: %s", e)
        return float(val)
    if isinstance(val, str):
        s = val.strip()
        if not s or s.lower() in ("null", "none"):
            return ""
        s_norm = _normalize_numeric_string_for_parse(s)
        try:
            d = Decimal(s_norm)
        except InvalidOperation:
            return val
        if not d.is_finite():
            return val
        try:
            as_int = int(d.to_integral_value(rounding=ROUND_HALF_UP))
            if Decimal(as_int) == d:
                return as_int
        except Exception as e:
            logger.debug("_excel_cell_value str decimal: %s", e)
        return val
    return val


def _coerce_value(raw: Any) -> Optional[str]:
    """Coerce value to string for form_data.value (max 255). Numeric values are stored as whole numbers."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.lower() in ("null", "none"):
        return None

    # If it's numeric-ish, force whole number (no decimals).
    s_norm = _normalize_numeric_string_for_parse(s)
    try:
        d = Decimal(s_norm)
        if d.is_finite():
            s = str(int(d.to_integral_value(rounding=ROUND_HALF_UP)))
        else:
            return None
    except InvalidOperation:
        # Non-numeric (e.g. "Male", free text): keep as trimmed string.
        pass

    if len(s) > 255:
        return s[:255]
    return s if s else None


def load_csv(path: str) -> List[Dict[str, str]]:
    """Load rows from CSV; first row = headers."""
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(_normalize_headers(row))
    return rows


def load_excel(path: str) -> List[Dict[str, str]]:
    """Load first sheet from Excel; first row = headers."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("Excel support requires openpyxl: pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb.active
    rows = []
    header = None
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            header = [str(c).strip() if c is not None else "" for c in row]
            continue
        if not header:
            continue
        d = {}
        for j, cell in enumerate(row):
            key = header[j] if j < len(header) else f"Col{j}"
            d[key] = str(cell) if cell is not None else ""
        rows.append(_normalize_headers(d))
    wb.close()
    return rows


def write_rows_to_excel(
    rows: List[Dict[str, str]],
    path: str,
    columns: Optional[Tuple[str, ...]] = None,
    extra_sheets: Optional[List[Tuple[str, List[Dict[str, Any]], Optional[Tuple[str, ...]]]]] = None,
) -> None:
    """Write list of row dicts to Excel; first row = headers. Optionally add extra sheets (name, rows, columns)."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("Excel support requires openpyxl: pip install openpyxl")
    cols = columns or (ALL_COLUMNS if rows else ())
    if rows and not cols:
        cols = tuple(rows[0].keys())
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Ready to import"
    for c, key in enumerate(cols, 1):
        sheet.cell(row=1, column=c, value=key)
    for r, row in enumerate(rows, 2):
        for c, key in enumerate(cols, 1):
            val = row.get(key, "")
            if isinstance(val, (dict, list)):
                val = json.dumps(_normalize_json_numbers_to_ints(val))
            else:
                val = _excel_cell_value(val)
            sheet.cell(row=r, column=c, value=val)
    for sheet_name, extra_rows, extra_cols in extra_sheets or []:
        if not extra_rows and not extra_cols:
            continue
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name max 31 chars
        ec = extra_cols or (tuple(extra_rows[0].keys()) if extra_rows else ())
        for c, key in enumerate(ec, 1):
            ws.cell(row=1, column=c, value=key)
        for r, row in enumerate(extra_rows, 2):
            for c, key in enumerate(ec, 1):
                val = row.get(key, "")
                if isinstance(val, (dict, list)):
                    val = json.dumps(_normalize_json_numbers_to_ints(val))
                else:
                    val = _excel_cell_value(val)
                ws.cell(row=r, column=c, value=val)
    wb.save(path)


def write_mapping_to_excel(rows: List[Dict[str, Any]], path: str) -> None:
    """Write mapping rows to Excel (sheet 'Mapping', columns fdrs_code, indicator_bank_id, item_id, indicator_name)."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("Excel support requires openpyxl: pip install openpyxl")
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Mapping"
    cols = MAPPING_EXCEL_COLUMNS
    for c, key in enumerate(cols, 1):
        sheet.cell(row=1, column=c, value=key)
    for r, row in enumerate(rows, 2):
        for c, key in enumerate(cols, 1):
            val = row.get(key, "")
            if isinstance(val, (dict, list)):
                val = json.dumps(_normalize_json_numbers_to_ints(val))
            else:
                val = _excel_cell_value(val)
            sheet.cell(row=r, column=c, value=val)
    wb.save(path)


def load_input(path: str) -> List[Dict[str, str]]:
    """Load CSV or Excel by extension."""
    path_lower = path.lower()
    if path_lower.endswith(".csv"):
        return load_csv(path)
    if path_lower.endswith((".xlsx", ".xls")):
        return load_excel(path)
    raise ValueError("Input file must be .csv or .xlsx")


def _normalize_api_row(obj: Dict[str, Any]) -> Dict[str, str]:
    """Convert one API response object to the same string-keyed shape as file rows."""
    def _str(v: Any) -> str:
        if v is None:
            return ""
        if isinstance(v, (dict, list)):
            return json.dumps(v)
        return str(v)

    return _normalize_headers({
        COL_ASSIGNMENT: _str(obj.get(COL_ASSIGNMENT) or obj.get("assignment_entity_status_id")),
        COL_PUBLIC: _str(obj.get(COL_PUBLIC) or obj.get("public_submission_id")),
        COL_ITEM: _str(obj.get(COL_ITEM) or obj.get(COL_ITEM_ALT) or obj.get("form_item_id")),
        COL_VALUE: _str(obj.get(COL_VALUE) or obj.get("value")),
        COL_DISAGG: _str(obj.get(COL_DISAGG) or obj.get("disagg_data") or obj.get("disaggregation_data")),
        COL_DATA_NA: _str(obj.get(COL_DATA_NA) or obj.get("data_not_available")),
        COL_NA: _str(obj.get(COL_NA) or obj.get("not_applicable")),
        COL_PREFILLED: _str(obj.get(COL_PREFILLED) or obj.get("prefilled_value")),
        COL_IMPUTED: _str(obj.get(COL_IMPUTED) or obj.get("imputed_value")),
        COL_SUBMITTED: _str(obj.get(COL_SUBMITTED) or obj.get("submitted_at")),
    })


def fetch_fdrs_api(url: str, api_key: Optional[str] = None) -> List[Dict[str, str]]:
    """
    Fetch FDRS data from an API. Expects JSON response to be:
    - A list of objects with Ready-to-import fields, or
    - An object with a key 'data' or 'rows' (or 'form_data') that is such a list.
    Returns list of normalized row dicts (same shape as file rows).
    """
    req = urllib.request.Request(url)
    req.add_header("Accept", "application/json")
    if api_key:
        if api_key.startswith("Bearer ") or api_key.lower().startswith("bearer "):
            req.add_header("Authorization", api_key)
        else:
            req.add_header("Authorization", f"Bearer {api_key}")
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            body = resp.read().decode("utf-8")
    except Exception as e:
        raise RuntimeError(f"FDRS API request failed: {e}") from e

    try:
        data = json.loads(body)
    except json.JSONDecodeError as e:
        raise ValueError(f"FDRS API did not return valid JSON: {e}") from e

    if isinstance(data, list):
        rows = data
    elif isinstance(data, dict):
        for key in ("data", "rows", "form_data", "items", "results"):
            if key in data and isinstance(data[key], list):
                rows = data[key]
                break
        else:
            raise ValueError(
                "FDRS API response object must contain a list under one of: data, rows, form_data, items, results"
            )
    else:
        raise ValueError("FDRS API response must be a JSON array or an object with a list field")

    return [_normalize_api_row(r) for r in rows if isinstance(r, dict)]


def _databank_get(
    base_url: str,
    path: str,
    api_key: Optional[str],
    params: Optional[Dict[str, str]] = None,
    timeout: int = 120,
) -> Any:
    """GET databank API (assigned-forms, form-items, indicator-bank)."""
    url = f"{base_url.rstrip('/')}{path}"
    if params:
        url += "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url)
    req.add_header("Accept", "application/json")
    if api_key:
        req.add_header("Authorization", f"Bearer {api_key}" if not api_key.startswith("Bearer ") else api_key)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read().decode("utf-8"))


def fetch_assignment_id_table(
    databank_base: str,
    api_key: Optional[str],
    template_id: int = 21,
) -> List[Dict[str, Any]]:
    """
    assignment_id query: GET /api/v1/assigned-forms?per_page=10000.
    Expand country_assignments; filter template_id=21.
    Return list of {period_name, iso3, assignment_entity_status_id}.
    """
    out = []
    try:
        data = _databank_get(
            databank_base,
            "/api/v1/assigned-forms",
            api_key,
            params={"per_page": "10000", "template_id": str(template_id)},
        )
        forms = (data or {}).get("assigned_forms") or []
        for af in forms:
            if af.get("template_id") != template_id:
                continue
            period_name = (af.get("period_name") or "").strip()
            for ca in af.get("country_assignments") or []:
                aes_id = ca.get("assignment_entity_status_id")
                iso3 = (ca.get("iso3") or "").strip()
                if aes_id is not None and period_name and iso3:
                    out.append({
                        "period_name": period_name,
                        "iso3": iso3,
                        "assignment_entity_status_id": aes_id,
                    })
    except Exception as e:
        if "run_import" in str(e):
            raise
        logger.warning("fetch_assignment_id_table failed: %s", e)
    return out


def fetch_form_item_table(
    databank_base: str,
    api_key: Optional[str],
    template_id: int = 21,
) -> List[Dict[str, Any]]:
    """
    form_item query: GET /api/v1/form-items?per_page=10000.
    Filter status=published and template_id=21; return {item_id, bank_id}.
    """
    out = []
    try:
        data = _databank_get(
            databank_base,
            "/api/v1/form-items",
            api_key,
            params={"per_page": "10000", "template_id": str(template_id)},
        )
        items = (data or {}).get("form_items") or []
        for it in items:
            if it.get("template_id") != template_id:
                continue
            tv = it.get("template_version") or {}
            if (tv.get("status") or "").lower() != "published":
                continue
            item_id = it.get("id")
            bank_id = it.get("indicator_bank_id")
            if item_id is not None and bank_id is not None:
                out.append({"item_id": item_id, "bank_id": bank_id})
    except Exception as e:
        logger.warning("fetch_form_item_table failed: %s", e)
    return out


def fetch_indicator_bank_table(
    databank_base: str,
    api_key: Optional[str],
    include_archived: bool = False,
) -> List[Dict[str, Any]]:
    """
    Indicator_Bank query: GET /api/v1/indicator-bank?per_page=10000.
    By default requests only non-archived indicators so that when multiple indicators
    share the same fdrs_kpi_code (e.g. KPI_ReachM), the active one is used.
    Return list of {id, fdrs_kpi_code, name}.
    """
    out = []
    try:
        params = {"per_page": "10000"}
        if not include_archived:
            params["archived"] = "false"
        data = _databank_get(
            databank_base,
            "/api/v1/indicator-bank",
            api_key,
            params=params,
        )
        indicators = (data or {}).get("indicators") or []
        for ind in indicators:
            iid = ind.get("id")
            code = ind.get("fdrs_kpi_code")
            if iid is not None and code is not None:
                out.append({
                    "id": iid,
                    "fdrs_kpi_code": (code or "").strip(),
                    "name": (ind.get("name") or "").strip(),
                })
    except Exception as e:
        logger.warning("fetch_indicator_bank_table failed: %s", e)
    return out


def _fetch_databank_tables_local(template_id: int) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Same row shapes as fetch_assignment_id_table / fetch_form_item_table / fetch_indicator_bank_table,
    loaded from the current app's database (no HTTP). Used when FDRS import runs in-process without
    a databank Bearer token (e.g. admin FDRS sync modal — urllib cannot send browser session cookies).
    """
    from sqlalchemy import func

    from app.extensions import db
    from app.models.assignments import AssignedForm, AssignmentEntityStatus
    from app.models.core import Country
    from app.models.form_items import FormItem
    from app.models.forms import FormTemplateVersion
    from app.models.indicator_bank import IndicatorBank

    # country_statuses on AssignedForm is lazy="dynamic" — cannot joinedload; use one SQL join instead.
    assignment_rows: List[Dict[str, Any]] = []
    for period_name, aes_id, iso3_raw in (
        db.session.query(AssignedForm.period_name, AssignmentEntityStatus.id, Country.iso3)
        .join(AssignmentEntityStatus, AssignmentEntityStatus.assigned_form_id == AssignedForm.id)
        .join(Country, Country.id == AssignmentEntityStatus.entity_id)
        .filter(
            AssignedForm.template_id == template_id,
            AssignmentEntityStatus.entity_type == "country",
        )
        .all()
    ):
        pn = (period_name or "").strip()
        iso3 = (iso3_raw or "").strip() if iso3_raw else ""
        if aes_id is not None and pn and iso3:
            assignment_rows.append(
                {
                    "period_name": pn,
                    "iso3": iso3,
                    "assignment_entity_status_id": int(aes_id),
                }
            )

    items = (
        db.session.query(FormItem)
        .join(FormTemplateVersion, FormItem.version_id == FormTemplateVersion.id)
        .filter(FormItem.template_id == template_id)
        .filter(func.lower(FormTemplateVersion.status) == "published")
        .all()
    )
    form_item_rows: List[Dict[str, Any]] = []
    for item in items:
        if item.indicator_bank_id is None:
            continue
        form_item_rows.append({"item_id": item.id, "bank_id": item.indicator_bank_id})

    indicators = IndicatorBank.query.filter_by(archived=False).order_by(IndicatorBank.name.asc()).all()
    indicator_bank_rows: List[Dict[str, Any]] = []
    for ind in indicators:
        code = getattr(ind, "fdrs_kpi_code", None)
        if code is None or not str(code).strip():
            continue
        indicator_bank_rows.append(
            {
                "id": ind.id,
                "fdrs_kpi_code": str(code).strip(),
                "name": (ind.name or "").strip(),
            }
        )

    return assignment_rows, form_item_rows, indicator_bank_rows


def _pick_total_row_for_base(rows: List[Dict[str, Any]], base_kpi: str) -> Dict[str, Any]:
    """
    From rows that share the same (iso3, year, BaseKPI), pick the one to use for the main value field.
    Prefer: base_kpi + '_Tot' or base_kpi; then base_kpi + '_CPD' (CPD is the total when there is no _Tot).
    Tot_M, Tot_F and other sex/age breakdowns are already in disagg_data and must not go into the main value.
    """
    if not rows:
        raise ValueError("empty rows")
    # 1. Aggregate total: base_kpi + "_Tot" or base_kpi
    agg_tot = base_kpi + "_Tot"
    for r in rows:
        kpi = (r.get("KPI_code") or "").strip()
        if kpi == agg_tot or kpi == base_kpi:
            return r
    # 1b. CPD row is the total when there is no _Tot (e.g. KPI_Climate_CPD for base KPI_Climate)
    cpd_row = base_kpi + "_CPD"
    for r in rows:
        kpi = (r.get("KPI_code") or "").strip()
        if kpi == cpd_row:
            return r
    # 2. No aggregate total row: use first row that is not Tot_M/Tot_F, but with value cleared so we don't put sex subtotals in main value
    for r in rows:
        kpi = (r.get("KPI_code") or "").strip()
        if not (kpi.endswith("_Tot_M") or kpi.endswith("_Tot_F")):
            out = dict(r)
            out["Value"] = None  # main value stays empty; disagg_data has the breakdown
            return out
    # 3. Only Tot_M/Tot_F rows: return first with value cleared
    out = dict(rows[0])
    out["Value"] = None
    return out


# Max items to keep per exclusion list (avoid huge payloads)
_EXCLUSION_LIST_CAP = 500


def _normalize_sex_choice(value: Any) -> str:
    """Normalize FDRS value for sex single-choice questions (KPI_pr_sex, KPI_sg_sex) to Male, Female, or Other."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    lower = s.lower()
    if lower in ("male", "m"):
        return "Male"
    if lower in ("female", "f"):
        return "Female"
    if lower in ("other", "unknown", "non-binary", "nonbinary", "x", "prefer not to say"):
        return "Other"
    # Pass through if already one of the three
    if s in ("Male", "Female", "Other"):
        return s
    return "Other"  # fallback


def _data_availability_from_group_rows(group_rows: List[Dict[str, Any]]) -> tuple:
    """
    Check FDRS group_rows for data availability KPIs. Map to databank flags (data_not_available, not_applicable).
    Returns (data_not_available: bool, not_applicable: bool).
    For _IsDataNotAvailable / _isDataNotCollected, null or empty Value is treated as true (row presence = flag set).
    """
    data_not_available = False
    not_applicable = False
    for r in group_rows or []:
        kpi = (r.get("KPI_code") or "").strip()
        val = r.get("Value")
        if kpi.endswith("_IsDataNotAvailable"):
            truthy = val is True or (isinstance(val, (int, float)) and val != 0)
            if not truthy and val is not None and str(val).strip():
                truthy = str(val).strip().lower() in ("1", "true", "yes", "on")
            if not truthy and (val is None or (isinstance(val, str) and not val.strip())):
                truthy = True  # presence of row with no value = treat as set
            if truthy:
                data_not_available = True
        elif kpi.endswith("_isDataNotCollected"):
            truthy = val is True or (isinstance(val, (int, float)) and val != 0)
            if not truthy and val is not None and str(val).strip():
                truthy = str(val).strip().lower() in ("1", "true", "yes", "on")
            if not truthy and (val is None or (isinstance(val, str) and not val.strip())):
                truthy = True
            if truthy:
                not_applicable = True
    return (data_not_available, not_applicable)


def _main_value_empty_or_zero(value: Any) -> bool:
    """True if the main value should be excluded: empty or 0 for numeric KPIs."""
    if value is None:
        return True
    s = str(value).strip()
    if not s:
        return True
    try:
        n = float(s)
        return n == 0
    except (ValueError, TypeError):
        return False  # non-numeric: keep (e.g. text)


def format_exclusion_summary(exclusion_summary: Dict[str, Any]) -> str:
    """Return a human-readable summary of what was excluded in the FDRS pipeline."""
    if not exclusion_summary:
        return ""
    lines = ["--- FDRS exclusion summary (what was left out) ---"]
    codebook = exclusion_summary.get("codebook") or []
    if codebook:
        by_reason = {}
        for item in codebook:
            r = item.get("reason", "?")
            by_reason[r] = by_reason.get(r, 0) + 1
        lines.append(f"  Codebook excluded: {len(codebook)} rows")
        for reason, count in sorted(by_reason.items(), key=lambda x: -x[1]):
            lines.append(f"    - {reason}: {count}")
        sample = [item.get("KPI_Code") for item in codebook[:5]]
        if sample:
            lines.append(f"    Sample KPI_Code: {sample}")
    ip_list = exclusion_summary.get("valid_kpi_excluded_ip") or []
    if ip_list:
        lines.append(f"  Valid KPI (section type IP) excluded: {len(ip_list)} — {ip_list[:10]}{'...' if len(ip_list) > 10 else ''}")
    fdrs = exclusion_summary.get("fdrs_data") or {}
    if fdrs.get("total_excluded"):
        lines.append(f"  FDRS Data excluded: {fdrs['total_excluded']} rows")
        for reason, count in (fdrs.get("count_by_reason") or {}).items():
            lines.append(f"    - {reason}: {count}")
    disagg = exclusion_summary.get("disagg") or {}
    if disagg.get("no_sex_or_age") or disagg.get("age_not_in_mapping"):
        lines.append(f"  Disagg: no_sex_or_age={disagg.get('no_sex_or_age', 0)}, age_not_in_mapping={disagg.get('age_not_in_mapping', 0)}")
        ag = disagg.get("age_groups_not_in_mapping") or []
        if ag:
            lines.append(f"    age_groups_not_in_mapping: {ag}")
    rti = exclusion_summary.get("ready_to_import") or {}
    if (
        rti.get("no_assignment_count")
        or rti.get("no_indicator_count")
        or rti.get("no_form_item_count")
        or rti.get("main_value_empty_or_zero_count")
    ):
        lines.append(
            f"  Ready-to-import skipped: no_assignment={rti.get('no_assignment_count', 0)}, "
            f"no_indicator={rti.get('no_indicator_count', 0)}, no_form_item={rti.get('no_form_item_count', 0)}, "
            f"main_value_empty_or_zero={rti.get('main_value_empty_or_zero_count', 0)}"
        )
        if rti.get("no_assignment"):
            lines.append(f"    Sample no_assignment (iso3,year): {rti['no_assignment'][:5]}")
        if rti.get("no_indicator"):
            lines.append(f"    Sample no_indicator (base_kpi): {rti['no_indicator'][:5]}")
        if rti.get("no_form_item"):
            lines.append(f"    Sample no_form_item (base_kpi): {rti['no_form_item'][:5]}")
    lines.append("---")
    return "\n".join(lines)


def build_ready_to_import_from_new_pipeline(
    fdrs_data: List[Dict[str, Any]],
    disagg_by_key: Dict[Tuple[str, str, str], str],
    assignment_rows: List[Dict[str, Any]],
    form_item_rows: List[Dict[str, Any]],
    indicator_bank_rows: List[Dict[str, Any]],
    submitted_at_default: Optional[datetime] = None,
    debug_iso3_year_item: Optional[Tuple[str, str, int]] = None,
    exclusion_summary: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, str]]:
    """
    Ready to import: merge FDRS Data + disagg + Indicator_Bank (BaseKPI = fdrs_kpi_code)
    + form_item (indicatorId = bank_id) + assignment_id (year = period_name, ISO3 = iso3).
    One row per (iso3, year, BaseKPI) using the Total variant value to avoid duplicate (aes_id, form_item_id).
    """
    # Lookups: BaseKPI -> indicator id; indicator id -> item_id; (period_name, iso3) -> assignment_entity_status_id
    base_to_indicator_id: Dict[str, int] = {}
    for r in indicator_bank_rows:
        code = (r.get("fdrs_kpi_code") or "").strip()
        if code:
            base_to_indicator_id[code] = int(r["id"])
    bank_to_item_id: Dict[int, int] = {}
    for r in form_item_rows:
        bank_to_item_id[int(r["bank_id"])] = int(r["item_id"])
    assignment_by_key: Dict[Tuple[str, str], int] = {}
    for r in assignment_rows:
        assignment_by_key[(r["period_name"], r["iso3"])] = int(r["assignment_entity_status_id"])

    # Group fdrs_data by (iso3, year, base_kpi); only include groups that map to an item and assignment
    from collections import defaultdict
    groups: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    rti_excl = (exclusion_summary or {}).get("ready_to_import") if exclusion_summary else None
    no_assignment_set = set()
    no_indicator_set = set()
    no_form_item_set = set()
    main_value_empty_or_zero_set = set()

    for r in fdrs_data:
        iso3 = (r.get("ISO3") or "").strip()
        year = (r.get("year") or "").strip()
        base_kpi = (r.get("BaseKPI") or "").strip()
        # Question KPIs: direct form_item link (no indicator bank)
        if base_kpi in FDRS_QUESTION_KPI_TO_ITEM:
            if assignment_by_key.get((year, iso3)) is None:
                if rti_excl is not None:
                    no_assignment_set.add((iso3, year))
                continue
            groups[(iso3, year, base_kpi)].append(r)
            continue
        indicator_id = base_to_indicator_id.get(base_kpi)
        if indicator_id is None:
            if rti_excl is not None:
                no_indicator_set.add(base_kpi)
            continue
        item_id = bank_to_item_id.get(indicator_id)
        if item_id is None:
            if rti_excl is not None:
                no_form_item_set.add(base_kpi)
            continue
        if assignment_by_key.get((year, iso3)) is None:
            if rti_excl is not None:
                no_assignment_set.add((iso3, year))
            continue
        groups[(iso3, year, base_kpi)].append(r)

    if rti_excl is not None and isinstance(rti_excl, dict):
        rti_excl["no_assignment"] = sorted([f"{iso3},{yr}" for (iso3, yr) in no_assignment_set])[:_EXCLUSION_LIST_CAP]
        rti_excl["no_indicator"] = sorted(no_indicator_set)[:_EXCLUSION_LIST_CAP]
        rti_excl["no_form_item"] = sorted(no_form_item_set)[:_EXCLUSION_LIST_CAP]
        rti_excl["main_value_empty_or_zero"] = []
        rti_excl["no_assignment_count"] = len(no_assignment_set)
        rti_excl["no_indicator_count"] = len(no_indicator_set)
        rti_excl["no_form_item_count"] = len(no_form_item_set)
        rti_excl["main_value_empty_or_zero_count"] = 0

    out = []
    if debug_iso3_year_item:
        debug_iso3, debug_year, _ = debug_iso3_year_item
        afg_item_ids = sorted(
            set(
                FDRS_QUESTION_KPI_TO_ITEM.get(base_kpi) or bank_to_item_id.get(base_to_indicator_id.get(base_kpi))
                for (iso3, year, base_kpi) in groups
                if iso3 == debug_iso3 and str(year) == str(debug_year)
            )
        )
        afg_item_ids = [x for x in afg_item_ids if x is not None]
        logger.debug("For %s %s available item_ids: %s", debug_iso3, debug_year, afg_item_ids)
    for (iso3, year, base_kpi), group_rows in groups.items():
        is_question_kpi = base_kpi in FDRS_QUESTION_KPI_TO_ITEM
        if is_question_kpi:
            item_id = FDRS_QUESTION_KPI_TO_ITEM[base_kpi]
            # Pick row with KPI_code == base_kpi (single-choice question: one value per iso3/year)
            r = next((x for x in group_rows if (x.get("KPI_code") or "").strip() == base_kpi), group_rows[0])
            value = _normalize_sex_choice(r.get("Value"))
            if not value:
                if rti_excl is not None and isinstance(rti_excl, dict):
                    rti_excl["main_value_empty_or_zero_count"] = rti_excl.get("main_value_empty_or_zero_count", 0) + 1
                    main_value_empty_or_zero_set.add((iso3, year, base_kpi))
                continue
            disagg_data = ""
            data_not_available_flag = False
            not_applicable_flag = False
        else:
            indicator_id = base_to_indicator_id[base_kpi]
            item_id = bank_to_item_id[indicator_id]
            data_not_available_flag, not_applicable_flag = _data_availability_from_group_rows(group_rows)
            r = _pick_total_row_for_base(group_rows, base_kpi)
            value = r.get("Value")
            # Emit row when value present OR when data_not_available/not_applicable set (FDRS data availability KPIs)
            if not (data_not_available_flag or not_applicable_flag) and _main_value_empty_or_zero(value):
                if rti_excl is not None and isinstance(rti_excl, dict):
                    rti_excl["main_value_empty_or_zero_count"] = rti_excl.get("main_value_empty_or_zero_count", 0) + 1
                    main_value_empty_or_zero_set.add((iso3, year, base_kpi))
                continue
            if data_not_available_flag or not_applicable_flag:
                value = ""  # UI shows checkbox only; no value
            else:
                value = str(value) if value is not None else ""
            disagg_data = disagg_by_key.get((iso3, year, base_kpi)) or ""
        aes_id = assignment_by_key[(year, iso3)]
        sub_at = submitted_at_default.strftime("%d/%m/%Y  %H:%M:%S") if submitted_at_default else ""
        kpi_code = (r.get("KPI_code") or base_kpi or "").strip()

        if debug_iso3_year_item:
            debug_iso3, debug_year, debug_item = debug_iso3_year_item
            match = (iso3 == debug_iso3 and str(year) == str(debug_year) and item_id == debug_item)
            if match:
                logger.info(
                    "[DEBUG] Match %s %s item_id=%s: aes_id=%s base_kpi=%r total_row KPI_code=%r value=%r group_size=%d disagg_len=%d",
                    debug_iso3, debug_year, debug_item, aes_id, base_kpi, kpi_code, value, len(group_rows), len(disagg_data)
                )
                if disagg_data:
                    try:
                        d = json.loads(disagg_data)
                        values = d.get("values") or {}
                        direct = values.get("direct") or {}
                        indirect = values.get("indirect")
                        logger.info("[DEBUG] disagg_data.values.direct:")
                        for k in sorted(direct.keys()):
                            logger.info("  %s: %s", k, direct[k])
                        if indirect is not None:
                            logger.info("[DEBUG] disagg_data.values.indirect: %s", indirect)
                    except Exception as e:
                        logger.debug("disagg_data parse failed: %s; raw: %s...", e, disagg_data[:200])
                for i, gr in enumerate(group_rows[:20]):
                    logger.info("  [%d] KPI_code=%r Value=%r", i, gr.get('KPI_code'), gr.get('Value'))
                if len(group_rows) > 20:
                    logger.info("  ... and %d more rows", len(group_rows) - 20)

        out.append(_normalize_headers({
            "_debug_year": year,
            "_debug_iso3": iso3,
            "_debug_kpi_code": kpi_code,
            COL_ASSIGNMENT: str(aes_id),
            COL_PUBLIC: "",
            COL_ITEM: str(item_id),
            COL_VALUE: (str(value) if value is not None else "") or "",
            COL_DISAGG: disagg_data,
            COL_DATA_NA: "true" if data_not_available_flag else "",
            COL_NA: "true" if not_applicable_flag else "",
            COL_PREFILLED: "",
            COL_IMPUTED: "",
            COL_SUBMITTED: sub_at,
        }))
    if rti_excl is not None and isinstance(rti_excl, dict):
        rti_excl["main_value_empty_or_zero"] = [
            f"{iso3},{yr},{base}"
            for (iso3, yr, base) in sorted(main_value_empty_or_zero_set)
        ][: _EXCLUSION_LIST_CAP]
    return out


def _load_indicator_mapping(csv_path: str) -> Dict[str, int]:
    """Load fdrs_code -> indicator_bank_id from CSV (columns: fdrs_code, indicator_bank_id)."""
    mapping = {}
    with open(csv_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            row = {k.strip(): v for k, v in row.items()}
            code = (row.get("fdrs_code") or row.get("FDRS") or row.get("kpi_code") or "").strip()
            if not code:
                continue
            try:
                mapping[code] = int(row.get("indicator_bank_id") or row.get("indicatorId") or 0)
            except (ValueError, TypeError):
                continue
    return mapping


def _load_indicator_mapping_from_excel(excel_path: str) -> Dict[str, int]:
    """
    Load fdrs_code -> item_id from the mapping Excel (sheet 'Mapping', columns fdrs_code, item_id).
    Use this when --indicator-mapping points to an Excel file (e.g. after manual edits).
    """
    try:
        import openpyxl
    except ImportError:
        sys.exit("Excel support requires openpyxl: pip install openpyxl")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet = wb["Mapping"] if "Mapping" in wb.sheetnames else wb.active
    code_to_item_id: Dict[str, int] = {}
    headers = []
    for row in sheet.iter_rows(values_only=True):
        if not headers:
            headers = [str(c).strip() if c is not None else "" for c in (row or [])]
            continue
        if not row:
            continue
        row = list(row) if row else []
        row_dict = dict(zip(headers, row + [None] * (len(headers) - len(row))))
        code = (row_dict.get("fdrs_code") or row_dict.get("FDRS") or row_dict.get("kpi_code") or "").strip()
        if not code:
            continue
        raw_item = row_dict.get("item_id") or row_dict.get("form_item_id") or ""
        try:
            item_id = int(raw_item) if raw_item not in (None, "") else None
        except (ValueError, TypeError):
            continue
        if item_id is not None:
            code_to_item_id[code] = item_id
    wb.close()
    return code_to_item_id


def _resolve_api_mapping_to_db_ids(
    mapping: Dict[str, int],
    form_item_by_bank: Dict[int, int],
    indicator_bank_api_base: str,
    indicator_bank_api_key: Optional[str],
) -> Dict[str, int]:
    """
    Resolve FDRS->API_id mapping to FDRS->our indicator_bank.id by matching API indicator
    name/title to our IndicatorBank.name. The external API uses its own id space; form items
    use our DB indicator_bank.id. Returns mapping that only contains our ids present in form_item_by_bank.
    """
    from fdrs_data_fetcher import fetch_indicator_bank_api
    from app.extensions import db
    from app.models.indicator_bank import IndicatorBank

    api_rows = fetch_indicator_bank_api(
        base_url=indicator_bank_api_base,
        api_key=indicator_bank_api_key,
    )
    api_id_to_name: Dict[int, str] = {}
    for r in api_rows:
        try:
            api_id = int(r.get("id"))
        except (TypeError, ValueError):
            continue
        name = (r.get("title") or r.get("name") or "").strip()
        if name:
            api_id_to_name[api_id] = name

    our_bank_ids = list(form_item_by_bank.keys())
    if not our_bank_ids:
        return {}
    our_id_to_name: Dict[int, str] = dict(
        db.session.query(IndicatorBank.id, IndicatorBank.name).filter(
            IndicatorBank.id.in_(our_bank_ids),
            IndicatorBank.name.isnot(None),
        ).all()
    )
    name_to_our_id: Dict[str, int] = {}
    for our_id, name in our_id_to_name.items():
        if name:
            name_to_our_id[name] = our_id  # first occurrence wins if duplicates

    api_id_to_our_id: Dict[int, int] = {}
    for api_id, api_name in api_id_to_name.items():
        our_id = name_to_our_id.get(api_name)
        if our_id is not None:
            api_id_to_our_id[api_id] = our_id

    resolved: Dict[str, int] = {}
    for fdrs_code, api_id in mapping.items():
        our_id = api_id_to_our_id.get(api_id)
        if our_id is not None and our_id in form_item_by_bank:
            resolved[fdrs_code] = our_id
    return resolved


# First-segment bases that match too many FDRS KPIs; never use as fallback.
_FDRS_GENERIC_BASE = frozenset({"KPI"})

# Only strip these suffixes when they are the *trailing* part of the code (exact match to mapping keys).
_FDRS_STRIP_SUFFIXES = ("_Tot", "_I", "_CPD", "_M", "_F")


def _disagg_base(kpi: str) -> str:
    """Strip trailing _Tot, _CPD, _I, _M, _F to get base for disagg lookup (e.g. KPI_DonBlood_Tot -> KPI_DonBlood)."""
    if not kpi:
        return kpi
    s = kpi.strip()
    while True:
        reduced = False
        for suffix in _FDRS_STRIP_SUFFIXES:
            if s.endswith(suffix):
                s = s[: -len(suffix)].rstrip("_")
                reduced = True
                break
        if not reduced:
            break
    return s


def _fdrs_lookup_code(kpi_code: str, mapping: Dict[str, int]) -> Optional[int]:
    """
    Find indicator_bank_id for this KPI. Match only when:
    1. Full KPI code is in mapping, or
    2. KPI ends with a known suffix and the rest is in mapping (e.g. KPI_Climate_CPD -> KPI_Climate), or
    3. KPI is exactly a single-segment base that is in mapping (e.g. 'ar' exact), and base is not generic.
    Do NOT match by base when KPI has more segments (e.g. 'ar_xxx' must not match mapping['ar']).
    """
    if not kpi_code or not mapping:
        return None
    kpi = kpi_code.strip()
    if kpi in mapping:
        return mapping[kpi]
    for suffix in _FDRS_STRIP_SUFFIXES:
        if kpi.endswith(suffix):
            candidate = kpi[: -len(suffix)].rstrip("_")
            if candidate and candidate in mapping:
                return mapping[candidate]
    # Base fallback only when KPI has no underscore (single segment) so we don't match ar_xxx to 'ar'
    if "_" not in kpi:
        if kpi in mapping and kpi not in _FDRS_GENERIC_BASE:
            return mapping[kpi]
    return None


def _fdrs_lookup_item_id(kpi_code: str, code_to_item_id: Dict[str, int]) -> Optional[int]:
    """Same key logic as _fdrs_lookup_code but for fdrs_code -> item_id mapping (e.g. from Excel)."""
    return _fdrs_lookup_code(kpi_code, code_to_item_id)


def _row_contributes_to_disagg_direct(kpi: str, base: str) -> bool:
    """True if this KPI row contributes to disagg_data.values.direct (sex/age breakdown)."""
    if not kpi or not base or kpi == base + "_I":
        return False
    # Has age segment and sex (M/F) in code - same pattern as fetcher's build_disagg
    return "age" in kpi and ("M" in kpi or "F" in kpi)


def build_fdrs_snapshot_export_rows(
    fdrs_snapshot_rows: List[Dict[str, Any]],
    ready_rows: List[Dict[str, str]],
    exclusion_summary: Optional[Dict[str, Any]] = None,
    disagg_by_key: Optional[Dict[Tuple[str, str, str], str]] = None,
) -> List[Dict[str, Any]]:
    """
    Build row-level snapshot export:
    - One row per raw FDRS combined record
    - Keeps original KPI_code/BaseKPI/Value
    - Adds import_status = imported | filtered_out and import_filter_reason
    - Adds in_disagg_data (yes/no) and disagg_role (direct/indirect) when disagg_by_key is provided
    """
    imported_base_keys = set()
    for r in ready_rows or []:
        iso3 = (r.get("_debug_iso3") or "").strip()
        year = str(r.get("_debug_year") or "").strip()
        kpi = (r.get("_debug_kpi_code") or "").strip()
        base = _disagg_base(kpi)
        if iso3 and year and base:
            imported_base_keys.add((iso3, year, base))

    rti = (exclusion_summary or {}).get("ready_to_import") if exclusion_summary else {}
    no_assignment = set(rti.get("no_assignment") or [])
    no_indicator = set(rti.get("no_indicator") or [])
    no_form_item = set(rti.get("no_form_item") or [])
    main_value_empty_or_zero = set(rti.get("main_value_empty_or_zero") or [])

    out: List[Dict[str, Any]] = []
    for r in fdrs_snapshot_rows or []:
        iso3 = (r.get("ISO3") or "").strip()
        year = str(r.get("year") or "").strip()
        base = (r.get("BaseKPI") or "").strip()
        key_base = (iso3, year, base)
        key_iso_year = f"{iso3},{year}"
        key_iso_year_base = f"{iso3},{year},{base}"

        if key_base in imported_base_keys:
            status = "imported"
            reason = ""
        elif (r.get("in_fdrs_data_stage") or "").strip().lower() != "yes":
            status = "filtered_out"
            reason = (r.get("fdrs_data_filter_reason") or "").strip() or "filtered_in_fdrs_data_stage"
        elif key_iso_year in no_assignment:
            status = "filtered_out"
            reason = "no_assignment_entity_status_id"
        elif base in no_indicator:
            status = "filtered_out"
            reason = "no_indicator_bank_match"
        elif base in no_form_item:
            status = "filtered_out"
            reason = "no_form_item_for_template"
        elif key_iso_year_base in main_value_empty_or_zero:
            status = "filtered_out"
            reason = "main_value_empty_or_zero"
        else:
            # Included in FDRS Data stage but not inserted as a standalone main row.
            # Typically variant rows that are only represented inside disagg_data.
            status = "filtered_out"
            reason = "not_selected_as_main_import_row"

        # Whether this row's value made it into disagg_data (direct or indirect)
        in_disagg = "no"
        disagg_role = ""
        disagg_data_json = ""
        if disagg_by_key:
            disagg_json = (disagg_by_key.get((iso3, year, base)) or "").strip()
            if disagg_json:
                disagg_data_json = disagg_json
                kpi_code = (r.get("KPI_code") or "").strip()
                if kpi_code == base + "_I":
                    in_disagg = "yes"
                    disagg_role = "indirect"
                elif _row_contributes_to_disagg_direct(kpi_code, base):
                    in_disagg = "yes"
                    disagg_role = "direct"

        out.append({
            "year": year,
            "ISO3": iso3,
            "DonCode": (r.get("DonCode") or "").strip(),
            "KPI_code": (r.get("KPI_code") or "").strip(),
            "BaseKPI": base,
            "Value": r.get("Value"),
            "ValueStatus": r.get("ValueStatus"),
            "State": r.get("State"),
            "in_fdrs_data_stage": r.get("in_fdrs_data_stage") or "",
            "fdrs_data_filter_reason": r.get("fdrs_data_filter_reason") or "",
            "import_status": status,
            "import_filter_reason": reason,
            "in_disagg_data": in_disagg,
            "disagg_role": disagg_role,
            "disagg_data_json": disagg_data_json,
        })
    return out


def print_debug_kpi_summary(
    snapshot_rows: List[Dict[str, Any]],
    base_kpi: str,
    max_sample_codes: int = 15,
) -> None:
    """
    Print to stdout why a BaseKPI was or wasn't imported: counts by import_status and
    import_filter_reason, plus sample KPI_codes. Use with --debug-kpi BASE_KPI.
    """
    base_kpi = (base_kpi or "").strip()
    if not base_kpi:
        logger.info("[DEBUG-KPI] No BaseKPI given.")
        return
    subset = [r for r in (snapshot_rows or []) if (r.get("BaseKPI") or "").strip() == base_kpi]
    if not subset:
        logger.info("[DEBUG-KPI] BaseKPI %r: no rows in snapshot (not in FDRS data or wrong spelling).", base_kpi)
        return
    by_status: Dict[str, int] = {}
    by_reason: Dict[str, int] = {}
    kpi_codes_seen: set = set()
    for r in subset:
        st = (r.get("import_status") or "").strip() or "?"
        by_status[st] = by_status.get(st, 0) + 1
        reason = (r.get("import_filter_reason") or "").strip() or "(imported)"
        if reason != "(imported)":
            by_reason[reason] = by_reason.get(reason, 0) + 1
        kpi_codes_seen.add((r.get("KPI_code") or "").strip())
    logger.info("[DEBUG-KPI] BaseKPI %r: %d rows", base_kpi, len(subset))
    logger.info("  import_status: %s", dict(by_status))
    if by_reason:
        logger.info("  import_filter_reason (filtered_out): %s", dict(by_reason))
    sample = sorted(kpi_codes_seen)[:max_sample_codes]
    logger.info("  sample KPI_codes: %s", sample)


def _normalize_age_part(s: str) -> str:
    """Normalize age segment to match form; FDRS 'Other' age -> 'unknown' (system uses Unknown)."""
    if not s:
        return s
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9_]", "_", s)
    if s == "80":
        return "80_"
    if s == "other":
        return "unknown"  # FDRS Other age -> system unknown
    return s


def _direct_key_for_form(key: str) -> str:
    """
    Normalize values.direct key to match entry_form / form_data_service.
    Form uses: sex_slug = male, female, non_binary, unknown (system uses Unknown, not Other).
    sex_age key = sex_slug + '_' + age_slug -> male_5_17, female_unknown, unknown_unknown.
    """
    if not key:
        return key
    # Sex-only: match form slug; FDRS "Other" -> system "unknown"
    if key in ("Male", "Female", "Unknown"):
        return key.lower()
    if key == "Other":
        return "unknown"
    if key in ("Non-binary", "NonBinary"):
        return "non_binary"
    # Sex_age: normalize sex prefix + age slug (Male_5_17 -> male_5_17, Female_Unknown -> female_unknown)
    for sex, slug in (
        ("Male_", "male_"),
        ("Female_", "female_"),
        ("Unknown_", "unknown_"),
        ("Other_", "unknown_"),  # FDRS Other -> system unknown
        ("Non_binary_", "non_binary_"),
        ("Non-binary_", "non_binary_"),
        ("NonBinary_", "non_binary_"),
    ):
        if key.startswith(sex):
            age_part = _normalize_age_part(key[len(sex):])
            return slug + age_part if age_part else key
    return key


def _cpd_variant_to_direct_label(variant_kpi: str, base_prefix: str) -> str:
    """
    Map a CPD variant KPI suffix to a values.direct label (sex/age/sex_age only).
    Keys match entry_form / form_data_service: lowercase sex (male, female), sex_age (male_30_39).
    No 'variants' key is used; only direct and indirect.
    Returns "" to skip this variant (do not add to direct or variants).
    """
    if not variant_kpi.startswith(base_prefix):
        return ""
    suffix = variant_kpi[len(base_prefix):].strip("_")
    if not suffix:
        return ""

    # D_M_age_XXX / D_F_age_XXX / D_NonBinary_age_XXX / D_UnknownSex_age_XXX / D_OtherSex_age_XXX -> sex_age (match form)
    if suffix.startswith("D_M_age_"):
        age_part = _normalize_age_part(suffix[8:].strip("_"))
        if age_part:
            return f"male_{age_part}"
    if suffix.startswith("D_F_age_"):
        age_part = _normalize_age_part(suffix[8:].strip("_"))
        if age_part:
            return f"female_{age_part}"
    if suffix.startswith("D_NonBinary_age_"):
        age_part = _normalize_age_part(suffix[16:].strip("_"))
        if age_part:
            return f"non_binary_{age_part}"
    if suffix.startswith("D_UnknownSex_age_"):
        age_part = _normalize_age_part(suffix[17:].strip("_"))
        if age_part:
            return f"unknown_{age_part}"
    if suffix.startswith("D_OtherSex_age_"):
        age_part = _normalize_age_part(suffix[14:].strip("_"))
        if age_part:
            return f"unknown_{age_part}"  # FDRS OtherSex -> system unknown
    # D_Tot_age_XXX / Tot_age_XXX -> age only (e.g. 13_17, 30_39)
    if suffix.startswith("D_Tot_age_"):
        age_part = _normalize_age_part(suffix[10:].strip("_"))
        if age_part:
            return age_part
    if suffix.startswith("Tot_age_"):
        age_part = _normalize_age_part(suffix[8:].strip("_"))
        if age_part:
            return age_part
    # Tot_0_5, Tot_13_17 etc. (age without "age_" in suffix)
    if suffix.startswith("Tot_") and suffix not in ("Tot_M", "Tot_F"):
        age_part = _normalize_age_part(suffix[4:].strip("_"))
        if age_part:
            return age_part if age_part != "Other" else "unknown"

    # Sex-only (D_Tot_M, Tot_M, M, etc.) -> male, female, non_binary, unknown, other (match form)
    if suffix in ("D_Tot_M", "Tot_M", "M"):
        return "male"
    if suffix in ("D_Tot_F", "Tot_F", "F"):
        return "female"
    if suffix.endswith("_M") and "age" not in suffix:
        return "male"
    if suffix.endswith("_F") and "age" not in suffix:
        return "female"
    if "UnknownSex" in suffix:
        return "unknown"
    if "OtherSex" in suffix:
        return "unknown"  # FDRS OtherSex -> system unknown
    if "NonBinary" in suffix:
        return "non_binary"

    # Standalone age keys: use _normalize_age_part so Unknown -> unknown, 50+ -> 50_, etc. (match form)
    age_norm = _normalize_age_part(suffix)
    _age_keys = ("0_5", "6_12", "13_17", "18_29", "0_29", "30_39", "40_49", "50_59", "60_69", "70_79", "80_", "unknown")
    if age_norm in _age_keys:
        return age_norm
    if suffix in ("age_Other", "Other"):
        return "unknown"  # FDRS Other age -> system unknown
    if suffix in ("Unknown", "age_Unknown"):
        return "unknown"
    if age_norm:
        return age_norm
    return ""


def _fdrs_table_to_ready_to_import(
    fdrs_rows: List[Dict[str, Any]],
    code_to_item_id: Dict[str, int],
    assignment_by_key: Dict[Tuple[Any, str], int],
    template_id: int = 21,
    submitted_at_default: Optional[datetime] = None,
) -> Tuple[List[Dict[str, str]], List[Dict[str, Any]]]:
    """
    Convert merged FDRS table to ready-to-import rows (filter, mode, disagg, join).
    Returns (list of string-keyed dicts compatible with row_to_payload, dropped_rows for preview).

    code_to_item_id: fdrs_code -> form_item_id (from CSV/API mapping+form_item_by_bank, or from Excel mapping).
    Logic mirrors the Power Query pipeline:
    - FDRS Data: filter (published_num_value / Male|Female), exclude _Public/_RollAvg/_D_Tot, finance;
      mode = Indirect | Total | age | sex | sex_age from Sex/AgeGroup and KPI_code.
    - disagg: non-Total/non-Indirect → group by (iso_3, year, KPI_code), build values.direct (attr keys
      normalized: hyphen→underscore, 80→80_); values.indirect from Indirect rows.
    - Ready to import: Total rows only; join disagg + indirect; map KPI_code->item_id;
      (year, iso_3)->assignment_entity_status_id. CPD compound rows get variant values in values.direct only (no variants key).
    """
    # Filter: meaningful value; exclude _Public, _RollAvg, _D_Tot; finance filter
    def keep_row(r: Dict[str, Any]) -> bool:
        kpi = r.get("KPI_code") or ""
        if kpi.endswith("_Public") or kpi.endswith("_RollAvg") or kpi.endswith("_D_Tot"):
            return False
        section = r.get("KPI_Section") or ""
        if section == "NS Finance & Partnerships" and not kpi.endswith("LC_CHF"):
            return False
        pub_num = r.get("published_num_value")
        pub_val = r.get("published_value")
        if pub_num is not None and pub_num != "" and pub_num != 0:
            return True
        if str(pub_val).strip() in ("Male", "Female"):
            return True
        return False

    filtered = [r for r in fdrs_rows if keep_row(r)]
    # Debug: filter stage
    try:
        _years = set()
        _isos = set()
        for r in filtered[:500]:
            y = r.get("year")
            _years.add(str(y) if y is not None else "")
            _isos.add((r.get("iso_3") or "").strip())
        logger.info("[DEBUG] After keep_row: filtered=%d (from %d); sample years=%s, sample iso3=%s", len(filtered), len(fdrs_rows), sorted(_years)[:10], sorted(_isos)[:10])
    except Exception as e:
        logger.debug("keep_row iteration exception: %s; filtered=%d (from %d)", e, len(filtered), len(fdrs_rows))

    # Add mode: Total, Indirect, age, sex, sex_age
    for r in filtered:
        kpi = r.get("KPI_code") or ""
        sex = (r.get("Sex") or "").strip()
        age = (r.get("AgeGroup") or "").strip()
        if kpi.endswith("_I"):
            r["_mode"] = "Indirect"
        elif not sex and not age:
            r["_mode"] = "Total"
        elif not sex and age:
            r["_mode"] = "age"
        elif sex and not age:
            r["_mode"] = "sex"
        else:
            r["_mode"] = "sex_age"

    def _year_key(y: Any) -> str:
        if y is None:
            return ""
        if hasattr(y, "year"):
            return str(getattr(y, "year", y))
        return str(y)

    # Disagg: group non-Total non-Indirect by (iso_3, year, KPI_code), build values.
    # Match Power Query: attr keys use underscore (0_5, 6_12, 80_ for 80+).
    def _normalize_disagg_attr(s: str) -> str:
        if not s:
            return ""
        s = s.strip().replace("-", "_")
        if s == "80":
            s = "80_"
        return s

    disagg_by_key: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    for r in filtered:
        if r.get("_mode") in ("Total", "Indirect"):
            continue
        key = ((r.get("iso_3") or "").strip(), _year_key(r.get("year")), (r.get("KPI_code") or "").strip())
        if key not in disagg_by_key:
            disagg_by_key[key] = {"mode": r["_mode"], "values": {"direct": {}, "indirect": None}}
        raw_attr = (r.get("AgeGroup") or r.get("Sex") or "").strip()
        if r.get("_mode") == "sex_age" and r.get("Sex") and r.get("AgeGroup"):
            raw_attr = f"{r.get('Sex')}_{r.get('AgeGroup')}"
        attr = _normalize_disagg_attr(raw_attr)
        val = r.get("published_value") or r.get("value")
        val = _to_whole_number_int(val)
        if attr:
            # Use form-compatible key (lowercase sex: male, female; sex_age: male_13_17)
            direct_key = _direct_key_for_form(attr)
            disagg_by_key[key]["values"]["direct"][direct_key] = val

    # Indirect by (iso_3, year, KPI_code)
    indirect_by_key: Dict[Tuple[str, str, str], Any] = {}
    for r in filtered:
        if r.get("_mode") != "Indirect":
            continue
        val = r.get("published_value") or r.get("value")
        val = _to_whole_number_int(val)
        key = ((r.get("iso_3") or "").strip(), _year_key(r.get("year")), (r.get("KPI_code") or "").strip())
        indirect_by_key[key] = val

    # Value by (iso_3, year, kpi) for all filtered rows (for CPD variant lookup)
    value_by_key: Dict[Tuple[str, str, str], Any] = {}
    kpis_by_iso_year: Dict[Tuple[str, str], set] = {}
    for r in filtered:
        iso = (r.get("iso_3") or "").strip()
        yr = _year_key(r.get("year"))
        kpi = (r.get("KPI_code") or "").strip()
        key = (iso, yr, kpi)
        val = r.get("published_value") or r.get("value")
        val = _to_whole_number_int(val)
        if val is not None:
            value_by_key[key] = val
        kpis_by_iso_year.setdefault((iso, yr), set()).add(kpi)

    # Total rows only; attach disagg and indirect
    total_rows = [r for r in filtered if r.get("_mode") == "Total"]
    # Debug: drop reasons in the loop; collect unmatched for preview Excel second sheet
    drop_no_bank_id = 0
    drop_no_aes_id = 0
    out = []
    dropped_rows: List[Dict[str, Any]] = []
    for r in total_rows:
        iso_3 = (r.get("iso_3") or "").strip()
        year = _year_key(r.get("year"))
        kpi = (r.get("KPI_code") or "").strip()
        raw_value = r.get("published_value") or r.get("value")
        value_str = str(raw_value) if raw_value is not None else ""
        item_id = _fdrs_lookup_item_id(kpi, code_to_item_id)
        if item_id is None:
            drop_no_bank_id += 1
            dropped_rows.append({"year": year, "iso_3": iso_3, "kpi_code": kpi, "value": value_str, "drop_reason": "no mapping (KPI not in mapping or no item_id)"})
            continue
        aes_id = assignment_by_key.get((year, iso_3))
        if aes_id is None:
            drop_no_aes_id += 1
            dropped_rows.append({"year": year, "iso_3": iso_3, "kpi_code": kpi, "value": value_str, "drop_reason": "no assignment_entity_status_id (year/iso3 not assigned)"})
            continue
        # Disagg base: strip trailing _Tot, _CPD, _I, _M, _F so KPI_DonBlood_Tot -> KPI_DonBlood for lookup
        base = _disagg_base(kpi)
        prefix = base + "_" if base else ""
        dkey = (iso_3, year, kpi)
        disagg_rec = disagg_by_key.get(dkey) or (disagg_by_key.get((iso_3, year, base)) if base != kpi else None)
        indirect = indirect_by_key.get(dkey)
        if indirect is None and base != kpi:
            indirect = indirect_by_key.get((iso_3, year, base))
        values: Dict[str, Any] = {}
        if disagg_rec:
            mode = disagg_rec.get("mode") or "total"
            values = dict(disagg_rec.get("values") or {})
            if indirect is not None:
                values["indirect"] = indirect
        else:
            mode = "total"
        if "direct" not in values:
            values["direct"] = {}
        direct = values["direct"]
        # Merge disagg from all variant KPIs with same base (KPI_DonBlood_Tot_M, KPI_Climate_CPD variants, etc.)
        if prefix:
            # (1) From disagg_by_key: age/sex rows keyed by variant code
            for (dk_iso, dk_yr, dk_kpi), rec in disagg_by_key.items():
                if (dk_iso, dk_yr) != (iso_3, year):
                    continue
                if dk_kpi != kpi and dk_kpi != base and not dk_kpi.startswith(prefix):
                    continue
                for sub_key, sub_val in ((rec.get("values") or {}).get("direct") or {}).items():
                    if sub_val is not None and sub_key:
                        direct[sub_key] = sub_val
            # (2) From value_by_key: variant KPIs -> direct label (D_Tot_M -> male, etc.)
            kpis_here = kpis_by_iso_year.get((iso_3, year), set())
            for vkpi in kpis_here:
                if vkpi == kpi or vkpi == base:
                    continue
                if not vkpi.startswith(prefix):
                    continue
                vval = value_by_key.get((iso_3, year, vkpi))
                if vval is None:
                    continue
                if vkpi.endswith("_I"):
                    values["indirect"] = vval
                    continue
                label = _cpd_variant_to_direct_label(vkpi, prefix)
                if label:
                    direct[label] = vval
        if indirect is not None:
            values["indirect"] = indirect
        # Only add disagg_data when there is actual breakdown (direct or indirect); keep empty for total-only
        has_direct = values.get("direct") and len(values.get("direct") or {}) > 0
        has_indirect = values.get("indirect") is not None
        if has_direct or has_indirect:
            if "indirect" not in values:
                values["indirect"] = None
            disagg_data = json.dumps({"mode": mode, "values": values})
        else:
            disagg_data = ""
        value = _coerce_value(r.get("published_value") or r.get("value"))
        sub_at = submitted_at_default
        out.append(_normalize_headers({
            COL_ASSIGNMENT: str(aes_id),
            COL_PUBLIC: "",
            COL_ITEM: str(item_id),
            COL_VALUE: value or "",
            COL_DISAGG: disagg_data or "",
            COL_DATA_NA: "",
            COL_NA: "",
            COL_PREFILLED: "",
            COL_IMPUTED: "",
            COL_SUBMITTED: sub_at.strftime("%d/%m/%Y  %H:%M:%S") if sub_at else "",
            # Debug metadata for preview Excel (ignored by row_to_payload)
            "_debug_year": year,
            "_debug_iso3": iso_3,
            "_debug_kpi_code": kpi,
        }))
    # Debug: conversion funnel
    logger.info(
        "[DEBUG] Conversion: total_rows=%d; dropped: no item_id (mapping)=%d, no assignment (year,iso3)=%d; out=%d",
        len(total_rows), drop_no_bank_id, drop_no_aes_id, len(out)
    )
    if total_rows and drop_no_aes_id and len(out) == 0:
        sample_lookups = set()
        for r in total_rows[:20]:
            sample_lookups.add((_year_key(r.get("year")), (r.get("iso_3") or "").strip()))
        logger.info("[DEBUG] Sample (year, iso3) lookups: %s", sorted(sample_lookups)[:10])
        sample_item_ids = set()
        for r in total_rows[:50]:
            kpi = (r.get("KPI_code") or "").strip()
            iid = _fdrs_lookup_item_id(kpi, code_to_item_id)
            if iid is not None:
                sample_item_ids.add(iid)
        logger.info("[DEBUG] Sample item_ids from code_to_item_id: %s", sorted(sample_item_ids)[:15])
    return out, dropped_rows


def row_to_payload(row: Dict[str, str]) -> Tuple[Optional[int], Optional[int], Optional[int], Dict[str, Any]]:
    """
    Convert a normalized row to (assignment_entity_status_id, public_submission_id, form_item_id, payload).
    Returns (None, None, None, {}) if row is invalid (e.g. missing required ids).
    """
    aes_raw = row.get(COL_ASSIGNMENT)
    item_raw = row.get(COL_ITEM) or row.get(COL_ITEM_ALT)
    if not aes_raw and not row.get(COL_PUBLIC):
        return None, None, None, {}
    try:
        assignment_entity_status_id = int(aes_raw) if aes_raw else None
    except (ValueError, TypeError):
        assignment_entity_status_id = None
    try:
        form_item_id = int(item_raw) if item_raw else None
    except (ValueError, TypeError):
        form_item_id = None
    public_raw = row.get(COL_PUBLIC)
    try:
        public_submission_id = int(public_raw) if public_raw else None
    except (ValueError, TypeError):
        public_submission_id = None

    if not form_item_id:
        return None, None, None, {}
    if not assignment_entity_status_id and not public_submission_id:
        return None, None, None, {}

    disagg_data = _parse_json_field(row.get(COL_DISAGG))
    prefilled_value = _parse_json_field(row.get(COL_PREFILLED))
    imputed_value = _parse_json_field(row.get(COL_IMPUTED))

    payload = {
        "value": _coerce_value(row.get(COL_VALUE)),
        # Ensure any numeric values inside JSON payloads are whole numbers too.
        "disagg_data": _normalize_json_numbers_to_ints(disagg_data),
        "data_not_available": _parse_bool(row.get(COL_DATA_NA)),
        "not_applicable": _parse_bool(row.get(COL_NA)),
        "prefilled_value": _normalize_json_numbers_to_ints(prefilled_value),
        "imputed_value": _normalize_json_numbers_to_ints(imputed_value),
        "submitted_at": _parse_submitted_at(row.get(COL_SUBMITTED)),
    }
    return assignment_entity_status_id, public_submission_id, form_item_id, payload


def run_import(
    input_path: Optional[str] = None,
    fdrs_api_url: Optional[str] = None,
    fdrs_api_key: Optional[str] = None,
    fdrs_from_data_api: bool = False,
    fdrs_data_api_base: Optional[str] = None,
    fdrs_data_api_key: Optional[str] = None,
    fdrs_imputed_url: Optional[str] = None,
    fdrs_imputed_from_api: bool = False,
    fdrs_imputed_kpi_codes_path: Optional[str] = None,
    fdrs_imputed_use_cache: bool = True,
    fdrs_years: Optional[List[int]] = None,
    fdrs_reported_import_states: Optional[List[int]] = None,
    indicator_mapping_path: Optional[str] = None,
    indicator_bank_api_base: Optional[str] = None,
    indicator_bank_api_key: Optional[str] = None,
    databank_base_url: Optional[str] = None,
    databank_api_key: Optional[str] = None,
    preview_excel_path: Optional[str] = None,
    snapshot_excel_path: Optional[str] = None,
    debug_kpi: Optional[str] = None,
    test_limit: Optional[int] = None,
    dry_run: bool = False,
    batch_size: int = 1000,
    template_id: Optional[int] = 21,
    progress_cb: Optional[Callable[[Dict[str, Any]], None]] = None,
) -> Dict[str, int]:
    """Load from file, FDRS API (ready-to-import), or data-api.ifrc.org pipeline; upsert into form_data.

    fdrs_reported_import_states: optional IFRC State codes (0,100,200,300,400,500) for reported values when not imputed;
    None uses FDRS_REPORTED_IMPORT_STATES env or default all except Not filled (0).
    """
    from app import create_app
    from app.extensions import db
    from app.models.forms import FormData
    from app.models.form_items import FormItem
    from app.models.assignments import AssignmentEntityStatus
    from app.models.assignments import AssignedForm, AssignmentEntityStatus
    from app.models.core import Country

    sources = sum([bool(input_path), bool(fdrs_api_url), bool(fdrs_from_data_api)])
    if sources != 1:
        raise ValueError("Provide exactly one of: input_path, fdrs_api_url, or fdrs_from_data_api")
    if fdrs_from_data_api and indicator_mapping_path and not os.path.isfile(indicator_mapping_path):
        raise ValueError(f"indicator_mapping file not found: {indicator_mapping_path}")

    app = create_app()
    stats = {"loaded": 0, "skipped": 0, "inserted": 0, "updated": 0, "errors": 0}

    def _progress(
        stage: str,
        message: str = "",
        *,
        current: Optional[int] = None,
        total: Optional[int] = None,
        percent: Optional[float] = None,
        extra: Optional[Dict[str, Any]] = None,
    ) -> None:
        """Best-effort progress reporting (must never break import)."""
        if not progress_cb:
            return
        payload: Dict[str, Any] = {
            "stage": stage,
            "message": message,
            "current": current,
            "total": total,
            "percent": percent,
        }
        if extra:
            payload.update(extra)
        try:
            progress_cb(payload)
        except Exception as e:
            logger.debug("progress_cb failed: %s", e)

    with app.app_context():
        if fdrs_from_data_api:
            from flask import current_app
            from fdrs_data_fetcher import build_fdrs_table
            from app.utils.datetime_helpers import utcnow

            base_data = (
                (fdrs_data_api_base or "").strip().rstrip("/")
                or (current_app.config.get("FDRS_DATA_API_BASE") or "").strip().rstrip("/")
                or DEFAULT_FDRS_DATA_API_BASE
            ).rstrip("/")
            fdrs_api_key = (
                (fdrs_data_api_key or "").strip()
                or (os.environ.get("FDRS_DATA_API_KEY") or "").strip()
                or (current_app.config.get("FDRS_DATA_API_KEY") or "").strip()
                or (DEFAULT_FDRS_DATA_API_KEY or "").strip()
            )
            if not fdrs_api_key:
                raise ValueError(
                    "FDRS_DATA_API_KEY is not set. IFRC data-api.ifrc.org returns HTTP 400 without it. "
                    "Add FDRS_DATA_API_KEY to Backoffice/.env (see env.example) or the server environment, then restart the app."
                )
            # Databank = current app (localhost or databank.ifrc.org)
            databank_base = (databank_base_url or current_app.config.get("BASE_URL", "http://localhost:5000")).rstrip("/")
            explicit_databank_base = bool((databank_base_url or "").strip())
            databank_key: Optional[str] = (databank_api_key or "").strip() or None
            if not databank_key:
                for cand in (
                    os.environ.get("FDRS_DATABANK_API_KEY"),
                    os.environ.get("DATABANK_API_KEY"),
                    current_app.config.get("FDRS_DATABANK_API_KEY"),
                    current_app.config.get("DATABANK_API_KEY"),
                    current_app.config.get("MOBILE_APP_API_KEY"),
                    DEFAULT_DATABANK_API_KEY,
                    current_app.config.get("API_KEY"),
                ):
                    if cand and str(cand).strip():
                        databank_key = str(cand).strip()
                        break
            if explicit_databank_base and not databank_key:
                raise ValueError(
                    "databank_base_url is set but no databank API key was found. Set FDRS_DATABANK_API_KEY "
                    "(or DATABANK_API_KEY / MOBILE_APP_API_KEY) for Bearer auth to /api/v1/*, "
                    "or omit databank_base_url to use this application's database for lookups."
                )
            # Default (admin FDRS sync, CLI without --databank-base-url): read assignments / form items /
            # indicator bank via SQLAlchemy. Do not use HTTP here even if env has a databank/mobile key:
            # wrong or non-matching keys cause 401 and zero rows; urllib cannot send a browser session.
            use_local_databank = not explicit_databank_base

            _progress(stage="fetch_fdrs", message="Fetching FDRS data (codebook, fdrsdata, disagg)...", percent=1.0)
            logger.info("Fetching FDRS data (codebook, fdrsdata, country map, FDRS Data + disagg)...")
            cache_dir = os.path.join(current_app.instance_path, "fdrs_imputed_cache")
            fdrs_data, disagg_by_key, exclusion_summary, fdrs_snapshot_rows = build_fdrs_table(
                base_url=base_data,
                api_key=fdrs_api_key,
                years=fdrs_years,
                imputed_url=fdrs_imputed_url,
                imputed_api_key=fdrs_api_key,
                use_imputed_cache=fdrs_imputed_use_cache,
                cache_dir=cache_dir,
                progress_cb=progress_cb,
                reported_import_states=fdrs_reported_import_states,
            )
            logger.info("FDRS Data rows: %d, disagg keys: %d", len(fdrs_data), len(disagg_by_key))

            _progress(stage="fetch_databank", message="Fetching databank lookups (assignments, form items, indicator bank)...", percent=8.0)
            if use_local_databank:
                logger.info(
                    "Databank lookups: in-process SQL (same app; databank_base_url not overridden). "
                    "HTTP /api/v1/* is not used for lookups."
                )
                assignment_rows, form_item_rows, indicator_bank_rows = _fetch_databank_tables_local(template_id or 21)
            else:
                logger.info("Fetching databank APIs (assigned-forms, form-items, indicator-bank)...")
                assignment_rows = fetch_assignment_id_table(databank_base, databank_key, template_id=template_id or 21)
                form_item_rows = fetch_form_item_table(databank_base, databank_key, template_id=template_id or 21)
                indicator_bank_rows = fetch_indicator_bank_table(databank_base, databank_key)
            logger.info(
                "Databank: assignments=%d, form_items=%d, indicators=%d",
                len(assignment_rows), len(form_item_rows), len(indicator_bank_rows)
            )

            _progress(stage="build_rows", message="Building rows to import (merge + mapping)...", percent=12.0)
            logger.info("Building ready-to-import rows (merge FDRS Data + disagg + indicator bank + form items + assignments)...")
            debug_row = os.environ.get("FDRS_DEBUG_ROW")
            debug_iso3_year_item = None
            if debug_row:
                parts = [p.strip() for p in debug_row.split(",")]
                if len(parts) >= 3:
                    try:
                        debug_iso3_year_item = (parts[0], parts[1], int(parts[2]))
                    except ValueError:
                        pass
            rows = build_ready_to_import_from_new_pipeline(
                fdrs_data,
                disagg_by_key,
                assignment_rows,
                form_item_rows,
                indicator_bank_rows,
                submitted_at_default=utcnow(),
                debug_iso3_year_item=debug_iso3_year_item,
                exclusion_summary=exclusion_summary,
            )
            dropped_rows = []
            stats["exclusion_summary"] = exclusion_summary
            summary_text = format_exclusion_summary(exclusion_summary)
            if summary_text:
                logger.info("%s", summary_text)
            if snapshot_excel_path or debug_kpi:
                snapshot_rows = build_fdrs_snapshot_export_rows(
                    fdrs_snapshot_rows=fdrs_snapshot_rows,
                    ready_rows=rows,
                    exclusion_summary=exclusion_summary,
                    disagg_by_key=disagg_by_key,
                )
                if snapshot_excel_path:
                    _progress(stage="snapshot_export", message="Writing FDRS snapshot Excel...", percent=17.0)
                    logger.info("Writing FDRS snapshot (%d rows) to %s...", len(snapshot_rows), snapshot_excel_path)
                    write_rows_to_excel(snapshot_rows, snapshot_excel_path, columns=SNAPSHOT_EXCEL_COLUMNS)
                if debug_kpi:
                    print_debug_kpi_summary(snapshot_rows, debug_kpi)
        elif fdrs_api_url:
            _progress(stage="fetch_api", message="Fetching ready-to-import rows from FDRS API...", percent=1.0)
            rows = fetch_fdrs_api(fdrs_api_url, api_key=fdrs_api_key)
            dropped_rows = []
        else:
            _progress(stage="load_file", message="Loading input file...", percent=1.0)
            rows = load_input(input_path)
            dropped_rows = []
        stats["loaded"] = len(rows)
        _progress(stage="rows_loaded", message=f"Rows loaded: {stats['loaded']}", current=stats["loaded"], total=stats["loaded"], percent=15.0)
        if not rows:
            # Dry-run preview uses a NamedTemporaryFile; if we return without writing, the file stays
            # 0 bytes and Excel reports "file format or extension is not valid".
            if preview_excel_path:
                _progress(stage="preview_export", message="Writing preview Excel (no importable rows)...", percent=18.0)
                logger.info("Writing 0 ready-to-import rows to %s...", preview_excel_path)
                extra_preview_sheets: List[Tuple[str, List[Dict[str, Any]], Optional[Tuple[str, ...]]]] = []
                if dropped_rows:
                    extra_preview_sheets.append(
                        ("Unmatched FDRS", dropped_rows, ("year", "iso_3", "kpi_code", "value", "drop_reason"))
                    )
                excl = stats.get("exclusion_summary")
                if excl:
                    txt = format_exclusion_summary(excl)
                    if txt.strip():
                        extra_preview_sheets.append(
                            ("Exclusion summary", [{"line": ln} for ln in txt.splitlines()], ("line",))
                        )
                write_rows_to_excel(
                    [],
                    preview_excel_path,
                    columns=PREVIEW_EXCEL_COLUMNS,
                    extra_sheets=extra_preview_sheets or None,
                )
                _progress(
                    stage="complete",
                    message="Preview exported (no importable rows).",
                    current=0,
                    total=0,
                    percent=100.0,
                    extra={"stats": dict(stats)},
                )
            return stats

        if test_limit is not None and test_limit > 0 and len(rows) > test_limit:
            rows = rows[:test_limit]
            logger.info("Test limit: using first %d records only.", test_limit)
            stats["loaded"] = len(rows)
            _progress(stage="rows_loaded", message=f"Test limit applied: {stats['loaded']} rows", current=stats["loaded"], total=stats["loaded"], percent=15.0)

        if preview_excel_path:
            _progress(stage="preview_export", message="Writing preview Excel...", percent=18.0)
            logger.info("Writing %d rows to %s...", len(rows), preview_excel_path)
            extra_sheets = []
            if dropped_rows:
                extra_sheets.append(("Unmatched FDRS", dropped_rows, ("year", "iso_3", "kpi_code", "value", "drop_reason")))
            write_rows_to_excel(rows, preview_excel_path, columns=PREVIEW_EXCEL_COLUMNS, extra_sheets=extra_sheets or None)
            # Show what values made it into disagg_data (first 5 non-empty, varied)
            disagg_col = COL_DISAGG
            seen = set()
            count = 0
            for r in rows[:100]:
                d = (r.get(disagg_col) or "").strip()
                if not d or d in seen:
                    continue
                seen.add(d)
                count += 1
                kpi = r.get("_debug_kpi_code", "")
                iso3 = r.get("_debug_iso3", "")
                yr = r.get("_debug_year", "")
                try:
                    parsed = json.loads(d)
                    values = parsed.get("values") or {}
                    direct = values.get("direct") or {}
                    indirect = values.get("indirect")
                    direct_preview = ", ".join(f"{k}={direct[k]}" for k in sorted(direct.keys())[:8])
                    if len(direct) > 8:
                        direct_preview += f" ... (+{len(direct) - 8} more)"
                    indirect_str = f", indirect={indirect}" if indirect is not None else ""
                    logger.info("[disagg_data] %s %s %r: direct(%d): {%s}%s", iso3, yr, kpi, len(direct), direct_preview, indirect_str)
                except Exception as e:
                    logger.debug("[disagg_data] %s %s %r parse failed: %s; raw: %s%s", iso3, yr, kpi, e, d[:100], '...' if len(d) > 100 else '')
                if count >= 5:
                    break
            _progress(stage="complete", message="Preview exported.", current=len(rows), total=len(rows), percent=100.0, extra={"stats": dict(stats)})
            return stats

        # Optional: restrict to form items of this template
        valid_form_item_ids = None
        if template_id is not None:
            valid_form_item_ids = set(
                fid for (fid,) in db.session.query(FormItem.id).filter(FormItem.template_id == template_id).all()
            )
        valid_aes_ids = set(
            aid for (aid,) in db.session.query(AssignmentEntityStatus.id).all()
        )

        # NOTE: Upsert performance
        # The naive approach (querying FormData per row) is extremely slow for large imports.
        # We batch-prefetch existing FormData rows for the next N input rows, then do in-memory matching.
        from sqlalchemy import tuple_

        total_rows = len(rows)
        _progress(stage="upsert", message="Starting upsert...", current=0, total=total_rows, percent=20.0, extra={"stats": dict(stats)})

        def _maybe_report(i: int, row: Dict[str, Any]) -> None:
            if not progress_cb:
                return
            if not (i == 1 or i % 50 == 0 or i == total_rows):
                return
            pct = 20.0 + (80.0 * (i / total_rows)) if total_rows else 100.0
            kpi = (row.get("_debug_kpi_code") or "").strip()
            iso3 = (row.get("_debug_iso3") or "").strip()
            yr = (row.get("_debug_year") or "").strip()
            details = " ".join(b for b in (iso3, yr, kpi) if b).strip()
            msg = f"Processing {i}/{total_rows} ({pct:.1f}%)" + (f" - {details}" if details else "")
            _progress(
                stage="upsert",
                message=msg,
                current=i,
                total=total_rows,
                percent=pct,
                extra={"stats": dict(stats)},
            )

        # Batch size for prefetching existing records (not the DB commit batch size).
        prefetch_size = max(2000, int(batch_size or 1000))

        for batch_start in range(0, total_rows, prefetch_size):
            batch = rows[batch_start: batch_start + prefetch_size]

            # Prefetch existing records for this batch
            aes_pairs_set = set()
            pub_pairs_set = set()
            for r in batch:
                try:
                    aes_id, pub_id, item_id, _ = row_to_payload(r)
                except Exception as e:
                    logger.debug("row_to_payload failed: %s", e)
                    continue
                if not item_id:
                    continue
                if aes_id:
                    aes_pairs_set.add((int(aes_id), int(item_id)))
                elif pub_id:
                    pub_pairs_set.add((int(pub_id), int(item_id)))

            existing_by_aes: Dict[Tuple[int, int], FormData] = {}
            existing_by_pub: Dict[Tuple[int, int], FormData] = {}
            if aes_pairs_set:
                q = (
                    FormData.query.filter(
                        tuple_(FormData.assignment_entity_status_id, FormData.form_item_id).in_(list(aes_pairs_set))
                    )
                )
                for fd in q.all():
                    key = (int(fd.assignment_entity_status_id), int(fd.form_item_id))
                    # If duplicates exist, keep the first one we see (shouldn't happen ideally).
                    existing_by_aes.setdefault(key, fd)
            if pub_pairs_set:
                q = (
                    FormData.query.filter(
                        tuple_(FormData.public_submission_id, FormData.form_item_id).in_(list(pub_pairs_set))
                    )
                )
                for fd in q.all():
                    key = (int(fd.public_submission_id), int(fd.form_item_id))
                    existing_by_pub.setdefault(key, fd)

            for j, row in enumerate(batch, start=batch_start + 1):
                assignment_entity_status_id, public_submission_id, form_item_id, payload = row_to_payload(row)
                if not form_item_id or (not assignment_entity_status_id and not public_submission_id):
                    stats["skipped"] += 1
                    _maybe_report(j, row)
                    continue
                if valid_form_item_ids is not None and form_item_id not in valid_form_item_ids:
                    stats["skipped"] += 1
                    _maybe_report(j, row)
                    continue
                if assignment_entity_status_id and assignment_entity_status_id not in valid_aes_ids:
                    stats["skipped"] += 1
                    _maybe_report(j, row)
                    continue

                if assignment_entity_status_id:
                    existing = existing_by_aes.get((int(assignment_entity_status_id), int(form_item_id)))
                else:
                    existing = existing_by_pub.get((int(public_submission_id), int(form_item_id)))

                if dry_run:
                    if existing:
                        stats["updated"] += 1
                    else:
                        stats["inserted"] += 1
                    _maybe_report(j, row)
                    continue

                try:
                    # IMPORTANT (PostgreSQL JSON/JSONB): Python None may serialize to JSON literal `null`.
                    # Use db.null() to store a real SQL NULL when there's no disaggregation payload.
                    disagg_for_db = _disagg_data_for_db(payload["disagg_data"])
                    if disagg_for_db is None:
                        disagg_for_db = db.null()
                    # Same rule for other JSON columns: store SQL NULL (not JSON null) when empty.
                    prefilled_for_db = payload["prefilled_value"]
                    if prefilled_for_db is None or prefilled_for_db == {} or prefilled_for_db == []:
                        prefilled_for_db = db.null()
                    imputed_for_db = payload["imputed_value"]
                    if imputed_for_db is None or imputed_for_db == {} or imputed_for_db == []:
                        imputed_for_db = db.null()
                    if existing:
                        existing.value = payload["value"]
                        existing.disagg_data = disagg_for_db
                        existing.data_not_available = payload["data_not_available"]
                        existing.not_applicable = payload["not_applicable"]
                        existing.prefilled_value = prefilled_for_db
                        existing.imputed_value = imputed_for_db
                        if payload["submitted_at"] is not None:
                            existing.submitted_at = payload["submitted_at"]
                        db.session.add(existing)
                        stats["updated"] += 1
                    else:
                        entry = FormData(
                            assignment_entity_status_id=assignment_entity_status_id,
                            public_submission_id=public_submission_id,
                            form_item_id=form_item_id,
                            value=payload["value"],
                            disagg_data=disagg_for_db,
                            data_not_available=payload["data_not_available"],
                            not_applicable=payload["not_applicable"],
                            prefilled_value=prefilled_for_db,
                            imputed_value=imputed_for_db,
                            submitted_at=payload["submitted_at"],
                        )
                        db.session.add(entry)
                        stats["inserted"] += 1
                        # So same (aes, item) later in this batch updates this instead of inserting again
                        if assignment_entity_status_id:
                            existing_by_aes[(int(assignment_entity_status_id), int(form_item_id))] = entry
                        else:
                            existing_by_pub[(int(public_submission_id), int(form_item_id))] = entry
                except Exception as e:
                    stats["errors"] += 1
                    if j < 5 or stats["errors"] <= 3:
                        logger.error("Row %d error: %s", j, e)

                if batch_size and ((stats["inserted"] + stats["updated"]) % batch_size == 0) and (stats["inserted"] + stats["updated"]) > 0:
                    db.session.commit()
                _maybe_report(j, row)

        if not dry_run and (stats["inserted"] + stats["updated"]) > 0:
            db.session.commit()

        _progress(stage="complete", message="Sync completed.", current=total_rows, total=total_rows, percent=100.0, extra={"stats": dict(stats)})
    return stats


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Import or sync FDRS-prepared data into form_data (insert new / update existing)."
    )
    parser.add_argument(
        "--input", "-i",
        help="Path to CSV or Excel file (Ready to import shape). Use this OR --fdrs-api-url.",
    )
    parser.add_argument(
        "--fdrs-api-url",
        help="URL to fetch FDRS data as JSON (same Ready-to-import shape). Use this OR --input.",
    )
    parser.add_argument(
        "--fdrs-api-key",
        help="Optional API key or Bearer token for FDRS API.",
    )
    parser.add_argument(
        "--fdrs-from-data-api",
        action="store_true",
        help="Fetch FDRS from data-api.ifrc.org and build ready-to-import via databank APIs (assigned-forms, form-items, indicator-bank).",
    )
    parser.add_argument(
        "--fdrs-data-api-base",
        default="https://data-api.ifrc.org",
        help="Base URL for FDRS data API (default: https://data-api.ifrc.org).",
    )
    parser.add_argument(
        "--fdrs-data-api-key",
        help="API key for data-api.ifrc.org (or set FDRS_DATA_API_KEY).",
    )
    parser.add_argument(
        "--databank-base-url",
        help="Override base URL for databank API (default: current app BASE_URL, e.g. http://localhost:5000 or https://databank.ifrc.org).",
    )
    parser.add_argument(
        "--databank-api-key",
        help="Bearer token for databank API. Default from env FDRS_DATABANK_API_KEY or DATABANK_API_KEY.",
    )
    parser.add_argument(
        "--fdrs-imputed-url",
        help="Optional URL that returns a full JSON array of imputed rows (alternative to --fdrs-imputed-from-api).",
    )
    parser.add_argument(
        "--fdrs-imputed-from-api",
        action="store_true",
        help="Fetch imputed values from data-api KpiImputedValue API (per kpicode/year). Use with --fdrs-from-data-api.",
    )
    parser.add_argument(
        "--fdrs-imputed-kpi-codes",
        help="Optional CSV with column KPI or kpicode to limit which KPIs to request from KpiImputedValue API. If omitted, all non-IP KPIs from codebook are used.",
    )
    parser.add_argument(
        "--fdrs-years",
        help="Comma-separated years for FDRS fetch (e.g. 2019,2020,2021). Default: 2010-2024.",
    )
    parser.add_argument(
        "--fdrs-reported-states",
        metavar="CODES",
        dest="fdrs_reported_import_states",
        help=(
            "Comma-separated IFRC fdrsdata State codes to import as reported values when not imputed "
            "(0 Not filled, 100 Saved, 200 Reopened, 300 Submitted, 400 Validated, 500 Published). "
            "Default: 100,200,300,400,500 (all except Not filled) or FDRS_REPORTED_IMPORT_STATES."
        ),
    )
    parser.add_argument(
        "--test",
        action="store_true",
        help="Testing mode: fetch only year 2024 for faster runs.",
    )
    parser.add_argument(
        "--indicator-mapping",
        help="CSV or Excel with fdrs_code, indicator_bank_id, item_id. Optional if INDICATOR_BANK_API_* is set in local config.",
    )
    parser.add_argument(
        "--export-mapping",
        metavar="FILE",
        help="Export current KPI->indicator_bank_id->item_id mapping to Excel and exit. Use with --indicator-mapping CSV or --fdrs-from-data-api. Edit the file then pass it back via --indicator-mapping.",
    )
    parser.add_argument(
        "--preview-excel",
        nargs="?",
        const="fdrs_import_preview.xlsx",
        metavar="FILE",
        help="Export ready-to-import rows to Excel and exit without writing to DB (default: fdrs_import_preview.xlsx).",
    )
    parser.add_argument(
        "--snapshot-excel",
        nargs="?",
        const="fdrs_raw_snapshot.xlsx",
        metavar="FILE",
        help="Export raw FDRS snapshot with import_status/import_filter_reason (default: fdrs_raw_snapshot.xlsx).",
    )
    parser.add_argument(
        "--debug-kpi",
        metavar="BASE_KPI",
        help="Print why this BaseKPI was/wasn't imported (counts by import_status and import_filter_reason). Use with --fdrs-from-data-api.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Preview changes without writing to the database")
    parser.add_argument("--batch-size", type=int, default=1000, help="Commit every N rows (default: 1000)")
    parser.add_argument(
        "--template-id",
        type=int,
        default=21,
        help="Only allow form_item_id from this template (default: 21). Use 0 to disable.",
    )
    args = parser.parse_args()

    # No args = full API pipeline (no files, no local config)
    no_source = not args.input and not args.fdrs_api_url and not args.fdrs_from_data_api
    if no_source:
        args.fdrs_from_data_api = True
        args.fdrs_data_api_base = args.fdrs_data_api_base or DEFAULT_FDRS_DATA_API_BASE
        args.fdrs_data_api_key = args.fdrs_data_api_key or DEFAULT_FDRS_DATA_API_KEY
        args.databank_api_key = getattr(args, "databank_api_key", None) or DEFAULT_DATABANK_API_KEY
    else:
        args.fdrs_data_api_base = args.fdrs_data_api_base or DEFAULT_FDRS_DATA_API_BASE
        args.fdrs_data_api_key = args.fdrs_data_api_key or DEFAULT_FDRS_DATA_API_KEY
        if args.fdrs_from_data_api:
            args.databank_api_key = getattr(args, "databank_api_key", None) or DEFAULT_DATABANK_API_KEY

    if sum([bool(args.input), bool(args.fdrs_api_url), bool(args.fdrs_from_data_api)]) != 1:
        logger.error("Error: provide at most one of --input, --fdrs-api-url, or --fdrs-from-data-api")
        return 1
    if args.fdrs_from_data_api and not (args.fdrs_data_api_key or "").strip():
        logger.error(
            "Error: FDRS data API key required — set environment variable FDRS_DATA_API_KEY "
            "or pass --fdrs-data-api-key."
        )
        return 1
    if args.input and not os.path.isfile(args.input):
        logger.error("Error: input file not found: %s", args.input)
        return 1
    if args.fdrs_from_data_api and args.indicator_mapping and not os.path.isfile(args.indicator_mapping):
        logger.error("Error: indicator-mapping file not found: %s", args.indicator_mapping)
        return 1

    # Export mapping only: build current mapping and write Excel for manual edit
    if args.export_mapping:
        has_mapping_source = (
            (args.indicator_mapping and os.path.isfile(args.indicator_mapping))
            or (getattr(args, "databank_api_key", None) or DEFAULT_DATABANK_API_KEY)
        )
        if not has_mapping_source:
            logger.error(
                "Error: --export-mapping requires either --indicator-mapping CSV or --fdrs-from-data-api with indicator bank API configured."
            )
            return 1
        from app import create_app
        app = create_app()
        template_id = args.template_id if args.template_id else 21
        with app.app_context():
            from app.extensions import db
            from app.models.form_items import FormItem
            from app.models.indicator_bank import IndicatorBank
            rows = []
            if args.indicator_mapping and os.path.isfile(args.indicator_mapping):
                path_lower = args.indicator_mapping.lower()
                if path_lower.endswith((".xlsx", ".xls")):
                    # Re-export from Excel: fdrs_code -> item_id; look up bank_id and name from DB
                    code_to_item_id = _load_indicator_mapping_from_excel(args.indicator_mapping)
                    item_ids = list(set(code_to_item_id.values()))
                    item_to_bank: Dict[int, int] = {}
                    if item_ids:
                        for (iid, bid) in db.session.query(FormItem.id, FormItem.indicator_bank_id).filter(
                            FormItem.template_id == template_id, FormItem.id.in_(item_ids)
                        ).all():
                            if bid is not None:
                                item_to_bank[int(iid)] = int(bid)
                    bank_ids = list(set(item_to_bank.values()))
                    id_to_name: Dict[int, str] = {}
                    if bank_ids:
                        id_to_name = dict(
                            db.session.query(IndicatorBank.id, IndicatorBank.name).filter(
                                IndicatorBank.id.in_(bank_ids), IndicatorBank.name.isnot(None)
                            ).all()
                        )
                    for fdrs_code in sorted(code_to_item_id.keys()):
                        item_id = code_to_item_id[fdrs_code]
                        bank_id = item_to_bank.get(item_id) or ""
                        indicator_name = id_to_name.get(bank_id, "") if isinstance(bank_id, int) else ""
                        rows.append({
                            "fdrs_code": fdrs_code,
                            "indicator_bank_id": bank_id,
                            "item_id": item_id,
                            "indicator_name": indicator_name,
                        })
                else:
                    mapping = _load_indicator_mapping(args.indicator_mapping)
                    form_item_by_bank = {}
                    for (fid, bid) in db.session.query(FormItem.id, FormItem.indicator_bank_id).filter(
                        FormItem.template_id == template_id, FormItem.indicator_bank_id.isnot(None)
                    ).all():
                        if bid is not None:
                            form_item_by_bank[int(bid)] = int(fid)
                    bank_ids = list(set(mapping.values()))
                    id_to_name = {}
                    if bank_ids:
                        id_to_name = dict(
                            db.session.query(IndicatorBank.id, IndicatorBank.name).filter(
                                IndicatorBank.id.in_(bank_ids), IndicatorBank.name.isnot(None)
                            ).all()
                        )
                    for fdrs_code, bank_id in sorted(mapping.items()):
                        item_id = form_item_by_bank.get(bank_id) or ""
                        indicator_name = id_to_name.get(bank_id) or ""
                        rows.append({
                            "fdrs_code": fdrs_code,
                            "indicator_bank_id": bank_id,
                            "item_id": item_id,
                            "indicator_name": indicator_name,
                        })
            else:
                # Use current app API (indicator-bank, form-items) for export when no mapping file provided
                databank_base = (getattr(args, "databank_base_url", None) or app.config.get("BASE_URL", "http://localhost:5000")).rstrip("/")
                databank_key = getattr(args, "databank_api_key", None) or DEFAULT_DATABANK_API_KEY or app.config.get("API_KEY")
                indicator_bank_rows = fetch_indicator_bank_table(databank_base, databank_key)
                form_item_rows = fetch_form_item_table(databank_base, databank_key, template_id=template_id)
                bank_to_item = {int(r["bank_id"]): int(r["item_id"]) for r in form_item_rows}
                for ind in sorted(indicator_bank_rows, key=lambda x: (x.get("fdrs_kpi_code") or "")):
                    fdrs_code = (ind.get("fdrs_kpi_code") or "").strip()
                    bank_id = ind.get("id")
                    item_id = bank_to_item.get(bank_id) if bank_id is not None else None
                    indicator_name = (ind.get("name") or "").strip()
                    rows.append({
                        "fdrs_code": fdrs_code,
                        "indicator_bank_id": bank_id if bank_id is not None else "",
                        "item_id": item_id if item_id is not None else "",
                        "indicator_name": indicator_name,
                    })
            write_mapping_to_excel(rows, args.export_mapping)
        logger.info("Mapping exported to %s", args.export_mapping)
        return 0

    fdrs_years = None
    test_limit = None
    if getattr(args, "test", False):
        fdrs_years = [2024]
        test_limit = 1000
        logger.info("Test mode: year 2024 only, max 1000 records.")
    elif args.fdrs_years:
        try:
            fdrs_years = [int(y.strip()) for y in args.fdrs_years.split(",") if y.strip()]
        except ValueError:
            logger.error("Error: --fdrs-years must be comma-separated integers")
            return 1

    template_id = args.template_id if args.template_id else None
    fdrs_reported_import_states_cli = None
    raw_states = getattr(args, "fdrs_reported_import_states", None)
    if raw_states:
        try:
            fdrs_reported_import_states_cli = [int(x.strip()) for x in str(raw_states).split(",") if x.strip()]
        except ValueError:
            logger.error("Error: --fdrs-reported-states must be comma-separated integers")
            return 1

    try:
        stats = run_import(
            input_path=args.input,
            fdrs_api_url=args.fdrs_api_url,
            fdrs_api_key=args.fdrs_api_key or None,
            fdrs_from_data_api=args.fdrs_from_data_api,
            fdrs_data_api_base=args.fdrs_data_api_base or None,
            fdrs_data_api_key=args.fdrs_data_api_key or None,
            fdrs_imputed_url=args.fdrs_imputed_url or None,
            fdrs_imputed_from_api=args.fdrs_imputed_from_api,
            fdrs_imputed_kpi_codes_path=args.fdrs_imputed_kpi_codes or None,
            fdrs_years=fdrs_years,
            fdrs_reported_import_states=fdrs_reported_import_states_cli,
            indicator_mapping_path=args.indicator_mapping or None,
            indicator_bank_api_base=getattr(args, "indicator_bank_api_base", None),
            indicator_bank_api_key=getattr(args, "indicator_bank_api_key", None),
            databank_base_url=getattr(args, "databank_base_url", None),
            databank_api_key=getattr(args, "databank_api_key", None),
            preview_excel_path=args.preview_excel,
            snapshot_excel_path=args.snapshot_excel,
            debug_kpi=getattr(args, "debug_kpi", None),
            test_limit=test_limit,
            dry_run=args.dry_run,
            batch_size=args.batch_size,
            template_id=template_id,
        )
    except (ValueError, RuntimeError) as e:
        logger.error("Error: %s", e)
        return 1

    logger.info("Loaded: %s, Skipped: %s, Inserted: %s, Updated: %s, Errors: %s",
                stats['loaded'], stats['skipped'], stats['inserted'], stats['updated'], stats['errors'])
    if getattr(args, "preview_excel", None):
        logger.info("(preview only - exported to %s, no DB write)", args.preview_excel)
    elif args.dry_run:
        logger.info("(dry run — no changes written)")
    if getattr(args, "snapshot_excel", None) and args.fdrs_from_data_api:
        logger.info("(FDRS snapshot exported to %s)", args.snapshot_excel)
    return 0 if stats["errors"] == 0 else 1


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    sys.exit(main())
