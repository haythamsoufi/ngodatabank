"""
Shared CSV and Excel file parsing utilities.
Reduces duplication of import/encoding logic across lookup lists, indicators, etc.
"""
import csv
import codecs
from typing import List, Dict, Any, Tuple
from werkzeug.datastructures import FileStorage


# Extensions for CSV and Excel
CSV_EXCEL_EXTENSIONS = {".csv", ".xlsx", ".xls"}
EXCEL_EXTENSIONS = {".xlsx", ".xls"}


def _decode_csv_content(file_content: bytes) -> str:
    """Decode CSV file content, handling BOM and common encodings."""
    if file_content.startswith(codecs.BOM_UTF8):
        return file_content.decode("utf-8-sig")
    try:
        return file_content.decode("utf-8")
    except UnicodeDecodeError:
        try:
            return file_content.decode("utf-8-sig")
        except UnicodeDecodeError:
            return file_content.decode("latin-1")


def parse_csv_to_rows(file: FileStorage) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Parse a CSV file into columns and row dicts.

    Args:
        file: FileStorage object (position will be consumed)

    Returns:
        Tuple of (columns: list of header names, rows: list of dicts)
    """
    file_content = file.stream.read()
    text = _decode_csv_content(file_content)
    reader = csv.DictReader(text.splitlines())
    columns = list(reader.fieldnames or [])
    rows = list(reader)
    return columns, rows


def parse_excel_to_rows(file: FileStorage) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Parse an Excel file (.xlsx, .xls) into columns and row dicts.

    Args:
        file: FileStorage object (position will be consumed)

    Returns:
        Tuple of (columns: list of header names, rows: list of dicts)
    """
    from openpyxl import load_workbook

    wb = load_workbook(file, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None) or []
    columns = [str(h) for h in headers]
    rows = []
    for values in rows_iter:
        row = {columns[i]: (values[i] if i < len(values) else None) for i in range(len(columns))}
        rows.append(row)
    return columns, rows


def parse_csv_or_excel_to_rows(file: FileStorage, filename: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """
    Parse a CSV or Excel file into columns and row dicts.
    Chooses parser based on filename extension.

    Args:
        file: FileStorage object (position will be consumed)
        filename: Original filename (used to determine format)

    Returns:
        Tuple of (columns: list of header names, rows: list of dicts)

    Raises:
        ValueError: If file format is not .csv, .xlsx, or .xls
    """
    fn = filename.lower()
    if fn.endswith(".csv"):
        return parse_csv_to_rows(file)
    if fn.endswith((".xlsx", ".xls")):
        return parse_excel_to_rows(file)
    raise ValueError(f"Unsupported file format: {filename}. Expected CSV or Excel.")
