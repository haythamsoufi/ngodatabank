# ========== Template Excel Import/Export Service ==========
from app.utils.datetime_helpers import utcnow
"""
Service for exporting and importing form templates to/from Excel.

This service handles:
- Exporting template structure (pages, sections, items) to Excel with exact DB column mapping
- Importing template structure from Excel with validation and ID mapping
- Preserving all configurations, translations, and skip logic rules
"""

from flask import current_app
from flask_login import current_user
from app import db
from app.models import (
    FormTemplate, FormPage, FormSection, FormItem, FormTemplateVersion
)
from app.models.indicator_bank import IndicatorBank
from contextlib import suppress
from app.utils.memory_monitor import memory_tracker
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
import io
import json
import re
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple


class TemplateExcelService:
    """Service for template Excel export/import operations."""

    # IFRC Color Scheme (matching ExcelService)
    IFRC_COLORS = {
        'RED': "FFED1B2E",
        'DARK_RED': "FFAF0E1B",
        'LIGHT_GRAY': "FFF5F5F5",
        'MEDIUM_GRAY': "FFE0E0E0",
        'DARK_GRAY': "FF666666",
        'WHITE': "FFFFFFFF",
        'YELLOW': "FFFFF9E6",
        'BLUE': "FF0066CC",
        'DARK_BLUE': "FF004499",
    }

    # Required vs Optional columns for each sheet
    REQUIRED_COLUMNS = {
        'Template': ['name'],
        'Pages': ['id', 'name', 'order'],
        'Sections': ['id', 'name', 'order'],
        'Items': ['id', 'section_id', 'item_type', 'label', 'order']
    }

    # Excel export version
    EXCEL_EXPORT_VERSION = 'V1'

    # Dropdown options for data validation (static options)
    DROPDOWN_OPTIONS = {
        'section_type': ['standard', 'repeat', 'dynamic_indicators'],
        # Note: item_type is now dynamic - see _get_item_type_options()
        'archived': ['TRUE', 'FALSE'],
        'add_to_self_report': ['TRUE', 'FALSE'],
        'display_order_visible': ['TRUE', 'FALSE'],
        'is_paginated': ['TRUE', 'FALSE'],
        'enable_export_pdf': ['TRUE', 'FALSE'],
        'enable_export_excel': ['TRUE', 'FALSE'],
        'enable_import_excel': ['TRUE', 'FALSE'],
        'enable_ai_validation': ['TRUE', 'FALSE'],
        'allow_data_not_available': ['TRUE', 'FALSE'],
        'allow_not_applicable': ['TRUE', 'FALSE'],
    }

    # Columns that should not have duplicate values (for conditional formatting)
    UNIQUE_COLUMNS = {
        'Pages': ['id'],
        'Sections': ['id'],
        'Items': ['id'],
    }

    # Column definitions matching database tables (excluding template_id, version_id, and audit fields)
    TEMPLATE_COLUMNS = [
        'name', 'description',
        'add_to_self_report', 'display_order_visible',
        'is_paginated', 'enable_export_pdf', 'enable_export_excel',
        'enable_import_excel', 'enable_ai_validation', 'name_translations', 'variables'
    ]

    PAGE_COLUMNS = [
        'id', 'name', 'order', 'name_translations'
    ]

    SECTION_COLUMNS = [
        'id', 'name', 'order', 'parent_section_id', 'page_id',
        'section_type', 'max_dynamic_indicators', 'allowed_sectors',
        'indicator_filters', 'allow_data_not_available', 'allow_not_applicable',
        'allowed_disaggregation_options', 'data_entry_display_filters',
        'add_indicator_note', 'name_translations', 'relevance_condition', 'archived'
    ]

    ITEM_COLUMNS = [
        'id', 'section_id', 'item_type', 'label', 'order',
        'relevance_condition', 'archived', 'config', 'indicator_bank_id',
        'type', 'unit', 'validation_condition', 'validation_message',
        'definition', 'options_json', 'lookup_list_id', 'list_display_column',
        'list_filters_json', 'label_translations', 'definition_translations',
        'options_translations', 'description_translations', 'description'
    ]

    @classmethod
    def _get_items_for_version(cls, template: FormTemplate, version: FormTemplateVersion) -> List[FormItem]:
        """Get items for a template version in a deterministic order.

        Note: We use a stable secondary sort on FormItem.id to keep export IDs deterministic.
        """
        return FormItem.query.join(FormSection).filter(
            FormItem.template_id == template.id,
            FormSection.version_id == version.id
        ).order_by(FormItem.order, FormItem.id).all()

    @classmethod
    def _build_item_db_to_export_map(cls, template: FormTemplate, version: FormTemplateVersion) -> Dict[int, int]:
        """Build mapping from DB item IDs -> sequential export IDs (1, 2, 3...)."""
        items = cls._get_items_for_version(template, version)
        return {item.id: idx + 1 for idx, item in enumerate(items)}

    @classmethod
    def _rewrite_rule_json_item_ids(cls, rule_json: Any, id_map: Dict[int, int]) -> Any:
        """Rewrite item references inside a relevance/validation rule JSON.

        The rule builder stores references under the key 'item_id'. Values can be:
        - numeric strings (e.g., "66") for regular items
        - prefixed strings (e.g., "plugin_123" or "plugin_123_measure_id")
        - legacy prefixed strings (e.g., "question_66")

        This method is used in two directions depending on the provided id_map:
        - Export: db_id -> export_id
        - Import: export_id -> new_db_id
        """
        if rule_json is None:
            return None
        if isinstance(rule_json, str) and rule_json.strip() == '':
            return rule_json

        # Parse (handle occasional double-encoded JSON)
        parsed = None
        if isinstance(rule_json, (dict, list)):
            parsed = rule_json
        else:
            raw = str(rule_json)
            with suppress(Exception):
                parsed = json.loads(raw)
                if isinstance(parsed, str):
                    parsed2 = json.loads(parsed)
                    parsed = parsed2
        if parsed is None:
            return rule_json

        def _rewrite_item_id_value(val: Any) -> Any:
            if val is None:
                return val
            # Numeric id stored as int/float/string
            if isinstance(val, (int, float)) or (isinstance(val, str) and val.isdigit()):
                old_id = int(val)
                new_id = id_map.get(old_id)
                return str(new_id) if new_id is not None else val

            if isinstance(val, str):
                # plugin_123 or plugin_123_measure_id
                if val.startswith('plugin_'):
                    parts = val.split('_')
                    if len(parts) >= 2 and parts[1].isdigit():
                        old_id = int(parts[1])
                        new_id = id_map.get(old_id)
                        if new_id is not None:
                            parts[1] = str(new_id)
                            return '_'.join(parts)
                    return val

                # Legacy prefixed IDs (question_66, indicator_12, document_field_7, form_item_99)
                m = re.match(r'^(question|indicator|document_field|form_item)_(\d+)$', val)
                if m:
                    old_id = int(m.group(2))
                    new_id = id_map.get(old_id)
                    if new_id is not None:
                        return f"{m.group(1)}_{new_id}"
                    return val

            return val

        def _walk(obj: Any) -> Any:
            if isinstance(obj, list):
                for i in range(len(obj)):
                    obj[i] = _walk(obj[i])
                return obj
            if isinstance(obj, dict):
                for k, v in list(obj.items()):
                    if k == 'item_id':
                        obj[k] = _rewrite_item_id_value(v)
                    else:
                        obj[k] = _walk(v)
                return obj
            return obj

        parsed = _walk(parsed)
        try:
            return json.dumps(parsed)
        except Exception as e:
            current_app.logger.debug("Rule JSON serialization failed, using original: %s", e)
            # Fall back to the original if serialization fails
            return rule_json

    @classmethod
    @memory_tracker("Template Excel Export", log_top_allocations=True)
    def export_template(cls, template_id: int, version_id: Optional[int] = None) -> io.BytesIO:
        """
        Export template structure to Excel.

        Args:
            template_id: Template ID to export
            version_id: Optional version ID (defaults to published or latest)

        Returns:
            BytesIO object containing Excel file
        """
        template = FormTemplate.query.get_or_404(template_id)

        # Determine version to export
        if version_id:
            version = FormTemplateVersion.query.filter_by(
                id=version_id, template_id=template.id
            ).first()
        else:
            if template.published_version_id:
                version = FormTemplateVersion.query.get(template.published_version_id)
            else:
                version = FormTemplateVersion.query.filter_by(
                    template_id=template.id
                ).order_by(FormTemplateVersion.created_at.desc()).first()

        if not version:
            raise ValueError(f"No version found for template {template_id}")

        # Create workbook
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        # Export Instructions sheet first (so it appears first)
        cls._export_instructions_sheet(workbook)

        # Export hidden metadata sheet
        cls._export_metadata_sheet(workbook, template, version)

        # Export Template sheet
        cls._export_template_sheet(workbook, template, version)

        # Export Pages sheet
        cls._export_pages_sheet(workbook, template, version)

        # Export Sections sheet
        cls._export_sections_sheet(workbook, template, version)

        # Export Items sheet
        cls._export_items_sheet(workbook, template, version)

        # Save to BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return output

    @classmethod
    def _export_template_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export template metadata to Template sheet."""
        sheet = workbook.create_sheet("Template")

        # Write headers with required/optional styling
        headers = cls.TEMPLATE_COLUMNS
        required_cols = cls.REQUIRED_COLUMNS.get('Template', [])
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            is_required = header in required_cols
            cls._style_header_cell(cell, is_required=is_required)

        # Write data row
        #
        # Versioning note:
        # Template configuration fields (description/etc.) are version-scoped
        # and live on FormTemplateVersion. Some older DBs may still have legacy columns
        # on FormTemplate, so we use safe getattr() fallbacks to avoid export crashes.
        version_name = version.name if version.name else template.name
        version_name_translations = version.name_translations if version.name_translations else template.name_translations

        version_description = version.description
        if version_description is None:
            version_description = getattr(template, 'description', None)

        version_add_to_self_report = version.add_to_self_report if version.add_to_self_report is not None else getattr(template, 'add_to_self_report', False)
        version_display_order_visible = version.display_order_visible if version.display_order_visible is not None else getattr(template, 'display_order_visible', False)
        version_is_paginated = version.is_paginated if version.is_paginated is not None else getattr(template, 'is_paginated', False)
        version_enable_export_pdf = version.enable_export_pdf if version.enable_export_pdf is not None else getattr(template, 'enable_export_pdf', False)
        version_enable_export_excel = version.enable_export_excel if version.enable_export_excel is not None else getattr(template, 'enable_export_excel', False)
        version_enable_import_excel = version.enable_import_excel if version.enable_import_excel is not None else getattr(template, 'enable_import_excel', False)
        version_enable_ai_validation = version.enable_ai_validation if getattr(version, 'enable_ai_validation', None) is not None else getattr(template, 'enable_ai_validation', False)

        version_variables = version.variables if version.variables else None

        row_data = [
            version_name,  # Export version-specific name (or template name as fallback)
            version_description,  # Export version-specific description
            version_add_to_self_report,  # Export version-specific add_to_self_report
            version_display_order_visible,  # Export version-specific display_order_visible
            version_is_paginated,  # Export version-specific is_paginated
            version_enable_export_pdf,  # Export version-specific enable_export_pdf
            version_enable_export_excel,  # Export version-specific enable_export_excel
            version_enable_import_excel,  # Export version-specific enable_import_excel
            version_enable_ai_validation,  # Export version-specific enable_ai_validation
            json.dumps(version_name_translations) if version_name_translations else None,  # Export version-specific translations (or template translations as fallback)
            json.dumps(version_variables) if version_variables else None,  # Export version-specific template variables
        ]

        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=2, column=col_idx, value=value)

        # Auto-size columns
        cls._auto_size_columns(sheet, len(headers))

        # Add data validation dropdowns
        add_to_self_report_col = headers.index('add_to_self_report') + 1
        cls._add_dropdown_validation(sheet, add_to_self_report_col, cls.DROPDOWN_OPTIONS['add_to_self_report'],
                                    start_row=2, end_row=2)

        display_order_visible_col = headers.index('display_order_visible') + 1
        cls._add_dropdown_validation(sheet, display_order_visible_col, cls.DROPDOWN_OPTIONS['display_order_visible'],
                                    start_row=2, end_row=2)

        is_paginated_col = headers.index('is_paginated') + 1
        cls._add_dropdown_validation(sheet, is_paginated_col, cls.DROPDOWN_OPTIONS['is_paginated'],
                                    start_row=2, end_row=2)

        enable_export_pdf_col = headers.index('enable_export_pdf') + 1
        cls._add_dropdown_validation(sheet, enable_export_pdf_col, cls.DROPDOWN_OPTIONS['enable_export_pdf'],
                                    start_row=2, end_row=2)

        enable_export_excel_col = headers.index('enable_export_excel') + 1
        cls._add_dropdown_validation(sheet, enable_export_excel_col, cls.DROPDOWN_OPTIONS['enable_export_excel'],
                                    start_row=2, end_row=2)

        enable_import_excel_col = headers.index('enable_import_excel') + 1
        cls._add_dropdown_validation(sheet, enable_import_excel_col, cls.DROPDOWN_OPTIONS['enable_import_excel'],
                                    start_row=2, end_row=2)

        enable_ai_validation_col = headers.index('enable_ai_validation') + 1
        cls._add_dropdown_validation(sheet, enable_ai_validation_col, cls.DROPDOWN_OPTIONS['enable_ai_validation'],
                                    start_row=2, end_row=2)

        # Create Excel table
        cls._create_excel_table(sheet, "TemplateTable", len(headers), 2)  # 1 header row + 1 data row

    @classmethod
    def _export_pages_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export pages to Pages sheet with sequential IDs."""
        sheet = workbook.create_sheet("Pages")

        # Get pages for this version (include archived)
        pages = FormPage.query.filter_by(
            template_id=template.id, version_id=version.id
        ).order_by(FormPage.order).all()

        # Write headers with required/optional styling
        headers = cls.PAGE_COLUMNS
        required_cols = cls.REQUIRED_COLUMNS.get('Pages', [])
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            is_required = header in required_cols
            cls._style_header_cell(cell, is_required=is_required)

        # Create mapping: sequential export ID -> database ID
        page_export_id_map = {}  # export_id -> db_id

        # Write data rows with sequential IDs
        for row_idx, page in enumerate(pages, start=2):
            export_id = row_idx - 1  # Sequential ID starting from 1
            page_export_id_map[export_id] = page.id  # Store mapping for reference

            row_data = [
                export_id,  # Sequential export ID (1, 2, 3...)
                page.name,
                page.order,
                json.dumps(page.name_translations) if page.name_translations else None,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Store mapping in sheet for reference (not visible, but can be used if needed)
        sheet._page_export_id_map = page_export_id_map

        # Auto-size columns
        cls._auto_size_columns(sheet, len(headers))

        # Create Excel table (always create, even with 0 rows)
        num_data_rows = len(pages)
        cls._create_excel_table(sheet, "PagesTable", len(headers), num_data_rows + 1)  # +1 for header

        # Add conditional formatting for duplicate IDs (only if there are data rows)
        if num_data_rows > 0:
            cls._add_duplicate_highlighting(sheet, 'id', headers, num_data_rows + 1)

    @classmethod
    def _export_sections_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export sections to Sections sheet with sequential IDs."""
        sheet = workbook.create_sheet("Sections")

        # Get sections for this version (include archived)
        sections = FormSection.query.filter_by(
            template_id=template.id, version_id=version.id
        ).order_by(FormSection.order).all()

        # Get page export ID mapping from Pages sheet
        pages_sheet = workbook['Pages']
        page_db_to_export = {}  # db_id -> export_id
        if hasattr(pages_sheet, '_page_export_id_map'):
            # Reverse the map: export_id -> db_id becomes db_id -> export_id
            for exp_id, db_id in pages_sheet._page_export_id_map.items():
                page_db_to_export[db_id] = exp_id

        # Build section database ID to sequential export ID mapping
        section_db_to_export = {}  # db_id -> export_id

        # Build item DB -> export mapping for rewriting relevance/validation rules.
        # This must match the Items sheet export IDs (which are sequential as well).
        item_db_to_export = cls._build_item_db_to_export_map(template, version)

        # First pass: build mapping
        for row_idx, section in enumerate(sections, start=2):
            export_id = row_idx - 1  # Sequential ID starting from 1
            section_db_to_export[section.id] = export_id

        # Write headers with required/optional styling
        headers = cls.SECTION_COLUMNS
        required_cols = cls.REQUIRED_COLUMNS.get('Sections', [])
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            is_required = header in required_cols
            cls._style_header_cell(cell, is_required=is_required)

        # Second pass: write data rows with sequential IDs and mapped references
        for row_idx, section in enumerate(sections, start=2):
            export_id = section_db_to_export[section.id]

            # Map page_id to export ID
            page_export_id = None
            if section.page_id and section.page_id in page_db_to_export:
                page_export_id = page_db_to_export[section.page_id]

            # Map parent_section_id to export ID
            parent_export_id = None
            if section.parent_section_id and section.parent_section_id in section_db_to_export:
                parent_export_id = section_db_to_export[section.parent_section_id]

            row_data = [
                export_id,  # Sequential export ID (1, 2, 3...)
                section.name,
                section.order,
                parent_export_id,  # Use export ID instead of database ID
                page_export_id,  # Use export ID instead of database ID
                section.section_type,
                section.max_dynamic_indicators,
                json.dumps(section.allowed_sectors) if section.allowed_sectors else None,
                json.dumps(section.indicator_filters) if section.indicator_filters else None,
                section.allow_data_not_available,
                section.allow_not_applicable,
                json.dumps(section.allowed_disaggregation_options) if section.allowed_disaggregation_options else None,
                json.dumps(section.data_entry_display_filters) if section.data_entry_display_filters else None,
                section.add_indicator_note,
                json.dumps(section.name_translations) if section.name_translations else None,
                cls._rewrite_rule_json_item_ids(section.relevance_condition, item_db_to_export),
                section.archived,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Store mapping for items sheet (export_id -> db_id)
        sheet._section_export_id_map = {exp_id: db_id for db_id, exp_id in section_db_to_export.items()}

        # Auto-size columns
        cls._auto_size_columns(sheet, len(headers))

        # Add data validation dropdowns
        section_type_col = headers.index('section_type') + 1  # +1 because Excel is 1-indexed
        cls._add_dropdown_validation(sheet, section_type_col, cls.DROPDOWN_OPTIONS['section_type'],
                                    start_row=2, end_row=len(sections) + 1)

        # Add boolean dropdowns
        archived_col = headers.index('archived') + 1
        cls._add_dropdown_validation(sheet, archived_col, cls.DROPDOWN_OPTIONS['archived'],
                                    start_row=2, end_row=len(sections) + 1)

        allow_data_not_available_col = headers.index('allow_data_not_available') + 1
        cls._add_dropdown_validation(sheet, allow_data_not_available_col, cls.DROPDOWN_OPTIONS['allow_data_not_available'],
                                    start_row=2, end_row=len(sections) + 1)

        allow_not_applicable_col = headers.index('allow_not_applicable') + 1
        cls._add_dropdown_validation(sheet, allow_not_applicable_col, cls.DROPDOWN_OPTIONS['allow_not_applicable'],
                                    start_row=2, end_row=len(sections) + 1)

        # Add dropdown for page_id that references Pages sheet (if Pages sheet exists)
        if 'Pages' in workbook.sheetnames:
            page_id_col = headers.index('page_id') + 1
            pages_sheet = workbook['Pages']
            cls._add_sheet_reference_dropdown(sheet, page_id_col, pages_sheet, 'id',
                                             start_row=2, end_row=len(sections) + 1)

        # Create Excel table (always create, even with 0 rows)
        num_data_rows = len(sections)
        cls._create_excel_table(sheet, "SectionsTable", len(headers), num_data_rows + 1)  # +1 for header

        # Add conditional formatting for duplicate IDs (only if there are data rows)
        if num_data_rows > 0:
            cls._add_duplicate_highlighting(sheet, 'id', headers, num_data_rows + 1)

    @classmethod
    def _export_items_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export items to Items sheet with sequential IDs."""
        sheet = workbook.create_sheet("Items")

        # Get items for this version (include archived)
        items = cls._get_items_for_version(template, version)

        # Build DB -> export mapping for rewriting rule JSON inside exported strings
        item_db_to_export = {item.id: idx + 1 for idx, item in enumerate(items)}

        # Get section export ID mapping from Sections sheet
        sections_sheet = workbook['Sections']
        section_db_to_export = {}  # db_id -> export_id
        if hasattr(sections_sheet, '_section_export_id_map'):
            # Reverse the map: export_id -> db_id becomes db_id -> export_id
            for exp_id, db_id in sections_sheet._section_export_id_map.items():
                section_db_to_export[db_id] = exp_id

        # Write headers with required/optional styling
        headers = cls.ITEM_COLUMNS
        required_cols = cls.REQUIRED_COLUMNS.get('Items', [])
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            is_required = header in required_cols
            cls._style_header_cell(cell, is_required=is_required)

        # Write data rows with sequential IDs
        for row_idx, item in enumerate(items, start=2):
            export_id = row_idx - 1  # Sequential ID starting from 1

            # Map section_id to export ID
            section_export_id = None
            if item.section_id and item.section_id in section_db_to_export:
                section_export_id = section_db_to_export[item.section_id]

            row_data = [
                export_id,  # Sequential export ID (1, 2, 3...)
                section_export_id,  # Use export ID instead of database ID
                item.item_type,
                item.label,
                item.order,
                cls._rewrite_rule_json_item_ids(item.relevance_condition, item_db_to_export),
                item.archived,
                json.dumps(item.config) if item.config else None,
                item.indicator_bank_id,
                item.type,
                item.unit,
                cls._rewrite_rule_json_item_ids(item.validation_condition, item_db_to_export),
                item.validation_message,
                item.definition,
                json.dumps(item.options_json) if item.options_json else None,
                item.lookup_list_id,
                item.list_display_column,
                json.dumps(item.list_filters_json) if item.list_filters_json else None,
                json.dumps(item.label_translations) if item.label_translations else None,
                json.dumps(item.definition_translations) if item.definition_translations else None,
                json.dumps(item.options_translations) if item.options_translations else None,
                json.dumps(item.description_translations) if item.description_translations else None,
                item.description,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns
        cls._auto_size_columns(sheet, len(headers))

        # Add data validation dropdowns
        item_type_col = headers.index('item_type') + 1
        item_type_options = cls._get_item_type_options()
        cls._add_dropdown_validation(sheet, item_type_col, item_type_options,
                                    start_row=2, end_row=len(items) + 1)

        archived_col = headers.index('archived') + 1
        cls._add_dropdown_validation(sheet, archived_col, cls.DROPDOWN_OPTIONS['archived'],
                                    start_row=2, end_row=len(items) + 1)

        # Add dropdown for type column with dynamic values from database
        if 'type' in headers:
            type_col = headers.index('type') + 1
            type_options = cls._get_type_options_from_database()
            if type_options:
                cls._add_dropdown_validation(sheet, type_col, type_options,
                                            start_row=2, end_row=len(items) + 1)

        # Add dropdown for section_id that references Sections sheet
        if 'Sections' in workbook.sheetnames:
            section_id_col = headers.index('section_id') + 1
            sections_sheet = workbook['Sections']
            cls._add_sheet_reference_dropdown(sheet, section_id_col, sections_sheet, 'id',
                                             start_row=2, end_row=len(items) + 1)

        # Create Excel table (always create, even with 0 rows)
        num_data_rows = len(items)
        cls._create_excel_table(sheet, "ItemsTable", len(headers), num_data_rows + 1)  # +1 for header

        # Add conditional formatting for duplicate IDs (only if there are data rows)
        if num_data_rows > 0:
            cls._add_duplicate_highlighting(sheet, 'id', headers, num_data_rows + 1)

    @classmethod
    def _export_metadata_sheet(cls, workbook, template: FormTemplate, version: FormTemplateVersion):
        """Export very hidden metadata sheet with system information."""
        sheet = workbook.create_sheet("_Metadata")
        sheet.sheet_state = 'veryHidden'  # Very hidden - cannot be unhidden via Excel UI

        # Metadata information
        metadata = {
            'Excel Export Version': cls.EXCEL_EXPORT_VERSION,
            'Export Timestamp': utcnow().isoformat(),
            'Template ID': template.id,
            'Template Name': template.name,
            'Version ID': version.id,
            'Version Number': version.version_number,
            'Version Status': version.status,
            'Exported By': current_user.email if current_user else 'System',
        }

        # Write metadata as key-value pairs
        row = 1
        for key, value in metadata.items():
            sheet.cell(row=row, column=1, value=key)
            sheet.cell(row=row, column=2, value=value)
            # Style the key column
            key_cell = sheet.cell(row=row, column=1)
            key_cell.font = Font(bold=True)
            row += 1

        # Auto-size columns
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 50

    @classmethod
    def _export_instructions_sheet(cls, workbook):
        """Export instructions sheet with formatting guide."""
        sheet = workbook.create_sheet("Instructions", 0)  # Insert at position 0 (first sheet)

        # Title
        title_cell = sheet.cell(row=1, column=1, value="Template Excel Import/Export Instructions")
        title_cell.font = Font(bold=True, size=16, color=cls.IFRC_COLORS['WHITE'])
        title_cell.fill = PatternFill(start_color=cls.IFRC_COLORS['DARK_RED'],
                                     end_color=cls.IFRC_COLORS['DARK_RED'],
                                     fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells('A1:D1')
        sheet.row_dimensions[1].height = 35

        # Section: Overview
        row = 3
        section_rows = []  # Track section header rows for border styling
        overview_cell = sheet.cell(row=row, column=1, value="📋 Overview")
        overview_cell.font = Font(bold=True, size=14, color=cls.IFRC_COLORS['DARK_RED'])
        overview_cell.fill = PatternFill(start_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                        end_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                        fill_type='solid')
        sheet.merge_cells(f'A{row}:D{row}')
        section_rows.append(row)
        row += 1

        instructions = [
            "This Excel file contains the structure of a form template that can be imported back into the system.",
            "",
            "The file contains the following sheets:",
            "  • Instructions (this sheet) - Ignored during import",
            "  • Template - Template metadata and configuration",
            "  • Pages - Page definitions (if template is paginated)",
            "  • Sections - Section definitions",
            "  • Items - Form items (indicators, questions, document fields, etc.)",
            "",
            "Any sheets other than the recognized ones (Template, Pages, Sections, Items) will be ignored during import.",
        ]

        for instruction in instructions:
            cell = sheet.cell(row=row, column=1, value=instruction)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if instruction.startswith("  •"):
                cell.font = Font(size=10)
            row += 1

        # Section: Column Headers
        row += 2
        headers_cell = sheet.cell(row=row, column=1, value="🎨 Column Headers")
        headers_cell.font = Font(bold=True, size=14, color=cls.IFRC_COLORS['DARK_RED'])
        headers_cell.fill = PatternFill(start_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                       end_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                       fill_type='solid')
        sheet.merge_cells(f'A{row}:D{row}')
        section_rows.append(row)
        row += 1

        header_instructions = [
            "Column headers in each sheet are color-coded to indicate whether they are required or optional:",
            "",
            "  🔴 RED HEADERS (Required): These columns must have values. They are essential for the template to function.",
            "     Missing values in required columns will cause the import to fail.",
            "",
            "  🔵 BLUE HEADERS (Optional): These columns can be left empty. They provide additional configuration or metadata.",
            "     Empty optional columns will use default values or be set to NULL.",
            "",
            "Required columns for each sheet:",
        ]

        for instruction in header_instructions:
            cell = sheet.cell(row=row, column=1, value=instruction)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            row += 1

        # List required columns for each sheet
        for sheet_name in ['Template', 'Pages', 'Sections', 'Items']:
            required = cls.REQUIRED_COLUMNS.get(sheet_name, [])
            if required:
                cell = sheet.cell(row=row, column=1, value=f"  {sheet_name}: {', '.join(required)}")
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                row += 1

        # Section: Important Notes
        row += 2
        notes_cell = sheet.cell(row=row, column=1, value="⚠️ Important Notes")
        notes_cell.font = Font(bold=True, size=14, color=cls.IFRC_COLORS['DARK_RED'])
        notes_cell.fill = PatternFill(start_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                     end_color=cls.IFRC_COLORS['LIGHT_GRAY'],
                                     fill_type='solid')
        sheet.merge_cells(f'A{row}:D{row}')
        section_rows.append(row)
        row += 1

        notes = [
            "1. ID Fields: The 'id' columns in Pages, Sections, and Items are sequential export IDs (1, 2, 3...).",
            "   They are NOT database IDs. The system will create new IDs when importing.",
            "",
            "2. References: When referencing other records (e.g., section_id in Items), use the export IDs from the same file.",
            "",
            "3. JSON Fields: Fields containing JSON (like name_translations, config, options_json) should be valid JSON strings.",
            "   Use double quotes for keys and string values.",
            "",
            "4. Boolean Fields: Use TRUE/FALSE or 1/0 for boolean values.",
            "",
            "5. Order Fields: Use numeric values (integers or decimals like 1, 1.1, 1.2) to control display order.",
            "",
            "6. Import Behavior:",
            "   • If importing into a published version, a new draft version will be created automatically.",
            "   • Existing items/sections with matching order+name will be updated, others will be created.",
            "   • The import will fail if required columns are missing or invalid.",
            "",
            "7. Do NOT modify the Instructions sheet or add unrecognized sheets if you plan to re-import the file.",
        ]

        for note in notes:
            cell = sheet.cell(row=row, column=1, value=note)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if note and note[0].isdigit():
                cell.font = Font(bold=True, size=10)
            row += 1

        # Auto-size columns
        sheet.column_dimensions['A'].width = 100
        sheet.column_dimensions['B'].width = 5
        sheet.column_dimensions['C'].width = 5
        sheet.column_dimensions['D'].width = 5

        # Add borders to section headers for better visual separation
        for section_row in section_rows:
            for col in ['A', 'B', 'C', 'D']:
                cell = sheet[f'{col}{section_row}']
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='medium'),
                    bottom=Side(style='thin')
                )

        # Freeze first row
        sheet.freeze_panes = 'A2'

        # Protect the instructions sheet (read-only, but allow formatting)
        sheet.protection.sheet = True
        sheet.protection.formatCells = False
        sheet.protection.formatColumns = False
        sheet.protection.formatRows = False

    @classmethod
    def _create_excel_table(cls, sheet, table_name: str, num_columns: int, num_rows: int):
        """Create an Excel table from the data range.

        Args:
            sheet: The worksheet to add the table to
            table_name: Name for the table
            num_columns: Number of columns
            num_rows: Number of rows (including header)
            Note: Excel requires at least 2 rows (header + 1 data row), so if num_rows=1, we create with 2 rows
        """
        if num_rows < 1:  # Need at least header row
            return

        # Excel requires at least 2 rows for a table (header + at least one data row)
        # If we only have a header, create table with header + 1 empty data row
        if num_rows == 1:
            num_rows = 2  # Add one empty data row

        # Calculate range (A1 to last column, last row)
        start_col = get_column_letter(1)
        end_col = get_column_letter(num_columns)
        table_range = f"{start_col}1:{end_col}{num_rows}"

        # Create table
        table = Table(displayName=table_name, ref=table_range)

        # Style the table
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style

        # Add table to sheet
        sheet.add_table(table)

    @classmethod
    def _add_dropdown_validation(cls, sheet, column: int, options: List[str],
                                 start_row: int = 2, end_row: int = 1000):
        """Add data validation dropdown to a column.

        Args:
            sheet: The worksheet
            column: Column number (1-indexed)
            options: List of options for the dropdown
            start_row: Starting row (default 2, after header)
            end_row: Ending row
        """
        if not options:
            return

        # Create comma-separated list of options
        # Excel list validation formula format: "option1,option2,option3"
        options_str = ','.join(options)
        formula = f'"{options_str}"'

        # Create data validation
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = 'Invalid value'
        dv.errorTitle = 'Invalid Entry'
        dv.prompt = 'Please select from the dropdown'
        dv.promptTitle = 'Select Value'

        # Apply to column range
        col_letter = get_column_letter(column)
        dv_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
        dv.add(dv_range)

        # Add validation to sheet
        sheet.add_data_validation(dv)

    @classmethod
    def _add_sheet_reference_dropdown(cls, sheet, column: int, source_sheet,
                                      source_column_name: str,
                                      start_row: int = 2, end_row: int = 1000):
        """Add data validation dropdown that references values from another sheet using table column reference.

        Uses structured table references (e.g., PagesTable[id]) which automatically expand when new rows are added.

        Args:
            sheet: The worksheet to add validation to
            column: Column number (1-indexed) in the target sheet
            source_sheet: The source worksheet to reference
            source_column_name: Name of the column in source sheet to reference
            start_row: Starting row (default 2, after header)
            end_row: Ending row
        """
        try:
            # Find the source column index in the source sheet
            source_headers = [cell.value for cell in source_sheet[1]]
            if source_column_name not in source_headers:
                current_app.logger.warning(f"Source column '{source_column_name}' not found in source sheet")
                return

            source_col_idx = source_headers.index(source_column_name) + 1
            source_col_letter = get_column_letter(source_col_idx)

            # Prepare sheet reference
            source_sheet_name = source_sheet.title
            # Escape sheet name if it contains spaces or special characters
            if ' ' in source_sheet_name or '-' in source_sheet_name:
                source_sheet_ref = f"'{source_sheet_name}'"
            else:
                source_sheet_ref = source_sheet_name

            # Use OFFSET with COUNTA to create a dynamic range that only includes non-empty cells
            # OFFSET(start_cell, rows, cols, height, width)
            # COUNTA counts all non-empty cells in the column, subtract 1 to exclude header row
            # This creates a range that automatically expands/contracts based on actual data
            # Format: OFFSET(Pages!$A$2,0,0,COUNTA(Pages!$A:$A)-1,1)
            source_range = f"OFFSET({source_sheet_ref}!${source_col_letter}$2,0,0,COUNTA({source_sheet_ref}!${source_col_letter}:${source_col_letter})-1,1)"

            current_app.logger.debug(f"Creating dropdown with OFFSET formula: {source_range}")

            # Create data validation with reference formula
            dv = DataValidation(type="list", formula1=source_range, allow_blank=True)
            # Generic error messages since this is used for both page_id and section_id
            dv.error = f'Invalid {source_column_name}. Please select from the dropdown.'
            dv.errorTitle = 'Invalid Entry'
            dv.prompt = f'Please select a {source_column_name} from the {source_sheet.title} sheet'
            dv.promptTitle = f'Select {source_column_name.replace("_", " ").title()}'

            # Apply to column range
            col_letter = get_column_letter(column)
            dv_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
            dv.add(dv_range)

            # Add validation to sheet
            sheet.add_data_validation(dv)

            current_app.logger.info(f"Successfully added sheet reference dropdown: {source_range} for column {col_letter}, range {dv_range}")

        except Exception as e:
            current_app.logger.error(f"Could not add sheet reference dropdown: {e}", exc_info=True)

    @classmethod
    def _get_item_type_options(cls) -> List[str]:
        """Get all supported item types including plugin types.

        Returns:
            List of all item types: standard types (indicator, question, document_field, matrix)
            plus all active plugin field types (prefixed with 'plugin_')

        Note: These match the hardcoded values in form_builder.html dropdown:
        - indicator (Indicator from Bank)
        - question (New Question)
        - document_field (Document Field - stored as document_field in DB, shown as 'document' in UI)
        - matrix (Matrix Table)
        - plugin_* (dynamically added from plugin system)
        """
        item_types = []

        # Add standard item types - these are the hardcoded fallback values currently in the system
        # Matching the form_builder.html dropdown options (lines 1624-1627)
        # Note: UI shows 'document' but DB stores 'document_field', so we use 'document_field' here
        standard_types = ['indicator', 'question', 'document_field', 'matrix']
        item_types.extend(standard_types)

        # Add plugin field types if plugin manager is available
        try:
            if hasattr(current_app, 'plugin_manager'):
                plugin_manager = current_app.plugin_manager
                active_field_types = plugin_manager.list_active_field_types()

                # Prefix each plugin field type with 'plugin_'
                for field_type in active_field_types:
                    plugin_item_type = f'plugin_{field_type}'
                    if plugin_item_type not in item_types:
                        item_types.append(plugin_item_type)
        except Exception as e:
            current_app.logger.warning(f"Could not fetch plugin field types: {e}")

        # Sort for consistent ordering
        return sorted(item_types)

    @classmethod
    def _get_type_options_from_database(cls) -> List[str]:
        """Get distinct type values from FormItem table dynamically.

        Returns:
            List of unique type values found in the database, sorted alphabetically
        """
        try:
            # Query distinct non-null type values from FormItem
            distinct_types = db.session.query(FormItem.type).filter(
                FormItem.type.isnot(None),
                FormItem.type != ''
            ).distinct().all()

            # Extract values and sort
            type_options = sorted([t[0] for t in distinct_types if t[0]])

            # If no types found in database, return common defaults
            if not type_options:
                type_options = ['Number', 'Percentage', 'Text', 'Boolean', 'Date', 'Choice']

            return type_options

        except Exception as e:
            current_app.logger.warning(f"Could not fetch type options from database: {e}")
            # Return common defaults as fallback
            return ['Number', 'Percentage', 'Text', 'Boolean', 'Date', 'Choice']

    @classmethod
    def _add_duplicate_highlighting(cls, sheet, column_name: str, headers: List[str],
                                     num_rows: int):
        """Add conditional formatting to highlight duplicate values in a column.

        Args:
            sheet: The worksheet
            column_name: Name of the column to check for duplicates
            headers: List of header names (to find column index)
            num_rows: Total number of rows (including header)
        """
        if num_rows < 3:  # Need at least header + 2 data rows to have duplicates
            return

        try:
            # Find column index
            if column_name not in headers:
                return

            col_idx = headers.index(column_name) + 1  # +1 because Excel is 1-indexed
            col_letter = get_column_letter(col_idx)

            # Create range for data rows (skip header row)
            data_range = f"{col_letter}2:{col_letter}{num_rows}"

            # Create duplicate values rule with yellow background
            # Use COUNTIF formula to detect duplicates: COUNTIF($A$2:$A$100, A2)>1
            # This formula checks if the current cell value appears more than once in the range
            duplicate_fill = PatternFill(start_color=cls.IFRC_COLORS['YELLOW'],
                                        end_color=cls.IFRC_COLORS['YELLOW'],
                                        fill_type='solid')

            # Formula: COUNTIF(absolute_range, relative_cell) > 1
            # $col_letter$2:$col_letter$num_rows is the absolute range
            # col_letter2 is the relative cell reference (will adjust per row)
            formula = f'COUNTIF(${col_letter}$2:${col_letter}${num_rows},{col_letter}2)>1'

            duplicate_rule = FormulaRule(formula=[formula], fill=duplicate_fill)

            # Add conditional formatting rule
            sheet.conditional_formatting.add(data_range, duplicate_rule)

        except Exception as e:
            # Log but don't fail if conditional formatting can't be added
            current_app.logger.warning(f"Could not add duplicate highlighting for column {column_name}: {e}")

    @classmethod
    def _style_header_cell(cls, cell, is_required=True):
        """Apply IFRC styling to header cell.

        Args:
            cell: The cell to style
            is_required: If True, style as required (red), if False, style as optional (blue)
        """
        cell.font = Font(bold=True, color=cls.IFRC_COLORS['WHITE'])
        # Use red for required columns, blue for optional columns
        fill_color = cls.IFRC_COLORS['DARK_RED'] if is_required else cls.IFRC_COLORS['DARK_BLUE']
        cell.fill = PatternFill(start_color=fill_color,
                               end_color=fill_color,
                               fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    @classmethod
    def _auto_size_columns(cls, sheet, num_columns):
        """Auto-size columns for better readability."""
        for col_idx in range(1, num_columns + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in sheet[column_letter]:
                with suppress(Exception):
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            sheet.column_dimensions[column_letter].width = adjusted_width

    @classmethod
    @memory_tracker("Template Excel Import", log_top_allocations=True)
    def import_template(cls, template_id: int, excel_file, version_id: Optional[int] = None) -> Dict[str, Any]:
        """
        Import template structure from Excel.

        Args:
            template_id: Template ID to import into
            excel_file: File-like object containing Excel file
            version_id: Optional version ID to import into (defaults to active version or creates draft)

        Returns:
            Dict with 'success', 'message', 'errors', 'created_count' keys
        """
        current_app.logger.info(f"=== TEMPLATE EXCEL IMPORT START ===")
        current_app.logger.info(f"Template ID: {template_id}, Version ID: {version_id}")

        template = FormTemplate.query.get_or_404(template_id)
        current_app.logger.info(f"Template found: '{template.name}'")

        errors = []
        created_counts = {'pages': 0, 'sections': 0, 'items': 0}

        try:
            # Load workbook
            current_app.logger.info("Loading Excel workbook...")
            workbook = openpyxl.load_workbook(io.BytesIO(excel_file.read()), data_only=True)
            current_app.logger.info(f"Workbook loaded. Sheets found: {workbook.sheetnames}")

            # Validate required sheets exist (ignore Instructions, _Metadata and any other unrecognized sheets)
            required_sheets = ['Template', 'Pages', 'Sections', 'Items']
            recognized_sheets = set(required_sheets + ['Instructions', '_Metadata'])  # Instructions and _Metadata are recognized but optional
            missing_sheets = [s for s in required_sheets if s not in workbook.sheetnames]

            # Log any unrecognized sheets (they will be ignored)
            unrecognized_sheets = [s for s in workbook.sheetnames if s not in recognized_sheets]
            if unrecognized_sheets:
                current_app.logger.info(f"Ignoring unrecognized sheets during import: {unrecognized_sheets}")

            if missing_sheets:
                current_app.logger.error(f"Missing required sheets: {missing_sheets}")
                return {
                    'success': False,
                    'message': f"Missing required sheets: {', '.join(missing_sheets)}",
                    'errors': errors,
                    'created_count': created_counts
                }

            # Determine target version
            current_app.logger.info("Determining target version...")
            target_version = None
            if version_id:
                target_version = FormTemplateVersion.query.filter_by(
                    id=version_id, template_id=template.id
                ).first()
                if target_version:
                    current_app.logger.info(f"Using specified version ID {version_id} (Status: {target_version.status}, Version #: {target_version.version_number})")

            if not target_version:
                # Use published version or latest as fallback
                if template.published_version_id:
                    target_version = FormTemplateVersion.query.get(template.published_version_id)
                    current_app.logger.info(f"Using published version ID {template.published_version_id}")
                else:
                    target_version = FormTemplateVersion.query.filter_by(
                        template_id=template.id
                    ).order_by(FormTemplateVersion.created_at.desc()).first()
                    if target_version:
                        current_app.logger.info(f"Using latest version ID {target_version.id} (Status: {target_version.status})")

            if not target_version:
                current_app.logger.error("No version found for template")
                return {
                    'success': False,
                    'message': "No version found for this template. Please create a version first.",
                    'errors': errors,
                    'created_count': created_counts
                }

            # If target version is published, create a new draft version
            if target_version.status == 'published':
                current_app.logger.info(f"Target version is published. Creating new draft version...")

                # Get the next version number
                from sqlalchemy import func
                max_version = db.session.query(func.max(FormTemplateVersion.version_number)).filter_by(
                    template_id=template.id
                ).scalar()
                next_version_number = (max_version + 1) if max_version else 1

                # Create new draft version
                new_draft = FormTemplateVersion(
                    template_id=template.id,
                    version_number=next_version_number,
                    status='draft',
                    created_by=current_user.id,
                    updated_by=current_user.id,
                    based_on_version_id=target_version.id,
                    comment='Created automatically for Excel import',
                    name=target_version.name,
                    name_translations=target_version.name_translations.copy() if target_version.name_translations else None,
                    description=target_version.description,
                    add_to_self_report=target_version.add_to_self_report,
                    display_order_visible=target_version.display_order_visible,
                    is_paginated=target_version.is_paginated,
                    enable_export_pdf=target_version.enable_export_pdf,
                    enable_export_excel=target_version.enable_export_excel,
                    enable_import_excel=target_version.enable_import_excel,
                    enable_ai_validation=getattr(target_version, 'enable_ai_validation', False),
                    variables=target_version.variables.copy() if target_version.variables else None
                )
                db.session.add(new_draft)
                db.session.flush()

                current_app.logger.info(f"Created new draft version: ID={new_draft.id}, Version #={new_draft.version_number}")

                # Clone structure from published version to draft
                cls._clone_template_structure(template.id, target_version.id, new_draft.id)

                # Use the new draft as target
                target_version = new_draft
                current_app.logger.info(f"Using new draft version: ID={target_version.id}, Version #={target_version.version_number}")

                # Update version_id to return the new draft version ID
                version_id = new_draft.id

            current_app.logger.info(f"Target version: ID={target_version.id}, Status={target_version.status}, Version #={target_version.version_number}")

            # Import Template metadata first (including version-specific name)
            current_app.logger.info("=== IMPORTING TEMPLATE METADATA ===")
            template_sheet = workbook['Template']
            template_errors = cls._import_template_metadata(template_sheet, template, target_version)
            errors.extend(template_errors)
            if template_errors:
                current_app.logger.warning(f"Template metadata import errors: {template_errors}")
            else:
                current_app.logger.info("Template metadata imported successfully")

            # ID mapping dictionaries (export ID -> new database ID)
            # Export IDs are sequential (1, 2, 3...) and map to new database IDs
            page_id_map = {}  # export_id -> new_db_id
            section_id_map = {}  # export_id -> new_db_id

            # Import Pages
            current_app.logger.info("=== IMPORTING PAGES ===")
            pages_sheet = workbook['Pages']
            page_errors = cls._import_pages(pages_sheet, template, target_version, page_id_map)
            errors.extend(page_errors)
            created_counts['pages'] = len(page_id_map)
            current_app.logger.info(f"Pages imported: {created_counts['pages']} pages created")
            current_app.logger.info(f"Page ID mapping: {page_id_map}")
            if page_errors:
                current_app.logger.warning(f"Page import errors: {page_errors}")

            # Import Sections
            current_app.logger.info("=== IMPORTING SECTIONS ===")
            sections_sheet = workbook['Sections']
            section_errors = cls._import_sections(
                sections_sheet, template, target_version, page_id_map, section_id_map
            )
            errors.extend(section_errors)
            created_counts['sections'] = len(section_id_map)
            current_app.logger.info(f"Sections imported: {created_counts['sections']} sections created")
            current_app.logger.info(f"Section ID mapping: {section_id_map}")
            if section_errors:
                current_app.logger.warning(f"Section import errors: {section_errors}")

            # Import Items
            current_app.logger.info("=== IMPORTING ITEMS ===")
            items_sheet = workbook['Items']
            item_errors = cls._import_items(
                items_sheet, template, target_version, section_id_map
            )
            errors.extend(item_errors)
            created_counts['items'] = db.session.query(FormItem).filter_by(
                template_id=template.id, version_id=target_version.id
            ).count()
            current_app.logger.info(f"Items imported: {created_counts['items']} items created")
            if item_errors:
                current_app.logger.warning(f"Item import errors: {item_errors}")

            if errors:
                current_app.logger.error(f"Import completed with {len(errors)} errors. Rolling back...")
                db.session.rollback()
                return {
                    'success': False,
                    'message': f"Import completed with {len(errors)} errors",
                    'errors': errors,
                    'created_count': created_counts
                }

            # Update version timestamp
            current_app.logger.info("Updating version timestamp...")
            target_version.updated_at = utcnow()
            target_version.updated_by = current_user.id

            current_app.logger.info("Committing database changes...")
            db.session.commit()
            current_app.logger.info("=== TEMPLATE EXCEL IMPORT SUCCESS ===")
            current_app.logger.info(f"Final counts - Pages: {created_counts['pages']}, Sections: {created_counts['sections']}, Items: {created_counts['items']}")

            return {
                'success': True,
                'message': f"Successfully imported {created_counts['pages']} pages, "
                          f"{created_counts['sections']} sections, "
                          f"{created_counts['items']} items",
                'errors': [],
                'created_count': created_counts,
                'version_id': version_id  # Return the version ID (may be new draft if published was selected)
            }

        except Exception as e:
            current_app.logger.error(f"=== TEMPLATE EXCEL IMPORT FAILED ===")
            current_app.logger.error(f"Error importing template from Excel: {e}", exc_info=True)
            db.session.rollback()
            return {
                'success': False,
                'message': "Error importing template. Check the file format and try again.",
                'errors': ['Import failed. See logs for details.'],
                'created_count': created_counts
            }

    @classmethod
    def _import_template_metadata(cls, sheet, template: FormTemplate, version: FormTemplateVersion) -> List[str]:
        """Import template metadata from Template sheet, including version-specific name."""
        current_app.logger.info("Starting template metadata import...")
        errors = []

        # Read headers
        headers = [cell.value for cell in sheet[1]]
        current_app.logger.info(f"Template sheet headers: {headers}")

        # Validate headers match expected columns.
        # Backwards-compatible: allow extra columns (e.g. legacy 'template_type') and ignore them.
        # Only require columns from REQUIRED_COLUMNS; all others (including 'variables') are optional.
        required_headers = cls.REQUIRED_COLUMNS.get('Template', ['name'])
        header_set = {h for h in headers if isinstance(h, str) and h}
        required_set = set(required_headers)

        # Check if all required columns are present
        if not required_set.issubset(header_set):
            missing = required_set - header_set
            error_msg = f"Template sheet headers missing required columns. Required: {required_headers}, Missing: {list(missing)}, Got: {headers}"
            current_app.logger.error(error_msg)
            errors.append(error_msg)
            return errors

        # Log if there are extra columns (legacy fields) or missing optional columns
        expected_headers = cls.TEMPLATE_COLUMNS
        expected_set = set(expected_headers)
        extra_columns = header_set - expected_set
        missing_optional = expected_set - header_set - required_set

        if extra_columns:
            current_app.logger.info(
                f"Template sheet contains extra/legacy columns (will be ignored): {list(extra_columns)}"
            )
        if missing_optional:
            current_app.logger.info(
                f"Template sheet missing optional columns (will use defaults): {list(missing_optional)}"
            )

        # Read first data row (should only be one row)
        row = next(sheet.iter_rows(min_row=2, values_only=True), None)
        if not row or all(cell is None for cell in row):
            current_app.logger.warning("Template sheet has no data row")
            return errors

        try:
            # Map row data to columns (use actual headers, then pick expected fields).
            # This supports legacy files that include extra columns like 'template_type'.
            row_data_all = dict(zip(headers, row))
            row_data = {h: row_data_all.get(h) for h in expected_headers}

            current_app.logger.info(f"Updating template metadata for template ID {template.id}")

            # Update version-specific fields (name is now only stored in versions)
            if 'name' in row_data and row_data['name']:
                version.name = row_data['name']
                current_app.logger.info(f"Updated version name: '{version.name}'")

            if 'description' in row_data:
                version.description = row_data['description']
                current_app.logger.info(f"Updated version/template description")

            if 'add_to_self_report' in row_data:
                version.add_to_self_report = bool(row_data['add_to_self_report'])
                current_app.logger.info(f"Updated version/template add_to_self_report: {version.add_to_self_report}")

            if 'display_order_visible' in row_data:
                version.display_order_visible = bool(row_data['display_order_visible'])
                current_app.logger.info(f"Updated version/template display_order_visible: {version.display_order_visible}")

            if 'is_paginated' in row_data:
                version.is_paginated = bool(row_data['is_paginated'])
                current_app.logger.info(f"Updated version/template is_paginated: {version.is_paginated}")

            if 'enable_export_pdf' in row_data:
                version.enable_export_pdf = bool(row_data['enable_export_pdf'])
                current_app.logger.info(f"Updated version/template enable_export_pdf: {version.enable_export_pdf}")

            if 'enable_export_excel' in row_data:
                version.enable_export_excel = bool(row_data['enable_export_excel'])
                current_app.logger.info(f"Updated version/template enable_export_excel: {version.enable_export_excel}")

            if 'enable_import_excel' in row_data:
                version.enable_import_excel = bool(row_data['enable_import_excel'])
                current_app.logger.info(f"Updated version/template enable_import_excel: {version.enable_import_excel}")

            if 'enable_ai_validation' in row_data:
                version.enable_ai_validation = bool(row_data['enable_ai_validation'])
                current_app.logger.info(f"Updated version/template enable_ai_validation: {version.enable_ai_validation}")

            if 'name_translations' in row_data:
                name_translations = cls._parse_json(row_data.get('name_translations'))
                if name_translations:
                    # Save to version (version-specific translations)
                    version.name_translations = name_translations
                    current_app.logger.info(f"Updated version name_translations")

            if 'variables' in row_data:
                variables = cls._parse_json(row_data.get('variables'))
                if variables is not None:  # Allow empty dict {} to clear variables
                    # Save to version (version-specific template variables)
                    version.variables = variables if variables else {}
                    current_app.logger.info(f"Updated version variables: {len(variables) if variables else 0} variable(s)")

            db.session.add(template)
            db.session.add(version)
            current_app.logger.info("Template metadata import complete")

        except Exception as e:
            current_app.logger.error("Template metadata row: %s", e, exc_info=True)
            errors.append("Template metadata row: Validation error.")

        return errors

    @classmethod
    def _import_pages(cls, sheet, template: FormTemplate, version: FormTemplateVersion,
                     page_id_map: Dict[int, int]) -> List[str]:
        """Import pages from sheet."""
        current_app.logger.info("Starting pages import...")
        errors = []

        # Read headers
        headers = [cell.value for cell in sheet[1]]
        current_app.logger.info(f"Pages sheet headers: {headers}")

        # Validate headers match expected columns
        expected_headers = cls.PAGE_COLUMNS
        if headers != expected_headers:
            error_msg = f"Pages sheet headers don't match expected columns. Expected: {expected_headers}, Got: {headers}"
            current_app.logger.error(error_msg)
            errors.append(error_msg)
            return errors

        current_app.logger.info("Headers validated. Processing page rows...")
        row_count = 0

        # Read data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue  # Skip empty rows

            row_count += 1
            try:
                # Map row data to columns
                row_data = dict(zip(expected_headers, row))

                # Get export ID (sequential ID from Excel, e.g., 1, 2, 3...)
                export_id = int(row_data['id']) if row_data.get('id') else None

                if export_id is None:
                    error_msg = f"Pages row {row_idx}: Missing export ID"
                    current_app.logger.warning(error_msg)
                    errors.append(error_msg)
                    continue

                page_name = row_data['name'] or ''
                page_order = int(row_data['order']) if row_data['order'] is not None else 1
                current_app.logger.info(f"Processing page row {row_idx}: export_id={export_id}, name='{page_name}', order={page_order}")

                # Create new page
                new_page = FormPage(
                    template_id=template.id,
                    version_id=version.id,
                    name=page_name,
                    order=page_order,
                    name_translations=cls._parse_json(row_data.get('name_translations'))
                )

                db.session.add(new_page)
                db.session.flush()

                # Map export ID to new database ID
                page_id_map[export_id] = new_page.id
                current_app.logger.info(f"Page created: export_id={export_id} -> db_id={new_page.id}, name='{page_name}'")

            except Exception as e:
                current_app.logger.error("Pages row %s: %s", row_idx, e, exc_info=True)
                errors.append(f"Pages row {row_idx}: Validation error.")

        current_app.logger.info(f"Pages import complete: {row_count} rows processed, {len(page_id_map)} pages created, {len(errors)} errors")
        return errors

    @classmethod
    def _import_sections(cls, sheet, template: FormTemplate, version: FormTemplateVersion,
                        page_id_map: Dict[int, int], section_id_map: Dict[int, int]) -> List[str]:
        """Import sections from sheet."""
        current_app.logger.info("Starting sections import...")
        current_app.logger.info(f"Page ID mapping available: {len(page_id_map)} pages")
        errors = []

        # Read headers
        headers = [cell.value for cell in sheet[1]]
        current_app.logger.info(f"Sections sheet headers: {headers}")

        # Validate headers
        expected_headers = cls.SECTION_COLUMNS
        if headers != expected_headers:
            error_msg = f"Sections sheet headers don't match expected columns. Expected: {expected_headers}, Got: {headers}"
            current_app.logger.error(error_msg)
            errors.append(error_msg)
            return errors

        # First pass: collect all section data
        current_app.logger.info("First pass: Collecting section data...")
        sections_data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue

            try:
                row_data = dict(zip(expected_headers, row))
                sections_data.append((row_idx, row_data))
            except Exception as e:
                current_app.logger.error("Sections row %s: %s", row_idx, e)
                errors.append(f"Sections row {row_idx}: Validation error.")

        current_app.logger.info(f"Collected {len(sections_data)} section rows to process")

        # Get existing sections for this version (for matching)
        existing_sections = FormSection.query.filter_by(
            template_id=template.id, version_id=version.id
        ).all()
        # Create lookup: (order, name) -> existing section (use order+name as match key)
        existing_sections_by_key = {(s.order, s.name): s for s in existing_sections}

        # Second pass: create or update sections and build ID map
        current_app.logger.info("Second pass: Creating/updating sections...")
        sections_created = 0
        sections_updated = 0

        for row_idx, row_data in sections_data:
            try:
                # Get export ID (sequential ID from Excel)
                export_id = int(row_data['id']) if row_data.get('id') else None

                if export_id is None:
                    error_msg = f"Sections row {row_idx}: Missing export ID"
                    current_app.logger.warning(error_msg)
                    errors.append(error_msg)
                    continue

                # Resolve page_id using export ID mapping
                page_id = None
                page_export_id = None
                if row_data.get('page_id'):
                    page_export_id = int(row_data['page_id']) if row_data.get('page_id') else None
                    if page_export_id and page_export_id in page_id_map:
                        page_id = page_id_map[page_export_id]
                        current_app.logger.debug(f"Section row {row_idx}: Mapped page export_id={page_export_id} -> db_id={page_id}")
                    elif page_export_id:
                        current_app.logger.warning(f"Section row {row_idx}: Page export_id={page_export_id} not found in page_id_map")

                section_name = row_data['name'] or ''
                section_order = float(row_data['order']) if row_data['order'] is not None else 0.0
                section_type = row_data.get('section_type') or 'standard'
                current_app.logger.info(f"Processing section row {row_idx}: export_id={export_id}, name='{section_name}', order={section_order}, type={section_type}, page_export_id={page_export_id}")

                # Check if section already exists (match by order + name)
                section_key = (section_order, section_name)
                existing_section = existing_sections_by_key.get(section_key)

                if existing_section:
                    # Update existing section
                    current_app.logger.info(f"Updating existing section: db_id={existing_section.id}, order={section_order}, name='{section_name}'")
                    existing_section.page_id = page_id
                    existing_section.section_type = section_type
                    existing_section.max_dynamic_indicators = int(row_data['max_dynamic_indicators']) if row_data.get('max_dynamic_indicators') else None
                    existing_section.allowed_sectors = cls._parse_json(row_data.get('allowed_sectors'))
                    existing_section.indicator_filters = cls._parse_json(row_data.get('indicator_filters'))
                    existing_section.allow_data_not_available = bool(row_data.get('allow_data_not_available', False))
                    existing_section.allow_not_applicable = bool(row_data.get('allow_not_applicable', False))
                    existing_section.allowed_disaggregation_options = cls._parse_json(row_data.get('allowed_disaggregation_options'))
                    existing_section.data_entry_display_filters = cls._parse_json(row_data.get('data_entry_display_filters'))
                    existing_section.add_indicator_note = row_data.get('add_indicator_note')
                    existing_section.name_translations = cls._parse_json(row_data.get('name_translations'))
                    existing_section.relevance_condition = row_data.get('relevance_condition')
                    existing_section.archived = bool(row_data.get('archived', False))
                    # Note: parent_section_id will be updated in third pass
                    sections_updated += 1
                    section_id_map[export_id] = existing_section.id
                    current_app.logger.info(f"Section updated: export_id={export_id} -> db_id={existing_section.id}, name='{section_name}'")
                else:
                    # Create new section (parent_section_id will be resolved in third pass)
                    new_section = FormSection(
                        template_id=template.id,
                        version_id=version.id,
                        name=section_name,
                        order=section_order,
                        page_id=page_id,
                        parent_section_id=None,  # Will be set in third pass
                        section_type=section_type,
                        max_dynamic_indicators=int(row_data['max_dynamic_indicators']) if row_data.get('max_dynamic_indicators') else None,
                        allowed_sectors=cls._parse_json(row_data.get('allowed_sectors')),
                        indicator_filters=cls._parse_json(row_data.get('indicator_filters')),
                        allow_data_not_available=bool(row_data.get('allow_data_not_available', False)),
                        allow_not_applicable=bool(row_data.get('allow_not_applicable', False)),
                        allowed_disaggregation_options=cls._parse_json(row_data.get('allowed_disaggregation_options')),
                        data_entry_display_filters=cls._parse_json(row_data.get('data_entry_display_filters')),
                        add_indicator_note=row_data.get('add_indicator_note'),
                        name_translations=cls._parse_json(row_data.get('name_translations')),
                        relevance_condition=row_data.get('relevance_condition'),
                        archived=bool(row_data.get('archived', False))
                    )

                    db.session.add(new_section)
                    db.session.flush()

                    # Map export ID to new database ID
                    section_id_map[export_id] = new_section.id
                    sections_created += 1
                    current_app.logger.info(f"Section created: export_id={export_id} -> db_id={new_section.id}, name='{section_name}'")

            except Exception as e:
                current_app.logger.error("Sections row %s: %s", row_idx, e, exc_info=True)
                errors.append(f"Sections row {row_idx}: Validation error.")

        current_app.logger.info(f"Sections created/updated: {sections_created} created, {sections_updated} updated")

        current_app.logger.info(f"Sections created: {len(section_id_map)} sections")

        # Third pass: update parent_section_id references using export ID mapping
        current_app.logger.info("Third pass: Updating parent_section_id references...")
        parent_updates = 0
        for row_idx, row_data in sections_data:
            if row_data.get('parent_section_id'):
                try:
                    parent_export_id = int(row_data['parent_section_id']) if row_data.get('parent_section_id') else None
                    section_export_id = int(row_data['id']) if row_data.get('id') else None

                    if parent_export_id and section_export_id:
                        # Find new database IDs using export IDs
                        new_section_id = section_id_map.get(section_export_id)
                        new_parent_id = section_id_map.get(parent_export_id)

                        if new_section_id and new_parent_id:
                            # Update parent reference
                            section = FormSection.query.get(new_section_id)
                            if section:
                                section.parent_section_id = new_parent_id
                                parent_updates += 1
                                current_app.logger.debug(f"Updated parent: section export_id={section_export_id} (db_id={new_section_id}) -> parent export_id={parent_export_id} (db_id={new_parent_id})")
                        else:
                            current_app.logger.warning(f"Sections row {row_idx}: Could not resolve parent reference (section_export_id={section_export_id}, parent_export_id={parent_export_id})")
                except Exception as e:
                    current_app.logger.error("Sections row %s (parent update): %s", row_idx, e)
                    errors.append(f"Sections row {row_idx} (parent update): Validation error.")

        current_app.logger.info(f"Sections import complete: {len(sections_data)} rows processed, {len(section_id_map)} sections created, {parent_updates} parent relationships updated, {len(errors)} errors")
        return errors

    @classmethod
    def _import_items(cls, sheet, template: FormTemplate, version: FormTemplateVersion,
                     section_id_map: Dict[int, int]) -> List[str]:
        """Import items from sheet."""
        current_app.logger.info("Starting items import...")
        current_app.logger.info(f"Section ID mapping available: {len(section_id_map)} sections")
        errors = []

        def _parse_int_like(value) -> Optional[int]:
            """Parse ints from Excel/JSON-ish values (e.g., 233, 233.0, '233')."""
            if value is None:
                return None
            try:
                # Excel frequently provides whole numbers as floats (e.g. 233.0)
                if isinstance(value, bool):
                    return None
                if isinstance(value, int):
                    return int(value)
                if isinstance(value, float):
                    if value.is_integer():
                        return int(value)
                    return None
                if isinstance(value, str):
                    s = value.strip()
                    if not s:
                        return None
                    if s.isdigit():
                        return int(s)
                    # Handle numeric strings like "233.0"
                    with suppress(Exception):
                        f = float(s)
                        if f.is_integer():
                            return int(f)
                return None
            except Exception as e:
                current_app.logger.debug("_parse_config_int failed: %s", e)
                return None

        def _ensure_import_issue(config: Any, *, code: str, message: str, meta: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
            """Attach an import issue to config['_import_issues'] and return the config dict."""
            cfg = config if isinstance(config, dict) else {}
            issues = cfg.get('_import_issues')
            if not isinstance(issues, list):
                issues = []
            issues.append({
                'code': code,
                'message': message,
                'meta': meta or {},
            })
            cfg['_import_issues'] = issues
            return cfg

        def _clear_import_issue_codes(config: Any, codes: List[str]) -> Any:
            """Remove matching import issue codes from config if present."""
            if not isinstance(config, dict):
                return config
            issues = config.get('_import_issues')
            if not isinstance(issues, list) or not issues:
                return config
            filtered = [i for i in issues if not (isinstance(i, dict) and i.get('code') in codes)]
            if filtered:
                config['_import_issues'] = filtered
            else:
                config.pop('_import_issues', None)
            return config

        # Read headers
        headers = [cell.value for cell in sheet[1]]
        current_app.logger.info(f"Items sheet headers: {headers}")

        # Validate headers
        expected_headers = cls.ITEM_COLUMNS
        if headers != expected_headers:
            error_msg = f"Items sheet headers don't match expected columns. Expected: {expected_headers}, Got: {headers}"
            current_app.logger.error(error_msg)
            errors.append(error_msg)
            return errors

        current_app.logger.info("Headers validated. Processing item rows...")

        # Get existing items for this version (for matching)
        existing_items_lookup = cls._get_existing_items_lookup(template.id, version.id)
        current_app.logger.info(f"Found {len(existing_items_lookup)} existing items to match against")

        # Pre-scan rows for indicator_bank_id values so we can validate in one DB query.
        # If an indicator references a missing IndicatorBank ID, we will import the item with
        # indicator_bank_id=NULL and flag it in config so the builder can show an issue and publishing can be blocked.
        candidate_indicator_bank_ids: set[int] = set()
        rows_buffer: List[Tuple[int, Dict[str, Any]]] = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue
            row_data = dict(zip(expected_headers, row))
            rows_buffer.append((row_idx, row_data))
            item_type = (row_data.get('item_type') or 'indicator')
            if item_type == 'indicator':
                raw_ib = row_data.get('indicator_bank_id')
                parsed_ib = _parse_int_like(raw_ib)
                if parsed_ib is not None:
                    candidate_indicator_bank_ids.add(parsed_ib)

        existing_indicator_bank_ids: set[int] = set()
        if candidate_indicator_bank_ids:
            existing_indicator_bank_ids = {
                int(x[0]) for x in db.session.query(IndicatorBank.id)
                .filter(IndicatorBank.id.in_(candidate_indicator_bank_ids))
                .all()
            }

        row_count = 0
        items_created = 0
        items_updated = 0

        # Map Items sheet export IDs -> actual FormItem objects (existing or newly created)
        item_export_to_obj: Dict[int, FormItem] = {}

        # Process buffered rows
        for row_idx, row_data in rows_buffer:
            row_count += 1
            try:
                # Get export ID
                export_id = int(row_data['id']) if row_data.get('id') else None

                # Resolve section_id using export ID mapping
                section_id = None
                section_export_id = None
                if row_data.get('section_id'):
                    section_export_id = int(row_data['section_id']) if row_data.get('section_id') else None
                    if section_export_id and section_export_id in section_id_map:
                        section_id = section_id_map[section_export_id]
                        current_app.logger.debug(f"Item row {row_idx}: Mapped section export_id={section_export_id} -> db_id={section_id}")
                    elif section_export_id:
                        current_app.logger.warning(f"Item row {row_idx}: Section export_id={section_export_id} not found in section_id_map")

                if not section_id:
                    error_msg = f"Items row {row_idx}: Could not resolve section_id (export_id: {section_export_id})"
                    current_app.logger.warning(error_msg)
                    errors.append(error_msg)
                    continue

                item_type = row_data['item_type'] or 'indicator'
                item_label = row_data['label'] or ''
                item_order = float(row_data['order']) if row_data['order'] is not None else 0.0
                current_app.logger.info(f"Processing item row {row_idx}: export_id={export_id}, type={item_type}, label='{item_label[:50]}...', order={item_order}, section_export_id={section_export_id}")

                # Check if item already exists (match by section_id + order + item_type + label)
                item_key = (section_id, item_order, item_type, item_label)
                existing_item = existing_items_lookup.get(item_key)

                if existing_item:
                    # Update existing item
                    current_app.logger.info(f"Updating existing item: db_id={existing_item.id}, section_id={section_id}, order={item_order}, type={item_type}")
                    existing_item.relevance_condition = row_data.get('relevance_condition')
                    existing_item.archived = bool(row_data.get('archived', False))
                    existing_item.config = cls._parse_json(row_data.get('config'))

                    # Validate indicator_bank_id (only meaningful for indicator items)
                    if item_type == 'indicator':
                        raw_ib = row_data.get('indicator_bank_id')
                        parsed_ib = _parse_int_like(raw_ib)
                        if parsed_ib is None:
                            # Missing/invalid ID: import as NULL + flag issue in config
                            existing_item.indicator_bank_id = None
                            existing_item.config = _ensure_import_issue(
                                existing_item.config,
                                code='missing_or_invalid_indicator_bank_id',
                                message='Indicator is missing a valid Indicator Bank ID. Please select a valid indicator before deploying.',
                                meta={'raw_indicator_bank_id': raw_ib},
                            )
                        elif parsed_ib not in existing_indicator_bank_ids:
                            # References a non-existent IndicatorBank row: import as NULL + flag issue in config
                            existing_item.indicator_bank_id = None
                            existing_item.config = _ensure_import_issue(
                                existing_item.config,
                                code='missing_indicator_bank_reference',
                                message=f'Indicator references missing Indicator Bank ID {parsed_ib}. Please select a valid indicator before deploying.',
                                meta={'raw_indicator_bank_id': raw_ib, 'parsed_indicator_bank_id': parsed_ib},
                            )
                        else:
                            # Valid reference: clear any prior import issue flags and set the FK
                            existing_item.indicator_bank_id = parsed_ib
                            existing_item.config = _clear_import_issue_codes(
                                existing_item.config,
                                ['missing_or_invalid_indicator_bank_id', 'missing_indicator_bank_reference'],
                            )
                    else:
                        # Non-indicator: clear any stale indicator import issues
                        existing_item.config = _clear_import_issue_codes(
                            existing_item.config,
                            ['missing_or_invalid_indicator_bank_id', 'missing_indicator_bank_reference'],
                        )
                        existing_item.indicator_bank_id = _parse_int_like(row_data.get('indicator_bank_id'))
                    existing_item.type = row_data.get('type')
                    existing_item.unit = row_data.get('unit')
                    existing_item.validation_condition = row_data.get('validation_condition')
                    existing_item.validation_message = row_data.get('validation_message')
                    existing_item.definition = row_data.get('definition')
                    existing_item.options_json = cls._parse_json(row_data.get('options_json'))
                    existing_item.lookup_list_id = str(row_data['lookup_list_id']) if row_data.get('lookup_list_id') else None
                    existing_item.list_display_column = row_data.get('list_display_column')
                    existing_item.list_filters_json = cls._parse_json(row_data.get('list_filters_json'))
                    existing_item.label_translations = cls._parse_json(row_data.get('label_translations'))
                    existing_item.definition_translations = cls._parse_json(row_data.get('definition_translations'))
                    existing_item.options_translations = cls._parse_json(row_data.get('options_translations'))
                    existing_item.description_translations = cls._parse_json(row_data.get('description_translations'))
                    existing_item.description = row_data.get('description')
                    items_updated += 1
                    current_app.logger.info(f"Item updated: db_id={existing_item.id}, label='{item_label[:50]}...'")
                    if export_id is not None:
                        item_export_to_obj[export_id] = existing_item
                else:
                    # Parse and validate indicator bank reference up-front (so we never violate FK constraints)
                    parsed_indicator_bank_id: Optional[int] = None
                    config_payload = cls._parse_json(row_data.get('config'))
                    if item_type == 'indicator':
                        raw_ib = row_data.get('indicator_bank_id')
                        parsed_ib = _parse_int_like(raw_ib)
                        if parsed_ib is None:
                            parsed_indicator_bank_id = None
                            config_payload = _ensure_import_issue(
                                config_payload,
                                code='missing_or_invalid_indicator_bank_id',
                                message='Indicator is missing a valid Indicator Bank ID. Please select a valid indicator before deploying.',
                                meta={'raw_indicator_bank_id': raw_ib},
                            )
                        elif parsed_ib not in existing_indicator_bank_ids:
                            parsed_indicator_bank_id = None
                            config_payload = _ensure_import_issue(
                                config_payload,
                                code='missing_indicator_bank_reference',
                                message=f'Indicator references missing Indicator Bank ID {parsed_ib}. Please select a valid indicator before deploying.',
                                meta={'raw_indicator_bank_id': raw_ib, 'parsed_indicator_bank_id': parsed_ib},
                            )
                        else:
                            parsed_indicator_bank_id = parsed_ib
                            config_payload = _clear_import_issue_codes(
                                config_payload,
                                ['missing_or_invalid_indicator_bank_id', 'missing_indicator_bank_reference'],
                            )
                    else:
                        parsed_indicator_bank_id = _parse_int_like(row_data.get('indicator_bank_id'))

                    # Create new item
                    new_item = FormItem(
                        section_id=section_id,
                        template_id=template.id,
                        version_id=version.id,
                        item_type=item_type,
                        label=item_label,
                        order=item_order,
                        relevance_condition=row_data.get('relevance_condition'),
                        archived=bool(row_data.get('archived', False)),
                        config=config_payload,
                        indicator_bank_id=parsed_indicator_bank_id,
                        type=row_data.get('type'),
                        unit=row_data.get('unit'),
                        validation_condition=row_data.get('validation_condition'),
                        validation_message=row_data.get('validation_message'),
                        definition=row_data.get('definition'),
                        options_json=cls._parse_json(row_data.get('options_json')),
                        lookup_list_id=str(row_data['lookup_list_id']) if row_data.get('lookup_list_id') else None,
                        list_display_column=row_data.get('list_display_column'),
                        list_filters_json=cls._parse_json(row_data.get('list_filters_json')),
                        label_translations=cls._parse_json(row_data.get('label_translations')),
                        definition_translations=cls._parse_json(row_data.get('definition_translations')),
                        options_translations=cls._parse_json(row_data.get('options_translations')),
                        description_translations=cls._parse_json(row_data.get('description_translations')),
                        description=row_data.get('description')
                    )

                    db.session.add(new_item)
                    items_created += 1
                    if export_id is not None:
                        item_export_to_obj[export_id] = new_item

                    if items_created % 10 == 0:
                        current_app.logger.debug(f"Items progress: {items_created} items created so far...")

            except Exception as e:
                current_app.logger.error("Items row %s: %s", row_idx, e, exc_info=True)
                errors.append(f"Items row {row_idx}: Validation error.")

        current_app.logger.info(f"Items import complete: {row_count} rows processed, {items_created} created, {items_updated} updated, {len(errors)} errors")

        # Rewrite rule JSON item IDs from export IDs -> new DB IDs so relevance/validation rules survive imports.
        try:
            db.session.flush()  # ensure new items have IDs
            item_id_map: Dict[int, int] = {}
            for exp_id, obj in item_export_to_obj.items():
                if obj is not None and getattr(obj, 'id', None) is not None:
                    item_id_map[int(exp_id)] = int(obj.id)

            if item_id_map:
                rewritten_items = 0
                for obj in item_export_to_obj.values():
                    if not obj:
                        continue
                    before_rel = obj.relevance_condition
                    before_val = getattr(obj, 'validation_condition', None)
                    obj.relevance_condition = cls._rewrite_rule_json_item_ids(before_rel, item_id_map)
                    if hasattr(obj, 'validation_condition'):
                        obj.validation_condition = cls._rewrite_rule_json_item_ids(before_val, item_id_map)
                    if before_rel != obj.relevance_condition or before_val != getattr(obj, 'validation_condition', None):
                        rewritten_items += 1

                rewritten_sections = 0
                sections = FormSection.query.filter_by(template_id=template.id, version_id=version.id).all()
                for s in sections:
                    before = s.relevance_condition
                    s.relevance_condition = cls._rewrite_rule_json_item_ids(before, item_id_map)
                    if before != s.relevance_condition:
                        rewritten_sections += 1

                current_app.logger.info(
                    f"Rewrote rule item IDs using item_id_map (size={len(item_id_map)}): "
                    f"items_updated={rewritten_items}, sections_updated={rewritten_sections}"
                )
        except Exception as e:
            # Don't fail the import for rule rewrite issues; log and continue.
            current_app.logger.warning(f"Could not rewrite rule item IDs after item import: {e}", exc_info=True)

        return errors

    @classmethod
    def _clone_template_structure(cls, template_id: int, source_version_id: int, target_version_id: int) -> None:
        """Clone pages, sections, and items from source_version_id to target_version_id preserving order."""
        current_app.logger.info(f"Cloning template structure from version {source_version_id} to {target_version_id}")
        # Maps for old->new IDs
        page_id_map = {}
        section_id_map = {}

        # Clone pages
        src_pages = FormPage.query.filter_by(template_id=template_id, version_id=source_version_id).order_by(FormPage.order).all()
        current_app.logger.info(f"Cloning {len(src_pages)} pages")
        for p in src_pages:
            new_p = FormPage(
                template_id=template_id,
                version_id=target_version_id,
                name=p.name,
                order=p.order,
                name_translations=p.name_translations
            )
            db.session.add(new_p)
            db.session.flush()
            page_id_map[p.id] = new_p.id

        # Clone sections (first pass: main sections)
        src_sections = FormSection.query.filter_by(template_id=template_id, version_id=source_version_id).order_by(FormSection.order).all()
        current_app.logger.info(f"Cloning {len(src_sections)} sections")
        # Create all sections without parent refs first
        for s in src_sections:
            new_s = FormSection(
                template_id=template_id,
                version_id=target_version_id,
                name=s.name,
                order=s.order,
                parent_section_id=None,  # set later
                page_id=page_id_map.get(s.page_id) if s.page_id else None,
                section_type=s.section_type,
                max_dynamic_indicators=s.max_dynamic_indicators,
                allowed_sectors=s.allowed_sectors,
                indicator_filters=s.indicator_filters,
                allow_data_not_available=s.allow_data_not_available,
                allow_not_applicable=s.allow_not_applicable,
                allowed_disaggregation_options=s.allowed_disaggregation_options,
                data_entry_display_filters=s.data_entry_display_filters,
                add_indicator_note=s.add_indicator_note,
                name_translations=s.name_translations,
                relevance_condition=s.relevance_condition,
                archived=s.archived
            )
            db.session.add(new_s)
            db.session.flush()
            section_id_map[s.id] = new_s.id

        # Second pass: set parent_section_id now that all new IDs exist
        parent_updates = 0
        for s in src_sections:
            if s.parent_section_id:
                new_id = section_id_map[s.id]
                new_parent_id = section_id_map.get(s.parent_section_id)
                if new_parent_id:
                    FormSection.query.filter_by(id=new_id).update({'parent_section_id': new_parent_id})
                    parent_updates += 1
        current_app.logger.info(f"Cloned {len(section_id_map)} sections, updated {parent_updates} parent relationships")

        # Clone items
        src_items = FormItem.query.join(FormSection, FormItem.section_id == FormSection.id).\
            filter(FormItem.template_id == template_id, FormItem.version_id == source_version_id).\
            order_by(FormItem.order).all()
        current_app.logger.info(f"Cloning {len(src_items)} items")
        items_cloned = 0
        for it in src_items:
            # Deep copy config to avoid cross-version mutations
            _new_config = None
            try:
                _new_config = json.loads(json.dumps(it.config)) if it.config is not None else None
            except Exception as e:
                current_app.logger.debug("JSON roundtrip for item config failed: %s", e)
                try:
                    import copy as _copy
                    _new_config = _copy.deepcopy(it.config) if it.config is not None else None
                except Exception as e2:
                    current_app.logger.debug("deepcopy item config failed: %s", e2)
                    _new_config = it.config.copy() if isinstance(it.config, dict) else it.config

            new_it = FormItem(
                template_id=template_id,
                version_id=target_version_id,
                section_id=section_id_map.get(it.section_id),
                item_type=it.item_type,
                label=it.label,
                order=it.order,
                relevance_condition=it.relevance_condition,
                archived=it.archived,
                config=_new_config,
                indicator_bank_id=it.indicator_bank_id,
                type=it.type,
                unit=it.unit,
                validation_condition=it.validation_condition,
                validation_message=it.validation_message,
                definition=it.definition,
                options_json=it.options_json,
                lookup_list_id=getattr(it, 'lookup_list_id', None),
                list_display_column=getattr(it, 'list_display_column', None),
                list_filters_json=getattr(it, 'list_filters_json', None),
                label_translations=getattr(it, 'label_translations', None),
                definition_translations=getattr(it, 'definition_translations', None),
                options_translations=getattr(it, 'options_translations', None),
                description_translations=getattr(it, 'description_translations', None),
                description=getattr(it, 'description', None)
            )
            db.session.add(new_it)
            items_cloned += 1

        current_app.logger.info(f"Successfully cloned structure: {len(page_id_map)} pages, {len(section_id_map)} sections, {items_cloned} items")

    @classmethod
    def _get_existing_items_lookup(cls, template_id: int, version_id: int) -> Dict[Tuple, FormItem]:
        """
        Create a lookup dictionary for existing items.
        Key: (section_id, order, item_type, label) - combination to uniquely identify an item
        Value: FormItem object
        """
        existing_items = FormItem.query.join(FormSection).filter(
            FormItem.template_id == template_id,
            FormSection.version_id == version_id
        ).all()

        # Create lookup: (section_id, order, item_type, label) -> item
        # Using label as part of key since it's more stable than just order
        items_lookup = {}
        for item in existing_items:
            key = (item.section_id, item.order, item.item_type, item.label or '')
            items_lookup[key] = item

        return items_lookup

    @classmethod
    def _parse_json(cls, value) -> Optional[Any]:
        """Parse JSON string to object, return None if invalid or empty."""
        if not value or value == '' or value == 'None':
            return None
        if isinstance(value, (dict, list)):
            return value
        try:
            parsed = json.loads(str(value))
            return parsed if parsed else None
        except (json.JSONDecodeError, TypeError):
            return None
