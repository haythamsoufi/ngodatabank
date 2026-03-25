# ========== Kobo XLSForm Import Service ==========
"""
Service for importing Kobo Toolbox XLSForm (.xlsx/.xls) files to create form templates.

Parses the standard XLSForm structure:
- survey worksheet: type, name, label (plus optional: relevant, required, appearance, etc.)
- choices worksheet: list_name, name, label (for select_one, select_multiple)
- settings worksheet: form_title (optional)

Maps to IFRC form structure:
- begin_group / end_group -> FormSection (standard)
- begin_repeat / end_repeat -> FormSection (repeat)
- Question types -> FormItem (question) with appropriate question_type

Supported Kobo question types -> IFRC mapping:
- text -> text (or textarea if appearance contains multiline)
- integer, decimal -> number
- select_one -> single_choice
- select_multiple -> multiple_choice
- date -> date
- datetime -> datetime
- time -> datetime (closest match)
- note -> blank (note/instruction)
- image, audio, video, file -> document_field (file upload)

Unsupported types (skipped with warning): geopoint, geotrace, geoshape, calculate,
hidden, barcode, background-audio, rank, acknowledge, select_one_from_file, select_multiple_from_file
"""

from __future__ import annotations

import io
from typing import Dict, List, Any, Optional
from contextlib import suppress

from flask import current_app
from flask_login import current_user
from app import db
from app.models import FormTemplate, FormSection, FormItem, FormTemplateVersion
from app.utils.datetime_helpers import utcnow

try:
    import openpyxl
except ImportError:
    openpyxl = None


# Kobo question types we support and their IFRC equivalents
KOBO_TO_QUESTION_TYPE = {
    'text': 'text',
    'textarea': 'textarea',  # Kobo uses appearance for multiline
    'integer': 'number',
    'decimal': 'number',
    'range': 'number',
    'select_one': 'single_choice',
    'select_multiple': 'multiple_choice',
    'date': 'date',
    'datetime': 'datetime',
    'time': 'datetime',  # closest match
    'note': 'blank',
    'image': 'document_field',
    'photo': 'document_field',
    'audio': 'document_field',
    'video': 'document_field',
    'file': 'document_field',
}

# Kobo types we skip (log warning, don't create item)
KOBO_SKIP_TYPES = {
    'geopoint', 'geotrace', 'geoshape',
    'calculate', 'hidden',
    'barcode', 'background-audio',
    'rank', 'acknowledge',
    'select_one_from_file', 'select_multiple_from_file',
}


def _parse_str(val: Any) -> str:
    """Safely coerce to string."""
    if val is None:
        return ''
    s = str(val).strip()
    return s if s else ''


def _get_label_from_row(row: Dict[str, Any], name_fallback: str = '') -> str:
    """Extract display label from a Kobo survey row. Prefer label over name."""
    if not isinstance(row, dict):
        return _parse_str(name_fallback) or 'Untitled'
    label = ''
    for k, v in row.items():
        if v is None or v == '':
            continue
        kn = str(k).strip().lower() if k else ''
        if kn == 'label':
            label = _parse_str(v)
            break
        if kn.startswith('label::'):
            label = _parse_str(v)
            if label:
                break
    return label or _parse_str(name_fallback) or 'Untitled'


def _get_type_from_row(row: Dict[str, Any]) -> str:
    """Extract type value from Kobo survey row (handles column name variations)."""
    if not isinstance(row, dict):
        return ''
    v = row.get('type')
    if v is not None and str(v).strip():
        return str(v).strip()
    for k, val in row.items():
        if k and str(k).strip().lower() == 'type' and val is not None:
            return str(val).strip() if val else ''
    return ''


def _parse_repeat_count(val: Any) -> Optional[int]:
    """Parse repeat_count from Kobo (can be number or reference like ${question_name})."""
    if val is None:
        return None
    try:
        if isinstance(val, (int, float)) and not isinstance(val, bool):
            n = int(float(val))
            return n if n > 0 else None
        s = str(val).strip()
        if not s or s.startswith('${'):  # dynamic reference, skip
            return None
        n = int(float(s))
        return n if n > 0 else None
    except (ValueError, TypeError):
        return None


def _parse_float(val: Any) -> float:
    """Safely parse float for order values."""
    if val is None:
        return 0.0
    try:
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip()
        return float(s) if s else 0.0
    except (ValueError, TypeError):
        return 0.0


class KoboXlsImportService:
    """Import Kobo XLSForm files into IFRC form templates."""

    @classmethod
    def import_kobo_xls(
        cls,
        excel_file,
        *,
        template_name: Optional[str] = None,
        owned_by: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        Parse a Kobo XLSForm file and create a new form template with sections and items.

        Args:
            excel_file: File-like object (Excel .xlsx or .xls)
            template_name: Optional override for template name (otherwise from form_title in settings)
            owned_by: User ID for template ownership

        Returns:
            Dict with keys: success, message, template_id, version_id, created_counts, errors, warnings
        """
        if openpyxl is None:
            return {
                'success': False,
                'message': 'openpyxl is required for Kobo import',
                'template_id': None,
                'version_id': None,
                'created_counts': {'sections': 0, 'items': 0},
                'errors': ['openpyxl package not installed'],
                'warnings': [],
            }

        errors: List[str] = []
        warnings: List[str] = []
        created_counts = {'sections': 0, 'items': 0}

        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(excel_file.read()),
                data_only=True,
                read_only=False,
            )
        except Exception as e:
            current_app.logger.error(f"Failed to load Kobo XLS file: {e}", exc_info=True)
            return {
                'success': False,
                'message': 'Invalid Excel file. Check the file format and try again.',
                'template_id': None,
                'version_id': None,
                'created_counts': created_counts,
                'errors': ['Failed to load Excel file.'],
                'warnings': [],
            }

        # Normalize sheet names (Kobo uses lowercase: survey, choices, settings)
        sheet_names_lower = {name.lower(): name for name in workbook.sheetnames}
        survey_sheet_name = sheet_names_lower.get('survey')
        choices_sheet_name = sheet_names_lower.get('choices')
        settings_sheet_name = sheet_names_lower.get('settings')

        if not survey_sheet_name:
            return {
                'success': False,
                'message': 'Kobo XLSForm must contain a "survey" worksheet',
                'template_id': None,
                'version_id': None,
                'created_counts': created_counts,
                'errors': ['Missing survey worksheet'],
                'warnings': [],
            }

        # Read form title from settings if available
        form_title = template_name
        if settings_sheet_name and not form_title:
            form_title = cls._read_form_title(workbook[settings_sheet_name])
        if not form_title or not _parse_str(form_title):
            form_title = 'Imported from Kobo'

        # Load choices for select_one / select_multiple
        choices_by_list: Dict[str, List[Tuple[str, str]]] = {}
        if choices_sheet_name:
            choices_by_list = cls._load_choices(workbook[choices_sheet_name])

        # Parse survey rows into a flat list with hierarchy
        survey_rows = cls._parse_survey_sheet(workbook[survey_sheet_name])
        if not survey_rows:
            return {
                'success': False,
                'message': 'Survey worksheet is empty or has no valid rows',
                'template_id': None,
                'version_id': None,
                'created_counts': created_counts,
                'errors': ['No valid survey rows'],
                'warnings': [],
            }

        # Build hierarchy: groups/repeats become sections, questions become items
        stack: List[Dict[str, Any]] = []
        root_sections: List[Dict] = []
        root_items: List[Dict] = []

        for row in survey_rows:
            type_val = _get_type_from_row(row)
            kobo_type_norm = type_val.replace(' ', '_').strip().lower()
            name = _parse_str(row.get('name') or '')
            label = _get_label_from_row(row, name)
            relevant = _parse_str(row.get('relevant') or row.get('required') or '')
            repeat_count = _parse_repeat_count(row.get('repeat_count'))

            if kobo_type_norm in ('begin_group', 'begin_repeat'):
                is_repeat = kobo_type_norm == 'begin_repeat'
                section_info = {
                    'name': name or f"group_{len(stack) + 1}",
                    'label': label,
                    'section_type': 'repeat' if is_repeat else 'standard',
                    'relevant': relevant,
                    'order': _parse_float(row.get('order', len(stack) + len(root_sections) + 1)),
                    'repeat_count': repeat_count if is_repeat else None,
                }
                stack.append({
                    'section_info': section_info,
                    'children': [],
                })
            elif kobo_type_norm in ('end_group', 'end_repeat'):
                if stack:
                    closed = stack.pop()
                    parent = stack[-1] if stack else None
                    node = {
                        'type': 'section',
                        'section_info': closed['section_info'],
                        'children': closed['children'],
                    }
                    if parent:
                        parent['children'].append(node)
                    else:
                        root_sections.append(node)
            else:
                # Question or note
                item_info = cls._map_kobo_row_to_item(row, choices_by_list, warnings)
                if item_info is None:
                    continue  # skipped (unsupported type)
                node = {'type': 'item', 'item_info': item_info}
                if stack:
                    stack[-1]['children'].append(node)
                else:
                    root_items.append(node)

        # Any unclosed groups: treat as root sections
        while stack:
            closed = stack.pop()
            root_sections.append({
                'type': 'section',
                'section_info': closed['section_info'],
                'children': closed['children'],
            })

        # Create template and version
        owner_id = owned_by or (current_user.id if current_user.is_authenticated else None)
        if not owner_id:
            return {
                'success': False,
                'message': 'Cannot determine template owner',
                'template_id': None,
                'version_id': None,
                'created_counts': created_counts,
                'errors': ['User not authenticated'],
                'warnings': warnings,
            }

        template = FormTemplate(created_by=owner_id, owned_by=owner_id)
        db.session.add(template)
        db.session.flush()

        now = utcnow()
        version = FormTemplateVersion(
            template_id=template.id,
            version_number=1,
            status='draft',
            name=form_title,
            description=f'Imported from Kobo XLSForm',
            add_to_self_report=False,
            display_order_visible=False,
            is_paginated=False,
            enable_export_pdf=False,
            enable_export_excel=False,
            enable_import_excel=False,
            enable_ai_validation=False,
            created_by=owner_id,
            updated_by=owner_id,
            created_at=now,
            updated_at=now,
        )
        db.session.add(version)
        db.session.flush()

        # Create sections and items
        section_order = 0.0

        def create_sections_and_items(
            nodes: List[Dict],
            parent_section_id: Optional[int] = None,
        ) -> None:
            nonlocal section_order
            for idx, node in enumerate(nodes):
                if node['type'] == 'section':
                    si = node['section_info']
                    section_order += 1.0
                    section_name = si.get('label') or si['name']  # use label (display) over name (id)
                    section_type = si.get('section_type') or 'standard'
                    section = FormSection(
                        template_id=template.id,
                        version_id=version.id,
                        name=section_name,
                        order=section_order,
                        parent_section_id=parent_section_id,
                        page_id=None,
                        section_type=section_type,
                        relevance_condition=si.get('relevant') or None,
                    )
                    if section_type == 'repeat' and si.get('repeat_count'):
                        section.config = section.config or {}
                        section.config['max_entries'] = si['repeat_count']
                    db.session.add(section)
                    db.session.flush()
                    created_counts['sections'] += 1
                    create_sections_and_items(node['children'], section.id)
                else:
                    item_info = node['item_info']
                    item_order = item_info.get('order', idx)
                    section_id = parent_section_id
                    if not section_id:
                        # Should not happen if hierarchy is correct; create fallback Main section
                        default_section = FormSection(
                            template_id=template.id,
                            version_id=version.id,
                            name='Main',
                            order=0.0,
                            parent_section_id=None,
                            page_id=None,
                            section_type='standard',
                        )
                        db.session.add(default_section)
                        db.session.flush()
                        created_counts['sections'] += 1
                        section_id = default_section.id
                    item = FormItem(
                        section_id=section_id,
                        template_id=template.id,
                        version_id=version.id,
                        item_type=item_info['item_type'],
                        label=item_info['label'],
                        order=float(item_order),
                        relevance_condition=item_info.get('relevance') or None,
                        type=item_info.get('type'),  # question type for questions
                        options_json=item_info.get('options_json'),
                        config=item_info.get('config') or {},
                    )
                    db.session.add(item)
                    created_counts['items'] += 1

        # Process root: questions before any group go into "Main" section; groups become nested sections
        if root_items:
            main_section = FormSection(
                template_id=template.id,
                version_id=version.id,
                name='Main',
                order=0.0,
                parent_section_id=None,
                page_id=None,
                section_type='standard',
            )
            db.session.add(main_section)
            db.session.flush()
            created_counts['sections'] += 1
            main_section_id = main_section.id
            create_sections_and_items(root_items, main_section_id)
        create_sections_and_items(root_sections, None)

        version.updated_at = utcnow()
        version.updated_by = owner_id
        db.session.commit()

        return {
            'success': True,
            'message': f"Template '{form_title}' created with {created_counts['sections']} sections and {created_counts['items']} items",
            'template_id': template.id,
            'version_id': version.id,
            'created_counts': created_counts,
            'errors': [],
            'warnings': warnings,
        }

    @classmethod
    def _read_form_title(cls, sheet) -> Optional[str]:
        """Read form_title from settings sheet (column form_title, row 2)."""
        try:
            headers = [cell.value for cell in sheet[1]]
            idx = next((i for i, h in enumerate(headers) if h and str(h).strip().lower() == 'form_title'), None)
            if idx is not None and sheet.max_row >= 2:
                val = sheet.cell(row=2, column=idx + 1).value
                return _parse_str(val) if val else None
        except Exception as e:
            current_app.logger.debug("Could not read form_title from Kobo settings sheet: %s", e)
        return None

    @classmethod
    def _load_choices(cls, sheet) -> Dict[str, List[Tuple[str, str]]]:
        """Load choices worksheet: list_name -> [(name, label), ...]."""
        result: Dict[str, List[Tuple[str, str]]] = {}
        try:
            headers = [cell.value for cell in sheet[1]]
            list_idx = next((i for i, h in enumerate(headers) if h and str(h).strip().lower() == 'list_name'), None)
            name_idx = next((i for i, h in enumerate(headers) if h and str(h).strip().lower() == 'name'), None)
            label_idx = next((i for i, h in enumerate(headers) if h and str(h).strip().lower() == 'label'), None)
            if list_idx is None or name_idx is None or label_idx is None:
                return result
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= max(list_idx, name_idx, label_idx):
                    continue
                list_name = _parse_str(row[list_idx])
                choice_name = _parse_str(row[name_idx])
                choice_label = _parse_str(row[label_idx]) or choice_name
                if list_name and choice_name:
                    result.setdefault(list_name, []).append((choice_name, choice_label))
        except Exception as e:
            current_app.logger.warning(f"Error loading Kobo choices: {e}")
        return result

    @classmethod
    def _parse_survey_sheet(cls, sheet) -> List[Dict[str, Any]]:
        """Parse survey sheet into list of row dicts."""
        rows = []
        try:
            header_row = [cell.value for cell in sheet[1]]
            headers = [str(h).strip().lower() if h else '' for h in header_row]
            type_idx = next((i for i, h in enumerate(headers) if h == 'type'), None)
            if type_idx is None:
                return rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= type_idx:
                    continue
                type_val = row[type_idx]
                if type_val is None or not str(type_val).strip():
                    continue
                row_data = dict(zip(headers, row))
                row_data['_order'] = len(rows)
                rows.append(row_data)
        except Exception as e:
            current_app.logger.warning(f"Error parsing survey sheet: {e}")
        return rows

    @classmethod
    def _map_kobo_row_to_item(
        cls,
        row: Dict[str, Any],
        choices_by_list: Dict[str, List[Tuple[str, str]]],
        warnings: List[str],
    ) -> Optional[Dict[str, Any]]:
        """Map a Kobo survey row to IFRC item dict. Returns None if skipped."""
        type_raw = (row.get('type') or '').strip().lower()
        name = _parse_str(row.get('name') or '')
        label = _get_label_from_row(row, name)
        relevant = _parse_str(row.get('relevant') or '')
        required = _parse_str(row.get('required') or '').lower() in ('yes', '1', 'true')
        appearance = _parse_str(row.get('appearance') or '').lower()

        # Extract base type (e.g. select_one listname -> select_one)
        base_type = type_raw.split()[0] if type_raw else ''
        list_name = ''
        if ' ' in type_raw:
            rest = type_raw.split(None, 1)[1]
            list_name = rest.strip() if rest else ''

        if base_type in KOBO_SKIP_TYPES:
            warnings.append(f"Skipped unsupported type '{base_type}': {label[:50]}...")
            return None

        if base_type not in KOBO_TO_QUESTION_TYPE:
            warnings.append(f"Skipped unknown type '{base_type}': {label[:50]}...")
            return None

        ifc_type = KOBO_TO_QUESTION_TYPE[base_type]

        # text with multiline/min_length appearance -> textarea
        if base_type == 'text' and ('multiline' in appearance or 'min_length' in appearance):
            ifc_type = 'textarea'

        item_type = 'question'
        options_json = None
        config: Dict[str, Any] = {'is_required': required}
        if ifc_type == 'document_field':
            item_type = 'document_field'
            ifc_type = None
        elif base_type in ('select_one', 'select_multiple'):
            options = choices_by_list.get(list_name, [])
            options_json = [{'value': n, 'label': lb} for n, lb in options]
            if not options_json:
                warnings.append(f"No choices for list '{list_name}' ({label[:50]}...), creating empty options")

        return {
            'item_type': item_type,
            'type': ifc_type,
            'label': label,
            'order': row.get('_order', 0),
            'relevance': relevant or None,
            'options_json': options_json,
            'config': config,
        }
