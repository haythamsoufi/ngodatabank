"""
KoBo Data Import Service

Imports submission data from KoBo Toolbox data export Excel files.
Handles: analyze structure, create templates, match entities, import data.

KoBo data exports have a specific format:
- Row 1: headers in "Group/Question" hierarchical format
- Rows 2+: one row per submission
- System columns at start (start, end, today) and end (_id, _uuid, etc.)
- Disaggregation columns (gender_X.Y, sexage_X.Y, age/*)
- Calculated/display columns (<span>, ${var}, total_*, cal_total_*)
"""

from __future__ import annotations

import re
import io
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Tuple, Set
from collections import OrderedDict

from flask import current_app

try:
    import openpyxl
except ImportError:
    openpyxl = None

KOBO_SYSTEM_EXACT = {
    'start', 'end', 'today', 'deviceid', 'phonenumber',
    'username', 'simserial', 'subscriberid',
}

_HIDDEN_RE = re.compile(r'<span\s+style=["\']display:\s*none', re.I)
_VALIDATION_RE = re.compile(r'<span\s+style=["\']color:\s*red', re.I)
_FORMULA_RE = re.compile(r'\$\{[^}]+\}')
_CALC_VAR_RE = re.compile(
    r'^(total_|cal_total_|orgs_total$|unique_indiv_|general_\d)',
    re.I,
)
_GENDER_GROUP_RE = re.compile(r'^gender_[\w.]+$', re.I)
_SEXAGE_GROUP_RE = re.compile(r'^sexage_[\w.]+$', re.I)
_AGE_RANGE_RE = re.compile(r'^\d+-\d+$|^\d+\+$')

_NS_STRIP = re.compile(
    r'\b(national\s+)?red\s+cross(\s+society)?'
    r'|\bred\s+crescent(\s+society)?'
    r'|\bcroix[- ]rouge'
    r'|\bcroissant[- ]rouge'
    r'|\bof\s*$'
    r'|\bsociety\s*$',
    re.I,
)


def _s(val: Any) -> str:
    if val is None:
        return ''
    return str(val).strip()


def _is_date(v: Any) -> bool:
    return isinstance(v, (datetime, date))


def _is_numeric(v: Any) -> bool:
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    s = _s(v)
    if not s:
        return False
    try:
        float(s)
        return True
    except (ValueError, TypeError):
        return False


def _classify_header(header: str) -> Tuple[str, str]:
    """Return (category, reason) for a column header.

    Categories: system, hidden, calculated, note, file_url, data
    """
    h = _s(header)
    if not h:
        return 'system', 'Empty header'

    if h.lower() in KOBO_SYSTEM_EXACT:
        return 'system', 'KoBo system field'

    if h.startswith('__') or (h.startswith('_') and '/' not in h):
        return 'system', 'KoBo metadata field'

    if h.lower().startswith('meta/'):
        return 'system', 'KoBo meta field'

    if _HIDDEN_RE.search(h):
        return 'hidden', 'Hidden display field'

    if _VALIDATION_RE.search(h):
        return 'hidden', 'Validation message'

    if h.endswith('_URL'):
        return 'file_url', 'File URL companion'

    last_segment = h.rsplit('/', 1)[-1] if '/' in h else h
    if _CALC_VAR_RE.match(last_segment):
        return 'calculated', 'Calculated variable'

    star_inner = last_segment.strip('*')
    if last_segment.startswith('*') and last_segment.endswith('*') and _FORMULA_RE.search(star_inner):
        return 'calculated', 'Calculated display'

    return 'data', ''


def _detect_disagg(header: str) -> Optional[str]:
    """Detect if a data column is a disaggregation sub-column."""
    parts = header.split('/')
    if len(parts) < 2:
        return None

    for p in parts[:-1]:
        if _GENDER_GROUP_RE.match(p):
            return 'sex'
        if _SEXAGE_GROUP_RE.match(p):
            return 'sex_age'

    last = parts[-1].strip('*').strip()
    if last in ('Male', 'Female'):
        for p in parts[:-1]:
            if _GENDER_GROUP_RE.match(p) or _SEXAGE_GROUP_RE.match(p):
                return 'sex' if _GENDER_GROUP_RE.match(p) else 'sex_age'

    for i, p in enumerate(parts[:-1]):
        if p.lower() == 'age' and _AGE_RANGE_RE.match(parts[-1].strip()):
            return 'age'

    return None


def _detect_data_type(values: List[Any]) -> Tuple[str, Optional[List[str]]]:
    """Detect the data type from a list of non-null values.

    Returns (type_name, options_or_none).
    """
    if not values:
        return 'text', None

    non_null = [v for v in values if v is not None and _s(v) != '']
    if not non_null:
        return 'text', None

    if all(_is_date(v) for v in non_null):
        return 'date', None

    if all(_is_numeric(v) for v in non_null):
        return 'number', None

    str_vals = [_s(v) for v in non_null if _s(v)]
    if not str_vals:
        return 'text', None

    unique = set(v.lower() for v in str_vals)

    if unique <= {'yes', 'no'}:
        return 'yesno', None

    if 2 <= len(unique) <= 15 and all(len(v) < 100 for v in str_vals) and len(unique) < len(str_vals) * 0.8:
        options = sorted(set(str_vals))
        return 'single_choice', options

    max_len = max(len(v) for v in str_vals) if str_vals else 0
    if max_len > 500:
        return 'textarea', None

    return 'text', None


def _extract_group_and_label(header: str) -> Tuple[Optional[str], str]:
    """Split header into (group_name, question_label)."""
    if '/' not in header:
        return None, header.strip()
    first_slash = header.index('/')
    group = header[:first_slash].strip()
    label = header[first_slash + 1:].strip()
    return group, label


def _find_validation_status_column(headers: List[Any]) -> Optional[int]:
    """Column index for KoBo / ODK validation status (metadata), if present."""
    for i, h in enumerate(headers):
        s = _s(h)
        if not s:
            continue
        low = s.lower()
        if low in ('_validation_status', 'validation_status'):
            return i
        if '/' in low:
            tail = low.rsplit('/', 1)[-1].strip('_').lower()
            if tail == 'validation_status':
                return i
    return None


def _kobo_validation_bucket(cell_val: Any) -> str:
    """Map _validation_status cell text to a coarse bucket (KoBo / ODK label variants).

    Buckets: ``approved``, ``on_hold``, ``not_validated``, ``pending``, ``draft``,
    ``rejected``, ``empty``, ``unknown``.
    """
    s = _s(cell_val)
    if not s:
        return 'empty'
    norm = s.lower().replace(' ', '_').replace('-', '_')

    if 'not_approved' in norm or 'notapproved' in norm:
        return 'rejected'
    if 'rejected' in norm or 'denied' in norm or 'declined' in norm:
        return 'rejected'
    if 'flagged_for_removal' in norm:
        return 'rejected'
    if 'removal' in norm and 'approved' not in norm:
        return 'rejected'

    if 'on_hold' in norm or 'onhold' in norm:
        return 'on_hold'

    if 'not_validated' in norm or 'unvalidated' in norm:
        return 'not_validated'

    if norm in ('draft',) or norm.endswith('_draft'):
        return 'draft'
    if 'approval_requested' in norm:
        return 'pending'
    if 'pending' in norm and 'approved' not in norm:
        return 'pending'

    if 'approved' in norm:
        return 'approved'

    return 'unknown'


def _is_kobo_submission_approved(cell_val: Any) -> bool:
    """True when validation status is treated as KoBo *Approved*."""
    return _kobo_validation_bucket(cell_val) == 'approved'


def _row_matches_submission_filter(cell_val: Any, submission_filter: str) -> bool:
    """Whether a row should be kept for the given ``submission_filter`` mode."""
    sf = (submission_filter or 'all').strip()
    if sf == 'all':
        return True

    bucket = _kobo_validation_bucket(cell_val)

    if sf == 'approved_only':
        return bucket == 'approved'

    if sf == 'exclude_rejected':
        # Approved, on hold, not validated, pending, draft, unknown, empty — drop rejected / flagged only.
        return bucket != 'rejected'

    if sf == 'approved_or_on_hold':
        return bucket in ('approved', 'on_hold')

    if sf == 'on_hold_only':
        return bucket == 'on_hold'

    if sf == 'not_validated_only':
        return bucket in ('not_validated', 'pending')

    if sf == 'draft_only':
        return bucket == 'draft'

    return True


def _sub_time_sort_value(data_rows: List[list], ri: int, sub_time_col: Optional[int]) -> Any:
    if sub_time_col is None:
        return datetime.min
    v = data_rows[ri][sub_time_col] if sub_time_col < len(data_rows[ri]) else None
    return v if isinstance(v, (datetime, date)) else datetime.min


def _eligible_row_indices(
    data_rows: List[list],
    *,
    submission_filter: str,
    validation_col: Optional[int],
) -> Tuple[List[int], int, Optional[str]]:
    """Rows to consider for import. Returns (indices, excluded_count, error_message)."""
    n = len(data_rows)
    all_idx = list(range(n))
    sf = (submission_filter or 'all').strip()
    if sf == 'all':
        return all_idx, 0, None

    if validation_col is None:
        return [], n, (
            f'This export has no _validation_status column; cannot use submission filter “{sf}”. '
            'Use an XLS export that includes validation metadata, or choose “All rows”.'
        )

    eligible: List[int] = []
    excluded = 0
    for ri in all_idx:
        cell = data_rows[ri][validation_col] if validation_col < len(data_rows[ri]) else None
        if _row_matches_submission_filter(cell, sf):
            eligible.append(ri)
        else:
            excluded += 1
    return eligible, excluded, None


def _pick_duplicate_winner(
    indices: List[int],
    *,
    data_rows: List[list],
    sub_time_col: Optional[int],
    duplicate_strategy: str,
    validation_col: Optional[int],
) -> Tuple[int, bool]:
    """Choose one row index from duplicates. Returns (keep_row_index, used_unapproved_fallback)."""
    if len(indices) == 1:
        return indices[0], False

    sorted_idx = sorted(indices, key=lambda ri: _sub_time_sort_value(data_rows, ri, sub_time_col))

    if duplicate_strategy == 'latest':
        return sorted_idx[-1], False
    if duplicate_strategy == 'first':
        return sorted_idx[0], False

    if duplicate_strategy in ('latest_approved', 'first_approved'):
        approved_list = []
        if validation_col is not None:
            for ri in sorted_idx:
                cell = data_rows[ri][validation_col] if validation_col < len(data_rows[ri]) else None
                if _is_kobo_submission_approved(cell):
                    approved_list.append(ri)
        if approved_list:
            if duplicate_strategy == 'latest_approved':
                return approved_list[-1], False
            return approved_list[0], False
        if duplicate_strategy == 'latest_approved':
            return sorted_idx[-1], True
        return sorted_idx[0], True

    # 'all' — caller should not use this for a single pick
    return sorted_idx[-1], False


def _resolve_import_row_indices(
    data_rows: List[list],
    headers: List[Any],
    *,
    entity_column_index: int,
    duplicate_strategy: str,
    submission_filter: str,
    submission_time_column_index: Optional[int],
    validation_status_column_index: Optional[int],
) -> Tuple[List[int], Dict[str, Any]]:
    """Apply submission filter + per-entity deduplication. Returns (flat row indices, stats dict)."""
    stats: Dict[str, Any] = {
        'rows_excluded_by_submission_filter': 0,
        'rows_excluded_not_approved': 0,  # same as rows_excluded_by_submission_filter (legacy key)
        'duplicate_count': 0,
        'duplicate_details': [],
        'error': None,
    }

    val_col = validation_status_column_index
    if val_col is not None and (val_col < 0 or val_col >= len(headers)):
        val_col = None
    if val_col is None:
        val_col = _find_validation_status_column(headers)

    eligible, excluded, err = _eligible_row_indices(
        data_rows, submission_filter=submission_filter, validation_col=val_col,
    )
    stats['rows_excluded_by_submission_filter'] = excluded
    stats['rows_excluded_not_approved'] = excluded
    if err:
        stats['error'] = err
        return [], stats

    eligible_set = set(eligible)
    entity_rows: Dict[str, List[int]] = OrderedDict()
    for ri in eligible:
        row = data_rows[ri]
        ename = _s(row[entity_column_index]) if entity_column_index < len(row) else ''
        if not ename:
            continue
        entity_rows.setdefault(ename, []).append(ri)

    kept_indices: List[int] = []
    dup_details: List[Dict[str, Any]] = []

    if duplicate_strategy == 'all':
        for ename, row_indices in entity_rows.items():
            kept_indices.extend(row_indices)
        stats['duplicate_count'] = 0
        stats['duplicate_details'] = []
        return kept_indices, stats

    for ename, row_indices in entity_rows.items():
        if len(row_indices) <= 1:
            kept_indices.append(row_indices[0])
            continue

        keep, fallback_unapproved = _pick_duplicate_winner(
            row_indices,
            data_rows=data_rows,
            sub_time_col=submission_time_column_index,
            duplicate_strategy=duplicate_strategy,
            validation_col=val_col,
        )
        stats['duplicate_count'] += len(row_indices) - 1

        detail: Dict[str, Any] = {
            'entity': ename,
            'total_submissions': len(row_indices),
            'removed': len(row_indices) - 1,
            'kept_row': keep + 2,
            'kept_reason': duplicate_strategy,
            'all_rows': [ri + 2 for ri in sorted(row_indices)],
            'fallback_unapproved': fallback_unapproved,
        }
        if submission_time_column_index is not None:
            stc = submission_time_column_index
            sorted_for_ts = sorted(row_indices, key=lambda ri: _sub_time_sort_value(data_rows, ri, stc))
            ts_list = []
            for ri in sorted_for_ts:
                v = data_rows[ri][stc] if stc < len(data_rows[ri]) else None
                ts_list.append(v.isoformat() if isinstance(v, (datetime, date)) else _s(v))
            detail['timestamps'] = ts_list
            if keep in sorted_for_ts:
                kept_ts_idx = sorted_for_ts.index(keep)
                detail['kept_timestamp'] = ts_list[kept_ts_idx] if kept_ts_idx < len(ts_list) else None
        dup_details.append(detail)
        kept_indices.append(keep)

    stats['duplicate_details'] = dup_details
    return kept_indices, stats


def _slugify_sex_category(sex_cat: str) -> str:
    return str(sex_cat).strip().lower().replace(' ', '_').replace('-', '_')


def _normalize_column_to_item_mapping(raw: Dict[str, Any]) -> Dict[int, Dict[str, Any]]:
    """Parse column_to_item_mapping from client (legacy int or structured dict)."""
    out: Dict[int, Dict[str, Any]] = {}
    if not raw:
        return out
    for k, v in raw.items():
        try:
            ci = int(k)
        except (TypeError, ValueError):
            continue
        if isinstance(v, dict) and v.get('item_id'):
            out[ci] = {'item_id': int(v['item_id']), 'disagg': v.get('disagg')}
        elif v is not None:
            try:
                out[ci] = {'item_id': int(v), 'disagg': None}
            except (TypeError, ValueError):
                continue
    return out


def _disagg_slice_label(mode: Optional[str], key: Optional[str], form_item: Any) -> str:
    """Human-readable suffix for preview column headers."""
    if not mode or mode == 'total':
        return 'total'
    if mode == 'sex' and key:
        for s in getattr(form_item, 'effective_sex_categories', []) or []:
            if _slugify_sex_category(s) == key:
                return str(s)
        return key
    if mode == 'age' and key:
        for a in getattr(form_item, 'effective_age_groups', []) or []:
            from app.utils.form_processing import slugify_age_group
            if slugify_age_group(a) == key:
                return str(a)
        return key
    if mode == 'sex_age' and key:
        from app.utils.form_processing import slugify_age_group
        sex_cats = getattr(form_item, 'effective_sex_categories', []) or []
        age_grps = getattr(form_item, 'effective_age_groups', []) or []
        for sx in sex_cats:
            sx_slug = _slugify_sex_category(sx)
            for ag in age_grps:
                ag_slug = slugify_age_group(ag)
                if f'{sx_slug}_{ag_slug}' == key:
                    return f'{sx} · {ag}'
        return key.replace('_', ' ')
    return mode or ''


def _validate_disagg_for_item(form_item: Any, disagg: Any) -> Optional[str]:
    """Return error message if disagg is invalid for this item, else None."""
    from app.utils.form_processing import slugify_age_group
    from app.models.form_items import FormItem

    if not isinstance(form_item, FormItem):
        return 'Invalid form item'

    opts = set(form_item.allowed_disaggregation_options or []) if form_item.is_indicator else set()
    has_slice = bool(form_item.is_indicator and any(o in opts for o in ('sex', 'age', 'sex_age')))

    if not disagg or not isinstance(disagg, dict):
        if form_item.is_indicator and form_item.supports_disaggregation and has_slice and 'total' not in opts:
            return 'Select a disaggregation slice for this column'
        return None

    if not form_item.is_indicator:
        return 'Disaggregation applies only to numeric indicators'

    if not form_item.supports_disaggregation:
        return 'Disaggregation slice is set but this indicator unit does not support disaggregation'

    mode = disagg.get('mode')
    key = disagg.get('key')

    if mode == 'total':
        if 'total' not in opts:
            return "Template item does not allow 'total' reporting"
        return None

    if mode == 'sex':
        if 'sex' not in opts:
            return "Template item does not allow sex disaggregation"
        if not key:
            return 'Missing sex slice key'
        valid = {_slugify_sex_category(s) for s in (form_item.effective_sex_categories or [])}
        if key not in valid:
            return f'Sex slice {key!r} is not configured on this indicator'
        return None

    if mode == 'age':
        if 'age' not in opts:
            return "Template item does not allow age disaggregation"
        if not key:
            return 'Missing age slice key'
        valid = {slugify_age_group(a) for a in (form_item.effective_age_groups or [])}
        if key not in valid:
            return f'Age slice {key!r} is not configured on this indicator'
        return None

    if mode == 'sex_age':
        if 'sex_age' not in opts:
            return "Template item does not allow sex × age disaggregation"
        if not key:
            return 'Missing sex×age slice key'
        valid_keys = set()
        for sx in form_item.effective_sex_categories or []:
            sx_slug = _slugify_sex_category(sx)
            for ag in form_item.effective_age_groups or []:
                ag_slug = slugify_age_group(ag)
                valid_keys.add(f'{sx_slug}_{ag_slug}')
        if key not in valid_keys:
            return f'Sex×age slice {key!r} is not configured on this indicator'
        return None

    return f'Unknown disaggregation mode {mode!r}'


def _parse_numeric_for_indicator(raw: Any, form_item: Any) -> Optional[Any]:
    """Parse Excel cell to int/float for indicator types (aligned with form entry)."""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float):
        if raw == int(raw):
            return int(raw)
        return raw
    s = _s(raw)
    if not s:
        return None
    from app.utils.form_processing import FormItemProcessor
    cleaned = FormItemProcessor._unformat_numeric_string(s)
    if cleaned is None or cleaned == '':
        return None
    t = (form_item.type or '').strip()
    try:
        if t == 'Percentage':
            return float(cleaned)
        return int(cleaned)
    except (ValueError, TypeError):
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return None


def _accumulate_disagg_value(
    form_item: Any,
    mode: str,
    slice_key: Optional[str],
    num_val: Any,
    bucket: Dict[str, Any],
) -> Optional[str]:
    """Merge one numeric slice into bucket { 'mode': str, 'inner': dict }."""
    if 'mode' not in bucket:
        bucket['mode'] = mode
    elif bucket['mode'] != mode:
        return 'Mixed disaggregation modes for the same indicator in one row'

    inner: Dict[str, Any] = bucket.setdefault('inner', {})

    if mode == 'total':
        inner_key = 'direct' if form_item.indirect_reach else 'total'
        if inner.get(inner_key) is not None:
            return 'Duplicate total value for the same indicator in one row'
        inner[inner_key] = num_val
        return None

    if mode == 'sex':
        if not slice_key:
            return 'Missing sex slice'
        if slice_key in inner:
            return f'Duplicate sex slice {slice_key!r} for the same indicator in one row'
        inner[slice_key] = num_val
        return None

    if mode == 'age':
        if not slice_key:
            return 'Missing age slice'
        if slice_key in inner:
            return f'Duplicate age slice {slice_key!r} for the same indicator in one row'
        inner[slice_key] = num_val
        return None

    if mode == 'sex_age':
        if not slice_key:
            return 'Missing sex×age slice'
        if slice_key in inner:
            return f'Duplicate sex×age slice {slice_key!r} for the same indicator in one row'
        inner[slice_key] = num_val
        return None

    return f'Unsupported mode {mode!r}'


def _bucket_to_disagg_values(form_item: Any, bucket: Dict[str, Any]) -> Dict[str, Any]:
    """Convert accumulate bucket to `values` dict for FormData.set_disaggregated_data."""
    mode = bucket['mode']
    inner = dict(bucket.get('inner') or {})
    if form_item.indirect_reach and mode in ('sex', 'age', 'sex_age'):
        return {'direct': inner}
    return inner


def _item_uses_disagg_json(form_item: Any) -> bool:
    """True when saved FormData should use disagg_data (matches form save behaviour)."""
    if not getattr(form_item, 'is_indicator', False):
        return False
    opts = getattr(form_item, 'allowed_disaggregation_options', None) or []
    if any(o in ('sex', 'age', 'sex_age') for o in opts):
        return True
    return bool(getattr(form_item, 'indirect_reach', False))


def _coerce_mapping_disagg_mode_key(form_item: Any, disagg: Any) -> Tuple[str, Optional[str]]:
    if disagg and isinstance(disagg, dict):
        return (disagg.get('mode') or 'total', disagg.get('key'))
    return ('total', None)


def _match_entity_to_country(
    entity_name: str,
    countries_by_name: Dict[str, Any],
    ns_by_name: Dict[str, Any],
) -> Optional[Any]:
    """Try to match a KoBo entity name to a Country.

    Strategies: exact NS name, fuzzy NS name, extract country name, fuzzy country.
    """
    name = entity_name.strip()
    name_lower = name.lower()

    if name_lower in ns_by_name:
        return ns_by_name[name_lower].country

    for ns_name, ns_obj in ns_by_name.items():
        if name_lower in ns_name or ns_name in name_lower:
            return ns_obj.country

    cleaned = _NS_STRIP.sub('', name).strip().strip(',').strip()
    cleaned_lower = cleaned.lower()
    if cleaned_lower and cleaned_lower in countries_by_name:
        return countries_by_name[cleaned_lower]

    for c_name, c_obj in countries_by_name.items():
        if cleaned_lower and (cleaned_lower in c_name or c_name in cleaned_lower):
            return c_obj

    for c_name, c_obj in countries_by_name.items():
        if c_name in name_lower or name_lower in c_name:
            return c_obj

    return None


class KoboDataImportService:
    """Import submission data from KoBo Toolbox data export Excel files."""

    @classmethod
    def analyze(cls, file_bytes: bytes) -> Dict[str, Any]:
        """Analyze a KoBo data export and return its detected structure.

        Returns dict with: success, groups, skipped_columns, entity_candidates,
        total_rows, total_columns, sheet_name, etc.
        """
        if openpyxl is None:
            return {'success': False, 'message': 'openpyxl is required', 'errors': ['openpyxl not installed']}

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        except Exception as e:
            return {'success': False, 'message': f'Cannot open Excel file: {e}', 'errors': [str(e)]}

        ws = wb.worksheets[0]
        sheet_name = wb.sheetnames[0]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return {'success': False, 'message': 'File is empty', 'errors': ['No header row found']}

        headers = list(header_row)
        total_columns = len(headers)

        data_rows: List[Tuple] = []
        for row in rows_iter:
            if any(v is not None for v in row):
                data_rows.append(tuple(row))
        total_rows = len(data_rows)

        if total_rows == 0:
            return {'success': False, 'message': 'No data rows found', 'errors': ['File contains only headers']}

        groups: Dict[str, Dict] = OrderedDict()
        skipped_columns: List[Dict] = []
        ungrouped_columns: List[Dict] = []
        entity_candidates: List[Dict] = []

        for col_idx, header_val in enumerate(headers):
            header_str = _s(header_val) if header_val is not None else ''
            if not header_str:
                skipped_columns.append({
                    'index': col_idx, 'header': '', 'category': 'system', 'reason': 'Empty header',
                })
                continue

            category, reason = _classify_header(header_str)

            col_values = [row[col_idx] if col_idx < len(row) else None for row in data_rows]
            non_null_values = [v for v in col_values if v is not None and _s(v) != '']

            if category == 'data' and not non_null_values:
                pure_text = header_str.split('/')[-1] if '/' in header_str else header_str
                stripped = pure_text.strip('*').strip()
                if stripped and not any(c.isalnum() and c.islower() for c in stripped):
                    category, reason = 'note', 'Display-only label (no data)'
                elif not stripped:
                    category, reason = 'note', 'Empty display label'

            if category != 'data':
                skipped_columns.append({
                    'index': col_idx, 'header': header_str,
                    'category': category, 'reason': reason,
                })
                continue

            disagg_type = _detect_disagg(header_str)
            detected_type, detected_options = _detect_data_type(non_null_values[:200])

            non_null_count = len(non_null_values)
            unique_strs = set(_s(v) for v in non_null_values)
            unique_count = len(unique_strs)

            sample_values = []
            seen: Set[str] = set()
            for v in non_null_values[:50]:
                sv = _s(v)[:120]
                if sv and sv not in seen:
                    sample_values.append(sv)
                    seen.add(sv)
                if len(sample_values) >= 8:
                    break

            group_name, question_label = _extract_group_and_label(header_str)

            # Include all data columns by default. disagg_type is informational (sex / age / sex_age)
            # so the UI can badge KoBo breakdown columns; users may bulk-uncheck a group if redundant.
            col_info = {
                'index': col_idx,
                'header': header_str,
                'label': question_label,
                'group': group_name,
                'detected_type': detected_type,
                'detected_options': detected_options,
                'disagg_type': disagg_type,
                'non_null_count': non_null_count,
                'unique_count': unique_count,
                'sample_values': sample_values,
                'include': True,
            }

            avg_val_len = sum(len(sv) for sv in sample_values) / max(len(sample_values), 1)
            is_candidate = (
                detected_type in ('text', 'single_choice')
                and 3 <= unique_count <= total_rows * 0.3
                and all(len(sv) < 120 for sv in sample_values)
                and avg_val_len < 80
                and non_null_count >= total_rows * 0.5
                and disagg_type is None
                and len(question_label) < 100
            )
            if is_candidate:
                col_info['is_entity_candidate'] = True
                entity_candidates.append({
                    'index': col_idx,
                    'header': header_str,
                    'label': question_label,
                    'unique_count': unique_count,
                    'sample_values': sample_values[:5],
                })

            if group_name:
                if group_name not in groups:
                    groups[group_name] = {'name': group_name, 'columns': []}
                groups[group_name]['columns'].append(col_info)
            else:
                ungrouped_columns.append(col_info)

        wb.close()

        all_data_cols = []
        for g in groups.values():
            all_data_cols.extend(g['columns'])
        all_data_cols.extend(ungrouped_columns)

        vs_idx = _find_validation_status_column(headers)

        return {
            'success': True,
            'sheet_name': sheet_name,
            'total_rows': total_rows,
            'total_columns': total_columns,
            'data_column_count': len(all_data_cols),
            'skipped_column_count': len(skipped_columns),
            'groups': list(groups.values()),
            'ungrouped_columns': ungrouped_columns,
            'skipped_columns': skipped_columns,
            'entity_candidates': entity_candidates,
            'validation_status_column_index': vs_idx,
            'validation_status_header': _s(headers[vs_idx]) if vs_idx is not None else None,
        }

    @classmethod
    def extract_unique_entities(
        cls,
        file_bytes: bytes,
        entity_col_index: int,
        *,
        submission_filter: str = 'all',
        validation_status_column_index: Optional[int] = None,
    ) -> List[str]:
        """Read the file and return unique entity names from the given column.

        Respects ``submission_filter`` the same way as import/preview (requires
        ``_validation_status`` when filter is not ``all``).
        """
        if openpyxl is None:
            return []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            ws = wb.worksheets[0]
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter)
            headers = list(header_row)
            data_rows: List[list] = []
            for row in rows_iter:
                if any(v is not None for v in row):
                    data_rows.append(list(row))
            wb.close()

            val_col = validation_status_column_index
            if val_col is not None and (val_col < 0 or val_col >= len(headers)):
                val_col = None
            if val_col is None:
                val_col = _find_validation_status_column(headers)

            eligible, _, err = _eligible_row_indices(
                data_rows, submission_filter=submission_filter, validation_col=val_col,
            )
            if err:
                return []
            sf = (submission_filter or 'all').strip()
            if sf != 'all' and not eligible:
                return []

            seen: Set[str] = set()
            ordered: List[str] = []
            for ri in eligible:
                row = data_rows[ri]
                val = row[entity_col_index] if entity_col_index < len(row) else None
                if val is None:
                    continue
                s = _s(val)
                if s and s not in seen:
                    seen.add(s)
                    ordered.append(s)
            return ordered
        except Exception:
            return []

    @classmethod
    def try_match_entities(cls, entity_names: List[str]) -> Dict[str, Any]:
        """Try to match a list of entity names to countries.

        Returns dict mapping entity_name -> {country_id, country_name, matched}
        """
        from app.models.core import Country
        from app.models.organization import NationalSociety

        countries = Country.query.filter_by(status='Active').all()
        national_societies = NationalSociety.query.filter_by(is_active=True).all()

        countries_by_name = {c.name.lower(): c for c in countries}
        ns_by_name = {ns.name.lower(): ns for ns in national_societies}

        result = {}
        for name in entity_names:
            country = _match_entity_to_country(name, countries_by_name, ns_by_name)
            if country:
                result[name] = {
                    'country_id': country.id,
                    'country_name': country.name,
                    'matched': True,
                }
            else:
                result[name] = {
                    'country_id': None,
                    'country_name': None,
                    'matched': False,
                }
        return result

    @classmethod
    def map_columns_to_template(
        cls,
        kobo_columns: List[Dict[str, Any]],
        template_items: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Auto-map KoBo data columns to existing template items by label similarity.

        kobo_columns: [{index, header, label, group, detected_type, ...}]
        template_items: [{id, label, type, section_id, section_name}]

        Returns a list of mapping entries:
            [{col_index, col_label, col_group, item_id, item_label, section_name, confidence}]
        """
        import difflib

        mappings = []
        item_labels_lower = [(it, (it.get('label') or '').lower().strip()) for it in template_items]
        used_item_ids: Set[int] = set()

        for col in kobo_columns:
            col_label = (col.get('label') or '').strip()
            col_label_lower = col_label.lower()
            if not col_label_lower:
                mappings.append({
                    'col_index': col['index'],
                    'col_label': col_label,
                    'col_group': col.get('group'),
                    'item_id': None,
                    'item_label': None,
                    'section_name': None,
                    'confidence': 0,
                })
                continue

            best_match = None
            best_ratio = 0.0

            for item, item_lbl in item_labels_lower:
                if item['id'] in used_item_ids:
                    continue
                if not item_lbl:
                    continue

                if col_label_lower == item_lbl:
                    best_match = item
                    best_ratio = 1.0
                    break

                ratio = difflib.SequenceMatcher(None, col_label_lower, item_lbl).ratio()
                if ratio > best_ratio:
                    best_ratio = ratio
                    best_match = item

            threshold = 0.55
            if best_match and best_ratio >= threshold:
                used_item_ids.add(best_match['id'])
                mappings.append({
                    'col_index': col['index'],
                    'col_label': col_label,
                    'col_group': col.get('group'),
                    'item_id': best_match['id'],
                    'item_label': best_match.get('label', ''),
                    'section_name': best_match.get('section_name', ''),
                    'confidence': round(best_ratio, 2),
                })
            else:
                mappings.append({
                    'col_index': col['index'],
                    'col_label': col_label,
                    'col_group': col.get('group'),
                    'item_id': None,
                    'item_label': None,
                    'section_name': None,
                    'confidence': round(best_ratio, 2) if best_match else 0,
                })

        return mappings

    @classmethod
    def generate_preview(
        cls,
        file_bytes: bytes,
        *,
        entity_column_index: Optional[int],
        columns_to_import: List[int],
        entity_mapping: Dict[str, Any],
        duplicate_strategy: str = 'latest',
        submission_time_column_index: Optional[int] = None,
        submission_filter: str = 'all',
        validation_status_column_index: Optional[int] = None,
        max_rows: int = 100,
        existing_template_id: Optional[int] = None,
        column_to_item_mapping: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """Build a preview table (column defs + row data) for the AG Grid.

        Returns {success, columns: [{field, headerName, headerGroup}], rows: [{...}],
                 total_rows, preview_rows, duplicate_info}.
        """
        if openpyxl is None:
            return {'success': False, 'message': 'openpyxl is required'}

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        except Exception as e:
            return {'success': False, 'message': f'Cannot open file: {e}'}

        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        headers = list(next(rows_iter))

        data_rows: List[list] = []
        for row in rows_iter:
            if any(v is not None for v in row):
                data_rows.append(list(row))
        wb.close()

        if not data_rows:
            return {'success': False, 'message': 'No data rows'}

        entity_col = entity_column_index
        sub_time_col = submission_time_column_index

        if entity_col is None or entity_col < 0:
            return {'success': False, 'message': 'Entity column is required for preview'}

        kept_indices, row_stats = _resolve_import_row_indices(
            data_rows,
            headers,
            entity_column_index=entity_col,
            duplicate_strategy=duplicate_strategy,
            submission_filter=submission_filter or 'all',
            submission_time_column_index=sub_time_col,
            validation_status_column_index=validation_status_column_index,
        )
        if row_stats.get('error'):
            return {'success': False, 'message': row_stats['error']}

        duplicate_count = row_stats['duplicate_count']
        duplicate_details = row_stats['duplicate_details']
        total_after_dedup = len(kept_indices)

        preview_map = _normalize_column_to_item_mapping(column_to_item_mapping or {})
        items_by_id: Dict[int, Any] = {}
        if existing_template_id and preview_map:
            from app.models import FormItem
            _ids = {spec['item_id'] for spec in preview_map.values()}
            if _ids:
                items_by_id = {i.id: i for i in FormItem.query.filter(FormItem.id.in_(_ids)).all()}

        # Build column definitions
        cols_set = set(columns_to_import)
        if entity_col is not None:
            cols_set.add(entity_col)

        col_defs: List[Dict[str, Any]] = []
        col_order: List[int] = []

        if entity_col is not None:
            _, elabel = _extract_group_and_label(_s(headers[entity_col]))
            col_defs.append({
                'field': f'c{entity_col}',
                'headerName': elabel,
                'pinned': 'left',
                'minWidth': 180,
                '_colIndex': entity_col,
            })
            # Add matched country column
            col_defs.append({
                'field': '_country',
                'headerName': 'Matched Country',
                'pinned': 'left',
                'minWidth': 140,
            })
            col_order.append(entity_col)

        for ci in columns_to_import:
            if ci == entity_col:
                continue
            h = _s(headers[ci]) if ci < len(headers) else ''
            if not h:
                continue
            group_name, label = _extract_group_and_label(h)
            header_name = label
            spec = preview_map.get(ci)
            if spec:
                fi = items_by_id.get(spec['item_id'])
                if fi:
                    ilab = fi.label or f'Item #{fi.id}'
                    dis = spec.get('disagg')
                    if dis and isinstance(dis, dict):
                        mode = dis.get('mode') or 'total'
                        key = dis.get('key')
                        if mode == 'total' or mode is None:
                            header_name = f'{ilab} (total)'
                        elif key:
                            header_name = f'{ilab} ({_disagg_slice_label(mode, key, fi)})'
                        else:
                            header_name = ilab
                    else:
                        header_name = ilab
            col_defs.append({
                'field': f'c{ci}',
                'headerName': header_name,
                'columnGroupShow': None,
                '_group': group_name,
                '_colIndex': ci,
            })
            col_order.append(ci)

        # Build row data (limited to max_rows)
        preview_indices = kept_indices[:max_rows]
        row_data: List[Dict[str, Any]] = []

        for ri in preview_indices:
            dr = data_rows[ri]
            row_obj: Dict[str, Any] = {'_rowIndex': ri}

            # Entity and country match
            if entity_col is not None and entity_col < len(dr):
                ename = _s(dr[entity_col])
                row_obj[f'c{entity_col}'] = ename
                info = entity_mapping.get(ename, {})
                if info.get('matched'):
                    row_obj['_country'] = info.get('country_name', '')
                else:
                    row_obj['_country'] = '(unmatched)'

            for ci in col_order:
                if ci == entity_col:
                    continue
                val = dr[ci] if ci < len(dr) else None
                if val is None:
                    row_obj[f'c{ci}'] = None
                elif isinstance(val, (datetime, date)):
                    row_obj[f'c{ci}'] = val.isoformat()
                elif isinstance(val, float):
                    row_obj[f'c{ci}'] = int(val) if val == int(val) else val
                elif isinstance(val, int):
                    row_obj[f'c{ci}'] = val
                else:
                    row_obj[f'c{ci}'] = _s(val)

            row_data.append(row_obj)

        # Group column defs by header group for AG Grid column groups
        grouped_defs: List[Dict] = []
        current_group: Optional[str] = None
        current_children: List[Dict] = []

        pinned = [cd for cd in col_defs if cd.get('pinned')]
        unpinned = [cd for cd in col_defs if not cd.get('pinned')]

        for cd in pinned:
            clean = {k: v for k, v in cd.items() if not k.startswith('_')}
            grouped_defs.append(clean)

        for cd in unpinned:
            g = cd.get('_group')
            clean = {k: v for k, v in cd.items() if not k.startswith('_')}
            if g and g == current_group:
                current_children.append(clean)
            else:
                if current_children:
                    grouped_defs.append({
                        'headerName': current_group,
                        'children': current_children,
                    })
                if g:
                    current_group = g
                    current_children = [clean]
                else:
                    current_group = None
                    current_children = []
                    grouped_defs.append(clean)
        if current_children:
            grouped_defs.append({
                'headerName': current_group,
                'children': current_children,
            })

        return {
            'success': True,
            'columns': grouped_defs,
            'rows': row_data,
            'total_rows': len(data_rows),
            'preview_rows': len(row_data),
            'after_dedup': total_after_dedup,
            'duplicates_removed': duplicate_count,
            'duplicate_details': duplicate_details,
            'rows_excluded_not_approved': row_stats.get('rows_excluded_not_approved', 0),
            'rows_excluded_by_submission_filter': row_stats.get('rows_excluded_by_submission_filter', 0),
        }

    @classmethod
    def execute_import(
        cls,
        file_bytes: bytes,
        config: Dict[str, Any],
    ) -> Dict[str, Any]:
        """Execute the full import: create template, assignments, and data.

        Config keys:
            template_name: str
            period_name: str
            entity_column_index: int
            columns_to_import: list[int]
            entity_mapping: dict[str, {country_id: int}]
            create_template: bool (default True)
            import_data: bool (default True)
            owned_by: int (user id)
            submission_time_column_index: int | None
            submission_filter: 'all' | 'approved_only' | 'exclude_rejected' | 'approved_or_on_hold'
                | 'on_hold_only' | 'not_validated_only' | 'draft_only'
            validation_status_column_index: int | None
            duplicate_strategy: 'latest' | 'first' | 'all' | 'latest_approved' | 'first_approved'
        """
        from flask_login import current_user
        from app import db
        from app.models import FormTemplate, FormSection, FormItem, FormTemplateVersion
        from app.models.assignments import AssignedForm, AssignmentEntityStatus
        from app.models.forms import FormData
        from app.models.core import Country
        from app.utils.datetime_helpers import utcnow

        if openpyxl is None:
            return {'success': False, 'message': 'openpyxl is required'}

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        except Exception as e:
            return {'success': False, 'message': f'Cannot open file: {e}'}

        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        headers = list(next(rows_iter))

        data_rows = []
        for row in rows_iter:
            if any(v is not None for v in row):
                data_rows.append(list(row))
        wb.close()

        if not data_rows:
            return {'success': False, 'message': 'No data rows to import'}

        template_name = config.get('template_name', 'KoBo Data Import')
        period_name = config.get('period_name', 'Imported')
        entity_col_idx = config.get('entity_column_index')
        cols_to_import = config.get('columns_to_import', [])
        entity_mapping = config.get('entity_mapping', {})
        create_template = config.get('create_template', True)
        import_data = config.get('import_data', True)
        owned_by = config.get('owned_by') or (current_user.id if current_user.is_authenticated else None)
        dup_strategy = config.get('duplicate_strategy', 'latest')
        sub_time_col = config.get('submission_time_column_index')
        submission_filter = config.get('submission_filter', 'all')
        validation_col_cfg = config.get('validation_status_column_index')

        if not owned_by:
            return {'success': False, 'message': 'Cannot determine user'}

        if not cols_to_import:
            return {'success': False, 'message': 'No columns selected for import'}

        warnings: List[str] = []
        counts = {
            'sections': 0, 'items': 0, 'assignments': 0,
            'data_entries': 0, 'skipped_entities': 0, 'duplicate_rows_skipped': 0,
            'rows_excluded_not_approved': 0,
            'rows_excluded_by_submission_filter': 0,
        }

        col_infos = []
        for ci in cols_to_import:
            if ci == entity_col_idx:
                continue
            h = _s(headers[ci]) if ci < len(headers) else ''
            if not h:
                continue
            group_name, label = _extract_group_and_label(h)
            vals = [row[ci] if ci < len(row) else None for row in data_rows[:200]]
            non_null = [v for v in vals if v is not None and _s(v) != '']
            dt, opts = _detect_data_type(non_null)
            col_infos.append({
                'col_index': ci,
                'header': h,
                'group': group_name,
                'label': label,
                'detected_type': dt,
                'options': opts,
            })

        if not col_infos:
            return {'success': False, 'message': 'No valid columns to import after filtering'}

        existing_template_id = config.get('existing_template_id')
        column_to_item_mapping = config.get('column_to_item_mapping', {})

        template = None
        version = None
        col_to_item_id: Dict[int, int] = {}
        col_mapping_struct: Dict[int, Dict[str, Any]] = {}
        import_items_by_id: Dict[int, Any] = {}

        if existing_template_id:
            template = FormTemplate.query.get(existing_template_id)
            if not template:
                return {'success': False, 'message': f'Template {existing_template_id} not found'}

            version = template.published_version
            if not version:
                version = template.versions.order_by(
                    FormTemplateVersion.version_number.desc()
                ).first()
            if not version:
                return {'success': False, 'message': 'Selected template has no version'}

            col_mapping_struct = _normalize_column_to_item_mapping(column_to_item_mapping)
            for ci, spec in col_mapping_struct.items():
                col_to_item_id[ci] = spec['item_id']

            if not col_mapping_struct:
                return {'success': False, 'message': 'No column-to-item mappings provided for existing template'}

            _item_ids = {spec['item_id'] for spec in col_mapping_struct.values()}
            import_items_by_id = {i.id: i for i in FormItem.query.filter(FormItem.id.in_(_item_ids)).all()}
            for ci, spec in col_mapping_struct.items():
                fi = import_items_by_id.get(spec['item_id'])
                if not fi:
                    return {'success': False, 'message': f'Mapped item {spec["item_id"]} not found'}
                v_err = _validate_disagg_for_item(fi, spec.get('disagg'))
                if v_err:
                    return {'success': False, 'message': f'Column mapping (column index {ci}): {v_err}'}

            counts['items'] = len(col_mapping_struct)

        elif create_template:
            template = FormTemplate(created_by=owned_by, owned_by=owned_by)
            db.session.add(template)
            db.session.flush()

            now = utcnow()
            version = FormTemplateVersion(
                template_id=template.id,
                version_number=1,
                status='draft',
                name=template_name,
                description=f'Imported from KoBo data export',
                add_to_self_report=False,
                display_order_visible=False,
                is_paginated=True,
                enable_export_pdf=False,
                enable_export_excel=True,
                enable_import_excel=True,
                enable_ai_validation=False,
                created_by=owned_by,
                updated_by=owned_by,
                created_at=now,
                updated_at=now,
            )
            db.session.add(version)
            db.session.flush()

            grouped: Dict[Optional[str], List[Dict]] = OrderedDict()
            for ci in col_infos:
                g = ci['group']
                grouped.setdefault(g, []).append(ci)

            section_order = 0.0
            for group_name, items in grouped.items():
                section_order += 1.0
                sec_name = group_name or 'General'
                section = FormSection(
                    template_id=template.id,
                    version_id=version.id,
                    name=sec_name,
                    order=section_order,
                    parent_section_id=None,
                    page_id=None,
                    section_type='standard',
                )
                db.session.add(section)
                db.session.flush()
                counts['sections'] += 1

                for item_idx, ci_info in enumerate(items):
                    item_type = 'question'
                    q_type = ci_info['detected_type']
                    options_json = None

                    if q_type == 'single_choice' and ci_info.get('options'):
                        options_json = [
                            {'value': opt, 'label': opt}
                            for opt in ci_info['options']
                        ]
                    elif q_type == 'yesno':
                        q_type = 'yesno'

                    item = FormItem(
                        section_id=section.id,
                        template_id=template.id,
                        version_id=version.id,
                        item_type=item_type,
                        label=ci_info['label'],
                        order=float(item_idx + 1),
                        type=q_type,
                        options_json=options_json,
                        config={'is_required': False},
                    )
                    db.session.add(item)
                    db.session.flush()
                    counts['items'] += 1
                    col_to_item_id[ci_info['col_index']] = item.id

            db.session.flush()

        if import_data and template and entity_col_idx is not None:
            kept_row_indices, row_stats = _resolve_import_row_indices(
                data_rows,
                headers,
                entity_column_index=entity_col_idx,
                duplicate_strategy=dup_strategy,
                submission_filter=submission_filter or 'all',
                submission_time_column_index=sub_time_col,
                validation_status_column_index=validation_col_cfg,
            )
            if row_stats.get('error'):
                return {'success': False, 'message': row_stats['error']}
            rx = row_stats.get('rows_excluded_by_submission_filter', row_stats.get('rows_excluded_not_approved', 0))
            counts['rows_excluded_not_approved'] = rx
            counts['rows_excluded_by_submission_filter'] = rx
            counts['duplicate_rows_skipped'] = row_stats.get('duplicate_count', 0)

            entity_rows: Dict[str, List[int]] = OrderedDict()
            for row_idx in kept_row_indices:
                row = data_rows[row_idx]
                ename = _s(row[entity_col_idx]) if entity_col_idx < len(row) else ''
                if not ename:
                    continue
                entity_rows.setdefault(ename, []).append(row_idx)

            assigned_form = AssignedForm(
                template_id=template.id,
                period_name=period_name,
                is_active=True,
            )
            db.session.add(assigned_form)
            db.session.flush()

            for ename, row_indices in entity_rows.items():
                mapping_info = entity_mapping.get(ename, {})
                country_id = mapping_info.get('country_id')

                if not country_id:
                    counts['skipped_entities'] += 1
                    warnings.append(f"Skipped '{ename}': no country match")
                    continue

                country = Country.query.get(country_id)
                if not country:
                    counts['skipped_entities'] += 1
                    warnings.append(f"Skipped '{ename}': country ID {country_id} not found")
                    continue

                aes = AssignmentEntityStatus(
                    assigned_form_id=assigned_form.id,
                    entity_type='country',
                    entity_id=country.id,
                    status='Pending',
                )
                db.session.add(aes)
                db.session.flush()
                counts['assignments'] += 1

                for row_idx in row_indices:
                    row = data_rows[row_idx]
                    if col_mapping_struct:
                        merge_disagg: Dict[int, Dict[str, Any]] = {}
                        simple_by_item: Dict[int, str] = {}

                        for col_idx, spec in col_mapping_struct.items():
                            item_id = spec['item_id']
                            fi = import_items_by_id.get(item_id)
                            if not fi:
                                continue
                            val = row[col_idx] if col_idx < len(row) else None
                            if val is None:
                                continue

                            if _item_uses_disagg_json(fi):
                                if item_id in simple_by_item:
                                    simple_by_item.pop(item_id, None)
                                    warnings.append(
                                        f"Row {row_idx + 2} ({ename}): replaced simple column value with "
                                        f"disaggregated data for item {item_id}"
                                    )
                                mode, slice_key = _coerce_mapping_disagg_mode_key(fi, spec.get('disagg'))
                                num_val = _parse_numeric_for_indicator(val, fi)
                                if num_val is None:
                                    continue
                                bucket = merge_disagg.setdefault(item_id, {})
                                acc_err = _accumulate_disagg_value(
                                    fi, mode, slice_key, num_val, bucket,
                                )
                                if acc_err:
                                    warnings.append(
                                        f"Row {row_idx + 2} ({ename}), column {col_idx}: {acc_err}"
                                    )
                                continue

                            str_val = _s(val)
                            if not str_val:
                                continue
                            if isinstance(val, (datetime, date)):
                                str_val = val.isoformat()
                            elif isinstance(val, float):
                                str_val = str(int(val)) if val == int(val) else str(val)
                            elif isinstance(val, int):
                                str_val = str(val)
                            if len(str_val) > 255:
                                str_val = str_val[:255]
                            if item_id in merge_disagg:
                                warnings.append(
                                    f"Row {row_idx + 2} ({ename}): skipped non-disaggregated column for "
                                    f"item {item_id} (row already has disaggregated values for this item)"
                                )
                                continue
                            if item_id in simple_by_item:
                                warnings.append(
                                    f"Row {row_idx + 2} ({ename}): duplicate simple columns for item "
                                    f"{item_id}; keeping the last value"
                                )
                            simple_by_item[item_id] = str_val

                        for item_id, bucket in merge_disagg.items():
                            if not bucket or 'mode' not in bucket:
                                continue
                            fi = import_items_by_id[item_id]
                            values_payload = _bucket_to_disagg_values(fi, bucket)
                            fd = FormData(
                                assignment_entity_status_id=aes.id,
                                form_item_id=item_id,
                            )
                            fd.set_disaggregated_data(bucket['mode'], values_payload)
                            db.session.add(fd)
                            counts['data_entries'] += 1

                        for item_id, str_val in simple_by_item.items():
                            if item_id in merge_disagg:
                                continue
                            fd = FormData(
                                assignment_entity_status_id=aes.id,
                                form_item_id=item_id,
                                value=str_val,
                            )
                            db.session.add(fd)
                            counts['data_entries'] += 1
                    else:
                        for col_idx, item_id in col_to_item_id.items():
                            val = row[col_idx] if col_idx < len(row) else None
                            if val is None:
                                continue

                            str_val = _s(val)
                            if not str_val:
                                continue

                            if isinstance(val, (datetime, date)):
                                str_val = val.isoformat()
                            elif isinstance(val, float):
                                if val == int(val):
                                    str_val = str(int(val))
                                else:
                                    str_val = str(val)
                            elif isinstance(val, int):
                                str_val = str(val)

                            if len(str_val) > 255:
                                str_val = str_val[:255]

                            fd = FormData(
                                assignment_entity_status_id=aes.id,
                                form_item_id=item_id,
                                value=str_val,
                            )
                            db.session.add(fd)
                            counts['data_entries'] += 1

            db.session.flush()

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"KoBo data import commit failed: {e}", exc_info=True)
            return {'success': False, 'message': f'Database error: {e}'}

        return {
            'success': True,
            'message': (
                f"Import complete: {counts['sections']} sections, {counts['items']} items, "
                f"{counts['assignments']} entities, {counts['data_entries']} data entries"
            ),
            'template_id': template.id if template else None,
            'version_id': version.id if version else None,
            'counts': counts,
            'warnings': warnings,
        }
