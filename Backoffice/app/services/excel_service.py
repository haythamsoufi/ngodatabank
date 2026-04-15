from __future__ import annotations

import io
import json
from contextlib import suppress
from typing import Dict, Tuple, Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from app.services import get_formdata_map
from app.services.form_data_service import FormDataService
from app.utils.route_helpers import get_unified_form_item_id
from app.models.forms import FormData


class ExcelService:
    """Utility helpers for assignment Excel export/import."""

    # New format: form-like layout without table headers
    # Column A: item_id (hidden)
    # Column B: Label/Field name
    # Column C: Value/Input area
    # Column D: Disaggregation Mode (if applicable)
    # Column E: Disaggregation Values (if applicable)

    @staticmethod
    def load_workbook(file_storage) -> openpyxl.Workbook:
        """Load an Excel workbook from an uploaded file."""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_storage.read()), data_only=True)
        except Exception as exc:  # pragma: no cover - openpyxl specific
            raise ValueError(f"Unable to read the Excel file: {exc}") from exc
        finally:
            with suppress(Exception):
                file_storage.stream.seek(0)
        return workbook

    @classmethod
    def build_assignment_workbook(cls, assignment_entity_status) -> Tuple[io.BytesIO, str]:
        """Create an Excel workbook representing the assignment."""
        aes = assignment_entity_status
        template = aes.assigned_form.template
        version = template.published_version or template.versions.order_by('created_at').first()
        pages = list(template.pages.order_by("order")) if version and version.is_paginated else [None]

        # Resolve template variables for Excel export
        from app.services.variable_resolution_service import VariableResolutionService

        resolved_variables = {}
        variable_configs = {}
        if version:
            variable_configs = version.variables or {}
            resolved_variables = VariableResolutionService.resolve_variables(
                version,
                aes
            )

        # Get full FormData entries to access disagg_data
        form_data_entries = FormData.query.filter_by(
            assignment_entity_status_id=aes.id
        ).all()

        # Create a map of form_item_id -> FormData entry for quick lookup
        entries_map = {entry.form_item_id: entry for entry in form_data_entries}

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Data Entry"

        # Form-like styling
        section_fill = PatternFill(start_color="FFCC0000", end_color="FFCC0000", fill_type="solid")
        section_font = Font(name="Arial", size=14, bold=True, color="FFFFFFFF")
        label_font = Font(name="Arial", size=11, bold=True, color="FF111827")
        value_fill = PatternFill(start_color="FFFFF9E6", end_color="FFFFF9E6", fill_type="solid")
        value_font = Font(name="Arial", size=11, color="FF111827")
        disagg_label_font = Font(name="Arial", size=10, bold=True, color="FF4B5563")
        disagg_value_fill = PatternFill(start_color="FFE6F3FF", end_color="FFE6F3FF", fill_type="solid")
        left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center")

        first_sheet_handled = False
        for page in pages:
            if not first_sheet_handled:
                ws = sheet
                first_sheet_handled = True
            else:
                ws = workbook.create_sheet((page.name if page else "Data Entry")[:31])

            # Write page title if paginated
            if page:
                ws.cell(row=1, column=2).value = page.name
                ws.cell(row=1, column=2).font = Font(name="Arial", size=16, bold=True, color="FFCC0000")
                ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
                current_row = 3
            else:
                current_row = 1

            page_sections = (
                template.sections.filter_by(page_id=page.id).order_by("order").all()
                if page else template.sections.filter_by(page_id=None).order_by("order").all()
            )

            for section in page_sections:
                current_row = cls._write_section(ws, section, entries_map, value_fill, section_fill,
                                                 section_font, label_font, value_font, disagg_label_font,
                                                 disagg_value_fill, left_align, current_row,
                                                 resolved_variables, variable_configs)

            # Hide item_id column
            ws.column_dimensions["A"].hidden = True

            # Set column widths
            ws.column_dimensions["B"].width = 25  # Label column 1 (part of merged label)
            ws.column_dimensions["C"].width = 25  # Label column 2 (part of merged label)
            ws.column_dimensions["D"].width = 30  # Value column
            ws.column_dimensions["E"].width = 20  # Disaggregation Mode
            ws.column_dimensions["F"].width = 40  # Disaggregation Values

            # Set row heights for better spacing
            for row in range(1, ws.max_row + 1):
                ws.row_dimensions[row].height = 20

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        filename = f"data_entry_{aes.country.iso3}_{aes.assigned_form.period_name.replace(' ', '_')}.xlsx"
        return output, filename

    @staticmethod
    def _calculate_label_rows_needed(label_text, max_chars_per_row=40, min_rows=1, max_rows=5):
        """Calculate how many rows are needed for a label based on its length.

        Args:
            label_text: The label text to measure
            max_chars_per_row: Maximum characters per row (approximate, more conservative)
            min_rows: Minimum rows to use (even for short labels)
            max_rows: Maximum rows to merge (cap to prevent excessive merging)

        Returns:
            Number of rows to merge (between min_rows and max_rows)
        """
        if not label_text:
            return min_rows

        # Count characters (accounting for spaces and punctuation)
        char_count = len(label_text)

        # Use a more conservative estimate - columns B+C are 25 chars each = 50 total
        # But we want to account for word wrapping and readability
        # Use 40 chars per row as a safer estimate to ensure proper wrapping

        # Calculate estimated rows needed
        # Add 0.5 to round up more aggressively for values close to threshold
        estimated_rows = max(1, int((char_count / max_chars_per_row) + 0.5))

        # Always use at least 2 rows if label is longer than 35 chars (close to threshold)
        # This ensures labels near the threshold get merged
        if char_count > 35 and estimated_rows == 1:
            estimated_rows = 2

        # Clamp between min and max
        return max(min_rows, min(estimated_rows, max_rows))

    @staticmethod
    def _write_section(ws, section, entries_map, value_fill, section_fill, section_font,
                      label_font, value_font, disagg_label_font, disagg_value_fill,
                      left_align, start_row, resolved_variables=None, variable_configs=None):
        """Write a section in form-like format with spacing between items."""
        # Resolve variables in section name
        section_name = getattr(section, 'display_name', None) or section.name
        if resolved_variables and section_name:
            try:
                from app.services.variable_resolution_service import VariableResolutionService
                section_name = VariableResolutionService.replace_variables_in_text(
                    section_name,
                    resolved_variables,
                    variable_configs
                )
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(
                    f"Error resolving variables in section name for section {section.id}: {e}",
                    exc_info=True
                )
        current_row = start_row

        # Write section header
        ws.cell(row=current_row, column=2).value = section_name
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=5)
        sec_cell = ws.cell(row=current_row, column=2)
        sec_cell.fill = section_fill
        sec_cell.font = section_font
        sec_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 25
        current_row += 1

        # Add spacing after section header
        current_row += 1

        for item in section.form_items.order_by("order").all():
            if item.is_document_field:
                continue
            item_id = get_unified_form_item_id(item)
            if not item_id:
                continue

            # Get the FormData entry if it exists
            entry = entries_map.get(item_id)

            # Get value and disaggregation data
            value = ""
            disagg_mode = ""
            disagg_values = ""

            # Check if this is a matrix item
            is_matrix = hasattr(item, 'item_type') and str(item.item_type).lower() == 'matrix'

            if entry:
                # Check for disaggregation data first (for indicators with sex/age disaggregation)
                if entry.disagg_data and isinstance(entry.disagg_data, dict):
                    # Check if it's disaggregation format (has 'mode' and 'values' keys)
                    if 'mode' in entry.disagg_data and 'values' in entry.disagg_data:
                        disagg_mode = entry.disagg_data.get('mode', '')
                        values = entry.disagg_data.get('values', {})
                        if values and isinstance(values, dict):
                            # Only export if values dict is not empty
                            if values:
                                disagg_values = json.dumps(values, ensure_ascii=False)
                    elif is_matrix:
                        # Matrix data - store as JSON in disagg_values for export
                        # Matrix data doesn't have mode/values structure
                        disagg_values = json.dumps(entry.disagg_data, ensure_ascii=False)
                        # Mark as matrix type for display
                        disagg_mode = 'matrix'

                # Otherwise use simple value (only if no disaggregation mode and not matrix)
                if not disagg_mode and not is_matrix and entry.value:
                    value = str(entry.value)

            # Write item_id in hidden column A (in first row of this item)
            ws.cell(row=current_row, column=1).value = item_id

            # Build label text - use display_label if available, otherwise label
            item_label = getattr(item, 'display_label', None) or item.label
            # Resolve variables in label
            if resolved_variables and item_label:
                try:
                    from app.services.variable_resolution_service import VariableResolutionService
                    item_label = VariableResolutionService.replace_variables_in_text(
                        item_label,
                        resolved_variables,
                        variable_configs
                    )
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(
                        f"Error resolving variables in label for form_item {item.id}: {e}",
                        exc_info=True
                    )

            label_text = f"{item.order}. {item_label}"
            if hasattr(item, 'unit') and item.unit:
                label_text += f" ({item.unit})"

            # Calculate how many rows to merge based on label length
            # Use more conservative threshold to ensure labels near limit get merged
            rows_needed = ExcelService._calculate_label_rows_needed(label_text, max_chars_per_row=40, min_rows=1, max_rows=5)

            # Write label spanning 2 columns (B and C) and merge vertically based on length
            label_start_row = current_row
            label_end_row = current_row + rows_needed - 1

            # Merge cells for label (columns B and C, multiple rows if needed)
            # Always merge horizontally (B and C), and vertically if rows_needed > 1
            if rows_needed > 1:
                # Merge vertically across multiple rows
                ws.merge_cells(start_row=label_start_row, start_column=2,
                              end_row=label_end_row, end_column=3)
            else:
                # Still merge horizontally (B and C) even for single row
                ws.merge_cells(start_row=label_start_row, start_column=2,
                              end_row=label_start_row, end_column=3)

            # Write label in merged cell
            label_cell = ws.cell(row=label_start_row, column=2)
            label_cell.value = label_text
            label_cell.font = label_font
            label_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # Set row heights for merged label rows
            for row_idx in range(label_start_row, label_end_row + 1):
                ws.row_dimensions[row_idx].height = 20

            # Write value in column D (shifted right by 2 columns for label)
            # Value stays in a single cell (no vertical merging)
            value_cell = ws.cell(row=label_start_row, column=4)
            value_cell.value = value
            value_cell.font = value_font
            value_cell.fill = value_fill
            value_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Write disaggregation data if present (in columns E and F, single cells)
            if disagg_mode:
                disagg_mode_cell = ws.cell(row=label_start_row, column=5)
                disagg_mode_cell.value = f"Mode: {disagg_mode}"
                disagg_mode_cell.font = disagg_label_font
                disagg_mode_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                disagg_values_cell = ws.cell(row=label_start_row, column=6)
                disagg_values_cell.value = disagg_values
                disagg_values_cell.font = value_font
                disagg_values_cell.fill = disagg_value_fill
                disagg_values_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Move to next item (after the merged rows)
            current_row = label_end_row + 1

            # Add spacing between items (2 empty rows)
            current_row += 2

        # Add extra spacing after section
        current_row += 1

        return current_row

    @classmethod
    def extract_field_values(cls, workbook) -> Dict[int, Dict[str, Any]]:
        """Extract field values from an uploaded workbook.

        Supports both old table format and new form-like format.

        Returns:
            Dict mapping form_item_id to a dict with:
            - 'value': simple value (str or None)
            - 'disagg_data': disaggregation data dict with 'mode' and 'values' (or None)
        """
        field_data: Dict[int, Dict[str, Any]] = {}
        valid_sheet_found = False

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            try:
                # Try to detect format by checking first row
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
                if not first_row:
                    continue

                # Check if it's old table format (has headers) or new form format
                is_table_format = False
                if first_row and len(first_row) > 0:
                    first_cell = first_row[0]
                    # Old format has headers like "item_id", "Section", etc.
                    if isinstance(first_cell, str) and first_cell.lower() in ('item_id', 'item id'):
                        is_table_format = True

                valid_sheet_found = True

                if is_table_format:
                    # Old table format - skip header row
                    start_row = 2
                else:
                    # New form format - start from row 1
                    start_row = 1

                # Iterate through all rows
                for row in sheet.iter_rows(min_row=start_row, values_only=True):
                    if len(row) < 3:
                        continue

                    # Column A: item_id (hidden)
                    item_id = row[0]
                    if item_id is None:
                        continue

                    try:
                        item_id = int(item_id)
                    except (ValueError, TypeError):
                        # Skip rows without valid item_id
                        continue

                    # Columns B & C: Label (merged, we don't need this for import)
                    # Column D (index 3): Value
                    value = row[3] if len(row) > 3 else None

                    # Column E (index 4): Disaggregation Mode (optional)
                    # Column F (index 5): Disaggregation Values (optional)
                    disagg_mode = None
                    disagg_values_str = None

                    if len(row) > 4:
                        mode_cell = row[4]  # Column E
                        if mode_cell and isinstance(mode_cell, str) and mode_cell.startswith('Mode:'):
                            # Extract mode from "Mode: sex" format
                            disagg_mode = mode_cell.replace('Mode:', '').strip()
                            disagg_values_str = row[5] if len(row) > 5 else None  # Column F

                    # Parse disaggregation data if present
                    disagg_data = None
                    if disagg_mode and disagg_values_str:
                        try:
                            disagg_values = json.loads(disagg_values_str) if isinstance(disagg_values_str, str) else disagg_values_str
                            if disagg_values:
                                disagg_data = {
                                    'mode': str(disagg_mode).strip(),
                                    'values': disagg_values
                                }
                        except (json.JSONDecodeError, TypeError) as e:
                            # If JSON parsing fails, log but continue
                            import logging
                            logger = logging.getLogger(__name__)
                            logger.warning(f"Failed to parse disaggregation data for item {item_id}: {e}")

                    # Store data - handle empty strings as None
                    value_str = None
                    if value is not None:
                        value_str = str(value).strip()
                        if not value_str or value_str.lower() in ('none', 'null', ''):
                            value_str = None

                    # Check if value is actually a disaggregation mode (misplaced data)
                    # This can happen if user edited the Excel incorrectly
                    if value_str and isinstance(value_str, str) and value_str.startswith('Mode:'):
                        # Value cell contains mode - skip this row or handle as error
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.warning(f"Item {item_id}: Value cell contains 'Mode:' - possible Excel format issue")
                        # Don't store this as it's likely a formatting error
                        continue

                    # Store data - always store entry if we have item_id
                    # This allows clearing values by leaving them empty
                    field_data[item_id] = {
                        'value': value_str,
                        'disagg_data': disagg_data
                    }

            except StopIteration:
                # Empty sheet, skip
                continue
            except Exception as e:
                # Log error but continue with other sheets
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Error processing sheet {sheet_name}: {e}")
                continue

        if not valid_sheet_found:
            raise ValueError("The Excel file does not contain a valid data sheet.")

        if not field_data:
            raise ValueError("The Excel file contains no valid data rows with item_id values.")

        return field_data

    @classmethod
    def import_assignment_data(cls, assignment_entity_status, workbook):
        """Persist workbook values into the database, including disaggregation data."""
        field_data = cls.extract_field_values(workbook)
        return cls._bulk_save_fields_with_disagg(assignment_entity_status, field_data)

    @classmethod
    def _bulk_save_fields_with_disagg(cls, assignment_entity_status, field_data: Dict[int, Dict[str, Any]]) -> Dict[str, Any]:
        """Save multiple fields at once, handling both simple values and disaggregation data.

        Args:
            assignment_entity_status: AssignmentEntityStatus or PublicSubmission object
            field_data: Dict mapping form_item_id to dict with 'value' and/or 'disagg_data'

        Returns:
            Dict with success status, count of updates, and any errors
        """
        from app.services.form_data_service import FormDataService
        from app.models.forms import FormData
        from app.models.assignments import PublicSubmission
        import logging

        logger = logging.getLogger(__name__)
        updated_count = 0
        errors = []

        try:
            # Determine data model
            is_public = isinstance(assignment_entity_status, PublicSubmission)
            DataModel = FormData  # Using FormData for both types

            for form_item_id, data_dict in field_data.items():
                try:
                    # Get query filter
                    if is_public:
                        query_filter = {
                            'public_submission_id': assignment_entity_status.id,
                            'form_item_id': form_item_id
                        }
                    else:
                        query_filter = {
                            'assignment_entity_status_id': assignment_entity_status.id,
                            'form_item_id': form_item_id
                        }

                    # Get or create entry
                    data_entry = DataModel.query.filter_by(**query_filter).first()

                    if not data_entry:
                        # Create new entry
                        data_entry = DataModel(**query_filter)
                        from app import db
                        db.session.add(data_entry)

                    # Handle disaggregation data first (takes precedence)
                    if data_dict.get('disagg_data'):
                        disagg_data = data_dict['disagg_data']
                        mode = disagg_data.get('mode')
                        values = disagg_data.get('values', {})

                        if mode and values:
                            data_entry.set_disaggregated_data(mode, values)
                            updated_count += 1
                        else:
                            errors.append(f"Field {form_item_id}: Invalid disaggregation data structure")
                    elif data_dict.get('value') is not None:
                        # Simple value
                        value = data_dict['value']
                        if value:
                            data_entry.set_simple_value(value)
                        else:
                            # Empty value - clear entry
                            data_entry.set_simple_value(None)
                        updated_count += 1
                    else:
                        # Both are None - clear entry
                        data_entry.set_simple_value(None)
                        updated_count += 1

                except Exception as e:
                    error_msg = f"Field {form_item_id}: Validation error."
                    errors.append(error_msg)
                    logger.error(error_msg, exc_info=True)

            # Commit all changes
            from app import db
            db.session.commit()

            return {
                'success': True,
                'updated_count': updated_count,
                'errors': errors
            }

        except Exception as e:
            from app import db
            db.session.rollback()
            logger.error(f"Error in bulk save with disaggregation: {e}", exc_info=True)
            return {
                'success': False,
                'updated_count': 0,
                'errors': ['Bulk save failed.']
            }
