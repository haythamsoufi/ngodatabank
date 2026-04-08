"""PDF / Excel export and import routes for forms."""
from __future__ import annotations

from contextlib import suppress
import io
import json
import os

from flask import current_app, flash, redirect, render_template, request, send_file, url_for
from flask_babel import _
from flask_login import current_user, login_required
from sqlalchemy.orm import joinedload

from app import get_locale
from app.models import (
    db, AssignedForm, AssignmentEntityStatus, Country, DynamicIndicatorData,
    FormData, FormItem, FormPage, FormSection, QuestionType,
    SubmittedDocument,
)
from app.services.excel_service import ExcelService
from app.services.form_processing_service import slugify_age_group
from app.utils.datetime_helpers import utcnow
from app.utils.form_localization import (
    get_localized_country_name,
    get_localized_indicator_name,
    get_localized_page_name,
    get_localized_section_name,
    get_localized_template_name,
    get_translation_key,
)
from app.utils.transactions import request_transaction_rollback
from config import Config

from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def register_export_routes(bp):
    """Register export/import routes onto the forms blueprint."""

    @bp.route("/assignment_status/<int:aes_id>/export_pdf", methods=["GET"])
    @login_required
    def export_assignment_pdf(aes_id):
        """Generate a high-quality PDF for an assignment using a print-optimized HTML template."""
        return _export_pdf_impl(aes_id)

    @bp.route("/assignment_status/<int:aes_id>/export_excel", methods=["GET"])
    @login_required
    def export_focal_data_excel(aes_id):
        return _export_excel_impl(aes_id)

    @bp.route("/assignment_status/<int:aes_id>/import_excel", methods=["POST"])
    @login_required
    def handle_excel_import(aes_id):
        """Legacy Excel import endpoint (delegates to shared ExcelService)."""
        return _import_excel_impl(aes_id)


def _export_pdf_impl(aes_id):
    """Generate a high-quality PDF for an assignment using a print-optimized HTML template.

    Uses WeasyPrint (HTML to PDF) for faithful rendering with proper pagination and styles.
    """
    try:
        assignment_entity_status = AssignmentEntityStatus.query.options(
            db.joinedload(AssignmentEntityStatus.assigned_form).joinedload(AssignedForm.template)
        ).get_or_404(aes_id)

        from app.services.authorization_service import AuthorizationService
        if not AuthorizationService.can_access_assignment(assignment_entity_status, current_user):
            flash("You are not authorized to export data for this assignment and country.", "warning")
            return redirect(url_for("main.dashboard"))

        assignment = assignment_entity_status.assigned_form
        country = assignment_entity_status.country
        form_template_for_export = assignment.template

        from app.services.variable_resolution_service import VariableResolutionService
        from app.models import FormTemplateVersion

        template_version = None
        resolved_variables = {}
        variable_configs = {}
        if form_template_for_export.published_version_id:
            template_version = FormTemplateVersion.query.get(form_template_for_export.published_version_id)
            if template_version:
                variable_configs = template_version.variables or {}
                resolved_variables = VariableResolutionService.resolve_variables(
                    template_version,
                    assignment_entity_status
                )

        if country:
            country_iso = (getattr(country, 'iso3', None) or getattr(country, 'iso2', None) or '').strip().upper()
            if country_iso:
                try:
                    from plugins.emergency_operations.routes import get_emergency_operations_data
                    operations = get_emergency_operations_data(country_iso=country_iso)
                    for i, key in enumerate(('EO1', 'EO2', 'EO3')):
                        if i < len(operations):
                            op = operations[i]
                            name = (op.get('name') or '').strip()
                            code = (op.get('code') or '').strip()
                            resolved_variables[key] = f"{name} ({code})" if code else (name or '')
                        else:
                            resolved_variables[key] = ''
                except Exception as e:
                    current_app.logger.debug(
                        f"Could not resolve EO1/EO2/EO3 for PDF export (plugin or API): {e}"
                    )

        translation_key = get_translation_key()

        assignment_display_name = None
        with suppress(Exception):
            assignment_display_name = get_localized_template_name(
                form_template_for_export,
                locale=translation_key,
                version=template_version,
            )
        country_display_name = None
        with suppress(Exception):
            country_display_name = get_localized_country_name(country) if country else None

        sections_by_page = {}
        default_page_id = 0

        section_nodes_by_id = {}
        ordered_section_ids = []

        for section_model in form_template_for_export.sections.order_by(FormSection.order).all():
            section_display_name = None
            with suppress(Exception):
                section_display_name = get_localized_section_name(section_model)
            if not section_display_name:
                section_display_name = getattr(section_model, 'display_name', None) or section_model.name
            if resolved_variables and section_display_name:
                try:
                    section_display_name = VariableResolutionService.replace_variables_in_text(
                        section_display_name,
                        resolved_variables,
                        variable_configs
                    )
                except Exception as e:
                    current_app.logger.warning(
                        f"Error resolving variables in section name for section {section_model.id}: {e}",
                        exc_info=True
                    )

            section_data_for_export = {
                'name': section_model.name,
                'display_name': section_display_name,
                'id': section_model.id,
                'order': section_model.order,
                'page_id': section_model.page_id,
                'parent_section_id': section_model.parent_section_id,
                'relevance_condition': getattr(section_model, 'relevance_condition', None),
                'subsections': [],
                'fields_ordered': []
            }

            temp_fields = []
            form_items = FormItem.query.filter_by(section_id=section_model.id, archived=False).order_by(FormItem.order).all()
            if form_items:
                for form_item in form_items:
                    display_label = None
                    with suppress(Exception):
                        lt = getattr(form_item, 'label_translations', None)
                        if isinstance(lt, dict) and lt:
                            candidate = lt.get(translation_key) or lt.get('en')
                            if isinstance(candidate, str) and candidate.strip():
                                display_label = candidate.strip()
                    if not display_label:
                        display_label = getattr(form_item, 'display_label', None) or form_item.label
                    if resolved_variables and display_label:
                        try:
                            display_label = VariableResolutionService.replace_variables_in_text(
                                display_label,
                                resolved_variables,
                                variable_configs
                            )
                        except Exception as e:
                            current_app.logger.warning(
                                f"Error resolving variables in display_label for form_item {form_item.id}: {e}",
                                exc_info=True
                            )

                    base = {
                        'id': form_item.id,
                        'order': form_item.order,
                        'label': form_item.label,
                        'display_label': display_label,
                        'unit': getattr(form_item, 'unit', None),
                        'type': getattr(form_item, 'type', None),
                        'conditions': getattr(form_item, 'conditions', None),
                    }
                    if form_item.is_indicator:
                        base.update({'kind': 'indicator', 'model': form_item})
                        temp_fields.append(base)
                    elif form_item.is_question:
                        is_blank_note = (
                            getattr(form_item, 'type', None) == 'blank'
                            or (getattr(form_item, 'question_type', None) and getattr(form_item.question_type, 'value', None) == 'blank')
                        )
                        if is_blank_note:
                            base.update({'kind': 'note', 'model': form_item})
                        else:
                            base.update({'kind': 'question', 'model': form_item})
                        temp_fields.append(base)
                    elif getattr(form_item, 'item_type', None) == 'matrix' or getattr(form_item, 'is_matrix', False):
                        matrix_config = {}
                        try:
                            if isinstance(getattr(form_item, 'config', None), dict):
                                matrix_config = form_item.config.get('matrix_config') or form_item.config or {}
                        except Exception as e:
                            current_app.logger.debug("matrix_config parse failed: %s", e)
                            matrix_config = {}

                        matrix_rows = getattr(form_item, '_display_matrix_rows', None)
                        if not matrix_rows and isinstance(matrix_config, dict):
                            matrix_rows = matrix_config.get('rows', []) or []

                        try:
                            if isinstance(matrix_config, dict):
                                row_mode = matrix_config.get('row_mode', 'manual')
                                if row_mode == 'manual' or not row_mode:
                                    if resolved_variables and matrix_rows and isinstance(matrix_rows, list):
                                        resolved_rows = []
                                        for r in matrix_rows:
                                            if isinstance(r, str):
                                                resolved_rows.append(
                                                    VariableResolutionService.replace_variables_in_text(
                                                        r,
                                                        resolved_variables,
                                                        variable_configs
                                                    )
                                                )
                                            else:
                                                resolved_rows.append(r)
                                        matrix_rows = resolved_rows
                        except Exception as e:
                            current_app.logger.warning(
                                f"Error resolving variables in matrix row labels for form_item {form_item.id}: {e}",
                                exc_info=True
                            )
                        matrix_columns = matrix_config.get('columns', []) if isinstance(matrix_config, dict) else []

                        base.update({
                            'kind': 'matrix',
                            'model': form_item,
                            'matrix_config': matrix_config,
                            'matrix_rows': matrix_rows,
                            'matrix_columns': matrix_columns,
                        })
                        temp_fields.append(base)
                    elif form_item.is_document_field:
                        base.update({'kind': 'document', 'model': form_item})
                        temp_fields.append(base)

            section_type = getattr(section_model, 'section_type', None) or 'standard'
            if section_type == 'dynamic_indicators':
                dynamic_assignments = DynamicIndicatorData.query.filter_by(
                    assignment_entity_status_id=assignment_entity_status.id,
                    section_id=section_model.id,
                ).order_by(DynamicIndicatorData.order).all()
                for dyn in dynamic_assignments:
                    display_label = dyn.custom_label
                    if not (display_label and str(display_label).strip()):
                        with suppress(Exception):
                            display_label = get_localized_indicator_name(dyn.indicator_bank)
                    if not display_label:
                        display_label = getattr(dyn.indicator_bank, 'name', '') or ''
                    if resolved_variables and display_label:
                        try:
                            display_label = VariableResolutionService.replace_variables_in_text(
                                display_label, resolved_variables, variable_configs
                            )
                        except Exception as e:
                            current_app.logger.debug("replace_variables for display_label failed: %s", e)
                    temp_fields.append({
                        'id': f'dynamic_{dyn.id}',
                        'order': dyn.order,
                        'label': display_label,
                        'display_label': display_label,
                        'unit': getattr(dyn.indicator_bank, 'unit', None),
                        'type': getattr(dyn.indicator_bank, 'type', None),
                        'conditions': None,
                        'kind': 'indicator',
                        'model': None,
                    })

            temp_fields.sort(key=lambda x: (x.get('order') is None, x.get('order')))
            section_data_for_export['fields_ordered'] = temp_fields

            section_nodes_by_id[section_model.id] = section_data_for_export
            ordered_section_ids.append(section_model.id)

        for section_id in ordered_section_ids:
            node = section_nodes_by_id.get(section_id)
            if not node:
                continue
            parent_id = node.get('parent_section_id')
            if parent_id and parent_id in section_nodes_by_id:
                section_nodes_by_id[parent_id]['subsections'].append(node)

        for section_id in ordered_section_ids:
            node = section_nodes_by_id.get(section_id)
            if not node or node.get('parent_section_id') is not None:
                continue
            page_id = node.get('page_id') if node.get('page_id') is not None else default_page_id
            if page_id not in sections_by_page:
                sections_by_page[page_id] = []
            sections_by_page[page_id].append(node)

        existing_entries = FormData.query.filter_by(
            assignment_entity_status_id=assignment_entity_status.id
        ).all()
        existing_data_processed_for_export = {}
        for entry in existing_entries:
            if entry.form_item_id:
                value = entry.disagg_data if entry.disagg_data is not None else entry.value
                if isinstance(value, str):
                    v = value.strip()
                    if (v.startswith('{') and v.endswith('}')) or (v.startswith('[') and v.endswith(']')):
                        with suppress(Exception):
                            value = json.loads(v)
                existing_data_processed_for_export[f"form_item_{entry.form_item_id}"] = value

        dynamic_entries = DynamicIndicatorData.query.filter_by(
            assignment_entity_status_id=assignment_entity_status.id
        ).all()
        for dyn in dynamic_entries:
            key = f"form_item_dynamic_{dyn.id}"
            if dyn.disagg_data is not None:
                existing_data_processed_for_export[key] = dyn.disagg_data
            else:
                existing_data_processed_for_export[key] = {'values': {'total': dyn.value}}

        def _parse_hidden_ids_arg(arg_name):
            raw = (request.args.get(arg_name) or '').strip()
            if not raw:
                return set()
            out = set()
            for part in raw.split(','):
                part = (part or '').strip()
                if not part:
                    continue
                if part.isdigit():
                    try:
                        out.add(int(part))
                    except (ValueError, TypeError):
                        continue
            return out

        hidden_section_ids_from_client = _parse_hidden_ids_arg('hidden_sections')
        hidden_field_ids_from_client = _parse_hidden_ids_arg('hidden_fields')

        def _filter_section_node(section_node):
            if not isinstance(section_node, dict):
                return None

            try:
                if section_node.get('id') in hidden_section_ids_from_client:
                    return None
            except Exception as e:
                current_app.logger.debug("hidden section filter failed: %s", e)

            kept_fields = []
            for f in (section_node.get('fields_ordered') or []):
                if not isinstance(f, dict):
                    continue
                try:
                    if f.get('id') in hidden_field_ids_from_client:
                        continue
                except Exception as e:
                    current_app.logger.debug("hidden field filter failed: %s", e)
                kept_fields.append(f)
            section_node['fields_ordered'] = kept_fields

            kept_children = []
            for child in (section_node.get('subsections') or []):
                kept = _filter_section_node(child)
                if kept is not None:
                    kept_children.append(kept)
            section_node['subsections'] = kept_children
            return section_node

        filtered_sections_by_page = {}
        for page_id, root_sections in (sections_by_page or {}).items():
            kept_roots = []
            for sec in (root_sections or []):
                kept = _filter_section_node(sec)
                if kept is not None:
                    kept_roots.append(kept)
            filtered_sections_by_page[page_id] = kept_roots
        sections_by_page = filtered_sections_by_page

        def _infer_list_library_rows_and_labels(field_dict, matrix_data):
            try:
                if not isinstance(field_dict, dict):
                    return
                if field_dict.get('kind') != 'matrix':
                    return
                if not isinstance(matrix_data, dict) or not matrix_data:
                    return

                matrix_config = field_dict.get('matrix_config') if isinstance(field_dict.get('matrix_config'), dict) else {}
                row_mode = (matrix_config.get('row_mode') or '').strip().lower()
                if row_mode != 'list_library':
                    return

                cols = field_dict.get('matrix_columns') or []
                col_names = []
                for c in cols:
                    if isinstance(c, dict):
                        col_names.append(str(c.get('name') if c.get('name') else c))
                    else:
                        col_names.append(str(c))
                col_names = [c for c in col_names if c and c != 'None']
                if not col_names:
                    return

                col_names_sorted = sorted(col_names, key=len, reverse=True)
                row_ids = []
                seen = set()
                for k in matrix_data.keys():
                    if not isinstance(k, str):
                        continue
                    if k.startswith('_'):
                        continue
                    matched_row_id = None
                    for cn in col_names_sorted:
                        suffix = "_" + cn
                        if k.endswith(suffix):
                            matched_row_id = k[: -len(suffix)]
                            break
                    if matched_row_id and matched_row_id not in seen:
                        seen.add(matched_row_id)
                        row_ids.append(matched_row_id)

                if not row_ids:
                    return

                from flask import session
                lookup_list_id = (matrix_config.get('lookup_list_id') or '').strip()

                display_column = (matrix_config.get('display_column') or matrix_config.get('list_display_column') or 'name').strip() or 'name'

                row_labels = {}
                if lookup_list_id and str(lookup_list_id).isdigit():
                    from app.models import LookupListRow
                    for rid in row_ids:
                        try:
                            rid_int = int(rid)
                        except (ValueError, TypeError):
                            row_labels[rid] = rid
                            continue
                        row_obj = LookupListRow.query.get(rid_int)
                        if row_obj and isinstance(row_obj.data, dict):
                            row_labels[rid] = str(row_obj.data.get(display_column) or row_obj.data.get('name') or rid)
                        else:
                            row_labels[rid] = rid
                elif lookup_list_id in ('country_map', 'national_society', 'indicator_bank'):
                    current_locale = session.get('language', 'en') or (str(get_locale()) if get_locale() else 'en')
                    if isinstance(current_locale, str) and '_' in current_locale:
                        current_locale = current_locale.split('_', 1)[0]

                    if lookup_list_id == 'country_map':
                        for rid in row_ids:
                            try:
                                rid_int = int(rid)
                            except (ValueError, TypeError):
                                row_labels[rid] = rid
                                continue
                            obj = Country.query.get(rid_int)
                            row_labels[rid] = get_localized_country_name(obj) if obj else rid
                    elif lookup_list_id == 'national_society':
                        from app.models.organization import NationalSociety
                        for rid in row_ids:
                            try:
                                rid_int = int(rid)
                            except (ValueError, TypeError):
                                row_labels[rid] = rid
                                continue
                            obj = NationalSociety.query.get(rid_int)
                            if obj:
                                localized_name = obj.get_name_translation(current_locale) if hasattr(obj, 'get_name_translation') else None
                                row_labels[rid] = (localized_name.strip() if isinstance(localized_name, str) and localized_name.strip() else obj.name)
                            else:
                                row_labels[rid] = rid
                    else:
                        from app.models.indicator_bank import IndicatorBank
                        for rid in row_ids:
                            try:
                                rid_int = int(rid)
                            except (ValueError, TypeError):
                                row_labels[rid] = rid
                                continue
                            obj = IndicatorBank.query.get(rid_int)
                            row_labels[rid] = obj.name if obj else rid
                else:
                    row_labels = {rid: rid for rid in row_ids}

                field_dict['matrix_rows'] = row_ids
                field_dict['matrix_row_labels'] = row_labels
            except Exception as e:
                current_app.logger.warning(f"Failed to infer list-library matrix rows for PDF: {e}", exc_info=True)

        def _walk_sections_for_export(section_node):
            if not isinstance(section_node, dict):
                return
            fields = section_node.get('fields_ordered') or []
            if isinstance(fields, list):
                for f in fields:
                    if isinstance(f, dict) and f.get('kind') == 'matrix':
                        item_key = f"form_item_{f.get('id')}"
                        _infer_list_library_rows_and_labels(f, existing_data_processed_for_export.get(item_key))
            for child in section_node.get('subsections', []) or []:
                _walk_sections_for_export(child)

        for page_id, root_sections in (sections_by_page or {}).items():
            for sec in root_sections or []:
                _walk_sections_for_export(sec)

        pages = list(form_template_for_export.pages) if form_template_for_export.is_paginated else [None]

        html_content = render_template(
            'forms/entry_form/export_pdf.html',
            assignment=assignment,
            assignment_display_name=assignment_display_name,
            country=country,
            country_display_name=country_display_name,
            aes=assignment_entity_status,
            form_template=form_template_for_export,
            sections_by_page=sections_by_page,
            pages=pages,
            existing_data=existing_data_processed_for_export,
            generated_at=utcnow(),
            get_localized_page_name=get_localized_page_name,
            wide_matrix_strategy='scale',
        )

        try:
            from weasyprint import HTML, CSS  # type: ignore
        except Exception as e:
            current_app.logger.error(f"WeasyPrint not available: {e}", exc_info=True)
            return current_app.response_class(
                response="PDF generation is not available on this deployment.",
                status=503,
                mimetype='text/plain'
            )

        static_dir = os.path.join(current_app.root_path, 'static')

        pdf_css_string = '''
            @page {
                size: A4;
                margin: 20mm 15mm 20mm 15mm;
                @bottom-right { content: "Page " counter(page); font-size: 10pt; color: #6b7280; }
            }
            body {
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, Arial, sans-serif;
                color: #111827;
                line-height: 1.5;
            }
            h1, h2, h3, h4 { color: #111827; margin: 0 0 8px 0; }
            h1 { font-size: 20pt; }
            h2 { font-size: 14pt; border-bottom: 2px solid #cc0000; padding-bottom: 4px; margin-top: 16px; margin-bottom: 12px; }
            .form-page-title { page-break-after: avoid; }
            h3 { font-size: 12pt; margin-top: 10px; margin-bottom: 8px; color: #374151; }
            h4 { font-size: 11pt; margin-top: 8px; margin-bottom: 6px; color: #374151; }
            .meta {
                color: #374151;
                font-size: 10pt;
                margin-bottom: 16px;
                padding: 8px;
                background: #f9fafb;
                border-left: 3px solid #cc0000;
            }
            .meta div { margin: 4px 0; }
            .section {
                margin-bottom: 16px;
            }
            .section h2, .section h3, .section h4 {
                page-break-after: avoid;
            }
            .subsection {
                margin-left: 12px;
                padding-left: 10px;
                border-left: 2px solid #e5e7eb;
            }
            .section-empty-note {
                color: #6b7280;
                font-size: 10pt;
                font-style: italic;
                margin: 8px 0 0 0;
            }
            .form-note {
                color: #374151;
                font-size: 10pt;
                margin: 8px 0;
                padding: 8px 12px;
                background: #f9fafb;
                border-left: 3px solid #9ca3af;
                border-radius: 0 4px 4px 0;
            }

            .field-box {
                border: 1.5px solid #e5e7eb;
                border-radius: 4px;
                margin: 8px 0;
                page-break-inside: avoid;
                background: #ffffff;
                box-shadow: 0 1px 2px rgba(0, 0, 0, 0.05);
            }
            .field-box-matrix {
                page-break-inside: auto;
            }
            .field-box-matrix .table thead {
                display: table-header-group;
            }
            .field-box-matrix .table tr {
                page-break-inside: avoid;
            }

            .field-filled {
                border-left: 4px solid #10b981;
            }

            .field-empty {
                border-left: 4px solid #d1d5db;
                background: #f9fafb;
            }

            .field-empty-required {
                border-left: 4px solid #ef4444;
                background: #fef2f2;
            }

            .field-empty-optional {
                border-left: 4px solid #d1d5db;
                background: #f9fafb;
            }

            .field-header {
                background: #f9fafb;
                padding: 8px 12px;
                border-bottom: 1px solid #e5e7eb;
                font-weight: 600;
            }

            .field-label {
                color: #111827;
                font-size: 11pt;
                display: block;
            }

            .field-unit {
                color: #6b7280;
                font-size: 9pt;
                font-weight: normal;
                font-style: italic;
            }

            .field-content {
                padding: 10px 12px;
                min-height: 20px;
            }

            .field-value {
                color: #111827;
                font-size: 10pt;
                word-wrap: break-word;
                display: block;
            }

            .disaggregation-caption {
                font-size: 9pt;
                font-weight: 600;
                color: #374151;
                margin: 6px 0 4px 0;
                display: block;
            }

            .not-reported {
                color: #dc2626;
                font-size: 10pt;
                font-weight: 600;
                font-style: italic;
                display: block;
            }

            .not-reported-optional {
                color: #6b7280;
                font-size: 10pt;
                font-weight: 600;
                font-style: italic;
                display: block;
            }

            html[dir="rtl"] body {
                direction: rtl;
                font-family: "Tajawal", -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, Arial, sans-serif;
            }
            html[dir="rtl"] h1,
            html[dir="rtl"] h2,
            html[dir="rtl"] h3,
            html[dir="rtl"] h4,
            html[dir="rtl"] .meta,
            html[dir="rtl"] .field-header,
            html[dir="rtl"] .field-content,
            html[dir="rtl"] .field-value,
            html[dir="rtl"] .disaggregation-caption,
            html[dir="rtl"] .not-reported,
            html[dir="rtl"] .not-reported-optional {
                text-align: right;
                font-family: "Tajawal", -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, Arial, sans-serif;
            }
            html[dir="rtl"] .meta { border-left: none; border-right: 3px solid #cc0000; }
            html[dir="rtl"] .form-note { border-left: none; border-right: 3px solid #9ca3af; border-radius: 4px 0 0 4px; }
            html[dir="rtl"] .subsection { margin-left: 0; margin-right: 12px; padding-left: 0; padding-right: 10px; border-left: none; border-right: 2px solid #e5e7eb; }
            html[dir="rtl"] table th, html[dir="rtl"] table td {
                text-align: right;
                font-family: "Tajawal", -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, Arial, sans-serif;
            }

            .document-note {
                color: #6b7280;
                font-style: italic;
            }

            .table {
                width: 100%;
                border-collapse: collapse;
                margin: 8px 0;
                font-size: 9pt;
            }
            .table th, .table td {
                border: 1px solid #d1d5db;
                padding: 6px 8px;
            }
            .table th {
                background: #f3f4f6;
                text-align: left;
                font-weight: 600;
                color: #374151;
            }
            .table td {
                background: #ffffff;
            }
            .table tbody tr:nth-child(even) td {
                background: #f9fafb;
            }
            .table td.cell-tick {
                text-align: center;
            }
            .wide-matrix-scale {
                width: 100%;
                overflow: hidden;
            }
            .wide-matrix-scale table.table {
                transform: scale(0.72);
                transform-origin: top left;
            }
            @page wide {
                size: A4 landscape;
                margin: 20mm 15mm 20mm 15mm;
                @bottom-right { content: "Page " counter(page); font-size: 10pt; color: #6b7280; }
            }
            .wide-matrix-landscape {
                page: wide;
            }
            .page-break { page-break-before: always; }
        '''
        with suppress(Exception):
            pdf_css_string = pdf_css_string.replace('content: "Page "', f'content: "{_("Page")} "')
        pdf_css = CSS(string=pdf_css_string)

        pdf_buffer = io.BytesIO()
        HTML(string=html_content, base_url=static_dir).write_pdf(
            pdf_buffer,
            stylesheets=[pdf_css],
            optimize_size=('fonts', 'images')
        )

        pdf_buffer.seek(0)
        filename = f"assignment_{country.iso3 if country else 'country'}_{str(assignment.period_name).replace(' ', '_')}.pdf"
        return send_file(
            pdf_buffer,
            download_name=filename,
            as_attachment=True,
            mimetype='application/pdf'
        )
    except Exception as e:
        current_app.logger.error(f"Error generating PDF for ACS {aes_id}: {e}", exc_info=True)
        flash("Failed to generate PDF.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))


def _export_excel_impl(aes_id):
    assignment_entity_status = AssignmentEntityStatus.query.options(
        db.joinedload(AssignmentEntityStatus.assigned_form).joinedload(AssignedForm.template)
    ).get_or_404(aes_id)

    from app.services.authorization_service import AuthorizationService
    if not AuthorizationService.can_access_assignment(assignment_entity_status, current_user):
         flash("You are not authorized to export data for this assignment and country.", "warning")
         return redirect(url_for("main.dashboard"))

    assignment = assignment_entity_status.assigned_form
    country = assignment_entity_status.country
    form_template_for_export = assignment.template

    from app.services.variable_resolution_service import VariableResolutionService
    from app.models import FormTemplateVersion

    template_version = None
    resolved_variables = {}
    variable_configs = {}
    if form_template_for_export.published_version_id:
        template_version = FormTemplateVersion.query.get(form_template_for_export.published_version_id)
        if template_version:
            variable_configs = template_version.variables or {}
            resolved_variables = VariableResolutionService.resolve_variables(
                template_version,
                assignment_entity_status
            )

    sections_by_page = {}
    default_page_id = 0

    for section_model in form_template_for_export.sections.order_by(FormSection.order).all():
        page_id = section_model.page_id if section_model.page_id is not None else default_page_id
        if page_id not in sections_by_page:
            sections_by_page[page_id] = []

        section_display_name = getattr(section_model, 'display_name', None) or section_model.name
        if resolved_variables and section_display_name:
            try:
                section_display_name = VariableResolutionService.replace_variables_in_text(
                    section_display_name,
                    resolved_variables,
                    variable_configs
                )
            except Exception as e:
                current_app.logger.warning(
                    f"Error resolving variables in section name for section {section_model.id}: {e}",
                    exc_info=True
                )

        section_data_for_export = {'name': section_display_name, 'id': section_model.id, 'fields_ordered': []}
        temp_fields = []

        form_items = FormItem.query.filter_by(section_id=section_model.id, archived=False).order_by(FormItem.order).all()
        if form_items:
            for form_item in form_items:
                display_label = getattr(form_item, 'display_label', None) or form_item.label
                if resolved_variables and display_label:
                    try:
                        display_label = VariableResolutionService.replace_variables_in_text(
                            display_label,
                            resolved_variables,
                            variable_configs
                        )
                    except Exception as e:
                        current_app.logger.warning(
                            f"Error resolving variables in display_label for form_item {form_item.id}: {e}",
                            exc_info=True
                        )

                if form_item.is_indicator:
                    temp_fields.append({
                        'id': form_item.id, 'legacy_id': None, 'label': display_label,
                        'type': form_item.type, 'unit': form_item.unit, 'order': form_item.order,
                        'is_indicator': True, 'is_form_item': True, 'item_model': form_item
                    })
                elif form_item.is_question:
                    temp_fields.append({
                        'id': form_item.id, 'legacy_id': None, 'label': display_label,
                        'type': form_item.type, 'order': form_item.order,
                        'is_question': True, 'is_form_item': True, 'item_model': form_item
                    })
                elif form_item.is_document_field:
                    temp_fields.append({
                        'id': form_item.id, 'legacy_id': None, 'label': display_label,
                        'type': 'DOCUMENT', 'order': form_item.order, 'is_required': form_item.is_required,
                        'description': form_item.description, 'is_document': True, 'is_form_item': True, 'item_model': form_item
                    })

        temp_fields.sort(key=lambda x: x['order'])
        section_data_for_export['fields_ordered'] = temp_fields
        sections_by_page[page_id].append(section_data_for_export)

    existing_data_entries_for_export = FormData.query.filter_by(
        assignment_entity_status_id=assignment_entity_status.id
    ).all()
    existing_data_processed_for_export = {}
    for entry in existing_data_entries_for_export:
        if entry.form_item_id:
            item_key_suffix = f"form_item_{entry.form_item_id}"
            existing_data_processed_for_export[item_key_suffix] = entry.value

    workbook = openpyxl.Workbook()

    IFRC_RED = "FFED1B2E"
    IFRC_DARK_RED = "FFAF0E1B"
    IFRC_LIGHT_GRAY = "FFF5F5F5"
    IFRC_MEDIUM_GRAY = "FFE0E0E0"
    IFRC_DARK_GRAY = "FF666666"
    IFRC_WHITE = "FFFFFFFF"
    IFRC_YELLOW = "FFFFF9E6"

    title_font = Font(name='Arial', size=16, bold=True, color=IFRC_DARK_RED)
    section_title_font = Font(name='Arial', size=14, bold=True, color=IFRC_DARK_RED)
    item_label_font = Font(name='Arial', size=12, bold=True, color=IFRC_DARK_GRAY)
    header_font = Font(name='Arial', size=11, bold=True, color=IFRC_WHITE)
    normal_font = Font(name='Arial', size=11, color=IFRC_DARK_GRAY)

    header_fill = PatternFill(start_color=IFRC_DARK_GRAY, end_color=IFRC_DARK_GRAY, fill_type='solid')
    section_fill = PatternFill(start_color=IFRC_LIGHT_GRAY, end_color=IFRC_LIGHT_GRAY, fill_type='solid')
    data_entry_fill = PatternFill(start_color=IFRC_YELLOW, end_color=IFRC_YELLOW, fill_type='solid')
    alternate_row_fill = PatternFill(start_color=IFRC_LIGHT_GRAY, end_color=IFRC_LIGHT_GRAY, fill_type='solid')

    no_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None)
    )
    disagg_border = Border(
        left=Side(style='thin', color=IFRC_DARK_GRAY),
        right=Side(style='thin', color=IFRC_DARK_GRAY),
        top=Side(style='thin', color=IFRC_DARK_GRAY),
        bottom=Side(style='thin', color=IFRC_DARK_GRAY)
    )

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    pages = list(form_template_for_export.pages) if form_template_for_export.is_paginated else [None]
    first_sheet = True

    for page in pages:
        page_id = page.id if page else default_page_id
        sheet_name_raw = get_localized_page_name(page)

        invalid_chars = ['/', '\\', '?', '*', '[', ']']
        sheet_name = sheet_name_raw
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, '-')

        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        if first_sheet:
            data_sheet = workbook.active
            data_sheet.title = sheet_name
            first_sheet = False
        else:
            data_sheet = workbook.create_sheet(sheet_name)

            for row in range(1, 1000):
                data_sheet.row_dimensions[row].height = 17

        current_row = 1
        data_sheet.cell(row=current_row, column=1).value = f"Assignment: {form_template_for_export.name} - {assignment.period_name} for {country.name}"
        data_sheet.cell(row=current_row, column=1).font = title_font
        current_row += 2

        for section_data in sections_by_page.get(page_id, []):
            data_sheet.cell(row=current_row, column=1).value = section_data['name']
            data_sheet.cell(row=current_row, column=1).font = section_title_font
            current_row += 1

            for field_data in section_data['fields_ordered']:
                field_model = field_data['item_model']
                data_sheet.cell(row=current_row, column=2).value = f"{field_data['order']}. {field_data['label']}"
                data_sheet.cell(row=current_row, column=2).font = item_label_font
                current_row += 1
                col_offset = 3

                if field_data.get('is_indicator'):
                    indicator = field_model
                    item_key = f"form_item_{field_data['id']}"
                    entry_data = existing_data_processed_for_export.get(item_key, {})

                    allowed_modes = indicator.allowed_disaggregation_options if indicator.unit and indicator.unit in ['People', 'Volunteers', 'Staff'] else ['total']

                    if indicator.type == 'Number':
                        for mode in allowed_modes:
                            mode_display = Config.DISAGGREGATION_MODES.get(mode, mode.title())
                            mode_cell = data_sheet.cell(row=current_row, column=col_offset, value=mode_display)
                            mode_cell.font = header_font
                            mode_cell.fill = header_fill
                            mode_cell.border = disagg_border
                            mode_cell.alignment = center_align
                            current_row += 1

                            current_values = entry_data.get('values', {}) if isinstance(entry_data, dict) else {'value': entry_data if entry_data is not None else ''}

                            if mode == 'total':
                                val_cell = data_sheet.cell(row=current_row, column=col_offset, value=current_values.get('total', ''))
                                val_cell.fill = data_entry_fill
                                val_cell.border = disagg_border
                                val_cell.number_format = '0'
                                current_row += 1

                            elif mode == 'sex':
                                headers = indicator.effective_sex_categories
                                for col, header in enumerate(headers, col_offset):
                                    data_sheet.cell(row=current_row, column=col).value = header
                                    data_sheet.cell(row=current_row, column=col).font = header_font
                                current_row += 1

                                for col, category in enumerate(headers, col_offset):
                                    val_key = category.lower().replace(' ', '_')
                                    val_cell = data_sheet.cell(row=current_row, column=col, value=current_values.get(val_key, ''))
                                    val_cell.fill = data_entry_fill
                                    val_cell.border = disagg_border
                                    val_cell.number_format = '0'
                                    data_sheet.row_dimensions[current_row].height = 17
                                current_row += 1

                            elif mode == 'age':
                                headers = indicator.effective_age_groups
                                for col, header in enumerate(headers, col_offset):
                                    data_sheet.cell(row=current_row, column=col).value = header
                                    data_sheet.cell(row=current_row, column=col).font = header_font
                                current_row += 1

                                for col, category in enumerate(headers, col_offset):
                                    val_key = category.lower().replace(' ', '_').replace('+', 'plus')
                                    val_cell = data_sheet.cell(row=current_row, column=col, value=current_values.get(val_key, ''))
                                    val_cell.fill = data_entry_fill
                                    val_cell.border = disagg_border
                                    val_cell.number_format = '0'
                                current_row += 1

                            elif mode == 'sex_age':
                                age_groups = indicator.effective_age_groups
                                sex_categories = indicator.effective_sex_categories

                                category_cell = data_sheet.cell(row=current_row, column=col_offset, value="Category")
                                category_cell.font = header_font
                                category_cell.fill = header_fill
                                category_cell.border = disagg_border
                                category_cell.alignment = center_align

                                for col, age in enumerate(age_groups, col_offset + 1):
                                    header_cell = data_sheet.cell(row=current_row, column=col, value=age)
                                    header_cell.font = header_font
                                    header_cell.fill = header_fill
                                    header_cell.border = disagg_border
                                    header_cell.alignment = center_align
                                current_row += 1

                                for sex in sex_categories:
                                    sex_cell = data_sheet.cell(row=current_row, column=col_offset, value=sex)
                                    sex_cell.font = header_font
                                    sex_cell.fill = header_fill
                                    sex_cell.border = disagg_border
                                    sex_cell.alignment = center_align

                                    for col, age in enumerate(age_groups, col_offset + 1):
                                        val_key = f"{sex.lower().replace(' ', '_')}_{age.lower().replace(' ', '_').replace('+', 'plus')}"
                                        val_cell = data_sheet.cell(row=current_row, column=col, value=current_values.get(val_key, ''))
                                        val_cell.fill = data_entry_fill
                                        val_cell.border = disagg_border
                                        val_cell.number_format = '0'
                                    current_row += 1

                            current_row += 1
                    else:
                                                    # Non-numeric indicators just get a single value cell
                                val_cell = data_sheet.cell(row=current_row, column=col_offset, value=current_values.get('value', ''))
                                val_cell.fill = data_entry_fill
                                val_cell.border = no_border
                                current_row += 2

                elif field_data.get('is_question'):
                    question = field_model
                    item_key = f"form_item_{question.id}"
                    q_value = existing_data_processed_for_export.get(item_key, '')
                    val_cell = data_sheet.cell(row=current_row, column=col_offset, value=q_value)
                    val_cell.fill = data_entry_fill
                    val_cell.border = no_border

                    if question.type == QuestionType.number:
                        val_cell.number_format = '0'
                    elif question.type == QuestionType.date and q_value:
                        with suppress(Exception):
                            val_cell.value = datetime.strptime(q_value, '%Y-%m-%d').date()
                            val_cell.number_format = 'yyyy-mm-dd'

                    if question.type in [QuestionType.single_choice, QuestionType.multiple_choice] and question.options:
                        options_list = [str(opt.get('value', opt) if isinstance(opt, dict) else opt) for opt in question.options]
                        options_str = ','.join([f'"{opt}"' for opt in options_list])
                        dv = DataValidation(type="list", formula1=f"={options_str}", allow_blank=True)
                        data_sheet.add_data_validation(dv)
                        dv.add(val_cell)

                    current_row += 2

                elif field_data.get('is_document'):
                    data_sheet.cell(row=current_row, column=col_offset, value="(Manage in Web Form)")
                    current_row += 2

                current_row += 1

        data_sheet.column_dimensions['C'].width = 20
        for col_idx in range(1, data_sheet.max_column + 1):
            if get_column_letter(col_idx) != 'C':
                data_sheet.column_dimensions[get_column_letter(col_idx)].autosize = True

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"data_entry_{country.iso3}_{str(assignment.period_name).replace(' ', '_')}.xlsx"
    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    download_name=filename,
                    as_attachment=True)


def _import_excel_impl(aes_id):
    """Legacy Excel import endpoint (delegates to shared ExcelService).

    NOTE: This endpoint is maintained for backward compatibility and testing.
    New code should use excel.import_assignment_excel instead.
    """
    MAX_EXCEL_FILE_SIZE = 10 * 1024 * 1024

    assignment_entity_status = AssignmentEntityStatus.query.get_or_404(aes_id)

    from app.services.authorization_service import AuthorizationService
    if not AuthorizationService.can_edit_assignment(assignment_entity_status, current_user):
        flash("You are not authorized to import data for this assignment or it's not in an editable state.", "warning")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    excel_file = request.files.get('excel_file')
    if not excel_file or excel_file.filename == '':
        flash("No Excel file selected for import.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    if not excel_file.filename.lower().endswith('.xlsx'):
        flash("Invalid file type. Please upload a .xlsx file.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    file_size = excel_file.content_length
    if file_size is None:
        excel_file.seek(0, 2)
        file_size = excel_file.tell()
        excel_file.seek(0)

    if file_size > MAX_EXCEL_FILE_SIZE:
        flash(f"File size ({file_size / (1024*1024):.2f}MB) exceeds the maximum allowed size of 10MB.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    try:
        from app.utils.advanced_validation import AdvancedValidator
        file_ext = os.path.splitext(excel_file.filename)[1].lower()
        is_valid_mime, detected_mime = AdvancedValidator.validate_mime_type(excel_file, [file_ext])
        if not is_valid_mime:
            current_app.logger.warning(f"Excel import MIME mismatch: claimed {file_ext}, detected {detected_mime}")
            flash("File content does not match its extension. Please upload a valid Excel file.", "danger")
            return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))
    except Exception as e:
        current_app.logger.warning(f"MIME validation error for Excel import: {e}", exc_info=True)
        flash("Unable to validate file type. Please try again.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    try:
        workbook = ExcelService.load_workbook(excel_file)
    except ValueError as exc:
        flash("An error occurred. Please try again.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    result = ExcelService.import_assignment_data(assignment_entity_status, workbook)
    if result['success']:
        if result.get('errors'):
            error_msg = f"Excel import completed with {result['updated_count']} values saved. Errors: {', '.join(result['errors'][:5])}"
            if len(result['errors']) > 5:
                error_msg += f" (and {len(result['errors']) - 5} more)"
            flash(error_msg, "warning")
        else:
            flash(f"Excel import completed: {result['updated_count']} values saved.", "success")
    else:
        error_msg = f"Excel import failed: {', '.join(result.get('errors', [])[:5])}"
        if len(result.get('errors', [])) > 5:
            error_msg += f" (and {len(result.get('errors', [])) - 5} more)"
        flash(error_msg, "danger")

    return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))
