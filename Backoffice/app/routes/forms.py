from app.utils.transactions import request_transaction_rollback
from contextlib import suppress
# ========== File: app/routes/forms.py ==========
# Unified Forms Blueprint - Handles all form-related functionality
# Consolidates focal point forms, admin forms, and public forms
#
# FEATURES:
# ✅ Unified URL structure: /forms/<form_type>/<form_id>
# ✅ Assignment forms: /forms/assignment/123
# ✅ Public submissions: /forms/public-submission/456
# ✅ Backward compatibility with legacy routes
# ✅ Same template (entry_form.html) for all form types
# ✅ Admin-only access for public submissions
# ✅ Full CRUD operations for both form types
# ✅ Unified form item processing for indicators, questions, documents
# ✅ Centralized localization and authorization utilities
#
# REGISTRATION EXAMPLE:
# In your main Flask app file (app/__init__.py or similar):
#
#   from app.routes import forms
#   app.register_blueprint(forms.bp)
from app.utils.datetime_helpers import utcnow
from app.utils.sql_utils import safe_ilike_pattern


from flask import Blueprint, render_template, redirect, url_for, request, flash, current_app, send_from_directory, abort, send_file
from flask_login import current_user, login_required
from flask_babel import _
from app.models import db, AssignedForm, FormTemplate, Country, FormSection, FormData, IndicatorBank, AssignmentEntityStatus, SubmittedDocument, User, QuestionType, DynamicIndicatorData, SectionType, RepeatGroupInstance, RepeatGroupData, FormItem, FormItemType, PublicSubmission, PublicSubmissionStatus, FormPage # Added public submission models
from app.models.enums import EntityType
from app.services.entity_service import EntityService
# NEW: Import notifications utilities
from app.utils.notifications import notify_assignment_submitted, notify_document_uploaded, notify_form_data_updated, log_entity_activity

# Import unified form processing utilities
from app.utils.form_processing import FormItemProcessor, IndirectReachProcessor, calculate_disaggregation_total, should_create_data_availability_entry, get_form_items_for_section
from app.utils.form_localization import (
    get_localized_indicator_name,
    get_localized_indicator_definition,
    get_localized_indicator_type,
    get_localized_indicator_unit,
    get_localized_sector_name,
    get_localized_subsector_name,
    get_localized_page_name,
    get_localized_section_name,
    get_localized_template_name,
    get_translation_key,
)
from app.utils.form_authorization import check_assignment_access, check_assignment_edit_access, admin_required, has_country_access, can_edit_assignment
from app.utils.request_utils import is_json_request
from app.utils.plugin_data_processor import process_form_plugin_data, plugin_data_processor
from app.utils.template_preparation import TemplatePreparationService
from app.utils.excel_service import ExcelService
from app.utils.api_helpers import GENERIC_ERROR_MESSAGE, get_json_safe
from app.utils.api_responses import json_bad_request, json_not_found, json_ok, json_server_error, require_json_keys
from app.utils.request_validation import enforce_csrf_json

# Import business logic services
from app.services.form_data_service import FormDataService
from app.forms.shared import DeleteForm
from sqlalchemy.orm import joinedload
from flask_wtf import FlaskForm
from sqlalchemy import func
import json
from config import Config
from flask import session
from app.utils.constants import (
    DEFAULT_LOOKUP_ROW_LIMIT,
    SELECTED_COUNTRY_ID_SESSION_KEY,
    SELF_REPORT_PERIOD_NAME,
)
from app.utils.form_localization import get_localized_country_name
from app import get_locale
from werkzeug.utils import secure_filename
import os
import pandas as pd
import io
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import logging # Import logging here too
import sys # ADDED: Import sys module
from app.utils.debug_utils import debug_manager, performance_monitor, debug_form_data, log_user_action
from sqlalchemy import or_
from app.services import get_formdata_map
import re
import uuid
from werkzeug.security import safe_join

bp = Blueprint("forms", __name__, url_prefix="/forms")

# Register additional route modules to keep this file manageable
from app.routes.forms_validation_summary import register_validation_summary_routes  # noqa: E402
register_validation_summary_routes(bp)

# Legacy compatibility adapter removed; AES is the only assignment status now.

# Template global for getting frontend URL
@bp.app_template_global()
def get_frontend_url_global():
    """Template global for getting frontend URL."""
    try:
        from app.utils.app_settings import get_frontend_url as _get_frontend_url
        url = _get_frontend_url()
        if url is None:
            return "#"  # Return a safe fallback
        return url
    except Exception as e:
        current_app.logger.debug("get_frontend_url failed: %s", e)
        return "#"  # Return a safe fallback

# Old localization and authorization functions have been moved to utility modules
# for better organization and reusability

# Import URL and form helpers from route_helpers module
from app.utils.route_helpers import get_unified_form_url, get_unified_form_item_id
from app.utils.form_processing import slugify_age_group

def debug_numeric_value(logger, context, field_id, field_type, value, processed_value):
    """Helper function to log numeric value processing"""
    logger.debug(f"[NUMERIC DEBUG] {context}")
    logger.debug(f"  Field ID: {field_id}")
    logger.debug(f"  Field Type: {field_type}")
    logger.debug(f"  Original Value: {value} (type: {type(value)})")
    logger.debug(f"  Processed Value: {processed_value} (type: {type(processed_value)})")

# Add this helper function at the top with other helpers
def process_numeric_value(value):
    """Process a numeric value, ensuring proper handling of None and invalid values"""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None  # Return None so we can filter it out later

    # Handle string representations of None/null/undefined
    if isinstance(value, str):
        value_str = value.strip()
        if value_str.lower() in ('none', 'null', 'undefined', ''):
            return None

        # Remove commas, spaces, and other formatting characters
        clean_value = value_str.replace(',', '').replace(' ', '').replace('\u00A0', '').replace('\u202F', '')
        if not clean_value:
            return None

        try:
            # Try to parse as integer first, then float
            if '.' in clean_value or 'e' in clean_value.lower():
                return float(clean_value)
            else:
                return int(clean_value)
        except (ValueError, TypeError):
            return None

    with suppress((ValueError, TypeError)):
        if isinstance(value, (int, float)):
            return value

    return None

def process_existing_data_for_template(data_entry):
    """Process existing data entry for template rendering using the new structure.
    Be tolerant to lightweight placeholder objects (e.g., TempEntry) by using getattr with defaults.
    """
    if not data_entry:
        return ""

    # Handle data availability flags
    data_not_available = getattr(data_entry, 'data_not_available', False)
    not_applicable = getattr(data_entry, 'not_applicable', False)

    if data_not_available:
        return "data_not_available"
    elif not_applicable:
        return "not_applicable"

    # Handle disaggregated/matrix/plugin data (JSON payload)
    # Priority: reported disagg_data -> prefilled_disagg_data -> imputed_disagg_data
    disagg_data = getattr(data_entry, 'disagg_data', None)
    if disagg_data is not None:
        return disagg_data
    prefilled_disagg_data = getattr(data_entry, 'prefilled_disagg_data', None)
    if prefilled_disagg_data is not None:
        return prefilled_disagg_data
    imputed_disagg_data = getattr(data_entry, 'imputed_disagg_data', None)
    if imputed_disagg_data is not None:
        return imputed_disagg_data

    # Handle simple value - now VARCHAR, no JSON parsing needed
    value = getattr(data_entry, 'value', None)
    if value:
        # Since value column is now VARCHAR, just return the string value directly
        return value

    # If no main value, check for prefilled value (scalar)
    prefilled_value = getattr(data_entry, 'prefilled_value', None)
    if prefilled_value is not None:
        return prefilled_value

    # If no main value and no prefilled value, allow imputed_value to populate the UI (scalar)
    # (e.g., AI suggestions applied from Data Exploration).
    imputed_value = getattr(data_entry, 'imputed_value', None)
    if imputed_value is not None:
        # Normalize quoted-string artifacts so single-choice questions match option values
        # (store/display CHF, not "CHF").
        if isinstance(imputed_value, str):
            s = imputed_value.strip()
            if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
                try:
                    imputed_value = json.loads(s)
                except (json.JSONDecodeError, ValueError, TypeError):
                    imputed_value = s[1:-1]
            elif len(s) >= 2 and s[0] == "'" and s[-1] == "'":
                imputed_value = s[1:-1]
        return imputed_value

    return ""


def _process_form_data_entry(entry, form_item):
    """Process a single FormData/PublicFormData entry into existing_data_processed updates.
    Shared logic for _load_existing_data_for_assignment and _load_existing_data_for_public_submission.
    Returns dict of key-value pairs to merge into existing_data_processed.
    """
    key = f'field_value[{entry.form_item_id}]'
    data_not_available = entry.data_not_available if entry.data_not_available is not None else False
    not_applicable = entry.not_applicable if entry.not_applicable is not None else False

    result = {}

    def _checkbox_key(suffix):
        if form_item.is_indicator:
            return f'indicator_{entry.form_item_id}_{suffix}'
        if form_item.item_type == 'matrix':
            return f'matrix_{entry.form_item_id}_{suffix}'
        return f'question_{entry.form_item_id}_{suffix}'

    if data_not_available:
        result[_checkbox_key('data_not_available')] = True
    if not_applicable:
        result[_checkbox_key('not_applicable')] = True

    if not data_not_available and not not_applicable:
        has_reported = (
            (entry.value is not None and str(entry.value).strip() != "")
            or (getattr(entry, "disagg_data", None) is not None)
        )
        has_prefilled = (
            (getattr(entry, "prefilled_value", None) is not None)
            or (getattr(entry, "prefilled_disagg_data", None) is not None)
        )
        has_imputed = (
            (getattr(entry, "imputed_value", None) is not None)
            or (getattr(entry, "imputed_disagg_data", None) is not None)
        )
        if form_item.item_type == 'matrix' or form_item.item_type.startswith('plugin_'):
            dd = getattr(entry, "disagg_data", None)
            dd_source = "reported"
            if dd is None:
                dd = getattr(entry, "prefilled_disagg_data", None)
                dd_source = "prefilled"
            if dd is None:
                dd = getattr(entry, "imputed_disagg_data", None)
                dd_source = "imputed"
            if dd is not None:
                result[key] = dd
                if dd_source == "prefilled":
                    result[f'{key}_is_prefilled'] = True
                elif dd_source == "imputed":
                    result[f'{key}_is_imputed'] = True
            else:
                result[key] = {}
        else:
            if has_reported or has_prefilled or has_imputed:
                result[key] = process_existing_data_for_template(entry)
                if (not has_reported) and has_prefilled:
                    result[f'{key}_is_prefilled'] = True
                elif (not has_reported) and (not has_prefilled) and has_imputed:
                    result[f'{key}_is_imputed'] = True

    return result


def _load_existing_data_for_assignment(assignment_entity_status, form_template):
    """Load and process existing FormData and DynamicIndicatorData for an assignment.
    Returns existing_data_processed dict keyed by field_value[<id>].
    Matrix fields are handled inline during the FormData loop (item_type == 'matrix').
    """
    existing_data_entries = (
        FormData.query
        .filter_by(assignment_entity_status_id=assignment_entity_status.id)
        .options(joinedload(FormData.form_item))
        .all()
    )
    existing_data_processed = {}
    for entry in existing_data_entries:
        if entry.form_item_id:
            form_item = entry.form_item
            if not form_item:
                current_app.logger.warning(
                    f"[DATA_LOADING] FormItem not found for form_item_id={entry.form_item_id}"
                )
                continue
            existing_data_processed.update(_process_form_data_entry(entry, form_item))

    dynamic_data_entries = DynamicIndicatorData.query.filter(
        DynamicIndicatorData.assignment_entity_status_id == assignment_entity_status.id
    ).all()
    for dynamic_data_entry in dynamic_data_entries:
        dynamic_key = f'field_value[dynamic_{dynamic_data_entry.id}]'
        if dynamic_data_entry.disagg_data:
            existing_data_processed[dynamic_key] = dynamic_data_entry.disagg_data
        else:
            existing_data_processed[dynamic_key] = dynamic_data_entry.value
        if dynamic_data_entry.data_not_available:
            existing_data_processed[f'dynamic_{dynamic_data_entry.id}_data_not_available'] = True
        if dynamic_data_entry.not_applicable:
            existing_data_processed[f'dynamic_{dynamic_data_entry.id}_not_applicable'] = True

    return existing_data_processed


def _load_existing_data_for_public_submission(submission):
    """Load and process existing form data entries for a public submission.
    Returns existing_data_processed dict.
    """
    existing_data_processed = {}
    for entry in submission.data_entries.all():
        if not entry.form_item_id:
            continue
        form_item = FormItem.query.get(entry.form_item_id)
        if not form_item:
            current_app.logger.warning(
                f"FormItem {entry.form_item_id} not found for public submission {submission.id}"
            )
            continue
        existing_data_processed.update(_process_form_data_entry(entry, form_item))
    return existing_data_processed


def _prepare_submitted_documents_for_template(submission):
    """Prepare submitted documents dict for entry_form.html.
    Returns dict mapping field keys to single doc or list of docs (most recent first).
    """
    result = {}
    for doc in submission.submitted_documents.order_by(SubmittedDocument.uploaded_at.desc()).all():
        if not doc.form_item_id:
            continue
        doc_key = f"field_value[{doc.form_item_id}]"
        if doc_key not in result:
            result[doc_key] = doc
        else:
            current = result[doc_key]
            if isinstance(current, list):
                current.append(doc)
            else:
                result[doc_key] = [current, doc]
    return result


# Note: get_form_items_for_section is now imported directly from app.utils.form_processing
# No local wrapper needed - using the unified function directly

def map_unified_item_to_original(item_id, item_type):
    """Map a unified item ID to the FormItem.

    Args:
        item_id: The unified item ID from FormItem
        item_type: The FormItemType (indicator, question, document_field)

    Returns:
        tuple: (FormItem instance, item_id) or (None, None) if not found
    """
    # Validate input - item_id must be a valid integer
    if item_id is None:
        return (None, None)

    # Try to convert to int if it's a string
    try:
        if isinstance(item_id, str):
            # Check if it's a valid integer string
            item_id = int(item_id)
        elif not isinstance(item_id, int):
            return (None, None)
    except (ValueError, TypeError):
        return (None, None)

    # Validate item_type
    if not item_type:
        return (None, None)

    try:
        form_item = FormItem.query.filter_by(id=item_id, item_type=item_type).first()
        return (form_item, item_id) if form_item else (None, None)
    except Exception as e:  # SQLAlchemy/DB errors - keep broad for DB layer
        current_app.logger.warning("_resolve_form_item_from_request DB error: %s", e, exc_info=True)
        return (None, None)

# Unified form handling route
@bp.route("/<form_type>/<int:form_id>", methods=["GET", "POST"])
@login_required
def view_edit_form(form_type, form_id):
    """Unified form viewing/editing endpoint for all form types."""

    # Route to appropriate handler based on form type
    if form_type == "assignment":
        return handle_assignment_form(form_id)
    elif form_type == "public-submission":
        # For public submissions, redirect to the specific view route by default for safety
        return redirect(url_for("forms.view_public_submission", submission_id=form_id))
    else:
        flash(_("Invalid form type."), "danger")
        return redirect(url_for("main.dashboard"))

# Legacy compatibility route - redirects to new unified route
@bp.route("/assignment_status/<int:aes_id>", methods=["GET", "POST"])
@login_required
def enter_data(aes_id):
    """Legacy route for backward compatibility - redirects to unified route."""
    return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

# Main assignment form handler (extracted from original enter_data function)
@performance_monitor("Assignment Form Handling")
def handle_assignment_form(aes_id):
    """Handle assignment form viewing/editing for focal points - CLEANED VERSION."""
    forms_logger = debug_manager.get_logger('app.routes.forms')




    # Get assignment entity status with helpful error handling
    from app.services import AssignmentService
    assignment_entity_status = AssignmentService.get_assignment_entity_status_by_id(aes_id)
    if not assignment_entity_status:
        current_app.logger.error(f"AssignmentEntityStatus not found for ID: {aes_id}")
        current_app.logger.error(f"Request method: {request.method}")
        current_app.logger.error(f"Request URL: {request.url}")
        current_app.logger.error(f"User: {current_user.email}")

        # Check all available assignment IDs for this user (country-level AES only)
        available_assignments = AssignmentEntityStatus.query\
            .filter(AssignmentEntityStatus.entity_type == EntityType.country.value)\
            .filter(AssignmentEntityStatus.entity_id.in_([c.id for c in current_user.countries.all()]))\
            .all()
        current_app.logger.error(f"Available AssignmentEntityStatus IDs for current user: {[aes.id for aes in available_assignments]}")

        flash(_("Assignment status for this country not found (ID: %(id)d). Please check the URL or contact support.", id=aes_id), "danger")
        return redirect(url_for("main.dashboard"))

    assignment = assignment_entity_status.assigned_form

    # Deactivated assignment guard: block viewing/submitting inactive assignments
    if assignment is not None and getattr(assignment, "is_active", True) is False:
        flash(_("This assignment is currently inactive."), "warning")
        return redirect(url_for("main.dashboard"))

    # Check access using AuthorizationService (supports all entity types)
    from app.services.authorization_service import AuthorizationService

    if not AuthorizationService.can_access_assignment(assignment_entity_status, current_user):
        entity_type = assignment_entity_status.entity_type
        entity_id = assignment_entity_status.entity_id
        from app.services.entity_service import EntityService
        entity_name = EntityService.get_entity_display_name(entity_type, entity_id)
        current_app.logger.warning(
            f"Access denied for user {current_user.email} to AssignmentEntityStatus {aes_id} "
            f"(Entity: {entity_type} {entity_id} - {entity_name}) - user does not have entity access."
        )
        flash(_("You are not authorized to access this assignment for %(entity)s.", entity=entity_name), "warning")
        return redirect(url_for("main.dashboard"))

    can_edit = AuthorizationService.can_edit_assignment(assignment_entity_status, current_user)
    form_template = assignment.template

    # Use the unified template preparation service for consistent processing
    template, all_sections, available_indicators_by_section = TemplatePreparationService.prepare_template_for_rendering(
        form_template, assignment_entity_status, is_preview_mode=False
    )

    # Resolve template variables and apply to form items
    from app.services.variable_resolution_service import VariableResolutionService
    from app.models import FormTemplateVersion
    import re

    template_version = None
    # Always define these so downstream prefill logic can safely reference them.
    resolved_variables = {}
    variable_configs = {}
    if form_template.published_version_id:
        template_version = FormTemplateVersion.query.get(form_template.published_version_id)

        # Resolve template variables (includes built-in metadata tokens even if no variables are defined)
        if template_version:
            variable_configs = template_version.variables or {}  # Store configs for formatting
            # Only resolve variables if the template actually references any resolvable tokens.
            # This avoids expensive DB lookups for templates that only use plugin placeholders (e.g. [EO1])
            # which are resolved client-side.
            try:
                placeholder_pattern = re.compile(r'\[(\w+)\]')

                def _extract_placeholders_from_sections(sections):
                    names = set()
                    for section in (sections or []):
                        # Section display name
                        sn = getattr(section, 'display_name', None)
                        if sn and '[' in sn:
                            names.update(placeholder_pattern.findall(sn))

                        for field in getattr(section, 'fields_ordered', []) or []:
                            if not field:
                                continue

                            # Common label/description fields
                            for attr in ('label', 'display_label', 'definition', 'description'):
                                t = getattr(field, attr, None)
                                if t and '[' in t:
                                    names.update(placeholder_pattern.findall(t))

                            # Translation dicts
                            for attr in ('label_translations', 'definition_translations', 'description_translations'):
                                d = getattr(field, attr, None)
                                if isinstance(d, dict) and d:
                                    for v in d.values():
                                        if v and '[' in v:
                                            names.update(placeholder_pattern.findall(v))

                            # Matrix manual rows (may contain variables)
                            if getattr(field, 'item_type', None) == 'matrix' and getattr(field, 'config', None):
                                try:
                                    matrix_config = field.config.get('matrix_config') if isinstance(field.config, dict) and 'matrix_config' in field.config else field.config
                                    if isinstance(matrix_config, dict):
                                        row_mode = matrix_config.get('row_mode', 'manual')
                                        if row_mode == 'manual' or not row_mode:
                                            rows = matrix_config.get('rows', [])
                                            if isinstance(rows, list):
                                                for row in rows:
                                                    if isinstance(row, dict):
                                                        for v in row.values():
                                                            if v and '[' in str(v):
                                                                names.update(placeholder_pattern.findall(str(v)))
                                                    elif isinstance(row, str) and '[' in row:
                                                        names.update(placeholder_pattern.findall(row))
                                except Exception as e:
                                    current_app.logger.debug("Placeholder extraction from row failed: %s", e)
                    return names  # After processing all sections

                placeholders = _extract_placeholders_from_sections(all_sections)
                resolvable_token_names = set(variable_configs.keys()) | set(getattr(VariableResolutionService, "_BUILTIN_METADATA_TYPES", {}).keys())
                should_resolve_variables = bool(placeholders & resolvable_token_names)
            except Exception as e:
                current_app.logger.debug("Placeholder extraction failed: %s", e)
                should_resolve_variables = True  # Safe fallback

            try:
                if should_resolve_variables:
                    resolved_variables = VariableResolutionService.resolve_variables(
                        template_version,
                        assignment_entity_status
                    )
                    current_app.logger.debug(
                        f"Variable resolution: Resolved {len(resolved_variables)} variables: {list(resolved_variables.keys())}"
                    )
                else:
                    resolved_variables = {}
            except Exception as e:
                current_app.logger.warning(f"Variable resolution failed: {e}", exc_info=True)
                resolved_variables = {}

            # Apply variable replacements to form item labels and descriptions
            # IMPORTANT: Store resolved values in temporary display-only attributes (prefixed with _display_)
            # to avoid modifying the actual database fields. The template will use these for display.
            # For context-dependent variables (e.g. match_by_indicator_bank), resolve per-field on demand.
            _match_by_bank_cache = {}
            placeholder_pattern = re.compile(r'\[(\w+)\]')

            def _resolved_vars_for_field(field_obj, text_value: str):
                merged = resolved_variables or {}
                if not (template_version and variable_configs and isinstance(variable_configs, dict)):
                    return merged
                if not text_value or '[' not in str(text_value):
                    return merged

                names = set(placeholder_pattern.findall(str(text_value)))
                if not names:
                    return merged

                out = None
                for var_name in names:
                    cfg = variable_configs.get(var_name) if isinstance(variable_configs, dict) else None
                    if not (cfg and isinstance(cfg, dict) and cfg.get('match_by_indicator_bank')):
                        continue
                    val = VariableResolutionService.resolve_variable_by_indicator_bank(
                        cfg,
                        assignment_entity_status,
                        field_obj,
                        cache=_match_by_bank_cache
                    )
                    if out is None:
                        out = dict(merged)
                    out[var_name] = val

                return out if out is not None else merged

            for section in all_sections:
                # Resolve variables in section names (use the already-localized display_name)
                if hasattr(section, 'display_name') and section.display_name and resolved_variables and '[' in section.display_name:
                    try:
                        section._display_name_resolved = VariableResolutionService.replace_variables_in_text(
                            section.display_name,
                            _resolved_vars_for_field(section, section.display_name),
                            variable_configs
                        )
                    except Exception as e:
                        current_app.logger.warning(
                            f"Error resolving variables in section name for section {getattr(section, 'id', 'unknown')}: {e}",
                            exc_info=True
                        )

                for field in getattr(section, 'fields_ordered', []):
                    if not field or not hasattr(field, 'id'):
                        continue

                    # Store resolved label in temporary display attribute (won't be saved to DB)
                    if hasattr(field, 'label') and field.label and resolved_variables and '[' in field.label:
                        original_label = field.label
                        resolved_label = VariableResolutionService.replace_variables_in_text(
                            field.label,
                            _resolved_vars_for_field(field, field.label),
                            variable_configs
                        )
                        # Store in temporary attribute for display only
                        field._display_label = resolved_label
                        # Removed logging to improve performance - use DEBUG level if needed

                    # Store resolved display_label in temporary display attribute
                    if hasattr(field, 'display_label') and field.display_label and resolved_variables and '[' in field.display_label:
                        original_display_label = field.display_label
                        resolved_display_label = VariableResolutionService.replace_variables_in_text(
                            field.display_label,
                            _resolved_vars_for_field(field, field.display_label),
                            variable_configs
                        )
                        # Store in temporary attribute for display only
                        field._display_label_resolved = resolved_display_label
                        # Removed logging to improve performance - use DEBUG level if needed

                    # Store resolved label_translations in temporary display attribute
                    if hasattr(field, 'label_translations') and field.label_translations and resolved_variables:
                        resolved_label_translations = {}
                        for lang_code, translated_label in field.label_translations.items():
                            if translated_label and '[' in translated_label:
                                original_translated = translated_label
                                resolved_translated = VariableResolutionService.replace_variables_in_text(
                                    translated_label,
                                    _resolved_vars_for_field(field, translated_label),
                                    variable_configs
                                )
                                resolved_label_translations[lang_code] = resolved_translated
                            elif translated_label:
                                # Keep untouched translation for this language
                                resolved_label_translations[lang_code] = translated_label
                        # Store in temporary attribute for display only
                        field._display_label_translations = resolved_label_translations

                    # Store resolved definition in temporary display attribute
                    if hasattr(field, 'definition') and field.definition and resolved_variables and '[' in field.definition:
                        original_definition = field.definition
                        resolved_definition = VariableResolutionService.replace_variables_in_text(
                            field.definition,
                            _resolved_vars_for_field(field, field.definition),
                            variable_configs
                        )
                        # Store in temporary attribute for display only
                        field._display_definition = resolved_definition
                        # Removed logging to improve performance - use DEBUG level if needed

                    # Store resolved description in temporary display attribute
                    if hasattr(field, 'description') and field.description and resolved_variables and '[' in field.description:
                        original_description = field.description
                        resolved_description = VariableResolutionService.replace_variables_in_text(
                            field.description,
                            _resolved_vars_for_field(field, field.description),
                            variable_configs
                        )
                        # Store in temporary attribute for display only
                        field._display_description = resolved_description
                        # Removed logging to improve performance - use DEBUG level if needed

                    # Store resolved definition_translations and description_translations in temporary display attributes
                    if hasattr(field, 'definition_translations') and field.definition_translations:
                        resolved_definition_translations = {}
                        for lang_code, translated_def in field.definition_translations.items():
                            if translated_def and resolved_variables and '[' in translated_def:
                                original_translated = translated_def
                                resolved_translated = VariableResolutionService.replace_variables_in_text(
                                    translated_def,
                                    _resolved_vars_for_field(field, translated_def),
                                    variable_configs
                                )
                                resolved_definition_translations[lang_code] = resolved_translated
                            elif translated_def:
                                resolved_definition_translations[lang_code] = translated_def
                        # Store in temporary attribute for display only
                        field._display_definition_translations = resolved_definition_translations

                    if hasattr(field, 'description_translations') and field.description_translations:
                        resolved_description_translations = {}
                        for lang_code, translated_desc in field.description_translations.items():
                            if translated_desc and resolved_variables and '[' in translated_desc:
                                original_translated = translated_desc
                                resolved_translated = VariableResolutionService.replace_variables_in_text(
                                    translated_desc,
                                    _resolved_vars_for_field(field, translated_desc),
                                    variable_configs
                                )
                                resolved_description_translations[lang_code] = resolved_translated
                            elif translated_desc:
                                resolved_description_translations[lang_code] = translated_desc
                        # Store in temporary attribute for display only
                        field._display_description_translations = resolved_description_translations

                    # Resolve variables in matrix row labels (for manual mode matrices)
                    if hasattr(field, 'item_type') and field.item_type == 'matrix' and hasattr(field, 'config') and field.config:
                        try:
                            # Get matrix config (handle both nested and flat structure)
                            matrix_config = field.config.get('matrix_config') if isinstance(field.config, dict) and 'matrix_config' in field.config else field.config
                            if isinstance(matrix_config, dict):
                                row_mode = matrix_config.get('row_mode', 'manual')
                                # Only resolve variables for manual mode rows (list library mode uses lookup list data)
                                if row_mode == 'manual' or not row_mode:
                                    rows = matrix_config.get('rows', [])
                                    if rows and isinstance(rows, list):
                                        resolved_rows = []
                                        for row in rows:
                                            if isinstance(row, str) and resolved_variables and '[' in row:
                                                resolved_row = VariableResolutionService.replace_variables_in_text(
                                                    row,
                                                    _resolved_vars_for_field(field, row),
                                                    variable_configs
                                                )
                                                resolved_rows.append(resolved_row)
                                                # Removed logging to improve performance - use DEBUG level if needed
                                            else:
                                                resolved_rows.append(row)  # Keep non-string rows as-is
                                        # Store resolved rows in temporary attribute for display only
                                        field._display_matrix_rows = resolved_rows
                        except Exception as e:
                            current_app.logger.warning(f"Error resolving variables in matrix row labels for field {field.id}: {e}", exc_info=True)

    # Create the final sections list (main sections)
    db_sections = [s for s in all_sections if s.parent_section_id is None]

    # Published pages list for navigation/sorting
    published_pages = (
        FormPage.query
        .filter_by(template_id=form_template.id, version_id=form_template.published_version_id)
        .order_by(FormPage.order)
        .all()
    )

    csrf_form = FlaskForm()
    SEX_CATEGORIES = Config.DEFAULT_SEX_CATEGORIES

    existing_data_processed = _load_existing_data_for_assignment(
        assignment_entity_status, form_template
    )

    # Prefill number fields with defaults / variables if not already set
    # - Indicator default_value can be a literal or a template variable like [var_name]
    # - Variable-based prefill from label/description requires resolved_variables
    for section in all_sections:
        for field in getattr(section, 'fields_ordered', []):
            if not field or not hasattr(field, 'id'):
                continue

            field_key = f'field_value[{field.id}]'

            # Skip if field already has data
            if field_key in existing_data_processed:
                continue

            # Check if this is a numeric-like field (indicator/question)
            # NOTE: indicator bank types are not always normalized (e.g. 'Number', 'currency').
            def _is_numeric_type(t):
                try:
                    return str(t or '').strip().lower() in ('number', 'integer', 'float', 'currency', 'percentage')
                except Exception as e:
                    current_app.logger.debug("_is_numeric_type failed: %s", e)
                    return False

            is_number_field = False
            if hasattr(field, 'is_indicator') and field.is_indicator:
                is_number_field = _is_numeric_type(getattr(field, 'type', None))
            elif hasattr(field, 'is_question') and field.is_question:
                is_number_field = _is_numeric_type(getattr(field, 'type', None))

            if is_number_field:
                # 1) Explicit indicator default value (literal or template variable)
                # If the indicator is disaggregation-enabled, the default is applied to Total mode only.
                if getattr(field, 'is_indicator', False):
                    try:
                        cfg = getattr(field, 'config', None)
                        dv_raw = None
                        if isinstance(cfg, dict):
                            dv_raw = cfg.get('default_value')
                        dv_raw = str(dv_raw).strip() if dv_raw is not None else ''

                        if dv_raw:
                            # If the default value uses a variable token, we must have resolved_variables.
                            # The standard placeholder-scanner doesn't inspect config, so resolve on-demand.
                            if ('[' in dv_raw) and (not resolved_variables) and template_version:
                                try:
                                    resolved_variables = VariableResolutionService.resolve_variables(
                                        template_version,
                                        assignment_entity_status
                                    )
                                except Exception as e:
                                    current_app.logger.debug("resolve_variables for default_value failed: %s", e)
                                    resolved_variables = resolved_variables or {}

                            dv_resolved = dv_raw
                            if '[' in dv_raw:
                                dv_resolved = VariableResolutionService.replace_variables_in_text(
                                    dv_raw,
                                    _resolved_vars_for_field(field, dv_raw),
                                    variable_configs
                                )
                            dv_resolved_str = str(dv_resolved).strip()
                            # Normalize common numeric formats ("1,234" -> "1234")
                            dv_resolved_str = dv_resolved_str.replace(',', '')

                            num_value = float(dv_resolved_str)
                            # Preserve integer-ness when possible (nicer rendering)
                            if num_value.is_integer():
                                num_value = int(num_value)

                            allowed_opts = []
                            try:
                                allowed_opts = list(getattr(field, 'allowed_disaggregation_options', []) or [])
                            except Exception as e:
                                current_app.logger.debug("allowed_disaggregation_options failed: %s", e)
                                allowed_opts = []

                            # If disagg-enabled: only apply to Total mode (and only if total is an allowed mode)
                            if allowed_opts and len(allowed_opts) > 1:
                                if 'total' not in allowed_opts:
                                    # No Total mode to apply to
                                    continue
                                # Use dict structure so the UI starts in Total mode with the default applied.
                                if getattr(field, 'indirect_reach', False):
                                    existing_data_processed[field_key] = {'mode': 'total', 'values': {'direct': num_value}}
                                else:
                                    existing_data_processed[field_key] = {'mode': 'total', 'values': {'total': num_value}}
                            else:
                                # Non-disaggregated numeric indicators/questions can use a simple scalar value.
                                existing_data_processed[field_key] = num_value

                            existing_data_processed[f'{field_key}_is_prefilled'] = True
                            continue
                    except Exception as e:
                        current_app.logger.debug("Default value parsing failed: %s", e)

                # 2) Variable-driven prefill from label/description (legacy behavior)
                if resolved_variables:
                    # Check if label or description contains variable references
                    label_text = getattr(field, 'label', '') or ''
                    desc_text = getattr(field, 'definition', '') or getattr(field, 'description', '') or ''
                    combined_text = f"{label_text} {desc_text}"

                    # Find variable references in the text
                    import re
                    variable_pattern = r'\[(\w+)\]'
                    matches = re.findall(variable_pattern, combined_text)

                    if matches:
                        # Use the first matching variable that resolves to a number
                        for var_name in matches:
                            if var_name in resolved_variables:
                                var_value = resolved_variables[var_name]
                                if var_value is not None:
                                    try:
                                        # Try to convert to number
                                        num_value = float(var_value)
                                        existing_data_processed[field_key] = num_value

                                        # Mark as prefilled and check if variable is readonly
                                        existing_data_processed[f'{field_key}_is_prefilled'] = True
                                        if template_version and template_version.variables:
                                            var_config = template_version.variables.get(var_name, {})
                                            if var_config.get('is_readonly', False):
                                                existing_data_processed[f'{field_key}_is_readonly'] = True

                                        break  # Use first matching variable
                                    except (ValueError, TypeError):
                                        continue

    # Load existing repeat group data (simplified)
    repeat_data_entries = RepeatGroupData.query.join(
        RepeatGroupInstance,
        RepeatGroupData.repeat_instance_id == RepeatGroupInstance.id
    ).filter(
        RepeatGroupInstance.assignment_entity_status_id == assignment_entity_status.id
    ).all()

    # Load existing repeat group data with proper structure
    repeat_instances = RepeatGroupInstance.query.filter_by(
        assignment_entity_status_id=assignment_entity_status.id,
        is_hidden=False
    ).all()

    # Pre-load all repeat data entries in a single query to avoid N+1
    repeat_instance_ids = [instance.id for instance in repeat_instances]
    all_repeat_data_entries = {}
    if repeat_instance_ids:
        repeat_entries = RepeatGroupData.query.filter(
            RepeatGroupData.repeat_instance_id.in_(repeat_instance_ids)
        ).all()
        # Group by instance ID
        for entry in repeat_entries:
            if entry.repeat_instance_id not in all_repeat_data_entries:
                all_repeat_data_entries[entry.repeat_instance_id] = []
            all_repeat_data_entries[entry.repeat_instance_id].append(entry)

    repeat_groups_data = {}
    for repeat_instance in repeat_instances:
        section_id = repeat_instance.section_id
        instance_number = repeat_instance.instance_number

        if section_id not in repeat_groups_data:
            repeat_groups_data[section_id] = {}

        if instance_number not in repeat_groups_data[section_id]:
            repeat_groups_data[section_id][instance_number] = {
                'id': repeat_instance.id,
                'label': repeat_instance.instance_label or '',
                'hidden': repeat_instance.is_hidden,
                'data': {}
            }

        # Load repeat data entries for this instance (using pre-loaded data)
        repeat_data_entries = all_repeat_data_entries.get(repeat_instance.id, [])

        for repeat_data_entry in repeat_data_entries:
            # Use field ID as key (frontend expects this format)
            if repeat_data_entry.form_item_id:
                field_key = str(repeat_data_entry.form_item_id)
            else:
                continue

            # Check disagg_data first for JSON data (new structure)
            if repeat_data_entry.disagg_data:
                # JSON data is stored in disagg_data
                actual_value = repeat_data_entry.disagg_data
                data_not_available = repeat_data_entry.data_not_available
                not_applicable = repeat_data_entry.not_applicable
                display_value = actual_value  # Keep as dict for frontend processing
            else:
                # Simple value is stored in value column (now VARCHAR)
                actual_value = repeat_data_entry.value
                data_not_available = repeat_data_entry.data_not_available
                not_applicable = repeat_data_entry.not_applicable
                display_value = actual_value

            # Create field data structure with value and data availability flags
            field_data = {
                'value': display_value,
                'data_not_available': data_not_available,
                'not_applicable': not_applicable
            }

            repeat_groups_data[section_id][instance_number]['data'][field_key] = field_data

    # For repeat sections, also include original field data as instance 1
    for section in all_sections:
        if section.section_type == 'repeat':
            section_id = section.id

            # Check if we already have instance 1 data from database
            has_instance_1 = section_id in repeat_groups_data and 1 in repeat_groups_data[section_id]

            if not has_instance_1:
                # Look for original field data and add it as instance 1
                instance_1_data = {}
                for field in section.fields_ordered:
                    if hasattr(field, 'indicator_bank_id'):
                        # This is an indicator - use field ID as key
                        field_key = str(field.id)
                        original_key = f"field_value[{field.id}]"
                        if original_key in existing_data_processed:
                            original_value = existing_data_processed[original_key]
                            # For indicators with disaggregation, preserve the full JSON structure
                            if isinstance(original_value, dict) and 'mode' in original_value and 'values' in original_value:
                                # This is structured indicator data - preserve the full structure for frontend processing
                                instance_1_data[field_key] = json.dumps(original_value)
                            else:
                                instance_1_data[field_key] = original_value
                    elif hasattr(field, 'question_type'):
                        # This is a question - use field ID as key
                        field_key = str(field.id)
                        original_key = f"field_value[{field.id}]"
                        if original_key in existing_data_processed:
                            original_value = existing_data_processed[original_key]
                            instance_1_data[field_key] = original_value

                if instance_1_data:
                    # Add instance 1 data to repeat_groups_data
                    if section_id not in repeat_groups_data:
                        repeat_groups_data[section_id] = {}

                    repeat_groups_data[section_id][1] = {
                        'id': None,  # No database instance yet
                        'label': '',
                        'hidden': False,
                        'data': instance_1_data
                    }

    # Add repeat groups data to existing_data_processed for template access
    existing_data_processed['repeat_groups_data'] = repeat_groups_data

    # Load existing submitted documents (support multiple docs per field)
    submitted_docs = (
        SubmittedDocument.query.filter_by(assignment_entity_status_id=assignment_entity_status.id)
        .order_by(SubmittedDocument.uploaded_at.desc())
        .all()
    )

    # Values are either a single SubmittedDocument (back-compat) or a list[SubmittedDocument] (most recent first).
    existing_submitted_documents_dict = {}
    for doc in submitted_docs:
        if not doc.form_item_id:
            continue
        key = f'field_value[{doc.form_item_id}]'
        if key not in existing_submitted_documents_dict:
            existing_submitted_documents_dict[key] = doc
        else:
            current = existing_submitted_documents_dict[key]
            if isinstance(current, list):
                current.append(doc)
            else:
                existing_submitted_documents_dict[key] = [current, doc]

    # Load section completion statuses (simplified)
    section_statuses = calculate_section_completion_status(
        all_sections, existing_data_processed, existing_submitted_documents_dict
    )

    if request.method == "POST":
        # Must be available for all POST error paths (including CSRF failure branch).
        is_ajax = is_json_request()

        if not can_edit:
             entity_name = EntityService.get_entity_display_name(assignment_entity_status.entity_type, assignment_entity_status.entity_id)
             flash(_("This assignment for %(entity)s is in '%(status)s' status and cannot be edited by you at this time.", entity=entity_name, status=assignment_entity_status.status), "warning")
             return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=assignment_entity_status.id))

        if csrf_form.validate_on_submit():
            action = request.form.get('action')

            try:
                # Use FormDataService for all form processing
                submission_result = FormDataService.process_form_submission(
                    assignment_entity_status, all_sections, csrf_form
                )
                if not submission_result['success']:
                    for error in submission_result['validation_errors']:
                        flash(error, "danger")
                    # Check if this is an AJAX request
                    if is_ajax:
                        msg = '; '.join(submission_result['validation_errors'])
                        return json_bad_request(msg, success=False)
                    else:
                        redirect_url = url_for("forms.view_edit_form", form_type="assignment", form_id=assignment_entity_status.id)
                        return redirect(redirect_url)

                # Log field changes for activity tracking
                field_changes_tracker = submission_result['field_changes']

                # Helper function to parse field values for better readability
                def parse_field_value(value, data_not_available=None, not_applicable=None):
                    """Parse field value to extract meaningful information for display."""
                    # Handle data availability flags first
                    if data_not_available:
                        return "Data not available"
                    if not_applicable:
                        return "Not applicable"

                    if value is None:
                        return "N/A"

                    if isinstance(value, dict):
                        # Handle complex field structures
                        if 'mode' in value and 'values' in value:
                            # Handle disaggregation fields
                            mode = value.get('mode', '')
                            values = value.get('values', {})
                            if mode == 'total' and 'total' in values:
                                return f"Total: {values['total']}"
                            elif mode == 'disaggregated' and values:
                                # Show first few disaggregated values
                                items = list(values.items())[:3]  # Show first 3
                                result = ", ".join([f"{k}: {v}" for k, v in items])
                                if len(values) > 3:
                                    result += f" (+{len(values) - 3} more)"
                                return result
                            else:
                                return str(values)
                        elif 'values' in value:
                            # Handle other value structures
                            return str(value['values'])
                        else:
                            # Handle simple dictionaries like {'direct': 89} for non-disaggregated fields
                            # Extract the direct value if it exists, otherwise show the whole structure
                            if 'direct' in value and len(value) == 1:
                                direct_value = value['direct']
                                # Check if direct_value is a nested dictionary (age group breakdown)
                                if isinstance(direct_value, dict):
                                    # Format age group breakdown nicely
                                    parts = []
                                    for age_group, count in direct_value.items():
                                        if count and count != 0:  # Only show non-zero values
                                            parts.append(f"{age_group}: {count}")
                                    return ", ".join(parts) if parts else "0"
                                else:
                                    return str(direct_value)
                            elif 'total' in value and len(value) == 1:
                                total_value = value['total']
                                # Check if total_value is a nested dictionary (age group breakdown)
                                if isinstance(total_value, dict):
                                    # Format age group breakdown nicely
                                    parts = []
                                    for age_group, count in total_value.items():
                                        if count and count != 0:  # Only show non-zero values
                                            parts.append(f"{age_group}: {count}")
                                    return ", ".join(parts) if parts else "0"
                                else:
                                    return str(total_value)
                            elif value.get('_matrix_change') or any(
                                    isinstance(k, str) and '_' in k and not k.startswith('_')
                                    for k in value.keys()):
                                if not value.get('_matrix_change'):
                                    value = dict(value)
                                    value['_matrix_change'] = True
                                return value
                            else:
                                return str(value)
                    elif isinstance(value, str):
                        # Truncate long strings
                        return value[:100] + "..." if len(value) > 100 else value
                    else:
                        return str(value)

                # Group all field changes into a single activity
                meaningful_changes = []

                for change in field_changes_tracker:
                    # Get data availability flags from the change
                    old_data_not_available = change.get('old_data_not_available', False)
                    new_data_not_available = change.get('new_data_not_available', False)
                    old_not_applicable = change.get('old_not_applicable', False)
                    new_not_applicable = change.get('new_not_applicable', False)

                    # Parse values with data availability context
                    old_value_parsed = parse_field_value(change.get('old_value'), old_data_not_available, old_not_applicable)
                    new_value_parsed = parse_field_value(change.get('new_value'), new_data_not_available, new_not_applicable)

                    # Skip only if both values are exactly the same
                    if old_value_parsed == new_value_parsed:
                        continue

                    # Determine change type and format display values
                    if old_value_parsed == 'N/A':
                        # Field was empty, now has a value - this is an "added" operation
                        change_type = 'added'
                        display_old_value = None  # Don't show N/A for added fields
                        display_new_value = new_value_parsed
                    elif new_value_parsed == 'N/A':
                        # Field had a value, now is empty - this is a "removed" operation
                        change_type = 'removed'
                        display_old_value = old_value_parsed
                        display_new_value = None
                    else:
                        # Field had a value and now has a different value - this is an "updated" operation
                        change_type = 'updated'
                        display_old_value = old_value_parsed
                        display_new_value = new_value_parsed

                    # Format the field change for the template
                    formatted_change = {
                        'type': change_type,
                        'form_item_id': change.get('form_item_id'),  # Preserve form_item_id
                        # Optional hint for dynamic indicator sections (IndicatorBank id vs FormItem id)
                        'field_id_kind': change.get('field_id_kind'),
                        'field_name': change.get('field_name', f"Field {change.get('form_item_id', 'Unknown')}"),
                        'old_value': display_old_value,
                        'new_value': display_new_value
                    }
                    meaningful_changes.append(formatted_change)

                # Log a single activity capturing all changes in summary_params
                if meaningful_changes:
                    try:
                        if len(meaningful_changes) == 1:
                            ch = meaningful_changes[0]
                            summary_key = 'activity.form_data_updated.single'
                            summary_params = {
                                'field': ch['field_name'],
                                'field_id': ch.get('form_item_id'),
                                'field_id_kind': ch.get('field_id_kind'),
                                'old': ch.get('old_value') or '',
                                'new': ch.get('new_value') or '',
                                'change_type': ch.get('type') or 'updated',
                                # Include template name so activity cards can show which
                                # assignment/template this change belongs to.
                                'template': assignment_entity_status.assigned_form.template.name
                                            if assignment_entity_status
                                            and assignment_entity_status.assigned_form
                                            and assignment_entity_status.assigned_form.template
                                            else None,
                            }
                            # Use a generic description that will be translated by render_activity_summary
                            activity_description = "Field updated"
                        else:
                            # Aggregate into a single payload with detailed changes list
                            changes_list = []
                            # Count different change types for better description
                            change_type_counts = {'added': 0, 'updated': 0, 'removed': 0}
                            for ch in meaningful_changes:
                                ct = (ch.get('type') or 'updated')
                                change_type_counts[ct] = change_type_counts.get(ct, 0) + 1
                                changes_list.append({
                                    'field': ch.get('field_name', 'Field'),
                                    'field_id': ch.get('form_item_id'),
                                    'field_id_kind': ch.get('field_id_kind'),
                                    'old': ch.get('old_value') or '',
                                    'new': ch.get('new_value') or '',
                                    'change_type': ct
                                })

                            # Use a generic description that will be translated by render_activity_summary
                            total_changes = len(meaningful_changes)
                            if change_type_counts['added'] == total_changes:
                                activity_description = "Fields added"
                                dominant_type = 'added'
                            elif change_type_counts['removed'] == total_changes:
                                activity_description = "Fields removed"
                                dominant_type = 'removed'
                            elif change_type_counts['updated'] == total_changes:
                                activity_description = "Fields updated"
                                dominant_type = 'updated'
                            else:
                                # Mixed change types
                                activity_description = "Fields modified"
                                dominant_type = 'updated'  # Default to updated for mixed types

                            summary_key = 'activity.form_data_updated.multiple'
                            summary_params = {
                                'count': total_changes,
                                'template': assignment_entity_status.assigned_form.template.name,
                                'change_type': dominant_type,
                                'changes': changes_list
                            }

                        log_entity_activity(
                            assignment_entity_status.entity_type,
                            assignment_entity_status.entity_id,
                            'data_update',
                            activity_description,
                            summary_key=summary_key,
                            summary_params=summary_params,
                            related_object_type='form_data',
                            related_object_id=assignment_entity_status.id,
                            assignment_id=assignment_entity_status.id,
                            user_id=current_user.id
                        )
                    except Exception as e:
                        current_app.logger.error(f"Error logging grouped activity: {e}", exc_info=True)

                # Handle submission vs save actions
                if submission_result.get('submitted'):
                    # Form was submitted
                    notify_assignment_submitted(assignment_entity_status)
                    flash("Assignment submitted successfully!", "success")

                    # Check if this is an AJAX request
                    is_ajax = is_json_request()
                    if is_ajax:
                        return json_ok(
                            message="Assignment submitted successfully!",
                            redirect_url=url_for("main.dashboard")
                        )
                    else:
                        return redirect(url_for("main.dashboard"))
                else:
                    # Data was saved
                    # Re-check if this is an AJAX request
                    is_ajax = is_json_request()
                    if is_ajax:
                        return json_ok(message="Progress saved successfully.")
                    else:
                        flash("Progress saved successfully.", "success")
                        redirect_url = url_for("forms.view_edit_form", form_type="assignment", form_id=assignment_entity_status.id)
                        return redirect(redirect_url)

            except Exception as e:
                request_transaction_rollback()
                error_message = GENERIC_ERROR_MESSAGE
                current_app.logger.exception("Error during data save/submit for AssignmentEntityStatus %s: %s", aes_id, e)


                if is_ajax:
                    return json_server_error(error_message, success=False)
                else:
                    # Traditional flash message and redirect for non-AJAX requests
                    flash(error_message, "danger")
                    return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=assignment_entity_status.id))
        else:
            current_app.logger.error("CSRF validation failed")
            current_app.logger.error(f"CSRF form errors: {csrf_form.errors}")
            error_message = "Form submission failed due to a security issue or validation errors. Please try again."
            current_app.logger.error(f"CSRF or other form validation failed for ACS {aes_id}. Errors: {csrf_form.errors}")


            if is_ajax:
                return json_bad_request(error_message, success=False)
            else:
                # Traditional flash message and redirect for non-AJAX requests
                flash(error_message, "danger")
                return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=assignment_entity_status.id))


    # Use the original form template as template_structure (do NOT assign to ORM relationships)
    template_structure = form_template

    # Calculate section statuses using the corrected function
    # This function already returns a dictionary with section names as keys
    section_statuses = calculate_section_completion_status(all_sections, existing_data_processed, existing_submitted_documents_dict)

    # Add display filter configuration for dynamic sections
    for section in all_sections:
        if section.section_type == 'dynamic_indicators':
            section.data_entry_display_filters_config = getattr(section, 'data_entry_display_filters_list', [])



    # Apply page + section translations (used by entry_form.html for sidebar + section headers)
    # NOTE: FormSection has no persisted "display_name" column; we attach display_name at runtime.
    current_locale = (get_locale() or 'en')
    current_locale_short = current_locale.split('_', 1)[0] if isinstance(current_locale, str) else 'en'

    # Pages (for paginated templates)
    try:
        for page in (published_pages or []):
            page.display_name = get_localized_page_name(page)
    except Exception as e:
        # Never fail form rendering due to localization issues, but don't swallow silently.
        current_app.logger.debug("Page localization failed; continuing without localized page names: %s", e, exc_info=True)

    # Ensure section.page.display_name is localized as well (pages referenced from sections)
    page_ids_processed = set()
    for section in (all_sections or []):
        if getattr(section, 'page', None) and getattr(section.page, 'id', None) not in page_ids_processed:
            try:
                section.page.display_name = get_localized_page_name(section.page)
            except Exception as e:
                current_app.logger.debug(
                    "Page localization failed for page_id=%s; continuing: %s",
                    getattr(section.page, "id", None),
                    e,
                    exc_info=True,
                )
            page_ids_processed.add(section.page.id)

    # Sections (including sub-sections)
    for section in (all_sections or []):
        translated_name = None
        nt = getattr(section, 'name_translations', None)
        if isinstance(nt, dict) and nt:
            # ISO codes only
            for k in (current_locale_short, 'en'):
                val = nt.get(k)
                if val and str(val).strip():
                    translated_name = str(val).strip()
                    break
        section.display_name = translated_name if translated_name else getattr(section, 'name', '')

    # Render the template
    # Add template_id and assignment_entity_status_id to template context for JavaScript variable resolution
    return render_template(
        "forms/entry_form/entry_form.html",
        template_structure=template_structure,  # Pass the complete template with all sections
        sections=db_sections,  # Main sections only for display structure
        all_sections=all_sections,  # Full section list for navigation/grouping
        published_pages=published_pages,
        assignment_status=assignment_entity_status,  # AssignmentEntityStatus object for all entity types
        assignment=assignment,
        existing_data=existing_data_processed,
        existing_submitted_documents=existing_submitted_documents_dict,
        can_edit=can_edit,
        csrf_form=csrf_form,
        sex_categories=SEX_CATEGORIES,
        form_type="assignment",
        section_statuses=section_statuses,
        available_indicators_by_section=available_indicators_by_section,
        # Add missing variables from original
        slugify_age_group=slugify_age_group,
        config=Config,
        QuestionType=QuestionType,
        isinstance=isinstance,
        json=json,
        hasattr=hasattr,
        get_localized_country_name=get_localized_country_name,
        form_action=url_for("forms.view_edit_form", form_type="assignment", form_id=assignment_entity_status.id),
        translation_key=get_translation_key(),
        # Add localization functions for template use
        get_localized_indicator_definition=get_localized_indicator_definition,
        get_localized_indicator_type=get_localized_indicator_type,
        get_localized_indicator_unit=get_localized_indicator_unit,
        get_localized_template_name=get_localized_template_name,
        # Add plugin manager information for template use
        plugin_manager=current_app.plugin_manager if hasattr(current_app, 'plugin_manager') else None,
        form_integration=current_app.form_integration if hasattr(current_app, 'form_integration') else None,
        # Add IDs for JavaScript variable resolution
        template_id=form_template.id,
        assignment_entity_status_id=assignment_entity_status.id,
        # Add template variables for JavaScript auto-load functionality
        template_variables=variable_configs if 'variable_configs' in locals() else {}
    )


def calculate_section_completion_status(all_sections, existing_data_processed, existing_submitted_documents_dict):
    """Calculate completion status for sections - returns dict format expected by template."""
    section_statuses = {}
    for section in all_sections:
        total_items_in_section = 0
        filled_items_count = 0
        if hasattr(section, 'fields_ordered'):
            for field in section.fields_ordered:
                # Skip Blank/Note fields - they don't count toward completion at all
                if hasattr(field, 'field_type_for_js') and field.field_type_for_js.lower() == 'blank':
                    continue

                total_items_in_section +=1

                # Handle dynamic indicators differently
                if hasattr(field, 'dynamic_assignment_id'):
                    item_key = f"field_value[dynamic_{field.dynamic_assignment_id}]"
                    not_applicable_key = f"dynamic_{field.dynamic_assignment_id}_not_applicable"
                else:
                    item_key = f"field_value[{field.id}]"
                    # Determine the field type for not_applicable key
                    if field.is_indicator:
                        not_applicable_key = f"indicator_{field.id}_not_applicable"
                    elif field.is_question:
                        not_applicable_key = f"question_{field.id}_not_applicable"
                    else:
                        not_applicable_key = f"field_{field.id}_not_applicable"

                # Check if field is marked as "Not applicable" - if so, count as completed
                if existing_data_processed.get(not_applicable_key):
                    filled_items_count += 1
                elif field.is_document_field:
                    if field.is_required_for_js and item_key in existing_submitted_documents_dict:
                        filled_items_count +=1
                    elif not field.is_required_for_js and item_key in existing_submitted_documents_dict:
                        filled_items_count +=1
                else:
                    entry_data = existing_data_processed.get(item_key)
                    if entry_data is not None:
                        if isinstance(entry_data, dict) and 'values' in entry_data:
                             # Disaggregated indicator: filled if any breakdown value has data
                             if any(str(v).strip() for v in entry_data['values'].values() if v is not None):
                                  filled_items_count += 1
                        elif hasattr(field, 'is_matrix') and field.is_matrix and isinstance(entry_data, dict):
                            # Matrix table: filled if ANY cell has a non-empty value
                            # (ignore internal metadata keys)
                            if any(
                                v is not None and str(v).strip() != ''
                                for k, v in entry_data.items()
                                if not k.startswith('_')
                            ):
                                filled_items_count += 1
                        elif field.field_type_for_js == 'CHECKBOX':
                            if entry_data == 'true' or entry_data is True:
                                filled_items_count += 1
                        elif entry_data is not None and str(entry_data).strip():
                             filled_items_count += 1

        # Calculate status using original logic
        if total_items_in_section == 0:
            section_statuses[section.name] = 'N/A'
        elif filled_items_count == 0:
            section_statuses[section.name] = 'Not Started'
        elif filled_items_count < total_items_in_section:
            section_statuses[section.name] = 'In Progress'
        else:
            section_statuses[section.name] = 'Completed'

    return section_statuses



def handle_public_submission_form(submission_id, is_edit_mode=False):
    """Handle public submission form viewing/editing for admins and focal points."""
    submission = PublicSubmission.query.options(
        db.joinedload(PublicSubmission.assigned_form).joinedload(AssignedForm.template),
        db.joinedload(PublicSubmission.country)
    ).get_or_404(submission_id)

    from app.services.authorization_service import AuthorizationService
    # Determine edit mode based on RBAC
    if AuthorizationService.is_system_manager(current_user) or AuthorizationService.has_rbac_permission(current_user, 'admin.assignments.public_submissions.manage'):
        can_edit = True
    elif AuthorizationService.has_country_access(current_user, submission.country_id) and AuthorizationService.has_rbac_permission(current_user, 'assignment.enter'):
        can_edit = True
    else:
        can_edit = False

    # Override with URL parameter if provided
    if request.args.get('edit') == 'true':
        can_edit = True
    elif request.args.get('edit') == 'false':
        can_edit = False

    form_template = submission.assigned_form.template

    # Get all sections for the PUBLISHED version only
    all_sections = FormSection.query.filter_by(
        template_id=form_template.id,
        version_id=form_template.published_version_id
    ).order_by(FormSection.order).all()

    # Separate main sections from sub-sections for proper hierarchical processing (same as assignment forms)
    main_sections = []
    sub_sections_by_parent = {}

    for section_obj in all_sections:
        if section_obj.parent_section_id is None:
            main_sections.append(section_obj)
        else:
            parent_id = section_obj.parent_section_id
            if parent_id not in sub_sections_by_parent:
                sub_sections_by_parent[parent_id] = []
            sub_sections_by_parent[parent_id].append(section_obj)

    # Use all_sections instead of just main sections for complete processing
    sections = all_sections

    # can_edit is already determined above based on user role and permissions

    # Create a dummy assignment country status for compatibility with form processing functions
    class DummyACS:
        def __init__(self, submission_id, assigned_form, country):
            self.id = submission_id  # Use submission ID instead of None
            self.assigned_form = assigned_form  # Template expects this
            self.country = country  # Template expects this
            self.status = "Public Submission"  # Template expects this
            self.country_id = country.id  # Template might expect this

    dummy_acs = DummyACS(submission.id, submission.assigned_form, submission.country)

    # Populate fields_ordered for each section (required by entry_form.html template)
    for section in sections:
        section.fields_ordered = get_form_items_for_section(section, dummy_acs)

        # Layout properties are now handled by properties that read from config

        # Add display filter configuration for dynamic sections
        if section.section_type == 'dynamic_indicators':
            section.data_entry_display_filters_config = getattr(section, 'data_entry_display_filters_list', [])

    # Handle POST request (saving edited data)
    if request.method == "POST" and can_edit:
        csrf_form = FlaskForm()
        if csrf_form.validate_on_submit():
            try:
                # Process form data similar to assignment forms
                action = request.form.get('action', 'save')

                # Handle country change if submitted
                new_country_id = request.form.get('country_id')
                if new_country_id and new_country_id != str(submission.country_id):
                    try:
                        new_country = Country.query.get(int(new_country_id))
                        if new_country:
                            submission.country_id = new_country.id
                            db.session.flush()
                            flash(f'Country changed to {new_country.name}', 'success')
                    except (ValueError, TypeError):
                        flash('Invalid country selection', 'danger')

                # Use the unified FormDataService for comprehensive data processing
                from app.services.form_data_service import FormDataService

                # Get all sections for the assigned form
                sections = submission.assigned_form.sections_ordered

                # Process form data using the unified service (no CSRF needed for admin editing)
                submission_result = FormDataService.process_form_submission(
                    submission, sections, csrf_form=None
                )

                if not submission_result['success']:
                    for error in submission_result['validation_errors']:
                        flash(error, "danger")
                    return redirect(url_for("forms.edit_public_submission", submission_id=submission_id))

                # Log field changes for activity tracking
                field_changes_tracker = submission_result['field_changes']

                # Helper function to parse field values for better readability
                def parse_field_value_for_display(value):
                    if value is None:
                        return "None"
                    elif isinstance(value, dict) and 'values' in value:
                        # Disaggregated data
                        total = sum(v for v in value['values'].values() if isinstance(v, (int, float)))
                        return f"Total: {total} (Disaggregated: {value['mode']})"
                    elif isinstance(value, str) and value.startswith('{'):
                        with suppress(Exception):
                            import json
                            parsed = json.loads(value)
                            if isinstance(parsed, dict) and 'values' in parsed:
                                total = sum(v for v in parsed['values'].values() if isinstance(v, (int, float)))
                                return f"Total: {total} (Disaggregated: {parsed['mode']})"
                    return str(value)

                # Log field changes
                for change in field_changes_tracker:
                    if change.get('type') in ['added', 'updated']:
                        old_val = parse_field_value_for_display(change.get('old_value'))
                        new_val = parse_field_value_for_display(change.get('new_value'))
                        current_app.logger.info(f"Field '{change.get('field_name', 'Unknown')}' {change['type']}: {old_val} -> {new_val}")

                flash("Form data saved successfully.", "success")
                flash("Public submission updated successfully.", "success")

                if action == 'save':
                    return redirect(url_for("forms.view_public_submission", submission_id=submission_id))

            except Exception as e:
                request_transaction_rollback()
                flash("An error occurred. Please try again.", "danger")
                current_app.logger.error(f"Error saving public submission {submission_id}: {e}", exc_info=True)
        else:
            flash("Form submission failed due to a security issue. Please try again.", "danger")

    # Load existing data for repeat groups (same structure as assignment forms)
    repeat_groups_data = {}
    existing_data_processed = _load_existing_data_for_public_submission(submission)

    # Add repeat groups data to existing_data_processed for template access
    existing_data_processed['repeat_groups_data'] = repeat_groups_data

    existing_submitted_documents = _prepare_submitted_documents_for_template(submission)

    # Create section statuses
    section_statuses = {}
    for section in sections:
        if hasattr(section, 'fields_ordered') and section.fields_ordered:
            section_statuses[section.name] = 'Completed'
        else:
            section_statuses[section.name] = 'N/A'

    # available_indicators_by_section is already prepared above for dynamic sections

    # Create CSRF form for template compatibility
    csrf_form = FlaskForm()

    # Import WTForms components for dummy forms
    from wtforms import SelectField, StringField, EmailField, SubmitField
    from wtforms.validators import DataRequired, Email

    # Create dummy forms for template compatibility (not actually used in admin view)
    class DummyCountrySelectForm(FlaskForm):
        country_id = SelectField("Select Your Country", coerce=int, validators=[DataRequired()])

    class DummySubmissionDetailsForm(FlaskForm):
        submitter_name = StringField("Your Name", validators=[DataRequired()])
        submitter_email = EmailField("Your Email", validators=[DataRequired(), Email()])
        submit = SubmitField("Submit Form")

    country_select_form = DummyCountrySelectForm()
    # Use the same country list as new public form submissions (countries with public access enabled)
    sorted_countries = sorted(submission.assigned_form.public_countries, key=lambda c: c.name)
    country_choices = [(c.id, c.name) for c in sorted_countries]
    country_select_form.country_id.choices = country_choices
    country_select_form.country_id.data = submission.country.id

    submission_details_form = DummySubmissionDetailsForm()
    submission_details_form.submitter_name.data = submission.submitter_name
    submission_details_form.submitter_email.data = submission.submitter_email

    # Prepare available indicators data for dynamic sections (same as assignment forms)
    available_indicators_by_section = {}
    for section in all_sections:
        if section.section_type == 'dynamic_indicators':
            # Add display filter configuration for dynamic sections
            section.data_entry_display_filters_config = getattr(section, 'data_entry_display_filters_list', [])

            # For public submissions, we don't allow adding new indicators, so keep empty
            available_indicators_by_section[section.id] = []
        else:
            available_indicators_by_section[section.id] = []

    # Apply page translations to all pages in the PUBLISHED version
    for page in FormPage.query.filter_by(template_id=form_template.id, version_id=form_template.published_version_id).order_by(FormPage.order).all():
        page.display_name = get_localized_page_name(page)

    # Apply page translations to all page objects referenced by sections
    page_ids_processed = set()
    for section in all_sections:
        if section.page and section.page.id not in page_ids_processed:
            section.page.display_name = get_localized_page_name(section.page)
            page_ids_processed.add(section.page.id)

    # Apply section translations to all sections (ISO codes only)
    current_locale_short = (str(get_locale()) if get_locale() else 'en').split('_', 1)[0]
    for section in all_sections:
        translated_name = None
        if section.name_translations and isinstance(section.name_translations, dict):
            translated_name = section.name_translations.get(current_locale_short) or section.name_translations.get('en')
        section.display_name = translated_name.strip() if isinstance(translated_name, str) and translated_name.strip() else section.name

    # Ensure translation_key is defined for template use
    translation_key = get_translation_key()

    # Create template structure object that includes ALL sections (main and sub-sections) like assignment forms
    template_structure = form_template
    template_structure.sections = all_sections  # Pass all sections so template can organize them hierarchically

    # Determine page title based on mode
    if can_edit:
        title = f"Edit Submission: {submission.country.name} - {submission.submitted_at.strftime('%Y-%m-%d %H:%M')}"
    else:
        title = f"View Submission: {submission.country.name} - {submission.submitted_at.strftime('%Y-%m-%d %H:%M')}"

    return render_template("forms/entry_form/entry_form.html",
                         title=title,
                         assignment=submission.assigned_form,  # Use assigned_form
                         assignment_status=dummy_acs,  # AssignmentEntityStatus object for all entity types
                         template_structure=template_structure,  # Use proper template_structure
                         form=csrf_form,  # Add CSRF form
                         existing_data=existing_data_processed,  # Use the correct variable name
                         existing_submitted_documents=existing_submitted_documents,
                         section_statuses=section_statuses,  # Add section_statuses
                         slugify_age_group=slugify_age_group,
                         config=Config,  # Use Config class
                         can_edit=can_edit,  # Set edit capability
                         is_preview_mode=False,  # Public submissions are never in preview mode
                         QuestionType=QuestionType,
                         isinstance=isinstance,  # Add missing Python builtin
                         json=json,  # Add missing JSON module
                         hasattr=hasattr,  # Add missing Python builtin
                         available_indicators_by_section=available_indicators_by_section,
                         get_localized_country_name=get_localized_country_name,
                         translation_key=translation_key,  # Add translation key for options translation
                         # Add localization functions for template use
                         get_localized_indicator_definition=get_localized_indicator_definition,
                         get_localized_indicator_type=get_localized_indicator_type,
                         get_localized_indicator_unit=get_localized_indicator_unit,
                         get_localized_template_name=get_localized_template_name,
                         # Public submission specific variables
                         submission=submission,
                         is_public_submission=True,
                         country_select_form=country_select_form,  # Add country_select_form for template compatibility
                         submission_details_form=submission_details_form,  # Fixed: use actual submission_details_form
                         # Add plugin manager information for template use
                         plugin_manager=current_app.plugin_manager if hasattr(current_app, 'plugin_manager') else None,
                         form_integration=current_app.form_integration if hasattr(current_app, 'form_integration') else None)

@bp.route("/download_document/<int:submitted_document_id>", methods=["GET"])
@login_required
def download_document(submitted_document_id):
    from app.services.document_service import DocumentService
    try:
        directory, filename, download_name = DocumentService.get_assignment_download_paths(submitted_document_id, current_user)
        return send_from_directory(directory, filename, as_attachment=True, download_name=download_name)
    except PermissionError as e:
        flash("An error occurred. Please try again.", "warning")
        return redirect(url_for("main.dashboard"))
    except FileNotFoundError:
        current_app.logger.error(f"Attempted to download non-existent file for ID {submitted_document_id}")
        abort(404)
    except Exception as e:
        current_app.logger.error(f"Error serving document {submitted_document_id}: {e}", exc_info=True)
        flash("An error occurred while trying to download the file.", "danger")
        return redirect(url_for("main.dashboard"))

@bp.route("/delete_document/<int:submitted_document_id>", methods=["POST"])
@login_required
def delete_document(submitted_document_id):
    from app.services.document_service import DocumentService
    from app.utils.redirect_utils import is_safe_redirect_url
    csrf_form = FlaskForm()
    # SECURITY: Validate referrer to prevent open redirect attacks
    referrer = request.referrer
    safe_referrer = referrer if referrer and is_safe_redirect_url(referrer) else None
    if not csrf_form.validate_on_submit():
        flash("Document deletion failed due to a security issue. Please try again.", "danger")
        return redirect(safe_referrer or url_for("main.dashboard"))
    try:
        deleted_name = DocumentService.delete_assignment_document(submitted_document_id, current_user)
        flash(f"Document '{deleted_name}' deleted successfully.", "success")
    except PermissionError as e:
        flash("An error occurred. Please try again.", "warning")
    except Exception as e:
        current_app.logger.error(f"Error deleting document {submitted_document_id}: {e}", exc_info=True)
        flash("Error deleting document.", "danger")
    return redirect(safe_referrer or url_for("main.dashboard"))


@bp.route("/assignment_status/<int:aes_id>/export_pdf", methods=["GET"])
@login_required
def export_assignment_pdf(aes_id):
    """Generate a high-quality PDF for an assignment using a print-optimized HTML template.

    Uses WeasyPrint (HTML to PDF) for faithful rendering with proper pagination and styles.
    """
    try:
        assignment_entity_status = AssignmentEntityStatus.query.options(
            db.joinedload(AssignmentEntityStatus.assigned_form).joinedload(AssignedForm.template)
        ).get_or_404(aes_id)

        from app.services.authorization_service import AuthorizationService
        if not AuthorizationService.can_access_assignment(assignment_entity_status, current_user):
            flash("You are not authorized to export data for this assignment and country.", "warning")
            return redirect(url_for("main.dashboard"))

        assignment = assignment_entity_status.assigned_form
        country = assignment_entity_status.country
        form_template_for_export = assignment.template

        # Resolve template variables for PDF export
        from app.services.variable_resolution_service import VariableResolutionService
        from app.models import FormTemplateVersion

        template_version = None
        resolved_variables = {}
        variable_configs = {}
        if form_template_for_export.published_version_id:
            template_version = FormTemplateVersion.query.get(form_template_for_export.published_version_id)
            if template_version:
                variable_configs = template_version.variables or {}
                resolved_variables = VariableResolutionService.resolve_variables(
                    template_version,
                    assignment_entity_status
                )

        # Resolve EO1/EO2/EO3 from emergency_operations plugin for PDF (section names use [EO1] etc.)
        if country:
            country_iso = (getattr(country, 'iso3', None) or getattr(country, 'iso2', None) or '').strip().upper()
            if country_iso:
                try:
                    from plugins.emergency_operations.routes import get_emergency_operations_data
                    operations = get_emergency_operations_data(country_iso=country_iso)
                    for i, key in enumerate(('EO1', 'EO2', 'EO3')):
                        if i < len(operations):
                            op = operations[i]
                            name = (op.get('name') or '').strip()
                            code = (op.get('code') or '').strip()
                            resolved_variables[key] = f"{name} ({code})" if code else (name or '')
                        else:
                            resolved_variables[key] = ''
                except Exception as e:
                    current_app.logger.debug(
                        f"Could not resolve EO1/EO2/EO3 for PDF export (plugin or API): {e}"
                    )

        # Resolve localized names for the PDF header (template + country)
        translation_key = get_translation_key()
        assignment_display_name = None
        with suppress(Exception):
            assignment_display_name = get_localized_template_name(
                form_template_for_export,
                locale=translation_key,
                version=template_version,
            )
        country_display_name = None
        with suppress(Exception):
            country_display_name = get_localized_country_name(country) if country else None

        # Build a lightweight structure for PDF export that preserves sub-section nesting.
        # The entry form renders `FormSection.sub_sections` inside the parent section; PDF should match.
        sections_by_page = {}
        default_page_id = 0

        section_nodes_by_id = {}
        ordered_section_ids = []

        for section_model in form_template_for_export.sections.order_by(FormSection.order).all():
            # Localize section name then resolve variables (match entry form behavior)
            section_display_name = None
            with suppress(Exception):
                section_display_name = get_localized_section_name(section_model)
            if not section_display_name:
                section_display_name = getattr(section_model, 'display_name', None) or section_model.name
            if resolved_variables and section_display_name:
                try:
                    section_display_name = VariableResolutionService.replace_variables_in_text(
                        section_display_name,
                        resolved_variables,
                        variable_configs
                    )
                except Exception as e:
                    current_app.logger.warning(
                        f"Error resolving variables in section name for section {section_model.id}: {e}",
                        exc_info=True
                    )

            section_data_for_export = {
                'name': section_model.name,
                'display_name': section_display_name,
                'id': section_model.id,
                'order': section_model.order,
                'page_id': section_model.page_id,
                'parent_section_id': section_model.parent_section_id,
                # Used for conditional visibility (relevance rules) during PDF export
                'relevance_condition': getattr(section_model, 'relevance_condition', None),
                'subsections': [],
                'fields_ordered': []
            }

            temp_fields = []
            form_items = FormItem.query.filter_by(section_id=section_model.id, archived=False).order_by(FormItem.order).all()
            if form_items:
                for form_item in form_items:
                    # Localize item label then resolve variables (match entry form behavior)
                    display_label = None
                    with suppress(Exception):
                        lt = getattr(form_item, 'label_translations', None)
                        if isinstance(lt, dict) and lt:
                            candidate = lt.get(translation_key) or lt.get('en')
                            if isinstance(candidate, str) and candidate.strip():
                                display_label = candidate.strip()
                    if not display_label:
                        display_label = getattr(form_item, 'display_label', None) or form_item.label
                    if resolved_variables and display_label:
                        try:
                            display_label = VariableResolutionService.replace_variables_in_text(
                                display_label,
                                resolved_variables,
                                variable_configs
                            )
                        except Exception as e:
                            current_app.logger.warning(
                                f"Error resolving variables in display_label for form_item {form_item.id}: {e}",
                                exc_info=True
                            )

                    base = {
                        'id': form_item.id,
                        'order': form_item.order,
                        'label': form_item.label,
                        'display_label': display_label,
                        'unit': getattr(form_item, 'unit', None),
                        'type': getattr(form_item, 'type', None),
                        # Used for conditional visibility (relevance rules) during PDF export
                        'conditions': getattr(form_item, 'conditions', None),
                    }
                    if form_item.is_indicator:
                        base.update({'kind': 'indicator', 'model': form_item})
                        temp_fields.append(base)
                    elif form_item.is_question:
                        is_blank_note = (
                            getattr(form_item, 'type', None) == 'blank'
                            or (getattr(form_item, 'question_type', None) and getattr(form_item.question_type, 'value', None) == 'blank')
                        )
                        if is_blank_note:
                            base.update({'kind': 'note', 'model': form_item})
                        else:
                            base.update({'kind': 'question', 'model': form_item})
                        temp_fields.append(base)
                    elif getattr(form_item, 'item_type', None) == 'matrix' or getattr(form_item, 'is_matrix', False):
                        # Matrix fields are stored/rendered differently (data lives in disagg_data)
                        matrix_config = {}
                        try:
                            if isinstance(getattr(form_item, 'config', None), dict):
                                matrix_config = form_item.config.get('matrix_config') or form_item.config or {}
                        except Exception as e:
                            current_app.logger.debug("matrix_config parse failed: %s", e)
                            matrix_config = {}

                        # Prefer resolved matrix rows (if variable substitution was applied earlier),
                        # otherwise use configured rows (manual-mode matrices).
                        matrix_rows = getattr(form_item, '_display_matrix_rows', None)
                        if not matrix_rows and isinstance(matrix_config, dict):
                            matrix_rows = matrix_config.get('rows', []) or []

                        # Resolve variables in manual-mode matrix row labels (e.g. [period], [[period]+1]).
                        # This is critical because saved matrix cell keys are built from the rendered row labels.
                        try:
                            if isinstance(matrix_config, dict):
                                row_mode = matrix_config.get('row_mode', 'manual')
                                if row_mode == 'manual' or not row_mode:
                                    if resolved_variables and matrix_rows and isinstance(matrix_rows, list):
                                        resolved_rows = []
                                        for r in matrix_rows:
                                            if isinstance(r, str):
                                                resolved_rows.append(
                                                    VariableResolutionService.replace_variables_in_text(
                                                        r,
                                                        resolved_variables,
                                                        variable_configs
                                                    )
                                                )
                                            else:
                                                resolved_rows.append(r)
                                        matrix_rows = resolved_rows
                        except Exception as e:
                            current_app.logger.warning(
                                f"Error resolving variables in matrix row labels for form_item {form_item.id}: {e}",
                                exc_info=True
                            )
                        matrix_columns = matrix_config.get('columns', []) if isinstance(matrix_config, dict) else []

                        base.update({
                            'kind': 'matrix',
                            'model': form_item,
                            'matrix_config': matrix_config,
                            'matrix_rows': matrix_rows,
                            'matrix_columns': matrix_columns,
                        })
                        temp_fields.append(base)
                    elif form_item.is_document_field:
                        base.update({'kind': 'document', 'model': form_item})
                        temp_fields.append(base)

            # Include dynamic indicators for dynamic_indicators sections (including subsections)
            section_type = getattr(section_model, 'section_type', None) or 'standard'
            if section_type == 'dynamic_indicators':
                dynamic_assignments = DynamicIndicatorData.query.filter_by(
                    assignment_entity_status_id=assignment_entity_status.id,
                    section_id=section_model.id,
                ).order_by(DynamicIndicatorData.order).all()
                for dyn in dynamic_assignments:
                    display_label = dyn.custom_label
                    if not (display_label and str(display_label).strip()):
                        with suppress(Exception):
                            display_label = get_localized_indicator_name(dyn.indicator_bank)
                    if not display_label:
                        display_label = getattr(dyn.indicator_bank, 'name', '') or ''
                    if resolved_variables and display_label:
                        try:
                            display_label = VariableResolutionService.replace_variables_in_text(
                                display_label, resolved_variables, variable_configs
                            )
                        except Exception as e:
                            current_app.logger.debug("replace_variables for display_label failed: %s", e)
                    temp_fields.append({
                        'id': f'dynamic_{dyn.id}',
                        'order': dyn.order,
                        'label': display_label,
                        'display_label': display_label,
                        'unit': getattr(dyn.indicator_bank, 'unit', None),
                        'type': getattr(dyn.indicator_bank, 'type', None),
                        'conditions': None,
                        'kind': 'indicator',
                        'model': None,
                    })

            temp_fields.sort(key=lambda x: (x.get('order') is None, x.get('order')))
            section_data_for_export['fields_ordered'] = temp_fields

            section_nodes_by_id[section_model.id] = section_data_for_export
            ordered_section_ids.append(section_model.id)

        # Attach children to parents in template order
        for section_id in ordered_section_ids:
            node = section_nodes_by_id.get(section_id)
            if not node:
                continue
            parent_id = node.get('parent_section_id')
            if parent_id and parent_id in section_nodes_by_id:
                section_nodes_by_id[parent_id]['subsections'].append(node)

        # Only top-level sections go into sections_by_page; sub-sections are rendered recursively
        for section_id in ordered_section_ids:
            node = section_nodes_by_id.get(section_id)
            if not node or node.get('parent_section_id') is not None:
                continue
            page_id = node.get('page_id') if node.get('page_id') is not None else default_page_id
            if page_id not in sections_by_page:
                sections_by_page[page_id] = []
            sections_by_page[page_id].append(node)

        # Existing data mapping by form_item id
        existing_entries = FormData.query.filter_by(
            assignment_entity_status_id=assignment_entity_status.id
        ).all()
        existing_data_processed_for_export = {}
        for entry in existing_entries:
            if entry.form_item_id:
                # Prefer disagg_data when present (covers disaggregated indicators + matrices).
                value = entry.disagg_data if entry.disagg_data is not None else entry.value
                if isinstance(value, str):
                    v = value.strip()
                    if (v.startswith('{') and v.endswith('}')) or (v.startswith('[') and v.endswith(']')):
                        with suppress(Exception):
                            value = json.loads(v)
                existing_data_processed_for_export[f"form_item_{entry.form_item_id}"] = value

        # Add dynamic indicator values (keyed by form_item_dynamic_<id> to match fields_ordered)
        dynamic_entries = DynamicIndicatorData.query.filter_by(
            assignment_entity_status_id=assignment_entity_status.id
        ).all()
        for dyn in dynamic_entries:
            key = f"form_item_dynamic_{dyn.id}"
            if dyn.disagg_data is not None:
                existing_data_processed_for_export[key] = dyn.disagg_data
            else:
                existing_data_processed_for_export[key] = {'values': {'total': dyn.value}}

        # Optional: allow the client (entry form UI) to tell the server which sections/fields are currently hidden
        # due to relevance rules, so PDF matches the user's current view (including unsaved UI state).
        def _parse_hidden_ids_arg(arg_name):
            raw = (request.args.get(arg_name) or '').strip()
            if not raw:
                return set()
            out = set()
            for part in raw.split(','):
                part = (part or '').strip()
                if not part:
                    continue
                if part.isdigit():
                    try:
                        out.add(int(part))
                    except (ValueError, TypeError):
                        continue
            return out

        hidden_section_ids_from_client = _parse_hidden_ids_arg('hidden_sections')
        hidden_field_ids_from_client = _parse_hidden_ids_arg('hidden_fields')

        # Filter export structure using the client-provided hidden lists (UI state).
        # This keeps the relevance logic in one place (client-side `conditions.js`) while allowing
        # the server-side PDF export to match what the user is currently seeing.
        def _filter_section_node(section_node):
            if not isinstance(section_node, dict):
                return None

            try:
                if section_node.get('id') in hidden_section_ids_from_client:
                    return None
            except Exception as e:
                current_app.logger.debug("hidden section filter failed: %s", e)

            # Drop hidden fields within this section
            kept_fields = []
            for f in (section_node.get('fields_ordered') or []):
                if not isinstance(f, dict):
                    continue
                try:
                    if f.get('id') in hidden_field_ids_from_client:
                        continue
                except Exception as e:
                    current_app.logger.debug("hidden field filter failed: %s", e)
                kept_fields.append(f)
            section_node['fields_ordered'] = kept_fields

            # Recurse into subsections
            kept_children = []
            for child in (section_node.get('subsections') or []):
                kept = _filter_section_node(child)
                if kept is not None:
                    kept_children.append(kept)
            section_node['subsections'] = kept_children
            return section_node

        filtered_sections_by_page = {}
        for page_id, root_sections in (sections_by_page or {}).items():
            kept_roots = []
            for sec in (root_sections or []):
                kept = _filter_section_node(sec)
                if kept is not None:
                    kept_roots.append(kept)
            filtered_sections_by_page[page_id] = kept_roots
        sections_by_page = filtered_sections_by_page

        # For list-library (advanced) matrices, rows are dynamic and not stored in config.rows.
        # Infer row IDs from saved keys (e.g. "3_SP2") and resolve display labels so PDF can render a real table.
        def _infer_list_library_rows_and_labels(field_dict, matrix_data):
            try:
                if not isinstance(field_dict, dict):
                    return
                if field_dict.get('kind') != 'matrix':
                    return
                if not isinstance(matrix_data, dict) or not matrix_data:
                    return

                matrix_config = field_dict.get('matrix_config') if isinstance(field_dict.get('matrix_config'), dict) else {}
                row_mode = (matrix_config.get('row_mode') or '').strip().lower()
                if row_mode != 'list_library':
                    return

                # Get column names from config
                cols = field_dict.get('matrix_columns') or []
                col_names = []
                for c in cols:
                    if isinstance(c, dict):
                        col_names.append(str(c.get('name') if c.get('name') else c))
                    else:
                        col_names.append(str(c))
                col_names = [c for c in col_names if c and c != 'None']
                if not col_names:
                    return

                # Determine row ids by matching "{rowId}_{colName}" keys
                col_names_sorted = sorted(col_names, key=len, reverse=True)
                row_ids = []
                seen = set()
                for k in matrix_data.keys():
                    if not isinstance(k, str):
                        continue
                    if k.startswith('_'):
                        continue
                    matched_row_id = None
                    for cn in col_names_sorted:
                        suffix = "_" + cn
                        if k.endswith(suffix):
                            matched_row_id = k[: -len(suffix)]
                            break
                    if matched_row_id and matched_row_id not in seen:
                        seen.add(matched_row_id)
                        row_ids.append(matched_row_id)

                if not row_ids:
                    return

                # Resolve row labels based on lookup list id (regular lists) or system lists.
                lookup_list_id = (matrix_config.get('lookup_list_id') or '').strip()

                display_column = (matrix_config.get('display_column') or matrix_config.get('list_display_column') or 'name').strip() or 'name'

                row_labels = {}
                # Regular lookup list rows
                if lookup_list_id and str(lookup_list_id).isdigit():
                    from app.models import LookupListRow
                    for rid in row_ids:
                        try:
                            rid_int = int(rid)
                        except (ValueError, TypeError):
                            row_labels[rid] = rid
                            continue
                        row_obj = LookupListRow.query.get(rid_int)
                        if row_obj and isinstance(row_obj.data, dict):
                            row_labels[rid] = str(row_obj.data.get(display_column) or row_obj.data.get('name') or rid)
                        else:
                            row_labels[rid] = rid
                elif lookup_list_id in ('country_map', 'national_society', 'indicator_bank'):
                    current_locale = session.get('language', 'en') or (str(get_locale()) if get_locale() else 'en')
                    if isinstance(current_locale, str) and '_' in current_locale:
                        current_locale = current_locale.split('_', 1)[0]

                    if lookup_list_id == 'country_map':
                        for rid in row_ids:
                            try:
                                rid_int = int(rid)
                            except (ValueError, TypeError):
                                row_labels[rid] = rid
                                continue
                            obj = Country.query.get(rid_int)
                            row_labels[rid] = get_localized_country_name(obj) if obj else rid
                    elif lookup_list_id == 'national_society':
                        from app.models.organization import NationalSociety
                        for rid in row_ids:
                            try:
                                rid_int = int(rid)
                            except (ValueError, TypeError):
                                row_labels[rid] = rid
                                continue
                            obj = NationalSociety.query.get(rid_int)
                            if obj:
                                localized_name = obj.get_name_translation(current_locale) if hasattr(obj, 'get_name_translation') else None
                                row_labels[rid] = (localized_name.strip() if isinstance(localized_name, str) and localized_name.strip() else obj.name)
                            else:
                                row_labels[rid] = rid
                    else:
                        from app.models.indicator_bank import IndicatorBank
                        for rid in row_ids:
                            try:
                                rid_int = int(rid)
                            except (ValueError, TypeError):
                                row_labels[rid] = rid
                                continue
                            obj = IndicatorBank.query.get(rid_int)
                            row_labels[rid] = obj.name if obj else rid
                else:
                    # Unknown/system/plugin list: fallback to showing IDs
                    row_labels = {rid: rid for rid in row_ids}

                field_dict['matrix_rows'] = row_ids
                field_dict['matrix_row_labels'] = row_labels
            except Exception as e:
                current_app.logger.warning(f"Failed to infer list-library matrix rows for PDF: {e}", exc_info=True)

        def _walk_sections_for_export(section_node):
            if not isinstance(section_node, dict):
                return
            fields = section_node.get('fields_ordered') or []
            if isinstance(fields, list):
                for f in fields:
                    if isinstance(f, dict) and f.get('kind') == 'matrix':
                        item_key = f"form_item_{f.get('id')}"
                        _infer_list_library_rows_and_labels(f, existing_data_processed_for_export.get(item_key))
            for child in section_node.get('subsections', []) or []:
                _walk_sections_for_export(child)

        for page_id, root_sections in (sections_by_page or {}).items():
            for sec in root_sections or []:
                _walk_sections_for_export(sec)

        # Pages list (preserve template page order; if none, create a single default page)
        pages = list(form_template_for_export.pages) if form_template_for_export.is_paginated else [None]

        # Render print-optimized HTML
        # wide_matrix_strategy: 'scale' (shrink table to fit) or 'landscape' (use landscape page for wide matrices)
        html_content = render_template(
            'forms/entry_form/export_pdf.html',
            assignment=assignment,
            assignment_display_name=assignment_display_name,
            country=country,
            country_display_name=country_display_name,
            aes=assignment_entity_status,
            form_template=form_template_for_export,
            sections_by_page=sections_by_page,
            pages=pages,
            existing_data=existing_data_processed_for_export,
            generated_at=utcnow(),
            get_localized_page_name=get_localized_page_name,
            wide_matrix_strategy='scale',
        )

        # Lazy import WeasyPrint to avoid hard dependency unless used
        try:
            from weasyprint import HTML, CSS  # type: ignore
        except Exception as e:
            current_app.logger.error(f"WeasyPrint not available: {e}", exc_info=True)
            return current_app.response_class(
                response="PDF generation is not available on this deployment.",
                status=503,
                mimetype='text/plain'
            )

        # Base URL for static assets
        static_dir = os.path.join(current_app.root_path, 'static')

        # Enhanced print CSS with boxed field styling
        # NOTE: Keep this as a plain string (not an f-string) to avoid brace escaping issues.
        # We localize small UI labels (like "Page") via post-processing below.
        pdf_css_string = '''
            @page {
                size: A4;
                margin: 20mm 15mm 20mm 15mm;
                @bottom-right { content: "Page " counter(page); font-size: 10pt; color: #6b7280; }
            }
            body {
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, Arial, sans-serif;
                color: #111827;
                line-height: 1.5;
            }
            h1, h2, h3, h4 { color: #111827; margin: 0 0 8px 0; }
            h1 { font-size: 20pt; }
            h2 { font-size: 14pt; border-bottom: 2px solid #cc0000; padding-bottom: 4px; margin-top: 16px; margin-bottom: 12px; }
            .form-page-title { page-break-after: avoid; }
            h3 { font-size: 12pt; margin-top: 10px; margin-bottom: 8px; color: #374151; }
            h4 { font-size: 11pt; margin-top: 8px; margin-bottom: 6px; color: #374151; }
            .meta {
                color: #374151;
                font-size: 10pt;
                margin-bottom: 16px;
                padding: 8px;
                background: #f9fafb;
                border-left: 3px solid #cc0000;
            }
            .meta div { margin: 4px 0; }
            /* Allow sections to break across pages; only avoid breaking inside a single field */
            .section {
                margin-bottom: 16px;
            }
            .section h2, .section h3, .section h4 {
                page-break-after: avoid;
            }
            .subsection {
                margin-left: 12px;
                padding-left: 10px;
                border-left: 2px solid #e5e7eb;
            }
            .section-empty-note {
                color: #6b7280;
                font-size: 10pt;
                font-style: italic;
                margin: 8px 0 0 0;
            }
            .form-note {
                color: #374151;
                font-size: 10pt;
                margin: 8px 0;
                padding: 8px 12px;
                background: #f9fafb;
                border-left: 3px solid #9ca3af;
                border-radius: 0 4px 4px 0;
            }

            /* Field box styling */
            .field-box {
                border: 1.5px solid #e5e7eb;
                border-radius: 4px;
                margin: 8px 0;
                page-break-inside: avoid;
                background: #ffffff;
                box-shadow: 0 1px 2px rgba(0, 0, 0, 0.05);
            }
            /* Matrix tables can span many rows: allow breaking so all rows appear in PDF */
            .field-box-matrix {
                page-break-inside: auto;
            }
            .field-box-matrix .table thead {
                display: table-header-group;
            }
            .field-box-matrix .table tr {
                page-break-inside: avoid;
            }

            .field-filled {
                border-left: 4px solid #10b981;
            }

            .field-empty {
                border-left: 4px solid #d1d5db;
                background: #f9fafb;
            }

            .field-empty-required {
                border-left: 4px solid #ef4444;
                background: #fef2f2;
            }

            .field-empty-optional {
                border-left: 4px solid #d1d5db;
                background: #f9fafb;
            }

            .field-header {
                background: #f9fafb;
                padding: 8px 12px;
                border-bottom: 1px solid #e5e7eb;
                font-weight: 600;
            }

            .field-label {
                color: #111827;
                font-size: 11pt;
                display: block;
            }

            .field-unit {
                color: #6b7280;
                font-size: 9pt;
                font-weight: normal;
                font-style: italic;
            }

            .field-content {
                padding: 10px 12px;
                min-height: 20px;
            }

            .field-value {
                color: #111827;
                font-size: 10pt;
                word-wrap: break-word;
                display: block;
            }

            .disaggregation-caption {
                font-size: 9pt;
                font-weight: 600;
                color: #374151;
                margin: 6px 0 4px 0;
                display: block;
            }

            .not-reported {
                color: #dc2626;
                font-size: 10pt;
                font-weight: 600;
                font-style: italic;
                display: block;
            }

            .not-reported-optional {
                color: #6b7280;
                font-size: 10pt;
                font-weight: 600;
                font-style: italic;
                display: block;
            }

            /* RTL support (e.g. Arabic) - driven by <html dir="rtl"> */
            html[dir="rtl"] body {
                direction: rtl;
                font-family: "Tajawal", -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, Arial, sans-serif;
            }
            html[dir="rtl"] h1,
            html[dir="rtl"] h2,
            html[dir="rtl"] h3,
            html[dir="rtl"] h4,
            html[dir="rtl"] .meta,
            html[dir="rtl"] .field-header,
            html[dir="rtl"] .field-content,
            html[dir="rtl"] .field-value,
            html[dir="rtl"] .disaggregation-caption,
            html[dir="rtl"] .not-reported,
            html[dir="rtl"] .not-reported-optional {
                text-align: right;
                font-family: "Tajawal", -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, Arial, sans-serif;
            }
            html[dir="rtl"] .meta { border-left: none; border-right: 3px solid #cc0000; }
            html[dir="rtl"] .form-note { border-left: none; border-right: 3px solid #9ca3af; border-radius: 4px 0 0 4px; }
            html[dir="rtl"] .subsection { margin-left: 0; margin-right: 12px; padding-left: 0; padding-right: 10px; border-left: none; border-right: 2px solid #e5e7eb; }
            html[dir="rtl"] table th, html[dir="rtl"] table td {
                text-align: right;
                font-family: "Tajawal", -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Inter, Arial, sans-serif;
            }

            .document-note {
                color: #6b7280;
                font-style: italic;
            }

            .table {
                width: 100%;
                border-collapse: collapse;
                margin: 8px 0;
                font-size: 9pt;
            }
            .table th, .table td {
                border: 1px solid #d1d5db;
                padding: 6px 8px;
            }
            .table th {
                background: #f3f4f6;
                text-align: left;
                font-weight: 600;
                color: #374151;
            }
            .table td {
                background: #ffffff;
            }
            .table tbody tr:nth-child(even) td {
                background: #f9fafb;
            }
            .table td.cell-tick {
                text-align: center;
            }
            /* Wide matrix: scale to fit (default) or use landscape page */
            .wide-matrix-scale {
                width: 100%;
                overflow: hidden;
            }
            .wide-matrix-scale table.table {
                transform: scale(0.72);
                transform-origin: top left;
            }
            @page wide {
                size: A4 landscape;
                margin: 20mm 15mm 20mm 15mm;
                @bottom-right { content: "Page " counter(page); font-size: 10pt; color: #6b7280; }
            }
            .wide-matrix-landscape {
                page: wide;
            }
            .page-break { page-break-before: always; }
        '''
        with suppress(Exception):
            pdf_css_string = pdf_css_string.replace('content: "Page "', f'content: "{_("Page")} "')
        pdf_css = CSS(string=pdf_css_string)

        pdf_buffer = io.BytesIO()
        HTML(string=html_content, base_url=static_dir).write_pdf(
            pdf_buffer,
            stylesheets=[pdf_css],
            optimize_size=('fonts', 'images')
        )

        pdf_buffer.seek(0)
        filename = f"assignment_{country.iso3 if country else 'country'}_{str(assignment.period_name).replace(' ', '_')}.pdf"
        return send_file(
            pdf_buffer,
            download_name=filename,
            as_attachment=True,
            mimetype='application/pdf'
        )
    except Exception as e:
        current_app.logger.error(f"Error generating PDF for ACS {aes_id}: {e}", exc_info=True)
        flash("Failed to generate PDF.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))


@bp.route("/assignment_status/<int:aes_id>/export_excel", methods=["GET"])
@login_required
def export_focal_data_excel(aes_id):
    assignment_entity_status = AssignmentEntityStatus.query.options(
        db.joinedload(AssignmentEntityStatus.assigned_form).joinedload(AssignedForm.template)
    ).get_or_404(aes_id)

    from app.services.authorization_service import AuthorizationService
    if not AuthorizationService.can_access_assignment(assignment_entity_status, current_user):
         flash("You are not authorized to export data for this assignment and country.", "warning")
         return redirect(url_for("main.dashboard"))

    assignment = assignment_entity_status.assigned_form
    country = assignment_entity_status.country
    form_template_for_export = assignment.template

    # Resolve template variables for Excel export
    from app.services.variable_resolution_service import VariableResolutionService
    from app.models import FormTemplateVersion

    template_version = None
    resolved_variables = {}
    variable_configs = {}
    if form_template_for_export.published_version_id:
        template_version = FormTemplateVersion.query.get(form_template_for_export.published_version_id)
        if template_version:
            variable_configs = template_version.variables or {}
            resolved_variables = VariableResolutionService.resolve_variables(
                template_version,
                assignment_entity_status
            )

    # Group sections by page
    sections_by_page = {}
    default_page_id = 0  # For sections without a page

    for section_model in form_template_for_export.sections.order_by(FormSection.order).all():
        page_id = section_model.page_id if section_model.page_id is not None else default_page_id
        if page_id not in sections_by_page:
            sections_by_page[page_id] = []

        # Resolve variables in section display_name
        section_display_name = getattr(section_model, 'display_name', None) or section_model.name
        if resolved_variables and section_display_name:
            try:
                section_display_name = VariableResolutionService.replace_variables_in_text(
                    section_display_name,
                    resolved_variables,
                    variable_configs
                )
            except Exception as e:
                current_app.logger.warning(
                    f"Error resolving variables in section name for section {section_model.id}: {e}",
                    exc_info=True
                )

        section_data_for_export = {'name': section_display_name, 'id': section_model.id, 'fields_ordered': []}
        temp_fields = []

        # Process form items (exclude archived items)
        form_items = FormItem.query.filter_by(section_id=section_model.id, archived=False).order_by(FormItem.order).all()
        if form_items:
            for form_item in form_items:
                # Resolve variables in display_label
                display_label = getattr(form_item, 'display_label', None) or form_item.label
                if resolved_variables and display_label:
                    try:
                        display_label = VariableResolutionService.replace_variables_in_text(
                            display_label,
                            resolved_variables,
                            variable_configs
                        )
                    except Exception as e:
                        current_app.logger.warning(
                            f"Error resolving variables in display_label for form_item {form_item.id}: {e}",
                            exc_info=True
                        )

                if form_item.is_indicator:
                    temp_fields.append({
                        'id': form_item.id, 'legacy_id': None, 'label': display_label,
                        'type': form_item.type, 'unit': form_item.unit, 'order': form_item.order,
                        'is_indicator': True, 'is_form_item': True, 'item_model': form_item
                    })
                elif form_item.is_question:
                    temp_fields.append({
                        'id': form_item.id, 'legacy_id': None, 'label': display_label,
                        'type': form_item.type, 'order': form_item.order,
                        'is_question': True, 'is_form_item': True, 'item_model': form_item
                    })
                elif form_item.is_document_field:
                    temp_fields.append({
                        'id': form_item.id, 'legacy_id': None, 'label': display_label,
                        'type': 'DOCUMENT', 'order': form_item.order, 'is_required': form_item.is_required,
                        'description': form_item.description, 'is_document': True, 'is_form_item': True, 'item_model': form_item
                    })

        temp_fields.sort(key=lambda x: x['order'])
        section_data_for_export['fields_ordered'] = temp_fields
        sections_by_page[page_id].append(section_data_for_export)

    # Use service to get existing data mapping for efficiency
    # Map not required; we'll compute from entries below

    # Still get full FormData objects for compatibility with existing export logic
    existing_data_entries_for_export = FormData.query.filter_by(
        assignment_entity_status_id=assignment_entity_status.id
    ).all()
    existing_data_processed_for_export = {}
    for entry in existing_data_entries_for_export:
        if entry.form_item_id:
            item_key_suffix = f"form_item_{entry.form_item_id}"
            existing_data_processed_for_export[item_key_suffix] = entry.value

    # Create workbook and define styles
    workbook = openpyxl.Workbook()

    # IFRC Colors - Add FF prefix for openpyxl color format
    IFRC_RED = "FFED1B2E"  # Primary red
    IFRC_DARK_RED = "FFAF0E1B"  # Darker red for some elements
    IFRC_LIGHT_GRAY = "FFF5F5F5"  # Light gray for alternate rows
    IFRC_MEDIUM_GRAY = "FFE0E0E0"  # Medium gray for borders
    IFRC_DARK_GRAY = "FF666666"  # Dark gray for text
    IFRC_WHITE = "FFFFFFFF"  # White
    IFRC_YELLOW = "FFFFF9E6"  # Light yellow for input fields

    # Define styles
    title_font = Font(name='Arial', size=16, bold=True, color=IFRC_DARK_RED)
    section_title_font = Font(name='Arial', size=14, bold=True, color=IFRC_DARK_RED)
    item_label_font = Font(name='Arial', size=12, bold=True, color=IFRC_DARK_GRAY)
    header_font = Font(name='Arial', size=11, bold=True, color=IFRC_WHITE)
    normal_font = Font(name='Arial', size=11, color=IFRC_DARK_GRAY)

    # Fill patterns
    header_fill = PatternFill(start_color=IFRC_DARK_GRAY, end_color=IFRC_DARK_GRAY, fill_type='solid')
    section_fill = PatternFill(start_color=IFRC_LIGHT_GRAY, end_color=IFRC_LIGHT_GRAY, fill_type='solid')
    data_entry_fill = PatternFill(start_color=IFRC_YELLOW, end_color=IFRC_YELLOW, fill_type='solid')
    alternate_row_fill = PatternFill(start_color=IFRC_LIGHT_GRAY, end_color=IFRC_LIGHT_GRAY, fill_type='solid')

    # Borders
    no_border = Border(
        left=Side(style=None),
        right=Side(style=None),
        top=Side(style=None),
        bottom=Side(style=None)
    )
    disagg_border = Border(
        left=Side(style='thin', color=IFRC_DARK_GRAY),
        right=Side(style='thin', color=IFRC_DARK_GRAY),
        top=Side(style='thin', color=IFRC_DARK_GRAY),
        bottom=Side(style='thin', color=IFRC_DARK_GRAY)
    )

    # Alignments
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Create sheets for each page
    pages = list(form_template_for_export.pages) if form_template_for_export.is_paginated else [None]
    first_sheet = True

    for page in pages:
        page_id = page.id if page else default_page_id
        sheet_name_raw = get_localized_page_name(page)

        # Sanitize the sheet name to remove invalid characters
        invalid_chars = ['/', '\\', '?', '*', '[', ']']
        sheet_name = sheet_name_raw
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, '-')

        # Ensure the sheet name is not too long (Excel has a 31-character limit)
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        # Create or get sheet
        if first_sheet:
            data_sheet = workbook.active
            data_sheet.title = sheet_name
            first_sheet = False
        else:
            data_sheet = workbook.create_sheet(sheet_name)

            # Set default row height for all rows
            for row in range(1, 1000):  # Set a reasonable maximum
                data_sheet.row_dimensions[row].height = 17

            # Write header
        current_row = 1
        data_sheet.cell(row=current_row, column=1).value = f"Assignment: {form_template_for_export.name} - {assignment.period_name} for {country.name}"
        data_sheet.cell(row=current_row, column=1).font = title_font
        current_row += 2

        # Process sections for this page
        for section_data in sections_by_page.get(page_id, []):
            data_sheet.cell(row=current_row, column=1).value = section_data['name']
            data_sheet.cell(row=current_row, column=1).font = section_title_font
            current_row += 1

            for field_data in section_data['fields_ordered']:
                field_model = field_data['item_model']
                data_sheet.cell(row=current_row, column=2).value = f"{field_data['order']}. {field_data['label']}"
                data_sheet.cell(row=current_row, column=2).font = item_label_font
                current_row += 1  # Add space after field label
                col_offset = 3

                if field_data.get('is_indicator'):
                    indicator = field_model
                    item_key = f"form_item_{field_data['id']}"
                    entry_data = existing_data_processed_for_export.get(item_key, {})

                    # Get all available disaggregation modes for this indicator
                    allowed_modes = indicator.allowed_disaggregation_options if indicator.unit and indicator.unit in ['People', 'Volunteers', 'Staff'] else ['total']

                    if indicator.type == 'Number':
                        # Write disaggregation tables
                        for mode in allowed_modes:
                            mode_display = Config.DISAGGREGATION_MODES.get(mode, mode.title())
                            mode_cell = data_sheet.cell(row=current_row, column=col_offset, value=mode_display)
                            mode_cell.font = header_font
                            mode_cell.fill = header_fill
                            mode_cell.border = disagg_border
                            mode_cell.alignment = center_align
                            current_row += 1

                            # Get current values for this mode
                            current_values = entry_data.get('values', {}) if isinstance(entry_data, dict) else {'value': entry_data if entry_data is not None else ''}

                            if mode == 'total':
                                # Single total value
                                val_cell = data_sheet.cell(row=current_row, column=col_offset, value=current_values.get('total', ''))
                                val_cell.fill = data_entry_fill
                                val_cell.border = disagg_border
                                val_cell.number_format = '0'
                                current_row += 1

                            elif mode == 'sex':
                                # Sex disaggregation table
                                headers = indicator.effective_sex_categories
                                for col, header in enumerate(headers, col_offset):
                                    data_sheet.cell(row=current_row, column=col).value = header
                                    data_sheet.cell(row=current_row, column=col).font = header_font
                                current_row += 1

                                for col, category in enumerate(headers, col_offset):
                                    val_key = category.lower().replace(' ', '_')
                                    val_cell = data_sheet.cell(row=current_row, column=col, value=current_values.get(val_key, ''))
                                    val_cell.fill = data_entry_fill
                                    val_cell.border = disagg_border
                                    val_cell.number_format = '0'
                                    data_sheet.row_dimensions[current_row].height = 17
                                current_row += 1

                            elif mode == 'age':
                                # Age disaggregation table
                                headers = indicator.effective_age_groups
                                for col, header in enumerate(headers, col_offset):
                                    data_sheet.cell(row=current_row, column=col).value = header
                                    data_sheet.cell(row=current_row, column=col).font = header_font
                                current_row += 1

                                for col, category in enumerate(headers, col_offset):
                                    val_key = category.lower().replace(' ', '_').replace('+', 'plus')
                                    val_cell = data_sheet.cell(row=current_row, column=col, value=current_values.get(val_key, ''))
                                    val_cell.fill = data_entry_fill
                                    val_cell.border = disagg_border
                                    val_cell.number_format = '0'
                                current_row += 1

                            elif mode == 'sex_age':
                                # Sex-Age disaggregation table
                                age_groups = indicator.effective_age_groups
                                sex_categories = indicator.effective_sex_categories

                                # Write age group headers
                                category_cell = data_sheet.cell(row=current_row, column=col_offset, value="Category")
                                category_cell.font = header_font
                                category_cell.fill = header_fill
                                category_cell.border = disagg_border
                                category_cell.alignment = center_align

                                for col, age in enumerate(age_groups, col_offset + 1):
                                    header_cell = data_sheet.cell(row=current_row, column=col, value=age)
                                    header_cell.font = header_font
                                    header_cell.fill = header_fill
                                    header_cell.border = disagg_border
                                    header_cell.alignment = center_align
                                current_row += 1

                                # Write sex-age values
                                for sex in sex_categories:
                                    sex_cell = data_sheet.cell(row=current_row, column=col_offset, value=sex)
                                    sex_cell.font = header_font
                                    sex_cell.fill = header_fill
                                    sex_cell.border = disagg_border
                                    sex_cell.alignment = center_align

                                    for col, age in enumerate(age_groups, col_offset + 1):
                                        val_key = f"{sex.lower().replace(' ', '_')}_{age.lower().replace(' ', '_').replace('+', 'plus')}"
                                        val_cell = data_sheet.cell(row=current_row, column=col, value=current_values.get(val_key, ''))
                                        val_cell.fill = data_entry_fill
                                        val_cell.border = disagg_border
                                        val_cell.number_format = '0'
                                    current_row += 1

                            current_row += 1  # Add space between disaggregation tables
                    else:
                                                        # Non-numeric indicators just get a single value cell
                                val_cell = data_sheet.cell(row=current_row, column=col_offset, value=current_values.get('value', ''))
                                val_cell.fill = data_entry_fill
                                val_cell.border = no_border
                                current_row += 2

                elif field_data.get('is_question'):
                    question = field_model
                    item_key = f"form_item_{question.id}"
                    q_value = existing_data_processed_for_export.get(item_key, '')
                    val_cell = data_sheet.cell(row=current_row, column=col_offset, value=q_value)
                    val_cell.fill = data_entry_fill
                    val_cell.border = no_border

                    if question.type == QuestionType.number:
                        val_cell.number_format = '0'
                    elif question.type == QuestionType.date and q_value:
                        with suppress(Exception):
                            val_cell.value = datetime.strptime(q_value, '%Y-%m-%d').date()
                            val_cell.number_format = 'yyyy-mm-dd'

                    if question.type in [QuestionType.single_choice, QuestionType.multiple_choice] and question.options:
                        options_list = [str(opt.get('value', opt) if isinstance(opt, dict) else opt) for opt in question.options]
                        options_str = ','.join([f'"{opt}"' for opt in options_list])
                        dv = DataValidation(type="list", formula1=f"={options_str}", allow_blank=True)
                        data_sheet.add_data_validation(dv)
                        dv.add(val_cell)

                    current_row += 2

                elif field_data.get('is_document'):
                    data_sheet.cell(row=current_row, column=col_offset, value="(Manage in Web Form)")
                    current_row += 2

                current_row += 1  # Extra space between fields

        # Set column widths
        data_sheet.column_dimensions['C'].width = 20  # Mode column
        # Auto-adjust other column widths
        for col_idx in range(1, data_sheet.max_column + 1):
            if get_column_letter(col_idx) != 'C':  # Skip column C as it's manually set
                data_sheet.column_dimensions[get_column_letter(col_idx)].autosize = True

    # Save and return the file
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"data_entry_{country.iso3}_{str(assignment.period_name).replace(' ', '_')}.xlsx"
    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    download_name=filename,
                    as_attachment=True)

@bp.route("/assignment_status/<int:aes_id>/import_excel", methods=["POST"])
@login_required
def handle_excel_import(aes_id):
    """Legacy Excel import endpoint (delegates to shared ExcelService).

    NOTE: This endpoint is maintained for backward compatibility and testing.
    New code should use excel.import_assignment_excel instead.
    """
    # Maximum file size for Excel imports (10MB) - matches excel_routes.py
    MAX_EXCEL_FILE_SIZE = 10 * 1024 * 1024

    assignment_entity_status = AssignmentEntityStatus.query.get_or_404(aes_id)

    from app.services.authorization_service import AuthorizationService
    if not AuthorizationService.can_edit_assignment(assignment_entity_status, current_user):
        flash("You are not authorized to import data for this assignment or it's not in an editable state.", "warning")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    excel_file = request.files.get('excel_file')
    if not excel_file or excel_file.filename == '':
        flash("No Excel file selected for import.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    # Validate file extension (xlsx only)
    if not excel_file.filename.lower().endswith('.xlsx'):
        flash("Invalid file type. Please upload a .xlsx file.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    # Validate file size
    file_size = excel_file.content_length
    if file_size is None:
        excel_file.seek(0, 2)
        file_size = excel_file.tell()
        excel_file.seek(0)

    if file_size > MAX_EXCEL_FILE_SIZE:
        flash(f"File size ({file_size / (1024*1024):.2f}MB) exceeds the maximum allowed size of 10MB.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    # SECURITY: Validate MIME type to prevent file spoofing
    try:
        from app.utils.advanced_validation import AdvancedValidator
        import os
        file_ext = os.path.splitext(excel_file.filename)[1].lower()
        is_valid_mime, detected_mime = AdvancedValidator.validate_mime_type(excel_file, [file_ext])
        if not is_valid_mime:
            current_app.logger.warning(f"Excel import MIME mismatch: claimed {file_ext}, detected {detected_mime}")
            flash("File content does not match its extension. Please upload a valid Excel file.", "danger")
            return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))
    except Exception as e:
        current_app.logger.warning(f"MIME validation error for Excel import: {e}", exc_info=True)
        flash("Unable to validate file type. Please try again.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    try:
        workbook = ExcelService.load_workbook(excel_file)
    except ValueError as exc:
        flash("An error occurred. Please try again.", "danger")
        return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))

    result = ExcelService.import_assignment_data(assignment_entity_status, workbook)
    if result['success']:
        if result.get('errors'):
            error_msg = f"Excel import completed with {result['updated_count']} values saved. Errors: {', '.join(result['errors'][:5])}"
            if len(result['errors']) > 5:
                error_msg += f" (and {len(result['errors']) - 5} more)"
            flash(error_msg, "warning")
        else:
            flash(f"Excel import completed: {result['updated_count']} values saved.", "success")
    else:
        error_msg = f"Excel import failed: {', '.join(result.get('errors', [])[:5])}"
        if len(result.get('errors', [])) > 5:
            error_msg += f" (and {len(result.get('errors', [])) - 5} more)"
        flash(error_msg, "danger")

    return redirect(url_for("forms.view_edit_form", form_type="assignment", form_id=aes_id))


@bp.route("/delete_self_report_assignment/<int:aes_id>", methods=["POST"])
@login_required
def delete_self_report_assignment(aes_id):
    assignment_entity_status = AssignmentEntityStatus.query.get_or_404(aes_id)
    from app.services.authorization_service import AuthorizationService
    if not AuthorizationService.check_self_report_access(assignment_entity_status, current_user):
        flash("Access denied.", "warning"); return redirect(url_for("main.dashboard"))
    if assignment_entity_status.assigned_form.period_name != SELF_REPORT_PERIOD_NAME:
        flash("You can only delete self-reported assignments.", "warning"); return redirect(url_for("main.dashboard"))
    csrf_form = FlaskForm()
    if csrf_form.validate_on_submit():
        try:
            # Capture values before delete to avoid lazy-load on a detached instance
            template_name = (
                assignment_entity_status.assigned_form.template.name
                if assignment_entity_status.assigned_form and assignment_entity_status.assigned_form.template
                else "Template"
            )
            country_name = (
                assignment_entity_status.country.name
                if assignment_entity_status.country
                else "Unknown"
            )

            db.session.delete(assignment_entity_status)
            db.session.flush()
            flash(
                f"Self-reported assignment '{template_name}' for {country_name} deleted successfully.",
                "success",
            )
        except Exception as e:
            request_transaction_rollback(); flash("Error deleting self-reported assignment.", "danger")
    else:
        flash("Deletion failed due to a security issue.", "danger")
    return redirect(url_for("main.dashboard"))

# API endpoints moved to forms_api.py blueprint

# Additional routes for public submission management
@bp.route("/public-submission/<int:submission_id>/view", methods=["GET"])
@admin_required
def view_public_submission(submission_id):
    """View public submission details (read-only)."""
    return handle_public_submission_form(submission_id, is_edit_mode=False)

@bp.route("/public-submission/<int:submission_id>/edit", methods=["GET", "POST"])
@admin_required
def edit_public_submission(submission_id):
    """Edit public submission data."""
    return handle_public_submission_form(submission_id, is_edit_mode=True)

@bp.route("/public-submission/<int:submission_id>/approve", methods=["POST"])
@admin_required
def approve_public_submission(submission_id):
    """Approve a public submission."""
    submission = PublicSubmission.query.get_or_404(submission_id)
    csrf_form = FlaskForm()

    if csrf_form.validate_on_submit():
        try:
            submission.status = PublicSubmissionStatus.approved
            db.session.flush()
            flash(f"Public Submission for {submission.country.name} approved.", "success")
        except Exception as e:
            request_transaction_rollback()
            flash("An error occurred. Please try again.", "danger")
            current_app.logger.error(f"Error approving public submission {submission_id}: {e}", exc_info=True)
    else:
        flash("Approval failed due to a security issue. Please try again.", "danger")

    return redirect(url_for('forms.view_public_submission', submission_id=submission_id))

@bp.route("/public-submission/<int:submission_id>/reject", methods=["POST"])
@admin_required
def reject_public_submission(submission_id):
    """Reject a public submission."""
    submission = PublicSubmission.query.get_or_404(submission_id)
    csrf_form = FlaskForm()

    if csrf_form.validate_on_submit():
        try:
            submission.status = PublicSubmissionStatus.rejected
            db.session.flush()
            flash(f"Public Submission for {submission.country.name} rejected.", "success")
        except Exception as e:
            request_transaction_rollback()
            flash("An error occurred. Please try again.", "danger")
            current_app.logger.error(f"Error rejecting public submission {submission_id}: {e}", exc_info=True)
    else:
        flash("Rejection failed due to a security issue. Please try again.", "danger")

    return redirect(url_for('forms.view_public_submission', submission_id=submission_id))

@bp.route("/public-submission/<int:submission_id>/delete", methods=["POST"])
@admin_required
def delete_public_submission(submission_id):
    """Delete a public submission."""
    submission = PublicSubmission.query.get_or_404(submission_id)
    csrf_form = FlaskForm()

    if csrf_form.validate_on_submit():
        try:
            country_name = submission.country.name if submission.country else 'N/A'

            # Delete associated documents
            from app.services.document_service import DocumentService
            for doc in submission.submitted_documents:
                try:
                    abs_path = DocumentService._resolve_storage_path(doc.storage_path)
                    if os.path.exists(abs_path):
                        os.remove(abs_path)
                except Exception as e:
                    current_app.logger.error(f"Error deleting document file {doc.storage_path}: {e}", exc_info=True)

            db.session.delete(submission)
            db.session.flush()
            flash(f"Public Submission for {country_name} deleted successfully.", "success")

            # Redirect to admin dashboard or submissions list
            return redirect(url_for("main.dashboard"))

        except Exception as e:
            request_transaction_rollback()
            flash("An error occurred. Please try again.", "danger")
            current_app.logger.error(f"Error deleting public submission {submission_id}: {e}", exc_info=True)
    else:
        flash("Deletion failed due to a security issue. Please try again.", "danger")

    return redirect(url_for('forms.view_public_submission', submission_id=submission_id))

@bp.route("/public-submission/<int:submission_id>/status", methods=["POST"])
@admin_required
def update_public_submission_status(submission_id):
    """Update public submission status via AJAX."""
    submission = PublicSubmission.query.get_or_404(submission_id)
    csrf_form = FlaskForm()

    if csrf_form.validate_on_submit():
        try:
            new_status = request.form.get('status')
            if new_status in ['pending', 'approved', 'rejected']:
                submission.status = getattr(PublicSubmissionStatus, new_status)
                db.session.flush()
                flash(f"Public Submission for {submission.country.name} status updated to {new_status}.", "success")
                return {"success": True, "message": "Status updated successfully"}
            else:
                return json_bad_request("Invalid status value", success=False)
        except Exception as e:
            request_transaction_rollback()
            current_app.logger.error(f"Error updating public submission {submission_id} status: {e}", exc_info=True)
            return json_server_error(GENERIC_ERROR_MESSAGE, success=False)
    else:
        return json_bad_request("Security validation failed", success=False)

# ========================= PUBLIC FORM ROUTES =========================

@bp.route("/debug/public-form-test", methods=["GET", "POST"])
@login_required
def debug_public_form_test():
    """Debug route to test public form logging.

    Security: Protected by @login_required and only available in DEBUG mode.
    """
    # Security: Only allow in DEBUG mode
    if not current_app.config.get('DEBUG', False):
        abort(404)

    current_app.logger.debug("=== DEBUG PUBLIC FORM TEST ===")
    current_app.logger.debug(f"Method: {request.method}")
    current_app.logger.debug(f"Form data: {dict(request.form)}")
    current_app.logger.debug(f"Files data: {dict(request.files)}")

    if request.method == "POST":
        csrf_form = FlaskForm()
        current_app.logger.debug(f"CSRF token present: {'csrf_token' in request.form}")
        current_app.logger.debug(f"CSRF validation: {csrf_form.validate_on_submit()}")
        current_app.logger.debug(f"CSRF errors: {csrf_form.errors}")

        return json_ok(
            status="success",
            csrf_valid=csrf_form.validate_on_submit(),
            csrf_errors=csrf_form.errors,
            form_data=dict(request.form)
        )

    return json_ok(status="debug_route_working")

@bp.route("/public/<uuid:public_token>", methods=["GET", "POST"])
def fill_public_form(public_token):
    """Main public form filling route - allows external users to submit data."""
    current_app.logger.debug(f"=== PUBLIC FORM ROUTE ENTRY ===")
    current_app.logger.debug(f"Method: {request.method}")
    current_app.logger.debug(f"Public token: {public_token}")
    current_app.logger.debug(f"Request URL: {request.url}")

    # Initialize data structures for form state
    existing_data_processed = {}
    existing_submitted_documents_dict = {}

    # Load the AssignedForm with necessary relationships
    assigned_form = AssignedForm.query.filter_by(unique_token=str(public_token)).options(
        joinedload(AssignedForm.template)
    ).first()

    if not assigned_form:
        return render_template("admin/public/public_form_unavailable.html",
                               title="Form Not Found",
                               message="This form link is not valid or has been removed.")

    # Check if the public URL is active
    if not assigned_form.is_public_active:
        return render_template("admin/public/public_form_unavailable.html",
                               title="Form Unavailable",
                               message="This form is currently not active.")

    # Deactivated assignment guard: public intake disabled when assignment is inactive
    if getattr(assigned_form, "is_active", True) is False:
        return render_template("admin/public/public_form_unavailable.html",
                               title="Form Unavailable",
                               message="This form is currently inactive.")

    # Get form template and sections
    form_template = assigned_form.template
    sections = FormSection.query.filter_by(template_id=form_template.id, version_id=form_template.published_version_id).order_by(FormSection.order).all()

    # Import form classes for template compatibility
    from wtforms import SelectField, StringField, EmailField, SubmitField
    from wtforms.validators import DataRequired, Email

    # Create helper classes for template compatibility
    class PublicCountrySelectForm(FlaskForm):
        country_id = SelectField("Select Your Country", coerce=int, validators=[DataRequired()])

    class PublicSubmissionDetailsForm(FlaskForm):
        submitter_name = StringField("Your Name", validators=[DataRequired()])
        submitter_email = EmailField("Your Email", validators=[DataRequired(), Email()])
        submit = SubmitField("Submit Form")

    class DummyACS:
        def __init__(self):
            self.id = None

    class DummyStatus:
        def __init__(self, status, template, country, period_name):
            self.status = status
            self.id = None
            self.assigned_form = type('DummyAssignedForm', (), {
                'template': template,
                'period_name': period_name
            })()
            self.country = country

    # Process sections and their fields
    dummy_acs = DummyACS()

    for section in sections:
        form_items = get_form_items_for_section(section, dummy_acs)
        current_section_fields = []

        # Process all form items and set up properties for public form
        for item in form_items:
            # Common properties for all items
            item.conditions = []  # No relevance conditions for public form
            item.validations_from_db = []  # No validation conditions for public form
            item.is_required_for_js = item.is_required
            item.layout_column_width = getattr(item, 'layout_column_width', 12)
            item.layout_break_after = getattr(item, 'layout_break_after', False)

            current_section_fields.append(item)

        # Sort fields by order and attach to section
        current_section_fields.sort(key=lambda x: x.order)
        section.fields_ordered = current_section_fields

    # Set up country selection from AssignedForm public countries
    sorted_countries = sorted(assigned_form.public_countries, key=lambda c: c.name)

    country_choices = [(c.id, c.name) for c in sorted_countries]

    if not country_choices:
        current_app.logger.warning(f"Public form link {public_token} has no countries assigned.")
        return render_template("admin/public/public_form_unavailable.html",
                               title="Form Unavailable",
                               message="This form is not configured for any countries.")

    # Initialize forms
    country_select_form = PublicCountrySelectForm()
    country_select_form.country_id.choices = country_choices
    submission_details_form = PublicSubmissionDetailsForm()
    csrf_form = FlaskForm()

    selected_country = None

    # Handle form submission
    if request.method == "POST" and 'submit_form' in request.form:
        current_app.logger.debug(f"=== PUBLIC FORM POST DEBUG ===")
        current_app.logger.debug(f"Public form POST request received for token: {public_token}")
        current_app.logger.debug(f"Form data keys: {list(request.form.keys())}")
        current_app.logger.debug(f"Files data keys: {list(request.files.keys())}")
        current_app.logger.debug(f"CSRF token in form: {'csrf_token' in request.form}")

        # Test CSRF validation
        csrf_valid = csrf_form.validate_on_submit()
        current_app.logger.debug(f"CSRF form valid: {csrf_valid}")
        current_app.logger.debug(f"CSRF form errors: {csrf_form.errors}")

        # Test submission form validation
        submission_valid = submission_details_form.validate_on_submit()
        current_app.logger.debug(f"Submission form valid: {submission_valid}")
        current_app.logger.debug(f"Submission form errors: {submission_details_form.errors}")

        # Test country form validation
        country_valid = country_select_form.country_id.validate(country_select_form)
        current_app.logger.debug(f"Country form valid: {country_valid}")
        current_app.logger.debug(f"Country form errors: {country_select_form.errors}")

        # Log form data for debugging
        current_app.logger.debug(f"Submitter name: {request.form.get('submitter_name', 'NOT_FOUND')}")
        current_app.logger.debug(f"Submitter email: {request.form.get('submitter_email', 'NOT_FOUND')}")
        current_app.logger.debug(f"Country ID: {request.form.get('country_id', 'NOT_FOUND')}")

        if csrf_valid and submission_valid and country_valid:
            selected_country_id = country_select_form.country_id.data
            selected_country = Country.query.get(selected_country_id)

            # Validate country selection for AssignedForm
            valid_countries = assigned_form.public_countries

            if not selected_country or selected_country not in valid_countries:
                flash("Invalid country selection during submission.", "danger")
            else:
                try:
                    # Create submission for AssignedForm
                    submission = PublicSubmission(
                        assigned_form_id=assigned_form.id,
                        country_id=selected_country.id,
                        submitter_name=submission_details_form.submitter_name.data,
                        submitter_email=submission_details_form.submitter_email.data,
                        status=PublicSubmissionStatus.pending
                    )
                    db.session.add(submission)
                    db.session.flush()  # Get submission.id

                    # Use the unified FormDataService for comprehensive data processing
                    from app.services.form_data_service import FormDataService

                    # Process form data using the unified service (no CSRF needed for public submissions)
                    submission_result = FormDataService.process_form_submission(
                        submission, sections, csrf_form=None
                    )

                    if not submission_result['success']:
                        for error in submission_result['validation_errors']:
                            flash(error, "danger")
                        return redirect(url_for("forms.fill_public_form", public_token=public_token))

                    # Log field changes for activity tracking
                    field_changes_tracker = submission_result['field_changes']

                    # Helper function to parse field values for better readability
                    def parse_field_value_for_display(value):
                        if value is None:
                            return "None"
                        elif isinstance(value, dict) and 'values' in value:
                            # Disaggregated data
                            total = sum(v for v in value['values'].values() if isinstance(v, (int, float)))
                            return f"Total: {total} (Disaggregated: {value['mode']})"
                        elif isinstance(value, str) and value.startswith('{'):
                            with suppress(Exception):
                                import json
                                parsed = json.loads(value)
                                if isinstance(parsed, dict) and 'values' in parsed:
                                    total = sum(v for v in parsed['values'].values() if isinstance(v, (int, float)))
                                    return f"Total: {total} (Disaggregated: {parsed['mode']})"
                        return str(value)

                    # Log field changes
                    for change in field_changes_tracker:
                        if change.get('type') in ['added', 'updated']:
                            old_val = parse_field_value_for_display(change.get('old_value'))
                            new_val = parse_field_value_for_display(change.get('new_value'))
                            current_app.logger.info(f"Field '{change.get('field_name', 'Unknown')}' {change['type']}: {old_val} -> {new_val}")

                    # Check for required fields validation
                    missing_required_fields = []
                    all_required_fields_completed = True

                    # Check if submission was successful
                    if submission_result['success']:
                        # Send notification to admins about public submission
                        try:
                            from app.utils.notifications import notify_public_submission_received
                            notify_public_submission_received(submission)
                        except Exception as e:
                            current_app.logger.error(f"Error sending public submission notification: {e}", exc_info=True)
                            # Don't fail the submission if notification fails

                        return redirect(url_for('forms.public_submission_success', submission_id=submission.id))
                    else:
                        request_transaction_rollback()
                        for error in submission_result['validation_errors']:
                            flash(error, "warning")

                except Exception as e:
                    request_transaction_rollback()
                    flash("An error occurred during submission. Please try again.", "danger")
                    current_app.logger.error(f"Error during public form submission: {e}", exc_info=True)
        else:
            current_app.logger.debug(f"=== VALIDATION FAILED ===")
            if not csrf_valid:
                flash("Form submission failed due to a security issue. Please try again.", "danger")
                current_app.logger.warning(f"CSRF validation failed for public form {public_token}. Errors: {csrf_form.errors}")
            elif not submission_valid:
                flash("Form submission failed due to validation errors. Please check your entries.", "danger")
                current_app.logger.warning(f"Submission form validation failed for public form {public_token}. Errors: {submission_details_form.errors}")
            elif not country_valid:
                flash("Form submission failed due to validation errors. Please check your entries.", "danger")
                current_app.logger.warning(f"Country form validation failed for public form {public_token}. Errors: {country_select_form.errors}")
            else:
                flash("Form submission failed due to validation errors. Please check your entries.", "danger")
                current_app.logger.warning(f"Unknown validation failure for public form {public_token}")

    # Handle GET request or failed submission - render form
    # Get selected country from URL parameter if provided
    selected_country_id_from_args = request.args.get('country_id', type=int)
    if selected_country_id_from_args:
        selected_country = Country.query.get(selected_country_id_from_args)
        valid_countries = assigned_form.public_countries

        if selected_country and selected_country in valid_countries:
            country_select_form.country_id.data = selected_country.id

    # Create dummy status for template
    period_name = assigned_form.period_name or "Public Submission"

    assignment_status = DummyStatus(
        status="In Progress",
        template=form_template,
        country=sorted_countries[0] if sorted_countries else None,
        period_name=period_name
    )

    # Calculate section statuses (all empty for new form)
    section_statuses = {section.name: 'Not Started' for section in sections}

    # Apply page and section translations
    # Apply page translations to all pages in the template
    for page in FormPage.query.filter_by(template_id=form_template.id, version_id=form_template.published_version_id).order_by(FormPage.order).all():
        page.display_name = get_localized_page_name(page)

    # Apply page translations to all page objects referenced by sections
    page_ids_processed = set()
    for section in sections:
        if section.page and section.page.id not in page_ids_processed:
            section.page.display_name = get_localized_page_name(section.page)
            page_ids_processed.add(section.page.id)

    # Apply section translations to all sections (ISO codes only)
    current_locale_short = (str(get_locale()) if get_locale() else 'en').split('_', 1)[0]
    for section in sections:
        translated_name = None
        if section.name_translations and isinstance(section.name_translations, dict):
            translated_name = section.name_translations.get(current_locale_short) or section.name_translations.get('en')
        section.display_name = translated_name.strip() if isinstance(translated_name, str) and translated_name.strip() else section.name

    # Prepare available indicators by section
    available_indicators_by_section = {}
    for section in sections:
        available_indicators_by_section[section.id] = []

        # Add display filter configuration for dynamic sections
        if section.section_type == 'dynamic_indicators':
            section.data_entry_display_filters_config = getattr(section, 'data_entry_display_filters_list', [])

    return render_template("forms/entry_form/entry_form.html",
                           title=get_localized_template_name(form_template),
                           assignment=assigned_form,
                           assignment_status=assignment_status,
                           template_structure=form_template,
                           form=submission_details_form,
                           csrf_form=csrf_form,
                           existing_data={},
                           existing_submitted_documents={},
                           section_statuses=section_statuses,
                           slugify_age_group=slugify_age_group,
                           config=Config,
                           can_edit=True,
                           QuestionType=QuestionType,
                           isinstance=isinstance,
                           json=json,
                           hasattr=hasattr,
                           available_indicators_by_section=available_indicators_by_section,
                           get_localized_country_name=get_localized_country_name,
                           translation_key=get_translation_key(),
                           # Add localization functions for template use
                           get_localized_indicator_definition=get_localized_indicator_definition,
                           get_localized_indicator_type=get_localized_indicator_type,
                           get_localized_indicator_unit=get_localized_indicator_unit,
                           get_localized_template_name=get_localized_template_name,
                           # Public form specific variables
                           is_public_submission=True,
                           is_preview_mode=False,
                           country_select_form=country_select_form,
                           submission_details_form=submission_details_form,
                           public_token=public_token,
                           form_action=url_for('forms.fill_public_form', public_token=public_token),
                           # Add plugin manager information for template use
                           plugin_manager=current_app.plugin_manager if hasattr(current_app, 'plugin_manager') else None,
                           form_integration=current_app.form_integration if hasattr(current_app, 'form_integration') else None)

@bp.route("/public-submission/<int:submission_id>/success", methods=["GET"])
def public_submission_success(submission_id):
    """Show success page after public form submission."""
    submission = PublicSubmission.query.get_or_404(submission_id)
    return render_template("admin/public/public_submission_success.html",
                           title="Submission Successful",
                           submission=submission)

@bp.route("/public-document/<int:document_id>/download", methods=["GET"])
def download_public_document_public(document_id):
    """Download a document from a public submission (public access)."""
    from app.services.document_service import DocumentService
    try:
        directory, filename, download_name = DocumentService.get_public_download_paths(document_id)
        return send_from_directory(directory, filename, as_attachment=True, download_name=download_name)
    except PermissionError:
        abort(404)
    except Exception as e:
        current_app.logger.error(f"Error serving public document {document_id}: {e}", exc_info=True)
        flash("An error occurred while trying to download the file.", "danger")
        return redirect(url_for("main.dashboard"))

@bp.route("/templates/preview/<int:template_id>", methods=["GET"])
@admin_required
def preview_template(template_id):
    """Preview a form template using existing form processing logic."""
    # Require templates permission and access (defense in depth).
    # This endpoint lives outside /admin but is an admin-only tool.
    from app.services.authorization_service import AuthorizationService
    from app.routes.admin.shared import check_template_access
    if not AuthorizationService.has_rbac_permission(current_user, "admin.templates.view"):
        flash(_("Access denied."), "warning")
        return redirect(url_for("main.dashboard"))
    if not check_template_access(template_id, current_user.id):
        flash(_("Access denied."), "warning")
        return redirect(url_for("main.dashboard"))

    from app.models import FormTemplate, FormSection
    from app.models.forms import FormTemplateVersion, FormPage
    from flask import request
    from flask_wtf import FlaskForm

    # Get the template
    template = FormTemplate.query.get_or_404(template_id)

    # Determine which version to preview: explicit version_id > published > latest
    requested_version_id = request.args.get('version_id', type=int)
    selected_version = None
    if requested_version_id:
        selected_version = FormTemplateVersion.query.filter_by(id=requested_version_id, template_id=template.id).first()
    if not selected_version and template.published_version_id:
        selected_version = FormTemplateVersion.query.get(template.published_version_id)
    if not selected_version:
        selected_version = FormTemplateVersion.query.filter_by(template_id=template.id).order_by(FormTemplateVersion.created_at.desc()).first()

    selected_version_id = selected_version.id if selected_version else None

    # -----------------------------
    # Preview context ("view as")
    # -----------------------------
    # Optional query params:
    # - view_as: "entity_type:entity_id" (e.g. "country:123", "ns_branch:45")
    # - period_name: string (e.g. "2025 H2")
    from app.models.assignments import AssignedForm, AssignmentEntityStatus
    from app.models.enums import EntityType
    from app import db
    from app.services.entity_service import EntityService

    requested_period_name = (request.args.get("period_name") or "").strip()
    # In preview mode, the period name input should start empty (use placeholder text in the UI).
    selected_period_name = requested_period_name

    requested_view_as = (request.args.get("view_as") or "").strip()
    selected_entity_type = None
    selected_entity_id = None
    if requested_view_as and ":" in requested_view_as:
        try:
            et, eid_str = requested_view_as.split(":", 1)
            eid = int(eid_str)
            # Basic allowlist for entity types
            allowed_types = {e.value for e in EntityType}
            if et in allowed_types and eid > 0:
                selected_entity_type = et
                selected_entity_id = eid
        except Exception as e:
            current_app.logger.debug("view_as parse failed: %s", e)
            selected_entity_type = None
            selected_entity_id = None

    # Build dropdown options for "view as" from entities the current user can access.
    # (Admins/system managers can see all entities.)
    preview_view_as_options = [{"value": "", "label": _("Preview placeholders")}]
    try:
        option_rows = []

        if AuthorizationService.is_admin(current_user) or AuthorizationService.is_system_manager(current_user):
            # Admin/system manager: include all entities (active if supported by the model)
            for et, model_cls in EntityService.ENTITY_MODEL_MAP.items():
                try:
                    entities = EntityService.get_all_entities_by_type(et, filter_active=True)
                    for ent in entities:
                        if not ent or not getattr(ent, "id", None):
                            continue
                        option_rows.append((et, int(ent.id)))
                except Exception as e:
                    current_app.logger.debug("EntityService.get_all_entities failed: %s", e)
                    continue
        else:
            # Non-admin: include only explicitly permitted entities
            from app.models.core import UserEntityPermission
            perms = UserEntityPermission.query.filter_by(user_id=current_user.id).all()
            for p in perms:
                if not p:
                    continue
                et = getattr(p, "entity_type", None)
                eid = getattr(p, "entity_id", None)
                if et and eid:
                    option_rows.append((et, int(eid)))

        # Dedupe and turn into display labels
        seen = set()
        enriched = []
        for et, eid in option_rows:
            key = (et, eid)
            if key in seen:
                continue
            seen.add(key)
            try:
                name = EntityService.get_localized_entity_name(et, eid, include_hierarchy=True)
                label = f"{EntityService.get_entity_type_label(et)}: {name}"
            except Exception as e:
                current_app.logger.debug("get_localized_entity_name failed: %s", e)
                label = f"{et}:{eid}"
            enriched.append((label, et, eid))

        for label, et, eid in sorted(enriched, key=lambda x: (x[0] or "").lower()):
            preview_view_as_options.append({"value": f"{et}:{eid}", "label": label})
    except Exception as e:
        current_app.logger.debug("preview_view_as_options failed: %s", e)

    # Build dropdown options for period_name from periods already used for this template.
    preview_period_name_options = []
    try:
        period_rows = (
            db.session.query(AssignedForm.period_name)
            .filter(AssignedForm.template_id == template.id)
            .distinct()
            .order_by(AssignedForm.period_name.desc())
            .all()
        )
        preview_period_name_options = [r[0] for r in period_rows if r and r[0]]
    except Exception as e:
        current_app.logger.debug("preview_period_name_options failed: %s", e)
        preview_period_name_options = []

    # Get all sections (both parent and sub-sections) and organize them hierarchically for the selected version
    all_sections = FormSection.query.filter_by(template_id=template.id, version_id=selected_version_id).order_by(FormSection.order).all()

    # Create mock assignment/entity status for compatibility with form processing functions
    class MockACS:
        def __init__(self, template, period_name, entity_type=None, entity_id=None):
            self.id = 0  # Use integer 0 for preview mode
            self.status = 'Preview Mode'
            self.due_date = None
            self.entity_type = entity_type
            self.entity_id = entity_id

            # Mock assignment
            mock_assignment = type('MockAssignment', (), {})()
            mock_assignment.template = template
            mock_assignment.period_name = period_name or 'Preview Period'
            self.assigned_form = mock_assignment

            # Mock country with all required attributes
            mock_country = type('MockCountry', (), {})()
            mock_country.name = 'Preview Country'
            mock_country.iso3 = 'PRE'
            # Set translations using ISO codes
            mock_country.name_translations = {
                'fr': 'Pays de Prévisualisation',
                'es': 'País de Vista Previa',
                'ar': 'بلد المعاينة',
                'ru': 'Страна Предварительного Просмотра',
                'zh': '预览国家',
                'hi': 'पूर्वावलोकन देश'
            }
            self._placeholder_country = mock_country

        @property
        def entity(self):
            """Entity object for non-country entity previews."""
            if not self.entity_type or not self.entity_id:
                return None
            try:
                return EntityService.get_entity(self.entity_type, int(self.entity_id))
            except Exception as e:
                current_app.logger.debug("MockACS.entity get_entity failed: %s", e)
                return None

        @property
        def country(self):
            """Country for the preview context (used heavily by templates and JS)."""
            if not self.entity_type or not self.entity_id:
                return self._placeholder_country
            try:
                c = EntityService.get_country_for_entity(self.entity_type, int(self.entity_id))
                return c or self._placeholder_country
            except Exception as e:
                current_app.logger.debug("MockACS.country failed: %s", e)
                return self._placeholder_country

        @property
        def country_id(self):
            """Compatibility helper for legacy code that expects country_id."""
            try:
                if self.entity_type == EntityType.country.value and self.entity_id:
                    return int(self.entity_id)
                c = self.country
                return int(c.id) if c and hasattr(c, "id") else None
            except Exception as e:
                current_app.logger.debug("MockACS.country_id failed: %s", e)
                return None

    mock_acs = MockACS(
        template,
        selected_period_name,
        entity_type=selected_entity_type,
        entity_id=selected_entity_id,
    )

    # Use existing form processing logic
    for section in all_sections:
        # Process form items using existing function
        section.fields_ordered = get_form_items_for_section(section, mock_acs)

        # Layout properties are now handled by properties that read from config

    # Load pages for the selected version and apply translations
    published_pages = FormPage.query.filter_by(template_id=template.id, version_id=selected_version_id).order_by(FormPage.order).all()
    for page in published_pages:
        page.display_name = get_localized_page_name(page)

    # Apply page translations to all page objects referenced by sections
    page_ids_processed = set()
    for section in all_sections:
        if section.page and section.page.id not in page_ids_processed:
            section.page.display_name = get_localized_page_name(section.page)
            page_ids_processed.add(section.page.id)

    # Apply section translations to all sections
    for section in all_sections:
        if section.name_translations:
            current_locale = get_locale()
            # Use ISO codes directly - no conversion needed
            translated_name = section.name_translations.get(current_locale)

            # Use translated name if available, otherwise fall back to original name
            if translated_name and translated_name.strip():
                section.display_name = translated_name
            else:
                section.display_name = section.name
        else:
            section.display_name = section.name

    # Prepare available indicators by section
    # Add display filter configuration for dynamic sections
    for section in all_sections:
        if section.section_type == 'dynamic_indicators':
            section.data_entry_display_filters_config = getattr(section, 'data_entry_display_filters_list', [])

    # IMPORTANT: In preview mode we still need to populate the indicator picker list,
    # otherwise the UI will always show "No more indicators available for this section."
    available_indicators_by_section = TemplatePreparationService._prepare_available_indicators(all_sections)

    # Create CSRF form
    csrf_form = FlaskForm()

    # Resolve template variables and apply to form items (same approach as real assignments)
    variable_configs = {}
    try:
        from app.services.variable_resolution_service import VariableResolutionService
        if selected_version:
            variable_configs = getattr(selected_version, "variables", None) or {}
            resolved_variables = VariableResolutionService.resolve_variables(
                selected_version,
                mock_acs
            )

            # Apply variable replacements to section names and form item labels/descriptions
            for section in all_sections:
                # Section display name (already localized)
                if hasattr(section, 'display_name') and section.display_name and resolved_variables:
                    try:
                        section._display_name_resolved = VariableResolutionService.replace_variables_in_text(
                            section.display_name,
                            resolved_variables,
                            variable_configs
                        )
                    except Exception as e:
                        current_app.logger.debug("section display_name replace failed: %s", e)

                for field in getattr(section, 'fields_ordered', []) or []:
                    if not field or not hasattr(field, 'id'):
                        continue

                    # Label
                    if hasattr(field, 'label') and field.label and resolved_variables:
                        try:
                            field._display_label = VariableResolutionService.replace_variables_in_text(
                                field.label,
                                resolved_variables,
                                variable_configs
                            )
                        except Exception as e:
                            current_app.logger.debug("field label replace failed: %s", e)

                    # Display label
                    if hasattr(field, 'display_label') and field.display_label and resolved_variables:
                        try:
                            field._display_label_resolved = VariableResolutionService.replace_variables_in_text(
                                field.display_label,
                                resolved_variables,
                                variable_configs
                            )
                        except Exception as e:
                            current_app.logger.debug("field display_label replace failed: %s", e)

                    # Definition
                    if hasattr(field, 'definition') and field.definition and resolved_variables:
                        try:
                            field._display_definition = VariableResolutionService.replace_variables_in_text(
                                field.definition,
                                resolved_variables,
                                variable_configs
                            )
                        except Exception as e:
                            current_app.logger.debug("field definition replace failed: %s", e)

                    # Description
                    if hasattr(field, 'description') and field.description and resolved_variables:
                        try:
                            field._display_description = VariableResolutionService.replace_variables_in_text(
                                field.description,
                                resolved_variables,
                                variable_configs
                            )
                        except Exception as e:
                            current_app.logger.debug("field description replace failed: %s", e)

                    # Matrix manual row labels (manual mode matrices only)
                    if hasattr(field, 'item_type') and field.item_type == 'matrix' and hasattr(field, 'config') and field.config:
                        try:
                            matrix_config = field.config.get('matrix_config') if isinstance(field.config, dict) and 'matrix_config' in field.config else field.config
                            if isinstance(matrix_config, dict):
                                row_mode = matrix_config.get('row_mode', 'manual')
                                if row_mode == 'manual' or not row_mode:
                                    rows = matrix_config.get('rows', [])
                                    if rows and isinstance(rows, list):
                                        field._display_matrix_rows = [
                                            VariableResolutionService.replace_variables_in_text(r, resolved_variables, variable_configs)
                                            if isinstance(r, str) else r
                                            for r in rows
                                        ]
                        except Exception as e:
                            current_app.logger.debug("matrix row variable replace failed: %s", e)
    except Exception as e:
        current_app.logger.debug("Preview variable resolution failed: %s", e)
        variable_configs = {}

    # Set up section statuses for preview
    section_statuses = {section.name: 'Not Started' for section in all_sections}

    return render_template("forms/entry_form/entry_form.html",
                         title=_("Preview: %(name)s", name=get_localized_template_name(template)),
                         assignment=mock_acs.assigned_form,
                         assignment_status=mock_acs,
                         template_structure=template,
                         all_sections=all_sections,
                         form=csrf_form,
                         published_pages=published_pages,
                         existing_data={},
                         existing_submitted_documents={},
                         section_statuses=section_statuses,
                         slugify_age_group=slugify_age_group,
                         config=Config,
                         can_edit=True,  # Enable editing in preview mode
                         QuestionType=QuestionType,
                         isinstance=isinstance,
                         json=json,
                         hasattr=hasattr,
                         available_indicators_by_section=available_indicators_by_section,
                         get_localized_country_name=get_localized_country_name,
                         # Add localization functions for template use
                         get_localized_indicator_definition=get_localized_indicator_definition,
                         get_localized_indicator_type=get_localized_indicator_type,
                         get_localized_indicator_unit=get_localized_indicator_unit,
                         get_localized_template_name=get_localized_template_name,
                         is_preview_mode=True,
                         template_id=template.id,
                         template_variables=variable_configs,
                         preview_version_id=selected_version_id,
                         preview_selected_view_as=(f"{selected_entity_type}:{selected_entity_id}" if selected_entity_type and selected_entity_id else ""),
                         preview_selected_period_name=selected_period_name,
                         preview_view_as_options=preview_view_as_options,
                         preview_period_name_options=preview_period_name_options,
                         # Add plugin manager information for template use
                         plugin_manager=current_app.plugin_manager if hasattr(current_app, 'plugin_manager') else None,
                         form_integration=current_app.form_integration if hasattr(current_app, 'form_integration') else None)


@bp.route('/matrix/search-rows', methods=['POST'])
@login_required
def search_matrix_rows():
    """
    API endpoint to search for rows in a list library for advanced matrix tables.
    This endpoint is for internal frontend use and requires user session authentication.

    Request Body:
        - lookup_list_id: ID of the lookup list to search
        - display_column: Column to use for row labels
        - filters: List of filters to apply (optional)
        - search_term: Search term to filter results (optional)
        - existing_rows: List of already selected row labels (optional)

    Returns:
        JSON object containing:
        - success: Boolean indicating success
        - options: List of available options with value and description
    """
    try:
        from app.models import LookupList, LookupListRow

        csrf_error = enforce_csrf_json()
        if csrf_error is not None:
            return csrf_error

        data = get_json_safe()
        err = require_json_keys(data, ['lookup_list_id', 'display_column'])
        if err:
            return err

        lookup_list_id = data.get('lookup_list_id')
        display_column = data.get('display_column')
        if not lookup_list_id or not display_column:
            return json_bad_request('lookup_list_id and display_column are required')
        filters = data.get('filters', [])
        search_term = data.get('search_term', '').strip()
        existing_rows = data.get('existing_rows', [])
        limit = data.get('limit', 500)  # Default to 500 for backward compatibility
        plugin_config = data.get('plugin_config', {})  # Plugin-specific configuration (e.g., emergency operations)

        def _row_matches_filters(row_data):
            """Check if a row dict matches all provided filters."""
            if not filters:
                return True
            if not isinstance(row_data, dict):
                return False
            for filter_item in filters:
                column = filter_item.get('column')
                operator = filter_item.get('operator', 'equals')
                value = filter_item.get('value')

                if column is None or value is None:
                    return False
                if column not in row_data:
                    return False

                row_value = str(row_data[column]).strip().lower()
                filter_value = str(value).strip().lower()

                if operator == 'equals' and row_value != filter_value:
                    return False
                if operator == 'not_equals' and row_value == filter_value:
                    return False
                if operator == 'contains' and filter_value not in row_value:
                    return False
                if operator == 'not_contains' and filter_value in row_value:
                    return False
            return True

        def _build_options_from_rows(rows_data):
            """Convert normalized rows_data into matrix search options."""
            if not isinstance(rows_data, list):
                return []

            filtered_rows = []
            for row_data in rows_data:
                if _row_matches_filters(row_data):
                    filtered_rows.append(row_data)

            options_local = []
            for row_data in filtered_rows:
                if not isinstance(row_data, dict):
                    continue

                if display_column not in row_data:
                    continue

                row_value = str(row_data[display_column]).strip()
                if not row_value or row_value in existing_rows:
                    continue

                if search_term and search_term.lower() not in row_value.lower():
                    continue

                description = None
                for desc_field in ['description', 'desc', 'details', 'notes']:
                    if desc_field in row_data and row_data[desc_field]:
                        description = str(row_data[desc_field])
                        break

                row_id = row_data.get('_id') or row_data.get('id')

                if not row_id:
                    current_app.logger.warning(
                        f"Matrix search option missing ID - row_value: {row_value}, row_data: {row_data}"
                    )

                if row_id and (not row_data.get('_id') or not row_data.get('id')):
                    row_data['_id'] = row_id
                    row_data['id'] = row_id

                options_local.append({
                    'value': row_value,
                    'description': description,
                    'data': row_data,
                    'id': row_id
                })
            return options_local

        def _detect_country_iso_from_matrix_context():
            """Detect country ISO from matrix search request context."""
            try:
                from app.models.assignments import AssignmentEntityStatus

                # Try to get assignment_entity_status_id from request body
                assignment_entity_status_id = data.get('assignment_entity_status_id')
                if assignment_entity_status_id:
                    with suppress((ValueError, TypeError)):
                        aes = AssignmentEntityStatus.query.get(int(assignment_entity_status_id))
                        if aes and aes.country:
                            return aes.country.iso2 or aes.country.iso3

                # Try to get from referer URL
                referer = request.headers.get('Referer') or ''
                # Match /forms/assignment/<id> or /forms/entry/<id> patterns
                m = re.search(r"/forms/(?:assignment|entry)/(\d+)", referer)
                if m:
                    with suppress((ValueError, TypeError)):
                        aes_id = int(m.group(1))
                        aes = AssignmentEntityStatus.query.get(aes_id)
                        if aes and aes.country:
                            return aes.country.iso2 or aes.country.iso3

                # Try query parameter
                iso = request.args.get('iso') or request.args.get('country')
                if iso:
                    iso = iso.strip().upper()
                    country = Country.query.filter(or_(Country.iso3 == iso, Country.iso2 == iso)).first()
                    if country:
                        return country.iso2 or country.iso3

                return None
            except Exception as e:
                current_app.logger.debug("_detect_country_iso_from_matrix_context failed: %s", e)
                return None

        def _fetch_plugin_lookup_rows(list_id, config=None):
            """Reuse forms_api lookup endpoint logic for plugin/system lists."""
            try:
                from app.routes.forms_api import get_plugin_lookup_list_options  # Local import to avoid circular deps

                # For emergency operations, detect country ISO from matrix context
                country_iso = None
                if list_id == 'emergency_operations':
                    country_iso = _detect_country_iso_from_matrix_context()

                # Call plugin lookup function directly with country ISO and config
                plugin_response = get_plugin_lookup_list_options(list_id, country_iso=country_iso, config=config)

                # All json_* helpers return (Response, status_code) tuples; unpack accordingly
                if isinstance(plugin_response, tuple):
                    response_obj, status_code = plugin_response[0], plugin_response[1] if len(plugin_response) > 1 else 200
                else:
                    response_obj = plugin_response
                    status_code = getattr(plugin_response, 'status_code', 200)

                if status_code != 200:
                    return None, plugin_response
                payload = response_obj.get_json(silent=True) or {}
                rows = payload.get('rows') or payload.get('options') or []
                if not isinstance(rows, list):
                    rows = []
                return rows, None
            except Exception as plugin_exc:
                current_app.logger.error(
                    f"Error loading lookup list {list_id} for matrix search: {plugin_exc}",
                    exc_info=True
                )
                err_resp, _ = json_server_error('Failed to load lookup list options', success=False)
                return None, err_resp

        # Handle system lists (string IDs) vs regular lookup lists (integer IDs)
        is_system_list = not str(lookup_list_id).isdigit()

        if is_system_list:
            rows_data = []
            plugin_response = None

            if lookup_list_id in ('country_map', 'national_society', 'indicator_bank'):
                from app.models.organization import NationalSociety
                from app.models.indicator_bank import IndicatorBank
                from sqlalchemy import inspect

                def _get_model_columns_config(model_class):
                    inspector = inspect(model_class)
                    columns_config = []
                    for column in inspector.columns:
                        if column.name == 'id':
                            continue
                        columns_config.append({'name': column.name})
                    return columns_config

                def _model_to_dict(obj, columns_config):
                    data = {}
                    for col in columns_config:
                        col_name = col['name']
                        if hasattr(obj, col_name):
                            value = getattr(obj, col_name)
                            if value is None:
                                data[col_name] = ''
                            elif isinstance(value, dict):
                                data[col_name] = str(value) if value else ''
                            else:
                                data[col_name] = value
                        else:
                            data[col_name] = ''
                    return data

                current_locale = session.get('language', 'en')
                if not current_locale:
                    current_locale = str(get_locale()) if get_locale() else 'en'

                if lookup_list_id == 'country_map':
                    model_class = Country
                    query_results = model_class.query.order_by(model_class.name).all()
                elif lookup_list_id == 'national_society':
                    model_class = NationalSociety
                    query_results = model_class.query.options(
                        joinedload(NationalSociety.country)
                    ).order_by(model_class.name).all()
                else:
                    model_class = IndicatorBank
                    query_results = model_class.query.order_by(model_class.name).all()

                columns_config = _get_model_columns_config(model_class)

                for instance in query_results:
                    row_data = _model_to_dict(instance, columns_config)

                    if hasattr(instance, 'id'):
                        row_data['id'] = instance.id
                        row_data['_id'] = instance.id
                    elif 'id' in row_data:
                        row_data['_id'] = row_data['id']

                    if lookup_list_id == 'national_society' and hasattr(instance, 'country') and instance.country:
                        row_data['region'] = instance.country.region

                    if display_column == 'name':
                        if lookup_list_id == 'country_map':
                            row_data['name'] = get_localized_country_name(instance)
                        elif lookup_list_id == 'national_society':
                            localized_name = instance.get_name_translation(current_locale)
                            row_data['name'] = localized_name if localized_name and localized_name.strip() else instance.name

                    rows_data.append(row_data)
            else:
                rows_data, plugin_response = _fetch_plugin_lookup_rows(lookup_list_id, config=plugin_config)
                if plugin_response:
                    return plugin_response

            options = _build_options_from_rows(rows_data)
        else:
            # Handle regular lookup lists
            lookup_list = LookupList.query.get(int(lookup_list_id))
            if not lookup_list:
                return json_not_found('Lookup list not found')

            # Get all rows
            query = lookup_list.rows.order_by(LookupListRow.order)

            # Apply filters
            if filters:
                for filter_item in filters:
                    column = filter_item.get('column')
                    operator = filter_item.get('operator', 'equals')
                    value = filter_item.get('value')

                    if not column or not value:
                        continue

                    if operator == 'equals':
                        query = query.filter(LookupListRow.data[column].astext == value)
                    elif operator == 'not_equals':
                        query = query.filter(LookupListRow.data[column].astext != value)
                    elif operator == 'contains':
                        query = query.filter(LookupListRow.data[column].astext.ilike(safe_ilike_pattern(value)))
                    elif operator == 'not_contains':
                        query = query.filter(~LookupListRow.data[column].astext.ilike(safe_ilike_pattern(value)))

            # Get rows
            rows = query.all()

            # Extract options
            options = []
            for row in rows:
                if row.data and display_column in row.data:
                    row_value = str(row.data[display_column]).strip()

                    # Skip if empty or already exists
                    if not row_value or row_value in existing_rows:
                        continue

                    # Apply search term filter
                    if search_term and search_term.lower() not in row_value.lower():
                        continue

                    # Get additional description if available
                    description = None
                    # Try common description fields
                    for desc_field in ['description', 'desc', 'details', 'notes']:
                        if desc_field in row.data and row.data[desc_field]:
                            description = str(row.data[desc_field])
                            break

                    # Include the row ID for regular lookup lists
                    row_id = row.id if hasattr(row, 'id') else None
                    row_data = row.data if isinstance(row.data, dict) else {}
                    if row_id is not None:
                        row_data['_id'] = row_id

                    options.append({
                        'value': row_value,
                        'description': description,
                        'data': row_data,
                        'id': row_id  # Include ID at top level for easy access
                    })

        # Sort options by value
        options.sort(key=lambda x: x['value'].lower())

        # Limit results to prevent overwhelming UI
        # Ensure limit is a positive integer
        try:
            limit = int(limit)
            if limit < 1:
                limit = DEFAULT_LOOKUP_ROW_LIMIT
        except (ValueError, TypeError):
            limit = DEFAULT_LOOKUP_ROW_LIMIT

        if len(options) > limit:
            options = options[:limit]

        return json_ok(success=True, options=options, total=len(options))

    except Exception as e:
        current_app.logger.error(f"Error searching matrix rows: {e}", exc_info=True)
        return json_server_error('Could not search rows')
