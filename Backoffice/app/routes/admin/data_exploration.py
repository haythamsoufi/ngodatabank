# File: Backoffice/app/routes/admin/data_exploration.py
"""
Data Exploration Module - Explore form data with filters
"""

from functools import wraps
from flask import Blueprint, render_template, request, current_app, send_file, abort
from flask_login import current_user
from sqlalchemy import distinct, func, and_, or_, tuple_
from sqlalchemy.orm import joinedload
from typing import Any, Dict, List
from io import BytesIO
from app import db
from app.models import (
    FormTemplate, AssignedForm, Country, FormItem, FormData, AIFormDataValidation,
    AssignmentEntityStatus, SubmittedDocument, FormSection, FormPage
)
from app.utils.api_responses import json_bad_request, json_error, json_forbidden, json_not_found, json_ok, json_server_error
from app.utils.api_helpers import GENERIC_ERROR_MESSAGE, get_json_safe
from app.routes.admin.shared import admin_required, permission_required
from app.services.security.api_authentication import get_user_allowed_template_ids
from app.utils.datetime_helpers import utcnow
from app.services.authorization_service import AuthorizationService
from flask_babel import gettext as _
import json
import logging

logger = logging.getLogger(__name__)
bp = Blueprint("data_exploration", __name__, url_prefix="/admin")

# Data Explorer permission codes (granular per tab)
DATA_EXPLORER_PERMISSIONS = [
    'admin.data_explore.data_table',
    'admin.data_explore.analysis',
    'admin.data_explore.compliance',
]


def _ai_beta_denied_response():
    """Return a JSON error response when AI beta access is restricted for this user."""
    try:
        from app.services.app_settings_service import is_ai_beta_restricted, user_has_ai_beta_access

        if not is_ai_beta_restricted():
            return None
        if not getattr(current_user, "is_authenticated", False):
            return json_forbidden("AI beta access is limited to selected users.")
        if not user_has_ai_beta_access(current_user):
            return json_forbidden("AI beta access is limited to selected users.")
    except Exception as e:
        logger.debug("data_exploration AI beta gate check failed: %s", e, exc_info=True)
    return None


def has_any_data_explorer_permission(user) -> bool:
    """Check if user has any Data Explorer permission."""
    if AuthorizationService.is_system_manager(user):
        return True
    for perm in DATA_EXPLORER_PERMISSIONS:
        if AuthorizationService.has_rbac_permission(user, perm):
            return True
    return False


def data_explorer_required(f):
    """
    Decorator that requires at least one Data Explorer permission.
    Used for the main explore_data page.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            abort(401)
        if not has_any_data_explorer_permission(current_user):
            abort(403)
        return f(*args, **kwargs)
    # Metadata for startup-time guard auditing
    try:
        decorated_function._rbac_permissions_any_required = list(DATA_EXPLORER_PERMISSIONS)  # type: ignore[attr-defined]
        decorated_function._rbac_permissions_required = list(getattr(f, "_rbac_permissions_required", []) or [])  # type: ignore[attr-defined]
        decorated_function._rbac_admin_required = bool(getattr(f, "_rbac_admin_required", False))  # type: ignore[attr-defined]
        decorated_function._rbac_system_manager_required = bool(getattr(f, "_rbac_system_manager_required", False))  # type: ignore[attr-defined]
    except Exception as e:
        logger.debug("data_explorer_required: metadata assignment failed: %s", e)
    return decorated_function

# === Data Exploration Routes ===
@bp.route("/data-exploration", methods=["GET"])
@data_explorer_required
def explore_data():
    """Display data exploration page with filters for template and assignment."""
    try:
        # Determine which tabs the user can access
        can_access_data_table = AuthorizationService.is_system_manager(current_user) or \
            AuthorizationService.has_rbac_permission(current_user, 'admin.data_explore.data_table')
        can_access_analysis = AuthorizationService.is_system_manager(current_user) or \
            AuthorizationService.has_rbac_permission(current_user, 'admin.data_explore.analysis')
        can_access_compliance = AuthorizationService.is_system_manager(current_user) or \
            AuthorizationService.has_rbac_permission(current_user, 'admin.data_explore.compliance')
        # System managers can see all templates regardless of ownership and sharing
        # Use joinedload for published_version to avoid N+1 queries when accessing template.name
        if AuthorizationService.is_system_manager(current_user):
            templates = (
                FormTemplate.query
                .options(joinedload(FormTemplate.published_version))
                .all()
            )
        else:
            # Get templates that the user has access to (owned or shared)
            allowed_template_ids = get_user_allowed_template_ids(current_user.id)
            if allowed_template_ids:
                templates = (
                    FormTemplate.query
                    .filter(FormTemplate.id.in_(allowed_template_ids))
                    .options(joinedload(FormTemplate.published_version))
                    .all()
                )
            else:
                templates = []
        # Sort by name (from published version) in Python since it's a property
        # Note: published_version is already eager-loaded, avoiding N+1 queries
        templates.sort(key=lambda t: t.name if t.name else "")

        # Optimize period names query - use distinct on column directly
        # This is more efficient than selecting tuples and extracting
        period_names = (
            db.session.query(distinct(AssignedForm.period_name))
            .filter(AssignedForm.period_name.isnot(None))
            .order_by(AssignedForm.period_name.desc())
            .all()
        )
        # Extract from single-element tuples more efficiently
        period_names = [p[0] for p in period_names if p[0]]

        # Get all countries for filter dropdown
        countries = Country.query.order_by(Country.name).all()

        return render_template("admin/data_exploration/explore_data.html",
                             templates=templates,
                             period_names=period_names,
                             countries=countries,
                             can_access_data_table=can_access_data_table,
                             can_access_analysis=can_access_analysis,
                             can_access_compliance=can_access_compliance,
                             title=_("Explore Data"))
    except Exception as e:
        logger.error(f"Error loading data exploration page: {str(e)}", exc_info=True)
        # Return empty data rather than crashing
        return render_template("admin/data_exploration/explore_data.html",
                             templates=[],
                             period_names=[],
                             countries=[],
                             can_access_data_table=True,
                             can_access_analysis=True,
                             can_access_compliance=True,
                             title=_("Explore Data"),
                             error="Failed to load filter options. Please refresh the page.")

@bp.route("/data-exploration/form-items", methods=["GET"])
@permission_required('admin.data_explore.data_table')
def get_form_items_for_template():
    """Get form items for a specific template."""
    try:
        template_id = request.args.get('template_id', type=int)

        if not template_id:
            return json_error('template_id is required', 400)

        logger.info(
            "Data Explorer: form-items request template_id=%s user_id=%s",
            template_id,
            getattr(current_user, "id", None),
        )

        # System managers have access to all templates
        if not AuthorizationService.is_system_manager(current_user):
            # Validate template exists and user has access
            allowed_template_ids = get_user_allowed_template_ids(current_user.id)
            if template_id not in allowed_template_ids:
                return json_forbidden('Forbidden: no access to requested template')

        template = FormTemplate.query.get(template_id)
        if not template:
            return json_not_found('Template not found')

        # IMPORTANT: Only return items for the published version (or earliest version fallback).
        # Templates can have multiple versions, and FormItem.template_id is denormalized across versions.
        # Without scoping by version_id, the dropdown can show "duplicates" (same label, different ids)
        # coming from different versions.
        version_id = None
        try:
            if getattr(template, "published_version_id", None):
                version_id = int(template.published_version_id)
            else:
                first_version = template.versions.order_by('created_at').first()
                if first_version and getattr(first_version, "id", None):
                    version_id = int(first_version.id)
        except Exception as e:
            logger.debug("version_id extraction failed: %s", e)
            version_id = None

        logger.info(
            "Data Explorer: form-items template_id=%s resolved_version_id=%s",
            template_id,
            version_id,
        )

        # Get form items for the template (load section, parent section, and page for full display order)
        q = (
            FormItem.query
            .options(
                joinedload(FormItem.form_section).options(
                    joinedload(FormSection.parent_section),
                    joinedload(FormSection.page),
                )
            )
            .filter_by(template_id=template_id, archived=False)
        )
        if version_id:
            q = q.filter(FormItem.version_id == int(version_id))
        form_items = q.order_by(FormItem.order).all()
        # Sort by template display order: page -> section hierarchy (parent then section) -> item order
        def _template_order_key(item):
            sec = item.form_section
            if not sec:
                return (0, 0, 0, float(item.order or 0))
            page = getattr(sec, 'page', None)
            page_order = int(page.order or 0) if page else 0
            parent = sec.parent_section
            parent_order = float((parent.order if parent else 0) or 0)
            sec_order = float(sec.order or 0)
            item_order = float(item.order or 0)
            return (page_order, parent_order, sec_order, item_order)
        form_items = sorted(form_items, key=_template_order_key)

        # Debug duplication signals (same ID should never repeat; labels may repeat depending on data)
        try:
            ids = [int(i.id) for i in form_items if getattr(i, "id", None) is not None]
            dup_ids = {v for v in ids if ids.count(v) > 1}
            if dup_ids:
                logger.warning(
                    "Data Explorer: duplicate FormItem IDs in query result template_id=%s dup_ids=%s",
                    template_id,
                    sorted(list(dup_ids))[:50],
                )

            labels = [str(i.label or "").strip() for i in form_items]
            dup_labels = sorted({v for v in labels if v and labels.count(v) > 1})[:25]
            if dup_labels:
                logger.info(
                    "Data Explorer: duplicate FormItem labels template_id=%s examples=%s",
                    template_id,
                    dup_labels,
                )
            logger.info(
                "Data Explorer: form-items response template_id=%s count=%s",
                template_id,
                len(form_items),
            )
        except Exception as e:
            logger.warning("Data Explorer: form-items debug logging failed: %s", e, exc_info=True)

        # Serialize form items with section/subsection, page, and order for dropdown (preserve template order)
        items_data = []
        for item in form_items:
            sec = item.form_section
            section_name = (sec.name if sec else None)
            parent_sec = (sec.parent_section if sec else None)
            parent_section_name = (parent_sec.name if parent_sec else None)
            parent_section_order = float(parent_sec.order or 0) if parent_sec else 0
            section_order = float(sec.order or 0) if sec else 0
            page = getattr(sec, 'page', None) if sec else None
            page_order = int(page.order or 0) if page else 0
            items_data.append({
                'item_id': item.id,
                'label': item.label,
                'item_type': item.item_type,
                'order': item.order,
                'section_id': item.section_id,
                'section_name': section_name,
                'parent_section_name': parent_section_name,
                'parent_section_order': parent_section_order,
                'section_order': section_order,
                'page_order': page_order,
            })

        return json_ok(form_items=items_data)
    except Exception as e:
        logger.error(f"Error loading form items for template {template_id}: {str(e)}", exc_info=True)
        return json_server_error('Failed to load form items')


@bp.route("/data-exploration/assignment-filters", methods=["GET"])
@permission_required('admin.data_explore.data_table')
def get_assignment_filters_for_template():
    """Get assignment periods and countries for a specific template (for filter dropdowns)."""
    try:
        template_id = request.args.get('template_id', type=int)

        if not template_id:
            return json_error('template_id is required', 400)

        if not AuthorizationService.is_system_manager(current_user):
            allowed_template_ids = get_user_allowed_template_ids(current_user.id)
            if template_id not in allowed_template_ids:
                return json_forbidden('Forbidden: no access to requested template')

        template = FormTemplate.query.get(template_id)
        if not template:
            return json_not_found('Template not found')

        # Distinct period names for this template's assignments
        period_names = (
            db.session.query(distinct(AssignedForm.period_name))
            .filter(AssignedForm.template_id == template_id, AssignedForm.period_name.isnot(None))
            .order_by(AssignedForm.period_name.desc())
            .all()
        )
        period_names = [p[0] for p in period_names if p[0]]

        # Distinct countries for this template (via AssignmentEntityStatus where entity_type='country')
        countries_query = (
            db.session.query(Country)
            .join(AssignmentEntityStatus, and_(
                AssignmentEntityStatus.entity_id == Country.id,
                AssignmentEntityStatus.entity_type == 'country'
            ))
            .join(AssignedForm, AssignedForm.id == AssignmentEntityStatus.assigned_form_id)
            .filter(AssignedForm.template_id == template_id)
            .distinct()
            .order_by(Country.name)
        )
        countries = countries_query.all()
        countries_data = [{'id': c.id, 'name': c.name} for c in countries]

        return json_ok(period_names=period_names, countries=countries_data)
    except Exception as e:
        logger.error(f"Error loading assignment filters for template {template_id}: {str(e)}", exc_info=True)
        return json_server_error('Failed to load assignment filters')


def _parse_ai_opinion_ids(raw_ids: str | List[str]) -> tuple[List[int], List[tuple[int, int]], List[str]]:
    """
    Parse a comma-separated string or list of row ids into form_data_ids, missing_pairs, missing_keys.
    Ids can be numeric (FormData id) or virtual (m:<aes_id>:<form_item_id>).
    Returns (form_data_ids, missing_pairs, missing_keys) with deduplication applied.
    """
    if not raw_ids:
        return [], [], []
    if isinstance(raw_ids, list):
        parts = [str(p).strip() for p in raw_ids if p is not None and str(p).strip()]
    else:
        parts = [p.strip() for p in str(raw_ids).split(",") if (p or "").strip()]

    form_data_ids: List[int] = []
    missing_pairs: List[tuple[int, int]] = []
    missing_keys: List[str] = []
    for part in parts:
        if not part:
            continue
        if part.startswith("m:"):
            try:
                _p = part.split(":")
                if len(_p) == 3:
                    aes_id = int(_p[1])
                    fi_id = int(_p[2])
                    if aes_id > 0 and fi_id > 0:
                        missing_pairs.append((aes_id, fi_id))
                        missing_keys.append(part)
                        continue
            except Exception as e:
                logger.debug("_parse_ai_opinion_ids: part parse failed for %r: %s", part, e)
        try:
            v = int(part)
            if v > 0:
                form_data_ids.append(v)
        except Exception as e:
            logger.debug("form_data_id int parse failed: %s", e)
            continue

    form_data_ids = list(dict.fromkeys(form_data_ids))
    seen_pairs = set()
    dedup_pairs: List[tuple[int, int]] = []
    dedup_keys: List[str] = []
    for k, pair in zip(missing_keys, missing_pairs):
        if pair in seen_pairs:
            continue
        seen_pairs.add(pair)
        dedup_pairs.append(pair)
        dedup_keys.append(k)
    return form_data_ids, dedup_pairs, dedup_keys


@bp.route("/data-exploration/ai-opinions", methods=["GET", "POST"])
@permission_required('admin.data_explore.data_table')
def get_ai_opinions_for_rows():
    """
    Fetch AI validation opinions for a set of row ids (single bulk query, no N+1).
    GET: query param ids=1,2,3,m:398:915
    POST: body { "ids": ["1", "2", "m:398:915"] } — use for large sets to avoid URL length limits.
    """
    try:
        denied = _ai_beta_denied_response()
        if denied is not None:
            return denied

        if request.method == "POST":
            payload = get_json_safe()
            raw_ids = payload.get("ids")
            if isinstance(raw_ids, list):
                form_data_ids, missing_pairs, missing_keys = _parse_ai_opinion_ids(raw_ids)
            else:
                form_data_ids, missing_pairs, missing_keys = _parse_ai_opinion_ids(
                    str(raw_ids) if raw_ids is not None else ""
                )
        else:
            raw_ids = (request.args.get("ids") or "").strip()
            form_data_ids, missing_pairs, missing_keys = _parse_ai_opinion_ids(raw_ids)

        if not form_data_ids and not missing_pairs:
            return json_ok(opinionsByFormDataId={})

        opinions: List[AIFormDataValidation] = []
        if form_data_ids:
            opinions.extend(AIFormDataValidation.query.filter(AIFormDataValidation.form_data_id.in_(form_data_ids)).all())
        if missing_pairs:
            # latest-only is enforced via unique constraint (aes_id, form_item_id)
            q = (
                AIFormDataValidation.query
                .filter(AIFormDataValidation.form_data_id.is_(None))
                .filter(
                    tuple_(
                        AIFormDataValidation.assignment_entity_status_id,
                        AIFormDataValidation.form_item_id,
                    ).in_(missing_pairs)
                )
            )
            opinions.extend(q.all())

        by_form_data_id = {int(o.form_data_id): o for o in opinions if getattr(o, "form_data_id", None)}
        by_missing_pair = {
            (int(o.assignment_entity_status_id), int(o.form_item_id)): o
            for o in opinions
            if (getattr(o, "form_data_id", None) is None)
            and getattr(o, "assignment_entity_status_id", None) is not None
            and getattr(o, "form_item_id", None) is not None
        }

        def _serialize(o: AIFormDataValidation) -> dict:
            suggestion = None
            opinion_ui = None
            try:
                suggestion = (o.evidence or {}).get("suggestion") if isinstance(o.evidence, dict) else None
            except Exception as e:
                logger.debug("suggestion from evidence failed: %s", e)
                suggestion = None
            try:
                opinion_ui = (o.evidence or {}).get("opinion_ui") if isinstance(o.evidence, dict) else None
            except Exception as e:
                logger.debug("opinion_ui from evidence failed: %s", e)
                opinion_ui = None
            return {
                "id": int(o.id),
                "form_data_id": int(o.form_data_id) if o.form_data_id else None,
                "status": o.status,
                "verdict": o.verdict,
                "confidence": o.confidence,
                "opinion_text": o.opinion_text,
                "opinion_summary": (opinion_ui or {}).get("summary") if isinstance(opinion_ui, dict) else o.opinion_text,
                "opinion_details": (opinion_ui or {}).get("details") if isinstance(opinion_ui, dict) else None,
                "opinion_sources": (opinion_ui or {}).get("sources") if isinstance(opinion_ui, dict) else None,
                "opinion_basis": (opinion_ui or {}).get("basis") if isinstance(opinion_ui, dict) else None,
                "decision": (opinion_ui or {}).get("decision") if isinstance(opinion_ui, dict) else None,
                "provider": o.provider,
                "model": o.model,
                "updated_at": o.updated_at.isoformat() if o.updated_at else None,
                "evidence": o.evidence,
                "suggestion": suggestion,
            }

        out: Dict[str, Any] = {}
        for fid in form_data_ids:
            o = by_form_data_id.get(int(fid))
            out[str(fid)] = _serialize(o) if o else None
        for key, pair in zip(missing_keys, missing_pairs):
            o = by_missing_pair.get(pair)
            out[str(key)] = _serialize(o) if o else None

        return json_ok(opinionsByFormDataId=out)
    except Exception as e:
        logger.error("Error fetching AI opinions: %s", e, exc_info=True)
        return json_server_error(GENERIC_ERROR_MESSAGE)


@bp.route("/data-exploration/ai-validate", methods=["POST"])
@permission_required('admin.data_explore.data_table')
def run_ai_validation_for_rows():
    """
    Run AI validation for selected FormData rows.
    Body: { form_data_ids: [1,2,3] }
    """
    try:
        denied = _ai_beta_denied_response()
        if denied is not None:
            return denied

        payload = get_json_safe()
        sources = payload.get("sources", None)  # optional: ['historical','system_documents','upr_documents']

        # New shape (preferred): rows=[{row_id, form_data_id? or submission_id+form_item_id?}]
        rows = payload.get("rows")
        if rows is not None:
            if not isinstance(rows, list) or not rows:
                return json_bad_request("rows is required")
        else:
            # Legacy: form_data_ids=[1,2,3]
            ids = payload.get("form_data_ids") or []
            if not isinstance(ids, list) or not ids:
                return json_bad_request("rows (or form_data_ids) is required")
            rows = [{"row_id": v, "form_data_id": v} for v in ids]

        from app.services.ai_formdata_validation_service import AIFormDataValidationService

        svc = AIFormDataValidationService()
        results: Dict[str, Any] = {}

        def _serialize_rec(rec: AIFormDataValidation) -> Dict[str, Any]:
            suggestion = None
            opinion_ui = None
            try:
                suggestion = (rec.evidence or {}).get("suggestion") if isinstance(rec.evidence, dict) else None
            except Exception as e:
                logger.debug("suggestion extract failed: %s", e)
                suggestion = None
            try:
                opinion_ui = (rec.evidence or {}).get("opinion_ui") if isinstance(rec.evidence, dict) else None
            except Exception as e:
                logger.debug("opinion_ui extract failed: %s", e)
                opinion_ui = None
            return {
                "id": int(rec.id),
                "form_data_id": int(rec.form_data_id) if rec.form_data_id else None,
                "status": rec.status,
                "verdict": rec.verdict,
                "confidence": rec.confidence,
                "opinion_text": rec.opinion_text,
                "opinion_summary": (opinion_ui or {}).get("summary") if isinstance(opinion_ui, dict) else rec.opinion_text,
                "opinion_details": (opinion_ui or {}).get("details") if isinstance(opinion_ui, dict) else None,
                "opinion_sources": (opinion_ui or {}).get("sources") if isinstance(opinion_ui, dict) else None,
                "opinion_basis": (opinion_ui or {}).get("basis") if isinstance(opinion_ui, dict) else None,
                "decision": (opinion_ui or {}).get("decision") if isinstance(opinion_ui, dict) else None,
                "provider": rec.provider,
                "model": rec.model,
                "updated_at": rec.updated_at.isoformat() if rec.updated_at else None,
                "evidence": rec.evidence,
                "suggestion": suggestion,
            }

        for row in rows:
            try:
                row_id = row.get("row_id") if isinstance(row, dict) else None
                form_data_id = row.get("form_data_id") if isinstance(row, dict) else None

                fd_id_int = None
                try:
                    if form_data_id is not None:
                        fd_id_int = int(form_data_id)
                except Exception as e:
                    logger.debug("fd_id_int parse failed: %s", e)
                    fd_id_int = None

                if not fd_id_int:
                    # Missing/virtual row: do NOT create placeholder FormData rows.
                    # Only validate rows that already have a persisted FormData id.
                    submission_id = row.get("submission_id") if isinstance(row, dict) else None
                    form_item_id = row.get("form_item_id") if isinstance(row, dict) else None
                    if not submission_id or not form_item_id:
                        raise ValueError("Row is missing form_data_id and (submission_id, form_item_id)")
                    aes_id = int(submission_id)
                    fi_id = int(form_item_id)

                    existing_fd = (
                        FormData.query
                        .filter(FormData.assignment_entity_status_id == aes_id, FormData.form_item_id == fi_id)
                        .first()
                    )
                    if existing_fd:
                        fd_id_int = int(existing_fd.id)
                    else:
                        # Non-reported item: run suggestion-only validation WITHOUT creating FormData,
                        # but persist the opinion keyed by (assignment_entity_status_id, form_item_id).
                        row_key = str(row_id if row_id is not None else f"m:{aes_id}:{fi_id}")
                        rec, _vr = svc.upsert_missing_assigned_validation(
                            assignment_entity_status_id=int(aes_id),
                            form_item_id=int(fi_id),
                            run_by_user_id=int(current_user.id),
                            sources=sources,
                        )
                        payload = _serialize_rec(rec)
                        payload["row_id"] = row_key
                        results[row_key] = payload
                        continue

                rec, _ = svc.upsert_validation(
                    form_data_id=int(fd_id_int),
                    run_by_user_id=int(current_user.id),
                    sources=sources,
                )
                results[str(row_id if row_id is not None else fd_id_int)] = _serialize_rec(rec)
            except Exception as e:
                logger.warning("AI validation failed for row %s: %s", row, e, exc_info=True)
                key = None
                try:
                    key = str(row.get("row_id")) if isinstance(row, dict) and row.get("row_id") is not None else None
                except Exception as e:
                    logger.debug("row_id key extract failed: %s", e)
                    key = None
                results[str(key or "unknown")] = {
                    "status": "failed",
                    "verdict": "uncertain",
                    "opinion_text": f"Validation failed: {e}",
                }

        return json_ok(resultsByRowId=results, resultsByFormDataId=results)
    except Exception as e:
        logger.error("Error running AI validation: %s", e, exc_info=True)
        return json_server_error(GENERIC_ERROR_MESSAGE)


@bp.route("/data-exploration/apply-imputed-value", methods=["POST"])
@permission_required('admin.data_explore.data_table')
def apply_imputed_value():
    """
    Apply an accepted AI-suggested value into FormData.imputed_value.
    Body:
      - { form_data_id: 123, imputed_value: <any JSON-serializable> }
      - { form_data_id: 123, imputed_disagg_data: <any JSON-serializable> }
      - { submission_id: 398, form_item_id: 915, imputed_value: ..., ... }  # creates FormData row if missing
      - or both
    """
    try:
        denied = _ai_beta_denied_response()
        if denied is not None:
            return denied

        payload = get_json_safe()
        form_data_id = payload.get("form_data_id")
        submission_id = payload.get("submission_id")  # AES id
        form_item_id = payload.get("form_item_id")

        created = False

        # Resolve/validate target: either an existing FormData id, or an (AES, FormItem) pair.
        fd = None
        if form_data_id is not None and str(form_data_id).strip() != "":
            try:
                form_data_id = int(form_data_id)
            except Exception as e:
                logger.debug("form_data_id int parse failed: %s", e)
                return json_bad_request("form_data_id must be an integer")
            if form_data_id <= 0:
                return json_bad_request("form_data_id must be positive")
            fd = FormData.query.get(int(form_data_id))
            if not fd:
                return json_not_found("FormData not found")
        else:
            try:
                submission_id = int(submission_id) if submission_id is not None else None
                form_item_id = int(form_item_id) if form_item_id is not None else None
            except Exception as e:
                logger.debug("submission_id/form_item_id parse failed: %s", e)
                return json_bad_request("submission_id and form_item_id must be integers")
            if not submission_id or not form_item_id:
                return json_bad_request("form_data_id or (submission_id and form_item_id) is required")

            fd = (
                FormData.query
                .filter(FormData.assignment_entity_status_id == int(submission_id), FormData.form_item_id == int(form_item_id))
                .first()
            )
            if not fd:
                # User-intent action: create a FormData row now (this is NOT a placeholder).
                fd = FormData(
                    assignment_entity_status_id=int(submission_id),
                    form_item_id=int(form_item_id),
                )
                # Ensure a clean "empty reported value" baseline.
                fd.value = None
                fd.disagg_data = db.null()
                fd.prefilled_value = db.null()
                fd.prefilled_disagg_data = db.null()
                fd.imputed_value = db.null()
                fd.imputed_disagg_data = db.null()
                fd.data_not_available = False
                fd.not_applicable = False
                db.session.add(fd)
                db.session.flush()
                created = True

        imputed_value = payload.get("imputed_value", None)
        imputed_disagg_data = payload.get("imputed_disagg_data", None)

        # Require at least one of the two payloads
        has_scalar = not (imputed_value is None or (isinstance(imputed_value, str) and not imputed_value.strip()))
        has_disagg = not (imputed_disagg_data is None or (isinstance(imputed_disagg_data, str) and not imputed_disagg_data.strip()))
        if not has_scalar and not has_disagg:
            return json_bad_request("imputed_value or imputed_disagg_data is required")

        # Normalize common quoted-string artifacts from the frontend.
        # For single-choice questions, the stored value should be CHF (not "CHF").
        try:
            if isinstance(imputed_value, str):
                s = imputed_value.strip()
                # Treat literal "null" string as empty/missing
                if s.lower() == "null":
                    imputed_value = None

                # If the value itself looks like a JSON-encoded string (e.g. '"CHF"'),
                # decode it so we store CHF without extra quotes.
                if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
                    try:
                        decoded = json.loads(s)
                        if isinstance(decoded, str):
                            imputed_value = decoded
                    except Exception as e:
                        logger.debug("imputed_value decode failed: %s", e)
                        imputed_value = s[1:-1]

                # Also tolerate single-quoted values (rare)
                s2 = str(imputed_value).strip()
                if len(s2) >= 2 and s2[0] == "'" and s2[-1] == "'":
                    imputed_value = s2[1:-1]
        except Exception as e:
            logger.debug("imputed_value parse failed: %s", e)

        # Normalize imputed_disagg_data if sent as a JSON string
        try:
            if isinstance(imputed_disagg_data, str):
                s = imputed_disagg_data.strip()
                if s.lower() == "null" or s == "":
                    imputed_disagg_data = None
                elif (s.startswith("{") and s.endswith("}")) or (s.startswith("[") and s.endswith("]")):
                    imputed_disagg_data = json.loads(s)
        except Exception as e:
            logger.debug("imputed_disagg_data json parse failed: %s", e)

        # Only set fields that were actually provided (allows scalar-only or disagg-only applies).
        if "imputed_value" in payload:
            fd.imputed_value = imputed_value
        if "imputed_disagg_data" in payload:
            fd.imputed_disagg_data = imputed_disagg_data

        # If the user is imputing a value for a previously "missing" row (no FormData existed when AI ran),
        # we likely already have an AIFormDataValidation record keyed by (assignment_entity_status_id, form_item_id).
        # Once we create/resolve the persisted FormData row, link that opinion to the new form_data_id so future
        # lookups by FormData id keep showing the same opinion/suggestion.
        try:
            aes_id = None
            fi_id = None
            if getattr(fd, "assignment_entity_status_id", None) and getattr(fd, "form_item_id", None):
                aes_id = int(fd.assignment_entity_status_id)
                fi_id = int(fd.form_item_id)
            if aes_id and fi_id and getattr(fd, "id", None):
                existing_for_fd = AIFormDataValidation.query.filter_by(form_data_id=int(fd.id)).first()
                missing_rec = (
                    AIFormDataValidation.query
                    .filter(AIFormDataValidation.form_data_id.is_(None))
                    .filter(AIFormDataValidation.assignment_entity_status_id == int(aes_id))
                    .filter(AIFormDataValidation.form_item_id == int(fi_id))
                    .first()
                )
                if missing_rec:
                    if existing_for_fd:
                        # If both exist, keep the FormData-linked one and remove the virtual duplicate.
                        db.session.delete(missing_rec)
                    else:
                        # Migrate the virtual opinion to be keyed by persisted FormData id.
                        missing_rec.form_data_id = int(fd.id)
                        missing_rec.assignment_entity_status_id = None
                        missing_rec.form_item_id = None
                        missing_rec.updated_at = utcnow()
        except Exception as e:
            logger.debug("opinion linking failed (non-fatal): %s", e)
        db.session.commit()

        return json_ok(
            success=True,
            form_data_id=int(fd.id),
            created=bool(created),
            imputed_value=fd.imputed_value,
            imputed_disagg_data=getattr(fd, "imputed_disagg_data", None),
        )
    except Exception as e:
        logger.error("Error applying imputed value: %s", e, exc_info=True)
        return json_server_error(GENERIC_ERROR_MESSAGE)


# FDRS Template ID for compliance check
FDRS_TEMPLATE_ID = 21

# Document types for compliance check (only the required documents)
COMPLIANCE_DOC_TYPES = [
    "Annual Report",
    "Audited Financial Statement"
]


@bp.route("/data-exploration/compliance", methods=["GET"])
@permission_required('admin.data_explore.compliance')
def get_compliance_data():
    """
    Get compliance data for FDRS template (ID 21).

    Query params:
    - reference_year: The most recent year to measure from (e.g., 2024 means check 2024, 2023, 2022)

    Returns document upload status for each country across the last 3 periods,
    and determines compliance based on:
    - At least one Annual Report in the past 3 years
    - At least one Audited Financial Statement in the past 3 years
    """
    try:
        # Get reference year from query params (optional)
        reference_year = request.args.get('reference_year', type=int)

        # Get all available periods for FDRS template
        all_periods_query = (
            db.session.query(distinct(AssignedForm.period_name))
            .filter(
                AssignedForm.template_id == FDRS_TEMPLATE_ID,
                AssignedForm.period_name.isnot(None)
            )
            .order_by(AssignedForm.period_name.desc())
            .all()
        )
        all_periods = [p[0] for p in all_periods_query if p[0]]

        # Extract available years from period names (assuming format like "2024" or "2024-2025")
        available_years = set()
        for period in all_periods:
            # Try to extract year from period name
            import re
            years_in_period = re.findall(r'20\d{2}', str(period))
            for y in years_in_period:
                available_years.add(int(y))
        available_years = sorted(available_years, reverse=True)

        # If reference_year is provided, filter periods to those containing that year or earlier (up to 3 years back)
        if reference_year:
            target_years = [reference_year, reference_year - 1, reference_year - 2]
            periods = []
            for period in all_periods:
                # Check if any target year is in this period
                years_in_period = re.findall(r'20\d{2}', str(period))
                for y in years_in_period:
                    if int(y) in target_years:
                        periods.append(period)
                        break
                if len(periods) >= 3:
                    break
        else:
            # Default: get the last 3 periods
            periods = all_periods[:3]

        if not periods:
            return json_ok(
                success=True,
                data=[],
                periods=[],
                available_years=available_years,
                reference_year=reference_year,
                doc_types=COMPLIANCE_DOC_TYPES,
                message="No FDRS assignment periods found",
            )

        # Get all countries
        countries = Country.query.order_by(Country.name).all()

        # Get document fields from FDRS template that match our compliance doc types
        doc_items = (
            FormItem.query
            .filter(
                FormItem.template_id == FDRS_TEMPLATE_ID,
                FormItem.item_type == 'document_field',
                FormItem.archived == False
            )
            .all()
        )

        # Map document field labels to their IDs
        doc_item_map = {}
        all_doc_item_ids = []
        for item in doc_items:
            label = item.label.strip() if item.label else ""
            for doc_type in COMPLIANCE_DOC_TYPES:
                if doc_type.lower() in label.lower():
                    if doc_type not in doc_item_map:
                        doc_item_map[doc_type] = []
                    doc_item_map[doc_type].append(item.id)
                    all_doc_item_ids.append(item.id)
                    break

        # Create reverse mapping: item_id -> doc_type
        item_id_to_doc_type = {}
        for doc_type, item_ids in doc_item_map.items():
            for item_id in item_ids:
                item_id_to_doc_type[item_id] = doc_type

        # ========== OPTIMIZED BULK QUERIES ==========
        # Instead of N+1 queries, we fetch everything in just 4 queries total

        # 1. Get all assignments for the selected periods (1 query)
        assignments = (
            AssignedForm.query
            .filter(
                AssignedForm.template_id == FDRS_TEMPLATE_ID,
                AssignedForm.period_name.in_(periods)
            )
            .all()
        )
        assignment_map = {a.period_name: a for a in assignments}
        assignment_ids = [a.id for a in assignments]

        # 2. Get all AssignmentEntityStatus records for these assignments (1 query)
        all_aes = []
        if assignment_ids:
            all_aes = (
                AssignmentEntityStatus.query
                .filter(
                    AssignmentEntityStatus.assigned_form_id.in_(assignment_ids),
                    AssignmentEntityStatus.entity_type == 'country'
                )
                .all()
            )

        # Build lookup: (assignment_id, country_id) -> aes
        aes_lookup = {}
        aes_ids = []
        for aes in all_aes:
            aes_lookup[(aes.assigned_form_id, aes.entity_id)] = aes
            aes_ids.append(aes.id)

        # 3. Get all SubmittedDocuments for these AES records and doc item types (1 query)
        submitted_docs = []
        if aes_ids and all_doc_item_ids:
            submitted_docs = (
                SubmittedDocument.query
                .filter(
                    SubmittedDocument.assignment_entity_status_id.in_(aes_ids),
                    SubmittedDocument.form_item_id.in_(all_doc_item_ids)
                )
                .all()
            )

        # Build lookup: (aes_id, doc_type) -> True
        doc_lookup = {}
        for doc in submitted_docs:
            doc_type = item_id_to_doc_type.get(doc.form_item_id)
            if doc_type:
                doc_lookup[(doc.assignment_entity_status_id, doc_type)] = True

        # ========== PROCESS DATA IN MEMORY ==========
        compliance_data = []

        for country in countries:
            country_periods = []
            has_annual_report = False
            has_audited_financial = False

            for period in periods:
                period_docs = {doc_type: False for doc_type in COMPLIANCE_DOC_TYPES}

                assignment = assignment_map.get(period)
                if assignment:
                    aes = aes_lookup.get((assignment.id, country.id))
                    if aes:
                        # Check for each document type using the lookup
                        for doc_type in COMPLIANCE_DOC_TYPES:
                            if doc_lookup.get((aes.id, doc_type)):
                                period_docs[doc_type] = True
                                if doc_type == "Annual Report":
                                    has_annual_report = True
                                elif doc_type == "Audited Financial Statement":
                                    has_audited_financial = True

                country_periods.append({
                    "period": period,
                    "documents": period_docs
                })

            # Determine compliance: must have at least 1 Annual Report AND 1 Audited Financial Statement
            is_compliant = has_annual_report and has_audited_financial

            compliance_data.append({
                "country_id": country.id,
                "country_name": country.name,
                "country_iso3": country.iso3,
                "region": country.region,
                "periods": country_periods,
                "is_compliant": is_compliant,
                "has_annual_report": has_annual_report,
                "has_audited_financial": has_audited_financial
            })

        return json_ok(
            success=True,
            data=compliance_data,
            periods=periods,
            available_years=available_years,
            reference_year=reference_year,
            doc_types=COMPLIANCE_DOC_TYPES,
        )

    except Exception as e:
        logger.error("Error fetching compliance data: %s", e, exc_info=True)
        return json_server_error(GENERIC_ERROR_MESSAGE)


@bp.route("/data-exploration/compliance/download", methods=["GET"])
@permission_required('admin.data_explore.compliance')
def download_compliance_excel():
    """
    Download compliance data as an Excel file.

    Query params:
    - reference_year: The most recent year to measure from (e.g., 2024 means check 2024, 2023, 2022)
    """
    try:
        import openpyxl
        import re
        from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime

        # Get reference year from query params (optional)
        reference_year = request.args.get('reference_year', type=int)

        # Get all available periods for FDRS template
        all_periods_query = (
            db.session.query(distinct(AssignedForm.period_name))
            .filter(
                AssignedForm.template_id == FDRS_TEMPLATE_ID,
                AssignedForm.period_name.isnot(None)
            )
            .order_by(AssignedForm.period_name.desc())
            .all()
        )
        all_periods = [p[0] for p in all_periods_query if p[0]]

        # If reference_year is provided, filter periods to those containing that year or earlier (up to 3 years back)
        if reference_year:
            target_years = [reference_year, reference_year - 1, reference_year - 2]
            periods = []
            for period in all_periods:
                # Check if any target year is in this period
                years_in_period = re.findall(r'20\d{2}', str(period))
                for y in years_in_period:
                    if int(y) in target_years:
                        periods.append(period)
                        break
                if len(periods) >= 3:
                    break
        else:
            # Default: get the last 3 periods
            periods = all_periods[:3]

        # Get all countries
        countries = Country.query.order_by(Country.name).all()

        # Get document fields from FDRS template
        doc_items = (
            FormItem.query
            .filter(
                FormItem.template_id == FDRS_TEMPLATE_ID,
                FormItem.item_type == 'document_field',
                FormItem.archived == False
            )
            .all()
        )

        # Map document field labels to their IDs
        doc_item_map = {}
        all_doc_item_ids = []
        for item in doc_items:
            label = item.label.strip() if item.label else ""
            for doc_type in COMPLIANCE_DOC_TYPES:
                if doc_type.lower() in label.lower():
                    if doc_type not in doc_item_map:
                        doc_item_map[doc_type] = []
                    doc_item_map[doc_type].append(item.id)
                    all_doc_item_ids.append(item.id)
                    break

        # Create reverse mapping: item_id -> doc_type
        item_id_to_doc_type = {}
        for doc_type, item_ids in doc_item_map.items():
            for item_id in item_ids:
                item_id_to_doc_type[item_id] = doc_type

        # ========== OPTIMIZED BULK QUERIES ==========
        # Get all assignments for the selected periods (1 query)
        assignments = (
            AssignedForm.query
            .filter(
                AssignedForm.template_id == FDRS_TEMPLATE_ID,
                AssignedForm.period_name.in_(periods)
            )
            .all()
        )
        assignment_map = {a.period_name: a for a in assignments}
        assignment_ids = [a.id for a in assignments]

        # Get all AssignmentEntityStatus records (1 query)
        all_aes = []
        if assignment_ids:
            all_aes = (
                AssignmentEntityStatus.query
                .filter(
                    AssignmentEntityStatus.assigned_form_id.in_(assignment_ids),
                    AssignmentEntityStatus.entity_type == 'country'
                )
                .all()
            )

        aes_lookup = {}
        aes_ids = []
        for aes in all_aes:
            aes_lookup[(aes.assigned_form_id, aes.entity_id)] = aes
            aes_ids.append(aes.id)

        # Get all SubmittedDocuments (1 query)
        submitted_docs = []
        if aes_ids and all_doc_item_ids:
            submitted_docs = (
                SubmittedDocument.query
                .filter(
                    SubmittedDocument.assignment_entity_status_id.in_(aes_ids),
                    SubmittedDocument.form_item_id.in_(all_doc_item_ids)
                )
                .all()
            )

        doc_lookup = {}
        for doc in submitted_docs:
            doc_type = item_id_to_doc_type.get(doc.form_item_id)
            if doc_type:
                doc_lookup[(doc.assignment_entity_status_id, doc_type)] = True

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FDRS Compliance"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        compliant_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        non_compliant_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        yes_font = Font(color="059669")
        no_font = Font(color="DC2626")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        # Build headers dynamically based on periods
        # Row 1: Country | Annual Report (merged) | Audited Financial Statement (merged) | Compliance Status
        # Row 2: (empty) | period1 | period2 | period3 | period1 | period2 | period3 | (empty)

        num_periods = len(periods)

        # First header row
        ws.cell(row=1, column=1, value="Country").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = center_align
        ws.cell(row=1, column=1).border = thin_border
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        # Annual Report header (spans num_periods columns)
        ar_start_col = 2
        ar_end_col = ar_start_col + num_periods - 1
        ws.cell(row=1, column=ar_start_col, value="Annual Report").font = header_font
        ws.cell(row=1, column=ar_start_col).fill = header_fill
        ws.cell(row=1, column=ar_start_col).alignment = center_align
        ws.cell(row=1, column=ar_start_col).border = thin_border
        if num_periods > 1:
            ws.merge_cells(start_row=1, start_column=ar_start_col, end_row=1, end_column=ar_end_col)

        # Audited Financial Statement header (spans num_periods columns)
        afs_start_col = ar_end_col + 1
        afs_end_col = afs_start_col + num_periods - 1
        ws.cell(row=1, column=afs_start_col, value="Audited Financial Statement").font = header_font
        ws.cell(row=1, column=afs_start_col).fill = header_fill
        ws.cell(row=1, column=afs_start_col).alignment = center_align
        ws.cell(row=1, column=afs_start_col).border = thin_border
        if num_periods > 1:
            ws.merge_cells(start_row=1, start_column=afs_start_col, end_row=1, end_column=afs_end_col)

        # Compliance Status header
        compliance_col = afs_end_col + 1
        ws.cell(row=1, column=compliance_col, value="Compliance Status").font = header_font
        ws.cell(row=1, column=compliance_col).fill = header_fill
        ws.cell(row=1, column=compliance_col).alignment = center_align
        ws.cell(row=1, column=compliance_col).border = thin_border
        ws.merge_cells(start_row=1, start_column=compliance_col, end_row=2, end_column=compliance_col)

        # Second header row with period names
        for idx, period in enumerate(periods):
            # Annual Report period column
            cell = ws.cell(row=2, column=ar_start_col + idx, value=period)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

            # Audited Financial Statement period column
            cell = ws.cell(row=2, column=afs_start_col + idx, value=period)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Add borders to merged cells in row 2 for Country and Compliance Status
        for col in [1, compliance_col]:
            cell = ws.cell(row=2, column=col)
            cell.border = thin_border

        # Write data (using pre-fetched lookups) - one row per country
        row_num = 3
        for country in countries:
            has_annual_report = False
            has_audited_financial = False

            # Build period_docs data and determine compliance
            period_docs_map = {}
            for period in periods:
                period_docs = {doc_type: False for doc_type in COMPLIANCE_DOC_TYPES}
                assignment = assignment_map.get(period)
                if assignment:
                    aes = aes_lookup.get((assignment.id, country.id))
                    if aes:
                        for doc_type in COMPLIANCE_DOC_TYPES:
                            if doc_lookup.get((aes.id, doc_type)):
                                period_docs[doc_type] = True
                                if doc_type == "Annual Report":
                                    has_annual_report = True
                                elif doc_type == "Audited Financial Statement":
                                    has_audited_financial = True
                period_docs_map[period] = period_docs

            is_compliant = has_annual_report and has_audited_financial

            # Write single row for country
            ws.cell(row=row_num, column=1, value=country.name).border = thin_border

            # Annual Report columns (one per period)
            for idx, period in enumerate(periods):
                cell = ws.cell(row=row_num, column=ar_start_col + idx)
                if period_docs_map.get(period, {}).get("Annual Report", False):
                    cell.value = "Yes"
                    cell.font = yes_font
                else:
                    cell.value = "No"
                    cell.font = no_font
                cell.alignment = center_align
                cell.border = thin_border

            # Audited Financial Statement columns (one per period)
            for idx, period in enumerate(periods):
                cell = ws.cell(row=row_num, column=afs_start_col + idx)
                if period_docs_map.get(period, {}).get("Audited Financial Statement", False):
                    cell.value = "Yes"
                    cell.font = yes_font
                else:
                    cell.value = "No"
                    cell.font = no_font
                cell.alignment = center_align
                cell.border = thin_border

            # Compliance Status
            cell = ws.cell(row=row_num, column=compliance_col)
            cell.value = "Compliant" if is_compliant else "Non-Compliant"
            cell.fill = compliant_fill if is_compliant else non_compliant_fill
            cell.alignment = center_align
            cell.border = thin_border

            row_num += 1

        # Adjust column widths dynamically
        ws.column_dimensions[get_column_letter(1)].width = 30  # Country
        for idx in range(num_periods):
            ws.column_dimensions[get_column_letter(ar_start_col + idx)].width = 12  # AR periods
            ws.column_dimensions[get_column_letter(afs_start_col + idx)].width = 12  # AFS periods
        ws.column_dimensions[get_column_letter(compliance_col)].width = 18  # Compliance Status

        # Freeze header rows (data starts at row 3)
        ws.freeze_panes = 'A3'

        # Add summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.cell(row=1, column=1, value="FDRS Document Compliance Summary").font = Font(bold=True, size=14)
        ws_summary.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        if reference_year:
            ws_summary.cell(row=3, column=1, value=f"Reference Year: {reference_year} (analyzing {reference_year}, {reference_year-1}, {reference_year-2})")
        ws_summary.cell(row=4, column=1, value=f"Periods analyzed: {', '.join(periods)}")

        ws_summary.cell(row=6, column=1, value="Total Countries:").font = Font(bold=True)
        ws_summary.cell(row=6, column=2, value=len(countries))
        ws_summary.cell(row=7, column=1, value="Compliance Rule:").font = Font(bold=True)
        ws_summary.cell(row=7, column=2, value="At least 1 Annual Report AND 1 Audited Financial Statement in the past 3 years")

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename with timestamp and reference year
        year_suffix = f"_{reference_year}" if reference_year else ""
        filename = f"FDRS_Compliance{year_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except ImportError:
        logger.error("openpyxl not installed for Excel export")
        return json_server_error("Excel export requires openpyxl library")
    except Exception as e:
        logger.error("Error generating compliance Excel: %s", e, exc_info=True)
        return json_server_error(GENERIC_ERROR_MESSAGE)
