from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, send_file, after_this_request
from flask_login import current_user
import tempfile
from app.models import db, FormTemplate, FormItem, FormSection, AssignedForm, FormData, Country, TemplateShare
from app.models.assignments import AssignmentEntityStatus
from app.routes.admin.shared import admin_permission_required, check_template_access
from app.services.imputation_service import ImputationService
import io
import os
import sys
import threading
import time
import uuid
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from contextlib import suppress
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from app.utils.transactions import request_transaction_rollback
from app.utils.api_helpers import GENERIC_ERROR_MESSAGE, get_json_safe
from app.utils.request_utils import get_json_or_form, is_json_request
from app.utils.error_handling import handle_json_view_exception
from app.utils.api_responses import json_accepted, json_bad_request, json_error, json_forbidden, json_not_found, json_ok, json_server_error
bp = Blueprint("template_special", __name__, url_prefix="/admin/templates/special")

# -----------------------------
# FDRS sync progress (in-memory)
# -----------------------------
_FDRS_SYNC_LOCK = threading.Lock()
_FDRS_SYNC_JOBS: Dict[str, Dict[str, Any]] = {}
_FDRS_SYNC_TTL_SECONDS = 6 * 60 * 60


def _utc_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _cleanup_fdrs_jobs_locked(now_ts: Optional[float] = None) -> None:
    """Remove old jobs to prevent unbounded memory growth (lock must be held)."""
    if now_ts is None:
        now_ts = time.time()
    expired = []
    for jid, job in _FDRS_SYNC_JOBS.items():
        updated_ts = float(job.get("updated_ts") or job.get("started_ts") or 0.0)
        if updated_ts and (now_ts - updated_ts) > _FDRS_SYNC_TTL_SECONDS:
            expired.append(jid)
    for jid in expired:
        job = _FDRS_SYNC_JOBS.pop(jid, None)
        # Best-effort cleanup of temporary preview files
        if job:
            path = job.get("preview_path")
            if path and isinstance(path, str):
                with suppress(Exception):
                    if os.path.isfile(path):
                        os.unlink(path)


_FDRS_SYNC_ALLOWED_STATES = frozenset({0, 100, 200, 300, 400, 500})


def _parse_fdrs_reported_import_states_from_request(data: Dict[str, Any]) -> Optional[List[int]]:
    """
    Parse JSON fdrs_reported_import_states (list of ints or comma-separated string).
    If the key is absent, return None (importer uses env / default allowlist).
    """
    if "fdrs_reported_import_states" not in data:
        return None
    raw = data.get("fdrs_reported_import_states")
    if raw is None:
        return None
    out: List[int] = []
    if isinstance(raw, str):
        parts = [p.strip() for p in raw.split(",") if p.strip()]
        for p in parts:
            try:
                out.append(int(p))
            except ValueError:
                raise ValueError("Each data status must be a whole number (IFRC State).")
    elif isinstance(raw, (list, tuple)):
        for p in raw:
            try:
                out.append(int(p))
            except (TypeError, ValueError):
                raise ValueError("Each data status must be a whole number (IFRC State).")
    else:
        raise ValueError("Data statuses must be sent as a list or comma-separated numbers.")
    if not out:
        raise ValueError("Select at least one data status to include.")
    bad = [x for x in out if x not in _FDRS_SYNC_ALLOWED_STATES]
    if bad:
        raise ValueError(
            "Unknown data status value(s): %s. Use only the statuses shown in the sync dialog."
            % (", ".join(str(x) for x in sorted(set(bad))),)
        )
    return out


@bp.route("/<int:template_id>", methods=["GET"])
@admin_permission_required('admin.templates.view')
def special_template_view(template_id: int):
    template = FormTemplate.query.get_or_404(template_id)

    # Check if user has access to this template (owner or shared with)
    if not check_template_access(template_id, current_user.id):
        flash("Access denied. You don't have permission to access this template.", "warning")
        return redirect(url_for("form_builder.manage_templates"))

    # Get all template sections ordered by their order field
    template_sections = FormSection.query.filter_by(template_id=template_id).order_by(FormSection.order).all()

    # Group items by section
    sections_with_items = []
    for section in template_sections:
        # Get items for this section, ordered by their order field
        section_items = FormItem.query.filter_by(
            template_id=template_id,
            section_id=section.id
        ).order_by(FormItem.order).all()

        if section_items:  # Only include sections that have items
            sections_with_items.append({
                'section': section,
                'section_items': list(section_items)  # Convert to list to ensure proper length calculation
            })

    return render_template(
        "admin/templates/special_template.html",
        template=template,
        sections_with_items=sections_with_items,
        title=f"Special View - {template.name}"
    )


@bp.route("/impute/template2", methods=["POST"])
@admin_permission_required('admin.templates.edit')
def impute_template2():
    # Accept from form or JSON; trim whitespace
    target_period = (request.form.get('target_period') or '').strip()
    if not target_period and is_json_request():
        try:
            payload = get_json_safe()
            target_period = str(payload.get('target_period') or '').strip()
        except Exception as e:
            current_app.logger.debug("target_period extraction failed: %s", e)
            target_period = ''
    if not target_period:
        flash("Target period is required (e.g., 2025)", "warning")
        return redirect(url_for('template_special.special_template_view', template_id=2))

    try:
        result = ImputationService.impute_template_2(target_period)
        if result.get("success"):
            flash(
                f"Imputation completed for {result['target_period']} from {result['source_period']}. "
                f"Countries: {result['countries_processed']}, Items: {result['items_imputed']}, "
                f"Rows created: {result['rows_created']}, updated: {result['rows_updated']}",
                "success"
            )
        else:
            flash(result.get("error") or "Imputation failed", "danger")
    except Exception as e:
        current_app.logger.error(f"Imputation error: {e}", exc_info=True)
        flash("An error occurred during imputation.", "danger")

    return redirect(url_for('template_special.special_template_view', template_id=2))


@bp.route("/update-imputation-methods-batch", methods=["POST"])
@admin_permission_required('admin.templates.edit')
def update_imputation_methods_batch():
    """Update imputation methods for multiple form items in a single request."""
    try:
        data = get_json_safe()
        updates = data.get('updates', [])

        if not updates:
            return json_bad_request('No updates provided')

        results = []
        for update in updates:
            item_id = update.get('item_id')
            method = update.get('method')

            if item_id is None or method is None:
                results.append({'item_id': item_id, 'success': False, 'error': 'Missing item_id or method'})
                continue

            # Coerce item_id to int
            try:
                item_id_int = int(item_id)
            except Exception as e:
                current_app.logger.debug("item_id int parse failed for %r: %s", item_id, e)
                results.append({'item_id': item_id, 'success': False, 'error': 'Invalid item_id'})
                continue

            if method not in ['last_year', 'three_year_avg', 'no_imputation']:
                results.append({'item_id': item_id_int, 'success': False, 'error': 'Invalid method'})
                continue

            # Get the form item
            form_item = FormItem.query.get(item_id_int)
            if not form_item:
                results.append({'item_id': item_id_int, 'success': False, 'error': 'Form item not found'})
                continue

            # Update the config - force a new dict to trigger SQLAlchemy change detection
            if form_item.config is None:
                new_config = {'imputation_method': method}
            else:
                new_config = form_item.config.copy()
                new_config['imputation_method'] = method

            # Force assignment to trigger change detection
            form_item.config = new_config
            db.session.add(form_item)
            results.append({'item_id': item_id_int, 'success': True})

        # Commit all changes at once
        db.session.flush()

        return json_ok(
            success=True,
            results=results,
            total_updated=len([r for r in results if r.get('success')])
        )

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/preview-data", methods=["GET"])
@admin_permission_required('admin.templates.view')
def preview_data(template_id: int):
    """Get current data for preview table with progress tracking."""
    try:
        year = request.args.get('year')
        if not year:
            return json_bad_request('Year parameter required')

        template = FormTemplate.query.get_or_404(template_id)

        # Get all assignments for this template and year
        assignments = AssignedForm.query.filter_by(
            template_id=template_id,
            period_name=year
        ).all()

        # Get all form items for this template (exclude items with no_imputation method)
        all_form_items = FormItem.query.filter_by(template_id=template_id).order_by(FormItem.order).all()
        form_items = []
        for item in all_form_items:
            # Skip items that are set to no_imputation
            imputation_method = item.config.get('imputation_method', 'last_year') if item.config else 'last_year'
            if imputation_method != 'no_imputation':
                form_items.append(item)

        preview_data = []
        total_operations = 0
        completed_operations = 0

        # Calculate total operations for progress tracking
        for assignment in assignments:
            total_operations += len(assignment.country_statuses.all()) * len(form_items)

        for assignment in assignments:
            for aes in assignment.country_statuses.all():
                country = aes.country
                if not country:
                    completed_operations += len(form_items)
                    continue

                for item in form_items:
                    # Get current data
                    form_data = FormData.query.filter_by(
                        assignment_entity_status_id=aes.id,
                        form_item_id=item.id
                    ).first()

                    current_value = None
                    if form_data:
                        try:
                            current_value = form_data.total_value
                        except AttributeError:
                            current_value = form_data.value

                    preview_data.append({
                        'country': country.name,
                        'item_label': item.label,
                        'item_unit': item.unit,
                        'current_value': current_value,
                        'imputed_value': None,  # Will be filled by preview imputation
                        'method': None,
                        'source_periods': None
                    })

                    completed_operations += 1

        return json_ok(
            success=True,
            data=preview_data,
            progress={
                'total': total_operations,
                'completed': completed_operations,
                'percentage': 100
            }
        )

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/preview-imputation", methods=["POST"])
@admin_permission_required('admin.templates.edit')
def preview_imputation(template_id: int):
    """Preview imputation results without saving to database."""
    try:
        data = get_json_safe()
        year = data.get('year')
        if not year:
            return json_bad_request('Year parameter required')

        template = FormTemplate.query.get_or_404(template_id)

        # Calculate previous year
        try:
            prev_year = str(int(year) - 1)
        except (ValueError, TypeError):
            return json_bad_request('Invalid year format')

        # Get all assignments for target and previous year
        target_assignments = AssignedForm.query.filter_by(
            template_id=template_id,
            period_name=year
        ).all()

        prev_assignments_by_country = {}
        for af in AssignedForm.query.filter_by(template_id=template_id, period_name=prev_year).all():
            for aes in af.country_statuses.all():
                if aes.country_id not in prev_assignments_by_country:
                    prev_assignments_by_country[aes.country_id] = af

        # Get all form items (exclude items with no_imputation method)
        all_form_items = FormItem.query.filter_by(template_id=template_id).order_by(FormItem.order).all()
        form_items = []
        for item in all_form_items:
            # Skip items that are set to no_imputation
            imputation_method = item.config.get('imputation_method', 'last_year') if item.config else 'last_year'
            if imputation_method != 'no_imputation':
                form_items.append(item)

        preview_data = []

        for assignment in target_assignments:
            for aes in assignment.country_statuses.all():
                country = aes.country
                if not country:
                    continue

                prev_af = prev_assignments_by_country.get(country.id)
                prev_aes = None
                if prev_af:
                    prev_aes = prev_af.country_statuses.filter_by(country_id=country.id).first()

                for item in form_items:
                    # Get current data
                    current_fd = FormData.query.filter_by(
                        assignment_entity_status_id=aes.id,
                        form_item_id=item.id
                    ).first()

                    current_value = None
                    if current_fd:
                        try:
                            current_value = current_fd.total_value
                        except AttributeError:
                            current_value = current_fd.value

                    # Calculate imputed value
                    imputed_value = None
                    method = None
                    source_periods = []

                    if prev_aes:
                        imputation_method = item.config.get('imputation_method', 'last_year') if item.config else 'last_year'

                        if imputation_method == 'three_year_avg':
                            # Get 3 years of data
                            values = []
                            for year_offset in range(1, 4):
                                source_year = str(int(year) - year_offset)
                                source_af = AssignedForm.query.filter_by(
                                    template_id=template_id,
                                    period_name=source_year
                                ).first()
                                if source_af:
                                    source_aes = source_af.country_statuses.filter_by(country_id=country.id).first()
                                    if source_aes:
                                        source_fd = FormData.query.filter_by(
                                            assignment_entity_status_id=source_aes.id,
                                            form_item_id=item.id
                                        ).first()
                                        if source_fd:
                                            with suppress(Exception):
                                                val = source_fd.total_value
                                                if val is not None:
                                                    values.append(float(val))
                                                    source_periods.append(source_year)

                            if values:
                                imputed_value = sum(values) / len(values)
                                method = 'Three Year Average'

                        else:  # last_year
                            prev_fd = FormData.query.filter_by(
                                assignment_entity_status_id=prev_aes.id,
                                form_item_id=item.id
                            ).first()
                            if prev_fd:
                                with suppress(Exception):
                                    # For single_choice items, prioritize value field over total_value
                                    # since text values like "Male" are stored in the value field
                                    val = prev_fd.value
                                    if val is not None and val != '':
                                        # Try to convert to float for numeric values, keep as string for text
                                        try:
                                            imputed_value = float(val)
                                        except (ValueError, TypeError):
                                            imputed_value = val  # Keep as string for text values
                                        method = 'Last Year\'s Value'
                                        source_periods = [prev_year]
                                    else:
                                        # Fallback to total_value for numeric items
                                        val = prev_fd.total_value
                                        if val is not None:
                                            imputed_value = float(val)
                                            method = 'Last Year\'s Value'
                                            source_periods = [prev_year]

                    preview_data.append({
                        'country': country.name,
                        'item_label': item.label,
                        'item_unit': item.unit,
                        'current_value': current_value,
                        'imputed_value': round(imputed_value, 2) if imputed_value is not None and isinstance(imputed_value, (int, float)) else imputed_value,
                        'method': method,
                        'source_periods': source_periods
                    })

        return json_ok(success=True, data=preview_data)

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/preview-data-chunked", methods=["POST"])
@admin_permission_required('admin.templates.view')
def preview_data_chunked(template_id: int):
    """Get current data for preview table with real progress tracking using chunked processing.

    Optimized to avoid building the full (ACS x items) combinations list on each request.
    """
    try:
        data = get_json_safe()
        year = data.get('year')
        chunk_size = data.get('chunk_size', 2000)
        offset = data.get('offset', 0)
        country_filter = data.get('country_filter')
        item_filter = data.get('item_filter')
        type_filter = data.get('type_filter')
        imputation_mode = data.get('imputation_mode', 'missing_only')  # Unused here but kept for parity

        if not year:
            return json_bad_request('Year parameter required')

        # Validate template
        _ = FormTemplate.query.get_or_404(template_id)

        # Gather filtered ACS (with country) for the target year
        assignments = AssignedForm.query.filter_by(template_id=template_id, period_name=year).all()
        filtered_aess = []  # List[Tuple[AssignmentEntityStatus, Country]]
        for assignment in assignments:
            for aes in assignment.country_statuses.all():
                country = aes.country
                if not country:
                    continue
                if country_filter and country.name != country_filter:
                    continue
                filtered_aess.append((aes, country))

        # Gather filtered items (exclude items with no_imputation method)
        items_query = FormItem.query.filter_by(template_id=template_id).order_by(FormItem.order)
        items = items_query.all()
        filtered_items = []
        for item in items:
            # Skip items that are set to no_imputation
            imputation_method = item.config.get('imputation_method', 'last_year') if item.config else 'last_year'
            if imputation_method == 'no_imputation':
                continue

            if item_filter and item.label != item_filter:
                continue
            if type_filter and item.type != type_filter:
                continue
            filtered_items.append(item)

        num_aess = len(filtered_aess)
        num_items = len(filtered_items)
        total = num_aess * num_items

        if total == 0:
            return json_ok(
                success=True,
                data=[],
                progress={
                    'total': 0,
                    'completed': 0,
                    'percentage': 100,
                    'is_complete': True,
                    'next_offset': None
                }
            )

        end_offset = min(offset + chunk_size, total)

        # Build needed sets for preloading FormData for this chunk only
        aes_ids_needed = set()
        item_ids_needed = set()
        for i in range(offset, end_offset):
            aes_idx = i // num_items
            item_idx = i % num_items
            aes, _country = filtered_aess[aes_idx]
            item = filtered_items[item_idx]
            aes_ids_needed.add(aes.id)
            item_ids_needed.add(item.id)

        formdata_map = {}
        if aes_ids_needed and item_ids_needed:
            fds = FormData.query.filter(
                FormData.assignment_entity_status_id.in_(aes_ids_needed),
                FormData.form_item_id.in_(item_ids_needed)
            ).all()
            for fd in fds:
                formdata_map[(fd.assignment_entity_status_id, fd.form_item_id)] = fd

        # Build chunk data using index math
        chunk_data = []
        for i in range(offset, end_offset):
            aes_idx = i // num_items
            item_idx = i % num_items
            aes, country = filtered_aess[aes_idx]
            item = filtered_items[item_idx]

            fd = formdata_map.get((aes.id, item.id))
            current_value = None
            if fd:
                try:
                    current_value = fd.total_value
                except Exception as e:
                    current_app.logger.debug("fd.total_value failed: %s", e)
                    current_value = fd.value

            chunk_data.append({
                'country': country.name,
                'item_label': item.label,
                'item_unit': item.unit,
                'current_value': current_value,
                'imputed_value': None,
                'method': None,
                'source_periods': None
            })

        is_complete = end_offset >= total
        progress_percentage = (end_offset / total) * 100 if total > 0 else 100

        return json_ok(
            success=True,
            data=chunk_data,
            progress={
                'total': total,
                'completed': end_offset,
                'percentage': round(progress_percentage, 1),
                'is_complete': is_complete,
                'next_offset': end_offset if not is_complete else None
            }
        )

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/preview-imputation-chunked", methods=["POST"])
@admin_permission_required('admin.templates.edit')
def preview_imputation_chunked(template_id: int):
    """Preview imputation results with real progress tracking using chunked processing.

    Optimized to avoid building the full (ACS x items) combinations list on each request,
    and preload only the FormData entries needed for the current chunk.
    """
    try:
        data = get_json_safe()
        year = data.get('year')
        chunk_size = data.get('chunk_size', 2000)
        offset = data.get('offset', 0)
        country_filter = data.get('country_filter')
        item_filter = data.get('item_filter')
        type_filter = data.get('type_filter')
        imputation_mode = data.get('imputation_mode', 'missing_only')

        if not year:
            return json_bad_request('Year parameter required')

        # Validate template and compute previous year
        _ = FormTemplate.query.get_or_404(template_id)
        try:
            prev_year = str(int(year) - 1)
        except Exception as e:
            current_app.logger.debug("prev_year parse failed: %s", e)
            return json_bad_request('Invalid year format')

        # Target AES list with countries (filtered)
        target_assignments = AssignedForm.query.filter_by(template_id=template_id, period_name=year).all()
        filtered_aess = []  # List[Tuple[AssignmentEntityStatus, Country]]
        for assignment in target_assignments:
            for aes in assignment.country_statuses.all():
                country = aes.country
                if not country:
                    continue
                if country_filter and country.name != country_filter:
                    continue
                filtered_aess.append((aes, country))

        # Items (filtered, exclude items with no_imputation method)
        items_query = FormItem.query.filter_by(template_id=template_id).order_by(FormItem.order)
        items = items_query.all()
        filtered_items = []
        for item in items:
            # Skip items that are set to no_imputation
            imputation_method = item.config.get('imputation_method', 'last_year') if item.config else 'last_year'
            if imputation_method == 'no_imputation':
                continue

            if item_filter and item.label != item_filter:
                continue
            if type_filter and item.type != type_filter:
                continue
            filtered_items.append(item)

        num_aess = len(filtered_aess)
        num_items = len(filtered_items)
        total = num_aess * num_items

        if total == 0:
            return json_ok(
                success=True,
                data=[],
                progress={
                    'total': 0,
                    'completed': 0,
                    'percentage': 100,
                    'is_complete': True,
                    'next_offset': None
                }
            )

        end_offset = min(offset + chunk_size, total)

        # Determine source years for three-year average
        source_years = []
        try:
            base_year = int(year)
            source_years = [str(base_year - k) for k in (1, 2, 3)]
        except Exception as e:
            current_app.logger.debug("source_years parse failed: %s", e)
            source_years = []

        # Preload mappings of (year -> country_id -> aes_id) for needed years, limited to countries present
        country_ids_in_scope = {country.id for (_aes, country) in filtered_aess}
        country_year_to_aes_id = {}
        if source_years:
            prev_afs = AssignedForm.query.filter(
                AssignedForm.template_id == template_id,
                AssignedForm.period_name.in_(source_years)
            ).all()
            for paf in prev_afs:
                for p_aes in paf.country_statuses.all():
                    if p_aes.country_id in country_ids_in_scope:
                        country_year_to_aes_id[(paf.period_name, p_aes.country_id)] = p_aes.id

        # Collect needed FormData pairs for this chunk only
        current_pairs = set()
        source_pairs = set()
        for i in range(offset, end_offset):
            aes_idx = i // num_items
            item_idx = i % num_items
            aes, country = filtered_aess[aes_idx]
            item = filtered_items[item_idx]

            current_pairs.add((aes.id, item.id))

            imputation_method = item.config.get('imputation_method', 'last_year') if item.config else 'last_year'
            if imputation_method == 'three_year_avg':
                for sy in source_years:
                    src_aes_id = country_year_to_aes_id.get((sy, country.id))
                    if src_aes_id:
                        source_pairs.add((src_aes_id, item.id))
            elif imputation_method == 'last_year':
                src_aes_id = country_year_to_aes_id.get((prev_year, country.id))
                if src_aes_id:
                    source_pairs.add((src_aes_id, item.id))

        # Preload FormData for all needed pairs
        needed_pairs = current_pairs | source_pairs
        all_aes_ids_needed = {aes_id for (aes_id, _iid) in needed_pairs}
        all_item_ids_needed = {iid for (_aes_id, iid) in needed_pairs}

        formdata_map = {}
        if all_aes_ids_needed and all_item_ids_needed:
            fds = FormData.query.filter(
                FormData.assignment_entity_status_id.in_(all_aes_ids_needed),
                FormData.form_item_id.in_(all_item_ids_needed)
            ).all()
            for fd in fds:
                formdata_map[(fd.assignment_entity_status_id, fd.form_item_id)] = fd

        # Build chunk data
        chunk_data = []
        for i in range(offset, end_offset):
            aes_idx = i // num_items
            item_idx = i % num_items
            aes, country = filtered_aess[aes_idx]
            item = filtered_items[item_idx]

            fd_current = formdata_map.get((aes.id, item.id))
            current_value = None
            if fd_current:
                try:
                    current_value = fd_current.total_value
                except Exception as e:
                    current_app.logger.debug("fd_current.total_value failed: %s", e)
                    current_value = fd_current.value

            imputed_value = None
            method = None
            source_periods = []

            imputation_method = item.config.get('imputation_method', 'last_year') if item.config else 'last_year'
            if imputation_method != 'no_imputation':
                if imputation_method == 'three_year_avg' and source_years:
                    values = []
                    used_years = []
                    for sy in source_years:
                        src_aes_id = country_year_to_aes_id.get((sy, country.id))
                        if not src_aes_id:
                            continue
                        fd_src = formdata_map.get((src_aes_id, item.id))
                        if not fd_src:
                            continue
                        with suppress(Exception):
                            val = fd_src.total_value
                            if val is None:
                                val = fd_src.value
                            if val is not None:
                                values.append(float(val))
                                used_years.append(sy)
                    if values:
                        imputed_value = sum(values) / len(values)
                        method = 'Three Year Average'
                        source_periods = used_years
                elif imputation_method == 'last_year':
                    src_aes_id = country_year_to_aes_id.get((prev_year, country.id))
                    if src_aes_id:
                        fd_prev = formdata_map.get((src_aes_id, item.id))
                        if fd_prev:
                            with suppress(Exception):
                                # For single_choice items, prioritize value field over total_value
                                # since text values like "Male" are stored in the value field
                                val = fd_prev.value
                                if val is not None and val != '':
                                    # Try to convert to float for numeric values, keep as string for text
                                    try:
                                        imputed_value = float(val)
                                    except (ValueError, TypeError):
                                        imputed_value = val  # Keep as string for text values
                                    method = 'Last Year\'s Value'
                                    source_periods = [prev_year]
                                else:
                                    # Fallback to total_value for numeric items
                                    val = fd_prev.total_value
                                    if val is not None:
                                        imputed_value = float(val)
                                        method = 'Last Year\'s Value'
                                        source_periods = [prev_year]

            # Respect imputation mode
            if imputation_mode == 'missing_only' and imputed_value is not None:
                has_existing_data = (
                    (current_value is not None and current_value != '') or
                    (fd_current and fd_current.imputed_value is not None and fd_current.imputed_value != '')
                )
                if has_existing_data:
                    imputed_value = None
                    method = None
                    source_periods = []

            chunk_data.append({
                'country': country.name,
                'item_label': item.label,
                'item_unit': item.unit,
                'current_value': current_value,
                'imputed_value': round(imputed_value, 2) if imputed_value is not None and isinstance(imputed_value, (int, float)) else imputed_value,
                'method': method,
                'source_periods': source_periods
            })

        is_complete = end_offset >= total
        progress_percentage = (end_offset / total) * 100 if total > 0 else 100

        # Log summary for this chunk
        imputed_count = len([row for row in chunk_data if row['imputed_value'] is not None])
        current_app.logger.info(f"Chunk {offset}-{end_offset}: {imputed_count}/{len(chunk_data)} items have imputed values")

        return json_ok(
            success=True,
            data=chunk_data,
            progress={
                'total': total,
                'completed': end_offset,
                'percentage': round(progress_percentage, 1),
                'is_complete': is_complete,
                'next_offset': end_offset if not is_complete else None
            }
        )

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/filter-options", methods=["GET"])
@admin_permission_required('admin.templates.view')
def get_filter_options(template_id: int):
    """Get available countries and items for filters."""
    try:
        year = request.args.get('year')
        if not year:
            return json_bad_request('Year parameter required')

        template = FormTemplate.query.get_or_404(template_id)

        # Get all assignments for this template and year
        assignments = AssignedForm.query.filter_by(
            template_id=template_id,
            period_name=year
        ).all()

        # Get all form items for this template
        form_items = FormItem.query.filter_by(template_id=template_id).order_by(FormItem.order).all()

        # Get unique countries from assignments
        countries = []
        for assignment in assignments:
            for aes in assignment.country_statuses.all():
                country = aes.country
                if country and country.name not in [c['name'] for c in countries]:
                    countries.append({
                        'id': country.id,
                        'name': country.name
                    })

        # Sort countries by name
        countries.sort(key=lambda x: x['name'])

        # Get all items (exclude items with no_imputation method)
        items = []
        for item in form_items:
            # Skip items that are set to no_imputation
            imputation_method = item.config.get('imputation_method', 'last_year') if item.config else 'last_year'
            if imputation_method == 'no_imputation':
                continue

            items.append({
                'id': item.id,
                'label': item.label,
                'unit': item.unit
            })

        return json_ok(success=True, countries=countries, items=items)

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/max-year", methods=["GET"])
@admin_permission_required('admin.templates.view')
def get_max_year(template_id: int):
    """Get the maximum year available for this template."""
    try:
        template = FormTemplate.query.get_or_404(template_id)

        # Get all assignments for this template and find the max year
        assignments = AssignedForm.query.filter_by(template_id=template_id).all()

        years = []
        for assignment in assignments:
            try:
                # Try to parse the period_name as a year
                year = int(assignment.period_name)
                years.append(year)
            except (ValueError, TypeError):
                # Skip non-numeric period names
                continue

        if years:
            max_year = max(years)
        else:
            # Default to current year if no assignments found
            from datetime import datetime
            max_year = datetime.now().year

        return json_ok(success=True, max_year=str(max_year))

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/run-imputation-filtered", methods=["POST"])
@admin_permission_required('admin.templates.edit')
def run_imputation_filtered(template_id: int):
    """Run imputation for a specific year with optional country and item filters."""
    try:
        data = get_json_safe()
        year = data.get('year')
        country_filter = data.get('country_filter')
        item_filter = data.get('item_filter')
        type_filter = data.get('type_filter')
        imputation_mode = data.get('imputation_mode', 'missing_only')

        if not year:
            return json_bad_request('Year parameter required')

        # For now, only support template 2
        if template_id != 2:
            return json_bad_request('Filtered imputation only supported for template 2')

        # Use the existing imputation service but with filters
        result = ImputationService.impute_template_2_filtered(
            target_year=year,
            country_filter=country_filter,
            item_filter=item_filter,
            type_filter=type_filter,
            imputation_mode=imputation_mode
        )

        if result.get("success"):
            return json_ok(
                success=True,
                target_period=result['target_period'],
                source_period=result['source_period'],
                countries_processed=result['countries_processed'],
                items_imputed=result['items_imputed'],
                rows_created=result['rows_created'],
                rows_updated=result['rows_updated']
            )
        else:
            return json_bad_request(result.get('error', 'Imputation failed'))

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/export-excel", methods=["POST"])
@admin_permission_required('admin.templates.export_excel')
def export_preview_excel(template_id: int):
    """Export preview data to Excel."""
    try:
        data = get_json_safe()
        year = data.get('year')
        preview_data = data.get('preview_data', [])

        if not year or not preview_data:
            return json_bad_request('Year and preview data required')

        # Create workbook with formatting
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Preview Data"

        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        imputed_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

        # Headers
        headers = ['Country', 'Item', 'Unit', 'Current Value', 'Imputed Value', 'Method', 'Source Periods']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, row_data in enumerate(preview_data, 2):
            sheet.cell(row=row_idx, column=1, value=row_data.get('country', ''))
            sheet.cell(row=row_idx, column=2, value=row_data.get('item_label', ''))
            sheet.cell(row=row_idx, column=3, value=row_data.get('item_unit', ''))

            # Current value
            current_val = row_data.get('current_value')
            sheet.cell(row=row_idx, column=4, value=current_val if current_val is not None else 'No data')

            # Imputed value with highlighting
            imputed_val = row_data.get('imputed_value')
            imputed_cell = sheet.cell(row=row_idx, column=5, value=imputed_val if imputed_val is not None else '-')
            if imputed_val is not None:
                imputed_cell.fill = imputed_fill

            sheet.cell(row=row_idx, column=6, value=row_data.get('method', '-'))
            sheet.cell(row=row_idx, column=7, value=', '.join(row_data.get('source_periods', [])) or '-')

        # Auto-size columns
        for column in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            column_letter = column[0].column_letter
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 30)

        # Freeze header row
        sheet.freeze_panes = "A2"

        # Save and return
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            download_name=f"preview_data_{year}.xlsx",
            as_attachment=True
        )

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/run-fdrs-sync", methods=["POST"])
@admin_permission_required("admin.templates.edit")
def run_fdrs_sync(template_id: int):
    """
    Trigger FDRS form data import/sync (import_fdrs_form_data.py pipeline).
    Expects JSON: dry_run (bool), batch_size (int), fdrs_years (str, comma-separated), test (bool),
    imputed_use_cache (bool), async (bool), and optional fdrs_reported_import_states (list of IFRC State ints).
    If fdrs_reported_import_states is omitted, the importer uses FDRS_REPORTED_IMPORT_STATES env or default all except Not filled (0).
    """
    try:
        template = FormTemplate.query.get_or_404(template_id)
        if not check_template_access(template_id, current_user.id):
            return json_forbidden("Access denied")

        data = get_json_safe()
        dry_run = bool(data.get("dry_run", False))
        batch_size_raw = data.get("batch_size", None)
        if batch_size_raw in (None, ""):
            batch_size = 1000
        else:
            try:
                batch_size = int(batch_size_raw)
            except Exception as e:
                current_app.logger.debug("batch_size parse failed: %s", e)
                return json_bad_request("Invalid batch_size: must be an integer (or omit it)")
        if batch_size < 100:
            return json_bad_request("Invalid batch_size: must be >= 100")
        fdrs_years_raw = (data.get("fdrs_years") or "").strip()
        test_mode = bool(data.get("test", False))
        async_mode = bool(data.get("async", False))
        imputed_use_cache = bool(data.get("imputed_use_cache", True))
        try:
            fdrs_reported_import_states = _parse_fdrs_reported_import_states_from_request(data)
        except ValueError as e:
            return json_bad_request(str(e))

        fdrs_years = None
        test_limit = None
        if test_mode:
            fdrs_years = [2024]
            test_limit = 1000
        elif fdrs_years_raw:
            try:
                fdrs_years = [int(y.strip()) for y in fdrs_years_raw.split(",") if y.strip()]
            except ValueError:
                return json_bad_request("Invalid fdrs_years: use comma-separated integers")

        scripts_dir = os.path.join(current_app.root_path, "..", "scripts")
        if scripts_dir not in sys.path:
            sys.path.insert(0, scripts_dir)
        from import_fdrs_form_data import run_import

        preview_path = None
        if dry_run:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp.close()
            preview_path = tmp.name

        # Async mode: run in background and expose progress via polling endpoint.
        if async_mode:
            job_id = uuid.uuid4().hex
            now_ts = time.time()
            logger = current_app.logger
            with _FDRS_SYNC_LOCK:
                _cleanup_fdrs_jobs_locked(now_ts)
                _FDRS_SYNC_JOBS[job_id] = {
                    "job_id": job_id,
                    "template_id": template_id,
                    "user_id": int(getattr(current_user, "id", 0) or 0),
                    "status": "queued",  # queued | running | completed | failed
                    "stage": "queued",
                    "message": "Queued",
                    "current": 0,
                    "total": None,
                    "percent": 0.0,
                    "stats": None,
                    "error": None,
                    "preview_path": preview_path,
                    "download_ready": False,
                    "started_at": _utc_iso(),
                    "updated_at": _utc_iso(),
                    "started_ts": now_ts,
                    "updated_ts": now_ts,
                    "last_logged_pct": None,
                }

            def _progress_cb(payload: Dict[str, Any]) -> None:
                # payload keys: stage, message, current, total, percent, stats...
                stage = payload.get("stage") or ""
                pct = payload.get("percent")
                msg = payload.get("message") or ""
                with _FDRS_SYNC_LOCK:
                    job = _FDRS_SYNC_JOBS.get(job_id)
                    if not job:
                        return
                    job["status"] = "running"
                    job["stage"] = payload.get("stage") or job.get("stage")
                    job["message"] = payload.get("message") or job.get("message")
                    job["current"] = payload.get("current") if payload.get("current") is not None else job.get("current")
                    job["total"] = payload.get("total") if payload.get("total") is not None else job.get("total")
                    job["percent"] = float(payload.get("percent") or job.get("percent") or 0.0)
                    if payload.get("stats") is not None:
                        job["stats"] = payload.get("stats")
                    job["updated_at"] = _utc_iso()
                    job["updated_ts"] = time.time()

                # Write progress to backend logs (throttled)
                try:
                    pct_f = float(pct) if pct is not None else None
                except Exception as e:
                    current_app.logger.debug("pct float parse failed: %s", e)
                    pct_f = None
                with _FDRS_SYNC_LOCK:
                    job = _FDRS_SYNC_JOBS.get(job_id) or {}
                    last_logged = job.get("last_logged_pct")
                    should_log = (stage and stage != "upsert") or (pct_f is not None and (last_logged is None or abs(pct_f - float(last_logged)) >= 5.0))
                    if should_log and pct_f is not None:
                        job["last_logged_pct"] = pct_f
                if should_log:
                    logger.info("FDRS sync %s: %s %s%% %s", job_id, stage or "-", f"{pct_f:.1f}" if pct_f is not None else "-", msg)

            def _run_job() -> None:
                log = logging.getLogger(__name__)
                with _FDRS_SYNC_LOCK:
                    job = _FDRS_SYNC_JOBS.get(job_id)
                    if job:
                        job["status"] = "running"
                        job["stage"] = "starting"
                        job["message"] = "Starting..."
                        job["updated_at"] = _utc_iso()
                        job["updated_ts"] = time.time()
                logger.info("FDRS sync %s: starting (template_id=%s, dry_run=%s, test=%s)", job_id, template_id, dry_run, test_mode)

                try:
                    stats = run_import(
                        input_path=None,
                        fdrs_api_url=None,
                        fdrs_from_data_api=True,
                        fdrs_data_api_base=None,
                        fdrs_data_api_key=None,
                        fdrs_imputed_url=None,
                        fdrs_imputed_from_api=False,
                        fdrs_imputed_kpi_codes_path=None,
                        fdrs_imputed_use_cache=imputed_use_cache,
                        fdrs_years=fdrs_years,
                        fdrs_reported_import_states=fdrs_reported_import_states,
                        indicator_mapping_path=None,
                        indicator_bank_api_base=None,
                        indicator_bank_api_key=None,
                        databank_base_url=None,
                        databank_api_key=None,
                        preview_excel_path=preview_path if dry_run else None,
                        test_limit=test_limit,
                        dry_run=dry_run,
                        batch_size=batch_size,
                        template_id=template_id,
                        progress_cb=_progress_cb,
                    )
                    with _FDRS_SYNC_LOCK:
                        job = _FDRS_SYNC_JOBS.get(job_id)
                        if job:
                            job["status"] = "completed"
                            job["stage"] = "complete"
                            job["message"] = "Completed"
                            job["percent"] = 100.0
                            job["stats"] = dict(stats or {})
                            job["download_ready"] = bool(dry_run and preview_path and os.path.isfile(preview_path))
                            job["updated_at"] = _utc_iso()
                            job["updated_ts"] = time.time()
                    logger.info(
                        "FDRS sync %s: completed loaded=%s skipped=%s inserted=%s updated=%s errors=%s",
                        job_id,
                        stats.get("loaded"),
                        stats.get("skipped"),
                        stats.get("inserted"),
                        stats.get("updated"),
                        stats.get("errors"),
                    )
                except Exception as e:
                    log.exception("FDRS async sync job failed: %s", e)
                    err_msg = str(e).strip() or type(e).__name__
                    if len(err_msg) > 2000:
                        err_msg = err_msg[:1997] + "..."
                    with _FDRS_SYNC_LOCK:
                        job = _FDRS_SYNC_JOBS.get(job_id)
                        if job:
                            job["status"] = "failed"
                            job["stage"] = "failed"
                            job["message"] = "Failed"
                            job["error"] = err_msg
                            job["updated_at"] = _utc_iso()
                            job["updated_ts"] = time.time()
                    logger.error("FDRS sync %s: failed: %s", job_id, e, exc_info=True)
                finally:
                    pass

            threading.Thread(target=_run_job, daemon=True).start()
            return json_accepted(job_id=job_id)

        stats = run_import(
            input_path=None,
            fdrs_api_url=None,
            fdrs_from_data_api=True,
            fdrs_data_api_base=None,
            fdrs_data_api_key=None,
            fdrs_imputed_url=None,
            fdrs_imputed_from_api=False,
            fdrs_imputed_kpi_codes_path=None,
            fdrs_imputed_use_cache=imputed_use_cache,
            fdrs_years=fdrs_years,
            fdrs_reported_import_states=fdrs_reported_import_states,
            indicator_mapping_path=None,
            indicator_bank_api_base=None,
            indicator_bank_api_key=None,
            databank_base_url=None,
            databank_api_key=None,
            preview_excel_path=preview_path if dry_run else None,
            test_limit=test_limit,
            dry_run=dry_run,
            batch_size=batch_size,
            template_id=template_id,
        )

        if dry_run and preview_path and os.path.isfile(preview_path):
            @after_this_request
            def _remove_preview(resp):
                try:
                    os.unlink(preview_path)
                except OSError:
                    pass
                return resp
            return send_file(
                preview_path,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="fdrs_import_preview.xlsx",
            )

        return json_ok(
            success=True,
            dry_run=dry_run,
            stats=stats,
            message=(
                f"Loaded: {stats['loaded']}, Skipped: {stats['skipped']}, "
                f"Inserted: {stats['inserted']}, Updated: {stats['updated']}, Errors: {stats['errors']}"
            ),
        )
    except (ValueError, RuntimeError) as e:
        current_app.logger.error(f"FDRS sync error: {e}", exc_info=True)
        msg = str(e).strip() or "Sync failed."
        return json_bad_request(msg[:2000] if len(msg) > 2000 else msg)
    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route("/<int:template_id>/fdrs-sync-status/<job_id>", methods=["GET"])
@admin_permission_required("admin.templates.edit")
def fdrs_sync_status(template_id: int, job_id: str):
    """Poll FDRS sync job status (for live UI progress)."""
    with _FDRS_SYNC_LOCK:
        _cleanup_fdrs_jobs_locked(time.time())
        job = _FDRS_SYNC_JOBS.get(job_id)
        if not job or int(job.get("template_id") or 0) != int(template_id):
            return json_not_found("Job not found")
        if int(job.get("user_id") or 0) != int(getattr(current_user, "id", 0) or 0):
            return json_forbidden("Access denied")

        resp = {
            "success": True,
            "job": {
                "job_id": job_id,
                "status": job.get("status"),
                "stage": job.get("stage"),
                "message": job.get("message"),
                "current": job.get("current"),
                "total": job.get("total"),
                "percent": job.get("percent"),
                "stats": job.get("stats"),
                "error": job.get("error"),
                "started_at": job.get("started_at"),
                "updated_at": job.get("updated_at"),
                "download_ready": bool(job.get("download_ready")),
            },
        }
    if resp["job"]["download_ready"]:
        resp["job"]["download_url"] = url_for("template_special.fdrs_sync_download", template_id=template_id, job_id=job_id)
    return json_ok(**resp) if isinstance(resp, dict) else json_ok(data=resp)


@bp.route("/<int:template_id>/fdrs-sync-download/<job_id>", methods=["GET"])
@admin_permission_required("admin.templates.edit")
def fdrs_sync_download(template_id: int, job_id: str):
    """Download preview Excel generated by an async dry-run sync."""
    with _FDRS_SYNC_LOCK:
        job = _FDRS_SYNC_JOBS.get(job_id)
        if not job or int(job.get("template_id") or 0) != int(template_id):
            return json_not_found("Job not found")
        if int(job.get("user_id") or 0) != int(getattr(current_user, "id", 0) or 0):
            return json_forbidden("Access denied")
        path = job.get("preview_path")
    if not path or not os.path.isfile(path):
        return json_not_found("Preview file not available")

    @after_this_request
    def _remove_preview(resp):
        with suppress(Exception):
            os.unlink(path)
        # Mark as consumed (best-effort)
        with _FDRS_SYNC_LOCK:
            j = _FDRS_SYNC_JOBS.get(job_id)
            if j:
                j["download_ready"] = False
        return resp

    return send_file(
        path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="fdrs_import_preview.xlsx",
    )
