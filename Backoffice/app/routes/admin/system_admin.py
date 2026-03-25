from app.utils.transactions import request_transaction_rollback
from contextlib import suppress
# File: Backoffice/app/routes/admin/system_admin.py
from app.utils.datetime_helpers import utcnow
from app.utils.advanced_validation import AdvancedValidator, validate_upload_extension_and_mime
from app.utils.file_parsing import parse_csv_or_excel_to_rows, CSV_EXCEL_EXTENSIONS, EXCEL_EXTENSIONS
from app.utils.api_helpers import GENERIC_ERROR_MESSAGE, get_json_safe
from app.utils.api_responses import json_bad_request, json_forbidden, json_ok, json_ok_result, json_server_error, json_form_errors, require_json_data, require_json_keys
from app.utils.constants import LOOKUP_ROW_TEMP_ORDER
"""
System Administration Module - Countries, Sectors, Indicator Bank, Translations, and Lookup Lists
"""

from flask import Blueprint, render_template, request, flash, redirect, url_for, current_app, make_response, send_from_directory, abort, send_file
from flask_login import current_user
from app import db
from config import Config
from collections import defaultdict
from datetime import datetime, timedelta
from app.models import (
    Country, Sector, SubSector, IndicatorBank, IndicatorBankHistory, IndicatorSuggestion,
    LookupList, LookupListRow, User, FormTemplate, FormSection, FormItem, FormItemType, FormData, UserSessionLog,
    CommonWord, NationalSociety
)
from app.forms.system import (
    CountryForm, SectorForm, SubSectorForm, IndicatorBankForm, CommonWordForm
)
from app.forms.content import TranslationForm
from app.forms.shared import DeleteForm
from app.routes.admin.shared import admin_required, permission_required, get_localized_sector_name, get_localized_subsector_name, rbac_guard_audit_exempt
from app.utils.request_utils import get_json_or_form, is_json_request
from app.utils.form_localization import get_localized_country_name
from app.utils.file_paths import (
    get_sector_logo_path, get_subsector_logo_path,
    save_system_logo, resolve_sector_logo, resolve_subsector_logo
)
from sqlalchemy import func, desc, or_, and_, inspect
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import json
import os
import uuid
from werkzeug.utils import secure_filename
import pandas as pd
import io

bp = Blueprint("system_admin", __name__, url_prefix="/admin")

# Reuse NS hierarchy WTForms from organization routes for unified edit template
from app.routes.admin.organization import (
    NSBranchForm,
    NSSubBranchForm,
    NSLocalUnitForm,
)

# === Country Management Routes ===
@bp.route("/countries", methods=["GET"])
@permission_required('admin.countries.view')
def manage_countries():
    # Redirect to organization page with countries tab active
    return redirect(url_for('organization.index', tab='countries'))

@bp.route("/countries/new", methods=["GET", "POST"])
@permission_required('admin.countries.edit')
def new_country():
    form = CountryForm()

    if form.validate_on_submit():
        try:
            # Prefer runtime-supported languages; used for translation fields
            translatable_langs = current_app.config.get("TRANSLATABLE_LANGUAGES", []) or []

            new_country = Country(
                name=form.name.data,
                short_name=(form.short_name.data or '').strip() or None,
                iso3=(form.iso3.data or '').upper(),
                status=form.status.data,
                preferred_language=Country.normalize_language_code(form.preferred_language.data),
                currency_code=form.currency_code.data
            )

            # Handle multilingual country name in JSON (name_translations)
            for code in translatable_langs:
                field = getattr(form, f"name_{code}", None)
                if field is not None:
                    new_country.set_name_translation(code, (field.data or "").strip())

            db.session.add(new_country)
            db.session.flush()

            flash(f"Country '{new_country.name}' created successfully.", "success")
            return redirect(url_for("system_admin.manage_countries"))

        except Exception as e:
            request_transaction_rollback()
            flash("An error occurred. Please try again.", "danger")
            current_app.logger.error(f"Error creating country: {e}", exc_info=True)

    return render_template("admin/countries/manage_country.html",
                         form=form,
                         title="Create New Country",
                         country=None)

@bp.route("/countries/<int:country_id>/data", methods=["GET"])
@permission_required('admin.countries.view')
def get_country_data_json(country_id):
    """API endpoint to get full country data for edit modal (used by AJAX)."""
    country = Country.query.get_or_404(country_id)
    return json_ok(
        id=country.id,
        name=country.name,
        short_name=country.short_name or '',
        iso3=country.iso3,
        status=country.status,
        preferred_language=country.preferred_language_code,
        currency_code=country.currency_code,
        name_translations=country.name_translations or {},
    )

@bp.route("/countries/<int:country_id>", methods=["GET"])
@permission_required('admin.countries.view')
def get_country_data(country_id):
    """API endpoint to get country data as JSON"""
    country = Country.query.get_or_404(country_id)
    return json_ok(id=country.id, name=country.name, region=country.region, iso3=country.iso3, status=country.status)

@bp.route("/countries/edit/<int:country_id>", methods=["GET", "POST"])
@permission_required('admin.countries.edit')
def edit_country(country_id):
    country = Country.query.get_or_404(country_id)
    form = CountryForm(request.form, obj=country)
    # Ensure ISO3 uniqueness validator excludes this country when editing
    form.original_country_id = country.id

    if form.validate_on_submit():
        try:
            country.name = form.name.data
            country.short_name = (form.short_name.data or '').strip() or None
            country.iso3 = (form.iso3.data or '').upper()
            country.status = form.status.data
            country.preferred_language = Country.normalize_language_code(form.preferred_language.data)
            country.currency_code = form.currency_code.data

            # Handle multilingual country name in JSON (name_translations)
            translatable_langs = current_app.config.get("TRANSLATABLE_LANGUAGES", []) or []
            for code in translatable_langs:
                field = getattr(form, f"name_{code}", None)
                if field is not None:
                    country.set_name_translation(code, (field.data or "").strip())

            db.session.flush()
            # If this is an AJAX request, return JSON
            if is_json_request():
                return json_ok(
                    message=f"Country '{country.name}' updated successfully.",
                    country={
                        'id': country.id,
                        'name': country.name,
                        'short_name': country.short_name or '',
                        'iso3': country.iso3,
                        'status': country.status,
                        'preferred_language': country.preferred_language_code,
                        'currency_code': country.currency_code,
                        'name_translations': country.name_translations or {},
                    },
                )
            # Fallback for non-AJAX form submissions
            flash(f"Country '{country.name}' updated successfully.", "success")
            return redirect(url_for("system_admin.manage_countries"))

        except Exception as e:
            request_transaction_rollback()
            current_app.logger.error(f"Error updating country {country_id}: {e}", exc_info=True)
            if is_json_request():
                return json_server_error(GENERIC_ERROR_MESSAGE)
            flash("An error occurred. Please try again.", "danger")

    # If validation failed and this is AJAX, return errors
    if request.method == 'POST' and is_json_request():
        return json_form_errors(form, "Validation failed.")

    return render_template("admin/countries/manage_country.html",
                         form=form,
                         country=country,
                         title=f"Edit Country: {country.name}")

@bp.route("/countries/delete/<int:country_id>", methods=["POST"])
@permission_required('admin.countries.edit')
def delete_country(country_id):
    country = Country.query.get_or_404(country_id)

    try:
        # Check if country is used in assignments or user associations
        if country.users.first() or country.assignment_statuses.first():
            flash(f"Cannot delete country '{country.name}' as it is associated with users or assignments.", "danger")
            return redirect(url_for("system_admin.manage_countries"))

        db.session.delete(country)
        db.session.flush()
        flash(f"Country '{country.name}' deleted successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("Error deleting country.", "danger")
        current_app.logger.error(f"Error deleting country {country_id}: {e}", exc_info=True)

    return redirect(url_for("system_admin.manage_countries"))

# === Sector and SubSector Management Routes ===
@bp.route("/sectors_subsectors", methods=["GET"])
@permission_required('admin.organization.manage')
def manage_sectors_subsectors():
    sectors = Sector.query.order_by(Sector.name).all()
    subsectors = SubSector.query.order_by(SubSector.sector_id, SubSector.name).all()
    return render_template("admin/indicator_bank/sectors_subsectors.html",
                         sectors=sectors,
                         subsectors=subsectors,
                         title="Manage Sectors & Sub-Sectors")

@bp.route("/sectors/new", methods=["POST"])
@permission_required('admin.organization.manage')
def new_sector():
    form = SectorForm()

    if form.validate():
        try:
            new_sector = Sector(
                name=form.name.data,
            )

            # Persist translations in JSONB (and sync legacy columns via model helper)
            languages = current_app.config.get("TRANSLATABLE_LANGUAGES", None) or getattr(Config, "TRANSLATABLE_LANGUAGES", []) or []
            for lang in languages:
                field = getattr(form, f"name_{lang}", None)
                if field is not None:
                    new_sector.set_name_translation(lang, field.data or "")

            # Handle logo upload
            if form.logo_file.data:
                logo_filename = _save_logo_file(
                    form.logo_file.data,
                    get_sector_logo_path(),
                    form.name.data,
                    'sector'
                )
                if logo_filename:
                    new_sector.logo_filename = logo_filename

            db.session.add(new_sector)
            db.session.flush()

            flash(f"Sector '{new_sector.name}' created successfully.", "success")

        except Exception as e:
            request_transaction_rollback()
            flash("An error occurred. Please try again.", "danger")
            current_app.logger.error(f"Error creating sector: {e}", exc_info=True)
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"Error in {field}: {error}", "danger")

    return redirect(url_for("system_admin.manage_sectors_subsectors"))

@bp.route("/sectors/edit/<int:sector_id>", methods=["POST"])
@permission_required('admin.organization.manage')
def edit_sector(sector_id):
    sector = Sector.query.get_or_404(sector_id)
    form = SectorForm(original_sector_id=sector_id)


    if form.validate():
        try:
            sector.name = form.name.data

            languages = current_app.config.get("TRANSLATABLE_LANGUAGES", None) or getattr(Config, "TRANSLATABLE_LANGUAGES", []) or []
            for lang in languages:
                field = getattr(form, f"name_{lang}", None)
                if field is not None:
                    sector.set_name_translation(lang, field.data or "")

            # Handle logo update
            if form.logo_file.data:
                # Delete old logo
                if sector.logo_filename:
                    _delete_logo_file(get_sector_logo_path(), sector.logo_filename)

                # Save new logo
                logo_filename = _save_logo_file(
                    form.logo_file.data,
                    get_sector_logo_path(),
                    sector.name,
                    'sector'
                )
                if logo_filename:
                    sector.logo_filename = logo_filename

            db.session.flush()
            flash(f"Sector '{sector.name}' updated successfully.", "success")

        except Exception as e:
            request_transaction_rollback()
            flash("An error occurred. Please try again.", "danger")
            current_app.logger.error(f"Error updating sector {sector_id}: {e}", exc_info=True)
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"Error in {field}: {error}", "danger")

    return redirect(url_for("system_admin.manage_sectors_subsectors"))

@bp.route("/sectors/<int:sector_id>", methods=["GET"])
@permission_required('admin.organization.manage')
def get_sector(sector_id):
    """API endpoint to get sector data for editing"""
    sector = Sector.query.get_or_404(sector_id)

    return json_ok(
        id=sector.id,
        name=sector.name,
        description=sector.description or '',
        display_order=sector.display_order or 0,
        icon_class=sector.icon_class or '',
        logo_filename=sector.logo_filename or '',
        is_active=sector.is_active,
        name_translations=sector.name_translations or {},
    )

@bp.route("/sectors/delete/<int:sector_id>", methods=["POST"])
@permission_required('admin.organization.manage')
def delete_sector(sector_id):
    sector = Sector.query.get_or_404(sector_id)

    try:
        # Check if sector has subsectors
        if sector.subsectors.first():
            flash(f"Cannot delete sector '{sector.name}' as it has associated sub-sectors.", "danger")
            return redirect(url_for("system_admin.manage_sectors_subsectors"))

        # Delete logo file
        if sector.logo_filename:
            _delete_logo_file(get_sector_logo_path(), sector.logo_filename)

        db.session.delete(sector)
        db.session.flush()
        flash(f"Sector '{sector.name}' deleted successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("An error occurred. Please try again.", "danger")
        current_app.logger.error(f"Error deleting sector {sector_id}: {e}", exc_info=True)

    return redirect(url_for("system_admin.manage_sectors_subsectors"))

@bp.route("/subsectors/new", methods=["POST"])
@permission_required('admin.organization.manage')
def new_subsector():
    form = SubSectorForm()

    if form.validate():
        try:
            new_subsector = SubSector(
                name=form.name.data,
                sector_id=form.sector_id.data
            )

            languages = current_app.config.get("TRANSLATABLE_LANGUAGES", None) or getattr(Config, "TRANSLATABLE_LANGUAGES", []) or []
            for lang in languages:
                field = getattr(form, f"name_{lang}", None)
                if field is not None:
                    new_subsector.set_name_translation(lang, field.data or "")

            # Handle logo upload
            if form.logo_file.data:
                logo_filename = _save_logo_file(
                    form.logo_file.data,
                    get_subsector_logo_path(),
                    form.name.data,
                    'subsector'
                )
                if logo_filename:
                    new_subsector.logo_filename = logo_filename

            db.session.add(new_subsector)
            db.session.flush()

            flash(f"Sub-sector '{new_subsector.name}' created successfully.", "success")

        except Exception as e:
            request_transaction_rollback()
            flash("An error occurred. Please try again.", "danger")
            current_app.logger.error(f"Error creating subsector: {e}", exc_info=True)
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"Error in {field}: {error}", "danger")

    return redirect(url_for("system_admin.manage_sectors_subsectors"))

@bp.route("/subsectors/edit/<int:subsector_id>", methods=["POST"])
@permission_required('admin.organization.manage')
def edit_subsector(subsector_id):
    subsector = SubSector.query.get_or_404(subsector_id)
    form = SubSectorForm(original_subsector_id=subsector_id)

    if form.validate():
        try:
            subsector.name = form.name.data
            languages = current_app.config.get("TRANSLATABLE_LANGUAGES", None) or getattr(Config, "TRANSLATABLE_LANGUAGES", []) or []
            for lang in languages:
                field = getattr(form, f"name_{lang}", None)
                if field is not None:
                    subsector.set_name_translation(lang, field.data or "")
            subsector.sector_id = form.sector_id.data

            # Handle logo update
            if form.logo_file.data:
                # Delete old logo
                if subsector.logo_filename:
                    _delete_logo_file(get_subsector_logo_path(), subsector.logo_filename)

                # Save new logo
                logo_filename = _save_logo_file(
                    form.logo_file.data,
                    get_subsector_logo_path(),
                    subsector.name,
                    'subsector'
                )
                if logo_filename:
                    subsector.logo_filename = logo_filename

            db.session.flush()
            flash(f"Sub-sector '{subsector.name}' updated successfully.", "success")

        except Exception as e:
            request_transaction_rollback()
            flash("An error occurred. Please try again.", "danger")
            current_app.logger.error(f"Error updating subsector {subsector_id}: {e}", exc_info=True)
    else:
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"Error in {field}: {error}", "danger")

    return redirect(url_for("system_admin.manage_sectors_subsectors"))

@bp.route("/subsectors/<int:subsector_id>", methods=["GET"])
@permission_required('admin.organization.manage')
def get_subsector(subsector_id):
    """API endpoint to get subsector data for editing"""
    subsector = SubSector.query.get_or_404(subsector_id)

    return json_ok(
        id=subsector.id,
        name=subsector.name,
        description=subsector.description or '',
        sector_id=subsector.sector_id,
        display_order=subsector.display_order or 0,
        icon_class=subsector.icon_class or '',
        logo_filename=subsector.logo_filename or '',
        is_active=subsector.is_active,
        name_translations=subsector.name_translations or {},
    )

@bp.route("/subsectors/delete/<int:subsector_id>", methods=["POST"])
@permission_required('admin.organization.manage')
def delete_subsector(subsector_id):
    subsector = SubSector.query.get_or_404(subsector_id)

    try:
        # Delete logo file
        if subsector.logo_filename:
            _delete_logo_file(get_subsector_logo_path(), subsector.logo_filename)

        db.session.delete(subsector)
        db.session.flush()
        flash(f"Sub-sector '{subsector.name}' deleted successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("An error occurred. Please try again.", "danger")
        current_app.logger.error(f"Error deleting subsector {subsector_id}: {e}", exc_info=True)

    return redirect(url_for("system_admin.manage_sectors_subsectors"))

# === Indicator Bank Management Routes ===
@bp.route("/indicator_bank", methods=["GET"])
@permission_required('admin.indicator_bank.view')
def manage_indicator_bank():
    # Get search parameters
    search = request.args.get('search', '')
    sector_filter = request.args.get('sector', '')
    type_filter = request.args.get('type', '')

    # Build query
    query = IndicatorBank.query

    if search:
        query = query.filter(
            or_(
                IndicatorBank.name.contains(search),
                IndicatorBank.definition.contains(search),
                IndicatorBank.type.contains(search)
            )
        )

    if sector_filter:
        query = query.filter(IndicatorBank.sector.contains(sector_filter))

    if type_filter:
        query = query.filter(IndicatorBank.type == type_filter)

    # Get all indicators (no server-side pagination - AG Grid handles pagination client-side)
    indicators = query.order_by(IndicatorBank.name).all()
    total_count = len(indicators)

    # Prefetch sector and subsector data to avoid N+1 queries
    sector_ids = set()
    subsector_ids = set()

    for indicator in indicators:
        if indicator.sector:
            for level in ['primary', 'secondary', 'tertiary']:
                if level in indicator.sector:
                    sector_ids.add(indicator.sector[level])
        if indicator.sub_sector:
            for level in ['primary', 'secondary', 'tertiary']:
                if level in indicator.sub_sector:
                    subsector_ids.add(indicator.sub_sector[level])

    # Fetch all sectors and subsectors in bulk
    sectors_dict = {}
    if sector_ids:
        sectors = Sector.query.filter(Sector.id.in_(sector_ids)).all()
        sectors_dict = {sector.id: sector for sector in sectors}

    subsectors_dict = {}
    if subsector_ids:
        subsectors = SubSector.query.filter(SubSector.id.in_(subsector_ids)).all()
        subsectors_dict = {subsector.id: subsector for subsector in subsectors}

    # Prefetch usage counts to avoid N+1 queries
    from sqlalchemy import func, select
    from app.models.form_items import FormItem
    indicator_ids = [ind.id for ind in indicators]
    usage_counts = {}
    if indicator_ids:
        usage_subquery = db.session.query(
            FormItem.indicator_bank_id,
            func.count(FormItem.id).label('count')
        ).filter(
            FormItem.indicator_bank_id.in_(indicator_ids)
        ).group_by(FormItem.indicator_bank_id).all()
        usage_counts = {row.indicator_bank_id: row.count for row in usage_subquery}

    # Cache the sector and subsector data on each indicator, and prefetch usage counts
    for indicator in indicators:
        indicator._cached_sectors = {}
        indicator._cached_subsectors = {}
        indicator._cached_usage_count = usage_counts.get(indicator.id, 0)

        if indicator.sector:
            for level in ['primary', 'secondary', 'tertiary']:
                if level in indicator.sector:
                    sector_id = indicator.sector[level]
                    if sector_id in sectors_dict:
                        indicator._cached_sectors[level] = sectors_dict[sector_id]

        if indicator.sub_sector:
            for level in ['primary', 'secondary', 'tertiary']:
                if level in indicator.sub_sector:
                    subsector_id = indicator.sub_sector[level]
                    if subsector_id in subsectors_dict:
                        indicator._cached_subsectors[level] = subsectors_dict[subsector_id]

    # Return JSON for API requests (mobile app)
    if is_json_request():
        indicators_data = []
        for indicator in indicators:
            # Get sector and subsector names
            sector_name = None
            subsector_name = None
            if indicator.sector:
                for level in ['primary', 'secondary', 'tertiary']:
                    if level in indicator.sector and indicator.sector[level]:
                        sector_id = indicator.sector[level]
                        if sector_id in sectors_dict:
                            sector_name = sectors_dict[sector_id].name
                            break
            if indicator.sub_sector:
                for level in ['primary', 'secondary', 'tertiary']:
                    if level in indicator.sub_sector and indicator.sub_sector[level]:
                        subsector_id = indicator.sub_sector[level]
                        if subsector_id in subsectors_dict:
                            subsector_name = subsectors_dict[subsector_id].name
                            break

            indicators_data.append({
                'id': indicator.id,
                'name': indicator.name or '',
                'definition': indicator.definition if hasattr(indicator, 'definition') else None,
                'type': indicator.type if hasattr(indicator, 'type') else None,
                'sector': sector_name,
                'sub_sector': subsector_name,
                'unit': indicator.unit if hasattr(indicator, 'unit') else None,
                'fdrs_kpi_code': getattr(indicator, 'fdrs_kpi_code', None),
                'usage_count': usage_counts.get(indicator.id, 0),
            })
        return json_ok(indicators=indicators_data, count=len(indicators_data), total_count=total_count)

    # Get filter options
    sectors = db.session.query(Sector.name).distinct().order_by(Sector.name).all()
    types = db.session.query(IndicatorBank.type).distinct().filter(IndicatorBank.type.isnot(None)).order_by(IndicatorBank.type).all()

    # Count pending indicator suggestions
    # Check for 'Pending', 'pending', or 'Pending Review' (in case status was stored incorrectly)
    pending_suggestions_count = IndicatorSuggestion.query.filter(
        or_(
            func.lower(IndicatorSuggestion.status) == 'pending',
            IndicatorSuggestion.status == 'Pending Review'
        )
    ).count()

    return render_template("admin/indicator_bank/indicator_bank.html",
                         indicators=indicators,
                         sectors=[s[0] for s in sectors],
                         types=[t[0] for t in types if t[0]],
                         search=search,
                         sector_filter=sector_filter,
                         type_filter=type_filter,
                         title="Manage Indicator Bank",
                         total_count=total_count,
                         pending_suggestions_count=pending_suggestions_count)


@bp.route("/indicator_bank/neural_map", methods=["GET"])
@permission_required('admin.indicator_bank.view')
def indicator_bank_neural_map():
    """Interactive neural map visualization of indicator bank vector embeddings."""
    return render_template("admin/indicator_bank/neural_map.html", title="Indicators Neural Map")


@bp.route("/indicator_bank/neural_map/data", methods=["GET"])
@permission_required('admin.indicator_bank.view')
def indicator_bank_neural_map_data():
    """JSON: 2D/3D scatter data for the indicators neural map."""
    from app.services.indicator_neural_map import build_embedding_scatter
    max_nodes = request.args.get("max_nodes", type=int) or 5000
    n_neighbors = request.args.get("neighbors", type=int) or 8
    method = request.args.get("method", "pca")
    dimensions = request.args.get("dimensions", 3, type=int)
    if method not in ("pca", "tsne"):
        method = "pca"
    if dimensions not in (2, 3):
        dimensions = 3
    try:
        result = build_embedding_scatter(
            max_nodes=max_nodes,
            n_neighbors=n_neighbors,
            method=method,
            dimensions=dimensions,
            exclude_archived=True,
        )
        return json_ok_result(result)
    except Exception as e:
        current_app.logger.exception("Neural map data failed")
        return json_server_error(GENERIC_ERROR_MESSAGE, nodes=[], groups=[], count=0)


@bp.route("/indicator_bank/neural_map/probe", methods=["POST"])
@permission_required('admin.indicator_bank.view')
def indicator_bank_neural_map_probe():
    """Embed a free-text query and return nearest indicator neighbours."""
    from app.services.indicator_neural_map import probe_query_embedding
    data = get_json_safe()
    query = (data.get("query") or "").strip()
    if not query:
        return json_bad_request("query is required")
    try:
        result = probe_query_embedding(query, top_k=10)
        return json_ok_result(result)
    except Exception as e:
        current_app.logger.exception("Neural map probe failed")
        return json_server_error(GENERIC_ERROR_MESSAGE)


@bp.route("/indicator_bank/sync_remote", methods=["POST"])
@permission_required("admin.indicator_bank.edit")
def sync_indicator_bank_remote():
    """Trigger background sync from the external IFRC Indicator Bank platform."""
    from app.services.indicatorbank_remote_sync_service import start_remote_sync, get_remote_sync_state

    api_key = os.getenv("IFRC_INDICATORBANK_API_KEY", "").strip()
    api_url = os.getenv("IFRC_INDICATORBANK_API_URL", "https://ifrc-indicatorbank.azurewebsites.net/api/indicator").strip()
    data = get_json_or_form()
    limit_raw = data.get("limit") or request.form.get("limit")
    try:
        limit = int(limit_raw) if limit_raw not in (None, "", "null") else None
    except Exception as e:
        current_app.logger.debug("indicator bank sync limit parse failed: %s", e)
        limit = None

    ok, msg = start_remote_sync(current_app._get_current_object(), api_url=api_url, api_key=api_key, limit=limit)
    return (json_ok(success=bool(ok), message=msg, state=get_remote_sync_state())
            if ok else json_bad_request(msg, success=False, state=get_remote_sync_state()))


@bp.route("/indicator_bank/sync_remote/status", methods=["GET"])
@permission_required("admin.indicator_bank.edit")
def sync_indicator_bank_remote_status():
    """Fetch background sync status/result."""
    from app.services.indicatorbank_remote_sync_service import get_remote_sync_state
    return json_ok(state=get_remote_sync_state())

@bp.route("/indicator_bank/view/<int:id>", methods=["GET"])
@permission_required('admin.indicator_bank.view')
def view_indicator_bank(id):
    indicator = IndicatorBank.query.get_or_404(id)
    return render_template("admin/indicator_bank/view_indicator_bank.html",
                         indicator=indicator,
                         title=f"View Indicator: {indicator.name}")

@bp.route("/indicator_bank/add", methods=["GET", "POST"])
@permission_required('admin.indicator_bank.create')
def add_indicator_bank():
    # Archive is a separate permission; don't allow archiving via "create" flows unless permitted.
    try:
        from app.services.authorization_service import AuthorizationService
        can_archive = (
            AuthorizationService.is_system_manager(current_user)
            or AuthorizationService.has_rbac_permission(current_user, 'admin.indicator_bank.archive')
        )
    except Exception as e:
        current_app.logger.debug("can_archive check failed (add_indicator_bank): %s", e)
        can_archive = False

    if request.method == 'POST':
        # For POST requests, create form with form data
        form = IndicatorBankForm(request.form)
    else:
        # For GET requests, create empty form
        form = IndicatorBankForm()

    # Always ensure choices are populated
    form._populate_choices()

    if form.validate_on_submit():
        try:
            new_indicator = IndicatorBank(
                name=form.name.data,
                definition=form.definition.data,
                type=form.type.data,
                unit=form.unit.data,
                fdrs_kpi_code=(form.fdrs_kpi_code.data or '').strip() or None,
                emergency=form.emergency.data,
                related_programs=form.related_programs.data,
                archived=(form.archived.data if can_archive else False)
            )

            # Use form's populate method to handle sector and sub-sector data
            form.populate_indicator_bank(new_indicator)

            # Enforce archive permission server-side (in case populate touches it).
            if not can_archive:
                new_indicator.archived = False

            db.session.add(new_indicator)
            db.session.flush()  # Get the ID of the new indicator

            # Create history record for the new indicator
            history = IndicatorBankHistory(
                indicator_bank_id=new_indicator.id,
                user_id=current_user.id,
                name=new_indicator.name,
                type=new_indicator.type,
                unit=new_indicator.unit,
                fdrs_kpi_code=new_indicator.fdrs_kpi_code,
                definition=new_indicator.definition,
                name_translations=new_indicator.name_translations,
                definition_translations=new_indicator.definition_translations,
                archived=new_indicator.archived,
                comments=new_indicator.comments,
                emergency=new_indicator.emergency,
                related_programs=new_indicator.related_programs,
                sector=new_indicator.sector,
                sub_sector=new_indicator.sub_sector,
                change_type='CREATED',
                change_description=f'Indicator "{new_indicator.name}" created by {current_user.name or current_user.email}'
            )
            db.session.add(history)

            db.session.flush()

            flash(f"Indicator '{new_indicator.name}' added successfully.", "success")
            return redirect(url_for("system_admin.manage_indicator_bank"))

        except Exception as e:
            request_transaction_rollback()
            flash("Error adding indicator.", "danger")
            current_app.logger.error(f"Error adding indicator: {e}", exc_info=True)

    return render_template("admin/indicator_bank/add_indicator_bank.html",
                         form=form,
                         title="Add New Indicator")

def track_indicator_changes(old_indicator, new_form_data, user):
    """
    Track specific changes between old indicator data and new form data.
    Returns a list of change descriptions.
    """
    from app.models import Sector, SubSector
    changes = []

    # Define fields to track
    fields_to_track = [
        ('name', 'Name'),
        ('type', 'Type'),
        ('unit', 'Unit'),
        ('fdrs_kpi_code', 'FDRS KPI Code'),
        ('definition', 'Definition'),
        # Translations are stored in JSONB; track the full dict rather than hardcoded language columns
        ('name_translations', 'Name Translations'),
        ('definition_translations', 'Definition Translations'),
        ('comments', 'Comments'),
        ('related_programs', 'Related Programs'),
        ('emergency', 'Emergency'),
        ('archived', 'Archived'),
        ('sector_primary', 'Sector Primary'),
        ('sector_secondary', 'Sector Secondary'),
        ('sector_tertiary', 'Sector Tertiary'),
        ('sub_sector_primary', 'Sub-Sector Primary'),
        ('sub_sector_secondary', 'Sub-Sector Secondary'),
        ('sub_sector_tertiary', 'Sub-Sector Tertiary')
    ]

    for field_name, display_name in fields_to_track:
        old_value = getattr(old_indicator, field_name, None)
        new_value = new_form_data.get(field_name)

        # Handle JSON translation dict fields sent as JSON strings from forms
        if field_name in ("name_translations", "definition_translations"):
            try:
                if isinstance(new_value, str) and new_value.strip():
                    import json as _json
                    parsed = _json.loads(new_value)
                    new_value = parsed if isinstance(parsed, dict) else None
                else:
                    new_value = None
            except Exception as e:
                current_app.logger.debug("name_translations/definition_translations parse failed: %s", e)
                new_value = None

        # Handle boolean fields
        if field_name in ['emergency', 'archived']:
            old_value = bool(old_value)
            new_value = bool(new_value)

        # Handle sector and sub_sector fields - compare with the corresponding level in the indicator's JSON
        if field_name.startswith('sector_'):
            level = field_name.replace('sector_', '')
            old_value = old_indicator.sector.get(level) if old_indicator.sector else None
        elif field_name.startswith('sub_sector_'):
            level = field_name.replace('sub_sector_', '')
            old_value = old_indicator.sub_sector.get(level) if old_indicator.sub_sector else None

        # Normalize empty values to None for comparison
        if new_value == '' or new_value == 'None':
            new_value = None
        if old_value == '' or old_value == 'None':
            old_value = None

        # Skip tracking if both values are None/empty
        if old_value is None and new_value is None:
            continue

        # Compare values
        if old_value != new_value:
            # Handle type field case sensitivity
            if field_name == 'type':
                # Normalize type values for comparison
                old_type = str(old_value).lower() if old_value else None
                new_type = str(new_value).lower() if new_value else None
                if old_type == new_type:
                    continue  # Skip if they're the same after normalization

            # Get readable names for sector and sub_sector fields
            if field_name.startswith('sector_') and new_value:
                try:
                    sector = Sector.query.get(new_value)
                    new_display_value = sector.name if sector else f"ID: {new_value}"
                except Exception as e:
                    current_app.logger.debug("sector lookup for new_value failed: %s", e)
                    new_display_value = f"ID: {new_value}"
            elif field_name.startswith('sub_sector_') and new_value:
                try:
                    subsector = SubSector.query.get(new_value)
                    new_display_value = subsector.name if subsector else f"ID: {new_value}"
                except Exception as e:
                    current_app.logger.debug("subsector lookup for new_value failed: %s", e)
                    new_display_value = f"ID: {new_value}"
            else:
                new_display_value = str(new_value) if new_value is not None else None

            if field_name.startswith('sector_') and old_value:
                try:
                    sector = Sector.query.get(old_value)
                    old_display_value = sector.name if sector else f"ID: {old_value}"
                except Exception as e:
                    current_app.logger.debug("sector lookup for old_value failed: %s", e)
                    old_display_value = f"ID: {old_value}"
            elif field_name.startswith('sub_sector_') and old_value:
                try:
                    subsector = SubSector.query.get(old_value)
                    old_display_value = subsector.name if subsector else f"ID: {old_value}"
                except Exception as e:
                    current_app.logger.debug("subsector lookup for old_value failed: %s", e)
                    old_display_value = f"ID: {old_value}"
            else:
                old_display_value = str(old_value) if old_value is not None else None

            # Format the change description
            if old_value is None and new_value is not None:
                changes.append(f"{display_name}: Added '{new_display_value}'")
            elif old_value is not None and new_value is None:
                changes.append(f"{display_name}: Removed '{old_display_value}'")
            elif old_value != new_value:
                # Truncate long values for readability
                old_display = old_display_value[:50] + "..." if len(str(old_display_value)) > 50 else str(old_display_value)
                new_display = new_display_value[:50] + "..." if len(str(new_display_value)) > 50 else str(new_display_value)
                changes.append(f"{display_name}: Changed from '{old_display}' to '{new_display}'")

    return changes

@bp.route("/indicator_bank/edit/<int:id>", methods=["GET", "POST"])
@permission_required('admin.indicator_bank.edit')
def edit_indicator_bank(id):
    indicator = IndicatorBank.query.get_or_404(id)

    # Archive is a separate permission; don't allow archiving via "edit" flows.
    try:
        from app.services.authorization_service import AuthorizationService
        can_archive = (
            AuthorizationService.is_system_manager(current_user)
            or AuthorizationService.has_rbac_permission(current_user, 'admin.indicator_bank.archive')
        )
    except Exception as e:
        current_app.logger.debug("can_archive check failed (edit_indicator_bank): %s", e)
        can_archive = False

    if request.method == 'POST':
        # Check if this is an AJAX/API request
        is_ajax = is_json_request()

        if is_ajax:
            # Handle AJAX request for auto-translate updates
            try:
                data = get_json_safe()
                err = require_json_data(data)
                if err:
                    return err

                # Prevent privilege escalation: block attempts to toggle archive state via AJAX.
                if 'archived' in data and not can_archive:
                    return json_forbidden('Archive permission required.')



                # Update only the provided fields
                updated_fields = []
                for field_name, value in data.items():


                    if value is not None:
                        # Prevent privilege escalation: block archive toggles via edit endpoint.
                        if field_name == 'archived' and not can_archive:
                            continue
                        # Handle translation fields specially for JSONB
                        if field_name.startswith('name_') and field_name != 'name':
                            # Extract language from field name (e.g., 'name_fr' -> 'fr')
                            language = field_name.replace('name_', '')

                            indicator.set_name_translation(language, value)
                            updated_fields.append(field_name)
                        elif field_name.startswith('definition_') and field_name != 'definition':
                            # Extract language from field name (e.g., 'definition_fr' -> 'fr')
                            language = field_name.replace('definition_', '')

                            indicator.set_definition_translation(language, value)
                            updated_fields.append(field_name)
                        elif hasattr(indicator, field_name):
                            # Handle regular fields normally
                            setattr(indicator, field_name, value)
                            updated_fields.append(field_name)
                        else:
                            # Field doesn't exist, skip it
                            continue

                # CRITICAL FIX: Mark the indicator as modified so SQLAlchemy knows to save it
                if updated_fields:
                    # Force SQLAlchemy to detect JSONB field changes
                    from sqlalchemy.orm.attributes import flag_modified

                    # Mark JSONB fields as modified
                    if any(field.startswith('name_') for field in updated_fields):
                        flag_modified(indicator, 'name_translations')
                    if any(field.startswith('definition_') for field in updated_fields):
                        flag_modified(indicator, 'definition_translations')

                    db.session.add(indicator)

                if updated_fields:
                    # Create history record for the update
                    history = IndicatorBankHistory(
                        indicator_bank_id=indicator.id,
                        user_id=current_user.id,
                        name=indicator.name,
                        type=indicator.type,
                        unit=indicator.unit,
                        fdrs_kpi_code=indicator.fdrs_kpi_code,
                        definition=indicator.definition,
                        name_translations=indicator.name_translations,
                        definition_translations=indicator.definition_translations,
                        archived=indicator.archived,
                        comments=indicator.comments,
                        emergency=indicator.emergency,
                        related_programs=indicator.related_programs,
                        sector=indicator.sector,
                        sub_sector=indicator.sub_sector,
                        change_type='UPDATED',
                        change_description=f"Auto-translate update: {', '.join(updated_fields)} updated by {current_user.name or current_user.email}"
                    )
                    db.session.add(history)
                    # Commit both the indicator changes and the history record
                    db.session.flush()

                    return json_ok(
                        message=f'Indicator updated successfully. Updated fields: {", ".join(updated_fields)}'
                    )
                else:
                    return json_bad_request('No valid fields to update')

            except Exception as e:
                request_transaction_rollback()
                current_app.logger.error(f"Error updating indicator {id} via AJAX: {e}", exc_info=True)
                return json_server_error(GENERIC_ERROR_MESSAGE)

        # Regular form submission
        form = IndicatorBankForm(request.form)

    else:
        # For GET requests, create form with object
        form = IndicatorBankForm(obj=indicator)

    # Always ensure choices are populated
    form._populate_choices()

    # Populate form with existing data (only for GET requests)
    if request.method == 'GET':
        form.populate_from_indicator_bank(indicator)

    if form.validate_on_submit():
        try:
            # Get form data for change tracking
            try:
                from flask import current_app
                _langs = current_app.config.get("TRANSLATABLE_LANGUAGES") or []
            except Exception as e:
                current_app.logger.debug("TRANSLATABLE_LANGUAGES config failed: %s", e)
                _langs = []

            name_translations = {"en": form.name.data}
            for _lang in _langs:
                _lang = str(_lang or "").strip().lower()
                if not _lang or _lang == "en":
                    continue
                _field = getattr(form, f"name_{_lang}", None)
                if _field is not None:
                    name_translations[_lang] = _field.data

            # If user can't archive, ignore any submitted archived changes.
            prev_archived = indicator.archived
            if not can_archive:
                try:
                    form.archived.data = prev_archived
                except Exception as e:
                    current_app.logger.debug("form.archived reset failed: %s", e)

            form_data = {
                'name': form.name.data,
                'type': form.type.data,
                'unit': form.unit.data,
                'fdrs_kpi_code': (form.fdrs_kpi_code.data or '').strip() or None,
                'definition': form.definition.data,
                'name_translations': name_translations,
                'comments': form.comments.data,
                'related_programs': form.related_programs.data,
                'emergency': form.emergency.data,
                'archived': prev_archived if not can_archive else form.archived.data,
                'sector_primary': form.sector_primary.data,
                'sector_secondary': form.sector_secondary.data,
                'sector_tertiary': form.sector_tertiary.data,
                'sub_sector_primary': form.sub_sector_primary.data,
                'sub_sector_secondary': form.sub_sector_secondary.data,
                'sub_sector_tertiary': form.sub_sector_tertiary.data
            }

            # Track changes
            changes = track_indicator_changes(indicator, form_data, current_user)

            # Create change description
            if changes:
                change_description = "; ".join(changes)
            else:
                change_description = f"Indicator updated by {current_user.name or current_user.email} (no specific changes detected)"

            # Create history record before updating
            history = IndicatorBankHistory(
                indicator_bank_id=indicator.id,
                user_id=current_user.id,
                name=indicator.name,
                type=indicator.type,
                unit=indicator.unit,
                fdrs_kpi_code=indicator.fdrs_kpi_code,
                definition=indicator.definition,
                name_translations=indicator.name_translations,
                definition_translations=indicator.definition_translations,
                archived=indicator.archived,
                comments=indicator.comments,
                emergency=indicator.emergency,
                related_programs=indicator.related_programs,
                sector=indicator.sector,
                sub_sector=indicator.sub_sector,
                change_type='UPDATED',
                change_description=change_description
            )
            db.session.add(history)

            # Update indicator using form's populate method
            form.populate_indicator_bank(indicator)

            # Enforce archive permission server-side (in case populate changed it).
            if not can_archive:
                indicator.archived = prev_archived

            db.session.flush()
            flash(f"Indicator '{indicator.name}' updated successfully.", "success")
            return redirect(url_for("system_admin.manage_indicator_bank"))

        except Exception as e:
            request_transaction_rollback()
            flash("An error occurred. Please try again.", "danger")
            current_app.logger.error(f"Error updating indicator {id}: {e}", exc_info=True)

    return render_template("admin/indicator_bank/edit_indicator_bank.html",
                         form=form,
                         indicator=indicator,
                         title=f"Edit Indicator: {indicator.name}")

@bp.route("/indicator_bank/delete/<int:id>", methods=["POST"])
@permission_required('admin.indicator_bank.edit')
def delete_indicator_bank(id):
    indicator = IndicatorBank.query.get_or_404(id)

    try:
        # Check if indicator is used in form items
        # (This depends on your FormItem model structure)

        # Create history record
        history = IndicatorBankHistory(
            indicator_bank_id=indicator.id,
            user_id=current_user.id,
            name=indicator.name,
            type=indicator.type,
            unit=indicator.unit,
            fdrs_kpi_code=indicator.fdrs_kpi_code,
            definition=indicator.definition,
            name_translations=indicator.name_translations,
            definition_translations=indicator.definition_translations,
            archived=indicator.archived,
            comments=indicator.comments,
            emergency=indicator.emergency,
            related_programs=indicator.related_programs,
            sector=indicator.sector,
            sub_sector=indicator.sub_sector,
            change_type='DELETED',
            change_description=f'Indicator "{indicator.name}" deleted by {current_user.name or current_user.email}'
        )
        db.session.add(history)

        db.session.delete(indicator)
        db.session.flush()

        flash(f"Indicator '{indicator.name}' deleted successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("An error occurred. Please try again.", "danger")
        current_app.logger.error(f"Error deleting indicator {id}: {e}", exc_info=True)

    return redirect(url_for("system_admin.manage_indicator_bank"))

@bp.route("/indicator_bank/archive/<int:id>", methods=["POST"])
@permission_required('admin.indicator_bank.archive')
def archive_indicator_bank(id):
    indicator = IndicatorBank.query.get_or_404(id)

    try:
        # Toggle archived status
        new_archived_status = not indicator.archived

        # Create history record
        history = IndicatorBankHistory(
            indicator_bank_id=indicator.id,
            user_id=current_user.id,
            name=indicator.name,
            type=indicator.type,
            unit=indicator.unit,
            fdrs_kpi_code=indicator.fdrs_kpi_code,
            definition=indicator.definition,
            name_translations=indicator.name_translations,
            definition_translations=indicator.definition_translations,
            archived=indicator.archived,
            comments=indicator.comments,
            emergency=indicator.emergency,
            related_programs=indicator.related_programs,
            sector=indicator.sector,
            sub_sector=indicator.sub_sector,
            change_type='ARCHIVED' if new_archived_status else 'UNARCHIVED',
            change_description=f'Indicator "{indicator.name}" {"archived" if new_archived_status else "unarchived"} by {current_user.name or current_user.email}'
        )
        db.session.add(history)

        # Update the archived status
        indicator.archived = new_archived_status
        db.session.flush()

        action = "archived" if new_archived_status else "unarchived"
        flash(f"Indicator '{indicator.name}' {action} successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("Error archiving indicator.", "danger")
        current_app.logger.error(f"Error archiving indicator {id}: {e}", exc_info=True)

    return redirect(url_for("system_admin.manage_indicator_bank"))

@bp.route("/indicator_bank/translations/<int:id>", methods=["POST"])
@permission_required('admin.indicator_bank.edit')
def update_indicator_translations(id):
    indicator = IndicatorBank.query.get_or_404(id)

    try:
        # Get translation data from form
        translation_changes = []
        languages = current_app.config.get("TRANSLATABLE_LANGUAGES", None) or getattr(Config, "TRANSLATABLE_LANGUAGES", []) or []

        for lang in languages:
            name_key = f'name_{lang}'
            definition_key = f'definition_{lang}'

            # Check name translations
            if name_key in request.form:
                new_name = request.form[name_key].strip()
                old_name = indicator.get_name_translation(lang)
                if old_name != new_name:
                    if old_name is None and new_name:
                        translation_changes.append(f"{lang.upper()} Name: Added '{new_name}'")
                    elif old_name and new_name != old_name:
                        translation_changes.append(f"{lang.upper()} Name: Changed from '{old_name}' to '{new_name}'")
                    indicator.set_name_translation(lang, new_name)

            # Check definition translations
            if definition_key in request.form:
                new_definition = request.form[definition_key].strip()
                old_definition = indicator.get_definition_translation(lang)
                if old_definition != new_definition:
                    if old_definition is None and new_definition:
                        translation_changes.append(f"{lang.upper()} Definition: Added '{new_definition[:50]}{'...' if len(new_definition) > 50 else ''}'")
                    elif old_definition and new_definition != old_definition:
                        translation_changes.append(f"{lang.upper()} Definition: Changed from '{old_definition[:50]}{'...' if len(old_definition) > 50 else ''}' to '{new_definition[:50]}{'...' if len(new_definition) > 50 else ''}'")
                    indicator.set_definition_translation(lang, new_definition)

        # Create change description
        if translation_changes:
            change_description = "; ".join(translation_changes)
        else:
            change_description = f"Translations updated by {current_user.name or current_user.email} (no specific changes detected)"

        # Create history record
        history = IndicatorBankHistory(
            indicator_bank_id=indicator.id,
            user_id=current_user.id,
            name=indicator.name,
            type=indicator.type,
            unit=indicator.unit,
            fdrs_kpi_code=indicator.fdrs_kpi_code,
            definition=indicator.definition,
            name_translations=indicator.name_translations,
            definition_translations=indicator.definition_translations,
            archived=indicator.archived,
            comments=indicator.comments,
            emergency=indicator.emergency,
            related_programs=indicator.related_programs,
            sector=indicator.sector,
            sub_sector=indicator.sub_sector,
            change_type='UPDATED',
            change_description=change_description
        )
        db.session.add(history)

        db.session.flush()
        flash("Translations updated successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("An error occurred. Please try again.", "danger")
        current_app.logger.error(f"Error updating indicator translations {id}: {e}", exc_info=True)

    return redirect(url_for("system_admin.view_indicator_bank", id=id))

# === Lookup List Management Routes ===
@bp.route("/lists", methods=["GET"])
@permission_required('admin.templates.edit')
def manage_lists():
    """Manage lookup lists"""
    lists = LookupList.query.order_by(LookupList.name).all()

    # Define system lists that read from their own tables
    system_lists = []

    # Country Map system list
    country_count = Country.query.count()
    system_lists.append({
        'name': 'Country Map',
        'description': 'List of all countries in the system',
        'columns': 4,  # name, iso3, iso2, region
        'rows': country_count,
        'url': url_for('system_admin.view_system_list', system_list_type='country_map')
    })

    # Indicator Bank system list
    indicator_count = IndicatorBank.query.count()
    system_lists.append({
        'name': 'Indicator Bank',
        'description': 'List of all indicators in the indicator bank',
        'columns': 3,  # name, type, unit
        'rows': indicator_count,
        'url': url_for('system_admin.view_system_list', system_list_type='indicator_bank')
    })

    # National Society system list
    ns_count = NationalSociety.query.count()
    system_lists.append({
        'name': 'National Society',
        'description': 'List of all national societies in the system',
        'columns': 4,  # name, code, description, country_id
        'rows': ns_count,
        'url': url_for('system_admin.view_system_list', system_list_type='national_society')
    })

    return render_template("admin/lists/manage_lists.html",
                         lists=lists,
                         system_lists=system_lists,
                         title="Manage Lookup Lists")

@bp.route("/lists/create", methods=["GET", "POST"])
@permission_required('admin.templates.edit')
def create_lookup_list():
    """Create a new lookup list"""
    if request.method == 'POST':
        try:
            name = request.form.get('name', '').strip()
            description = request.form.get('description', '').strip()

            if not name:
                flash("Name is required.", "danger")
                # SECURITY: Use request.path instead of request.url to prevent query param injection
                return redirect(request.path)

            # Check if name already exists
            existing = LookupList.query.filter_by(name=name).first()
            if existing:
                flash(f"A lookup list with the name '{name}' already exists.", "danger")
                # SECURITY: Use request.path instead of request.url to prevent query param injection
                return redirect(request.path)

            # Process columns configuration
            columns_str = request.form.get('columns', '').strip()
            columns_config = []
            if columns_str:
                columns = [col.strip() for col in columns_str.split(',') if col.strip()]
                columns_config = [{"name": col, "type": "string"} for col in columns]

            new_list = LookupList(
                name=name,
                description=description,
                columns_config=columns_config
            )

            db.session.add(new_list)
            db.session.flush()

            flash(f"Lookup list '{name}' created successfully.", "success")
            return redirect(url_for("system_admin.manage_lists"))

        except Exception as e:
            request_transaction_rollback()
            flash("An error occurred. Please try again.", "danger")
            current_app.logger.error(f"Error creating lookup list: {e}", exc_info=True)

    return render_template("admin/lists/manage_lists.html",
                         title="Create Lookup List")

@bp.route("/lists/view/<int:list_id>", methods=["GET"])
@permission_required('admin.templates.edit')
def view_lookup_list(list_id):
    """View lookup list details"""
    lookup_list = LookupList.query.get_or_404(list_id)
    rows = lookup_list.rows.order_by(LookupListRow.order).all()
    return render_template("admin/lists/list_detail.html",
                         lookup_list=lookup_list,
                         rows=rows,
                         title=f"View Lookup List: {lookup_list.name}")

def _get_model_columns_config(model_class):
    """Get all column names from a SQLAlchemy model, excluding relationships and internal fields."""
    inspector = inspect(model_class)
    columns_config = []

    # Get all columns from the model
    for column in inspector.columns:
        # Skip primary key 'id' column as it's usually not needed for display
        if column.name == 'id':
            continue

        # Skip name_translations field - we use 'name' with multilingual support instead
        if column.name == 'name_translations':
            continue

        # Determine column type
        col_type = "string"  # default
        if hasattr(column.type, 'python_type'):
            py_type = column.type.python_type
            if py_type == int or py_type == float:
                col_type = "number"
            elif py_type == bool:
                col_type = "boolean"
            elif py_type == datetime:
                col_type = "date"

        columns_config.append({
            "name": column.name,
            "type": col_type
        })

    return columns_config

def _model_to_dict(obj, columns_config):
    """Convert a model instance to a dictionary using the columns_config."""
    data = {}
    for col in columns_config:
        col_name = col['name']
        if hasattr(obj, col_name):
            value = getattr(obj, col_name)
            # Handle None values
            if value is None:
                data[col_name] = ''
            # Handle JSONB fields (like name_translations)
            elif isinstance(value, dict):
                # For JSONB fields, we might want to serialize or extract a value
                # For now, convert to string representation
                data[col_name] = str(value) if value else ''
            else:
                data[col_name] = value
        else:
            data[col_name] = ''
    return data

@bp.route("/lists/system/<system_list_type>", methods=["GET"])
@permission_required('admin.templates.edit')
def view_system_list(system_list_type):
    """View system list details (Country Map or Indicator Bank)"""
    if system_list_type == 'country_map':
        # Read from Country table
        countries = Country.query.order_by(Country.name).all()
        columns_config = _get_model_columns_config(Country)

        # Convert countries to row-like format
        rows_data = []
        for idx, country in enumerate(countries):
            country_data = _model_to_dict(country, columns_config)
            rows_data.append({
                'id': country.id,
                'order': idx,
                'data': country_data
            })

        # Create a mock lookup_list object for the template
        class MockLookupList:
            def __init__(self):
                self.id = 'country_map'
                self.name = 'Country Map'
                self.description = 'List of all countries in the system'
                self.columns_config = columns_config
                self.is_system_list = True

        lookup_list = MockLookupList()
        return render_template("admin/lists/list_detail.html",
                             lookup_list=lookup_list,
                             rows=rows_data,
                             title="View System List: Country Map")

    elif system_list_type == 'indicator_bank':
        # Read from IndicatorBank table
        indicators = IndicatorBank.query.order_by(IndicatorBank.name).all()
        columns_config = _get_model_columns_config(IndicatorBank)

        # Convert indicators to row-like format
        rows_data = []
        for idx, indicator in enumerate(indicators):
            indicator_data = _model_to_dict(indicator, columns_config)
            rows_data.append({
                'id': indicator.id,
                'order': idx,
                'data': indicator_data
            })

        # Create a mock lookup_list object for the template
        class MockLookupList:
            def __init__(self):
                self.id = 'indicator_bank'
                self.name = 'Indicator Bank'
                self.description = 'List of all indicators in the indicator bank'
                self.columns_config = columns_config
                self.is_system_list = True

        lookup_list = MockLookupList()
        return render_template("admin/lists/list_detail.html",
                             lookup_list=lookup_list,
                             rows=rows_data,
                             title="View System List: Indicator Bank")

    elif system_list_type == 'national_society':
        # Read from NationalSociety table
        national_societies = NationalSociety.query.order_by(NationalSociety.name).all()
        columns_config = _get_model_columns_config(NationalSociety)

        # Convert national societies to row-like format
        rows_data = []
        for idx, ns in enumerate(national_societies):
            ns_data = _model_to_dict(ns, columns_config)
            rows_data.append({
                'id': ns.id,
                'order': idx,
                'data': ns_data
            })

        # Create a mock lookup_list object for the template
        class MockLookupList:
            def __init__(self):
                self.id = 'national_society'
                self.name = 'National Society'
                self.description = 'List of all national societies in the system'
                self.columns_config = columns_config
                self.is_system_list = True

        lookup_list = MockLookupList()
        return render_template("admin/lists/list_detail.html",
                             lookup_list=lookup_list,
                             rows=rows_data,
                             title="View System List: National Society")

    else:
        flash("Invalid system list type.", "danger")
        return redirect(url_for("system_admin.manage_lists"))

@bp.route("/lists/edit/<int:list_id>", methods=["GET", "POST"])
@permission_required('admin.templates.edit')
def edit_lookup_list(list_id):
    """Edit basic lookup list properties"""
    lookup_list = LookupList.query.get_or_404(list_id)

    if request.method == 'POST':
        try:
            # Update basic info
            lookup_list.name = request.form.get('name', '').strip()
            lookup_list.description = request.form.get('description', '').strip()

            # Update columns config if provided
            columns_str = request.form.get('columns', '').strip()
            if columns_str:
                columns = [col.strip() for col in columns_str.split(',') if col.strip()]
                lookup_list.columns_config = [{"name": col, "type": "string"} for col in columns]

            db.session.flush()
            flash(f"Lookup list '{lookup_list.name}' updated successfully.", "success")
            return redirect(url_for("system_admin.manage_lists"))

        except Exception as e:
            request_transaction_rollback()
            flash("An error occurred. Please try again.", "danger")
            current_app.logger.error(f"Error updating lookup list {list_id}: {e}", exc_info=True)

    return redirect(url_for("system_admin.manage_lists"))

@bp.route("/lists/delete/<int:list_id>", methods=["POST"])
@permission_required('admin.templates.edit')
def delete_lookup_list(list_id):
    """Delete lookup list"""
    lookup_list = LookupList.query.get_or_404(list_id)

    try:
        # Delete all rows first
        LookupListRow.query.filter_by(lookup_list_id=list_id).delete()

        # Delete the list
        db.session.delete(lookup_list)
        db.session.flush()

        flash(f"Lookup list '{lookup_list.name}' deleted successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("An error occurred. Please try again.", "danger")
        current_app.logger.error(f"Error deleting lookup list {list_id}: {e}", exc_info=True)

    return redirect(url_for("system_admin.manage_lists"))

# === Lookup List Import/Export Routes ===
@bp.route("/lists/import", methods=["POST"])
@permission_required('admin.templates.edit')
def import_lookup_list():
    """Create a new lookup list from an uploaded CSV/XLS/XLSX file."""
    try:
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        file = request.files.get('file')
        if not name or not file or file.filename == '':
            flash("Name and file are required.", "danger")
            return redirect(url_for("system_admin.manage_lists"))

        # Validate file type and MIME - security check
        filename = secure_filename(file.filename)
        valid, error_msg, ext = validate_upload_extension_and_mime(file, CSV_EXCEL_EXTENSIONS)
        if not valid:
            flash(error_msg or "Unsupported file type. Please upload CSV or XLSX.", "danger")
            if ext:
                current_app.logger.warning(
                    f"Rejected lookup list import - {error_msg} (ext: {ext})"
                )
            return redirect(url_for("system_admin.manage_lists"))

        # Create the list first
        new_list = LookupList(name=name, description=description)
        db.session.add(new_list)
        db.session.flush()

        columns, rows_data = parse_csv_or_excel_to_rows(file, filename)

        new_list.columns_config = [{"name": c, "type": "string"} for c in columns]
        order_counter = 1
        for row in rows_data:
            db.session.add(LookupListRow(lookup_list_id=new_list.id, data=row, order=order_counter))
            order_counter += 1

        db.session.flush()
        flash(f"Lookup list '{new_list.name}' imported successfully.", "success")
    except Exception as e:
        request_transaction_rollback()
        current_app.logger.error(f"Error importing lookup list: {e}", exc_info=True)
        flash("Error importing lookup list.", "danger")
    return redirect(url_for("system_admin.manage_lists"))

@bp.route("/lists/<int:list_id>/import", methods=["POST"])
@permission_required('admin.templates.edit')
def import_into_lookup_list(list_id):
    """Import data into an existing list (append or replace)."""
    lookup_list = LookupList.query.get_or_404(list_id)
    mode = request.form.get('mode', 'append')
    file = request.files.get('file')
    if not file or file.filename == '':
        flash("No file selected.", "danger")
        return redirect(url_for("system_admin.view_lookup_list", list_id=list_id))
    try:
        filename = secure_filename(file.filename)

        # SECURITY: Validate file type and MIME
        valid, error_msg, ext = validate_upload_extension_and_mime(file, CSV_EXCEL_EXTENSIONS)
        if not valid:
            flash(error_msg or "Unsupported file type. Please upload CSV or XLSX.", "danger")
            if ext:
                current_app.logger.warning(
                    f"Rejected lookup list import - {error_msg} (ext: {ext})"
                )
            return redirect(url_for("system_admin.view_lookup_list", list_id=list_id))

        columns = [c.get('name') for c in (lookup_list.columns_config or [])]
        parsed_columns, parsed_rows = parse_csv_or_excel_to_rows(file, filename)
        if not columns:
            columns = parsed_columns
            lookup_list.columns_config = [{"name": c, "type": "string"} for c in columns]
        new_rows = [{k: row.get(k) for k in columns} for row in parsed_rows]

        if mode == 'replace':
            LookupListRow.query.filter_by(lookup_list_id=list_id).delete()
            db.session.flush()

        current_highest = db.session.query(func.coalesce(func.max(LookupListRow.order), 0)).filter_by(lookup_list_id=list_id).scalar() or 0
        order_counter = current_highest + 1
        for row in new_rows:
            db.session.add(LookupListRow(lookup_list_id=list_id, data=row, order=order_counter))
            order_counter += 1

        db.session.flush()
        flash("Data imported successfully.", "success")
    except Exception as e:
        request_transaction_rollback()
        current_app.logger.error(f"Error importing data into lookup list {list_id}: {e}", exc_info=True)
        flash("Error importing data.", "danger")
    return redirect(url_for("system_admin.view_lookup_list", list_id=list_id))

@bp.route("/lists/<int:list_id>/export", methods=["GET"])
@permission_required('admin.templates.edit')
def export_lookup_list(list_id):
    """Export a lookup list to CSV for download with proper Arabic encoding."""
    lookup_list = LookupList.query.get_or_404(list_id)
    import csv
    from io import StringIO, BytesIO
    import codecs

    # Create a StringIO buffer for CSV writing
    output = StringIO()
    columns = [c.get('name') for c in (lookup_list.columns_config or [])]
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()

    for row in lookup_list.rows.order_by(LookupListRow.order).all():
        writer.writerow({k: (row.data.get(k, '') if isinstance(row.data, dict) else '') for k in columns})

    # Get the CSV content and encode it properly for Arabic
    csv_content = output.getvalue()

    # Create BytesIO with UTF-8 BOM for Excel compatibility
    bytes_output = BytesIO()
    bytes_output.write(codecs.BOM_UTF8)  # Add BOM for Excel
    bytes_output.write(csv_content.encode('utf-8'))
    bytes_output.seek(0)

    response = make_response(bytes_output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f"attachment; filename={secure_filename((lookup_list.name or 'lookup_list'))}.csv"
    return response

# === Lookup List Row API Routes ===
@bp.route("/templates/lists/<int:list_id>/rows/<int:row_id>", methods=["PATCH"])
@permission_required('admin.templates.edit')
def update_lookup_list_row(list_id, row_id):
    """Update a specific row in a lookup list"""
    try:
        current_app.logger.info(f"Updating lookup list row: list_id={list_id}, row_id={row_id}")

        lookup_list = LookupList.query.get_or_404(list_id)
        row = LookupListRow.query.filter_by(id=row_id, lookup_list_id=list_id).first_or_404()

        data = get_json_safe()
        current_app.logger.info(f"Received data: {data}")
        err = require_json_data(data)
        if err:
            return err

        # Update the row data
        if isinstance(row.data, dict):
            current_app.logger.info(f"Updating existing data: {row.data} with {data}")
            # Create a new dict to ensure proper JSON serialization
            updated_data = dict(row.data)
            updated_data.update(data)
            row.data = updated_data
            current_app.logger.info(f"Updated data before commit: {row.data}")
        else:
            current_app.logger.info(f"Setting new data: {data}")
            row.data = data

        # Force the change to be detected by SQLAlchemy
        db.session.add(row)
        current_app.logger.info(f"Row data before commit: {row.data}")

        try:
            db.session.flush()
            current_app.logger.info("Commit successful")
        except Exception as commit_error:
            current_app.logger.error(f"Commit failed: {commit_error}")
            request_transaction_rollback()
            raise commit_error

        # Refresh the row to ensure we have the latest data
        db.session.refresh(row)
        current_app.logger.info(f"Row updated successfully. New data: {row.data}")

        # Double-check by querying the database directly
        verification_row = LookupListRow.query.get(row_id)
        current_app.logger.info(f"Verification - row data from DB: {verification_row.data}")

        return json_ok(message='Row updated successfully')

    except Exception as e:
        request_transaction_rollback()
        current_app.logger.error(f"Error updating lookup list row {row_id}: {e}", exc_info=True)
        return json_server_error(GENERIC_ERROR_MESSAGE, success=False)

@bp.route("/templates/lists/<int:list_id>/rows/<int:row_id>/move", methods=["POST"])
@permission_required('admin.templates.edit')
def move_lookup_list_row(list_id, row_id):
    """Move a row to a new position in the list"""
    try:
        current_app.logger.info(f"Moving lookup list row: list_id={list_id}, row_id={row_id}")

        lookup_list = LookupList.query.get_or_404(list_id)
        row = LookupListRow.query.filter_by(id=row_id, lookup_list_id=list_id).first_or_404()

        data = get_json_safe()
        current_app.logger.info(f"Move data: {data}")
        err = require_json_keys(data, ['target_row_id'])
        if err:
            return err

        target_row_id = data.get('target_row_id')
        position = data.get('position', 'after')  # 'before' or 'after'

        if not target_row_id:
            return json_bad_request('Target row ID required')

        target_row = LookupListRow.query.filter_by(id=target_row_id, lookup_list_id=list_id).first_or_404()

        # Get all rows in current order
        all_rows = lookup_list.rows.order_by(LookupListRow.order).all()
        current_app.logger.info(f"All rows before move: {[(r.id, r.order) for r in all_rows]}")

        # Calculate new order for the moving row
        target_order = target_row.order
        if position == 'before':
            new_order = target_order
        else:  # after
            new_order = target_order + 1

        current_order = row.order

        # If moving to the same position, do nothing
        if current_order == new_order:
            return json_ok(message='Row already in position')

        # Step 1: Set the moving row to a temporary high order to avoid conflicts
        row.order = LOOKUP_ROW_TEMP_ORDER
        db.session.flush()

        # Step 2: Shift other rows as needed
        if current_order < new_order:
            # Moving down - shift rows up
            lookup_list.rows.filter(
                LookupListRow.order > current_order,
                LookupListRow.order <= new_order
            ).update({
                LookupListRow.order: LookupListRow.order - 1
            })
        else:
            # Moving up - shift rows down
            lookup_list.rows.filter(
                LookupListRow.order >= new_order,
                LookupListRow.order < current_order
            ).update({
                LookupListRow.order: LookupListRow.order + 1
            })

        # Step 3: Set the final order for the moving row
        row.order = new_order

        db.session.flush()

        # Log the final result
        final_rows = lookup_list.rows.order_by(LookupListRow.order).all()
        current_app.logger.info(f"All rows after move: {[(r.id, r.order) for r in final_rows]}")
        return json_ok(message='Row moved successfully')

    except Exception as e:
        request_transaction_rollback()
        current_app.logger.error(f"Error moving lookup list row {row_id}: {e}", exc_info=True)
        return json_server_error(GENERIC_ERROR_MESSAGE, success=False)

@bp.route("/templates/lists/<int:list_id>/rows/<int:row_id>", methods=["DELETE"])
@permission_required('admin.templates.edit')
def delete_lookup_list_row(list_id, row_id):
    """Delete a specific row from a lookup list"""
    try:
        lookup_list = LookupList.query.get_or_404(list_id)
        row = LookupListRow.query.filter_by(id=row_id, lookup_list_id=list_id).first_or_404()

        # Reorder remaining rows
        remaining_rows = lookup_list.rows.filter(LookupListRow.id != row_id).order_by(LookupListRow.order).all()
        for i, remaining_row in enumerate(remaining_rows, 1):
            remaining_row.order = i

        db.session.delete(row)
        db.session.flush()

        return json_ok(message='Row deleted successfully')

    except Exception as e:
        request_transaction_rollback()
        current_app.logger.error(f"Error deleting lookup list row {row_id}: {e}", exc_info=True)
        return json_server_error(GENERIC_ERROR_MESSAGE, success=False)

@bp.route("/api/templates/lists/<int:list_id>/rows", methods=["POST"])
@permission_required('admin.templates.edit')
def add_lookup_list_row(list_id):
    """Add a new row to a lookup list"""
    try:
        lookup_list = LookupList.query.get_or_404(list_id)

        data = get_json_safe()
        order = data.get('order')
        insert_after_order = data.get('insert_after_order')

        # Calculate the new order
        if insert_after_order is not None:
            # Insert after a specific order
            new_order = insert_after_order + 1
            # Shift all rows with higher order
            lookup_list.rows.filter(LookupListRow.order > insert_after_order).update({
                LookupListRow.order: LookupListRow.order + 1
            })
        elif order is not None:
            # Use provided order
            new_order = order
            # Shift all rows with higher or equal order
            lookup_list.rows.filter(LookupListRow.order >= order).update({
                LookupListRow.order: LookupListRow.order + 1
            })
        else:
            # Add to the end
            max_order = db.session.query(func.max(LookupListRow.order)).filter_by(lookup_list_id=list_id).scalar() or 0
            new_order = max_order + 1

        # Create new row with empty data
        new_row = LookupListRow(
            lookup_list_id=list_id,
            data={},  # Initialize with empty dict
            order=new_order
        )

        # Ensure the data field is properly initialized
        if new_row.data is None:
            new_row.data = {}

        db.session.add(new_row)
        db.session.flush()

        return json_ok(message='Row added successfully', row_id=new_row.id)

    except Exception as e:
        request_transaction_rollback()
        current_app.logger.error(f"Error adding row to lookup list {list_id}: {e}", exc_info=True)
        return json_server_error(GENERIC_ERROR_MESSAGE, success=False)

@bp.route("/ns_hierarchy/branch/new", methods=["GET", "POST"])
@permission_required('admin.organization.manage')
def new_ns_branch():
    """Create a new NS branch using unified edit template"""
    from app.models import NSBranch, Country

    form = NSBranchForm()
    # Filter countries based on user role
    from app.services.authorization_service import AuthorizationService
    is_focal_point_only = AuthorizationService.has_role(current_user, "assignment_editor_submitter") and not AuthorizationService.is_admin(current_user)
    if is_focal_point_only:
        countries = list(current_user.countries)
    else:
        countries = Country.query.order_by(Country.name).all()
    form.country_id.choices = [(c.id, c.name) for c in countries]

    if form.validate_on_submit():
        try:
            new_branch = NSBranch(
                name=form.name.data,
                code=form.code.data or None,
                description=form.description.data or None,
                country_id=form.country_id.data,
                address=form.address.data or None,
                city=form.city.data or None,
                postal_code=form.postal_code.data or None,
                coordinates=form.coordinates.data or None,
                phone=form.phone.data or None,
                email=form.email.data or None,
                website=form.website.data or None,
                is_active=form.is_active.data,
                established_date=form.established_date.data,
                display_order=form.display_order.data or 0
            )
            db.session.add(new_branch)
            db.session.flush()
            flash(f"Branch '{new_branch.name}' created successfully.", "success")
            return redirect(url_for("main.manage_ns_hierarchy"))
        except Exception as e:
            request_transaction_rollback()
            flash("Error creating branch.", "danger")
            current_app.logger.error(f"Error creating NS branch: {e}", exc_info=True)

    return render_template(
        "admin/organization/edit_entity.html",
        form=form,
        is_edit=False,
        entity=None,
        entity_label='NS Branch',
        icon='fas fa-code-branch',
        cancel_url=url_for('main.manage_ns_hierarchy')
    )

@bp.route("/ns_hierarchy/branch/edit/<int:branch_id>", methods=["GET", "POST"])
@permission_required('admin.organization.manage')
def edit_ns_branch(branch_id):
    """Edit an existing NS branch using unified edit template"""
    from app.models import NSBranch, Country

    branch = NSBranch.query.get_or_404(branch_id)

    # Check if focal point has access to this branch's country
    from app.services.authorization_service import AuthorizationService
    is_focal_point_only = AuthorizationService.has_role(current_user, "assignment_editor_submitter") and not AuthorizationService.is_admin(current_user)
    if is_focal_point_only:
        user_country_ids = [country.id for country in current_user.countries]
        if branch.country_id not in user_country_ids:
            flash("Access denied. You can only manage branches in your assigned countries.", "danger")
            return redirect(url_for("main.manage_ns_hierarchy"))

    form = NSBranchForm(obj=branch)
    if is_focal_point_only:
        countries = list(current_user.countries)
    else:
        countries = Country.query.order_by(Country.name).all()
    form.country_id.choices = [(c.id, c.name) for c in countries]

    if form.validate_on_submit():
        try:
            branch.name = form.name.data
            branch.code = form.code.data or None
            branch.description = form.description.data or None
            branch.country_id = form.country_id.data
            branch.address = form.address.data or None
            branch.city = form.city.data or None
            branch.postal_code = form.postal_code.data or None
            branch.coordinates = form.coordinates.data or None
            branch.phone = form.phone.data or None
            branch.email = form.email.data or None
            branch.website = form.website.data or None
            branch.is_active = form.is_active.data
            branch.established_date = form.established_date.data
            branch.display_order = form.display_order.data or 0
            db.session.flush()
            flash(f"Branch '{branch.name}' updated successfully.", "success")
            return redirect(url_for("main.manage_ns_hierarchy"))
        except Exception as e:
            request_transaction_rollback()
            flash("An error occurred. Please try again.", "danger")
            current_app.logger.error(f"Error updating NS branch {branch_id}: {e}", exc_info=True)

    return render_template(
        "admin/organization/edit_entity.html",
        form=form,
        is_edit=True,
        entity=branch,
        entity_label='NS Branch',
        icon='fas fa-code-branch',
        cancel_url=url_for('main.manage_ns_hierarchy')
    )

@bp.route("/ns_hierarchy/branch/delete/<int:branch_id>", methods=["POST"])
@permission_required('admin.organization.manage')
def delete_ns_branch(branch_id):
    """Delete an NS branch"""
    from app.models import NSBranch

    branch = NSBranch.query.get_or_404(branch_id)

    # Check if focal point has access to this branch's country
    from app.services.authorization_service import AuthorizationService
    is_focal_point_only = AuthorizationService.has_role(current_user, "assignment_editor_submitter") and not AuthorizationService.is_admin(current_user)
    if is_focal_point_only:
        user_country_ids = [country.id for country in current_user.countries]
        if branch.country_id not in user_country_ids:
            flash("Access denied. You can only manage branches in your assigned countries.", "danger")
            return redirect(url_for("main.manage_ns_hierarchy"))

    try:
        # Check if branch has sub-branches or local units
        if branch.subbranches.first() or branch.local_units.first():
            flash(f"Cannot delete branch '{branch.name}' as it has associated sub-branches or local units.", "danger")
            return redirect(url_for("main.manage_ns_hierarchy"))

        db.session.delete(branch)
        db.session.flush()

        flash(f"Branch '{branch.name}' deleted successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("Error deleting branch.", "danger")
        current_app.logger.error(f"Error deleting NS branch {branch_id}: {e}", exc_info=True)

    return redirect(url_for("main.manage_ns_hierarchy"))

@bp.route("/ns_hierarchy/subbranch/new", methods=["GET", "POST"])
@permission_required('admin.organization.manage')
def new_ns_subbranch():
    """Create a new NS sub-branch using unified edit template"""
    from app.models import NSSubBranch, NSBranch

    form = NSSubBranchForm()
    branches = NSBranch.query.filter_by(is_active=True).order_by(NSBranch.name).all()
    form.branch_id.choices = [(b.id, b.name) for b in branches]

    if form.validate_on_submit():
        try:
            new_subbranch = NSSubBranch(
                name=form.name.data,
                code=form.code.data or None,
                description=form.description.data or None,
                branch_id=form.branch_id.data,
                address=form.address.data or None,
                city=form.city.data or None,
                postal_code=form.postal_code.data or None,
                coordinates=form.coordinates.data or None,
                phone=form.phone.data or None,
                email=form.email.data or None,
                is_active=form.is_active.data,
                established_date=form.established_date.data,
                display_order=form.display_order.data or 0
            )
            db.session.add(new_subbranch)
            db.session.flush()
            flash(f"Sub-branch '{new_subbranch.name}' created successfully.", "success")
            return redirect(url_for("main.manage_ns_hierarchy"))
        except Exception as e:
            request_transaction_rollback()
            flash("An error occurred. Please try again.", "danger")
            current_app.logger.error(f"Error creating NS sub-branch: {e}", exc_info=True)

    return render_template(
        "admin/organization/edit_entity.html",
        form=form,
        is_edit=False,
        entity=None,
        entity_label='NS Sub-branch',
        icon='fas fa-network-wired',
        cancel_url=url_for('main.manage_ns_hierarchy')
    )

@bp.route("/ns_hierarchy/subbranch/edit/<int:subbranch_id>", methods=["GET", "POST"])
@permission_required('admin.organization.manage')
def edit_ns_subbranch(subbranch_id):
    """Edit an existing NS sub-branch using unified edit template"""
    from app.models import NSSubBranch, NSBranch

    subbranch = NSSubBranch.query.get_or_404(subbranch_id)
    form = NSSubBranchForm(obj=subbranch)
    branches = NSBranch.query.filter_by(is_active=True).order_by(NSBranch.name).all()
    form.branch_id.choices = [(b.id, b.name) for b in branches]

    if form.validate_on_submit():
        try:
            subbranch.name = form.name.data
            subbranch.code = form.code.data or None
            subbranch.description = form.description.data or None
            subbranch.branch_id = form.branch_id.data
            subbranch.address = form.address.data or None
            subbranch.city = form.city.data or None
            subbranch.postal_code = form.postal_code.data or None
            subbranch.coordinates = form.coordinates.data or None
            subbranch.phone = form.phone.data or None
            subbranch.email = form.email.data or None
            subbranch.is_active = form.is_active.data
            subbranch.established_date = form.established_date.data
            subbranch.display_order = form.display_order.data or 0
            db.session.flush()
            flash(f"Sub-branch '{subbranch.name}' updated successfully.", "success")
            return redirect(url_for("main.manage_ns_hierarchy"))
        except Exception as e:
            request_transaction_rollback()
            flash("Error updating sub-branch.", "danger")
            current_app.logger.error(f"Error updating NS sub-branch {subbranch_id}: {e}", exc_info=True)

    return render_template(
        "admin/organization/edit_entity.html",
        form=form,
        is_edit=True,
        entity=subbranch,
        entity_label='NS Sub-branch',
        icon='fas fa-network-wired',
        cancel_url=url_for('main.manage_ns_hierarchy')
    )

@bp.route("/ns_hierarchy/subbranch/delete/<int:subbranch_id>", methods=["POST"])
@permission_required('admin.organization.manage')
def delete_ns_subbranch(subbranch_id):
    """Delete an NS sub-branch"""
    from app.models import NSSubBranch

    subbranch = NSSubBranch.query.get_or_404(subbranch_id)

    try:
        # Check if sub-branch has local units
        if subbranch.local_units.first():
            flash(f"Cannot delete sub-branch '{subbranch.name}' as it has associated local units.", "danger")
            return redirect(url_for("main.manage_ns_hierarchy"))

        db.session.delete(subbranch)
        db.session.flush()

        flash(f"Sub-branch '{subbranch.name}' deleted successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("An error occurred. Please try again.", "danger")
        current_app.logger.error(f"Error deleting NS sub-branch {subbranch_id}: {e}", exc_info=True)

    return redirect(url_for("main.manage_ns_hierarchy"))

@bp.route("/ns_hierarchy/localunit/new", methods=["GET", "POST"])
@permission_required('admin.organization.manage')
def new_ns_localunit():
    """Create a new NS local unit using unified edit template"""
    from app.models import NSLocalUnit, NSBranch, NSSubBranch

    form = NSLocalUnitForm()
    branches = NSBranch.query.filter_by(is_active=True).order_by(NSBranch.name).all()
    form.branch_id.choices = [(b.id, b.name) for b in branches]
    subbranches = NSSubBranch.query.filter_by(is_active=True).order_by(NSSubBranch.name).all()
    # Include an explicit none option similar to organization routes
    form.subbranch_id.choices = [('', 'None (Direct to Branch)')] + [(sb.id, sb.name) for sb in subbranches]

    if form.validate_on_submit():
        try:
            new_localunit = NSLocalUnit(
                name=form.name.data,
                code=form.code.data or None,
                description=form.description.data or None,
                branch_id=form.branch_id.data,
                subbranch_id=form.subbranch_id.data or None,
                address=form.address.data or None,
                city=form.city.data or None,
                postal_code=form.postal_code.data or None,
                coordinates=form.coordinates.data or None,
                phone=form.phone.data or None,
                email=form.email.data or None,
                is_active=form.is_active.data,
                established_date=form.established_date.data,
                display_order=form.display_order.data or 0
            )
            db.session.add(new_localunit)
            db.session.flush()
            flash(f"Local unit '{new_localunit.name}' created successfully.", "success")
            return redirect(url_for("main.manage_ns_hierarchy"))
        except Exception as e:
            request_transaction_rollback()
            flash("Error creating local unit.", "danger")
            current_app.logger.error(f"Error creating NS local unit: {e}", exc_info=True)

    return render_template(
        "admin/organization/edit_entity.html",
        form=form,
        is_edit=False,
        entity=None,
        entity_label='NS Local Unit',
        icon='fas fa-map-marker-alt',
        cancel_url=url_for('main.manage_ns_hierarchy')
    )

@bp.route("/ns_hierarchy/localunit/edit/<int:localunit_id>", methods=["GET", "POST"])
@permission_required('admin.organization.manage')
def edit_ns_localunit(localunit_id):
    """Edit an existing NS local unit using unified edit template"""
    from app.models import NSLocalUnit, NSBranch, NSSubBranch

    localunit = NSLocalUnit.query.get_or_404(localunit_id)
    form = NSLocalUnitForm(obj=localunit)
    branches = NSBranch.query.filter_by(is_active=True).order_by(NSBranch.name).all()
    form.branch_id.choices = [(b.id, b.name) for b in branches]
    subbranches = NSSubBranch.query.filter_by(is_active=True).order_by(NSSubBranch.name).all()
    form.subbranch_id.choices = [('', 'None (Direct to Branch)')] + [(sb.id, sb.name) for sb in subbranches]

    if form.validate_on_submit():
        try:
            localunit.name = form.name.data
            localunit.code = form.code.data or None
            localunit.description = form.description.data or None
            localunit.branch_id = form.branch_id.data
            localunit.subbranch_id = form.subbranch_id.data or None
            localunit.address = form.address.data or None
            localunit.city = form.city.data or None
            localunit.postal_code = form.postal_code.data or None
            localunit.coordinates = form.coordinates.data or None
            localunit.phone = form.phone.data or None
            localunit.email = form.email.data or None
            localunit.is_active = form.is_active.data
            localunit.established_date = form.established_date.data
            localunit.display_order = form.display_order.data or 0
            db.session.flush()
            flash(f"Local unit '{localunit.name}' updated successfully.", "success")
            return redirect(url_for("main.manage_ns_hierarchy"))
        except Exception as e:
            request_transaction_rollback()
            flash("An error occurred. Please try again.", "danger")
            current_app.logger.error(f"Error updating NS local unit {localunit_id}: {e}", exc_info=True)

    return render_template(
        "admin/organization/edit_entity.html",
        form=form,
        is_edit=True,
        entity=localunit,
        entity_label='NS Local Unit',
        icon='fas fa-map-marker-alt',
        cancel_url=url_for('main.manage_ns_hierarchy')
    )

@bp.route("/ns_hierarchy/localunit/delete/<int:localunit_id>", methods=["POST"])
@permission_required('admin.organization.manage')
def delete_ns_localunit(localunit_id):
    """Delete an NS local unit"""
    from app.models import NSLocalUnit

    localunit = NSLocalUnit.query.get_or_404(localunit_id)

    try:
        db.session.delete(localunit)
        db.session.flush()

        flash(f"Local unit '{localunit.name}' deleted successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("Error deleting local unit.", "danger")
        current_app.logger.error(f"Error deleting NS local unit {localunit_id}: {e}", exc_info=True)

    return redirect(url_for("main.manage_ns_hierarchy"))

# === Static file serving for sector/subsector logos ===
@bp.route("/sectors/<int:sector_id>/logo", methods=["GET"])
@rbac_guard_audit_exempt("Intentionally public to allow logo rendering without admin session.")
def sector_logo(sector_id):
    sector = Sector.query.get_or_404(sector_id)
    if not sector.logo_filename:
        return ("", 404)
    return send_from_directory(get_sector_logo_path(), sector.logo_filename)

@bp.route("/subsectors/<int:subsector_id>/logo", methods=["GET"])
@rbac_guard_audit_exempt("Intentionally public to allow logo rendering without admin session.")
def subsector_logo(subsector_id):
    subsector = SubSector.query.get_or_404(subsector_id)
    if not subsector.logo_filename:
        return ("", 404)
    return send_from_directory(get_subsector_logo_path(), subsector.logo_filename)

# === Session Management Routes ===
@bp.route("/sessions/cleanup", methods=["POST"])
@permission_required('admin.analytics.view')
def cleanup_sessions():
    """Cleanup inactive sessions"""
    try:
        from app.utils.user_analytics import cleanup_inactive_sessions
        count = cleanup_inactive_sessions()
        flash(f"Successfully cleaned up {count} inactive sessions.", "success")
    except Exception as e:
        request_transaction_rollback()
        flash("An error occurred during session cleanup.", "danger")
        current_app.logger.error(f"Error during session cleanup: {e}", exc_info=True)

    # Canonical UI lives in Analytics → Session Logs
    return redirect(url_for("analytics.session_logs"))

@bp.route("/sessions/status", methods=["GET"])
@permission_required('admin.analytics.view')
def session_status():
    """Legacy endpoint. Redirect to analytics session logs."""
    return redirect(url_for("analytics.session_logs"))

# === Export/Import Routes ===
@bp.route("/indicator_bank/export", methods=["GET", "POST"])
@permission_required('admin.indicator_bank.view')
def export_indicators():
    """Export indicators to Excel.

    Supports:
    - GET: export all indicators
    - POST: export only selected IDs passed as comma-separated 'selected_ids'
    """
    try:
        # Optional selected IDs from POST
        selected_ids = None
        if request.method == "POST":
            ids_str = request.form.get('selected_ids')
            if ids_str:
                try:
                    selected_ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
                except Exception as e:
                    current_app.logger.debug("selected_ids parse failed: %s", e)
                    selected_ids = None

        # Build query
        query = IndicatorBank.query
        if selected_ids:
            query = query.filter(IndicatorBank.id.in_(selected_ids))
        indicators = query.order_by(IndicatorBank.name).all()

        # Create workbook
        wb = Workbook()

        # --- Main user-friendly sheet (visible) ---
        ws = wb.active
        ws.title = "Indicators"

        languages = current_app.config.get("SUPPORTED_LANGUAGES", getattr(Config, "LANGUAGES", ["en"])) or ["en"]
        name_lang_headers = [f"Name ({code})" for code in languages]
        def_lang_headers = [f"Definition ({code})" for code in languages]
        sector_subsector_headers = [
            "Sector Primary", "Sector Secondary", "Sector Tertiary",
            "SubSector Primary", "SubSector Secondary", "SubSector Tertiary",
        ]
        main_headers = (
            ['ID', 'Name', 'Definition', 'Type', 'Unit', 'FDRS KPI Code', 'Emergency', 'Related Programs', 'Archived', 'Created At']
            + name_lang_headers
            + def_lang_headers
            + sector_subsector_headers
        )
        for col, header in enumerate(main_headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Prefetch sector/subsector names for all indicators
        sector_ids = set()
        subsector_ids = set()
        for ind in indicators:
            for level in ("primary", "secondary", "tertiary"):
                if ind.sector and ind.sector.get(level):
                    sector_ids.add(ind.sector.get(level))
                if ind.sub_sector and ind.sub_sector.get(level):
                    subsector_ids.add(ind.sub_sector.get(level))
        sector_names = {}
        if sector_ids:
            for s in Sector.query.filter(Sector.id.in_(sector_ids)).all():
                sector_names[s.id] = s.name
        subsector_names = {}
        if subsector_ids:
            for ss in SubSector.query.filter(SubSector.id.in_(subsector_ids)).all():
                subsector_names[ss.id] = ss.name

        for row, indicator in enumerate(indicators, 2):
            ws.cell(row=row, column=1, value=indicator.id)
            ws.cell(row=row, column=2, value=indicator.name)
            ws.cell(row=row, column=3, value=indicator.definition)
            ws.cell(row=row, column=4, value=indicator.type)
            ws.cell(row=row, column=5, value=indicator.unit)
            ws.cell(row=row, column=6, value=getattr(indicator, 'fdrs_kpi_code', None) or '')
            ws.cell(row=row, column=7, value=indicator.emergency)
            ws.cell(row=row, column=8, value=indicator.related_programs)
            ws.cell(row=row, column=9, value=indicator.archived)
            ws.cell(row=row, column=10, value=indicator.created_at.strftime('%Y-%m-%d') if indicator.created_at else '')

            nt = indicator.name_translations or {}
            dt = indicator.definition_translations or {}
            col = 11
            for code in languages:
                ws.cell(row=row, column=col, value=nt.get(code) if code != "en" else (nt.get("en") or indicator.name))
                col += 1
            for code in languages:
                ws.cell(row=row, column=col, value=dt.get(code) if code != "en" else (dt.get("en") or indicator.definition))
                col += 1
            sec = indicator.sector or {}
            subsec = indicator.sub_sector or {}
            for level in ("primary", "secondary", "tertiary"):
                sid = sec.get(level)
                ws.cell(row=row, column=col, value=sector_names.get(sid, "") if sid else "")
                col += 1
            for level in ("primary", "secondary", "tertiary"):
                ssid = subsec.get(level)
                ws.cell(row=row, column=col, value=subsector_names.get(ssid, "") if ssid else "")
                col += 1

        # Auto-adjust column widths (main sheet only)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                with suppress(Exception):
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # --- Hidden relational sheets (DB snapshot for full import) ---
        def _json_dump(val):
            try:
                if val is None:
                    return ""
                return json.dumps(val, ensure_ascii=False)
            except Exception as e:
                current_app.logger.debug("indicator export _json_dump failed: %s", e)
                return ""

        # 1) Indicators (DB relational fields + translations)
        ws_db_ind = wb.create_sheet(title="DB_Indicators")
        ws_db_ind.sheet_state = "hidden"

        db_ind_headers = [
            "id",
            "name",
            "definition",
            "type",
            "unit",
            "fdrs_kpi_code",
            "emergency",
            "related_programs",
            "archived",
            "comments",
            "sector_primary_id",
            "sector_secondary_id",
            "sector_tertiary_id",
            "subsector_primary_id",
            "subsector_secondary_id",
            "subsector_tertiary_id",
            "name_translations_json",
            "definition_translations_json",
            "created_at",
            "updated_at",
        ]
        for col, header in enumerate(db_ind_headers, 1):
            cell = ws_db_ind.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        for row, indicator in enumerate(indicators, 2):
            sector = indicator.sector or {}
            subsector = indicator.sub_sector or {}
            ws_db_ind.cell(row=row, column=1, value=indicator.id)
            ws_db_ind.cell(row=row, column=2, value=indicator.name)
            ws_db_ind.cell(row=row, column=3, value=indicator.definition)
            ws_db_ind.cell(row=row, column=4, value=indicator.type)
            ws_db_ind.cell(row=row, column=5, value=indicator.unit)
            ws_db_ind.cell(row=row, column=6, value=getattr(indicator, 'fdrs_kpi_code', None) or '')
            ws_db_ind.cell(row=row, column=7, value=indicator.emergency)
            ws_db_ind.cell(row=row, column=8, value=indicator.related_programs)
            ws_db_ind.cell(row=row, column=9, value=indicator.archived)
            ws_db_ind.cell(row=row, column=10, value=indicator.comments)
            ws_db_ind.cell(row=row, column=11, value=sector.get("primary"))
            ws_db_ind.cell(row=row, column=12, value=sector.get("secondary"))
            ws_db_ind.cell(row=row, column=13, value=sector.get("tertiary"))
            ws_db_ind.cell(row=row, column=14, value=subsector.get("primary"))
            ws_db_ind.cell(row=row, column=15, value=subsector.get("secondary"))
            ws_db_ind.cell(row=row, column=16, value=subsector.get("tertiary"))
            ws_db_ind.cell(row=row, column=17, value=_json_dump(indicator.name_translations or {}))
            ws_db_ind.cell(row=row, column=18, value=_json_dump(indicator.definition_translations or {}))
            ws_db_ind.cell(row=row, column=19, value=indicator.created_at.isoformat() if indicator.created_at else "")
            ws_db_ind.cell(row=row, column=20, value=indicator.updated_at.isoformat() if getattr(indicator, "updated_at", None) else "")

        # 2) Sectors + SubSectors (single hidden sheet, record_type distinguishes rows)
        ws_db_ss = wb.create_sheet(title="DB_Sectors_SubSectors")
        ws_db_ss.sheet_state = "hidden"

        db_ss_headers = [
            "record_type",  # "sector" | "subsector"
            "id",
            "name",
            "description",
            "sector_id",  # for subsector only (parent sector)
            "display_order",
            "is_active",
            "icon_class",
            "logo_filename",
            "logo_path",
            "name_translations_json",
            "created_at",
            "updated_at",
        ]
        for col, header in enumerate(db_ss_headers, 1):
            cell = ws_db_ss.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        ss_row = 2
        sectors = Sector.query.order_by(Sector.display_order, Sector.name).all()
        for s in sectors:
            ws_db_ss.cell(row=ss_row, column=1, value="sector")
            ws_db_ss.cell(row=ss_row, column=2, value=s.id)
            ws_db_ss.cell(row=ss_row, column=3, value=s.name)
            ws_db_ss.cell(row=ss_row, column=4, value=s.description)
            ws_db_ss.cell(row=ss_row, column=5, value="")
            ws_db_ss.cell(row=ss_row, column=6, value=s.display_order)
            ws_db_ss.cell(row=ss_row, column=7, value=s.is_active)
            ws_db_ss.cell(row=ss_row, column=8, value=s.icon_class)
            ws_db_ss.cell(row=ss_row, column=9, value=s.logo_filename)
            ws_db_ss.cell(row=ss_row, column=10, value=s.logo_path)
            ws_db_ss.cell(row=ss_row, column=11, value=_json_dump(s.name_translations or {}))
            ws_db_ss.cell(row=ss_row, column=12, value=s.created_at.isoformat() if getattr(s, "created_at", None) else "")
            ws_db_ss.cell(row=ss_row, column=13, value=s.updated_at.isoformat() if getattr(s, "updated_at", None) else "")
            ss_row += 1

        subsectors = SubSector.query.order_by(SubSector.display_order, SubSector.name).all()
        for ss in subsectors:
            ws_db_ss.cell(row=ss_row, column=1, value="subsector")
            ws_db_ss.cell(row=ss_row, column=2, value=ss.id)
            ws_db_ss.cell(row=ss_row, column=3, value=ss.name)
            ws_db_ss.cell(row=ss_row, column=4, value=ss.description)
            ws_db_ss.cell(row=ss_row, column=5, value=ss.sector_id)
            ws_db_ss.cell(row=ss_row, column=6, value=ss.display_order)
            ws_db_ss.cell(row=ss_row, column=7, value=ss.is_active)
            ws_db_ss.cell(row=ss_row, column=8, value=ss.icon_class)
            ws_db_ss.cell(row=ss_row, column=9, value=ss.logo_filename)
            ws_db_ss.cell(row=ss_row, column=10, value=ss.logo_path)
            ws_db_ss.cell(row=ss_row, column=11, value=_json_dump(ss.name_translations or {}))
            ws_db_ss.cell(row=ss_row, column=12, value=ss.created_at.isoformat() if getattr(ss, "created_at", None) else "")
            ws_db_ss.cell(row=ss_row, column=13, value=ss.updated_at.isoformat() if getattr(ss, "updated_at", None) else "")
            ss_row += 1

        # 3) Common Words
        ws_db_cw = wb.create_sheet(title="DB_CommonWords")
        ws_db_cw.sheet_state = "hidden"

        db_cw_headers = [
            "id",
            "term",
            "meaning",
            "is_active",
            "meaning_translations_json",
            "created_at",
            "updated_at",
        ]
        for col, header in enumerate(db_cw_headers, 1):
            cell = ws_db_cw.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        common_words = CommonWord.query.order_by(CommonWord.term).all()
        for row, cw in enumerate(common_words, 2):
            ws_db_cw.cell(row=row, column=1, value=cw.id)
            ws_db_cw.cell(row=row, column=2, value=cw.term)
            ws_db_cw.cell(row=row, column=3, value=cw.meaning)
            ws_db_cw.cell(row=row, column=4, value=cw.is_active)
            ws_db_cw.cell(row=row, column=5, value=_json_dump(cw.meaning_translations or {}))
            ws_db_cw.cell(row=row, column=6, value=cw.created_at.isoformat() if getattr(cw, "created_at", None) else "")
            ws_db_cw.cell(row=row, column=7, value=cw.updated_at.isoformat() if getattr(cw, "updated_at", None) else "")

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=indicators_export.xlsx'

        return response

    except Exception as e:
        current_app.logger.error(f"Error exporting indicators: {e}", exc_info=True)
        flash("Error exporting indicators.", "danger")
        return redirect(url_for("system_admin.manage_indicator_bank"))

# === Helper Functions ===

def _save_logo_file(file_storage, base_path, item_name, item_type):
    """Save a logo file using standardized path functions.

    Note: This function is kept for backward compatibility but now uses
    standardized path functions internally.
    """
    try:
        if not file_storage or not file_storage.filename:
            return None

        # Determine if this is a sector or subsector based on base_path
        is_sector = 'sectors' in base_path or base_path == get_sector_logo_path()

        # Use standardized save function
        return save_system_logo(file_storage, item_name, item_type, is_sector=is_sector)

    except Exception as e:
        current_app.logger.exception("Error saving logo file: %s", e)
        return None

def _delete_logo_file(base_path, filename):
    """Delete a logo file"""
    try:
        # Determine if this is a sector or subsector
        is_sector = 'sectors' in base_path or base_path == get_sector_logo_path()

        if is_sector:
            file_path = resolve_sector_logo(filename)
        else:
            file_path = resolve_subsector_logo(filename)

        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception as e:
        current_app.logger.exception("Error deleting logo file: %s", e)

# ========================
# API Endpoints
# ========================

@bp.route("/api/indicator-count", methods=["POST"])
@permission_required('admin.indicator_bank.view')
def get_filtered_indicator_count():
    """API endpoint to get count of indicators matching filters"""
    try:
        data = request.json
        filters = data.get('filters', [])
        section_id = data.get('section_id')
        include_indicators = data.get('include_indicators', False)

        # Start with all indicators
        query = db.session.query(IndicatorBank)

        # If section_id is provided, get the section's filters
        if section_id:
            from app.models import FormSection
            section = FormSection.query.get(section_id)
            if section and section.indicator_filters_list:
                filters = section.indicator_filters_list

        # Apply each filter
        for filter_obj in filters:
            field = filter_obj.get('field')
            values = filter_obj.get('values', [])

            if not field or not values:
                continue

            try:
                if field == 'type':
                    query = query.filter(IndicatorBank.type.in_(values))
                elif field == 'unit':
                    query = query.filter(IndicatorBank.unit.in_(values))
                elif field == 'emergency':
                    # Convert string values to boolean
                    bool_values = [v.lower() == 'true' for v in values]
                    query = query.filter(IndicatorBank.emergency.in_(bool_values))
                elif field == 'archived':
                    # Convert string values to boolean
                    bool_values = [v.lower() == 'true' for v in values]
                    query = query.filter(IndicatorBank.archived.in_(bool_values))
                elif field == 'related_programs':
                    # For related programs, we need to check if any of the selected values
                    # are contained in the comma-separated related_programs field
                    conditions = []
                    for value in values:
                        conditions.append(IndicatorBank.related_programs.like(f'%{value}%'))
                    if conditions:
                        query = query.filter(db.or_(*conditions))
                elif field == 'sector':
                    # For sector, check JSONB field across all levels (primary, secondary, tertiary)
                    # or only primary level if primary_only is set
                    conditions = []
                    primary_only = filter_obj.get('primary_only', False)

                    for value in values:
                        if primary_only:
                            # Only check primary level using JSONB operators
                            conditions.append(IndicatorBank.sector.cast(db.Text).like(f'%"primary": "{value}"%'))
                        else:
                            # Check all levels (primary, secondary, tertiary) using JSONB operators
                            conditions.append(IndicatorBank.sector.cast(db.Text).like(f'%"primary": "{value}"%'))
                            conditions.append(IndicatorBank.sector.cast(db.Text).like(f'%"secondary": "{value}"%'))
                            conditions.append(IndicatorBank.sector.cast(db.Text).like(f'%"tertiary": "{value}"%'))
                    if conditions:
                        query = query.filter(db.or_(*conditions))
                elif field == 'subsector':
                    # For subsector, check JSONB field across all levels (primary, secondary, tertiary)
                    # or only primary level if primary_only is set
                    conditions = []
                    primary_only = filter_obj.get('primary_only', False)

                    for value in values:
                        if primary_only:
                            # Only check primary level using JSONB operators
                            conditions.append(IndicatorBank.sub_sector.cast(db.Text).like(f'%"primary": "{value}"%'))
                        else:
                            # Check all levels (primary, secondary, tertiary) using JSONB operators
                            conditions.append(IndicatorBank.sub_sector.cast(db.Text).like(f'%"primary": "{value}"%'))
                            conditions.append(IndicatorBank.sub_sector.cast(db.Text).like(f'%"secondary": "{value}"%'))
                            conditions.append(IndicatorBank.sub_sector.cast(db.Text).like(f'%"tertiary": "{value}"%'))
                    if conditions:
                        query = query.filter(db.or_(*conditions))
            except Exception as filter_error:
                current_app.logger.error(f"Error applying filter {field}: {str(filter_error)}")
                # Continue with other filters instead of failing completely
                continue

        # Order by name for consistent results
        query = query.order_by(IndicatorBank.name)

        try:
            count = query.count()
            result = {
                'count': count,
                'success': True
            }
        except Exception as count_error:
            current_app.logger.error(f"Error executing count query: {str(count_error)}")
            # Return a safe default
            result = {
                'count': 0,
                'success': True
            }

        # Include the actual indicators if requested
        if include_indicators:
            try:
                indicators = query.all()
                result['indicators'] = [{
                    'id': indicator.id,
                    'name': indicator.name,
                    'type': indicator.type,
                    'unit': indicator.unit,
                    'fdrs_kpi_code': getattr(indicator, 'fdrs_kpi_code', None),
                    'definition': indicator.definition,
                    'related_programs': indicator.related_programs,
                    'emergency': indicator.emergency,
                    'archived': indicator.archived,
                    'sector': indicator.sector,
                    'sub_sector': indicator.sub_sector
                } for indicator in indicators]
            except Exception as indicators_error:
                current_app.logger.error(f"Error retrieving indicators: {str(indicators_error)}")
                result['indicators'] = []

        return json_ok(**result)

    except Exception as e:
        # Ensure database session is clean on error
        with suppress(Exception):
            request_transaction_rollback()

        current_app.logger.error(f"Error in get_filtered_indicator_count: {str(e)}", exc_info=True)
        return json_server_error(GENERIC_ERROR_MESSAGE)

# === Common Words Management Routes ===
@bp.route("/common_words", methods=["GET"])
@permission_required('admin.indicator_bank.edit')
def manage_common_words():
    """Manage common words used in indicators."""
    search = request.args.get('search', '')

    query = CommonWord.query

    if search:
        query = query.filter(
            or_(
                CommonWord.term.contains(search),
                CommonWord.meaning.contains(search)
            )
        )

    common_words = query.order_by(CommonWord.term).all()

    return render_template("admin/common_words/manage_common_words.html",
                         common_words=common_words,
                         search=search,
                         title="Manage Common Words")

@bp.route("/common_words/add", methods=["POST"])
@permission_required('admin.indicator_bank.edit')
def add_common_word_modal():
    """Add a new common word via modal."""
    form = CommonWordForm(request.form)

    if form.validate_on_submit():
        try:
            new_common_word = CommonWord(
                created_by_user_id=current_user.id
            )

            form.populate_common_word(new_common_word)

            db.session.add(new_common_word)
            db.session.flush()

            return json_ok(
                success=True,
                message=f"Common word '{new_common_word.term}' added successfully.",
                common_word={
                    'id': new_common_word.id,
                    'term': new_common_word.term,
                    'meaning': new_common_word.meaning,
                    'is_active': new_common_word.is_active,
                    'created_at': new_common_word.created_at.strftime('%Y-%m-%d') if new_common_word.created_at else None
                }
            )

        except Exception as e:
            request_transaction_rollback()
            error_msg = GENERIC_ERROR_MESSAGE
            current_app.logger.error(f"Error adding common word: {e}", exc_info=True)

            return json_bad_request(error_msg, success=False)
    else:
        return json_form_errors(form, "Please correct the errors in the form.")

@bp.route("/common_words/edit/<int:id>", methods=["POST"])
@permission_required('admin.indicator_bank.edit')
def edit_common_word_modal(id):
    """Edit an existing common word via modal."""
    common_word = CommonWord.query.get_or_404(id)
    form = CommonWordForm(request.form)

    if form.validate_on_submit():
        try:
            form.populate_common_word(common_word)
            db.session.flush()

            return json_ok(
                success=True,
                message=f"Common word '{common_word.term}' updated successfully.",
                common_word={
                    'id': common_word.id,
                    'term': common_word.term,
                    'meaning': common_word.meaning,
                    'is_active': common_word.is_active,
                    'updated_at': common_word.updated_at.strftime('%Y-%m-%d') if common_word.updated_at else None
                }
            )

        except Exception as e:
            request_transaction_rollback()
            current_app.logger.error(f"Error updating common word {id}: {e}", exc_info=True)
            return json_bad_request("Error updating common word.", success=False)
    else:
        return json_form_errors(form, "Please correct the errors in the form.")


@bp.route("/common_words/delete/<int:id>", methods=["POST"])
@permission_required('admin.indicator_bank.edit')
def delete_common_word(id):
    """Delete a common word."""
    common_word = CommonWord.query.get_or_404(id)

    try:
        term_name = common_word.term
        db.session.delete(common_word)
        db.session.flush()

        flash(f"Common word '{term_name}' deleted successfully.", "success")

    except Exception as e:
        request_transaction_rollback()
        flash("An error occurred. Please try again.", "danger")
        current_app.logger.error(f"Error deleting common word {id}: {e}", exc_info=True)

    return redirect(url_for("system_admin.manage_common_words"))

@bp.route("/common_words/export", methods=["GET"])
@permission_required('admin.indicator_bank.view')
def export_common_words():
    """Export common words to Excel file."""
    try:
        # Get all active common words
        common_words = CommonWord.query.filter_by(is_active=True).all()

        # Prepare data for Excel (dynamic translation columns)
        translatable = current_app.config.get("TRANSLATABLE_LANGUAGES") or []
        display_names = getattr(Config, "ALL_LANGUAGES_DISPLAY_NAMES", {}) or {}

        data = []
        for word in common_words:
            row = {
                'Term': word.term,
                'Meaning': word.meaning,
            }
            for code in translatable:
                header = display_names.get(code, code.upper())
                row[header] = word.get_meaning_translation(code) if word.meaning_translations else ''
            row.update({
                'Active': 'TRUE' if word.is_active else 'FALSE',
                'Created': word.created_at.strftime('%Y-%m-%d') if word.created_at else '',
                'Updated': word.updated_at.strftime('%Y-%m-%d') if word.updated_at else ''
            })
            data.append(row)

        # Create DataFrame and export to Excel
        df = pd.DataFrame(data)

        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Common Words', index=False)

            # Auto-adjust column widths
            worksheet = writer.sheets['Common Words']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    with suppress(Exception):
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)

        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'common_words_export_{timestamp}.xlsx'

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        current_app.logger.exception("Error exporting common words: %s", e)
        flash('Error exporting common words.', 'danger')
        return redirect(url_for('system_admin.manage_common_words'))

@bp.route("/common_words/import", methods=["POST"])
@permission_required('admin.indicator_bank.edit')
def import_common_words():
    """Import common words from Excel file."""
    try:
        if 'excel_file' not in request.files:
            flash('No file selected', 'danger')
            return redirect(url_for('system_admin.manage_common_words'))

        file = request.files['excel_file']
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(url_for('system_admin.manage_common_words'))

        valid, error_msg, ext = validate_upload_extension_and_mime(file, EXCEL_EXTENSIONS)
        if not valid:
            flash(error_msg or 'Invalid file format. Please upload an Excel file (.xlsx or .xls).', 'danger')
            return redirect(url_for('system_admin.manage_common_words'))

        # Read Excel file
        df = pd.read_excel(file, engine='openpyxl')

        # Validate required columns
        required_columns = ['Term', 'Meaning']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            flash(f'Missing required columns: {", ".join(missing_columns)}', 'danger')
            return redirect(url_for('system_admin.manage_common_words'))

        # Process the data
        overwrite = request.form.get('overwrite_existing') == 'on'
        imported_count = 0
        updated_count = 0
        errors = []

        for index, row in df.iterrows():
            try:
                term = str(row['Term']).strip()
                meaning = str(row['Meaning']).strip()

                if not term or not meaning:
                    continue

                # Check if term already exists
                existing_word = CommonWord.query.filter_by(term=term).first()

                if existing_word and not overwrite:
                    errors.append(f'Term "{term}" already exists (row {index + 2})')
                    continue

                # Prepare translations
                translations = {}
                # Accept translation columns by display name or ISO code
                translatable = current_app.config.get("TRANSLATABLE_LANGUAGES") or []
                display_names = getattr(Config, "ALL_LANGUAGES_DISPLAY_NAMES", {}) or {}
                # Build candidate column names for each lang
                candidates = {}
                for code in translatable:
                    candidates[code] = {
                        code,  # "fr"
                        code.upper(),  # "FR"
                        display_names.get(code, '').strip(),  # "French"
                    }
                for lang_code, possible_cols in candidates.items():
                    for col in possible_cols:
                        if not col:
                            continue
                        if col in df.columns and pd.notna(row[col]):
                            translation = str(row[col]).strip()
                            if translation:
                                translations[lang_code] = translation
                            break

                # Set active status
                is_active = True
                if 'Active' in df.columns and pd.notna(row['Active']):
                    active_value = str(row['Active']).strip().upper()
                    is_active = active_value in ['TRUE', 'YES', '1', 'ACTIVE']

                if existing_word and overwrite:
                    # Update existing word
                    existing_word.meaning = meaning
                    existing_word.meaning_translations = translations
                    existing_word.is_active = is_active
                    existing_word.updated_at = utcnow()
                    updated_count += 1
                else:
                    # Create new word
                    new_word = CommonWord(
                        term=term,
                        meaning=meaning,
                        meaning_translations=translations,
                        is_active=is_active,
                        created_by_user_id=current_user.id
                    )
                    db.session.add(new_word)
                    imported_count += 1

            except Exception as e:
                current_app.logger.warning("Error processing common words row %d: %s", index + 2, e, exc_info=True)
                errors.append(f'Error processing row {index + 2}.')
                continue

        # Commit changes
        db.session.flush()

        # Show results
        if imported_count > 0 or updated_count > 0:
            success_msg = f'Successfully imported {imported_count} new common words'
            if updated_count > 0:
                success_msg += f' and updated {updated_count} existing ones'
            flash(success_msg, 'success')

        if errors:
            error_msg = f'Import completed with {len(errors)} errors: ' + '; '.join(errors[:5])
            if len(errors) > 5:
                error_msg += f'... and {len(errors) - 5} more'
            flash(error_msg, 'warning')

        return redirect(url_for('system_admin.manage_common_words'))

    except Exception as e:
        current_app.logger.exception("Error importing common words: %s", e)
        flash('Error importing common words.', 'danger')
        return redirect(url_for('system_admin.manage_common_words'))

@bp.route("/common_words/template", methods=["GET"])
@permission_required('admin.indicator_bank.view')
def download_common_words_template():
    """Download Excel template for common words import."""
    try:
        translatable = current_app.config.get("TRANSLATABLE_LANGUAGES") or []
        display_names = getattr(Config, "ALL_LANGUAGES_DISPLAY_NAMES", {}) or {}

        # Create sample data for template (translation columns included, values left blank by default)
        base_rows = [
            {
                'Term': 'Emergency',
                'Meaning': 'A serious, unexpected, and often dangerous situation requiring immediate action.',
                'Active': 'TRUE'
            },
            {
                'Term': 'Response',
                'Meaning': 'An answer or reply, especially in the context of humanitarian aid or disaster relief.',
                'Active': 'TRUE'
            }
        ]
        sample_data = []
        for r in base_rows:
            row = dict(r)
            for code in translatable:
                header = display_names.get(code, code.upper())
                row.setdefault(header, '')
            sample_data.append(row)

        # Create DataFrame
        df = pd.DataFrame(sample_data)

        # Create Excel file in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Common Words Template', index=False)

            # Auto-adjust column widths
            worksheet = writer.sheets['Common Words Template']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    with suppress(Exception):
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='common_words_template.xlsx'
        )

    except Exception as e:
        current_app.logger.exception("Error downloading template: %s", e)
        flash("An error occurred. Please try again.", "danger")
        return redirect(url_for('system_admin.manage_common_words'))
