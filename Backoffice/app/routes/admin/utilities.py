from app.utils.transactions import request_transaction_rollback
from contextlib import suppress
# File: Backoffice/app/routes/admin/utilities.py
from app.utils.datetime_helpers import utcnow
from app.utils.advanced_validation import validate_upload_extension_and_mime
from app.utils.file_parsing import EXCEL_EXTENSIONS
"""
Utilities Module - Import/Export, Session Management, and Utility Functions

Note: For full translation file management functionality, install the polib package:
    pip install polib
"""

from flask import Blueprint, request, flash, redirect, url_for, current_app, make_response, render_template, send_file
from flask_login import current_user
from flask_wtf import csrf
from flask_babel import _
from app import db
from app.models import (
    IndicatorBank, IndicatorBankHistory, IndicatorSuggestion,
    Sector, SubSector,
    UserSessionLog, SecurityEvent
)
from app.forms.content import TranslationForm
from app.routes.admin.shared import admin_required, permission_required, permission_required_any
from app.utils.request_utils import is_json_request, get_request_data
from app.utils.auto_translator import translate_text as auto_translate_text
from app.extensions import limiter
from config.config import Config
from app.utils.api_helpers import GENERIC_ERROR_MESSAGE, get_json_safe
from app.utils.error_handling import handle_json_view_exception, handle_view_exception
from app.utils.api_responses import json_bad_request, json_error, json_forbidden, json_ok, json_server_error, require_json_data, require_json_keys
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
import json
from sqlalchemy import and_, inspect
import logging
import io
import zipfile
import shutil

logger = logging.getLogger(__name__)

bp = Blueprint("utilities", __name__, url_prefix="/admin")

# === Import/Export Routes ===
@bp.route("/indicator_bank/import", methods=["GET", "POST"])
@permission_required('admin.indicator_bank.edit')
def import_indicators():
    """Import indicators from Excel file"""
    if request.method == 'POST':
        temp_path = None
        try:
            if 'file' not in request.files:
                return json_bad_request('No file selected.')

            file = request.files['file']
            if file.filename == '':
                return json_bad_request('No file selected.')

            valid, error_msg, ext = validate_upload_extension_and_mime(file, EXCEL_EXTENSIONS)
            if valid:
                filename = secure_filename(file.filename)

                # Save temporary file
                from app.utils.file_paths import get_temp_upload_path
                temp_dir = get_temp_upload_path()
                os.makedirs(temp_dir, exist_ok=True)
                temp_path = os.path.join(temp_dir, filename)
                file.save(temp_path)

                # Process the Excel file
                current_app.logger.info(f"Processing import file: {temp_path}")
                result = _process_indicator_import(temp_path)
                current_app.logger.info(f"Import result: {result}")

                if result['success']:
                    message = f"Successfully imported {result['imported']} indicators. {result['updated']} indicators were updated."
                    if result['errors']:
                        message += f" Encountered {len(result['errors'])} errors during import."
                    return json_ok(message=message)
                else:
                    current_app.logger.error(f"Import failed: {result['message']}")
                    return json_bad_request(f"Import failed: {result['message']}")
            else:
                return json_bad_request(error_msg or _('Please upload an Excel file (.xlsx or .xls).'))

        except Exception as e:
            return handle_json_view_exception(e, 'Error processing import file.', status_code=500)
        finally:
            # Clean up temporary file with error handling
            if temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except (PermissionError, OSError) as e:
                    current_app.logger.warning(f"Could not remove temporary file {temp_path}: {e}")
                    # Schedule cleanup for later (Windows file locking issue)
                    import threading
                    def delayed_cleanup():
                        import time
                        time.sleep(5)  # Wait 5 seconds
                        with suppress(Exception):
                            os.remove(temp_path)
                    threading.Thread(target=delayed_cleanup, daemon=True).start()

    # For GET requests, redirect to the indicator bank page
    return redirect(url_for("system_admin.manage_indicator_bank"))

@bp.route("/indicator_bank/change_history", methods=["GET"])
@permission_required('admin.audit.view')
def indicator_change_history():
    """View indicator change history"""
    # Get all change history records (no pagination - AG Grid handles pagination client-side)
    changes = IndicatorBankHistory.query.order_by(
        IndicatorBankHistory.created_at.desc()
    ).all()

    return render_template("admin/indicator_bank/change_history.html",
                         changes=changes,
                         title="Indicator Change History")

# === Session Management Routes ===
@bp.route("/utilities/sessions/cleanup", methods=["POST"])
@permission_required('admin.analytics.view')
def cleanup_sessions():
    """Cleanup expired sessions"""
    try:
        # Define session timeout (2 hours)
        timeout_threshold = utcnow() - timedelta(hours=2)

        # Clean up expired sessions if session tracking exists
        if inspect(db.engine).has_table(UserSessionLog.__tablename__):
            expired_sessions = UserSessionLog.query.filter(
                and_(
                    UserSessionLog.last_activity < timeout_threshold,
                    UserSessionLog.ended_at.is_(None)
                )
            ).all()

            for session in expired_sessions:
                session.ended_at = utcnow()
                session.end_reason = 'timeout'

            db.session.flush()

            flash(_("Cleaned up %(count)d expired sessions.", count=len(expired_sessions)), "success")
        else:
            flash(_("Session tracking is not configured."), "warning")

    except Exception as e:
        return handle_view_exception(e, _("Error during session cleanup."), redirect_endpoint="admin.admin_dashboard")

    return redirect(url_for("admin.admin_dashboard"))

@bp.route("/utilities/sessions/show_all", methods=["GET"])
@permission_required('admin.analytics.view')
def show_all_sessions():
    """Show all active sessions"""
    try:
        sessions = []

        if inspect(db.engine).has_table(UserSessionLog.__tablename__):
            active_sessions = UserSessionLog.query.filter(
                UserSessionLog.ended_at.is_(None)
            ).order_by(UserSessionLog.last_activity.desc()).all()

            for session in active_sessions:
                sessions.append({
                    'id': session.id,
                    'user_id': session.user_id,
                    'user_name': session.user.name if session.user else 'Unknown',
                    'started_at': session.started_at,
                    'last_activity': session.last_activity,
                    'ip_address': getattr(session, 'ip_address', None),
                    'user_agent': getattr(session, 'user_agent', None)
                })

        return render_template("admin/utilities/all_sessions.html",
                             sessions=sessions,
                             title="All Active Sessions")

    except Exception as e:
        return handle_view_exception(e, _("Error retrieving session data."), redirect_endpoint="admin.admin_dashboard")

# === Translation Management Routes ===
@bp.route("/translations/manage", methods=["GET"])
@permission_required("admin.translations.manage")
def manage_translations():
    """Manage translations for the application"""
    from config import Config
    # Use the dynamically loaded languages from app config, not static Config
    languages = current_app.config.get('SUPPORTED_LANGUAGES', Config.LANGUAGES)
    language_names = Config.LANGUAGE_DISPLAY_NAMES

    # Get all translation files (robustly via polib)
    translation_data = {}
    all_msgids = set()
    msgid_sources = {}

    # Try to import polib, but make it optional
    try:
        import polib  # type: ignore
    except ImportError:
        current_app.logger.warning("polib not available - translation file management will be limited")
        polib = None
        # Show a flash message to the user about the missing dependency
        flash(_("Warning: polib package not available. Translation file management will be limited. Install with: pip install polib"), "warning")

    for lang in languages:
        # Use absolute path based on Flask app root path
        po_file_path = os.path.join(current_app.root_path, 'translations', lang, 'LC_MESSAGES', 'messages.po')
        translations = {}
        if os.path.exists(po_file_path) and polib:
            with suppress(Exception):
                po = polib.pofile(po_file_path)
                for entry in po:
                    # Skip header entry
                    if not entry.msgid or entry.msgid == "":
                        continue
                    # Skip PO file metadata entries (header entries that look like metadata)
                    # These typically contain strings like "Project-Id-Version", "POT-Creation-Date", etc.
                    # Check if the msgid contains multiple metadata keys, which indicates it's the entire header block
                    msgid_lower = entry.msgid.lower()
                    metadata_keys = [
                        'project-id-version', 'report-msgid-bugs-to', 'pot-creation-date',
                        'po-revision-date', 'last-translator', 'language-team', 'mime-version',
                        'content-type', 'content-transfer-encoding', 'plural-forms', 'generated-by'
                    ]
                    metadata_key_count = sum(1 for key in metadata_keys if key in msgid_lower)
                    # If msgid contains 3+ metadata keys, it's likely the entire header block
                    if metadata_key_count >= 3:
                        continue
                    # Record msgid regardless of whether msgstr is empty
                    all_msgids.add(entry.msgid)

                    # Handle plural forms: check msgstr_plural if msgstr is empty
                    if entry.msgstr:
                        translations[entry.msgid] = entry.msgstr
                    elif hasattr(entry, 'msgstr_plural') and entry.msgstr_plural:
                        # For plural forms, use the first plural form as the translation
                        # (or concatenate all forms, but first is usually sufficient for display)
                        first_plural = entry.msgstr_plural.get(0, '') or (list(entry.msgstr_plural.values())[0] if entry.msgstr_plural else '')
                        translations[entry.msgid] = first_plural
                    else:
                        translations[entry.msgid] = ""

                    # Capture source page once per msgid
                    if entry.occurrences and entry.msgid not in msgid_sources:
                        src_path, _ = entry.occurrences[0]
                        page_name = _extract_page_name(src_path)
                        msgid_sources[entry.msgid] = page_name
        translation_data[lang] = {
            'name': language_names.get(lang, lang.upper()),
            'translations': translations
        }

    # Sort all message IDs for consistent display
    all_msgids = sorted(list(all_msgids))

    # Count empty translations for each language (for frontend auto-translate)
    empty_translation_counts = {}
    empty_translation_msgids = {}  # Store the actual msgids that need translation
    for lang in languages:
        if lang == 'en':  # Skip English as it's the source language
            continue
        empty_count = 0
        empty_msgids = []
        lang_translations = translation_data.get(lang, {}).get('translations', {})
        # Also check the actual PO file for plural forms that might not be in translations dict
        po_file_path = os.path.join(current_app.root_path, 'translations', lang, 'LC_MESSAGES', 'messages.po')
        po_cache = None
        if polib and os.path.exists(po_file_path):
            with suppress(Exception):
                po_cache = polib.pofile(po_file_path)

        for msgid in all_msgids:
            translation = lang_translations.get(msgid, '').strip()

            # If translation appears empty, check if it's a plural form in the PO file
            if not translation and po_cache:
                with suppress(Exception):  # If we can't find the entry, fall back to checking the dict
                    entry = po_cache.find(msgid)
                    if entry and hasattr(entry, 'msgstr_plural') and entry.msgstr_plural:
                        # Check if any plural form has a non-empty translation
                        has_plural_translation = any(
                            v and str(v).strip()
                            for v in entry.msgstr_plural.values()
                        )
                        if has_plural_translation:
                            translation = "PLURAL_FORM"  # Mark as having translation

            if not translation:  # Count empty or whitespace-only translations
                empty_count += 1
                empty_msgids.append(msgid)
        empty_translation_counts[lang] = empty_count
        empty_translation_msgids[lang] = empty_msgids

    # Add variables needed by JavaScript auto-translate functionality
    from config import Config
    try:
        all_languages = current_app.config.get('SUPPORTED_LANGUAGES', Config.LANGUAGES)
        # Exclude English from translatable languages since it's the source language
        translatable_languages = [lang for lang in all_languages if lang != 'en']
    except Exception as e:
        logger.debug("SUPPORTED_LANGUAGES config lookup failed: %s", e)
        all_languages = getattr(Config, 'LANGUAGES', ['en'])
        # Exclude English from translatable languages since it's the source language
        translatable_languages = [lang for lang in (all_languages or []) if lang != 'en']

    # Return JSON for API requests (mobile app)
    if is_json_request():
        translations_list = []
        for msgid in all_msgids:
            translation_entry = {
                'msgid': msgid,
                'source': msgid_sources.get(msgid, 'unknown'),
                'translations': {}
            }
            for lang in languages:
                lang_data = translation_data.get(lang, {})
                lang_translations = lang_data.get('translations', {})
                translation_entry['translations'][lang] = {
                    'text': lang_translations.get(msgid, ''),
                    'language_name': lang_data.get('name', lang.upper())
                }
            translations_list.append(translation_entry)

        return json_ok(
            translations=translations_list,
            count=len(translations_list),
            languages=[{'code': lang, 'name': language_names.get(lang, lang.upper())} for lang in languages],
            empty_translation_counts=empty_translation_counts,
            polib_available=polib is not None,
        )

    return render_template('admin/translations/manage_translations.html',
                         translation_data=translation_data,
                         all_msgids=all_msgids,
                         languages=languages,
                         language_names=language_names,
                         msgid_sources=msgid_sources,
                         TRANSLATABLE_LANGUAGES=translatable_languages,
                         empty_translation_counts=empty_translation_counts,
                         empty_translation_msgids=empty_translation_msgids,
                         polib_available=polib is not None)


def _translations_po_path(lang_code: str) -> str:
    return os.path.join(current_app.root_path, "translations", lang_code, "LC_MESSAGES", "messages.po")


def _translations_pot_path() -> str:
    return os.path.join(current_app.root_path, "translations", "messages.pot")


def _entry_to_display_msgstr(entry) -> str:
    # Mirror the UI behavior: show msgstr if present, else first plural form if any.
    try:
        if getattr(entry, "msgstr", None):
            return entry.msgstr or ""
        plural = getattr(entry, "msgstr_plural", None)
        if plural:
            if isinstance(plural, dict) and plural:
                if plural.get(0):
                    return plural.get(0) or ""
                return next(iter(plural.values()), "") or ""
    except Exception as e:
        logger.debug("plural extraction failed: %s", e)
    return ""


@bp.route("/translations/export", methods=["GET"])
@permission_required("admin.translations.manage")
def export_translations():
    """
    Export translations in a safe, round-trippable format.

    Supported query params:
      - format: xlsx | po | po-zip
      - lang: required when format=po
    """
    from config import Config
    languages = current_app.config.get("SUPPORTED_LANGUAGES", Config.LANGUAGES)
    fmt = (request.args.get("format") or "xlsx").strip().lower()

    try:
        import polib  # type: ignore
    except ImportError:
        flash(_("polib is not installed; export is unavailable."), "danger")
        return redirect(url_for("utilities.manage_translations"))

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    if fmt == "po":
        lang = (request.args.get("lang") or "").strip().lower()
        if not lang:
            flash(_("Please select a language to export."), "warning")
            return redirect(url_for("utilities.manage_translations"))
        if lang not in languages:
            flash(_("Unknown/disabled language: %(lang)s", lang=lang), "danger")
            return redirect(url_for("utilities.manage_translations"))

        po_path = _translations_po_path(lang)
        if not os.path.exists(po_path):
            flash(_("PO file not found for %(lang)s.", lang=lang), "danger")
            return redirect(url_for("utilities.manage_translations"))

        return send_file(
            po_path,
            as_attachment=True,
            download_name=f"messages_{lang}_{timestamp}.po",
            mimetype="text/x-gettext-translation",
        )

    if fmt == "po-zip":
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for lang in languages:
                po_path = _translations_po_path(lang)
                if not os.path.exists(po_path):
                    continue
                arcname = f"{lang}/LC_MESSAGES/messages.po"
                zf.write(po_path, arcname=arcname)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"translations_po_{timestamp}.zip",
            mimetype="application/zip",
        )

    # Default: XLSX export (all languages in one workbook)
    # Source of truth for msgids is the POT when available.
    pot_path = _translations_pot_path()
    msgids: list[str] = []
    if os.path.exists(pot_path):
        with suppress(Exception):
            pot = polib.pofile(pot_path)
            for entry in pot:
                if getattr(entry, "obsolete", False):
                    continue
                if entry.msgid:
                    msgids.append(entry.msgid)

    if not msgids:
        # Fallback: union of msgids from existing PO files
        msgid_set = set()
        for lang in languages:
            po_path = _translations_po_path(lang)
            if not os.path.exists(po_path):
                continue
            with suppress(Exception):
                po = polib.pofile(po_path)
                for entry in po:
                    if getattr(entry, "obsolete", False):
                        continue
                    if entry.msgid:
                        msgid_set.add(entry.msgid)
        msgids = sorted(msgid_set)

    # Preload PO files per language for efficient lookup
    po_maps: dict[str, dict[str, object]] = {}
    for lang in languages:
        po_path = _translations_po_path(lang)
        entries: dict[str, object] = {}
        if os.path.exists(po_path):
            with suppress(Exception):
                po = polib.pofile(po_path)
                for entry in po:
                    if getattr(entry, "obsolete", False):
                        continue
                    if not entry.msgid:
                        continue
                    # Note: msgctxt is ignored here (current UI doesn't surface it).
                    entries[entry.msgid] = entry
        po_maps[lang] = entries

    wb = Workbook()
    ws = wb.active
    ws.title = "translations"

    # Add warning note above headers
    ws.merge_cells("A1:B1")
    note_cell = ws.cell(row=1, column=1)
    note_cell.value = "⚠️ DO NOT MODIFY: Source and msgid columns are read-only. Only edit translation columns."
    note_cell.font = Font(bold=True, color="DC2626")  # red-600
    note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    note_cell.fill = PatternFill("solid", fgColor="FEE2E2")  # red-100 background
    ws.row_dimensions[1].height = 30

    # Include a source column to help translators understand where a string comes from.
    # Order: source, msgid, then languages
    header = ["source", "msgid"] + list(languages)
    ws.append(header)
    # Freeze header row and source + msgid columns
    ws.freeze_panes = "C3"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E5E7EB")  # gray-200
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=2, column=c)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
    ws.row_dimensions[2].height = 24

    # Reasonable default widths
    ws.column_dimensions["A"].width = 28  # source
    ws.column_dimensions["B"].width = 70  # msgid
    for idx, lang in enumerate(languages, start=3):
        ws.column_dimensions[get_column_letter(idx)].width = 55

    # Build msgid -> source mapping (prefer POT occurrences; fall back to PO occurrences)
    msgid_sources: dict[str, str] = {}
    if os.path.exists(pot_path):
        with suppress(Exception):
            pot = polib.pofile(pot_path)
            for entry in pot:
                if getattr(entry, "obsolete", False):
                    continue
                if not entry.msgid:
                    continue
                if entry.occurrences:
                    src_path, _ = entry.occurrences[0]
                    msgid_sources[entry.msgid] = _extract_page_name(src_path)

    wrap = Alignment(vertical="top", wrap_text=True)
    for msgid in msgids:
        src = msgid_sources.get(msgid, "")
        if not src:
            # Fall back to any locale's PO entry occurrences
            for lang in languages:
                entry = po_maps.get(lang, {}).get(msgid)
                if entry is not None and getattr(entry, "occurrences", None):
                    with suppress(Exception):
                        src_path, _ = entry.occurrences[0]
                        src = _extract_page_name(src_path)
                        break

        row = [src, msgid]  # source first, then msgid
        for lang in languages:
            entry = po_maps.get(lang, {}).get(msgid)
            row.append(_entry_to_display_msgstr(entry) if entry is not None else "")
        ws.append(row)

    # Apply wrap alignment to all data rows (starting from row 3: note=1, header=2, data=3+)
    for r in range(3, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = wrap
            # Sheet will be protected; lock only source/msgid columns, keep translations editable
            ws.cell(row=r, column=c).protection = Protection(locked=(c in (1, 2)))

    # Protect the sheet but allow editing unlocked (translation) cells.
    #
    # Important: in OOXML, most `sheetProtection` boolean flags are *locks*:
    # setting them to True generally DISALLOWS that action (e.g., `selectUnlockedCells="1"` blocks selecting unlocked cells).
    # So we keep these flags False (defaults) and only enable sheet protection.
    with suppress(Exception):
        ws.protection.selectLockedCells = False
    with suppress(Exception):
        ws.protection.selectUnlockedCells = False
    with suppress(Exception):
        ws.protection.autoFilter = False
    with suppress(Exception):
        ws.protection.sort = False
    with suppress(Exception):
        ws.protection.formatColumns = False
    with suppress(Exception):
        ws.protection.formatRows = False
    # Enable protection last
    with suppress(Exception):
        ws.protection.sheet = True

    # Add thin grey borders for readability (including note row)
    try:
        border_color = "D1D5DB"  # gray-300
        thin = Side(style="thin", color=border_color)
        table_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).border = table_border
    except Exception as e:
        logger.debug("translation export border apply failed: %s", e)

    # Apply visual formatting to source and msgid columns (gray background) for clarity
    try:
        read_only_fill = PatternFill("solid", fgColor="F3F4F6")  # gray-100
        for r in range(3, ws.max_row + 1):
            ws.cell(row=r, column=1).fill = read_only_fill  # source column
            ws.cell(row=r, column=2).fill = read_only_fill  # msgid column
    except Exception as e:
        logger.debug("translation export fill apply failed: %s", e)

    # Enable filtering on all columns (starting from header row 2)
    try:
        ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    except Exception as e:
        logger.debug("translation export auto_filter failed: %s", e)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        as_attachment=True,
        download_name=f"translations_{timestamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/translations/import", methods=["POST"])
@permission_required("admin.translations.manage")
def import_translations():
    """
    Import translations safely:
    - Never changes msgid
    - Never creates new msgids in PO
    - By default only updates non-empty incoming translations (won't clear existing)
    """
    from config import Config
    languages = current_app.config.get("SUPPORTED_LANGUAGES", Config.LANGUAGES)

    fmt = (request.form.get("format") or request.form.get("import_format") or "").strip().lower()
    if not fmt:
        fmt = "xlsx"

    only_non_empty = (request.form.get("only_non_empty") or "").strip().lower() in ("1", "true", "on", "yes")
    allow_clear = (request.form.get("allow_clear") or "").strip().lower() in ("1", "true", "on", "yes")
    if only_non_empty:
        allow_clear = False  # strongest safety default

    file = request.files.get("file")
    if file is None or not getattr(file, "filename", ""):
        flash(_("No file selected."), "warning")
        return redirect(url_for("utilities.manage_translations"))

    try:
        import polib  # type: ignore
    except ImportError:
        flash(_("polib is not installed; import is unavailable."), "danger")
        return redirect(url_for("utilities.manage_translations"))

    def _backup(po_path: str) -> None:
        try:
            ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            backup_path = f"{po_path}.bak.{ts}"
            shutil.copy2(po_path, backup_path)
        except Exception as e:
            logger.debug("translation import backup failed: %s", e)

    def _apply_incoming_po(current, incoming):
        """
        Apply incoming translations to an existing PO object safely.
        Returns: (updated, skipped_missing, skipped_empty)
        """
        current_map = {}
        for e in current:
            if getattr(e, "obsolete", False):
                continue
            if not e.msgid:
                continue
            # Include msgctxt in the key if present (safer than msgid-only).
            current_map[(e.msgctxt or None, e.msgid)] = e

        updated = 0
        skipped_missing = 0
        skipped_empty = 0

        for inc in incoming:
            if getattr(inc, "obsolete", False):
                continue
            if not inc.msgid:
                continue

            key = (getattr(inc, "msgctxt", None) or None, inc.msgid)
            cur = current_map.get(key)
            if cur is None:
                skipped_missing += 1
                continue

            # Decide the incoming translation value(s)
            inc_singular = (inc.msgstr or "") if hasattr(inc, "msgstr") else ""
            inc_plural = getattr(inc, "msgstr_plural", None) or {}

            # If we only import non-empty, skip blanks (never clear existing translations).
            if only_non_empty:
                if (not inc_singular or not str(inc_singular).strip()) and (not inc_plural or all(not str(v).strip() for v in inc_plural.values())):
                    skipped_empty += 1
                    continue
            elif not allow_clear:
                # If not allowing clear, also skip empty updates.
                if (not inc_singular or not str(inc_singular).strip()) and (not inc_plural or all(not str(v).strip() for v in inc_plural.values())):
                    skipped_empty += 1
                    continue

            # Update only msgstr / msgstr_plural (never msgid)
            if getattr(cur, "msgid_plural", None):
                # Prefer plural updates if provided; otherwise update index 0 only.
                if inc_plural:
                    for k, v in inc_plural.items():
                        if (only_non_empty or (not allow_clear)) and (v is None or not str(v).strip()):
                            continue
                        cur.msgstr_plural[int(k)] = "" if (v is None) else str(v)
                    updated += 1
                else:
                    if allow_clear or (inc_singular and str(inc_singular).strip()) or (not only_non_empty and inc_singular is not None):
                        cur.msgstr_plural[0] = "" if (inc_singular is None) else str(inc_singular)
                        updated += 1
            else:
                if allow_clear or (inc_singular and str(inc_singular).strip()) or (not only_non_empty and inc_singular is not None):
                    cur.msgstr = "" if (inc_singular is None) else str(inc_singular)
                    updated += 1

        return updated, skipped_missing, skipped_empty

    if fmt == "po-zip":
        # Import a ZIP containing multiple locales: <lang>/LC_MESSAGES/messages.po
        if not file.filename.lower().endswith(".zip"):
            flash(_("Please upload a .zip file for PO ZIP import."), "warning")
            return redirect(url_for("utilities.manage_translations"))

        temp_zip_path = None
        temp_dir = None
        try:
            filename = secure_filename(file.filename)
            from app.utils.file_paths import get_temp_upload_path

            temp_dir = get_temp_upload_path()
            os.makedirs(temp_dir, exist_ok=True)
            temp_zip_path = os.path.join(temp_dir, filename)
            file.save(temp_zip_path)

            processed_locales = 0
            skipped_locales = 0
            total_updated = 0
            total_missing = 0
            total_empty = 0

            per_lang = {}

            with zipfile.ZipFile(temp_zip_path, "r") as zf:
                for name in zf.namelist():
                    if not name or name.endswith("/"):
                        continue
                    # Accept our export layout and similar layouts
                    if not name.replace("\\", "/").endswith("/LC_MESSAGES/messages.po"):
                        continue

                    parts = name.replace("\\", "/").split("/")
                    # Find the locale segment immediately before LC_MESSAGES
                    try:
                        i = parts.index("LC_MESSAGES")
                        lang = parts[i - 1].strip().lower() if i > 0 else ""
                    except ValueError:
                        continue

                    if not lang:
                        skipped_locales += 1
                        continue
                    if lang == "en":
                        # English is the source language; skip
                        skipped_locales += 1
                        continue
                    if lang not in languages:
                        skipped_locales += 1
                        continue

                    po_path = _translations_po_path(lang)
                    if not os.path.exists(po_path):
                        skipped_locales += 1
                        continue

                    # Write member to a temp .po file for polib
                    member_bytes = zf.read(name)
                    temp_po_path = os.path.join(temp_dir, f"_import_{lang}_messages.po")
                    with open(temp_po_path, "wb") as f:
                        f.write(member_bytes)

                    incoming = polib.pofile(temp_po_path)
                    current = polib.pofile(po_path)

                    updated, skipped_missing, skipped_empty = _apply_incoming_po(current, incoming)
                    per_lang[lang] = {
                        "updated": updated,
                        "skipped_missing": skipped_missing,
                        "skipped_empty": skipped_empty,
                    }

                    if updated > 0:
                        _backup(po_path)
                        current.save(po_path)

                    processed_locales += 1
                    total_updated += updated
                    total_missing += skipped_missing
                    total_empty += skipped_empty

                    with suppress(Exception):
                        os.remove(temp_po_path)

            if processed_locales == 0:
                flash(
                    _("No valid locale PO files found in ZIP. Expected files like: <lang>/LC_MESSAGES/messages.po"),
                    "warning",
                )
                return redirect(url_for("utilities.manage_translations"))

            flash(
                _(
                    "PO ZIP import finished. Locales processed: %(p)d, skipped: %(s)d. Updated: %(u)d, skipped missing: %(m)d, skipped empty: %(e)d.",
                    p=processed_locales,
                    s=skipped_locales,
                    u=total_updated,
                    m=total_missing,
                    e=total_empty,
                ),
                "success",
            )
        except Exception as e:
            current_app.logger.error(f"Translation PO ZIP import failed: {e}", exc_info=True)
            flash(_("An error occurred. Please try again."), "danger")
        finally:
            if temp_zip_path and os.path.exists(temp_zip_path):
                with suppress(Exception):
                    os.remove(temp_zip_path)

        return redirect(url_for("utilities.manage_translations"))

    if fmt == "po":
        lang = (request.form.get("lang") or "").strip().lower()
        if not lang:
            flash(_("Please select a language for PO import."), "warning")
            return redirect(url_for("utilities.manage_translations"))
        if lang not in languages:
            flash(_("Unknown/disabled language: %(lang)s", lang=lang), "danger")
            return redirect(url_for("utilities.manage_translations"))
        if lang == "en":
            flash(_("English (en) is the source language and should not be imported."), "warning")
            return redirect(url_for("utilities.manage_translations"))

        if not file.filename.lower().endswith(".po"):
            flash(_("Please upload a .po file for PO import."), "warning")
            return redirect(url_for("utilities.manage_translations"))

        po_path = _translations_po_path(lang)
        if not os.path.exists(po_path):
            flash(_("PO file not found for %(lang)s.", lang=lang), "danger")
            return redirect(url_for("utilities.manage_translations"))

        # Save to a temp path for polib
        temp_path = None
        try:
            filename = secure_filename(file.filename)
            from app.utils.file_paths import get_temp_upload_path

            temp_dir = get_temp_upload_path()
            os.makedirs(temp_dir, exist_ok=True)
            temp_path = os.path.join(temp_dir, filename)
            file.save(temp_path)

            incoming = polib.pofile(temp_path)
            current = polib.pofile(po_path)

            updated, skipped_missing, skipped_empty = _apply_incoming_po(current, incoming)

            if updated > 0:
                _backup(po_path)
                current.save(po_path)

            flash(
                _(
                    "PO import finished for %(lang)s. Updated: %(u)d, skipped missing: %(m)d, skipped empty: %(e)d.",
                    lang=lang,
                    u=updated,
                    m=skipped_missing,
                    e=skipped_empty,
                ),
                "success",
            )
        except Exception as e:
            current_app.logger.error(f"Translation PO import failed: {e}", exc_info=True)
            flash(_("An error occurred. Please try again."), "danger")
        finally:
            if temp_path and os.path.exists(temp_path):
                with suppress(Exception):
                    os.remove(temp_path)

        return redirect(url_for("utilities.manage_translations"))

    # XLSX import (all languages in one workbook)
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        flash(_("Please upload an Excel file (.xlsx or .xls)."), "warning")
        return redirect(url_for("utilities.manage_translations"))

    try:
        file.stream.seek(0)
    except Exception as e:
        logger.debug("file.stream.seek failed: %s", e)

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb["translations"] if "translations" in wb.sheetnames else wb.active

        # Detect header row: check if row 1 has "msgid", otherwise try row 2 (new format has note in row 1)
        header_row_num = 1
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row:
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            norm_headers = [h.strip().lower() for h in headers]
            if "msgid" not in norm_headers:
                # Try row 2 (new format with note row)
                header_row_num = 2
                header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)

        if not header_row:
            flash(_("Excel file is empty or header row not found."), "warning")
            return redirect(url_for("utilities.manage_translations"))

        headers = [str(h).strip() if h is not None else "" for h in header_row]
        norm_headers = [h.strip().lower() for h in headers]
        if "msgid" not in norm_headers:
            flash(_("Excel import requires a 'msgid' column header."), "danger")
            return redirect(url_for("utilities.manage_translations"))

        msgid_col_idx = norm_headers.index("msgid")

        # Map language columns by header exact match (lang code)
        lang_cols: dict[str, int] = {}
        for idx, h in enumerate(norm_headers):
            if h in languages:
                lang_cols[h] = idx

        if not lang_cols:
            flash(_("No language columns found in Excel header."), "danger")
            return redirect(url_for("utilities.manage_translations"))

        # Collect updates per language
        updates_by_lang: dict[str, dict[str, str]] = {lang: {} for lang in lang_cols.keys() if lang != "en"}

        # Data rows start after header (row 2 if header is row 1, row 3 if header is row 2)
        data_start_row = header_row_num + 1
        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if not row:
                continue
            raw_msgid = row[msgid_col_idx] if msgid_col_idx < len(row) else None
            msgid = (str(raw_msgid) if raw_msgid is not None else "").strip()
            if not msgid:
                continue

            for lang, col_idx in lang_cols.items():
                if lang == "en":
                    continue
                if col_idx >= len(row):
                    continue
                v = row[col_idx]
                if v is None:
                    continue
                s = str(v)
                if only_non_empty and not s.strip():
                    continue
                if (not allow_clear) and (not s.strip()):
                    continue
                updates_by_lang.setdefault(lang, {})[msgid] = s

        total_updated = 0
        total_missing = 0

        for lang, updates in updates_by_lang.items():
            if not updates:
                continue
            po_path = _translations_po_path(lang)
            if not os.path.exists(po_path):
                continue

            po = polib.pofile(po_path)
            entry_map = {}
            for e in po:
                if getattr(e, "obsolete", False):
                    continue
                if not e.msgid:
                    continue
                entry_map[(e.msgctxt or None, e.msgid)] = e

            updated_lang = 0
            missing_lang = 0
            for msgid, msgstr in updates.items():
                cur = entry_map.get((None, msgid))
                if cur is None:
                    missing_lang += 1
                    continue
                if getattr(cur, "msgid_plural", None):
                    cur.msgstr_plural[0] = "" if msgstr is None else str(msgstr)
                else:
                    cur.msgstr = "" if msgstr is None else str(msgstr)
                updated_lang += 1

            if updated_lang > 0:
                _backup(po_path)
                po.save(po_path)

            total_updated += updated_lang
            total_missing += missing_lang

        flash(
            _(
                "Excel import finished. Updated: %(u)d, skipped missing msgids: %(m)d. (Empty updates are ignored by default.)",
                u=total_updated,
                m=total_missing,
            ),
            "success",
        )
    except Exception as e:
        current_app.logger.error(f"Translation Excel import failed: {e}", exc_info=True)
        flash(_("An error occurred. Please try again."), "danger")

    return redirect(url_for("utilities.manage_translations"))

@bp.route("/translations/add", methods=["GET", "POST"])
@permission_required("admin.translations.manage")
def add_translation():
    """Add a new translation for all languages"""
    form = TranslationForm()

    if form.validate_on_submit():
        msgid = form.msgid.data

        # Add translations for all languages
        from config import Config
        languages = current_app.config.get('SUPPORTED_LANGUAGES', Config.LANGUAGES)
        added_count = 0

        # Try to import polib for this function
        try:
            import polib  # type: ignore
        except ImportError:
            current_app.logger.warning("polib not available - translation file updates will be skipped")
            polib = None

        for lang in languages:
            msgstr_field = getattr(form, f'msgstr_{lang}', None)
            msgstr = (msgstr_field.data if msgstr_field else None)
            if msgstr and str(msgstr).strip():  # Only add if translation is provided
                po_file_path = os.path.join(current_app.root_path, 'translations', lang, 'LC_MESSAGES', 'messages.po')

                # Check if polib is available
                if polib is None:
                    current_app.logger.warning(f"polib not available - skipping translation file update for {lang}")
                    continue

                if os.path.exists(po_file_path) and polib:
                    try:
                        po = polib.pofile(po_file_path)
                        entry = po.find(msgid)
                        if entry is None:
                            entry = polib.POEntry(msgid=msgid, msgstr=msgstr)
                            po.append(entry)
                        else:
                            entry.msgstr = msgstr
                        po.save(po_file_path)
                        added_count += 1
                    except Exception as _e:
                        current_app.logger.error(f"Failed to add translation for {lang}: {_e}")

        if added_count > 0:
            flash(_('Translation added successfully for %(count)d language(s)', count=added_count), 'success')
        else:
            flash(_('No translations were added. Please provide at least one translation.'), 'warning')

        return redirect(url_for('utilities.manage_translations'))

    # Pass language metadata to the template for client-side auto-translate
    try:
        from config import Config
        languages = current_app.config.get('SUPPORTED_LANGUAGES', Config.LANGUAGES)
        # ISO codes only for client-side auto-translate (no long keys)
        language_model_keys = {code: (None if code == 'en' else code) for code in languages}
        language_display_names = Config.LANGUAGE_DISPLAY_NAMES
    except Exception as e:
        logger.debug("SUPPORTED_LANGUAGES/LANGUAGE_DISPLAY_NAMES lookup failed: %s", e)
        languages = getattr(Config, 'LANGUAGES', ['en'])
        # ISO codes only
        language_model_keys = {code: (None if code == 'en' else code) for code in (languages or [])}
        language_display_names = getattr(Config, 'LANGUAGE_DISPLAY_NAMES', {}) or getattr(Config, 'ALL_LANGUAGES_DISPLAY_NAMES', {}) or {}

    return render_template(
        'admin/translations/edit_translation.html',
        form=form,
        languages=languages,
        language_model_keys=language_model_keys,
        language_display_names=language_display_names,
        is_add=True
    )

@bp.route("/translations/edit", methods=["GET", "POST"])
@permission_required("admin.translations.manage")
def edit_translation():
    """Edit translations for all languages"""
    is_ajax = is_json_request()

    # --- JSON POST path (WAF-safe: payload travels in JSON body) ---
    if request.method == "POST" and request.is_json:
        data = get_request_data()
        msgid = (data.get('msgid') or '').strip()
        if not msgid:
            return json_bad_request(_("msgid is required"))

        from config import Config
        languages = current_app.config.get('SUPPORTED_LANGUAGES', Config.LANGUAGES)
        updated_count = 0

        try:
            import polib  # type: ignore
        except ImportError:
            current_app.logger.warning("polib not available - translation file updates will be skipped")
            polib = None

        if polib is not None:
            for lang in languages:
                msgstr = data.get(f'msgstr_{lang}')
                if msgstr is None:
                    continue
                po_file_path = os.path.join(current_app.root_path, 'translations', lang, 'LC_MESSAGES', 'messages.po')
                if os.path.exists(po_file_path):
                    try:
                        po = polib.pofile(po_file_path)
                        entry = po.find(msgid)
                        if entry is None:
                            if str(msgstr).strip():
                                po.append(polib.POEntry(msgid=msgid, msgstr=msgstr))
                                updated_count += 1
                        else:
                            entry.msgstr = msgstr
                            updated_count += 1
                        po.save(po_file_path)
                    except Exception as _e:
                        current_app.logger.error(f"Failed to update translation for {lang}: {_e}")

        if updated_count > 0:
            return json_ok(
                message=_('Translation updated successfully for %(count)d language(s)', count=updated_count),
                updated_count=updated_count,
            )
        return json_ok(
            message=_('No translations were updated'),
            updated_count=0,
        )

    # --- Legacy WTForms POST path (non-JSON form submissions) ---
    form = TranslationForm()

    if form.validate_on_submit():
        msgid = form.msgid.data

        from config import Config
        languages = current_app.config.get('SUPPORTED_LANGUAGES', Config.LANGUAGES)
        updated_count = 0

        try:
            import polib  # type: ignore
        except ImportError:
            current_app.logger.warning("polib not available - translation file updates will be skipped")
            polib = None

        for lang in languages:
            msgstr_field = getattr(form, f'msgstr_{lang}', None)
            msgstr = (msgstr_field.data if msgstr_field else None)
            if msgstr is None:
                continue
            po_file_path = os.path.join(current_app.root_path, 'translations', lang, 'LC_MESSAGES', 'messages.po')

            if polib is None:
                current_app.logger.warning(f"polib not available - skipping translation file update for {lang}")
                continue

            if os.path.exists(po_file_path) and polib:
                try:
                    po = polib.pofile(po_file_path)
                    entry = po.find(msgid)
                    if entry is None:
                        if str(msgstr).strip():
                            po.append(polib.POEntry(msgid=msgid, msgstr=msgstr))
                            updated_count += 1
                    else:
                        entry.msgstr = msgstr
                        updated_count += 1
                    po.save(po_file_path)
                except Exception as _e:
                    current_app.logger.error(f"Failed to update translation for {lang}: {_e}")

        if is_ajax:
            if updated_count > 0:
                return json_ok(
                    message=_('Translation updated successfully for %(count)d language(s)', count=updated_count),
                    updated_count=updated_count,
                )
            else:
                return json_ok(
                    message=_('No translations were updated'),
                    updated_count=0,
                )
        else:
            if updated_count > 0:
                flash(_('Translation updated successfully for %(count)d language(s)', count=updated_count), 'success')
            else:
                flash(_('No translations were updated'), 'warning')
            return redirect(url_for('utilities.manage_translations'))

    # --- GET path: support base64-encoded msgid to avoid WAF triggers ---
    import base64 as _b64
    msgid_b64 = request.args.get('msgid_b64')
    if msgid_b64:
        try:
            msgid = _b64.b64decode(msgid_b64).decode('utf-8')
        except Exception:
            msgid = None
    else:
        msgid = request.args.get('msgid')

    from config import Config
    languages = current_app.config.get('SUPPORTED_LANGUAGES', Config.LANGUAGES)
    translations = {}

    if msgid:
        form.msgid.data = msgid

        try:
            import polib  # type: ignore
        except ImportError:
            current_app.logger.warning("polib not available - cannot pre-populate translations")
            polib = None

        if polib is not None:
            for lang in languages:
                po_file_path = os.path.join(current_app.root_path, 'translations', lang, 'LC_MESSAGES', 'messages.po')
                if os.path.exists(po_file_path) and polib:
                    try:
                        po = polib.pofile(po_file_path)
                        entry = po.find(msgid)
                        if entry is not None:
                            translations[lang] = entry.msgstr or ''
                            msgstr_field = getattr(form, f'msgstr_{lang}', None)
                            if msgstr_field is not None:
                                msgstr_field.data = entry.msgstr or ''
                    except Exception as _e:
                        current_app.logger.error(f"Failed to read translation for {lang}: {_e}")

    if is_ajax:
        try:
            from config import Config
            language_display_names = Config.LANGUAGE_DISPLAY_NAMES
        except Exception as e:
            logger.debug("Config.LANGUAGE_DISPLAY_NAMES lookup failed: %s", e)
            language_display_names = getattr(Config, 'LANGUAGE_DISPLAY_NAMES', {}) or getattr(Config, 'ALL_LANGUAGES_DISPLAY_NAMES', {}) or {}

        return json_ok(
            msgid=msgid or '',
            translations=translations,
            languages=languages,
            language_display_names=language_display_names,
        )

    return redirect(url_for('utilities.manage_translations'))

@bp.route("/translations/compile", methods=["POST"])
@permission_required("admin.translations.manage")
def compile_translations():
    """Compile translations and optionally restart the application"""
    try:
        # Run the compilation script
        import subprocess
        import os
        # Get the backoffice root directory (parent of app/)
        backoffice_root = os.path.abspath(os.path.join(current_app.root_path, '..'))
        scripts_dir = os.path.join(backoffice_root, 'scripts')
        result = subprocess.run(['python', 'compile_translations.py'],
                              capture_output=True, text=True, cwd=scripts_dir)

        if result.returncode == 0:
            # Force reload of translations
            from flask_babel import refresh
            refresh()

            # Check if restart was requested
            restart_requested = request.form.get('restart') == '1'

            if restart_requested:
                # Attempt to restart the application
                try:
                    import os
                    import time

                    # In debug mode with Werkzeug reloader, touch a watched Python module to trigger reload
                    if current_app.debug and os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
                        target_files = [
                            os.path.join(current_app.root_path, '__init__.py'),
                            os.path.join(os.path.dirname(current_app.root_path), 'config', 'config.py'),
                        ]
                        touched = False
                        for f in target_files:
                            if os.path.exists(f):
                                now = time.time()
                                os.utime(f, (now, now))
                                touched = True
                        if touched:
                            flash(_('Translations compiled successfully! Application restart triggered.'), 'success')
                        else:
                            flash(_('Translations compiled successfully! Please restart the application manually to apply changes.'), 'warning')
                    else:
                        # For production, we can't easily restart the application
                        # Instead, we'll try to reload the application context
                        try:
                            # Force reload of all modules by touching the main app file
                            app_file = os.path.join(current_app.root_path, '__init__.py')
                            if os.path.exists(app_file):
                                now = time.time()
                                os.utime(app_file, (now, now))
                            flash(_('Translations compiled successfully! Application context refreshed. Please restart the application via your process manager to apply changes across all workers.'), 'warning')
                        except Exception as e:
                            logger.debug("touch/restart hint failed: %s", e)
                            flash(_('Translations compiled successfully! Please restart the application via your process manager to apply changes.'), 'warning')
                except Exception as restart_error:
                    flash(_('Translations compiled successfully! However, automatic restart failed. Please restart the application manually.'), 'warning')
            else:
                flash(_('Translations compiled and reloaded successfully!'), 'success')
        else:
            flash(_('Error compiling translations: %(error)s', error=result.stderr), 'danger')

    except Exception as e:
        flash(_("An error occurred. Please try again."), "danger")

    return redirect(url_for('utilities.manage_translations'))

@bp.route("/translations/reload", methods=["POST"])
@permission_required("admin.translations.manage")
def reload_translations():
    """Manually reload translations without recompiling"""
    try:
        from flask_babel import refresh
        refresh()
        flash(_('Translations reloaded successfully!'), 'success')
    except Exception as e:
        flash(_("An error occurred. Please try again."), "danger")

    return redirect(url_for('utilities.manage_translations'))

@bp.route("/translations/extract-update", methods=["POST"])
@permission_required("admin.translations.manage")
def extract_update_translations():
    """Extract translatable strings and update PO files"""
    try:
        import subprocess
        import os
        import sys

        # Get the backoffice root directory (parent of app/)
        backoffice_root = os.path.abspath(os.path.join(current_app.root_path, '..'))
        scripts_dir = os.path.join(backoffice_root, 'scripts')
        script_path = os.path.join(scripts_dir, 'extract_update_translations.py')

        if not os.path.exists(script_path):
            flash(_('Extraction script not found. Please ensure scripts/extract_update_translations.py exists.'), 'danger')
            return redirect(url_for('utilities.manage_translations'))

        # Run the extraction script
        # Use sys.executable to ensure we use the same Python interpreter
        result = subprocess.run(
            [sys.executable, 'extract_update_translations.py'],
            capture_output=True,
            text=True,
            cwd=scripts_dir,
            timeout=300  # 5 minute timeout
        )

        if result.returncode == 0:
            # Parse output to get useful information
            output_lines = result.stdout.split('\n')
            obsolete_info = []
            for line in output_lines:
                if 'obsolete' in line.lower() and '/' in line:
                    obsolete_info.append(line.strip())

            # Force reload of translations
            from flask_babel import refresh
            refresh()

            # Build success message
            success_msg = _('Translations extracted and updated successfully!')
            if obsolete_info:
                success_msg += ' ' + ' '.join(obsolete_info[:3])  # Show first 3 locales
                if len(obsolete_info) > 3:
                    success_msg += _(' (and %(count)d more)', count=len(obsolete_info) - 3)

            flash(success_msg, 'success')

            # Show any warnings from stderr (non-fatal)
            if result.stderr and 'warning' in result.stderr.lower():
                flash(_('Note: %(note)s', note=result.stderr[:200]), 'info')
        else:
            error_msg = result.stderr or result.stdout or _('Unknown error')
            flash(_('Error extracting/updating translations: %(error)s', error=error_msg[:500]), 'danger')
            current_app.logger.error(f"Extraction script failed: {result.stderr}")

    except subprocess.TimeoutExpired:
        flash(_('Extraction timed out after 5 minutes. The process may still be running.'), 'warning')
    except Exception as e:
        flash(_("An error occurred. Please try again."), "danger")
        current_app.logger.error(f"Extraction error: {e}", exc_info=True)

    return redirect(url_for('utilities.manage_translations'))

# === Indicator Suggestions Management ===
@bp.route("/indicator_suggestions", methods=["GET"])
@permission_required('admin.indicator_bank.suggestions.review')
def manage_indicator_suggestions():
    """Manage indicator suggestions from users"""
    from app.utils.api_pagination import validate_pagination_params
    page, per_page = validate_pagination_params(request.args, default_per_page=20, max_per_page=100)
    status_filter = request.args.get('status', '')
    suggestion_type_filter = request.args.get('suggestion_type', '')

    query = IndicatorSuggestion.query

    if status_filter:
        query = query.filter(IndicatorSuggestion.status == status_filter)

    pagination = query.order_by(
        IndicatorSuggestion.submitted_at.desc()
    ).paginate(page=page, per_page=per_page, error_out=False)

    return render_template(
        "admin/indicator_suggestions.html",
        suggestions=pagination.items,
        pagination=pagination,
        status_filter=status_filter,
        suggestion_type_filter=suggestion_type_filter,
        title="Manage Indicator Suggestions",
    )

@bp.route("/indicator_suggestions/view/<int:suggestion_id>", methods=["GET"])
@permission_required('admin.indicator_bank.suggestions.review')
def view_indicator_suggestion(suggestion_id):
    """View individual indicator suggestion"""
    suggestion = IndicatorSuggestion.query.get_or_404(suggestion_id)
    return render_template("admin/view_indicator_suggestion.html",
                         suggestion=suggestion,
                         title=f"View Suggestion: {suggestion.indicator_name}")

@bp.route("/indicator_suggestions/update_status/<int:suggestion_id>", methods=["POST"])
@permission_required('admin.indicator_bank.suggestions.review')
def update_indicator_suggestion_status(suggestion_id):
    """Update indicator suggestion status"""
    suggestion = IndicatorSuggestion.query.get_or_404(suggestion_id)
    new_status = request.form.get('status')
    admin_notes = request.form.get('admin_notes', '').strip()

    try:
        if new_status in ['Pending', 'approved', 'rejected', 'implemented']:
            suggestion.status = new_status
            suggestion.admin_notes = admin_notes
            suggestion.reviewed_by = current_user
            suggestion.reviewed_at = utcnow()

            # If approved, optionally create the indicator automatically
            if new_status == 'approved':
                _create_indicator_from_suggestion(suggestion)

            db.session.flush()
            flash(_("Suggestion status updated to %(status)s.", status=new_status), "success")
        else:
            flash(_("Invalid status provided."), "danger")

    except Exception as e:
        request_transaction_rollback()
        current_app.logger.error(f"Error updating suggestion status: {e}", exc_info=True)
        flash(_("Error updating suggestion status."), "danger")

    return redirect(url_for("utilities.view_indicator_suggestion", suggestion_id=suggestion_id))

@bp.route("/indicator_suggestions/delete/<int:suggestion_id>", methods=["POST"])
@permission_required('admin.indicator_bank.suggestions.review')
def delete_indicator_suggestion(suggestion_id):
    """Delete indicator suggestion"""
    suggestion = IndicatorSuggestion.query.get_or_404(suggestion_id)

    try:
        db.session.delete(suggestion)
        db.session.flush()
        flash(_("Suggestion deleted successfully."), "success")

    except Exception as e:
        request_transaction_rollback()
        current_app.logger.error(f"Error deleting suggestion: {e}", exc_info=True)
        flash(_("Error deleting suggestion."), "danger")

    return redirect(url_for("utilities.manage_indicator_suggestions"))

# === API Routes ===
@bp.route("/api/refresh_csrf_token", methods=["POST"])
@admin_required
def refresh_csrf_token():
    """Refresh CSRF token for AJAX requests"""
    try:
        token = csrf.generate_csrf()
        return json_ok(csrf_token=token, status='success')
    except Exception as e:
        return handle_json_view_exception(e, 'Error refreshing CSRF token', status_code=500)

# Compatibility alias: support GET and dash-separated path used by existing JS
@bp.route("/api/refresh-csrf-token", methods=["GET"])
@admin_required
def refresh_csrf_token_get():
    try:
        token = csrf.generate_csrf()
        return json_ok(csrf_token=token, status='success')
    except Exception as e:
        return handle_json_view_exception(e, 'Error refreshing CSRF token', status_code=500)

# Auto-translate functionality is now handled by the centralized modal system

@bp.route("/api/translation_services", methods=["GET"])
@permission_required_any(
    'admin.templates.edit',
    'admin.templates.create',
    'admin.indicator_bank.create',
    'admin.indicator_bank.edit',
    'admin.resources.manage',
    'admin.translations.manage',
    'admin.organization.manage',
    'admin.settings.manage',
)
def api_translation_services():
    """Get available translation services status"""
    try:
        from app.utils.auto_translator import get_auto_translator

        auto_translator = get_auto_translator()
        available_services = auto_translator.get_available_services()
        default_service = auto_translator.get_default_service()
        service_status = auto_translator.check_service_status()

        # Map service names to display names
        service_display_names = {
            'ifrc': 'Hosted translation API',
            'libre': 'LibreTranslate AI',
            'google': 'Google Translate'
        }

        services = []
        for service in available_services:
            is_available = service_status.get(service, False)
            services.append({
                'value': service,
                'label': service_display_names.get(service, service.title()),
                'is_default': service == default_service,
                'is_available': is_available
            })

        return json_ok(services=services, default_service=default_service)

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)

# Bulk translation functionality is implemented above

# === Helper Functions ===

def _process_indicator_import(file_path):
    """Process Excel file for indicator import"""
    try:
        current_app.logger.info(f"Loading workbook from: {file_path}")
        wb = load_workbook(file_path, read_only=True, data_only=True)

        result = {
            'success': True,
            'imported': 0,  # indicators imported from main sheet OR DB_Indicators
            'updated': 0,   # indicators updated from main sheet OR DB_Indicators
            'sectors_imported': 0,
            'sectors_updated': 0,
            'subsectors_imported': 0,
            'subsectors_updated': 0,
            'common_words_imported': 0,
            'common_words_updated': 0,
            'errors': [],
            'message': ''
        }

        # --- Helpers ---
        def _norm_header(v):
            return str(v).strip().lower() if v is not None else ""

        def _to_int(v):
            try:
                if v is None or v == "":
                    return None
                if isinstance(v, bool):
                    return int(v)
                if isinstance(v, (int, float)):
                    return int(v)
                s = str(v).strip()
                if not s:
                    return None
                return int(float(s))
            except Exception as e:
                logger.debug("int parse failed for %r: %s", v, e)
                return None

        def _to_bool(v):
            if v is None or v == "":
                return None
            if isinstance(v, bool):
                return v
            if isinstance(v, (int, float)):
                return bool(int(v))
            s = str(v).strip().lower()
            if s in ("true", "t", "yes", "y", "1"):
                return True
            if s in ("false", "f", "no", "n", "0"):
                return False
            return None

        def _to_json_dict(v):
            if v is None or v == "":
                return None
            if isinstance(v, dict):
                return v
            try:
                parsed = json.loads(str(v))
                return parsed if isinstance(parsed, dict) else None
            except Exception as e:
                logger.debug("json parse failed for %r: %s", v, e)
                return None

        def _sheet_rows_as_dicts(ws):
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            raw_headers = list(rows[0])
            headers = [_norm_header(h) for h in raw_headers]
            out = []
            for r in rows[1:]:
                if not r:
                    continue
                # skip fully empty rows
                if all((c is None or str(c).strip() == "") for c in r):
                    continue
                d = {}
                for idx, key in enumerate(headers):
                    if not key:
                        continue
                    if idx < len(r):
                        d[key] = r[idx]
                out.append(d)
            return out

        def _build_levels_json(primary_id, secondary_id, tertiary_id):
            data = {}
            if primary_id is not None:
                data["primary"] = primary_id
            if secondary_id is not None:
                data["secondary"] = secondary_id
            if tertiary_id is not None:
                data["tertiary"] = tertiary_id
            return data or None

        sheetnames = set(wb.sheetnames or [])

        # If file contains hidden DB_* sheets (from export), use full relational import; otherwise use main sheet
        if "DB_Indicators" in sheetnames or "DB_Sectors_SubSectors" in sheetnames or "DB_CommonWords" in sheetnames:
            # 1) Sectors/SubSectors
            if "DB_Sectors_SubSectors" in sheetnames:
                from app.models import Sector, SubSector
                ws_ss = wb["DB_Sectors_SubSectors"]
                ss_rows = _sheet_rows_as_dicts(ws_ss)

                for idx, r in enumerate(ss_rows, 2):
                    try:
                        record_type = (r.get("record_type") or "").strip().lower()
                        rid = _to_int(r.get("id"))
                        name = (r.get("name") or "").strip()
                        description = (r.get("description") or None)
                        sector_id = _to_int(r.get("sector_id"))
                        display_order = _to_int(r.get("display_order"))
                        is_active = _to_bool(r.get("is_active"))
                        icon_class = (r.get("icon_class") or None)
                        logo_filename = (r.get("logo_filename") or None)
                        logo_path = (r.get("logo_path") or None)
                        name_translations = _to_json_dict(r.get("name_translations_json")) or {}

                        if record_type == "sector":
                            if not rid and not name:
                                continue
                            obj = Sector.query.get(rid) if rid else Sector.query.filter_by(name=name).first()
                            is_new = obj is None
                            if obj is None:
                                obj = Sector()
                                if rid:
                                    obj.id = rid
                                db.session.add(obj)
                            if name:
                                obj.name = name
                            obj.description = description
                            if display_order is not None:
                                obj.display_order = display_order
                            if is_active is not None:
                                obj.is_active = is_active
                            obj.icon_class = icon_class
                            obj.logo_filename = logo_filename
                            obj.logo_path = logo_path
                            obj.name_translations = name_translations or {}
                            if is_new:
                                result["sectors_imported"] += 1
                            else:
                                result["sectors_updated"] += 1

                        elif record_type == "subsector":
                            if not rid and not name:
                                continue
                            obj = SubSector.query.get(rid) if rid else SubSector.query.filter_by(name=name).first()
                            is_new = obj is None
                            if obj is None:
                                obj = SubSector()
                                if rid:
                                    obj.id = rid
                                db.session.add(obj)
                            if name:
                                obj.name = name
                            obj.description = description
                            if sector_id is not None:
                                obj.sector_id = sector_id
                            if display_order is not None:
                                obj.display_order = display_order
                            if is_active is not None:
                                obj.is_active = is_active
                            obj.icon_class = icon_class
                            obj.logo_filename = logo_filename
                            obj.logo_path = logo_path
                            obj.name_translations = name_translations or {}
                            if is_new:
                                result["subsectors_imported"] += 1
                            else:
                                result["subsectors_updated"] += 1
                    except Exception as e:
                        current_app.logger.error(f"Error processing DB_Sectors_SubSectors row {idx}: {e}", exc_info=True)
                        result["errors"].append(f"DB_Sectors_SubSectors row {idx}: error.")

            # 2) Common words
            if "DB_CommonWords" in sheetnames:
                from app.models import CommonWord
                ws_cw = wb["DB_CommonWords"]
                cw_rows = _sheet_rows_as_dicts(ws_cw)
                for idx, r in enumerate(cw_rows, 2):
                    try:
                        rid = _to_int(r.get("id"))
                        term = (r.get("term") or "").strip()
                        meaning = (r.get("meaning") or "").strip()
                        is_active = _to_bool(r.get("is_active"))
                        meaning_translations = _to_json_dict(r.get("meaning_translations_json")) or {}

                        if not term and not rid:
                            continue

                        obj = CommonWord.query.get(rid) if rid else CommonWord.query.filter_by(term=term).first()
                        is_new = obj is None
                        if obj is None:
                            obj = CommonWord()
                            if rid:
                                obj.id = rid
                            db.session.add(obj)
                        if term:
                            obj.term = term
                        if meaning:
                            obj.meaning = meaning
                        if is_active is not None:
                            obj.is_active = is_active
                        obj.meaning_translations = meaning_translations or {}

                        if is_new:
                            result["common_words_imported"] += 1
                        else:
                            result["common_words_updated"] += 1
                    except Exception as e:
                        current_app.logger.error(f"Error processing DB_CommonWords row {idx}: {e}", exc_info=True)
                        result["errors"].append(f"DB_CommonWords row {idx}: Validation error.")

            # 3) Indicators (from hidden DB_Indicators sheet when present)
            if "DB_Indicators" in sheetnames:
                ws_ind = wb["DB_Indicators"]
                ind_rows = _sheet_rows_as_dicts(ws_ind)
                for idx, r in enumerate(ind_rows, 2):
                    try:
                        rid = _to_int(r.get("id"))
                        name = (r.get("name") or "").strip()
                        if not name and not rid:
                            continue

                        definition = (r.get("definition") or "").strip()
                        indicator_type = (r.get("type") or "").strip() or "numeric"
                        unit = (r.get("unit") or "").strip()
                        fdrs_kpi_code = (r.get("fdrs_kpi_code") or "").strip() or None
                        emergency = _to_bool(r.get("emergency"))
                        archived = _to_bool(r.get("archived"))
                        comments = (r.get("comments") or None)
                        programs = (r.get("related_programs") or "")
                        programs = str(programs).strip() if programs is not None else ""

                        s_primary = _to_int(r.get("sector_primary_id"))
                        s_secondary = _to_int(r.get("sector_secondary_id"))
                        s_tertiary = _to_int(r.get("sector_tertiary_id"))
                        ss_primary = _to_int(r.get("subsector_primary_id"))
                        ss_secondary = _to_int(r.get("subsector_secondary_id"))
                        ss_tertiary = _to_int(r.get("subsector_tertiary_id"))

                        name_translations = _to_json_dict(r.get("name_translations_json")) or {}
                        definition_translations = _to_json_dict(r.get("definition_translations_json")) or {}

                        existing = IndicatorBank.query.get(rid) if rid else IndicatorBank.query.filter_by(name=name).first()
                        is_new = existing is None

                        if existing is None:
                            existing = IndicatorBank(
                                name=name,
                                definition=definition,
                                type=indicator_type,
                                unit=unit,
                                fdrs_kpi_code=fdrs_kpi_code,
                                emergency=bool(emergency) if emergency is not None else False,
                                related_programs=programs,
                            )
                            if rid:
                                existing.id = rid
                            db.session.add(existing)
                            db.session.flush()
                        else:
                            # update base fields
                            if name:
                                existing.name = name
                            existing.definition = definition
                            existing.type = indicator_type
                            existing.unit = unit
                            existing.fdrs_kpi_code = fdrs_kpi_code
                            if emergency is not None:
                                existing.emergency = emergency
                            if archived is not None:
                                existing.archived = archived
                            existing.comments = comments
                            existing.related_programs = programs

                        # update relational JSON fields + translations
                        existing.sector = _build_levels_json(s_primary, s_secondary, s_tertiary)
                        existing.sub_sector = _build_levels_json(ss_primary, ss_secondary, ss_tertiary)
                        existing.name_translations = name_translations or {}
                        existing.definition_translations = definition_translations or {}

                        # Create history record (minimal but consistent)
                        history = IndicatorBankHistory(
                            indicator_bank_id=existing.id,
                            user_id=current_user.id,
                            name=existing.name,
                            type=existing.type,
                            unit=existing.unit,
                            fdrs_kpi_code=existing.fdrs_kpi_code,
                            definition=existing.definition,
                            name_translations=existing.name_translations,
                            definition_translations=existing.definition_translations,
                            archived=existing.archived,
                            comments=existing.comments,
                            emergency=existing.emergency,
                            related_programs=existing.related_programs,
                            sector=existing.sector,
                            sub_sector=existing.sub_sector,
                            change_type='CREATED' if is_new else 'UPDATED',
                            change_description=(
                                f'Indicator "{existing.name}" created via DB import by {current_user.name or current_user.email}'
                                if is_new
                                else f'Indicator "{existing.name}" updated via DB import by {current_user.name or current_user.email}'
                            ),
                        )
                        db.session.add(history)

                        if is_new:
                            result["imported"] += 1
                        else:
                            result["updated"] += 1
                    except Exception as e:
                        current_app.logger.error(f"Error processing DB_Indicators row {idx}: {e}", exc_info=True)
                        result["errors"].append(f"DB_Indicators row {idx}: error.")

            db.session.flush()
            return result

        # --- Main sheet import (primary): use active sheet ---
        ws = wb.active

        # Build header -> column index (0-based) for optional columns
        headers_raw = [cell.value for cell in ws[1]]
        headers = [str(h).strip() if h is not None else "" for h in headers_raw]
        header_to_col = {}
        for idx, h in enumerate(headers):
            key = _norm_header(h) if h else ""
            if key and key not in header_to_col:
                header_to_col[key] = idx
        current_app.logger.info(f"Excel headers: {headers}")

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        current_app.logger.info(f"Total rows to process: {len(rows)}")

        def _val(row, idx):
            if row is None or idx is None or idx < 0 or idx >= len(row):
                return None
            return row[idx]

        def _resolve_sector_id(name_val):
            if name_val is None or str(name_val).strip() == "":
                return None
            s = Sector.query.filter_by(name=str(name_val).strip()).first()
            return s.id if s else None

        def _resolve_subsector_id(name_val):
            if name_val is None or str(name_val).strip() == "":
                return None
            ss = SubSector.query.filter_by(name=str(name_val).strip()).first()
            return ss.id if ss else None

        for row_num, row in enumerate(rows, 2):
            try:
                if row is None or len(row) < 7:
                    current_app.logger.warning(f"Skipping row {row_num}: row is None or has insufficient columns")
                    continue

                if _val(row, 1) is None or str(_val(row, 1)).strip() == "":
                    continue

                name = (str(_val(row, header_to_col.get("name", 1)) or "")).strip()
                definition = str(_val(row, header_to_col.get("definition", 2)) or "").strip() if _val(row, header_to_col.get("definition", 2)) is not None else ''
                indicator_type = str(_val(row, header_to_col.get("type", 3)) or "").strip() if _val(row, header_to_col.get("type", 3)) is not None else 'numeric'
                unit = str(_val(row, header_to_col.get("unit", 4)) or "").strip() if _val(row, header_to_col.get("unit", 4)) is not None else ''
                fdrs_kpi_code_col = header_to_col.get("fdrs kpi code")
                fdrs_kpi_code = (str(_val(row, fdrs_kpi_code_col) or "").strip() or None) if fdrs_kpi_code_col is not None else None
                _emergency = _to_bool(_val(row, header_to_col.get("emergency", 5))) if _val(row, header_to_col.get("emergency", 5)) is not None else False
                emergency = _emergency if _emergency is not None else False
                programs = str(_val(row, header_to_col.get("related programs", 6)) or "").strip() if _val(row, header_to_col.get("related programs", 6)) is not None else ''

                # Optional: localized names/definitions from "Name (code)" and "Definition (code)" columns
                name_translations = {}
                definition_translations = {}
                for key, col in header_to_col.items():
                    if key.startswith("name (") and key.endswith(")"):
                        code = key[6:-1].strip().lower()
                        if code and len(code) <= 6:
                            val = _val(row, col)
                            if val is not None and str(val).strip():
                                name_translations[code] = str(val).strip()
                    elif key.startswith("definition (") and key.endswith(")"):
                        code = key[12:-1].strip().lower()
                        if code and len(code) <= 6:
                            val = _val(row, col)
                            if val is not None and str(val).strip():
                                definition_translations[code] = str(val).strip()

                # Optional: sector/subsector names from "Sector Primary", "SubSector Primary", etc.
                s_primary = _resolve_sector_id(_val(row, header_to_col.get("sector primary")))
                s_secondary = _resolve_sector_id(_val(row, header_to_col.get("sector secondary")))
                s_tertiary = _resolve_sector_id(_val(row, header_to_col.get("sector tertiary")))
                ss_primary = _resolve_subsector_id(_val(row, header_to_col.get("subsector primary")))
                ss_secondary = _resolve_subsector_id(_val(row, header_to_col.get("subsector secondary")))
                ss_tertiary = _resolve_subsector_id(_val(row, header_to_col.get("subsector tertiary")))
                sector_json = _build_levels_json(s_primary, s_secondary, s_tertiary)
                sub_sector_json = _build_levels_json(ss_primary, ss_secondary, ss_tertiary)

                existing = IndicatorBank.query.filter_by(name=name).first()

                if existing:
                    changes = []
                    if existing.definition != definition:
                        old_def = existing.definition or ''
                        new_def = definition or ''
                        changes.append(f"Definition: Changed from '{old_def[:50]}{'...' if len(old_def) > 50 else ''}' to '{new_def[:50]}{'...' if len(new_def) > 50 else ''}'")
                    if existing.type != indicator_type:
                        changes.append(f"Type: Changed from '{existing.type or ''}' to '{indicator_type}'")
                    if existing.unit != unit:
                        changes.append(f"Unit: Changed from '{existing.unit or ''}' to '{unit}'")
                    if getattr(existing, 'fdrs_kpi_code', None) != fdrs_kpi_code:
                        changes.append(f"FDRS KPI Code: Changed from '{existing.fdrs_kpi_code or ''}' to '{fdrs_kpi_code or ''}'")
                    if existing.emergency != emergency:
                        changes.append(f"Emergency: Changed from '{existing.emergency}' to '{emergency}'")
                    if existing.related_programs != programs:
                        changes.append(f"Related Programs: Changed from '{existing.related_programs or ''}' to '{programs}'")

                    existing.definition = definition
                    existing.type = indicator_type
                    existing.unit = unit
                    existing.fdrs_kpi_code = fdrs_kpi_code
                    existing.emergency = emergency
                    existing.related_programs = programs
                    if name_translations:
                        existing.name_translations = {**(existing.name_translations or {}), **name_translations}
                    if definition_translations:
                        existing.definition_translations = {**(existing.definition_translations or {}), **definition_translations}
                    if sector_json is not None:
                        existing.sector = sector_json
                    if sub_sector_json is not None:
                        existing.sub_sector = sub_sector_json

                    change_description = "; ".join(changes) if changes else "Indicator updated via import (no specific changes detected)"

                    history = IndicatorBankHistory(
                        indicator_bank_id=existing.id,
                        user_id=current_user.id,
                        name=existing.name,
                        type=existing.type,
                        unit=existing.unit,
                        fdrs_kpi_code=existing.fdrs_kpi_code,
                        definition=existing.definition,
                        name_translations=existing.name_translations,
                        definition_translations=existing.definition_translations,
                        archived=existing.archived,
                        comments=existing.comments,
                        emergency=existing.emergency,
                        related_programs=existing.related_programs,
                        sector=existing.sector,
                        sub_sector=existing.sub_sector,
                        change_type='UPDATED',
                        change_description=change_description
                    )
                    db.session.add(history)
                    result['updated'] += 1
                else:
                    new_indicator = IndicatorBank(
                        name=name,
                        definition=definition,
                        type=indicator_type,
                        unit=unit,
                        fdrs_kpi_code=fdrs_kpi_code,
                        emergency=emergency,
                        related_programs=programs
                    )
                    if name_translations:
                        new_indicator.name_translations = name_translations
                    if definition_translations:
                        new_indicator.definition_translations = definition_translations
                    if sector_json is not None:
                        new_indicator.sector = sector_json
                    if sub_sector_json is not None:
                        new_indicator.sub_sector = sub_sector_json
                    db.session.add(new_indicator)
                    db.session.flush()

                    history = IndicatorBankHistory(
                        indicator_bank_id=new_indicator.id,
                        user_id=current_user.id,
                        name=new_indicator.name,
                        type=new_indicator.type,
                        unit=new_indicator.unit,
                        fdrs_kpi_code=new_indicator.fdrs_kpi_code,
                        definition=new_indicator.definition,
                        name_translations=new_indicator.name_translations,
                        definition_translations=new_indicator.definition_translations,
                        archived=new_indicator.archived,
                        comments=new_indicator.comments,
                        emergency=new_indicator.emergency,
                        related_programs=new_indicator.related_programs,
                        sector=new_indicator.sector,
                        sub_sector=new_indicator.sub_sector,
                        change_type='CREATED',
                        change_description=f'Indicator "{new_indicator.name}" created via import by {current_user.name or current_user.email}'
                    )
                    db.session.add(history)
                    result['imported'] += 1

            except Exception as e:
                current_app.logger.error(f"Error processing row {row_num}: {str(e)}", exc_info=True)
                current_app.logger.error(f"Row data: {row}")
                result['errors'].append(f"Row {row_num}: error.")
                continue

        db.session.flush()

    except Exception as e:
        result['success'] = False
        result['message'] = GENERIC_ERROR_MESSAGE
        request_transaction_rollback()

    return result

def _create_indicator_from_suggestion(suggestion):
    """Create an indicator from an approved suggestion"""
    try:
        new_indicator = IndicatorBank(
            name=suggestion.suggested_name,
            definition=suggestion.suggested_definition or '',
            type=suggestion.suggested_type or 'numeric',
            unit=suggestion.suggested_unit or '',
            fdrs_kpi_code=None,
            emergency=suggestion.suggested_emergency or False,
            related_programs=suggestion.suggested_programs or ''
        )

        db.session.add(new_indicator)
        db.session.flush()  # Get the ID

        # Create history record
        history = IndicatorBankHistory(
            indicator_bank_id=new_indicator.id,
            user_id=current_user.id,
            name=new_indicator.name,
            type=new_indicator.type,
            unit=new_indicator.unit,
            fdrs_kpi_code=new_indicator.fdrs_kpi_code,
            definition=new_indicator.definition,
            name_translations=new_indicator.name_translations,
            definition_translations=new_indicator.definition_translations,
            archived=new_indicator.archived,
            comments=new_indicator.comments,
            emergency=new_indicator.emergency,
            related_programs=new_indicator.related_programs,
            sector=new_indicator.sector,
            sub_sector=new_indicator.sub_sector,
            change_type='CREATED',
            change_description=f'Indicator "{new_indicator.name}" created from suggestion by {current_user.name or current_user.email}'
        )
        db.session.add(history)

        # Update suggestion status
        suggestion.status = 'implemented'
        suggestion.indicator_id = new_indicator.id

        return new_indicator

    except Exception as e:
        current_app.logger.error(f"Error creating indicator from suggestion: {e}", exc_info=True)
        return None

# Legacy bulk translation functions removed - now handled by centralized modal system

def _extract_page_name(source_path):
    """Return a concise page name from a PO source reference.

    Examples of inputs:
    - "app/templates/admin/api_management.html:100"
    - "app/templates/admin/api_management.html:100 app/templates/admin/api_management.html:110"
    - "app/routes/admin/utilities.py:716"

    Output should be only the base name without extension or line numbers, e.g.:
    - "api_management"
    - "utilities"
    """
    if not source_path:
        return "Unknown"

    # Use the first reference if multiple are provided on the same line
    first_ref = str(source_path).strip().split()[0]

    # Normalize path separators and take the last segment
    last_segment = first_ref.replace('\\', '/').split('/')[-1]

    # Strip any ":<line>" suffix if present
    file_with_no_line = last_segment.split(':', 1)[0]

    # Drop extension if present
    if '.' in file_with_no_line:
        base_name = file_with_no_line.rsplit('.', 1)[0]
    else:
        base_name = file_with_no_line

    return base_name or "Unknown"

# ========================
# Translation API Endpoints
# ========================

@bp.route('/api/auto-translate', methods=['POST'])
@limiter.exempt
@permission_required_any(
    'admin.templates.edit',
    'admin.templates.create',
    'admin.indicator_bank.create',
    'admin.indicator_bank.edit',
    'admin.resources.manage',
    'admin.translations.manage',
    'admin.organization.manage',
    'admin.settings.manage',
)
def api_auto_translate():
    """API endpoint for automatic translation"""


    try:
        from app.utils.auto_translator import get_auto_translator, translate_form_item_auto, translate_section_name_auto, translate_question_option_auto, translate_page_name_auto, translate_template_name_auto
        from app.services.authorization_service import AuthorizationService

        data = get_json_safe()
        err = require_json_data(data)
        if err:
            return err

        # --- RBAC enforcement (context-aware) ---
        # IMPORTANT: Do not infer permission from "type" alone. Multiple pages reuse the same
        # translation types (e.g., indicator bank uses "form_item"), so callers must declare
        # an explicit permission context + permission code which we validate against allowlists.
        permission_context = (data.get('permission_context') or data.get('context') or '').strip().lower()
        permission_code = (data.get('permission_code') or data.get('permissionCode') or '').strip()

        allowed_context_permissions = {
            # Template creation/editing flows
            'templates': {'admin.templates.create', 'admin.templates.edit'},
            # Indicator bank bulk actions
            'indicator_bank': {'admin.indicator_bank.create', 'admin.indicator_bank.edit'},
            # Resources management editor
            'resources': {'admin.resources.manage'},
            # Translation file/admin translations editor
            'translations': {'admin.translations.manage'},
            # Settings admin area (org branding translations, etc.)
            'settings': {'admin.settings.manage'},
            # Organization management editor (edit_entity helper page)
            'organization': {'admin.organization.manage'},
        }

        # Backwards compatibility: if caller didn't pass context, infer safe defaults.
        # (Prefer explicit context; new UI callers should always send it.)
        if not permission_context:
            if (data.get('type') or '').strip() == 'translation':
                permission_context = 'translations'
            else:
                permission_context = 'templates'

        if not permission_code:
            # Default permission per context (most common edit/manage action)
            default_permission_by_context = {
                'templates': 'admin.templates.edit',
                'indicator_bank': 'admin.indicator_bank.edit',
                'resources': 'admin.resources.manage',
                'translations': 'admin.translations.manage',
                'settings': 'admin.settings.manage',
                'organization': 'admin.organization.manage',
            }
            permission_code = default_permission_by_context.get(permission_context, '')

        allowed_permissions = allowed_context_permissions.get(permission_context)
        if not allowed_permissions or permission_code not in allowed_permissions:
            return json_forbidden(f'Auto-translate not allowed for context "{permission_context}"')

        # System managers bypass granular permission checks
        if not AuthorizationService.is_system_manager(current_user):
            if not AuthorizationService.has_rbac_permission(current_user, permission_code):
                return json_forbidden('You do not have permission to use auto-translate here.')

        translation_type = data.get('type')  # 'form_item', 'section_name', 'question_option', 'page_name', or 'template_name'
        text = data.get('text', '').strip()
        definition = data.get('definition', '').strip()
        # Get target languages from config (normalize to ISO codes)
        try:
            from config.config import Config
            # Normalize incoming target languages to ISO language codes expected by translator
            incoming_target_languages = data.get('target_languages')
            if incoming_target_languages:
                normalized_target_languages = []
                for lang in incoming_target_languages:
                    # Accept only ISO codes (e.g., 'fr'); strip any region suffix (e.g., 'fr_FR' -> 'fr')
                    if isinstance(lang, str):
                        lang_norm = lang.split('_', 1)[0].strip().lower()
                        if lang_norm:
                            normalized_target_languages.append(lang_norm)
                target_languages = normalized_target_languages
            else:
                # Default to all translatable language codes, skipping English/source
                target_languages = [
                    lc for lc in (current_app.config.get('TRANSLATABLE_LANGUAGES') or getattr(Config, 'TRANSLATABLE_LANGUAGES', []) or [])
                    if lc != 'en'
                ]
        except (ImportError, NameError, AttributeError):
            # Fallback defaults (language codes) from runtime config
            configured = current_app.config.get('TRANSLATABLE_LANGUAGES') or []
            target_languages = [lc for lc in (configured or []) if lc != 'en']
        translation_service = data.get('translation_service', 'ifrc')  # Default hosted service (internal id "ifrc")
        service_name = translation_service  # Map to existing parameter for backward compatibility

        if not text:
            return json_bad_request('Text is required')

        auto_translator = get_auto_translator()

        # Check if any translation services are available
        available_services = auto_translator.get_available_services()
        if not available_services:
            return json_bad_request('No translation service available. Please configure translation API keys. Set IFRC_TRANSLATE_API_KEY, GOOGLE_TRANSLATE_API_KEY, or LIBRE_TRANSLATE_URL environment variables.')

        # If a specific service was requested but is not available, try to use a fallback
        if service_name and service_name not in available_services:
            logger.warning(f"Requested translation service '{service_name}' is not available. Available services: {available_services}. Using fallback.")
            service_name = None  # Let it use the default service

        if translation_type == 'form_item':
            # Translate form item (label and definition)
            result = translate_form_item_auto(
                label=text,
                definition=definition,
                target_languages=target_languages,
                service_name=service_name
            )

            if result and (result.get('label_translations') or result.get('definition_translations')):
                return json_ok(translations=result, service_used=auto_translator.get_default_service())
            else:
                return json_bad_request(f'Translation failed. No translations were generated. Available services: {", ".join(available_services)}. Please check your API keys and try again.')

        elif translation_type == 'document_field':
            # Translate document field (label only; descriptions are optional)
            result = translate_form_item_auto(
                label=text,
                definition=None,
                target_languages=target_languages,
                service_name=service_name
            )
            if result and result.get('label_translations'):
                return json_ok(translations=result, service_used=auto_translator.get_default_service())
            else:
                return json_bad_request(f'Translation failed. No translations were generated. Available services: {", ".join(available_services)}. Please check your API keys and try again.')
        elif translation_type == 'section_name':
            # Translate section name only
            result = translate_section_name_auto(
                name=text,
                target_languages=target_languages,
                service_name=service_name
            )

            if result and len(result) > 0:
                return json_ok(translations=result, service_used=auto_translator.get_default_service())
            else:
                # Return partial success if some translations failed but we have at least one
                # This allows the frontend to show successful translations even if some languages failed
                return json_error(
                    f'Translation failed for some languages. Available services: {", ".join(available_services)}. Some translations may have failed due to service limitations.',
                    200,
                    success=False,
                    translations=result or {},
                )

        elif translation_type == 'question_option':
            # Translate question option
            result = translate_question_option_auto(
                option_text=text,
                target_languages=target_languages,
                service_name=service_name
            )

            if result and len(result) > 0:
                return json_ok(translations=result, service_used=auto_translator.get_default_service())
            else:
                return json_bad_request(f'Translation failed. No translations were generated. Available services: {", ".join(available_services)}. Please check your API keys and try again.')

        elif translation_type == 'page_name':
            # Translate page name only
            result = translate_page_name_auto(
                name=text,
                target_languages=target_languages,
                service_name=service_name
            )

            if result and len(result) > 0:
                return json_ok(translations=result, service_used=auto_translator.get_default_service())
            else:
                return json_bad_request(f'Translation failed. No translations were generated. Available services: {", ".join(available_services)}. Please check your API keys and try again.')

        elif translation_type == 'template_name':
            # Translate template name only
            result = translate_template_name_auto(
                name=text,
                target_languages=target_languages,
                service_name=service_name
            )

            if result and len(result) > 0:
                return json_ok(translations=result, service_used=auto_translator.get_default_service())
            else:
                # Translation produced no output (e.g. acronym/proper noun returned unchanged or
                # service unavailable for this text). Return the source text for each target language
                # so the caller can populate fields rather than showing a hard error.
                fallback = {lang: text for lang in target_languages if lang}
                if fallback:
                    return json_ok(
                        translations=fallback,
                        service_used=None,
                        untranslated=True,
                        message='No translation available; original text returned.',
                    )
                return json_bad_request(f'Translation failed. No translations were generated. Available services: {", ".join(available_services)}. Please check your API keys and try again.')

        elif translation_type == 'translation':
            # Translate a translation file entry (msgid to other languages)
            from app.utils.auto_translator import translate_text as auto_translate_text
            from config import Config

            # Try to import polib, but make it optional
            try:
                import polib  # type: ignore
            except ImportError:
                current_app.logger.warning("polib not available - translation file updates will be skipped")
                polib = None

            # Get message ID for the translation
            message_id = data.get('id', '').strip()
            if not message_id:
                return json_bad_request('Message ID is required for translation type')

            # Check if polib is available for file updates
            if polib is None:
                return json_bad_request('Translation file updates are not available. Please install polib package.')

            # Translate the text to the target languages
            translations = {}
            success_count = 0

            for lang in target_languages:
                try:
                    # ISO codes only (accept regional variants like fr_FR)
                    target_locale = str(lang or '').strip()
                    if '_' in target_locale:
                        target_locale = target_locale.split('_', 1)[0]
                    target_locale = target_locale.lower()
                    if target_locale == 'en':
                        continue  # Skip English as source language

                    # Translate the text
                    translated_text = auto_translate_text(
                        text=text,
                        target_language=target_locale,
                        service_name=service_name
                    )

                    if translated_text and translated_text.strip():
                        translations[target_locale] = translated_text

                        # Update (or create) the translation file
                        po_file_path = os.path.join(current_app.root_path, 'translations', target_locale, 'LC_MESSAGES', 'messages.po')
                        try:
                            # Ensure directory exists
                            po_dir = os.path.dirname(po_file_path)
                            with suppress(Exception):
                                os.makedirs(po_dir, exist_ok=True)

                            # Load existing PO or create a new one
                            if os.path.exists(po_file_path):
                                po = polib.pofile(po_file_path)
                            else:
                                po = polib.POFile()
                                # Try to copy metadata/header from English if available
                                en_po_path = os.path.join(current_app.root_path, 'translations', 'en', 'LC_MESSAGES', 'messages.po')
                                if os.path.exists(en_po_path):
                                    with suppress(Exception):
                                        en_po = polib.pofile(en_po_path)
                                        if getattr(en_po, "metadata", None):
                                            po.metadata = dict(en_po.metadata)
                                # Minimal metadata fallback
                                if not getattr(po, "metadata", None):
                                    po.metadata = {}
                                po.metadata.setdefault('Content-Type', 'text/plain; charset=utf-8')
                                po.metadata.setdefault('Content-Transfer-Encoding', '8bit')
                                po.metadata.setdefault('Language', target_locale)

                            entry = po.find(message_id)
                            if entry:
                                # Update existing entry (plural or non-plural)
                                if hasattr(entry, 'msgid_plural') and entry.msgid_plural:
                                    if not hasattr(entry, 'msgstr_plural') or not entry.msgstr_plural:
                                        entry.msgstr_plural = {}
                                    entry.msgstr_plural[0] = translated_text
                                    # If plural forms exist, mirror into second form as a simple fallback
                                    if len(entry.msgstr_plural) > 1:
                                        entry.msgstr_plural[1] = translated_text
                                    current_app.logger.info(f"Updated existing plural translation for {message_id} in {target_locale}: {translated_text}")
                                else:
                                    entry.msgstr = translated_text
                                    current_app.logger.info(f"Updated existing translation for {message_id} in {target_locale}: {translated_text}")
                            else:
                                # Create new entry if it doesn't exist
                                en_po_path = os.path.join(current_app.root_path, 'translations', 'en', 'LC_MESSAGES', 'messages.po')
                                is_plural = False
                                msgid_plural = None
                                if os.path.exists(en_po_path):
                                    with suppress(Exception):
                                        en_po = polib.pofile(en_po_path)
                                        en_entry = en_po.find(message_id)
                                        if en_entry and hasattr(en_entry, 'msgid_plural') and en_entry.msgid_plural:
                                            is_plural = True
                                            msgid_plural = en_entry.msgid_plural

                                if is_plural:
                                    entry = polib.POEntry(msgid=message_id, msgid_plural=msgid_plural or message_id)
                                    entry.msgstr_plural = {0: translated_text, 1: translated_text}
                                else:
                                    entry = polib.POEntry(msgid=message_id, msgstr=translated_text)
                                po.append(entry)
                                current_app.logger.info(f"Created new translation for {message_id} in {target_locale}: {translated_text}")

                            po.save(po_file_path)
                            success_count += 1
                        except Exception as e:
                            current_app.logger.error(f"Error updating/creating po file for {target_locale}: {e}", exc_info=True)

                except Exception as e:
                    current_app.logger.error(f"Error translating to {lang}: {e}", exc_info=True)
                    continue

            if success_count > 0:
                return json_ok(
                    translations=translations,
                    updated_count=success_count,
                    service_used=service_name,
                )
            else:
                return json_server_error('Failed to translate or update translation files')

        else:
            return json_bad_request('Invalid translation type')

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)

@bp.route('/api/bulk-update-translations', methods=['POST'])
@limiter.exempt
@permission_required('admin.templates.edit')
def api_bulk_update_translations():
    """API endpoint for bulk updating translations of template items"""
    try:
        from app.models import FormItem, FormSection, FormPage
        from app.extensions import db

        data = get_json_safe()
        err = require_json_keys(data, ['items'])
        if err:
            return err

        items = data.get('items', [])
        if not items:
            return json_bad_request('No items provided')

        current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Bulk update request received with {len(items)} items")
        current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Request data: {data}")

        success_count = 0
        error_count = 0
        errors = []

        # Process all items in a single transaction
        try:
            for item in items:
                try:
                    if not isinstance(item, dict):
                        error_count += 1
                        errors.append(f"Invalid item format: {type(item)}")
                        continue

                    item_id = item.get('id')
                    item_type = item.get('type')  # 'question', 'document_field', 'section', or 'page'
                    translations = item.get('translations', {})

                    current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Processing item {item_id} of type {item_type}")
                    current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Translations: {translations}")

                    if not item_id or not item_type:
                        error_count += 1
                        errors.append(f"Missing ID or type for item")
                        current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Missing ID or type for item: {item}")
                        continue

                    # Handle different item types (including plugin_* types)
                    is_form_item_type = (
                        item_type in ['indicator', 'question', 'document_field', 'matrix']
                        or (isinstance(item_type, str) and item_type.startswith('plugin_'))
                    )
                    if is_form_item_type:
                        # Map item type to string values (plugin_* types pass through as-is)
                        if item_type == 'indicator':
                            form_item_type = 'indicator'
                        elif item_type == 'question':
                            form_item_type = 'question'
                        elif item_type == 'document_field':
                            form_item_type = 'document_field'
                        elif item_type == 'matrix':
                            form_item_type = 'matrix'
                        else:
                            form_item_type = item_type  # e.g. plugin_interactive_map

                        # Find the form item
                        current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Looking for form item with item_id={item_id}, item_type={form_item_type}")
                        form_item = FormItem.query.filter_by(id=item_id, item_type=form_item_type).first()
                        if not form_item:
                            error_count += 1
                            errors.append(f"Form item not found: {item_id}")
                            current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Form item not found: {item_id}")
                            continue

                        current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Found form item: {form_item.id}, label: {form_item.label}")
                        current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Current label_translations: {form_item.label_translations}")
                        current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Current definition_translations: {form_item.definition_translations}")

                        # Update form item translations
                        # Handle both nested format and individual language format
                        if 'label_translations' in translations:
                            # Merge with existing translations instead of replacing
                            if not form_item.label_translations:
                                form_item.label_translations = {}
                            existing_translations = form_item.label_translations.copy()
                            new_translations = translations['label_translations']
                            existing_translations.update(new_translations)
                            form_item.label_translations = existing_translations
                            current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Merged label translations: {form_item.label_translations}")
                        else:
                            # Handle individual language fields (e.g., label_french, label_spanish)
                            current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Processing individual label translations")
                            if not form_item.label_translations:
                                form_item.label_translations = {}
                            label_translations = form_item.label_translations.copy()
                            current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Current label translations: {label_translations}")
                            for key, value in translations.items():
                                if key.startswith('label_') and len(key) > 6:
                                    lang = key[6:]  # Remove 'label_' prefix
                                    if value and value.strip():
                                        label_translations[lang] = value.strip()
                                        current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Added label translation {lang}: {value.strip()}")
                            if label_translations:
                                form_item.label_translations = label_translations
                                current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Updated form item label translations: {form_item.label_translations}")

                        if 'definition_translations' in translations:
                            # Merge with existing translations instead of replacing
                            if not form_item.definition_translations:
                                form_item.definition_translations = {}
                            existing_translations = form_item.definition_translations.copy()
                            new_translations = translations['definition_translations']
                            existing_translations.update(new_translations)
                            form_item.definition_translations = existing_translations
                            current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Merged definition translations: {form_item.definition_translations}")
                        elif 'description_translations' in translations:
                            # Merge with existing translations instead of replacing
                            if not form_item.description_translations:
                                form_item.description_translations = {}
                            existing_translations = form_item.description_translations.copy()
                            new_translations = translations['description_translations']
                            existing_translations.update(new_translations)
                            form_item.description_translations = existing_translations
                            current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Merged description translations: {form_item.description_translations}")
                        else:
                            # Handle individual description language fields
                            if form_item.is_indicator or form_item.is_question:
                                if not form_item.definition_translations:
                                    form_item.definition_translations = {}
                                definition_translations = form_item.definition_translations.copy()
                                for key, value in translations.items():
                                    if key.startswith('description_') and len(key) > 12:
                                        lang = key[12:]  # Remove 'description_' prefix
                                        if value and value.strip():
                                            definition_translations[lang] = value.strip()
                                if definition_translations:
                                    form_item.definition_translations = definition_translations
                            elif form_item.is_document_field or form_item.is_matrix:
                                if not form_item.description_translations:
                                    form_item.description_translations = {}
                                description_translations = form_item.description_translations.copy()
                                for key, value in translations.items():
                                    if key.startswith('description_') and len(key) > 12:
                                        lang = key[12:]  # Remove 'description_' prefix
                                        if value and value.strip():
                                            description_translations[lang] = value.strip()
                                if description_translations:
                                    form_item.description_translations = description_translations

                        if 'options_translations' in translations:
                            form_item.options_translations = translations['options_translations']

                        # Debug: Print what the translations look like after update
                        current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: AFTER UPDATE - label_translations: {form_item.label_translations}")
                        current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: AFTER UPDATE - definition_translations: {form_item.definition_translations}")
                        current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: AFTER UPDATE - description_translations: {form_item.description_translations}")

                    elif item_type == 'section':
                        # Find the section
                        section = FormSection.query.filter_by(id=item_id).first()
                        if not section:
                            error_count += 1
                            errors.append(f"Section not found: {item_id}")
                            continue

                        # Update section name translations
                        if 'name_translations' in translations:
                            # Merge with existing translations instead of replacing
                            if not section.name_translations:
                                section.name_translations = {}
                            existing_translations = section.name_translations.copy()
                            new_translations = translations['name_translations']
                            existing_translations.update(new_translations)
                            section.name_translations = existing_translations
                            current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Merged section name translations: {section.name_translations}")

                    elif item_type == 'page':
                        # Find the page
                        page = FormPage.query.filter_by(id=item_id).first()
                        if not page:
                            error_count += 1
                            errors.append(f"Page not found: {item_id}")
                            continue

                        # Update page name translations
                        if 'name_translations' in translations:
                            # Merge with existing translations instead of replacing
                            if not page.name_translations:
                                page.name_translations = {}
                            existing_translations = page.name_translations.copy()
                            new_translations = translations['name_translations']
                            existing_translations.update(new_translations)
                            page.name_translations = existing_translations
                            current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Merged page name translations: {page.name_translations}")

                    elif item_type == 'template_name':
                        # Find the template; name_translations live on FormTemplateVersion, not FormTemplate
                        from app.models import FormTemplate, FormTemplateVersion
                        template = FormTemplate.query.filter_by(id=item_id).first()
                        if not template:
                            error_count += 1
                            errors.append(f"Template not found: {item_id}")
                            continue
                        # Use the version that owns the template name (published or first)
                        version = template.published_version or template.versions.order_by(FormTemplateVersion.created_at).first()
                        if not version:
                            error_count += 1
                            errors.append(f"No version found for template: {item_id}")
                            continue
                        # Update name_translations on the version (FormTemplate.name_translations is a read-only property from version)
                        if 'name_translations' in translations:
                            if not version.name_translations:
                                version.name_translations = {}
                            existing_translations = version.name_translations.copy()
                            new_translations = translations['name_translations']
                            existing_translations.update(new_translations)
                            version.name_translations = existing_translations
                            current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Merged template name translations on version {version.id}: {version.name_translations}")

                    else:
                        error_count += 1
                        errors.append(f"Invalid item type: {item_type}")
                        continue

                    success_count += 1
                    current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Successfully processed item {item_id}")

                except Exception as e:
                    error_count += 1
                    errors.append(f"Error updating item {item.get('id', 'unknown')}.")
                    current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Error updating item {item.get('id', 'unknown')}: {str(e)}")
                    current_app.logger.error(f"Error updating translations for item {item.get('id', 'unknown')}: {e}")

            # Commit all changes at once
            current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: Committing {success_count} successful updates")
            db.session.flush()
            current_app.logger.debug("AUTO-TRANSLATE DEBUG: Database commit successful")

            # Verify the data was actually saved by re-querying one of the updated items
            if success_count > 0:
                current_app.logger.debug("AUTO-TRANSLATE DEBUG: Verifying data persistence...")
                for item in items:
                    item_type = item.get('type')
                    is_form_item = (
                        item_type in ['indicator', 'question', 'document_field', 'matrix']
                        or (isinstance(item_type, str) and item_type.startswith('plugin_'))
                    )
                    if is_form_item:
                        item_id = item.get('id')
                        if item_type == 'indicator':
                            form_item_type = 'indicator'
                        elif item_type == 'question':
                            form_item_type = 'question'
                        elif item_type == 'document_field':
                            form_item_type = 'document_field'
                        elif item_type == 'matrix':
                            form_item_type = 'matrix'
                        else:
                            form_item_type = item_type

                        verify_item = FormItem.query.filter_by(id=item_id, item_type=form_item_type).first()
                        if verify_item:
                            current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: VERIFICATION - Item {item_id} label_translations: {verify_item.label_translations}")
                            current_app.logger.debug(f"AUTO-TRANSLATE DEBUG: VERIFICATION - Item {item_id} definition_translations: {verify_item.definition_translations}")
                            break  # Just verify one item

        except Exception as e:
            return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)

        return json_ok(
            message=f'Successfully updated {success_count} items. {error_count} errors occurred.',
            success_count=success_count,
            error_count=error_count,
            errors=errors,
        )

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)
