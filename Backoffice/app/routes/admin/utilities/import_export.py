from contextlib import suppress
import json
import logging
import os
import threading

from flask import current_app, flash, redirect, render_template, request, url_for
from flask_babel import _
from flask_login import current_user
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from app import db
from app.models import (
    IndicatorBank,
    IndicatorBankHistory,
    IndicatorBankType,
    IndicatorBankUnit,
    IndicatorSuggestion,
    Sector,
    SubSector,
)
from app.routes.admin.shared import permission_required
from app.routes.admin.utilities import bp
from app.utils.advanced_validation import validate_upload_extension_and_mime
from app.utils.api_helpers import GENERIC_ERROR_MESSAGE
from app.utils.api_responses import json_bad_request, json_ok
from app.utils.datetime_helpers import utcnow
from app.utils.error_handling import handle_json_view_exception
from app.utils.file_parsing import EXCEL_EXTENSIONS
from app.utils.transactions import request_transaction_rollback
from app.services.indicator_measurement_sync import backfill_fk_from_strings_bank

logger = logging.getLogger(__name__)

# === Import/Export Routes ===
@bp.route("/indicator_bank/import", methods=["GET", "POST"])
@permission_required('admin.indicator_bank.edit')
def import_indicators():
    """Import indicators from Excel file"""
    if request.method == 'POST':
        temp_path = None
        try:
            if 'file' not in request.files:
                return json_bad_request('No file selected.')

            file = request.files['file']
            if file.filename == '':
                return json_bad_request('No file selected.')

            valid, error_msg, ext = validate_upload_extension_and_mime(file, EXCEL_EXTENSIONS)
            if valid:
                filename = secure_filename(file.filename)

                # Save temporary file
                from app.utils.file_paths import get_temp_upload_path
                temp_dir = get_temp_upload_path()
                os.makedirs(temp_dir, exist_ok=True)
                temp_path = os.path.join(temp_dir, filename)
                file.save(temp_path)

                # Process the Excel file
                current_app.logger.info(f"Processing import file: {temp_path}")
                result = _process_indicator_import(temp_path)
                current_app.logger.info(f"Import result: {result}")

                if result['success']:
                    message = (
                        f"Successfully imported {result['imported']} indicators. "
                        f"{result['updated']} indicators were updated."
                    )
                    mt_i = result.get("measurement_types_imported") or 0
                    mt_u = result.get("measurement_types_updated") or 0
                    mu_i = result.get("measurement_units_imported") or 0
                    mu_u = result.get("measurement_units_updated") or 0
                    if mt_i or mt_u:
                        message += f" Measurement types: {mt_i} new, {mt_u} updated."
                    if mu_i or mu_u:
                        message += f" Measurement units: {mu_i} new, {mu_u} updated."
                    if result['errors']:
                        message += f" Encountered {len(result['errors'])} errors during import."
                    return json_ok(message=message)
                else:
                    current_app.logger.error(f"Import failed: {result['message']}")
                    return json_bad_request(f"Import failed: {result['message']}")
            else:
                return json_bad_request(error_msg or _('Please upload an Excel file (.xlsx or .xls).'))

        except Exception as e:
            return handle_json_view_exception(e, 'Error processing import file.', status_code=500)
        finally:
            # Clean up temporary file with error handling
            if temp_path and os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except (PermissionError, OSError) as e:
                    current_app.logger.warning(f"Could not remove temporary file {temp_path}: {e}")
                    # Schedule cleanup for later (Windows file locking issue)
                    def delayed_cleanup():
                        import time
                        time.sleep(5)  # Wait 5 seconds
                        with suppress(Exception):
                            os.remove(temp_path)
                    threading.Thread(target=delayed_cleanup, daemon=True).start()

    # For GET requests, redirect to the indicator bank page
    return redirect(url_for("system_admin.manage_indicator_bank"))

@bp.route("/indicator_bank/change_history", methods=["GET"])
@permission_required('admin.audit.view')
def indicator_change_history():
    """View indicator change history"""
    # Get all change history records (no pagination - AG Grid handles pagination client-side)
    changes = IndicatorBankHistory.query.order_by(
        IndicatorBankHistory.created_at.desc()
    ).all()

    return render_template("admin/indicator_bank/change_history.html",
                         changes=changes,
                         title="Indicator Change History")

# === Indicator Suggestions Management ===
@bp.route("/indicator_suggestions", methods=["GET"])
@permission_required('admin.indicator_bank.suggestions.review')
def manage_indicator_suggestions():
    """Manage indicator suggestions from users"""
    from app.utils.api_pagination import validate_pagination_params
    page, per_page = validate_pagination_params(request.args, default_per_page=20, max_per_page=100)
    status_filter = request.args.get('status', '')
    suggestion_type_filter = request.args.get('suggestion_type', '')

    query = IndicatorSuggestion.query

    if status_filter:
        query = query.filter(IndicatorSuggestion.status == status_filter)

    pagination = query.order_by(
        IndicatorSuggestion.submitted_at.desc()
    ).paginate(page=page, per_page=per_page, error_out=False)

    return render_template(
        "admin/indicator_suggestions.html",
        suggestions=pagination.items,
        pagination=pagination,
        status_filter=status_filter,
        suggestion_type_filter=suggestion_type_filter,
        title="Manage Indicator Suggestions",
    )

@bp.route("/indicator_suggestions/view/<int:suggestion_id>", methods=["GET"])
@permission_required('admin.indicator_bank.suggestions.review')
def view_indicator_suggestion(suggestion_id):
    """View individual indicator suggestion"""
    suggestion = IndicatorSuggestion.query.get_or_404(suggestion_id)
    return render_template("admin/view_indicator_suggestion.html",
                         suggestion=suggestion,
                         title=f"View Suggestion: {suggestion.indicator_name}")

@bp.route("/indicator_suggestions/update_status/<int:suggestion_id>", methods=["POST"])
@permission_required('admin.indicator_bank.suggestions.review')
def update_indicator_suggestion_status(suggestion_id):
    """Update indicator suggestion status"""
    suggestion = IndicatorSuggestion.query.get_or_404(suggestion_id)
    new_status = request.form.get('status')
    admin_notes = request.form.get('admin_notes', '').strip()

    try:
        if new_status in ['Pending', 'approved', 'rejected', 'implemented']:
            suggestion.status = new_status
            suggestion.admin_notes = admin_notes
            suggestion.reviewed_by = current_user
            suggestion.reviewed_at = utcnow()

            # If approved, optionally create the indicator automatically
            if new_status == 'approved':
                _create_indicator_from_suggestion(suggestion)

            db.session.flush()
            flash(_("Suggestion status updated to %(status)s.", status=new_status), "success")
        else:
            flash(_("Invalid status provided."), "danger")

    except Exception as e:
        request_transaction_rollback()
        current_app.logger.error(f"Error updating suggestion status: {e}", exc_info=True)
        flash(_("Error updating suggestion status."), "danger")

    return redirect(url_for("utilities.view_indicator_suggestion", suggestion_id=suggestion_id))

@bp.route("/indicator_suggestions/delete/<int:suggestion_id>", methods=["POST"])
@permission_required('admin.indicator_bank.suggestions.review')
def delete_indicator_suggestion(suggestion_id):
    """Delete indicator suggestion"""
    suggestion = IndicatorSuggestion.query.get_or_404(suggestion_id)

    try:
        db.session.delete(suggestion)
        db.session.flush()
        flash(_("Suggestion deleted successfully."), "success")

    except Exception as e:
        request_transaction_rollback()
        current_app.logger.error(f"Error deleting suggestion: {e}", exc_info=True)
        flash(_("Error deleting suggestion."), "danger")

    return redirect(url_for("utilities.manage_indicator_suggestions"))

# === Helper Functions ===

def _process_indicator_import(file_path):
    """Process Excel file for indicator import"""
    try:
        current_app.logger.info(f"Loading workbook from: {file_path}")
        wb = load_workbook(file_path, read_only=True, data_only=True)

        result = {
            'success': True,
            'imported': 0,  # indicators imported from main sheet OR DB_Indicators
            'updated': 0,   # indicators updated from main sheet OR DB_Indicators
            'sectors_imported': 0,
            'sectors_updated': 0,
            'subsectors_imported': 0,
            'subsectors_updated': 0,
            'common_words_imported': 0,
            'common_words_updated': 0,
            'measurement_types_imported': 0,
            'measurement_types_updated': 0,
            'measurement_units_imported': 0,
            'measurement_units_updated': 0,
            'errors': [],
            'message': ''
        }

        # --- Helpers ---
        def _norm_header(v):
            return str(v).strip().lower() if v is not None else ""

        def _to_int(v):
            try:
                if v is None or v == "":
                    return None
                if isinstance(v, bool):
                    return int(v)
                if isinstance(v, (int, float)):
                    return int(v)
                s = str(v).strip()
                if not s:
                    return None
                return int(float(s))
            except Exception as e:
                logger.debug("int parse failed for %r: %s", v, e)
                return None

        def _to_bool(v):
            if v is None or v == "":
                return None
            if isinstance(v, bool):
                return v
            if isinstance(v, (int, float)):
                return bool(int(v))
            s = str(v).strip().lower()
            if s in ("true", "t", "yes", "y", "1"):
                return True
            if s in ("false", "f", "no", "n", "0"):
                return False
            return None

        def _to_json_dict(v):
            if v is None or v == "":
                return None
            if isinstance(v, dict):
                return v
            try:
                parsed = json.loads(str(v))
                return parsed if isinstance(parsed, dict) else None
            except Exception as e:
                logger.debug("json parse failed for %r: %s", v, e)
                return None

        def _sheet_rows_as_dicts(ws):
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            raw_headers = list(rows[0])
            headers = [_norm_header(h) for h in raw_headers]
            out = []
            for r in rows[1:]:
                if not r:
                    continue
                # skip fully empty rows
                if all((c is None or str(c).strip() == "") for c in r):
                    continue
                d = {}
                for idx, key in enumerate(headers):
                    if not key:
                        continue
                    if idx < len(r):
                        d[key] = r[idx]
                out.append(d)
            return out

        def _build_levels_json(primary_id, secondary_id, tertiary_id):
            data = {}
            if primary_id is not None:
                data["primary"] = primary_id
            if secondary_id is not None:
                data["secondary"] = secondary_id
            if tertiary_id is not None:
                data["tertiary"] = tertiary_id
            return data or None

        sheetnames = set(wb.sheetnames or [])

        # If file contains hidden DB_* sheets (from export), use full relational import; otherwise use main sheet
        if (
            "DB_Indicators" in sheetnames
            or "DB_Sectors_SubSectors" in sheetnames
            or "DB_CommonWords" in sheetnames
            or "DB_MeasurementTypes" in sheetnames
            or "DB_MeasurementUnits" in sheetnames
        ):
            # 1) Sectors/SubSectors
            if "DB_Sectors_SubSectors" in sheetnames:
                from app.models import Sector, SubSector
                ws_ss = wb["DB_Sectors_SubSectors"]
                ss_rows = _sheet_rows_as_dicts(ws_ss)

                for idx, r in enumerate(ss_rows, 2):
                    try:
                        record_type = (r.get("record_type") or "").strip().lower()
                        rid = _to_int(r.get("id"))
                        name = (r.get("name") or "").strip()
                        description = (r.get("description") or None)
                        sector_id = _to_int(r.get("sector_id"))
                        display_order = _to_int(r.get("display_order"))
                        is_active = _to_bool(r.get("is_active"))
                        icon_class = (r.get("icon_class") or None)
                        logo_filename = (r.get("logo_filename") or None)
                        logo_path = (r.get("logo_path") or None)
                        name_translations = _to_json_dict(r.get("name_translations_json")) or {}

                        if record_type == "sector":
                            if not rid and not name:
                                continue
                            obj = Sector.query.get(rid) if rid else Sector.query.filter_by(name=name).first()
                            is_new = obj is None
                            if obj is None:
                                obj = Sector()
                                if rid:
                                    obj.id = rid
                                db.session.add(obj)
                            if name:
                                obj.name = name
                            obj.description = description
                            if display_order is not None:
                                obj.display_order = display_order
                            if is_active is not None:
                                obj.is_active = is_active
                            obj.icon_class = icon_class
                            obj.logo_filename = logo_filename
                            obj.logo_path = logo_path
                            obj.name_translations = name_translations or {}
                            if is_new:
                                result["sectors_imported"] += 1
                            else:
                                result["sectors_updated"] += 1

                        elif record_type == "subsector":
                            if not rid and not name:
                                continue
                            obj = SubSector.query.get(rid) if rid else SubSector.query.filter_by(name=name).first()
                            is_new = obj is None
                            if obj is None:
                                obj = SubSector()
                                if rid:
                                    obj.id = rid
                                db.session.add(obj)
                            if name:
                                obj.name = name
                            obj.description = description
                            if sector_id is not None:
                                obj.sector_id = sector_id
                            if display_order is not None:
                                obj.display_order = display_order
                            if is_active is not None:
                                obj.is_active = is_active
                            obj.icon_class = icon_class
                            obj.logo_filename = logo_filename
                            obj.logo_path = logo_path
                            obj.name_translations = name_translations or {}
                            if is_new:
                                result["subsectors_imported"] += 1
                            else:
                                result["subsectors_updated"] += 1
                    except Exception as e:
                        current_app.logger.error(f"Error processing DB_Sectors_SubSectors row {idx}: {e}", exc_info=True)
                        result["errors"].append(f"DB_Sectors_SubSectors row {idx}: error.")

            # 2) Common words
            if "DB_CommonWords" in sheetnames:
                from app.models import CommonWord
                ws_cw = wb["DB_CommonWords"]
                cw_rows = _sheet_rows_as_dicts(ws_cw)
                for idx, r in enumerate(cw_rows, 2):
                    try:
                        rid = _to_int(r.get("id"))
                        term = (r.get("term") or "").strip()
                        meaning = (r.get("meaning") or "").strip()
                        is_active = _to_bool(r.get("is_active"))
                        meaning_translations = _to_json_dict(r.get("meaning_translations_json")) or {}

                        if not term and not rid:
                            continue

                        obj = CommonWord.query.get(rid) if rid else CommonWord.query.filter_by(term=term).first()
                        is_new = obj is None
                        if obj is None:
                            obj = CommonWord()
                            if rid:
                                obj.id = rid
                            db.session.add(obj)
                        if term:
                            obj.term = term
                        if meaning:
                            obj.meaning = meaning
                        if is_active is not None:
                            obj.is_active = is_active
                        obj.meaning_translations = meaning_translations or {}

                        if is_new:
                            result["common_words_imported"] += 1
                        else:
                            result["common_words_updated"] += 1
                    except Exception as e:
                        current_app.logger.error(f"Error processing DB_CommonWords row {idx}: {e}", exc_info=True)
                        result["errors"].append(f"DB_CommonWords row {idx}: Validation error.")

            if "DB_MeasurementTypes" in sheetnames or "DB_MeasurementUnits" in sheetnames:
                from sqlalchemy.orm.attributes import flag_modified

            # 2b) Measurement types (central catalog; run before DB_Indicators)
            if "DB_MeasurementTypes" in sheetnames:
                ws_mt = wb["DB_MeasurementTypes"]
                mt_rows = _sheet_rows_as_dicts(ws_mt)
                for idx, r in enumerate(mt_rows, 2):
                    try:
                        rid = _to_int(r.get("id"))
                        code = (r.get("code") or "").strip().lower()
                        name = (r.get("name") or "").strip()
                        name_translations = _to_json_dict(r.get("name_translations_json")) or {}
                        sort_order = _to_int(r.get("sort_order"))
                        is_active = _to_bool(r.get("is_active"))
                        if not code and not rid:
                            continue
                        obj = IndicatorBankType.query.get(rid) if rid else None
                        if obj is None and code:
                            obj = (
                                IndicatorBankType.query.filter(
                                    db.func.lower(IndicatorBankType.code) == code
                                ).first()
                            )
                        is_new = obj is None
                        if obj is None:
                            obj = IndicatorBankType()
                            if rid:
                                obj.id = rid
                            db.session.add(obj)
                        if code:
                            obj.code = code
                        if name:
                            obj.name = name
                        obj.name_translations = name_translations or {}
                        flag_modified(obj, "name_translations")
                        if sort_order is not None:
                            obj.sort_order = sort_order
                        if is_active is not None:
                            obj.is_active = is_active
                        if is_new:
                            result["measurement_types_imported"] += 1
                        else:
                            result["measurement_types_updated"] += 1
                    except Exception as e:
                        current_app.logger.error(
                            f"Error processing DB_MeasurementTypes row {idx}: {e}", exc_info=True
                        )
                        result["errors"].append(f"DB_MeasurementTypes row {idx}: error.")

            # 2c) Measurement units
            if "DB_MeasurementUnits" in sheetnames:
                ws_mu = wb["DB_MeasurementUnits"]
                mu_rows = _sheet_rows_as_dicts(ws_mu)
                for idx, r in enumerate(mu_rows, 2):
                    try:
                        rid = _to_int(r.get("id"))
                        code = (r.get("code") or "").strip().lower()
                        name = (r.get("name") or "").strip()
                        name_translations = _to_json_dict(r.get("name_translations_json")) or {}
                        sort_order = _to_int(r.get("sort_order"))
                        is_active = _to_bool(r.get("is_active"))
                        allows_disaggregation = _to_bool(r.get("allows_disaggregation"))
                        if not code and not rid:
                            continue
                        obj = IndicatorBankUnit.query.get(rid) if rid else None
                        if obj is None and code:
                            obj = (
                                IndicatorBankUnit.query.filter(
                                    db.func.lower(IndicatorBankUnit.code) == code
                                ).first()
                            )
                        is_new = obj is None
                        if obj is None:
                            obj = IndicatorBankUnit()
                            if rid:
                                obj.id = rid
                            db.session.add(obj)
                        if code:
                            obj.code = code
                        if name:
                            obj.name = name
                        obj.name_translations = name_translations or {}
                        flag_modified(obj, "name_translations")
                        if sort_order is not None:
                            obj.sort_order = sort_order
                        if is_active is not None:
                            obj.is_active = is_active
                        if allows_disaggregation is not None:
                            obj.allows_disaggregation = allows_disaggregation
                        if is_new:
                            result["measurement_units_imported"] += 1
                        else:
                            result["measurement_units_updated"] += 1
                    except Exception as e:
                        current_app.logger.error(
                            f"Error processing DB_MeasurementUnits row {idx}: {e}", exc_info=True
                        )
                        result["errors"].append(f"DB_MeasurementUnits row {idx}: error.")

            # 3) Indicators (from hidden DB_Indicators sheet when present)
            if "DB_Indicators" in sheetnames:
                ws_ind = wb["DB_Indicators"]
                ind_rows = _sheet_rows_as_dicts(ws_ind)
                for idx, r in enumerate(ind_rows, 2):
                    try:
                        rid = _to_int(r.get("id"))
                        name = (r.get("name") or "").strip()
                        if not name and not rid:
                            continue

                        definition = (r.get("definition") or "").strip()
                        indicator_type = (r.get("type") or "").strip() or "numeric"
                        unit = (r.get("unit") or "").strip()
                        fdrs_kpi_code = (r.get("fdrs_kpi_code") or "").strip() or None
                        emergency = _to_bool(r.get("emergency"))
                        archived = _to_bool(r.get("archived"))
                        comments = (r.get("comments") or None)
                        programs = (r.get("related_programs") or "")
                        programs = str(programs).strip() if programs is not None else ""

                        s_primary = _to_int(r.get("sector_primary_id"))
                        s_secondary = _to_int(r.get("sector_secondary_id"))
                        s_tertiary = _to_int(r.get("sector_tertiary_id"))
                        ss_primary = _to_int(r.get("subsector_primary_id"))
                        ss_secondary = _to_int(r.get("subsector_secondary_id"))
                        ss_tertiary = _to_int(r.get("subsector_tertiary_id"))

                        name_translations = _to_json_dict(r.get("name_translations_json")) or {}
                        definition_translations = _to_json_dict(r.get("definition_translations_json")) or {}
                        itid = _to_int(r.get("indicator_type_id"))
                        iuid = _to_int(r.get("indicator_unit_id"))

                        existing = IndicatorBank.query.get(rid) if rid else IndicatorBank.query.filter_by(name=name).first()
                        is_new = existing is None

                        if existing is None:
                            existing = IndicatorBank(
                                name=name,
                                definition=definition,
                                type=indicator_type,
                                unit=unit,
                                fdrs_kpi_code=fdrs_kpi_code,
                                emergency=bool(emergency) if emergency is not None else False,
                                related_programs=programs,
                            )
                            if rid:
                                existing.id = rid
                            db.session.add(existing)
                            db.session.flush()
                        else:
                            # update base fields
                            if name:
                                existing.name = name
                            existing.definition = definition
                            existing.type = indicator_type
                            existing.unit = unit
                            existing.fdrs_kpi_code = fdrs_kpi_code
                            if emergency is not None:
                                existing.emergency = emergency
                            if archived is not None:
                                existing.archived = archived
                            existing.comments = comments
                            existing.related_programs = programs

                        # update relational JSON fields + translations
                        existing.sector = _build_levels_json(s_primary, s_secondary, s_tertiary)
                        existing.sub_sector = _build_levels_json(ss_primary, ss_secondary, ss_tertiary)
                        existing.name_translations = name_translations or {}
                        existing.definition_translations = definition_translations or {}

                        backfill_fk_from_strings_bank(existing)
                        if itid is not None:
                            existing.indicator_type_id = itid
                        if iuid is not None:
                            existing.indicator_unit_id = iuid
                        if hasattr(existing, "sync_type_unit_string_columns"):
                            existing.sync_type_unit_string_columns()

                        # Create history record (minimal but consistent)
                        history = IndicatorBankHistory(
                            indicator_bank_id=existing.id,
                            user_id=current_user.id,
                            name=existing.name,
                            type=existing.type,
                            unit=existing.unit,
                            fdrs_kpi_code=existing.fdrs_kpi_code,
                            definition=existing.definition,
                            name_translations=existing.name_translations,
                            definition_translations=existing.definition_translations,
                            archived=existing.archived,
                            comments=existing.comments,
                            emergency=existing.emergency,
                            related_programs=existing.related_programs,
                            sector=existing.sector,
                            sub_sector=existing.sub_sector,
                            change_type='CREATED' if is_new else 'UPDATED',
                            change_description=(
                                f'Indicator "{existing.name}" created via DB import by {current_user.name or current_user.email}'
                                if is_new
                                else f'Indicator "{existing.name}" updated via DB import by {current_user.name or current_user.email}'
                            ),
                        )
                        db.session.add(history)

                        if is_new:
                            result["imported"] += 1
                        else:
                            result["updated"] += 1
                    except Exception as e:
                        current_app.logger.error(f"Error processing DB_Indicators row {idx}: {e}", exc_info=True)
                        result["errors"].append(f"DB_Indicators row {idx}: error.")

            db.session.flush()
            return result

        # --- Main sheet import (primary): use active sheet ---
        ws = wb.active

        # Build header -> column index (0-based) for optional columns
        headers_raw = [cell.value for cell in ws[1]]
        headers = [str(h).strip() if h is not None else "" for h in headers_raw]
        header_to_col = {}
        for idx, h in enumerate(headers):
            key = _norm_header(h) if h else ""
            if key and key not in header_to_col:
                header_to_col[key] = idx
        current_app.logger.info(f"Excel headers: {headers}")

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        current_app.logger.info(f"Total rows to process: {len(rows)}")

        def _val(row, idx):
            if row is None or idx is None or idx < 0 or idx >= len(row):
                return None
            return row[idx]

        def _resolve_sector_id(name_val):
            if name_val is None or str(name_val).strip() == "":
                return None
            s = Sector.query.filter_by(name=str(name_val).strip()).first()
            return s.id if s else None

        def _resolve_subsector_id(name_val):
            if name_val is None or str(name_val).strip() == "":
                return None
            ss = SubSector.query.filter_by(name=str(name_val).strip()).first()
            return ss.id if ss else None

        for row_num, row in enumerate(rows, 2):
            try:
                if row is None or len(row) < 7:
                    current_app.logger.warning(f"Skipping row {row_num}: row is None or has insufficient columns")
                    continue

                if _val(row, 1) is None or str(_val(row, 1)).strip() == "":
                    continue

                name = (str(_val(row, header_to_col.get("name", 1)) or "")).strip()
                definition = str(_val(row, header_to_col.get("definition", 2)) or "").strip() if _val(row, header_to_col.get("definition", 2)) is not None else ''
                indicator_type = str(_val(row, header_to_col.get("type", 3)) or "").strip() if _val(row, header_to_col.get("type", 3)) is not None else 'numeric'
                unit = str(_val(row, header_to_col.get("unit", 4)) or "").strip() if _val(row, header_to_col.get("unit", 4)) is not None else ''
                fdrs_kpi_code_col = header_to_col.get("fdrs kpi code")
                fdrs_kpi_code = (str(_val(row, fdrs_kpi_code_col) or "").strip() or None) if fdrs_kpi_code_col is not None else None
                _emergency = _to_bool(_val(row, header_to_col.get("emergency", 5))) if _val(row, header_to_col.get("emergency", 5)) is not None else False
                emergency = _emergency if _emergency is not None else False
                programs = str(_val(row, header_to_col.get("related programs", 6)) or "").strip() if _val(row, header_to_col.get("related programs", 6)) is not None else ''

                # Optional: localized names/definitions from "Name (code)" and "Definition (code)" columns
                name_translations = {}
                definition_translations = {}
                for key, col in header_to_col.items():
                    if key.startswith("name (") and key.endswith(")"):
                        code = key[6:-1].strip().lower()
                        if code and len(code) <= 6:
                            val = _val(row, col)
                            if val is not None and str(val).strip():
                                name_translations[code] = str(val).strip()
                    elif key.startswith("definition (") and key.endswith(")"):
                        code = key[12:-1].strip().lower()
                        if code and len(code) <= 6:
                            val = _val(row, col)
                            if val is not None and str(val).strip():
                                definition_translations[code] = str(val).strip()

                # Optional: sector/subsector names from "Sector Primary", "SubSector Primary", etc.
                s_primary = _resolve_sector_id(_val(row, header_to_col.get("sector primary")))
                s_secondary = _resolve_sector_id(_val(row, header_to_col.get("sector secondary")))
                s_tertiary = _resolve_sector_id(_val(row, header_to_col.get("sector tertiary")))
                ss_primary = _resolve_subsector_id(_val(row, header_to_col.get("subsector primary")))
                ss_secondary = _resolve_subsector_id(_val(row, header_to_col.get("subsector secondary")))
                ss_tertiary = _resolve_subsector_id(_val(row, header_to_col.get("subsector tertiary")))
                sector_json = _build_levels_json(s_primary, s_secondary, s_tertiary)
                sub_sector_json = _build_levels_json(ss_primary, ss_secondary, ss_tertiary)

                existing = IndicatorBank.query.filter_by(name=name).first()

                if existing:
                    changes = []
                    if existing.definition != definition:
                        old_def = existing.definition or ''
                        new_def = definition or ''
                        changes.append(f"Definition: Changed from '{old_def[:50]}{'...' if len(old_def) > 50 else ''}' to '{new_def[:50]}{'...' if len(new_def) > 50 else ''}'")
                    if existing.type != indicator_type:
                        changes.append(f"Type: Changed from '{existing.type or ''}' to '{indicator_type}'")
                    if existing.unit != unit:
                        changes.append(f"Unit: Changed from '{existing.unit or ''}' to '{unit}'")
                    if getattr(existing, 'fdrs_kpi_code', None) != fdrs_kpi_code:
                        changes.append(f"FDRS KPI Code: Changed from '{existing.fdrs_kpi_code or ''}' to '{fdrs_kpi_code or ''}'")
                    if existing.emergency != emergency:
                        changes.append(f"Emergency: Changed from '{existing.emergency}' to '{emergency}'")
                    if existing.related_programs != programs:
                        changes.append(f"Related Programs: Changed from '{existing.related_programs or ''}' to '{programs}'")

                    existing.definition = definition
                    existing.type = indicator_type
                    existing.unit = unit
                    existing.fdrs_kpi_code = fdrs_kpi_code
                    existing.emergency = emergency
                    existing.related_programs = programs
                    if name_translations:
                        existing.name_translations = {**(existing.name_translations or {}), **name_translations}
                    if definition_translations:
                        existing.definition_translations = {**(existing.definition_translations or {}), **definition_translations}
                    if sector_json is not None:
                        existing.sector = sector_json
                    if sub_sector_json is not None:
                        existing.sub_sector = sub_sector_json

                    backfill_fk_from_strings_bank(existing)

                    change_description = "; ".join(changes) if changes else "Indicator updated via import (no specific changes detected)"

                    history = IndicatorBankHistory(
                        indicator_bank_id=existing.id,
                        user_id=current_user.id,
                        name=existing.name,
                        type=existing.type,
                        unit=existing.unit,
                        fdrs_kpi_code=existing.fdrs_kpi_code,
                        definition=existing.definition,
                        name_translations=existing.name_translations,
                        definition_translations=existing.definition_translations,
                        archived=existing.archived,
                        comments=existing.comments,
                        emergency=existing.emergency,
                        related_programs=existing.related_programs,
                        sector=existing.sector,
                        sub_sector=existing.sub_sector,
                        change_type='UPDATED',
                        change_description=change_description
                    )
                    db.session.add(history)
                    result['updated'] += 1
                else:
                    new_indicator = IndicatorBank(
                        name=name,
                        definition=definition,
                        type=indicator_type,
                        unit=unit,
                        fdrs_kpi_code=fdrs_kpi_code,
                        emergency=emergency,
                        related_programs=programs
                    )
                    if name_translations:
                        new_indicator.name_translations = name_translations
                    if definition_translations:
                        new_indicator.definition_translations = definition_translations
                    if sector_json is not None:
                        new_indicator.sector = sector_json
                    if sub_sector_json is not None:
                        new_indicator.sub_sector = sub_sector_json
                    db.session.add(new_indicator)
                    db.session.flush()
                    backfill_fk_from_strings_bank(new_indicator)

                    history = IndicatorBankHistory(
                        indicator_bank_id=new_indicator.id,
                        user_id=current_user.id,
                        name=new_indicator.name,
                        type=new_indicator.type,
                        unit=new_indicator.unit,
                        fdrs_kpi_code=new_indicator.fdrs_kpi_code,
                        definition=new_indicator.definition,
                        name_translations=new_indicator.name_translations,
                        definition_translations=new_indicator.definition_translations,
                        archived=new_indicator.archived,
                        comments=new_indicator.comments,
                        emergency=new_indicator.emergency,
                        related_programs=new_indicator.related_programs,
                        sector=new_indicator.sector,
                        sub_sector=new_indicator.sub_sector,
                        change_type='CREATED',
                        change_description=f'Indicator "{new_indicator.name}" created via import by {current_user.name or current_user.email}'
                    )
                    db.session.add(history)
                    result['imported'] += 1

            except Exception as e:
                current_app.logger.error(f"Error processing row {row_num}: {str(e)}", exc_info=True)
                current_app.logger.error(f"Row data: {row}")
                result['errors'].append(f"Row {row_num}: error.")
                continue

        db.session.flush()

    except Exception as e:
        result['success'] = False
        result['message'] = GENERIC_ERROR_MESSAGE
        request_transaction_rollback()

    return result

def _create_indicator_from_suggestion(suggestion):
    """Create an indicator from an approved suggestion"""
    try:
        new_indicator = IndicatorBank(
            name=suggestion.suggested_name,
            definition=suggestion.suggested_definition or '',
            type=suggestion.suggested_type or 'numeric',
            unit=suggestion.suggested_unit or '',
            fdrs_kpi_code=None,
            emergency=suggestion.suggested_emergency or False,
            related_programs=suggestion.suggested_programs or ''
        )

        db.session.add(new_indicator)
        db.session.flush()  # Get the ID
        backfill_fk_from_strings_bank(new_indicator)

        # Create history record
        history = IndicatorBankHistory(
            indicator_bank_id=new_indicator.id,
            user_id=current_user.id,
            name=new_indicator.name,
            type=new_indicator.type,
            unit=new_indicator.unit,
            fdrs_kpi_code=new_indicator.fdrs_kpi_code,
            definition=new_indicator.definition,
            name_translations=new_indicator.name_translations,
            definition_translations=new_indicator.definition_translations,
            archived=new_indicator.archived,
            comments=new_indicator.comments,
            emergency=new_indicator.emergency,
            related_programs=new_indicator.related_programs,
            sector=new_indicator.sector,
            sub_sector=new_indicator.sub_sector,
            change_type='CREATED',
            change_description=f'Indicator "{new_indicator.name}" created from suggestion by {current_user.name or current_user.email}'
        )
        db.session.add(history)

        # Update suggestion status
        suggestion.status = 'implemented'
        suggestion.indicator_id = new_indicator.id

        return new_indicator

    except Exception as e:
        current_app.logger.error(f"Error creating indicator from suggestion: {e}", exc_info=True)
        return None
