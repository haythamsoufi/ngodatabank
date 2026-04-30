from contextlib import suppress
from datetime import datetime
import os
import json
import io
import zipfile
import shutil
import logging

from flask import request, flash, redirect, url_for, current_app, render_template, send_file
from flask_login import current_user
from flask_babel import _
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

from app.forms.content import TranslationForm
from app.routes.admin.shared import admin_required, permission_required, permission_required_any
from app.utils.request_utils import is_json_request, get_request_data
from app.services.translation.auto_translator import translate_text as auto_translate_text
from app.extensions import limiter
from app.utils.api_helpers import GENERIC_ERROR_MESSAGE, get_json_safe
from app.utils.error_handling import handle_json_view_exception
from app.utils.api_responses import json_bad_request, json_error, json_forbidden, json_ok, json_server_error, require_json_data, require_json_keys

from app.routes.admin.utilities import bp
from app.routes.admin.utilities.helpers import _translations_dir, _translations_po_path, _translations_pot_path, _entry_to_display_msgstr, _extract_page_name

logger = logging.getLogger(__name__)


def _update_po_translations(msgid, lang_to_msgstr):
    """Update PO translation files for the given msgid across languages.

    For each (lang, msgstr) pair: update the existing entry or create a new one
    (new entries are only created when msgstr is non-empty).
    Returns the count of languages successfully updated.
    """
    try:
        import polib  # type: ignore
    except ImportError:
        current_app.logger.warning("polib not available - translation file updates will be skipped")
        return 0

    updated = 0
    for lang, msgstr in lang_to_msgstr.items():
        po_file_path = _translations_po_path(lang)
        if not os.path.exists(po_file_path):
            continue
        try:
            po = polib.pofile(po_file_path)
            entry = po.find(msgid)
            if entry is None:
                if str(msgstr).strip():
                    po.append(polib.POEntry(msgid=msgid, msgstr=msgstr))
                    updated += 1
            else:
                entry.msgstr = msgstr
                updated += 1
            po.save(po_file_path)
        except Exception as e:
            current_app.logger.error("Failed to update translation for %s: %s", lang, e)
    return updated


def _decode_translation_payload(data):
    """Decode a base64-wrapped translation payload if present.

    Returns (decoded_data, error_response):
      - No payload field -> (data, None)  — original data unchanged
      - Decoded successfully -> (_JsonFormProxy(decoded), None)
      - Decode failed -> (None, json_bad_request response)
    """
    try:
        payload = data.get("payload") or data.get("payload_b64")
    except Exception:
        payload = None

    if not payload:
        return data, None

    try:
        from app.utils.request_utils import _JsonFormProxy
        import base64 as _b64

        decoded = _b64.b64decode(str(payload)).decode("utf-8")
        decoded_obj = json.loads(decoded)
        if not isinstance(decoded_obj, dict):
            return None, json_bad_request(_("Invalid translation payload"))
        return _JsonFormProxy(decoded_obj), None
    except Exception:
        return None, json_bad_request(_("Invalid translation payload"))


@bp.route("/translations/manage", methods=["GET"])
@permission_required("admin.translations.manage")
def manage_translations():
    """Manage translations for the application"""
    from config import Config
    # Use the dynamically loaded languages from app config, not static Config
    languages = current_app.config.get('SUPPORTED_LANGUAGES', Config.LANGUAGES)
    language_names = Config.LANGUAGE_DISPLAY_NAMES

    # Get all translation files (robustly via polib)
    translation_data = {}
    all_msgids = set()
    obsolete_msgids = set()   # msgids marked #~ in PO files
    msgid_sources = {}

    # Try to import polib, but make it optional
    try:
        import polib  # type: ignore
    except ImportError:
        current_app.logger.warning("polib not available - translation file management will be limited")
        polib = None
        # Show a flash message to the user about the missing dependency
        flash(_("Warning: polib package not available. Translation file management will be limited. Install with: pip install polib"), "warning")

    # Build msgid_sources from the POT file first (authoritative source of #: references).
    # PO files may lack occurrences for manually-added entries or stale sync states.
    if polib:
        pot_path = _translations_pot_path()
        if os.path.exists(pot_path):
            with suppress(Exception):
                pot = polib.pofile(pot_path)
                for entry in pot:
                    if not entry.msgid or entry.obsolete:
                        continue
                    if entry.occurrences and entry.msgid not in msgid_sources:
                        src_path, _ = entry.occurrences[0]
                        msgid_sources[entry.msgid] = _extract_page_name(src_path)

    for lang in languages:
        po_file_path = _translations_po_path(lang)
        translations = {}
        if os.path.exists(po_file_path) and polib:
            with suppress(Exception):
                po = polib.pofile(po_file_path)
                for entry in po:
                    if not entry.msgid or entry.msgid == "":
                        continue

                    # Collect obsolete entries separately so they display in the grid
                    # with a "removed" indicator rather than being silently hidden.
                    if entry.obsolete:
                        obsolete_msgids.add(entry.msgid)
                        if entry.msgid not in msgid_sources:
                            # Prefer the tagged translator comment "[Removed] was: <ref>"
                            # (stored in tcomment = plain "# " lines by polib)
                            raw_tcomment = getattr(entry, 'tcomment', '') or ''
                            if '[Removed]' in raw_tcomment and 'was:' in raw_tcomment:
                                ref = raw_tcomment.split('was:', 1)[-1].strip()
                                page_name = _extract_page_name(ref)
                            elif entry.occurrences:
                                src_path, _ = entry.occurrences[0]
                                page_name = _extract_page_name(src_path)
                            else:
                                page_name = None
                            if page_name:
                                msgid_sources[entry.msgid] = '\x00' + page_name  # \x00 prefix = removed
                        # Carry existing translation so it's still visible in the grid
                        if entry.msgid not in translations:
                            if entry.msgstr:
                                translations[entry.msgid] = entry.msgstr
                            elif hasattr(entry, 'msgstr_plural') and entry.msgstr_plural:
                                first_plural = (entry.msgstr_plural.get(0, '')
                                               or next(iter(entry.msgstr_plural.values()), ''))
                                translations[entry.msgid] = first_plural
                            else:
                                translations[entry.msgid] = ''
                        continue

                    # Skip PO file metadata entries
                    msgid_lower = entry.msgid.lower()
                    metadata_keys = [
                        'project-id-version', 'report-msgid-bugs-to', 'pot-creation-date',
                        'po-revision-date', 'last-translator', 'language-team', 'mime-version',
                        'content-type', 'content-transfer-encoding', 'plural-forms', 'generated-by'
                    ]
                    if sum(1 for k in metadata_keys if k in msgid_lower) >= 3:
                        continue

                    all_msgids.add(entry.msgid)

                    if entry.msgstr:
                        translations[entry.msgid] = entry.msgstr
                    elif hasattr(entry, 'msgstr_plural') and entry.msgstr_plural:
                        first_plural = entry.msgstr_plural.get(0, '') or (list(entry.msgstr_plural.values())[0] if entry.msgstr_plural else '')
                        translations[entry.msgid] = first_plural
                    else:
                        # Guard: a duplicate empty entry (e.g. the original plural entry from
                        # pybabel extraction appearing after an auto-translate non-plural entry)
                        # must never overwrite a translation already found earlier in this file.
                        existing = translations.get(entry.msgid, '')
                        if not existing:
                            translations[entry.msgid] = ""

                    # Fallback: capture source from PO if POT didn't have it
                    if entry.occurrences and entry.msgid not in msgid_sources:
                        src_path, _ = entry.occurrences[0]
                        msgid_sources[entry.msgid] = _extract_page_name(src_path)
        translation_data[lang] = {
            'name': language_names.get(lang, lang.upper()),
            'translations': translations
        }

    # Merge obsolete msgids into the full list so they appear in the grid.
    # Active entries always take precedence over obsolete ones with the same msgid.
    all_msgids.update(obsolete_msgids - all_msgids)

    # Sort all message IDs for consistent display
    all_msgids = sorted(list(all_msgids))
    # Active-only msgids for bulk auto-translate (exclude removed/obsolete entries)
    active_translation_msgids = [m for m in all_msgids if m not in obsolete_msgids]

    # Count empty translations for each language (for frontend auto-translate)
    empty_translation_counts = {}
    empty_translation_msgids = {}  # Store the actual msgids that need translation
    for lang in languages:
        if lang == 'en':  # Skip English as it's the source language
            continue
        empty_count = 0
        empty_msgids = []
        lang_translations = translation_data.get(lang, {}).get('translations', {})
        po_file_path = _translations_po_path(lang)
        po_cache = None
        if polib and os.path.exists(po_file_path):
            with suppress(Exception):
                po_cache = polib.pofile(po_file_path)

        for msgid in all_msgids:
            # Obsolete (#~) strings are kept in the grid for visibility but should not
            # count as "missing" for auto-translate or per-locale empty tallies.
            if msgid in obsolete_msgids:
                continue
            translation = lang_translations.get(msgid, '').strip()

            # If translation appears empty, check if it's a plural form in the PO file
            if not translation and po_cache:
                with suppress(Exception):  # If we can't find the entry, fall back to checking the dict
                    entry = po_cache.find(msgid)
                    if entry and hasattr(entry, 'msgstr_plural') and entry.msgstr_plural:
                        # Check if any plural form has a non-empty translation
                        has_plural_translation = any(
                            v and str(v).strip()
                            for v in entry.msgstr_plural.values()
                        )
                        if has_plural_translation:
                            translation = "PLURAL_FORM"  # Mark as having translation

            if not translation:  # Count empty or whitespace-only translations
                empty_count += 1
                empty_msgids.append(msgid)
        empty_translation_counts[lang] = empty_count
        empty_translation_msgids[lang] = empty_msgids

    # Add variables needed by JavaScript auto-translate functionality
    from config import Config
    try:
        all_languages = current_app.config.get('SUPPORTED_LANGUAGES', Config.LANGUAGES)
        # Exclude English from translatable languages since it's the source language
        translatable_languages = [lang for lang in all_languages if lang != 'en']
    except Exception as e:
        logger.debug("SUPPORTED_LANGUAGES config lookup failed: %s", e)
        all_languages = getattr(Config, 'LANGUAGES', ['en'])
        # Exclude English from translatable languages since it's the source language
        translatable_languages = [lang for lang in (all_languages or []) if lang != 'en']

    # Return JSON for API requests (mobile app)
    if is_json_request():
        translations_list = []
        for msgid in all_msgids:
            translation_entry = {
                'msgid': msgid,
                'source': msgid_sources.get(msgid, 'unknown'),
                'translations': {}
            }
            for lang in languages:
                lang_data = translation_data.get(lang, {})
                lang_translations = lang_data.get('translations', {})
                translation_entry['translations'][lang] = {
                    'text': lang_translations.get(msgid, ''),
                    'language_name': lang_data.get('name', lang.upper())
                }
            translations_list.append(translation_entry)

        return json_ok(
            translations=translations_list,
            count=len(translations_list),
            languages=[{'code': lang, 'name': language_names.get(lang, lang.upper())} for lang in languages],
            empty_translation_counts=empty_translation_counts,
            polib_available=polib is not None,
        )

    return render_template('admin/translations/manage_translations.html',
                         translation_data=translation_data,
                         all_msgids=all_msgids,
                         active_translation_msgids=active_translation_msgids,
                         obsolete_msgids=obsolete_msgids,
                         languages=languages,
                         language_names=language_names,
                         msgid_sources=msgid_sources,
                         TRANSLATABLE_LANGUAGES=translatable_languages,
                         empty_translation_counts=empty_translation_counts,
                         empty_translation_msgids=empty_translation_msgids,
                         polib_available=polib is not None)


@bp.route("/translations/export", methods=["GET"])
@permission_required("admin.translations.manage")
def export_translations():
    """
    Export translations in a safe, round-trippable format.

    Supported query params:
      - format: xlsx | po | po-zip
      - lang: required when format=po
    """
    from config import Config
    languages = current_app.config.get("SUPPORTED_LANGUAGES", Config.LANGUAGES)
    fmt = (request.args.get("format") or "xlsx").strip().lower()

    try:
        import polib  # type: ignore
    except ImportError:
        flash(_("polib is not installed; export is unavailable."), "danger")
        return redirect(url_for("utilities.manage_translations"))

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    if fmt == "po":
        lang = (request.args.get("lang") or "").strip().lower()
        if not lang:
            flash(_("Please select a language to export."), "warning")
            return redirect(url_for("utilities.manage_translations"))
        if lang not in languages:
            flash(_("Unknown/disabled language: %(lang)s", lang=lang), "danger")
            return redirect(url_for("utilities.manage_translations"))

        po_path = _translations_po_path(lang)
        if not os.path.exists(po_path):
            flash(_("PO file not found for %(lang)s.", lang=lang), "danger")
            return redirect(url_for("utilities.manage_translations"))

        return send_file(
            po_path,
            as_attachment=True,
            download_name=f"messages_{lang}_{timestamp}.po",
            mimetype="text/x-gettext-translation",
        )

    if fmt == "po-zip":
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for lang in languages:
                po_path = _translations_po_path(lang)
                if not os.path.exists(po_path):
                    continue
                arcname = f"{lang}/LC_MESSAGES/messages.po"
                zf.write(po_path, arcname=arcname)
        buf.seek(0)
        return send_file(
            buf,
            as_attachment=True,
            download_name=f"translations_po_{timestamp}.zip",
            mimetype="application/zip",
        )

    # Default: XLSX export (all languages in one workbook)
    # Source of truth for msgids is the POT when available.
    pot_path = _translations_pot_path()
    msgids: list[str] = []
    if os.path.exists(pot_path):
        with suppress(Exception):
            pot = polib.pofile(pot_path)
            for entry in pot:
                if getattr(entry, "obsolete", False):
                    continue
                if entry.msgid:
                    msgids.append(entry.msgid)

    if not msgids:
        # Fallback: union of msgids from existing PO files
        msgid_set = set()
        for lang in languages:
            po_path = _translations_po_path(lang)
            if not os.path.exists(po_path):
                continue
            with suppress(Exception):
                po = polib.pofile(po_path)
                for entry in po:
                    if getattr(entry, "obsolete", False):
                        continue
                    if entry.msgid:
                        msgid_set.add(entry.msgid)
        msgids = sorted(msgid_set)

    # Preload PO files per language for efficient lookup
    po_maps: dict[str, dict[str, object]] = {}
    for lang in languages:
        po_path = _translations_po_path(lang)
        entries: dict[str, object] = {}
        if os.path.exists(po_path):
            with suppress(Exception):
                po = polib.pofile(po_path)
                for entry in po:
                    if getattr(entry, "obsolete", False):
                        continue
                    if not entry.msgid:
                        continue
                    # Note: msgctxt is ignored here (current UI doesn't surface it).
                    entries[entry.msgid] = entry
        po_maps[lang] = entries

    wb = Workbook()
    ws = wb.active
    ws.title = "translations"

    # Add warning note above headers
    ws.merge_cells("A1:B1")
    note_cell = ws.cell(row=1, column=1)
    note_cell.value = "⚠️ DO NOT MODIFY: Source and msgid columns are read-only. Only edit translation columns."
    note_cell.font = Font(bold=True, color="DC2626")  # red-600
    note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    note_cell.fill = PatternFill("solid", fgColor="FEE2E2")  # red-100 background
    ws.row_dimensions[1].height = 30

    # Include a source column to help translators understand where a string comes from.
    # Order: source, msgid, then languages
    header = ["source", "msgid"] + list(languages)
    ws.append(header)
    # Freeze header row and source + msgid columns
    ws.freeze_panes = "C3"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E5E7EB")  # gray-200
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=2, column=c)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
    ws.row_dimensions[2].height = 24

    # Reasonable default widths
    ws.column_dimensions["A"].width = 28  # source
    ws.column_dimensions["B"].width = 70  # msgid
    for idx, lang in enumerate(languages, start=3):
        ws.column_dimensions[get_column_letter(idx)].width = 55

    # Build msgid -> source mapping (prefer POT occurrences; fall back to PO occurrences)
    msgid_sources: dict[str, str] = {}
    if os.path.exists(pot_path):
        with suppress(Exception):
            pot = polib.pofile(pot_path)
            for entry in pot:
                if getattr(entry, "obsolete", False):
                    continue
                if not entry.msgid:
                    continue
                if entry.occurrences:
                    src_path, _ = entry.occurrences[0]
                    msgid_sources[entry.msgid] = _extract_page_name(src_path)

    wrap = Alignment(vertical="top", wrap_text=True)
    for msgid in msgids:
        src = msgid_sources.get(msgid, "")
        if not src:
            # Fall back to any locale's PO entry occurrences
            for lang in languages:
                entry = po_maps.get(lang, {}).get(msgid)
                if entry is not None and getattr(entry, "occurrences", None):
                    with suppress(Exception):
                        src_path, _ = entry.occurrences[0]
                        src = _extract_page_name(src_path)
                        break

        row = [src, msgid]  # source first, then msgid
        for lang in languages:
            entry = po_maps.get(lang, {}).get(msgid)
            row.append(_entry_to_display_msgstr(entry) if entry is not None else "")
        ws.append(row)

    # Apply wrap alignment to all data rows (starting from row 3: note=1, header=2, data=3+)
    for r in range(3, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = wrap
            # Sheet will be protected; lock only source/msgid columns, keep translations editable
            ws.cell(row=r, column=c).protection = Protection(locked=(c in (1, 2)))

    # Protect the sheet but allow editing unlocked (translation) cells.
    #
    # Important: in OOXML, most `sheetProtection` boolean flags are *locks*:
    # setting them to True generally DISALLOWS that action (e.g., `selectUnlockedCells="1"` blocks selecting unlocked cells).
    # So we keep these flags False (defaults) and only enable sheet protection.
    with suppress(Exception):
        ws.protection.selectLockedCells = False
    with suppress(Exception):
        ws.protection.selectUnlockedCells = False
    with suppress(Exception):
        ws.protection.autoFilter = False
    with suppress(Exception):
        ws.protection.sort = False
    with suppress(Exception):
        ws.protection.formatColumns = False
    with suppress(Exception):
        ws.protection.formatRows = False
    # Enable protection last
    with suppress(Exception):
        ws.protection.sheet = True

    # Add thin grey borders for readability (including note row)
    try:
        border_color = "D1D5DB"  # gray-300
        thin = Side(style="thin", color=border_color)
        table_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).border = table_border
    except Exception as e:
        logger.debug("translation export border apply failed: %s", e)

    # Apply visual formatting to source and msgid columns (gray background) for clarity
    try:
        read_only_fill = PatternFill("solid", fgColor="F3F4F6")  # gray-100
        for r in range(3, ws.max_row + 1):
            ws.cell(row=r, column=1).fill = read_only_fill  # source column
            ws.cell(row=r, column=2).fill = read_only_fill  # msgid column
    except Exception as e:
        logger.debug("translation export fill apply failed: %s", e)

    # Enable filtering on all columns (starting from header row 2)
    try:
        ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    except Exception as e:
        logger.debug("translation export auto_filter failed: %s", e)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        as_attachment=True,
        download_name=f"translations_{timestamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/translations/import", methods=["POST"])
@permission_required("admin.translations.manage")
def import_translations():
    """
    Import translations safely:
    - Never changes msgid
    - Never creates new msgids in PO
    - By default only updates non-empty incoming translations (won't clear existing)
    """
    from config import Config
    languages = current_app.config.get("SUPPORTED_LANGUAGES", Config.LANGUAGES)

    fmt = (request.form.get("format") or request.form.get("import_format") or "").strip().lower()
    if not fmt:
        fmt = "xlsx"

    only_non_empty = (request.form.get("only_non_empty") or "").strip().lower() in ("1", "true", "on", "yes")
    allow_clear = (request.form.get("allow_clear") or "").strip().lower() in ("1", "true", "on", "yes")
    if only_non_empty:
        allow_clear = False  # strongest safety default

    file = request.files.get("file")
    if file is None or not getattr(file, "filename", ""):
        flash(_("No file selected."), "warning")
        return redirect(url_for("utilities.manage_translations"))

    try:
        import polib  # type: ignore
    except ImportError:
        flash(_("polib is not installed; import is unavailable."), "danger")
        return redirect(url_for("utilities.manage_translations"))

    def _backup(po_path: str) -> None:
        try:
            ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            backup_path = f"{po_path}.bak.{ts}"
            shutil.copy2(po_path, backup_path)
        except Exception as e:
            logger.debug("translation import backup failed: %s", e)

    def _apply_incoming_po(current, incoming):
        """
        Apply incoming translations to an existing PO object safely.
        Returns: (updated, skipped_missing, skipped_empty)
        """
        current_map = {}
        for e in current:
            if getattr(e, "obsolete", False):
                continue
            if not e.msgid:
                continue
            # Include msgctxt in the key if present (safer than msgid-only).
            current_map[(e.msgctxt or None, e.msgid)] = e

        updated = 0
        skipped_missing = 0
        skipped_empty = 0

        for inc in incoming:
            if getattr(inc, "obsolete", False):
                continue
            if not inc.msgid:
                continue

            key = (getattr(inc, "msgctxt", None) or None, inc.msgid)
            cur = current_map.get(key)
            if cur is None:
                skipped_missing += 1
                continue

            # Decide the incoming translation value(s)
            inc_singular = (inc.msgstr or "") if hasattr(inc, "msgstr") else ""
            inc_plural = getattr(inc, "msgstr_plural", None) or {}

            # If we only import non-empty, skip blanks (never clear existing translations).
            if only_non_empty:
                if (not inc_singular or not str(inc_singular).strip()) and (not inc_plural or all(not str(v).strip() for v in inc_plural.values())):
                    skipped_empty += 1
                    continue
            elif not allow_clear:
                # If not allowing clear, also skip empty updates.
                if (not inc_singular or not str(inc_singular).strip()) and (not inc_plural or all(not str(v).strip() for v in inc_plural.values())):
                    skipped_empty += 1
                    continue

            # Update only msgstr / msgstr_plural (never msgid)
            if getattr(cur, "msgid_plural", None):
                # Prefer plural updates if provided; otherwise update index 0 only.
                if inc_plural:
                    for k, v in inc_plural.items():
                        if (only_non_empty or (not allow_clear)) and (v is None or not str(v).strip()):
                            continue
                        cur.msgstr_plural[int(k)] = "" if (v is None) else str(v)
                    updated += 1
                else:
                    if allow_clear or (inc_singular and str(inc_singular).strip()) or (not only_non_empty and inc_singular is not None):
                        cur.msgstr_plural[0] = "" if (inc_singular is None) else str(inc_singular)
                        updated += 1
            else:
                if allow_clear or (inc_singular and str(inc_singular).strip()) or (not only_non_empty and inc_singular is not None):
                    cur.msgstr = "" if (inc_singular is None) else str(inc_singular)
                    updated += 1

        return updated, skipped_missing, skipped_empty

    if fmt == "po-zip":
        # Import a ZIP containing multiple locales: <lang>/LC_MESSAGES/messages.po
        if not file.filename.lower().endswith(".zip"):
            flash(_("Please upload a .zip file for PO ZIP import."), "warning")
            return redirect(url_for("utilities.manage_translations"))

        temp_zip_path = None
        temp_dir = None
        try:
            filename = secure_filename(file.filename)
            from app.utils.file_paths import get_temp_upload_path

            temp_dir = get_temp_upload_path()
            os.makedirs(temp_dir, exist_ok=True)
            temp_zip_path = os.path.join(temp_dir, filename)
            file.save(temp_zip_path)

            processed_locales = 0
            skipped_locales = 0
            total_updated = 0
            total_missing = 0
            total_empty = 0

            per_lang = {}

            with zipfile.ZipFile(temp_zip_path, "r") as zf:
                for name in zf.namelist():
                    if not name or name.endswith("/"):
                        continue
                    # Accept our export layout and similar layouts
                    if not name.replace("\\", "/").endswith("/LC_MESSAGES/messages.po"):
                        continue

                    parts = name.replace("\\", "/").split("/")
                    # Find the locale segment immediately before LC_MESSAGES
                    try:
                        i = parts.index("LC_MESSAGES")
                        lang = parts[i - 1].strip().lower() if i > 0 else ""
                    except ValueError:
                        continue

                    if not lang:
                        skipped_locales += 1
                        continue
                    if lang == "en":
                        # English is the source language; skip
                        skipped_locales += 1
                        continue
                    if lang not in languages:
                        skipped_locales += 1
                        continue

                    po_path = _translations_po_path(lang)
                    if not os.path.exists(po_path):
                        skipped_locales += 1
                        continue

                    # Write member to a temp .po file for polib
                    member_bytes = zf.read(name)
                    temp_po_path = os.path.join(temp_dir, f"_import_{lang}_messages.po")
                    with open(temp_po_path, "wb") as f:
                        f.write(member_bytes)

                    incoming = polib.pofile(temp_po_path)
                    current = polib.pofile(po_path)

                    updated, skipped_missing, skipped_empty = _apply_incoming_po(current, incoming)
                    per_lang[lang] = {
                        "updated": updated,
                        "skipped_missing": skipped_missing,
                        "skipped_empty": skipped_empty,
                    }

                    if updated > 0:
                        _backup(po_path)
                        current.save(po_path)

                    processed_locales += 1
                    total_updated += updated
                    total_missing += skipped_missing
                    total_empty += skipped_empty

                    with suppress(Exception):
                        os.remove(temp_po_path)

            if processed_locales == 0:
                flash(
                    _("No valid locale PO files found in ZIP. Expected files like: <lang>/LC_MESSAGES/messages.po"),
                    "warning",
                )
                return redirect(url_for("utilities.manage_translations"))

            flash(
                _(
                    "PO ZIP import finished. Locales processed: %(p)d, skipped: %(s)d. Updated: %(u)d, skipped missing: %(m)d, skipped empty: %(e)d.",
                    p=processed_locales,
                    s=skipped_locales,
                    u=total_updated,
                    m=total_missing,
                    e=total_empty,
                ),
                "success",
            )
        except Exception as e:
            current_app.logger.error(f"Translation PO ZIP import failed: {e}", exc_info=True)
            flash(_("An error occurred. Please try again."), "danger")
        finally:
            if temp_zip_path and os.path.exists(temp_zip_path):
                with suppress(Exception):
                    os.remove(temp_zip_path)

        return redirect(url_for("utilities.manage_translations"))

    if fmt == "po":
        lang = (request.form.get("lang") or "").strip().lower()
        if not lang:
            flash(_("Please select a language for PO import."), "warning")
            return redirect(url_for("utilities.manage_translations"))
        if lang not in languages:
            flash(_("Unknown/disabled language: %(lang)s", lang=lang), "danger")
            return redirect(url_for("utilities.manage_translations"))
        if lang == "en":
            flash(_("English (en) is the source language and should not be imported."), "warning")
            return redirect(url_for("utilities.manage_translations"))

        if not file.filename.lower().endswith(".po"):
            flash(_("Please upload a .po file for PO import."), "warning")
            return redirect(url_for("utilities.manage_translations"))

        po_path = _translations_po_path(lang)
        if not os.path.exists(po_path):
            flash(_("PO file not found for %(lang)s.", lang=lang), "danger")
            return redirect(url_for("utilities.manage_translations"))

        # Save to a temp path for polib
        temp_path = None
        try:
            filename = secure_filename(file.filename)
            from app.utils.file_paths import get_temp_upload_path

            temp_dir = get_temp_upload_path()
            os.makedirs(temp_dir, exist_ok=True)
            temp_path = os.path.join(temp_dir, filename)
            file.save(temp_path)

            incoming = polib.pofile(temp_path)
            current = polib.pofile(po_path)

            updated, skipped_missing, skipped_empty = _apply_incoming_po(current, incoming)

            if updated > 0:
                _backup(po_path)
                current.save(po_path)

            flash(
                _(
                    "PO import finished for %(lang)s. Updated: %(u)d, skipped missing: %(m)d, skipped empty: %(e)d.",
                    lang=lang,
                    u=updated,
                    m=skipped_missing,
                    e=skipped_empty,
                ),
                "success",
            )
        except Exception as e:
            current_app.logger.error(f"Translation PO import failed: {e}", exc_info=True)
            flash(_("An error occurred. Please try again."), "danger")
        finally:
            if temp_path and os.path.exists(temp_path):
                with suppress(Exception):
                    os.remove(temp_path)

        return redirect(url_for("utilities.manage_translations"))

    # XLSX import (all languages in one workbook)
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        flash(_("Please upload an Excel file (.xlsx or .xls)."), "warning")
        return redirect(url_for("utilities.manage_translations"))

    try:
        file.stream.seek(0)
    except Exception as e:
        logger.debug("file.stream.seek failed: %s", e)

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb["translations"] if "translations" in wb.sheetnames else wb.active

        # Detect header row: check if row 1 has "msgid", otherwise try row 2 (new format has note in row 1)
        header_row_num = 1
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row:
            headers = [str(h).strip() if h is not None else "" for h in header_row]
            norm_headers = [h.strip().lower() for h in headers]
            if "msgid" not in norm_headers:
                # Try row 2 (new format with note row)
                header_row_num = 2
                header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)

        if not header_row:
            flash(_("Excel file is empty or header row not found."), "warning")
            return redirect(url_for("utilities.manage_translations"))

        headers = [str(h).strip() if h is not None else "" for h in header_row]
        norm_headers = [h.strip().lower() for h in headers]
        if "msgid" not in norm_headers:
            flash(_("Excel import requires a 'msgid' column header."), "danger")
            return redirect(url_for("utilities.manage_translations"))

        msgid_col_idx = norm_headers.index("msgid")

        # Map language columns by header exact match (lang code)
        lang_cols: dict[str, int] = {}
        for idx, h in enumerate(norm_headers):
            if h in languages:
                lang_cols[h] = idx

        if not lang_cols:
            flash(_("No language columns found in Excel header."), "danger")
            return redirect(url_for("utilities.manage_translations"))

        # Collect updates per language
        updates_by_lang: dict[str, dict[str, str]] = {lang: {} for lang in lang_cols.keys() if lang != "en"}

        # Data rows start after header (row 2 if header is row 1, row 3 if header is row 2)
        data_start_row = header_row_num + 1
        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if not row:
                continue
            raw_msgid = row[msgid_col_idx] if msgid_col_idx < len(row) else None
            msgid = (str(raw_msgid) if raw_msgid is not None else "").strip()
            if not msgid:
                continue

            for lang, col_idx in lang_cols.items():
                if lang == "en":
                    continue
                if col_idx >= len(row):
                    continue
                v = row[col_idx]
                if v is None:
                    continue
                s = str(v)
                if only_non_empty and not s.strip():
                    continue
                if (not allow_clear) and (not s.strip()):
                    continue
                updates_by_lang.setdefault(lang, {})[msgid] = s

        total_updated = 0
        total_missing = 0

        for lang, updates in updates_by_lang.items():
            if not updates:
                continue
            po_path = _translations_po_path(lang)
            if not os.path.exists(po_path):
                continue

            po = polib.pofile(po_path)
            entry_map = {}
            for e in po:
                if getattr(e, "obsolete", False):
                    continue
                if not e.msgid:
                    continue
                entry_map[(e.msgctxt or None, e.msgid)] = e

            updated_lang = 0
            missing_lang = 0
            for msgid, msgstr in updates.items():
                cur = entry_map.get((None, msgid))
                if cur is None:
                    missing_lang += 1
                    continue
                if getattr(cur, "msgid_plural", None):
                    cur.msgstr_plural[0] = "" if msgstr is None else str(msgstr)
                else:
                    cur.msgstr = "" if msgstr is None else str(msgstr)
                updated_lang += 1

            if updated_lang > 0:
                _backup(po_path)
                po.save(po_path)

            total_updated += updated_lang
            total_missing += missing_lang

        flash(
            _(
                "Excel import finished. Updated: %(u)d, skipped missing msgids: %(m)d. (Empty updates are ignored by default.)",
                u=total_updated,
                m=total_missing,
            ),
            "success",
        )
    except Exception as e:
        current_app.logger.error(f"Translation Excel import failed: {e}", exc_info=True)
        flash(_("An error occurred. Please try again."), "danger")

    return redirect(url_for("utilities.manage_translations"))

@bp.route("/translations/add", methods=["GET", "POST"])
@permission_required("admin.translations.manage")
def add_translation():
    """Add a new translation for all languages"""
    form = TranslationForm()

    if form.validate_on_submit():
        msgid = form.msgid.data

        from config import Config
        languages = current_app.config.get('SUPPORTED_LANGUAGES', Config.LANGUAGES)

        lang_to_msgstr = {}
        for lang in languages:
            msgstr_field = getattr(form, f'msgstr_{lang}', None)
            msgstr = (msgstr_field.data if msgstr_field else None)
            if msgstr and str(msgstr).strip():
                lang_to_msgstr[lang] = msgstr
        added_count = _update_po_translations(msgid, lang_to_msgstr)

        if added_count > 0:
            flash(_('Translation added successfully for %(count)d language(s)', count=added_count), 'success')
        else:
            flash(_('No translations were added. Please provide at least one translation.'), 'warning')

        return redirect(url_for('utilities.manage_translations'))

    # Pass language metadata to the template for client-side auto-translate
    try:
        from config import Config
        languages = current_app.config.get('SUPPORTED_LANGUAGES', Config.LANGUAGES)
        # ISO codes only for client-side auto-translate (no long keys)
        language_model_keys = {code: (None if code == 'en' else code) for code in languages}
        language_display_names = Config.LANGUAGE_DISPLAY_NAMES
    except Exception as e:
        logger.debug("SUPPORTED_LANGUAGES/LANGUAGE_DISPLAY_NAMES lookup failed: %s", e)
        languages = getattr(Config, 'LANGUAGES', ['en'])
        # ISO codes only
        language_model_keys = {code: (None if code == 'en' else code) for code in (languages or [])}
        language_display_names = getattr(Config, 'LANGUAGE_DISPLAY_NAMES', {}) or getattr(Config, 'ALL_LANGUAGES_DISPLAY_NAMES', {}) or {}

    return render_template(
        'admin/translations/edit_translation.html',
        form=form,
        languages=languages,
        language_model_keys=language_model_keys,
        language_display_names=language_display_names,
        is_add=True
    )

@bp.route("/translations/edit", methods=["GET", "POST"])
@permission_required("admin.translations.manage")
def edit_translation():
    """Edit translations for all languages"""
    is_ajax = is_json_request()

    # --- JSON POST path (supports base64-wrapped payload to avoid WAF false positives) ---
    if request.method == "POST" and request.is_json:
        data = get_request_data()
        data, err = _decode_translation_payload(data)
        if err:
            return err

        msgid = (data.get('msgid') or '').strip()
        if not msgid:
            return json_bad_request(_("msgid is required"))

        from config import Config
        languages = current_app.config.get('SUPPORTED_LANGUAGES', Config.LANGUAGES)

        lang_to_msgstr = {}
        for lang in languages:
            msgstr = data.get(f'msgstr_{lang}')
            if msgstr is not None:
                lang_to_msgstr[lang] = msgstr
        updated_count = _update_po_translations(msgid, lang_to_msgstr)

        if updated_count > 0:
            return json_ok(
                message=_('Translation updated successfully for %(count)d language(s)', count=updated_count),
                updated_count=updated_count,
            )
        return json_ok(
            message=_('No translations were updated'),
            updated_count=0,
        )

    # --- Legacy WTForms POST path (non-JSON form submissions) ---
    form = TranslationForm()

    if form.validate_on_submit():
        msgid = form.msgid.data

        from config import Config
        languages = current_app.config.get('SUPPORTED_LANGUAGES', Config.LANGUAGES)

        lang_to_msgstr = {}
        for lang in languages:
            msgstr_field = getattr(form, f'msgstr_{lang}', None)
            msgstr = (msgstr_field.data if msgstr_field else None)
            if msgstr is not None:
                lang_to_msgstr[lang] = msgstr
        updated_count = _update_po_translations(msgid, lang_to_msgstr)

        if is_ajax:
            if updated_count > 0:
                return json_ok(
                    message=_('Translation updated successfully for %(count)d language(s)', count=updated_count),
                    updated_count=updated_count,
                )
            else:
                return json_ok(
                    message=_('No translations were updated'),
                    updated_count=0,
                )
        else:
            if updated_count > 0:
                flash(_('Translation updated successfully for %(count)d language(s)', count=updated_count), 'success')
            else:
                flash(_('No translations were updated'), 'warning')
            return redirect(url_for('utilities.manage_translations'))

    # --- GET path: support base64-encoded msgid to avoid WAF triggers ---
    import base64 as _b64
    msgid_b64 = request.args.get('msgid_b64')
    if msgid_b64:
        try:
            msgid = _b64.b64decode(msgid_b64).decode('utf-8')
        except Exception:
            msgid = None
    else:
        msgid = request.args.get('msgid')

    from config import Config
    languages = current_app.config.get('SUPPORTED_LANGUAGES', Config.LANGUAGES)
    translations = {}

    if msgid:
        form.msgid.data = msgid

        try:
            import polib  # type: ignore
        except ImportError:
            current_app.logger.warning("polib not available - cannot pre-populate translations")
            polib = None

        if polib is not None:
            for lang in languages:
                po_file_path = _translations_po_path(lang)
                if os.path.exists(po_file_path) and polib:
                    try:
                        po = polib.pofile(po_file_path)
                        entry = po.find(msgid)
                        if entry is not None:
                            translations[lang] = entry.msgstr or ''
                            msgstr_field = getattr(form, f'msgstr_{lang}', None)
                            if msgstr_field is not None:
                                msgstr_field.data = entry.msgstr or ''
                    except Exception as _e:
                        current_app.logger.error(f"Failed to read translation for {lang}: {_e}")

    if is_ajax:
        try:
            from config import Config
            language_display_names = Config.LANGUAGE_DISPLAY_NAMES
        except Exception as e:
            logger.debug("Config.LANGUAGE_DISPLAY_NAMES lookup failed: %s", e)
            language_display_names = getattr(Config, 'LANGUAGE_DISPLAY_NAMES', {}) or getattr(Config, 'ALL_LANGUAGES_DISPLAY_NAMES', {}) or {}

        return json_ok(
            msgid=msgid or '',
            translations=translations,
            languages=languages,
            language_display_names=language_display_names,
        )

    return redirect(url_for('utilities.manage_translations'))


@bp.route("/translations/delete-removed", methods=["POST"])
@permission_required("admin.translations.manage")
def delete_removed_translation():
    """Remove obsolete (#~) PO entries for a msgid from all locale files (active entries are never deleted)."""
    if not request.is_json:
        return json_bad_request(_("Expected JSON body"))

    data = get_request_data()
    data, err = _decode_translation_payload(data)
    if err:
        return err

    msgid = (data.get("msgid") or "").strip()
    if not msgid:
        return json_bad_request(_("msgid is required"))

    try:
        import polib  # type: ignore
    except ImportError:
        return json_server_error(_("Translation tools are not available (polib missing)."))

    from config import Config
    languages = current_app.config.get("SUPPORTED_LANGUAGES", Config.LANGUAGES)

    files_updated = 0
    entries_removed = 0
    file_errors = []

    for lang in languages:
        po_file_path = _translations_po_path(lang)
        if not os.path.exists(po_file_path):
            continue
        try:
            po = polib.pofile(po_file_path)
            removed_here = 0
            for entry in list(po):
                if entry.msgid == msgid and getattr(entry, "obsolete", False):
                    po.remove(entry)
                    removed_here += 1
            if removed_here:
                po.save(po_file_path)
                files_updated += 1
                entries_removed += removed_here
        except Exception as ex:
            logger.warning("delete_removed_translation failed for %s: %s", po_file_path, ex)
            file_errors.append(lang)

    if entries_removed == 0:
        if file_errors:
            return json_server_error(_("Could not update translation files."))
        return json_bad_request(
            _("No removed (obsolete) entries found for this message. Try refreshing the page."),
        )

    message = _("Removed obsolete translation entries from %(count)d file(s)", count=files_updated)
    if file_errors:
        message = message + " " + _(
            "Some languages could not be updated: %(langs)s",
            langs=", ".join(file_errors),
        )

    return json_ok(
        message=message,
        files_updated=files_updated,
        entries_removed=entries_removed,
        partial_errors=file_errors or None,
    )


@bp.route("/translations/compile", methods=["POST"])
@permission_required("admin.translations.manage")
def compile_translations():
    """Compile translations and optionally restart the application"""
    try:
        # Run the compilation script
        import subprocess
        import os
        # Get the backoffice root directory (parent of app/)
        backoffice_root = os.path.abspath(os.path.join(current_app.root_path, '..'))
        scripts_dir = os.path.join(backoffice_root, 'scripts')
        result = subprocess.run(['python', 'compile_translations.py'],
                              capture_output=True, text=True, cwd=scripts_dir)

        if result.returncode == 0:
            # Force reload of translations
            from flask_babel import refresh
            refresh()

            # Check if restart was requested
            restart_requested = request.form.get('restart') == '1'

            if restart_requested:
                # Attempt to restart the application
                try:
                    import os
                    import time

                    # In debug mode with Werkzeug reloader, touch a watched Python module to trigger reload
                    if current_app.debug and os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
                        target_files = [
                            os.path.join(current_app.root_path, '__init__.py'),
                            os.path.join(os.path.dirname(current_app.root_path), 'config', 'config.py'),
                        ]
                        touched = False
                        for f in target_files:
                            if os.path.exists(f):
                                now = time.time()
                                os.utime(f, (now, now))
                                touched = True
                        if touched:
                            flash(_('Translations compiled successfully! Application restart triggered.'), 'success')
                        else:
                            flash(_('Translations compiled successfully! Please restart the application manually to apply changes.'), 'warning')
                    else:
                        # For production, we can't easily restart the application
                        # Instead, we'll try to reload the application context
                        try:
                            # Force reload of all modules by touching the main app file
                            app_file = os.path.join(current_app.root_path, '__init__.py')
                            if os.path.exists(app_file):
                                now = time.time()
                                os.utime(app_file, (now, now))
                            flash(_('Translations compiled successfully! Application context refreshed. Please restart the application via your process manager to apply changes across all workers.'), 'warning')
                        except Exception as e:
                            logger.debug("touch/restart hint failed: %s", e)
                            flash(_('Translations compiled successfully! Please restart the application via your process manager to apply changes.'), 'warning')
                except Exception as restart_error:
                    flash(_('Translations compiled successfully! However, automatic restart failed. Please restart the application manually.'), 'warning')
            else:
                flash(_('Translations compiled and reloaded successfully!'), 'success')
        else:
            flash(_('Error compiling translations: %(error)s', error=result.stderr), 'danger')

    except Exception as e:
        flash(_("An error occurred. Please try again."), "danger")

    return redirect(url_for('utilities.manage_translations'))

@bp.route("/translations/reload", methods=["POST"])
@permission_required("admin.translations.manage")
def reload_translations():
    """Manually reload translations without recompiling"""
    try:
        from flask_babel import refresh
        refresh()
        flash(_('Translations reloaded successfully!'), 'success')
    except Exception as e:
        flash(_("An error occurred. Please try again."), "danger")

    return redirect(url_for('utilities.manage_translations'))

@bp.route("/translations/extract-update", methods=["POST"])
@permission_required("admin.translations.manage")
def extract_update_translations():
    """Extract translatable strings and update PO files"""
    try:
        import subprocess
        import os
        import sys

        # Get the backoffice root directory (parent of app/)
        backoffice_root = os.path.abspath(os.path.join(current_app.root_path, '..'))
        scripts_dir = os.path.join(backoffice_root, 'scripts')
        script_path = os.path.join(scripts_dir, 'extract_update_translations.py')

        if not os.path.exists(script_path):
            flash(_('Extraction script not found. Please ensure scripts/extract_update_translations.py exists.'), 'danger')
            return redirect(url_for('utilities.manage_translations'))

        # Run the extraction script
        # Use sys.executable to ensure we use the same Python interpreter
        result = subprocess.run(
            [sys.executable, 'extract_update_translations.py'],
            capture_output=True,
            text=True,
            cwd=scripts_dir,
            timeout=300  # 5 minute timeout
        )

        if result.returncode == 0:
            # Parse output to get useful information
            output_lines = result.stdout.split('\n')
            obsolete_info = []
            for line in output_lines:
                if 'obsolete' in line.lower() and '/' in line:
                    obsolete_info.append(line.strip())

            # Force reload of translations
            from flask_babel import refresh
            refresh()

            # Build success message
            success_msg = _('Translations extracted and updated successfully!')
            if obsolete_info:
                success_msg += ' ' + ' '.join(obsolete_info[:3])  # Show first 3 locales
                if len(obsolete_info) > 3:
                    success_msg += _(' (and %(count)d more)', count=len(obsolete_info) - 3)

            flash(success_msg, 'success')

            # Show any warnings from stderr (non-fatal)
            if result.stderr and 'warning' in result.stderr.lower():
                flash(_('Note: %(note)s', note=result.stderr[:200]), 'info')
        else:
            error_msg = result.stderr or result.stdout or _('Unknown error')
            flash(_('Error extracting/updating translations: %(error)s', error=error_msg[:500]), 'danger')
            current_app.logger.error(f"Extraction script failed: {result.stderr}")

    except subprocess.TimeoutExpired:
        flash(_('Extraction timed out after 5 minutes. The process may still be running.'), 'warning')
    except Exception as e:
        flash(_("An error occurred. Please try again."), "danger")
        current_app.logger.error(f"Extraction error: {e}", exc_info=True)

    return redirect(url_for('utilities.manage_translations'))


# ========================
# Translation API Endpoints
# ========================

@bp.route('/api/auto-translate', methods=['POST'])
@limiter.exempt
@permission_required_any(
    'admin.templates.edit',
    'admin.templates.create',
    'admin.indicator_bank.create',
    'admin.indicator_bank.edit',
    'admin.resources.manage',
    'admin.translations.manage',
    'admin.organization.manage',
    'admin.settings.manage',
)
def api_auto_translate():
    """API endpoint for automatic translation"""


    try:
        from app.services.translation.auto_translator import (
            get_auto_translator,
            translate_form_item_auto,
            translate_section_name_auto,
            translate_question_option_auto,
            translate_page_name_auto,
            translate_template_name_auto,
            translate_email_template_html_auto,
        )
        from app.services.authorization_service import AuthorizationService

        data = get_json_safe()
        err = require_json_data(data)
        if err:
            return err

        # Support base64-wrapped payloads to avoid WAF false positives on rich strings (HTML, SQL-like markers, etc.)
        # Prefer payload if present; keep plain JSON for backward compatibility.
        try:
            payload = data.get("payload") or data.get("payload_b64")
        except Exception:
            payload = None
        if payload:
            try:
                import base64 as _b64
                decoded = _b64.b64decode(str(payload)).decode("utf-8")
                decoded_obj = json.loads(decoded)
                if isinstance(decoded_obj, dict):
                    data = decoded_obj
            except Exception:
                # Keep original data so we can return a structured JSON error below
                pass

        # --- RBAC enforcement (context-aware) ---
        # IMPORTANT: Do not infer permission from "type" alone. Multiple pages reuse the same
        # translation types (e.g., indicator bank uses "form_item"), so callers must declare
        # an explicit permission context + permission code which we validate against allowlists.
        permission_context = (data.get('permission_context') or data.get('context') or '').strip().lower()
        permission_code = (data.get('permission_code') or data.get('permissionCode') or '').strip()

        allowed_context_permissions = {
            # Template creation/editing flows
            'templates': {'admin.templates.create', 'admin.templates.edit'},
            # Indicator bank bulk actions
            'indicator_bank': {'admin.indicator_bank.create', 'admin.indicator_bank.edit'},
            # Resources management editor
            'resources': {'admin.resources.manage'},
            # Translation file/admin translations editor
            'translations': {'admin.translations.manage'},
            # Settings admin area (org branding translations, etc.)
            'settings': {'admin.settings.manage'},
            # Organization management editor (edit_entity helper page)
            'organization': {'admin.organization.manage'},
        }

        # Backwards compatibility: if caller didn't pass context, infer safe defaults.
        # (Prefer explicit context; new UI callers should always send it.)
        if not permission_context:
            if (data.get('type') or '').strip() == 'translation':
                permission_context = 'translations'
            else:
                permission_context = 'templates'

        if not permission_code:
            # Default permission per context (most common edit/manage action)
            default_permission_by_context = {
                'templates': 'admin.templates.edit',
                'indicator_bank': 'admin.indicator_bank.edit',
                'resources': 'admin.resources.manage',
                'translations': 'admin.translations.manage',
                'settings': 'admin.settings.manage',
                'organization': 'admin.organization.manage',
            }
            permission_code = default_permission_by_context.get(permission_context, '')

        allowed_permissions = allowed_context_permissions.get(permission_context)
        if not allowed_permissions or permission_code not in allowed_permissions:
            return json_forbidden(f'Auto-translate not allowed for context "{permission_context}"')

        # System managers bypass granular permission checks
        if not AuthorizationService.is_system_manager(current_user):
            if not AuthorizationService.has_rbac_permission(current_user, permission_code):
                return json_forbidden('You do not have permission to use auto-translate here.')

        translation_type = data.get('type')  # 'form_item', 'section_name', 'question_option', 'page_name', or 'template_name'
        # Support base64-wrapped text to avoid WAF false positives on user-provided strings.
        # Prefer *_b64 when present; keep plain fields for backward compatibility.
        def _decode_b64_field(key: str) -> str:
            raw = data.get(key)
            if not raw:
                return ""
            try:
                import base64 as _b64
                return _b64.b64decode(str(raw)).decode("utf-8")
            except Exception:
                return ""

        _raw_text = _decode_b64_field('text_b64') or data.get('text') or ''
        _ttype = str(translation_type or '').strip()
        # PO gettext msgids are matched exactly; leading/trailing spaces are significant.
        # Stripping would break po.find() and append duplicate entries while the UI still shows the original empty row.
        if _ttype == 'translation':
            text = _raw_text if isinstance(_raw_text, str) else str(_raw_text)
            if not text.strip():
                return json_bad_request('Text is required')
        else:
            text = _raw_text.strip()
        definition = (_decode_b64_field('definition_b64') or data.get('definition') or '').strip()
        # Get target languages from config (normalize to ISO codes)
        try:
            from config.config import Config
            # Normalize incoming target languages to ISO language codes expected by translator
            incoming_target_languages = data.get('target_languages')
            if incoming_target_languages:
                normalized_target_languages = []
                for lang in incoming_target_languages:
                    # Accept only ISO codes (e.g., 'fr'); strip any region suffix (e.g., 'fr_FR' -> 'fr')
                    if isinstance(lang, str):
                        lang_norm = lang.split('_', 1)[0].strip().lower()
                        if lang_norm:
                            normalized_target_languages.append(lang_norm)
                target_languages = normalized_target_languages
            else:
                # Default to all translatable language codes, skipping English/source
                target_languages = [
                    lc for lc in (current_app.config.get('TRANSLATABLE_LANGUAGES') or getattr(Config, 'TRANSLATABLE_LANGUAGES', []) or [])
                    if lc != 'en'
                ]
        except (ImportError, NameError, AttributeError):
            # Fallback defaults (language codes) from runtime config
            configured = current_app.config.get('TRANSLATABLE_LANGUAGES') or []
            target_languages = [lc for lc in (configured or []) if lc != 'en']
        translation_service = data.get('translation_service', 'ifrc')  # Default hosted service (internal id "ifrc")
        service_name = translation_service  # Map to existing parameter for backward compatibility

        if not text:
            return json_bad_request('Text is required')

        auto_translator = get_auto_translator()

        # Check if any translation services are available
        available_services = auto_translator.get_available_services()
        if not available_services:
            return json_bad_request('No translation service available. Please configure translation API keys. Set IFRC_TRANSLATE_API_KEY, GOOGLE_TRANSLATE_API_KEY, or LIBRE_TRANSLATE_URL environment variables.')

        # If a specific service was requested but is not available, try to use a fallback
        if service_name and service_name not in available_services:
            logger.warning(f"Requested translation service '{service_name}' is not available. Available services: {available_services}. Using fallback.")
            service_name = None  # Let it use the default service

        if translation_type == 'form_item':
            # Translate form item (label and definition)
            result = translate_form_item_auto(
                label=text,
                definition=definition,
                target_languages=target_languages,
                service_name=service_name
            )

            if result and (result.get('label_translations') or result.get('definition_translations')):
                return json_ok(translations=result, service_used=auto_translator.get_default_service())
            else:
                return json_bad_request(f'Translation failed. No translations were generated. Available services: {", ".join(available_services)}. Please check your API keys and try again.')

        elif translation_type == 'document_field':
            # Translate document field (label only; descriptions are optional)
            result = translate_form_item_auto(
                label=text,
                definition=None,
                target_languages=target_languages,
                service_name=service_name
            )
            if result and result.get('label_translations'):
                return json_ok(translations=result, service_used=auto_translator.get_default_service())
            else:
                return json_bad_request(f'Translation failed. No translations were generated. Available services: {", ".join(available_services)}. Please check your API keys and try again.')
        elif translation_type == 'section_name':
            # Translate section name only
            result = translate_section_name_auto(
                name=text,
                target_languages=target_languages,
                service_name=service_name
            )

            if result and len(result) > 0:
                return json_ok(translations=result, service_used=auto_translator.get_default_service())
            else:
                # Return partial success if some translations failed but we have at least one
                # This allows the frontend to show successful translations even if some languages failed
                return json_error(
                    f'Translation failed for some languages. Available services: {", ".join(available_services)}. Some translations may have failed due to service limitations.',
                    200,
                    success=False,
                    translations=result or {},
                )

        elif translation_type == 'question_option':
            # Translate question option
            result = translate_question_option_auto(
                option_text=text,
                target_languages=target_languages,
                service_name=service_name
            )

            if result and len(result) > 0:
                return json_ok(translations=result, service_used=auto_translator.get_default_service())
            else:
                return json_bad_request(f'Translation failed. No translations were generated. Available services: {", ".join(available_services)}. Please check your API keys and try again.')

        elif translation_type == 'page_name':
            # Translate page name only
            result = translate_page_name_auto(
                name=text,
                target_languages=target_languages,
                service_name=service_name
            )

            if result and len(result) > 0:
                return json_ok(translations=result, service_used=auto_translator.get_default_service())
            else:
                return json_bad_request(f'Translation failed. No translations were generated. Available services: {", ".join(available_services)}. Please check your API keys and try again.')

        elif translation_type == 'template_name':
            # Translate template name only
            result = translate_template_name_auto(
                name=text,
                target_languages=target_languages,
                service_name=service_name
            )

            if result and len(result) > 0:
                return json_ok(translations=result, service_used=auto_translator.get_default_service())
            else:
                # Translation produced no output (e.g. acronym/proper noun returned unchanged or
                # service unavailable for this text). Return the source text for each target language
                # so the caller can populate fields rather than showing a hard error.
                fallback = {lang: text for lang in target_languages if lang}
                if fallback:
                    return json_ok(
                        translations=fallback,
                        service_used=None,
                        untranslated=True,
                        message='No translation available; original text returned.',
                    )
                return json_bad_request(f'Translation failed. No translations were generated. Available services: {", ".join(available_services)}. Please check your API keys and try again.')

        elif translation_type == 'email_template_html':
            # English HTML email body; Jinja {{ }} / {% %} preserved in auto_translator
            result = translate_email_template_html_auto(
                html=text,
                target_languages=target_languages,
                service_name=service_name,
            )
            if result and len(result) > 0:
                return json_ok(translations=result, service_used=auto_translator.get_default_service())
            return json_bad_request(
                f'Translation failed. No translations were generated. Available services: {", ".join(available_services)}. Please check your API keys and try again.'
            )

        elif translation_type == 'translation':
            # Translate a translation file entry (msgid to other languages)
            from app.services.translation.auto_translator import translate_text as auto_translate_text
            from config import Config

            # Try to import polib, but make it optional
            try:
                import polib  # type: ignore
            except ImportError:
                current_app.logger.warning("polib not available - translation file updates will be skipped")
                polib = None

            # Get message ID for the translation (exact string; do not strip — gettext msgids are exact)
            _mid = data.get('id')
            if _mid is None:
                return json_bad_request('Message ID is required for translation type')
            message_id = _mid if isinstance(_mid, str) else str(_mid)
            if not message_id.strip():
                return json_bad_request('Message ID is required for translation type')

            # Check if polib is available for file updates
            if polib is None:
                return json_bad_request('Translation file updates are not available. Please install polib package.')

            # Translate the text to the target languages
            translations = {}
            success_count = 0
            skipped_untranslatable = 0

            from config import Config as _TransCfg
            _supported_locales = set(current_app.config.get('SUPPORTED_LANGUAGES', _TransCfg.LANGUAGES))

            for lang in target_languages:
                try:
                    # ISO codes only (accept regional variants like fr_FR)
                    target_locale = str(lang or '').strip()
                    if '_' in target_locale:
                        target_locale = target_locale.split('_', 1)[0]
                    target_locale = target_locale.lower()
                    if target_locale == 'en':
                        continue  # Skip English as source language
                    # Validate against supported languages to prevent path injection
                    if target_locale not in _supported_locales:
                        logger.warning("Skipping unsupported locale for translation: %s", target_locale)
                        continue

                    # Translate the text
                    translated_text = auto_translate_text(
                        text=text,
                        target_language=target_locale,
                        service_name=service_name
                    )

                    if not translated_text or not translated_text.strip():
                        # Service responded but returned no usable translation (e.g. proper
                        # noun / acronym / technical term that the API left unchanged and the
                        # untranslated-output heuristic rejected). This is not a server error.
                        skipped_untranslatable += 1
                        continue

                    if translated_text and translated_text.strip():
                        translations[target_locale] = translated_text

                        po_file_path = _translations_po_path(target_locale)
                        try:
                            # Ensure directory exists
                            po_dir = os.path.dirname(po_file_path)
                            with suppress(Exception):
                                os.makedirs(po_dir, exist_ok=True)

                            # Load existing PO or create a new one
                            if os.path.exists(po_file_path):
                                po = polib.pofile(po_file_path)
                            else:
                                po = polib.POFile()
                                en_po_path = _translations_po_path('en')
                                if os.path.exists(en_po_path):
                                    with suppress(Exception):
                                        en_po = polib.pofile(en_po_path)
                                        if getattr(en_po, "metadata", None):
                                            po.metadata = dict(en_po.metadata)
                                # Minimal metadata fallback
                                if not getattr(po, "metadata", None):
                                    po.metadata = {}
                                po.metadata.setdefault('Content-Type', 'text/plain; charset=utf-8')
                                po.metadata.setdefault('Content-Transfer-Encoding', '8bit')
                                po.metadata.setdefault('Language', target_locale)

                            entry = po.find(message_id)
                            if entry:
                                # Update existing entry (plural or non-plural)
                                if hasattr(entry, 'msgid_plural') and entry.msgid_plural:
                                    if not hasattr(entry, 'msgstr_plural') or not entry.msgstr_plural:
                                        entry.msgstr_plural = {}
                                    entry.msgstr_plural[0] = translated_text
                                    # If plural forms exist, mirror into second form as a simple fallback
                                    if len(entry.msgstr_plural) > 1:
                                        entry.msgstr_plural[1] = translated_text
                                else:
                                    entry.msgstr = translated_text
                            else:
                                en_po_path = _translations_po_path('en')
                                is_plural = False
                                msgid_plural = None
                                if os.path.exists(en_po_path):
                                    with suppress(Exception):
                                        en_po = polib.pofile(en_po_path)
                                        en_entry = en_po.find(message_id)
                                        if en_entry and hasattr(en_entry, 'msgid_plural') and en_entry.msgid_plural:
                                            is_plural = True
                                            msgid_plural = en_entry.msgid_plural

                                if is_plural:
                                    entry = polib.POEntry(msgid=message_id, msgid_plural=msgid_plural or message_id)
                                    entry.msgstr_plural = {0: translated_text, 1: translated_text}
                                else:
                                    entry = polib.POEntry(msgid=message_id, msgstr=translated_text)
                                po.append(entry)

                            po.save(po_file_path)
                            success_count += 1
                        except Exception as e:
                            current_app.logger.error(f"Error updating/creating po file for {target_locale}: {e}", exc_info=True)

                except Exception as e:
                    current_app.logger.error(f"Error translating to {lang}: {e}", exc_info=True)
                    continue

            if success_count > 0:
                return json_ok(
                    translations={'label_translations': translations},
                    updated_count=success_count,
                    service_used=service_name,
                )
            elif skipped_untranslatable > 0:
                # The translation API responded successfully but the text could not be
                # translated (e.g. proper noun, technical term, or acronym returned
                # unchanged). This is not an error — return a soft 200 so
                # the UI can show an informational message rather than "Network error".
                return json_ok(
                    translations={},
                    updated_count=0,
                    skipped_untranslatable=skipped_untranslatable,
                    service_used=service_name,
                    untranslated=True,
                    message='No translation available: the text may be a proper noun or technical term that does not require translation.',
                )
            else:
                return json_server_error('Failed to translate or update translation files')

        else:
            return json_bad_request('Invalid translation type')

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)


@bp.route('/api/auto-translate-summary', methods=['POST'])
@limiter.exempt
@permission_required_any(
    'admin.templates.edit',
    'admin.templates.create',
    'admin.indicator_bank.create',
    'admin.indicator_bank.edit',
    'admin.resources.manage',
    'admin.translations.manage',
    'admin.organization.manage',
    'admin.settings.manage',
)
def api_auto_translate_summary():
    """Log a single activity record summarising a bulk auto-translate run.

    Called by the client once when the batch finishes, instead of logging
    one record per individual translation call.
    """
    try:
        data = get_json_safe()
        err = require_json_data(data)
        if err:
            return err

        total = int(data.get('total', 0) or 0)
        success_count = int(data.get('success_count', 0) or 0)
        error_count = int(data.get('error_count', 0) or 0)
        languages = data.get('languages') or []
        context = (data.get('context') or '').strip()
        service = (data.get('service') or '').strip()

        lang_list = ', '.join(str(l) for l in languages[:10]) if languages else 'all'
        description = (
            f"Bulk auto-translate: {success_count}/{total} translations completed"
            f" ({error_count} errors) for languages [{lang_list}]"
        )
        if service:
            description += f" using {service}"
        if context:
            description += f" in {context}"

        from app.services.user_analytics_service import log_user_activity
        log_user_activity(
            activity_type='request',
            description=description,
            context_data={
                'endpoint': 'utilities.api_auto_translate_summary',
                'method': 'POST',
                'total': total,
                'success_count': success_count,
                'error_count': error_count,
                'languages': languages,
                'permission_context': context,
                'service': service,
            },
        )

        return json_ok(logged=True)
    except Exception as e:
        logger.warning("Error logging auto-translate summary: %s", e)
        return json_ok(logged=False)


@bp.route('/api/bulk-update-translations', methods=['POST'])
@limiter.exempt
@permission_required('admin.templates.edit')
def api_bulk_update_translations():
    """API endpoint for bulk updating translations of template items"""
    try:
        from app.models import FormItem, FormSection, FormPage
        from app.extensions import db

        data = get_json_safe()
        err = require_json_keys(data, ['items'])
        if err:
            return err

        items = data.get('items', [])
        if not items:
            return json_bad_request('No items provided')

        success_count = 0
        error_count = 0
        errors = []

        # Process all items in a single transaction
        try:
            for item in items:
                try:
                    if not isinstance(item, dict):
                        error_count += 1
                        errors.append(f"Invalid item format: {type(item)}")
                        continue

                    item_id = item.get('id')
                    item_type = item.get('type')  # 'question', 'document_field', 'section', or 'page'
                    translations = item.get('translations', {})

                    if not item_id or not item_type:
                        error_count += 1
                        errors.append(f"Missing ID or type for item")
                        continue

                    # Handle different item types (including plugin_* types)
                    is_form_item_type = (
                        item_type in ['indicator', 'question', 'document_field', 'matrix']
                        or (isinstance(item_type, str) and item_type.startswith('plugin_'))
                    )
                    if is_form_item_type:
                        # Map item type to string values (plugin_* types pass through as-is)
                        if item_type == 'indicator':
                            form_item_type = 'indicator'
                        elif item_type == 'question':
                            form_item_type = 'question'
                        elif item_type == 'document_field':
                            form_item_type = 'document_field'
                        elif item_type == 'matrix':
                            form_item_type = 'matrix'
                        else:
                            form_item_type = item_type  # e.g. plugin_interactive_map

                        # Find the form item
                        form_item = FormItem.query.filter_by(id=item_id, item_type=form_item_type).first()
                        if not form_item:
                            error_count += 1
                            errors.append(f"Form item not found: {item_id}")
                            continue

                        # Update form item translations
                        # Handle both nested format and individual language format
                        if 'label_translations' in translations:
                            # Merge with existing translations instead of replacing
                            if not form_item.label_translations:
                                form_item.label_translations = {}
                            existing_translations = form_item.label_translations.copy()
                            new_translations = translations['label_translations']
                            existing_translations.update(new_translations)
                            form_item.label_translations = existing_translations
                        else:
                            # Handle individual language fields (e.g., label_french, label_spanish)
                            if not form_item.label_translations:
                                form_item.label_translations = {}
                            label_translations = form_item.label_translations.copy()
                            for key, value in translations.items():
                                if key.startswith('label_') and len(key) > 6:
                                    lang = key[6:]  # Remove 'label_' prefix
                                    if value and value.strip():
                                        label_translations[lang] = value.strip()
                            if label_translations:
                                form_item.label_translations = label_translations

                        if 'definition_translations' in translations:
                            # Merge with existing translations instead of replacing
                            if not form_item.definition_translations:
                                form_item.definition_translations = {}
                            existing_translations = form_item.definition_translations.copy()
                            new_translations = translations['definition_translations']
                            existing_translations.update(new_translations)
                            form_item.definition_translations = existing_translations
                        elif 'description_translations' in translations:
                            # Merge with existing translations instead of replacing
                            if not form_item.description_translations:
                                form_item.description_translations = {}
                            existing_translations = form_item.description_translations.copy()
                            new_translations = translations['description_translations']
                            existing_translations.update(new_translations)
                            form_item.description_translations = existing_translations
                        else:
                            # Handle individual description language fields
                            if form_item.is_indicator or form_item.is_question:
                                if not form_item.definition_translations:
                                    form_item.definition_translations = {}
                                definition_translations = form_item.definition_translations.copy()
                                for key, value in translations.items():
                                    if key.startswith('description_') and len(key) > 12:
                                        lang = key[12:]  # Remove 'description_' prefix
                                        if value and value.strip():
                                            definition_translations[lang] = value.strip()
                                if definition_translations:
                                    form_item.definition_translations = definition_translations
                            elif form_item.is_document_field or form_item.is_matrix:
                                if not form_item.description_translations:
                                    form_item.description_translations = {}
                                description_translations = form_item.description_translations.copy()
                                for key, value in translations.items():
                                    if key.startswith('description_') and len(key) > 12:
                                        lang = key[12:]  # Remove 'description_' prefix
                                        if value and value.strip():
                                            description_translations[lang] = value.strip()
                                if description_translations:
                                    form_item.description_translations = description_translations

                        if 'options_translations' in translations:
                            form_item.options_translations = translations['options_translations']

                    elif item_type == 'section':
                        # Find the section
                        section = FormSection.query.filter_by(id=item_id).first()
                        if not section:
                            error_count += 1
                            errors.append(f"Section not found: {item_id}")
                            continue

                        # Update section name translations
                        if 'name_translations' in translations:
                            # Merge with existing translations instead of replacing
                            if not section.name_translations:
                                section.name_translations = {}
                            existing_translations = section.name_translations.copy()
                            new_translations = translations['name_translations']
                            existing_translations.update(new_translations)
                            section.name_translations = existing_translations

                    elif item_type == 'page':
                        # Find the page
                        page = FormPage.query.filter_by(id=item_id).first()
                        if not page:
                            error_count += 1
                            errors.append(f"Page not found: {item_id}")
                            continue

                        # Update page name translations
                        if 'name_translations' in translations:
                            # Merge with existing translations instead of replacing
                            if not page.name_translations:
                                page.name_translations = {}
                            existing_translations = page.name_translations.copy()
                            new_translations = translations['name_translations']
                            existing_translations.update(new_translations)
                            page.name_translations = existing_translations

                    elif item_type == 'template_name':
                        # Find the template; name_translations live on FormTemplateVersion, not FormTemplate
                        from app.models import FormTemplate, FormTemplateVersion
                        template = FormTemplate.query.filter_by(id=item_id).first()
                        if not template:
                            error_count += 1
                            errors.append(f"Template not found: {item_id}")
                            continue
                        # Use the version that owns the template name (published or first)
                        version = template.published_version or template.versions.order_by(FormTemplateVersion.created_at).first()
                        if not version:
                            error_count += 1
                            errors.append(f"No version found for template: {item_id}")
                            continue
                        # Update name_translations on the version (FormTemplate.name_translations is a read-only property from version)
                        if 'name_translations' in translations:
                            if not version.name_translations:
                                version.name_translations = {}
                            existing_translations = version.name_translations.copy()
                            new_translations = translations['name_translations']
                            existing_translations.update(new_translations)
                            version.name_translations = existing_translations

                    else:
                        error_count += 1
                        errors.append(f"Invalid item type: {item_type}")
                        continue

                    success_count += 1

                except Exception as e:
                    error_count += 1
                    errors.append(f"Error updating item {item.get('id', 'unknown')}.")
                    current_app.logger.error(f"Error updating translations for item {item.get('id', 'unknown')}: {e}")

            # Commit all changes at once
            db.session.flush()

        except Exception as e:
            return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)

        return json_ok(
            message=f'Successfully updated {success_count} items. {error_count} errors occurred.',
            success_count=success_count,
            error_count=error_count,
            errors=errors,
        )

    except Exception as e:
        return handle_json_view_exception(e, GENERIC_ERROR_MESSAGE, status_code=500)
