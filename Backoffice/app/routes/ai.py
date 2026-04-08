from __future__ import annotations

import html as html_lib
import io
import json
import logging
import os
import queue
import re
import threading
import uuid
import time
from datetime import datetime
from typing import Any, Dict, Optional, Tuple

logger = logging.getLogger(__name__)

_ACTIVE_SSE_CANCEL_EVENTS: Dict[str, threading.Event] = {}
_ACTIVE_SSE_CANCEL_LOCK = threading.Lock()

from contextlib import suppress
from flask import Blueprint, Response, current_app, g, jsonify, request, send_file, stream_with_context, session
from flask_login import login_required, current_user, login_user, logout_user

from app.utils.ai_request_user import resolve_ai_identity
from app.utils.ai_tokens import issue_ai_token
from app.utils.ai_pricing import estimate_chat_cost
from app.utils.ai_utils import sanitize_page_context, openai_model_supports_sampling_params
from app.utils.api_helpers import GENERIC_ERROR_MESSAGE, get_json_safe
from app.utils.api_responses import json_auth_required, json_bad_request, json_error, json_forbidden, json_not_found, json_ok, json_server_error
from app.services.ai_chat_request import (
    parse_chat_request,
    resolve_conversation_and_history,
    get_idempotent_reply,
    apply_anonymous_rules,
    load_conversation_history_for_llm,
    replace_conversation_messages,
    find_existing_reply_for_client_message_id,
)
from sqlalchemy.orm.attributes import flag_modified
from app.utils.datetime_helpers import utcnow
from app.extensions import db, limiter
from app.models.ai_chat import AIConversation, AIMessage
from app.models import AIReasoningTrace
from app.services.ai_chat_retention import delete_archive_object, load_archived_conversation
from app.services.ai_dlp import evaluate_ai_message, log_dlp_audit_event

from app.services.ai_fastpath import try_answer_value_question

# Telemetry tracking
from app.services.chatbot_telemetry import ChatbotTelemetryService, ChatbotMetrics
from app.services.user_analytics_service import get_client_ip

# AI Agent integration (RAG + agentic reasoning)
# Lazy-loaded to avoid import errors if dependencies are missing
_ai_chat_integration = None

def _get_ai_chat_integration():
    """Lazy-load AIChatIntegration to avoid import errors at startup."""
    global _ai_chat_integration
    if _ai_chat_integration is None:
        try:
            from app.services.ai_chat_integration import AIChatIntegration
            _ai_chat_integration = AIChatIntegration()
        except (ImportError, Exception) as e:
            current_app.logger.warning("Failed to initialize AIChatIntegration: %s", e)
            _ai_chat_integration = False  # Mark as failed, don't retry
    return _ai_chat_integration if _ai_chat_integration else None


ai_bp = Blueprint("ai_v2", __name__, url_prefix="/api/ai/v2")


PUBLIC_AI_PROXY_HEADER = "X-NGO-Databank-AI-Proxy"


def _is_development_mode() -> bool:
    flask_config = (os.environ.get("FLASK_CONFIG") or "").strip().lower()
    return flask_config in {"development", "default", ""}


def _public_ai_proxy_secret() -> Optional[str]:
    # Read directly from environment to avoid relying on config object mapping.
    secret = (os.environ.get("AI_PUBLIC_PROXY_SECRET") or "").strip()
    return secret or None


def _is_allowed_public_proxy_request() -> bool:
    """
    Allow anonymous AI chat only when coming through the Website proxy.

    In production/staging, require a shared secret header (prevents direct public abuse).
    In development, allow anonymous without a secret for convenience.
    """
    secret = _public_ai_proxy_secret()
    if not secret:
        # Dev convenience: don't block anonymous when not configured.
        if _is_development_mode():
            current_app.logger.warning(
                "AI_PUBLIC_PROXY_SECRET is not set; allowing anonymous AI chat in development. "
                "Set AI_PUBLIC_PROXY_SECRET in production to prevent abuse."
            )
            return True
        current_app.logger.error(
            "AI_PUBLIC_PROXY_SECRET is not set in non-development mode; blocking anonymous AI chat. "
            "Set AI_PUBLIC_PROXY_SECRET to enable Website public chatbot safely."
        )
        return False

    presented = (request.headers.get(PUBLIC_AI_PROXY_HEADER) or "").strip()
    return bool(presented) and presented == secret


def _ai_beta_denied_response(identity):
    """Return a JSON denial response when AI beta access restriction blocks this identity."""
    try:
        from app.services.app_settings_service import is_ai_beta_restricted, user_has_ai_beta_access

        if not is_ai_beta_restricted():
            return None
        if not identity.is_authenticated or not identity.user:
            return json_auth_required("AI beta access is limited to selected users.")
        if not user_has_ai_beta_access(identity.user):
            return json_forbidden("AI beta access is limited to selected users.")
    except Exception as e:
        logger.debug("_ai_beta_denied_response check failed: %s", e, exc_info=True)
    return None


def _ai_chat_rate_limit_key() -> str:
    """
    Rate limit key that prefers authenticated user id, else client IP.
    Uses get_client_ip() so X-Forwarded-For / X-Real-IP works behind proxies.
    """
    try:
        identity = resolve_ai_identity()
        if identity.is_authenticated and identity.user:
            return f"ai_chat:{identity.auth_source}:{int(identity.user.id)}"
    except Exception as e:
        logger.debug("resolve_ai_identity failed, using IP fallback: %s", e, exc_info=True)
    return f"ai_chat:ip:{get_client_ip()}"


def _ai_chat_limit() -> str:
    # Keep static to reduce configuration surface.
    try:
        identity = resolve_ai_identity()
        if identity.is_authenticated:
            return "120 per minute"
    except Exception as e:
        logger.debug("resolve_ai_identity failed for rate limit: %s", e, exc_info=True)
    return "60 per minute"

def _ai_chat_daily_user_limit() -> str:
    # System managers should not be constrained by the per-user daily cap.
    try:
        identity = resolve_ai_identity()
        if identity.is_authenticated and identity.user:
            from app.services.authorization_service import AuthorizationService
            if AuthorizationService.is_system_manager(identity.user):
                # Effectively exempt (Flask-Limiter does not provide a reliable "disable limit"
                # signal via dynamic strings; use a very high ceiling instead).
                return "5000000 per day"
    except Exception as e:
        logger.debug("resolve_ai_identity or is_system_manager failed: %s", e, exc_info=True)
    return "1000000 per day"


def _ai_chat_daily_system_limit() -> str:
    return "5000000 per day"


def _ai_chat_system_rate_limit_key() -> str:
    # Single shared bucket for system-wide daily cap.
    return "ai_chat:system"


def _ai_cancel_limit() -> str:
    return "120 per minute"


def _ai_feedback_limit() -> str:
    return "120 per minute"


def _ai_export_limit() -> str:
    return "60 per minute"


def _ai_table_export_limit() -> str:
    return "60 per minute"


def _ai_delete_all_conversations_limit() -> str:
    return "30 per minute"


def _ai_clear_inflight_limit() -> str:
    return "30 per minute"


def _ai_append_message_limit() -> str:
    return "60 per minute"


def _ai_import_conversation_limit() -> str:
    return "10 per minute"


def _build_initial_conversation_title(message: Optional[str]) -> str:
    text = str(message or "").strip()
    if not text:
        return "New chat"
    return (text[:80] + "…") if len(text) > 80 else text


def _normalize_conversation_meta(meta: Any) -> Dict[str, Any]:
    if isinstance(meta, dict):
        return dict(meta)
    return {}


def _strip_html_to_text(value: Optional[str]) -> str:
    text = str(value or "")
    if not text:
        return ""
    text = re.sub(r"<[^>]+>", " ", text)
    text = html_lib.unescape(text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _clean_generated_title(raw: Optional[str]) -> Optional[str]:
    title = str(raw or "").strip()
    if not title:
        return None
    title = title.strip().strip('"').strip("'").strip()
    title = re.sub(r"^[Tt]itle\s*:\s*", "", title).strip()
    title = re.sub(r"\s+", " ", title).strip(" .,-:;")
    if not title:
        return None
    if len(title) > 120:
        title = title[:120].rstrip(" .,-:;")
    return title or None


def _generate_refined_title_with_openai(
    *,
    user_message: Optional[str],
    assistant_response: Optional[str],
) -> Optional[str]:
    api_key = current_app.config.get("OPENAI_API_KEY") or os.getenv("OPENAI_API_KEY")
    if not api_key:
        return None

    user_text = _strip_html_to_text(user_message)
    assistant_text = _strip_html_to_text(assistant_response)
    if not user_text and not assistant_text:
        return None

    try:
        from openai import OpenAI
    except Exception as e:
        logger.debug("OpenAI import failed for title refinement: %s", e, exc_info=True)
        return None

    model_name = str(
        current_app.config.get("OPENAI_TITLE_MODEL")
        or current_app.config.get("OPENAI_MODEL", "gpt-5-mini")
    )
    timeout_sec = int(current_app.config.get("AI_HTTP_TIMEOUT_SECONDS", 60))
    client = OpenAI(api_key=api_key, timeout=timeout_sec)

    system_prompt = (
        "You create concise conversation titles for a chat sidebar.\n"
        "Rules:\n"
        "- Return ONLY the title text.\n"
        "- Max 8 words.\n"
        "- No quotes, no markdown, no punctuation at the end.\n"
        "- Be specific to the user's intent."
    )
    user_prompt = (
        f"User message:\n{user_text[:700]}\n\n"
        f"Assistant reply:\n{assistant_text[:900]}\n\n"
        "Generate a better conversation title."
    )

    kwargs: Dict[str, Any] = {
        "model": model_name,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "max_completion_tokens": 32,
    }
    if openai_model_supports_sampling_params(model_name):
        kwargs["temperature"] = 0.2

    try:
        completion = client.chat.completions.create(**kwargs)
        raw_title = (
            (completion.choices[0].message.content or "")
            if completion and completion.choices
            else ""
        )
        return _clean_generated_title(raw_title)
    except Exception as e:
        current_app.logger.debug("Conversation title refinement failed: %s", e, exc_info=True)
        return None


def _maybe_refine_conversation_title(
    convo: Optional[AIConversation],
    *,
    user_message: Optional[str],
    assistant_response: Optional[str],
) -> bool:
    if not convo:
        return False

    meta = _normalize_conversation_meta(convo.meta)
    title_source = str(meta.get("title_source") or "").strip().lower()

    # Only auto-refine titles we generated automatically.
    if title_source and title_source not in {"auto_initial"}:
        return False

    refined = _generate_refined_title_with_openai(
        user_message=user_message,
        assistant_response=assistant_response,
    )
    if not refined:
        return False

    current_title = str(convo.title or "").strip()
    if refined == current_title:
        meta["title_source"] = "auto_refined"
        convo.meta = meta
        return False

    convo.title = refined
    meta["title_source"] = "auto_refined"
    meta["title_refined_at"] = utcnow().isoformat()
    convo.meta = meta
    return True


@ai_bp.route("/chat/cancel", methods=["POST"])
@limiter.limit(_ai_cancel_limit)
def cancel_chat_stream():
    """
    Best-effort cancel for SSE streaming requests.
    The client calls this before aborting the SSE fetch so we can distinguish
    "Stop" from a refresh/network drop when keep_running_on_disconnect is enabled.
    """
    identity = resolve_ai_identity()
    beta_denied = _ai_beta_denied_response(identity)
    if beta_denied is not None:
        return beta_denied
    if not identity.is_authenticated:
        return json_auth_required("Authentication required")

    data = get_json_safe()
    request_id = (data.get("request_id") or "").strip()
    if not request_id:
        return json_bad_request("request_id required")

    cancelled = False
    with _ACTIVE_SSE_CANCEL_LOCK:
        ev = _ACTIVE_SSE_CANCEL_EVENTS.get(request_id)
        if ev:
            try:
                ev.set()
                cancelled = True
            except Exception as e:
                logger.warning("Failed to set cancel event: %s", e, exc_info=True)
                cancelled = False

    return json_ok(cancelled=cancelled)


def _build_access_context(identity) -> Dict[str, Any]:
    """
    Build RBAC-aware context. This is used to decide which tools/data can be used.
    """
    if not identity.is_authenticated or not identity.user:
        return {
            "access_level": "public",
            "role": "public",
            "permissions": {},
            "allowed_country_ids": [],
        }

    user = identity.user
    from app.services.authorization_service import AuthorizationService

    # Permissions: store a small, stable set of RBAC permission checks for downstream tool gating.
    perms = {
        "*": bool(AuthorizationService.is_system_manager(user)),
        # Admin modules
        "admin.users.view": AuthorizationService.has_rbac_permission(user, "admin.users.view"),
        "admin.templates.view": AuthorizationService.has_rbac_permission(user, "admin.templates.view"),
        "admin.assignments.view": AuthorizationService.has_rbac_permission(user, "admin.assignments.view"),
        "admin.countries.view": AuthorizationService.has_rbac_permission(user, "admin.countries.view"),
        "admin.organization.manage": AuthorizationService.has_rbac_permission(user, "admin.organization.manage"),
        "admin.indicator_bank.view": AuthorizationService.has_rbac_permission(user, "admin.indicator_bank.view"),
        "admin.settings.manage": AuthorizationService.has_rbac_permission(user, "admin.settings.manage"),
        "admin.analytics.view": AuthorizationService.has_rbac_permission(user, "admin.analytics.view"),
        "admin.audit.view": AuthorizationService.has_rbac_permission(user, "admin.audit.view"),
        "admin.api.manage": AuthorizationService.has_rbac_permission(user, "admin.api.manage"),
        "admin.plugins.manage": AuthorizationService.has_rbac_permission(user, "admin.plugins.manage"),
        "admin.data_explore.data_table": AuthorizationService.has_rbac_permission(user, "admin.data_explore.data_table"),
        "admin.data_explore.analysis": AuthorizationService.has_rbac_permission(user, "admin.data_explore.analysis"),
        "admin.data_explore.compliance": AuthorizationService.has_rbac_permission(user, "admin.data_explore.compliance"),
        "admin.documents.manage": AuthorizationService.has_rbac_permission(user, "admin.documents.manage"),
        # Assignment participation
        "assignment.view": AuthorizationService.has_rbac_permission(user, "assignment.view"),
        "assignment.enter": AuthorizationService.has_rbac_permission(user, "assignment.enter"),
        "assignment.submit": AuthorizationService.has_rbac_permission(user, "assignment.submit"),
        "assignment.approve": AuthorizationService.has_rbac_permission(user, "assignment.approve"),
        "assignment.reopen": AuthorizationService.has_rbac_permission(user, "assignment.reopen"),
    }

    allowed_country_ids = []
    try:
        if AuthorizationService.is_system_manager(user) or AuthorizationService.is_admin(user):
            # Admins typically can see all countries, but still rely on tool-level checks.
            allowed_country_ids = []
        else:
            allowed_country_ids = [c.id for c in user.countries.all()]
    except Exception as e:
        logger.warning("Failed to resolve allowed_country_ids: %s", e, exc_info=True)
        allowed_country_ids = []

    return {
        "access_level": identity.access_level,
        "role": identity.access_level,
        "permissions": perms,
        "allowed_country_ids": allowed_country_ids,
    }


def _build_platform_context(identity, access_context: Dict[str, Any]) -> Dict[str, Any]:
    """Build platform context for the AI engine. Used by chat and chat_stream."""
    platform_context: Dict[str, Any] = {
        "user_info": {
            "id": int(getattr(identity.user, "id", 0) or 0) if identity.is_authenticated and identity.user else None,
            "role": access_context["role"],
            "access_level": access_context["access_level"],
        },
        "access": access_context,
        "user_data": {},
        "available_countries": [],
        "available_templates": [],
        "available_indicators": [],
        "platform_stats": {},
    }
    try:
        from app.models import User, Country, FormTemplate, IndicatorBank
        platform_context["platform_stats"] = {
            "total_users": User.query.count() if identity.is_authenticated else 0,
            "total_countries": Country.query.count(),
            "total_templates": FormTemplate.query.count(),
            "total_indicators": IndicatorBank.query.filter_by(archived=False).count(),
            "total_assignments": 0,
        }
    except Exception as e_stats:
        current_app.logger.warning("Could not populate platform stats: %s", e_stats)

    if identity.is_authenticated and identity.user:
        try:
            from app.services import get_user_data_context
            user_data_context = get_user_data_context(user_id=identity.user.id)
            platform_context["user_data"]["total_assignments"] = user_data_context.get("total_assignments", 0)
            platform_context["user_data"]["completed_assignments"] = user_data_context.get("completed_assignments", 0)
            platform_context["user_data"]["pending_assignments"] = user_data_context.get("pending_assignments", 0)
            platform_context["user_data"]["pending_assignment_details"] = user_data_context.get("pending_assignment_details", [])
            platform_context["user_data"]["countries"] = user_data_context.get("countries", [])
            if platform_context["user_data"]["total_assignments"]:
                platform_context["platform_stats"]["total_assignments"] = platform_context["user_data"]["total_assignments"]
            if access_context["role"] in ("admin", "system_manager"):
                platform_context["user_data"]["recent_submissions_count"] = user_data_context.get("recent_submissions_count", 0)
        except Exception as e_assign:
            current_app.logger.warning("Could not compute assignment stats: %s", e_assign)

    if identity.is_authenticated and identity.user:
        try:
            user_countries = identity.user.countries.all() if hasattr(identity.user, "countries") else []
            if access_context["role"] not in ("admin", "system_manager"):
                platform_context["available_countries"] = [
                    {"id": c.id, "name": c.name, "iso3": getattr(c, "iso3", ""), "national_society": getattr(c, "national_society_name", "") or ""}
                    for c in user_countries
                ]
            else:
                from app.models import Country
                platform_context["available_countries"] = [
                    {"id": c.id, "name": c.name, "iso3": getattr(c, "iso3", ""), "national_society": getattr(c, "national_society_name", "") or ""}
                    for c in Country.query.limit(50).all()
                ]
        except Exception as e_countries:
            current_app.logger.warning("Could not populate available countries: %s", e_countries)

    return platform_context


def _get_platform_context_cached(identity, access_context: Dict[str, Any]) -> Dict[str, Any]:
    """Return platform context, reusing from g when same identity within the request."""
    cache_key = f"ai_platform_{identity.user.id if identity.is_authenticated and identity.user else 'anon'}"
    if getattr(g, "_ai_platform_context_key", None) == cache_key and getattr(g, "_ai_platform_context", None) is not None:
        return g._ai_platform_context
    ctx = _build_platform_context(identity, access_context)
    g._ai_platform_context_key = cache_key
    g._ai_platform_context = ctx
    return ctx


def _persist_stream_conversation(
    app,
    *,
    user_id: int,
    conversation_id: str,
    message: str,
    response_text: str,
    page_context: Optional[Dict[str, Any]],
    client_message_id: Optional[str],
    branch_from_edit: bool,
    client: str,
    llm_provider: str,
    model_name: Optional[str],
    function_calls_used: list,
    map_payload: Optional[Dict[str, Any]] = None,
    chart_payload: Optional[Dict[str, Any]] = None,
    table_payload: Optional[Dict[str, Any]] = None,
    trace_id: Optional[int] = None,
) -> None:
    """
    Persist conversation + messages for streaming chats.

    IMPORTANT: This function is safe to call from a background worker thread
    (e.g., the SSE engine worker). It does NOT spawn another thread.
    """
    with app.app_context():
        try:
            convo = AIConversation.query.filter_by(id=conversation_id, user_id=user_id).first()
            if not convo:
                title = _build_initial_conversation_title(message)
                client_safe = client.strip().lower() if isinstance(client, str) else "unknown"
                if client_safe not in {"website", "mobile", "backoffice", "unknown"}:
                    client_safe = "unknown"
                convo = AIConversation(
                    id=conversation_id,
                    user_id=user_id,
                    title=title,
                    created_at=utcnow(),
                    updated_at=utcnow(),
                    last_message_at=utcnow(),
                    meta={"client": client_safe, "title_source": "auto_initial"},
                )
                db.session.add(convo)
            else:
                convo.updated_at = utcnow()
                convo.last_message_at = utcnow()

            # Insert user message only when we can be reasonably confident it doesn't already exist.
            # In the SSE flow, we usually insert the user message at stream start; this is a fallback.
            should_insert_user_msg = not branch_from_edit
            if should_insert_user_msg and client_message_id:
                existing = AIMessage.query.filter_by(
                    conversation_id=conversation_id,
                    user_id=user_id,
                    role="user",
                    client_message_id=client_message_id,
                ).first()
                if existing:
                    should_insert_user_msg = False
            else:
                # Without a client_message_id, avoid duplicating the early-inserted user row.
                should_insert_user_msg = False

            if should_insert_user_msg:
                db.session.add(
                    AIMessage(
                        conversation_id=conversation_id,
                        user_id=user_id,
                        role="user",
                        content=message,
                        client_message_id=client_message_id,
                        meta={"page_context": page_context} if page_context else None,
                    )
                )

            assistant_meta = {
                "provider": llm_provider,
                "model": model_name,
                "function_calls": function_calls_used,
                **({"map_payload": map_payload} if map_payload else {}),
                **({"chart_payload": chart_payload} if chart_payload else {}),
                **({"table_payload": table_payload} if table_payload else {}),
                **({"in_reply_to_client_message_id": client_message_id} if client_message_id else {}),
            }
            if trace_id is not None:
                assistant_meta["trace_id"] = trace_id
            db.session.add(
                AIMessage(
                    conversation_id=conversation_id,
                    user_id=user_id,
                    role="assistant",
                    content=response_text or "",
                    meta=assistant_meta,
                )
            )
            _maybe_refine_conversation_title(
                convo,
                user_message=message,
                assistant_response=response_text,
            )
            db.session.commit()
        except Exception as e:
            current_app.logger.exception(
                "Failed to persist AI conversation (stream) user_id=%s conversation_id=%s", user_id, conversation_id
            )
            try:
                db.session.rollback()
            except Exception as rb_e:
                logger.debug("Session rollback failed (non-fatal): %s", rb_e)
        finally:
            # Ensure thread-local sessions are cleaned up (no request teardown in background workers).
            with suppress(Exception):
                db.session.remove()


# Conversation DB helpers re-exported for ai_ws and _persist_stream_conversation
_load_conversation_history_for_llm = load_conversation_history_for_llm
_replace_conversation_messages = replace_conversation_messages
_find_existing_reply_for_client_message_id = find_existing_reply_for_client_message_id


@ai_bp.route("/token", methods=["GET"])
@login_required
def issue_token():
    """
    Issue a short-lived JWT that Website/Mobile can use as a Bearer token.
    Conversation persistence is for logged-in users only, so tokens are required for that.
    """
    try:
        from app.services.app_settings_service import user_has_ai_beta_access

        if not user_has_ai_beta_access(current_user):
            return json_forbidden("AI beta access is limited to selected users.")
        from app.services.authorization_service import AuthorizationService
        token_role = "system_manager" if AuthorizationService.is_system_manager(current_user) else ("admin" if AuthorizationService.is_admin(current_user) else "user")
        token = issue_ai_token(user_id=int(current_user.id), role=token_role)
        current_app.logger.debug(f"Issued AI token for user {current_user.id}")
        return json_ok(token=token)
    except Exception as e:
        current_app.logger.error(f"Failed to issue AI token: {e}", exc_info=True)
        return json_server_error("Failed to issue token")


@ai_bp.route("/health", methods=["GET"])
def ai_health():
    """
    Health check for AI stack: config, agent availability, and optional embedding probe.
    Use for load balancers or monitoring. No auth required.
    Query: ?probe=embedding to run a minimal embedding call (slower).
    """
    status = {"ok": True, "checks": {}}
    # Config checks
    status["checks"]["openai_key"] = bool(current_app.config.get("OPENAI_API_KEY"))
    status["checks"]["embedding_provider"] = "openai"
    if not status["checks"]["openai_key"]:
        status["ok"] = False

    # Agent availability (lazy init; best-effort)
    try:
        integration = _get_ai_chat_integration()
        status["checks"]["agent_available"] = integration is not None and getattr(integration, "agent_enabled", False) and getattr(integration, "agent", None) is not None
    except Exception as e:
        logger.debug("AI health agent check failed: %s", e, exc_info=True)
        status["checks"]["agent_available"] = False

    # Optional embedding probe
    if request.args.get("probe") == "embedding":
        try:
            from app.services.ai_embedding_service import AIEmbeddingService
            svc = AIEmbeddingService()
            _, _ = svc.generate_embedding("health")
            status["checks"]["embedding_probe"] = "ok"
        except Exception as e:
            status["checks"]["embedding_probe"] = str(e)[:200]
            status["ok"] = False
    code = 200 if status["ok"] else 503
    return jsonify(status), code


@ai_bp.route("/chat", methods=["POST"])
@limiter.limit(_ai_chat_daily_system_limit, key_func=_ai_chat_system_rate_limit_key)
@limiter.limit(_ai_chat_daily_user_limit, key_func=_ai_chat_rate_limit_key)
@limiter.limit(_ai_chat_limit, key_func=_ai_chat_rate_limit_key, override_defaults=True)
def chat():
    """
    V2 chat endpoint usable by Backoffice (cookie), Website (Bearer), Mobile (Bearer).
    Phase 1: non-streaming only.
    """
    identity = resolve_ai_identity()
    beta_denied = _ai_beta_denied_response(identity)
    if beta_denied is not None:
        return beta_denied
    # Anonymous access is only allowed via the Website proxy (shared-secret header).
    if not identity.is_authenticated and not _is_allowed_public_proxy_request():
        return json_auth_required("Authentication required")
    # Ensure current_user reflects Bearer auth so existing RBAC helpers (used by tools) work.
    did_login = False
    try:
        if identity.user and identity.auth_source == "bearer" and not current_user.is_authenticated:
            login_user(identity.user, remember=False)
            did_login = True
            # Avoid emitting a session cookie for bearer-token clients.
            try:
                session.modified = False
            except Exception as sess_e:
                logger.debug("session.modified = False failed: %s", sess_e)
    except Exception as e:
        # If login_user fails, proceed; RBAC in tool calls may deny access.
        logger.debug("login_user failed for bearer client: %s", e, exc_info=True)
    data = get_json_safe()

    parsed, err_msg, err_code = parse_chat_request(data)
    if err_msg:
        return json_error(err_msg, err_code)

    # Debug logging (opt-in): helps diagnose conversation awareness + idempotency issues.
    try:
        dbg = current_app.config.get("AI_CHAT_DEBUG_LOGS", None)
        if dbg is None:
            dbg = (os.getenv("AI_CHAT_DEBUG_LOGS") or "").strip().lower() in {"1", "true", "yes", "y", "on"}
        if bool(dbg):
            current_app.logger.debug(
                "AI v2 /chat request: user=%s auth=%s conv_id_in=%s client_msg_id=%s msg_preview=%r",
                (int(identity.user.id) if identity.is_authenticated and identity.user else None),
                str(identity.auth_source),
                (parsed.conversation_id or None),
                (parsed.client_message_id or None),
                (parsed.message or "")[:120],
            )
    except Exception as e:
        logger.debug("AI_CHAT_DEBUG_LOGS check failed: %s", e)

    # Optional: user-selected source gating for tools (databank vs system docs vs UPR docs).
    # Stored on request-scoped g so tool registry can enforce/filter deterministically.
    try:
        g.ai_sources_cfg = getattr(parsed, "sources_cfg", None)
    except Exception as e:
        logger.debug("g.ai_sources_cfg assignment failed: %s", e)

    allow_sensitive = bool(data.get("allow_sensitive"))

    # DLP: block/confirm before sending anything to external providers.
    allowed, dlp_err, dlp_findings = evaluate_ai_message(
        message=parsed.message,
        allow_sensitive=allow_sensitive,
    )
    if dlp_findings:
        log_dlp_audit_event(
            user_id=(int(identity.user.id) if identity.is_authenticated and identity.user else None),
            action=("blocked" if (not allowed and (dlp_err or {}).get("error_type") == "dlp_blocked") else
                    "confirm_required" if (not allowed) else
                    "send_anyway" if allow_sensitive else
                    "allowed"),
            transport="http",
            endpoint_path=(request.path or "/api/ai/v2/chat"),
            client=str((data.get("client") or "unknown")),
            conversation_id=(parsed.conversation_id or data.get("conversation_id")),
            client_message_id=parsed.client_message_id,
            allow_sensitive=allow_sensitive,
            findings=dlp_findings,
        )
    if not allowed and dlp_err:
        err_msg = dlp_err.get("error", "Request blocked") if isinstance(dlp_err, dict) else "Request blocked"
        extra = {k: v for k, v in (dlp_err or {}).items() if k != "error"} if isinstance(dlp_err, dict) else {}
        return json_error(err_msg, 400, **extra)

    if not identity.is_authenticated:
        parsed = apply_anonymous_rules(parsed)
        conversation_id = None
        conversation_history = parsed.conversation_history or []
    else:
        conversation_id, conversation_history = resolve_conversation_and_history(identity, parsed)

    access_context = _build_access_context(identity)

    existing = get_idempotent_reply(identity, conversation_id, parsed.client_message_id)
    if existing:
        reply_text, user_msg_id, assistant_msg_id = existing
        try:
            dbg = current_app.config.get("AI_CHAT_DEBUG_LOGS", None)
            if dbg is None:
                dbg = (os.getenv("AI_CHAT_DEBUG_LOGS") or "").strip().lower() in {"1", "true", "yes", "y", "on"}
            if bool(dbg):
                current_app.logger.debug(
                    "AI v2 /chat deduped reply: user=%s conv_id=%s client_msg_id=%s user_msg_id=%s assistant_msg_id=%s",
                    (int(identity.user.id) if identity.is_authenticated and identity.user else None),
                    conversation_id,
                    parsed.client_message_id,
                    user_msg_id,
                    assistant_msg_id,
                )
        except Exception as e:
            logger.debug("Dedup debug log failed: %s", e)
        return json_ok(
            success=True,
            conversation_id=conversation_id,
            reply=reply_text,
            meta={
                "provider": "deduped",
                "model": None,
                "response_time_ms": 0,
                "auth_source": identity.auth_source,
                "access_level": access_context["access_level"],
                "is_fallback": False,
                "deduped": True,
                "deduped_user_message_id": user_msg_id,
                "deduped_assistant_message_id": assistant_msg_id,
            },
        )

    message = parsed.message
    page_context = parsed.page_context
    preferred_language = parsed.preferred_language
    client_message_id = parsed.client_message_id
    branch_from_edit = parsed.branch_from_edit

    platform_context = _get_platform_context_cached(identity, access_context)
    # Pass conversation_id so agent traces and persistence can log it (do not mutate cached dict)
    run_platform_context = {**platform_context, "conversation_id": conversation_id} if conversation_id else platform_context

    # OpenAI-only provider chain (agent -> OpenAI). No other providers/fallbacks.
    start_time = utcnow()
    llm_provider = "openai"
    model_name: Optional[str] = None
    response_text: Optional[str] = None
    map_payload: Optional[Dict[str, Any]] = None
    chart_payload: Optional[Dict[str, Any]] = None
    table_payload: Optional[Dict[str, Any]] = None
    function_calls_used = []
    error_type: Optional[str] = None
    success = False

    try:
        # Centralized OpenAI-only chain (agent -> OpenAI)
        from app.services.ai_chat_engine import AIChatEngine

        engine = AIChatEngine()
        result = engine.run(
            message=message,
            platform_context=run_platform_context,
            page_context=page_context,
            preferred_language=preferred_language,
            conversation_history=conversation_history,
            enable_agent=True,
        )

        llm_provider = result.provider or "openai"
        model_name = result.model
        function_calls_used = result.function_calls_used or []
        response_text = result.response_html
        map_payload = getattr(result, "map_payload", None)
        chart_payload = getattr(result, "chart_payload", None)
        table_payload = getattr(result, "table_payload", None)
        success = bool(result.success)
        error_type = result.error_type if getattr(result, "error_type", None) else (None if success else "LLMError")
        trace_id_for_feedback = getattr(result, "trace_id", None) or getattr(g, "ai_trace_id", None)

        end_time = utcnow()
        response_time_ms = (end_time - start_time).total_seconds() * 1000

        # If OpenAI failed, return an error (no provider fallbacks).
        if not success:
            err_msg = (getattr(result, "error_message", None) or "AI is temporarily unavailable").strip()
            # Track telemetry for failures (best-effort), but don't persist conversation.
            try:
                telemetry_service = ChatbotTelemetryService()
                user_id = identity.user.id if identity.is_authenticated and identity.user else 0
                session_id = request.headers.get("X-Session-ID") or str(uuid.uuid4())
                input_tokens = result.input_tokens if getattr(result, "input_tokens", None) is not None else (len(message.split()) * 1.3)
                output_tokens = result.output_tokens if getattr(result, "output_tokens", None) is not None else 0
                estimated_cost = estimate_chat_cost(model_name or "", int(input_tokens or 0), int(output_tokens or 0)) if llm_provider == "openai" else 0.0
                telemetry_service.track_interaction(
                    ChatbotMetrics(
                        user_id=user_id,
                        session_id=session_id,
                        timestamp=utcnow(),
                        message_length=len(message),
                        language=preferred_language,
                        page_context=json.dumps(page_context) if page_context else None,
                        llm_provider=llm_provider,
                        model_name=model_name,
                        function_calls_made=function_calls_used,
                        response_time_ms=response_time_ms,
                        success=False,
                        error_type=error_type,
                        input_tokens=int(input_tokens) if input_tokens else None,
                        output_tokens=int(output_tokens) if output_tokens else None,
                        estimated_cost_usd=estimated_cost if estimated_cost > 0 else None,
                        response_length=0,
                        used_provenance=bool(function_calls_used),
                    )
                )
            except Exception as e:
                current_app.logger.warning("Failed to track telemetry for v2 chat failure: %s", e, exc_info=True)

            return json_error(
                err_msg,
                503,
                success=False,
                conversation_id=conversation_id,
                meta={
                    "provider": llm_provider,
                    "model": model_name,
                    "response_time_ms": response_time_ms,
                    "auth_source": identity.auth_source,
                    "access_level": access_context["access_level"],
                    "error_type": error_type,
                },
            )

        # Persist conversation + messages for logged-in users only
        if identity.is_authenticated and identity.user and conversation_id:
            try:
                convo = AIConversation.query.filter_by(id=conversation_id, user_id=identity.user.id).first()
                if not convo:
                    title = _build_initial_conversation_title(message)
                    client = (data.get("client") or "unknown")
                    if not isinstance(client, str):
                        client = "unknown"
                    client = client.strip().lower()
                    if client not in {"website", "mobile", "backoffice", "unknown"}:
                        client = "unknown"

                    convo = AIConversation(
                        id=conversation_id,
                        user_id=identity.user.id,
                        title=title,
                        created_at=utcnow(),
                        updated_at=utcnow(),
                        last_message_at=utcnow(),
                        meta={"client": client, "title_source": "auto_initial"},
                    )
                    db.session.add(convo)
                else:
                    convo.updated_at = utcnow()
                    convo.last_message_at = utcnow()

                # If branch_from_edit, we already replaced conversation with client history (incl. current user msg).
                # If client_message_id is present and the user message already exists, don't insert it again.
                should_insert_user_msg = not branch_from_edit
                if should_insert_user_msg and client_message_id:
                    existing_user = AIMessage.query.filter_by(
                        conversation_id=conversation_id,
                        user_id=identity.user.id,
                        role="user",
                        client_message_id=client_message_id,
                    ).first()
                    if existing_user:
                        should_insert_user_msg = False

                if should_insert_user_msg:
                    db.session.add(
                        AIMessage(
                            conversation_id=conversation_id,
                            user_id=identity.user.id,
                            role="user",
                            content=message,
                            client_message_id=client_message_id,
                            meta={"page_context": page_context} if page_context else None,
                        )
                    )
                db.session.add(
                    AIMessage(
                        conversation_id=conversation_id,
                        user_id=identity.user.id,
                        role="assistant",
                        content=response_text or "",
                        meta={
                            "provider": llm_provider,
                            "model": model_name,
                            "function_calls": function_calls_used,
                            **({"map_payload": map_payload} if map_payload else {}),
                            **({"chart_payload": chart_payload} if chart_payload else {}),
                            **({"table_payload": table_payload} if table_payload else {}),
                            **({"in_reply_to_client_message_id": client_message_id} if client_message_id else {}),
                        },
                    )
                )
                _maybe_refine_conversation_title(
                    convo,
                    user_message=message,
                    assistant_response=response_text,
                )
                db.session.commit()
            except Exception as e:
                current_app.logger.exception("Failed to persist AI conversation")
                with db.session.no_autoflush:
                    db.session.rollback()

        # Track telemetry
        try:
            telemetry_service = ChatbotTelemetryService()
            user_id = identity.user.id if identity.is_authenticated and identity.user else 0
            session_id = request.headers.get("X-Session-ID") or str(uuid.uuid4())

            # Use real token counts from result when available, else estimate
            input_tokens = result.input_tokens if getattr(result, "input_tokens", None) is not None else (len(message.split()) * 1.3)
            output_tokens = result.output_tokens if getattr(result, "output_tokens", None) is not None else (len((response_text or "").split()) * 1.3 if response_text else 0)

            # Estimate cost (centralized pricing)
            estimated_cost = estimate_chat_cost(model_name or "", int(input_tokens or 0), int(output_tokens or 0)) if llm_provider == "openai" else 0.0

            metrics = ChatbotMetrics(
                user_id=user_id,
                session_id=session_id,
                timestamp=utcnow(),
                message_length=len(message),
                language=preferred_language,
                page_context=json.dumps(page_context) if page_context else None,
                llm_provider=llm_provider,
                model_name=model_name,
                function_calls_made=function_calls_used,
                response_time_ms=response_time_ms,
                success=success,
                error_type=error_type,
                input_tokens=int(input_tokens) if input_tokens else None,
                output_tokens=int(output_tokens) if output_tokens else None,
                estimated_cost_usd=estimated_cost if estimated_cost > 0 else None,
                response_length=len(response_text or ""),
                used_provenance=bool(function_calls_used),
            )
            telemetry_service.track_interaction(metrics)
        except Exception as e:
            # Don't fail the request if telemetry fails
            current_app.logger.warning("Failed to track telemetry for v2 chat: %s", e, exc_info=True)

        # Log which provider was used for debugging
        current_app.logger.info(
            f"AI v2 chat response: provider={llm_provider}, model={model_name}, "
            f"response_length={len(response_text or '')}, user_id={identity.user.id if identity.user else 'anonymous'}"
        )

        return json_ok(
            success=True,
            conversation_id=conversation_id,
            reply=response_text,
            map_payload=map_payload,
            chart_payload=chart_payload,
            table_payload=table_payload,
            sources=getattr(result, "sources", None),
            trace_id=trace_id_for_feedback,
            meta={
                "provider": llm_provider,
                "model": model_name,
                "response_time_ms": response_time_ms,
                "auth_source": identity.auth_source,
                "access_level": access_context["access_level"],
                "deduped": False,
                "confidence": getattr(result, "confidence", None),
                "grounding_score": getattr(result, "grounding_score", None),
            },
        )

    except Exception as e:
        error_type = type(e).__name__
        error_message = GENERIC_ERROR_MESSAGE
        current_app.logger.exception("AI v2 chat failed: %s", error_message)

        # Track error in telemetry
        try:
            telemetry_service = ChatbotTelemetryService()
            user_id = identity.user.id if identity.is_authenticated and identity.user else 0
            session_id = request.headers.get("X-Session-ID") or str(uuid.uuid4())
            end_time = utcnow()
            response_time_ms = (end_time - start_time).total_seconds() * 1000

            metrics = ChatbotMetrics(
                user_id=user_id,
                session_id=session_id,
                timestamp=utcnow(),
                message_length=len(message),
                language=preferred_language,
                page_context=json.dumps(page_context) if page_context else None,
                llm_provider="error",
                model_name=None,
                function_calls_made=[],
                response_time_ms=response_time_ms,
                success=False,
                error_type=error_type,
                input_tokens=None,
                output_tokens=None,
                estimated_cost_usd=None,
                response_length=0,
                used_provenance=False,
            )
            telemetry_service.track_interaction(metrics)
        except Exception as e:
            current_app.logger.debug("AI telemetry tracking failed: %s", e, exc_info=True)

        # Build detailed error response
        error_response = {
            "success": False,
            "error": "Chat failed",
            "error_type": error_type,
            "message": error_message,
        }

        # Add additional context for common error types
        if "quota" in error_message.lower() or "rate limit" in error_message.lower() or "429" in error_message:
            error_response["error_type"] = "quota_exceeded"
            error_response["error"] = "API rate limit exceeded"
            error_response["message"] = "The AI service has reached its request limit. Please try again later."
        elif "timeout" in error_message.lower() or "TimeoutError" in error_type:
            error_response["error"] = "Request timeout"
            error_response["message"] = "The request took too long to process. Please try again."
        elif "connection" in error_message.lower() or "ConnectionError" in error_type:
            error_response["error"] = "Connection error"
            error_response["message"] = "Unable to connect to the AI service. Please check your connection and try again."
        elif "authentication" in error_message.lower() or "401" in error_message or "403" in error_message:
            error_response["error"] = "Authentication error"
            error_response["message"] = "Authentication failed. Please log in and try again."

        # SECURITY: Never expose tracebacks to clients, even in debug mode.
        # Log the traceback server-side only for debugging purposes.
        import traceback
        current_app.logger.error(f"AI chat error details: {traceback.format_exc()}")

        return json_server_error(
            error_response.get("message", "Chat failed"),
            success=False,
            error=error_response.get("error", "Chat failed"),
            error_type=error_response.get("error_type"),
            message=error_response.get("message"),
        )
    finally:
        if did_login:
            with suppress(Exception):
                logout_user()
            with suppress(Exception):
                session.modified = False


@ai_bp.route("/chat/stream", methods=["POST"])
@limiter.limit(_ai_chat_daily_system_limit, key_func=_ai_chat_system_rate_limit_key)
@limiter.limit(_ai_chat_daily_user_limit, key_func=_ai_chat_rate_limit_key)
@limiter.limit(_ai_chat_limit, key_func=_ai_chat_rate_limit_key, override_defaults=True)
def chat_stream():
    """
    V2 chat streaming endpoint (SSE).

    Emits events as `data: {json}\\n\\n` with types:
    - meta, step, delta, done, error
    """
    identity = resolve_ai_identity()
    beta_denied = _ai_beta_denied_response(identity)
    if beta_denied is not None:
        return beta_denied
    # Anonymous access is only allowed via the Website proxy (shared-secret header).
    if not identity.is_authenticated and not _is_allowed_public_proxy_request():
        return json_auth_required("Authentication required")
    did_login = False
    try:
        if identity.user and identity.auth_source == "bearer" and not current_user.is_authenticated:
            login_user(identity.user, remember=False)
            did_login = True
            # Avoid emitting a session cookie for bearer-token clients.
            try:
                session.modified = False
            except Exception as sess_e:
                logger.debug("session.modified = False failed (stream): %s", sess_e)
    except Exception as e:
        logger.debug("login_user failed for bearer client (stream): %s", e, exc_info=True)

    data = get_json_safe()
    parsed, err_msg, err_code = parse_chat_request(data)
    if err_msg:
        return json_error(err_msg, err_code)

    # Capture explicit sources selection for worker thread propagation.
    sources_cfg = getattr(parsed, "sources_cfg", None)

    try:
        g.ai_sources_cfg = sources_cfg
    except Exception as e:
        logger.debug("g.ai_sources_cfg assignment failed (stream): %s", e)

    allow_sensitive = bool(data.get("allow_sensitive"))

    # DLP: for streaming, reply with a single SSE error event (200 OK) so the client
    # can show a confirmation dialog and optionally resend with allow_sensitive/private mode.
    allowed, dlp_err, dlp_findings = evaluate_ai_message(
        message=parsed.message,
        allow_sensitive=allow_sensitive,
    )
    if dlp_findings:
        log_dlp_audit_event(
            user_id=(int(identity.user.id) if identity.is_authenticated and identity.user else None),
            action=("blocked" if (not allowed and (dlp_err or {}).get("error_type") == "dlp_blocked") else
                    "confirm_required" if (not allowed) else
                    "send_anyway" if allow_sensitive else
                    "allowed"),
            transport="sse",
            endpoint_path=(request.path or "/api/ai/v2/chat/stream"),
            client=str((data.get("client") or "unknown")),
            conversation_id=(parsed.conversation_id or data.get("conversation_id")),
            client_message_id=parsed.client_message_id,
            allow_sensitive=allow_sensitive,
            findings=dlp_findings,
        )
    if not allowed and dlp_err:
        request_id = str(uuid.uuid4())

        def _dlp_gen():
            yield f"data: {json.dumps({'type': 'meta', 'request_id': request_id, 'conversation_id': None})}\n\n"
            payload = dict(dlp_err)
            payload["type"] = "error"
            yield f"data: {json.dumps(payload)}\n\n"

        resp = Response(stream_with_context(_dlp_gen()), mimetype="text/event-stream")
        resp.headers["Cache-Control"] = "no-cache"
        resp.headers["X-Accel-Buffering"] = "no"
        return resp

    if not identity.is_authenticated:
        parsed = apply_anonymous_rules(parsed)
        conversation_id = None
        conversation_history = parsed.conversation_history or []
    else:
        conversation_id, conversation_history = resolve_conversation_and_history(identity, parsed)

    access_context = _build_access_context(identity)

    message = parsed.message
    page_context = parsed.page_context
    preferred_language = parsed.preferred_language
    client_message_id = parsed.client_message_id
    branch_from_edit = parsed.branch_from_edit

    keep_running_on_disconnect = bool(data.get("keep_running_on_disconnect"))

    # Ensure the conversation row exists immediately for new chats so the immersive UI can
    # refresh the sidebar list right away (stream persistence happens async later).
    if identity.is_authenticated and identity.user and conversation_id:
        try:
            convo = AIConversation.query.filter_by(id=conversation_id, user_id=int(identity.user.id)).first()
            if not convo:
                title = _build_initial_conversation_title(message)
                client = (data.get("client") or "unknown")
                if not isinstance(client, str):
                    client = "unknown"
                client = client.strip().lower()
                if client not in {"website", "mobile", "backoffice", "unknown"}:
                    client = "unknown"
                convo = AIConversation(
                    id=conversation_id,
                    user_id=int(identity.user.id),
                    title=title,
                    created_at=utcnow(),
                    updated_at=utcnow(),
                    last_message_at=utcnow(),
                    meta={"client": client, "source": "stream_init", "title_source": "auto_initial"},
                )
                db.session.add(convo)
                db.session.commit()
            else:
                # Touch updated_at so it sorts correctly in the sidebar even before messages persist.
                convo.updated_at = utcnow()
                db.session.commit()
        except Exception as e:
            current_app.logger.exception("Failed to pre-create AI conversation (stream init)")
            with suppress(Exception):
                db.session.rollback()

    existing = get_idempotent_reply(identity, conversation_id, client_message_id)
    if existing:
        reply_text, _, _ = existing
        request_id = str(uuid.uuid4())

        def _deduped_gen():
            yield f"data: {json.dumps({'type': 'meta', 'request_id': request_id, 'deduped': True, 'conversation_id': conversation_id})}\n\n"
            yield f"data: {json.dumps({'type': 'done', 'response': reply_text, 'provider': 'deduped', 'model': None})}\n\n"

        _deduped_resp = Response(stream_with_context(_deduped_gen()), mimetype="text/event-stream")
        _deduped_resp.headers["Cache-Control"] = "no-cache"
        _deduped_resp.headers["X-Accel-Buffering"] = "no"
        return _deduped_resp

    # Persist the USER message immediately for logged-in users (before long-running tool calls).
    # This is important for immersive UI so a refresh shows at least the submitted message.
    if identity.is_authenticated and identity.user and conversation_id and message and not branch_from_edit:
        try:
            convo = AIConversation.query.filter_by(id=conversation_id, user_id=int(identity.user.id)).first()
            if convo:
                convo.updated_at = utcnow()
                convo.last_message_at = utcnow()

            should_insert_user = True
            if client_message_id:
                existing_user = AIMessage.query.filter_by(
                    conversation_id=conversation_id,
                    user_id=int(identity.user.id),
                    role="user",
                    client_message_id=client_message_id,
                ).first()
                if existing_user:
                    should_insert_user = False

            if should_insert_user:
                db.session.add(
                    AIMessage(
                        conversation_id=conversation_id,
                        user_id=int(identity.user.id),
                        role="user",
                        content=message,
                        client_message_id=client_message_id,
                        meta={"page_context": page_context} if page_context else None,
                    )
                )
            # Commit even if we skipped insert so convo timestamp touches don't leak as pending work.
            db.session.commit()
        except Exception as e:
            current_app.logger.exception("Failed to persist user message at stream start")
            with suppress(Exception):
                db.session.rollback()

    platform_context = _get_platform_context_cached(identity, access_context)
    # Stream uses minimal country list; ensure keys exist for engine
    if "available_countries" not in platform_context:
        platform_context["available_countries"] = []

    request_id = str(uuid.uuid4())
    start_time = utcnow()
    cancelled = threading.Event()
    client_disconnected = threading.Event()
    q: "queue.Queue[Optional[Dict[str, Any]]]" = queue.Queue()
    # Accumulate deltas server-side so we can persist even if the engine
    # doesn't populate result.response_html (some streaming paths only emit deltas).
    delta_accum: list[str] = []

    def _emit(obj: Dict[str, Any]) -> None:
        try:
            if client_disconnected.is_set():
                return
            q.put(obj)
        except Exception as e:
            logger.debug("SSE _emit q.put failed: %s", e)

    app_obj = current_app._get_current_object()
    stream_debug = bool(
        current_app.config.get("AI_STREAM_DEBUG")
        or current_app.config.get("VERBOSE_FORM_DEBUG")
        or current_app.debug
    )
    user_id_for_persist: Optional[int] = None
    try:
        if identity.is_authenticated and identity.user:
            user_id_for_persist = int(identity.user.id)
    except Exception as e:
        logger.debug("Failed to get user_id_for_persist: %s", e)
        user_id_for_persist = None

    # Persist an "in-flight" progress snapshot on the conversation so immersive UI can
    # restore steps/progress after a refresh while generation continues server-side.
    if user_id_for_persist and conversation_id:
        try:
            convo = AIConversation.query.filter_by(id=conversation_id, user_id=user_id_for_persist).first()
            if convo:
                meta = convo.meta if isinstance(convo.meta, dict) else {}
                meta = dict(meta)
                meta["inflight"] = {
                    "request_id": request_id,
                    "status": "in_progress",
                    "started_at": utcnow().isoformat(),
                    "updated_at": utcnow().isoformat(),
                    "steps": [{"message": "Preparing query…", "detail_lines": []}],
                }
                convo.meta = meta
                convo.updated_at = utcnow()
                db.session.commit()
        except Exception as e:
            current_app.logger.exception("AI SSE[%s] failed to persist inflight snapshot", request_id)
            with suppress(Exception):
                db.session.rollback()

    current_app.logger.info(
        "AI SSE[%s] start user_id=%s conversation_id=%s client_message_id=%s",
        request_id,
        user_id_for_persist,
        conversation_id,
        client_message_id,
    )

    # Register this stream so /chat/cancel can signal it.
    with _ACTIVE_SSE_CANCEL_LOCK:
        _ACTIVE_SSE_CANCEL_EVENTS[request_id] = cancelled

    inflight_last_persist = 0.0
    inflight_last_steps_json = None  # debug / change detection (optional)

    def _persist_inflight_update(*, kind: str, step_message: Optional[str] = None, detail: Optional[str] = None, force: bool = False) -> None:
        """
        Update conversation.meta['inflight'] with the newest step/detail.
        Throttles detail writes to reduce DB churn.
        """
        nonlocal inflight_last_persist, inflight_last_steps_json
        if not user_id_for_persist or not conversation_id:
            return
        now_mono = time.monotonic()
        if not force and kind == "detail" and (now_mono - inflight_last_persist) < 1.5:
            return
        inflight_last_persist = now_mono

        try:
            convo = AIConversation.query.filter_by(id=conversation_id, user_id=user_id_for_persist).first()
            if not convo:
                return
            # json.loads(json.dumps(...)) produces a true deep copy so in-place
            # mutations of nested dicts don't alias the original convo.meta.
            # Without this, SQLAlchemy's JSON column comparator sees the old and
            # new values as equal (same nested references) and skips the UPDATE.
            meta = json.loads(json.dumps(convo.meta)) if isinstance(convo.meta, dict) else {}
            inflight = meta.get("inflight")
            if not isinstance(inflight, dict) or inflight.get("request_id") != request_id:
                return

            steps = inflight.get("steps")
            if not isinstance(steps, list):
                steps = []

            def _coalesce_progress_tick(lines: list, new_line: str) -> None:
                import re
                progress_re = re.compile(r"^(.+?):\s*(\d+)\s*/\s*(\d+)\s*$")
                if not lines:
                    lines.append(new_line)
                    return
                m_new = progress_re.match(new_line.strip())
                m_last = progress_re.match(str(lines[-1] or "").strip())
                if m_new and m_last:
                    if (m_new.group(1) or "").strip() == (m_last.group(1) or "").strip():
                        lines[-1] = new_line
                        return
                lines.append(new_line)

            if kind == "step":
                msg = (step_message or "").strip()
                if msg:
                    if steps and isinstance(steps[-1], dict) and str(steps[-1].get("message") or "").strip() == msg:
                        if detail and str(detail).strip():
                            dl = steps[-1].get("detail_lines")
                            if not isinstance(dl, list):
                                dl = []
                            _coalesce_progress_tick(dl, str(detail).strip())
                            steps[-1]["detail_lines"] = dl
                    else:
                        step_obj = {"message": msg, "detail_lines": []}
                        if detail and str(detail).strip():
                            step_obj["detail_lines"] = [str(detail).strip()]
                        steps.append(step_obj)
            elif kind == "detail":
                d = (detail or "").strip()
                if d:
                    if not steps:
                        steps.append({"message": "Preparing query…", "detail_lines": []})
                    if not isinstance(steps[-1], dict):
                        steps[-1] = {"message": str(steps[-1]), "detail_lines": []}
                    dl = steps[-1].get("detail_lines")
                    if not isinstance(dl, list):
                        dl = []
                    _coalesce_progress_tick(dl, d)
                    steps[-1]["detail_lines"] = dl

            inflight["steps"] = steps
            inflight["status"] = "in_progress"
            inflight["updated_at"] = utcnow().isoformat()
            meta["inflight"] = inflight
            convo.meta = meta
            flag_modified(convo, "meta")
            convo.updated_at = utcnow()

            # Avoid redundant commits when nothing changed (best-effort).
            try:
                steps_json = json.dumps(steps, ensure_ascii=False)
            except Exception as e:
                logger.debug("steps json.dumps failed: %s", e)
                steps_json = None
            if steps_json is not None and inflight_last_steps_json == steps_json and not force:
                return
            inflight_last_steps_json = steps_json

            db.session.commit()
        except Exception as e:
            current_app.logger.exception("AI SSE[%s] failed to update inflight progress", request_id)
            with suppress(Exception):
                db.session.rollback()

    def _clear_inflight_snapshot(*, force: bool = False) -> None:
        if not user_id_for_persist or not conversation_id:
            return
        try:
            convo = AIConversation.query.filter_by(id=conversation_id, user_id=user_id_for_persist).first()
            if not convo:
                return
            meta = convo.meta if isinstance(convo.meta, dict) else {}
            meta = dict(meta)
            inflight = meta.get("inflight")
            if not isinstance(inflight, dict):
                return
            if inflight.get("request_id") != request_id and not force:
                return
            meta.pop("inflight", None)
            convo.meta = meta or None
            convo.updated_at = utcnow()
            db.session.commit()
        except Exception as e:
            current_app.logger.exception("AI SSE[%s] failed to clear inflight snapshot", request_id)
            with suppress(Exception):
                db.session.rollback()

    def _on_step(step_message: Optional[str] = None, detail: Optional[str] = None) -> None:
        if cancelled.is_set():
            return
        if step_message is None and detail is not None:
            logger.info("AI SSE[%s] emitting step_detail: %s", request_id, detail)
            _persist_inflight_update(kind="detail", detail=detail)
            _emit({"type": "step_detail", "detail": detail})
            return
        if step_message:
            logger.info("AI SSE[%s] emitting step event: %s", request_id, step_message)
        payload = {"type": "step", "message": step_message or ""}
        if detail is not None:
            payload["detail"] = detail
        _persist_inflight_update(kind="step", step_message=step_message, detail=detail, force=True)
        _emit(payload)

    delta_stats = {"count": 0, "chars": 0}
    def _on_delta(delta_html: str) -> None:
        if cancelled.is_set():
            return
        try:
            if delta_html:
                delta_accum.append(str(delta_html))
        except Exception as e:
            logger.debug("delta_accum.append failed: %s", e)
        if delta_html:
            try:
                delta_stats["count"] += 1
                delta_stats["chars"] += len(str(delta_html))
                # Only log delta progress when a client is still connected; avoid flooding logs for background runs.
                if (
                    stream_debug
                    and not client_disconnected.is_set()
                    and (delta_stats["count"] == 1 or delta_stats["count"] % 200 == 0)
                ):
                    current_app.logger.info(
                        "AI SSE[%s] delta count=%s chars=%s sample=%r",
                        request_id,
                        delta_stats["count"],
                        delta_stats["chars"],
                        str(delta_html)[:80],
                    )
            except Exception as e:
                logger.debug("delta_stats update failed: %s", e)
        _emit({"type": "delta", "text": delta_html})

    # Run engine in a background thread and feed events into the queue.
    _worker_spawn_t = time.time()

    def _worker() -> None:
        try:
            _worker_start_t = time.time()
            current_app_logger = None  # will be set once app context is available
            from app.services.ai_chat_engine import AIChatEngine

            with app_obj.app_context():
                current_app.logger.info(
                    "AI SSE[%s] worker thread started: spawn_to_start=%dms",
                    request_id, int((_worker_start_t - _worker_spawn_t) * 1000),
                )
                with app_obj.test_request_context():
                    try:
                        if identity.is_authenticated and identity.user:
                            login_user(identity.user, remember=False)
                            try:
                                session.modified = False
                            except Exception as sess_e:
                                logger.debug("session.modified failed in worker: %s", sess_e)
                    except Exception as e:
                        logger.debug("login_user failed in worker: %s", e)
                    # Propagate per-request source selection into this synthetic request context
                    # so tool gating/filtering remains deterministic during streaming.
                    try:
                        g.ai_sources_cfg = sources_cfg
                    except Exception as e:
                        logger.debug("g.ai_sources_cfg failed in worker: %s", e)
                    # Pass conversation_id so agent traces can log it (do not mutate cached platform_context)
                    run_platform_context = {**platform_context, "conversation_id": conversation_id} if conversation_id else platform_context
                    current_app.logger.info(
                        "AI SSE[%s] worker setup done, calling engine.run: setup=%dms total=%dms",
                        request_id,
                        int((time.time() - _worker_start_t) * 1000),
                        int((time.time() - _worker_spawn_t) * 1000),
                    )
                    engine = AIChatEngine()
                    result = engine.run(
                        message=message,
                        platform_context=run_platform_context,
                        page_context=page_context,
                        preferred_language=preferred_language,
                        conversation_history=conversation_history,
                        enable_agent=True,
                        on_step=_on_step,
                        on_delta=_on_delta,
                        cancelled=cancelled,
                    )
                    response_text = result.response_html or ("".join(delta_accum) if delta_accum else "")
                    map_payload_result = getattr(result, "map_payload", None)
                    chart_payload_result = getattr(result, "chart_payload", None)
                    table_payload_result = getattr(result, "table_payload", None)
                    # When agent returns only a map/chart/table (no text), we still want to persist the assistant message.
                    has_visual = bool(map_payload_result or chart_payload_result or table_payload_result)
                    persist_content = response_text or ""
                    if not persist_content and result.success and has_visual:
                        if map_payload_result and chart_payload_result:
                            persist_content = "[Map and chart visualization]"
                        elif map_payload_result:
                            persist_content = "[Map visualization]"
                        elif table_payload_result:
                            persist_content = "[Data table]"
                        else:
                            persist_content = "[Chart visualization]"
                    if stream_debug:
                        current_app.logger.info(
                            "AI SSE[%s] worker engine done success=%s provider=%s model=%s response_len=%s delta_count=%s delta_chars=%s",
                            request_id,
                            bool(result.success),
                            result.provider,
                            result.model,
                            len(response_text or ""),
                            delta_stats.get("count"),
                            delta_stats.get("chars"),
                        )

                    trace_id_for_feedback = getattr(result, "trace_id", None) or getattr(g, "ai_trace_id", None)

                    # Emit _final BEFORE persistence so the UI can show the answer ASAP.
                    _emit(
                        {
                            "type": "_final",
                            "provider": result.provider,
                            "model": result.model,
                            "function_calls_used": result.function_calls_used,
                            "response": response_text or "",
                            "success": result.success,
                            "input_tokens": getattr(result, "input_tokens", None),
                            "output_tokens": getattr(result, "output_tokens", None),
                            "map_payload": map_payload_result,
                            "chart_payload": chart_payload_result,
                            "table_payload": table_payload_result,
                            "trace_id": trace_id_for_feedback,
                            "confidence": getattr(result, "confidence", None),
                            "grounding_score": getattr(result, "grounding_score", None),
                        }
                    )

                    # Persist in this worker thread so it still happens even if the client disconnects.
                    # Persist when we have response text OR a successful result with map/chart (agent map-only case).
                    try:
                        if (
                            user_id_for_persist
                            and conversation_id
                            and (response_text or (result.success and has_visual))
                            and identity.is_authenticated
                            and identity.user
                        ):
                            if stream_debug:
                                current_app.logger.info(
                                    "AI SSE[%s] persisting conversation_id=%s response_len=%s",
                                    request_id,
                                    conversation_id,
                                    len(persist_content or ""),
                                )
                            _persist_stream_conversation(
                                app_obj,
                                user_id=user_id_for_persist,
                                conversation_id=conversation_id,
                                message=message,
                                response_text=persist_content,
                                page_context=page_context,
                                client_message_id=client_message_id,
                                branch_from_edit=branch_from_edit,
                                client=(data.get("client") or "unknown"),
                                llm_provider=(result.provider or "openai"),
                                model_name=result.model,
                                function_calls_used=result.function_calls_used or [],
                                map_payload=map_payload_result,
                                chart_payload=chart_payload_result,
                                table_payload=table_payload_result,
                                trace_id=trace_id_for_feedback,
                            )
                            if stream_debug:
                                current_app.logger.info("AI SSE[%s] persisted OK conversation_id=%s", request_id, conversation_id)
                    except Exception as e:
                        current_app.logger.exception(
                            "AI SSE[%s] persistence failed user_id=%s conversation_id=%s",
                            request_id,
                            user_id_for_persist,
                            conversation_id,
                        )
                    finally:
                        # Always clear the inflight snapshot when the worker finishes (success or failure)
                        # so refreshes don't show stale "in progress" steps.
                        _clear_inflight_snapshot()
        except Exception as e:
            logger.exception("AI chat stream worker failed: %s", e)
            _emit({"type": "error", "message": GENERIC_ERROR_MESSAGE, "error_type": "Error"})
            with suppress(Exception):
                _clear_inflight_snapshot()
        finally:
            with suppress(Exception):
                with _ACTIVE_SSE_CANCEL_LOCK:
                    _ACTIVE_SSE_CANCEL_EVENTS.pop(request_id, None)
            # Background thread: ensure scoped session is cleaned up.
            with suppress(Exception):
                db.session.remove()
            q.put(None)

    threading.Thread(target=_worker, daemon=True).start()

    def _sse_gen():
        started_at = time.time()
        last_event_at = time.time()
        last_heartbeat_at = 0.0
        heartbeat_every_sec = float(current_app.config.get("AI_SSE_HEARTBEAT_SECONDS", 10))
        # Hard timeout: avoid cancelling healthy long generations too aggressively.
        # Default is intentionally larger than AI_HTTP_TIMEOUT_SECONDS because agent flows can
        # involve multiple tool calls + LLM turns and may have long single calls.
        try:
            ai_http_timeout = int(current_app.config.get("AI_HTTP_TIMEOUT_SECONDS", 60))
        except Exception as e:
            logger.debug("AI_HTTP_TIMEOUT_SECONDS parse failed: %s", e)
            ai_http_timeout = 60
        try:
            ai_agent_timeout = int(current_app.config.get("AI_AGENT_TIMEOUT_SECONDS", ai_http_timeout))
        except Exception as e:
            logger.debug("AI_AGENT_TIMEOUT_SECONDS parse failed: %s", e)
            ai_agent_timeout = ai_http_timeout
        hard_timeout_sec = float(
            current_app.config.get(
                "AI_SSE_IDLE_TIMEOUT_SECONDS",
                180,
            )
        )
        # Ensure the SSE stream lives at least as long as the agent timeout
        # (plus a buffer) so agent runs aren't cut short by the stream timeout.
        hard_timeout_sec = max(hard_timeout_sec, float(ai_agent_timeout) + 30.0)
        try:
            yield f"data: {json.dumps({'type': 'meta', 'request_id': request_id, 'conversation_id': conversation_id})}\n\n"
            while True:
                try:
                    item = q.get(timeout=2.0)
                except queue.Empty:
                    now = time.time()
                    elapsed = now - started_at
                    if elapsed >= hard_timeout_sec:
                        if keep_running_on_disconnect:
                            # Stream has run too long; close the SSE connection but let
                            # the worker continue in the background.  The worker will
                            # persist the result and clear inflight; the frontend
                            # recovers via polling.
                            client_disconnected.set()
                            current_app.logger.info(
                                "AI SSE[%s] stream timeout after %.1fs; detaching "
                                "(worker continues in background, conversation_id=%s)",
                                request_id,
                                elapsed,
                                conversation_id,
                            )
                        else:
                            cancelled.set()
                            current_app.logger.warning(
                                "AI SSE[%s] hard timeout after %.1fs without completion",
                                request_id,
                                elapsed,
                            )
                            yield f"data: {json.dumps({'type': 'error', 'message': 'The request timed out while generating a response. Please try again.', 'error_type': 'stream_idle_timeout'})}\n\n"
                        break
                    if (now - last_heartbeat_at) >= heartbeat_every_sec:
                        last_heartbeat_at = now
                        # Count heartbeats as stream activity so the connection does not get cut
                        # while the worker is doing a long LLM call (which emits no intermediate events).
                        # Use type 'heartbeat' so the client does not display any text (spinner/steps suffice).
                        last_event_at = now
                        yield f"data: {json.dumps({'type': 'heartbeat'})}\n\n"
                    continue

                last_event_at = time.time()
                if item is None:
                    if stream_debug:
                        current_app.logger.info("AI SSE[%s] queue closed before final (no done emitted)", request_id)
                    break

                if item.get("type") == "_final":
                    end_time = utcnow()
                    response_time_ms = (end_time - start_time).total_seconds() * 1000
                    response_text = item.get("response") or ""
                    llm_provider = item.get("provider") or "openai"
                    model_name = item.get("model")
                    function_calls_used = item.get("function_calls_used") or []
                    success = bool(item.get("success", True))
                    error_type = None

                    # Track telemetry: use real token counts from engine when available, else estimate
                    input_tokens = item.get("input_tokens")
                    output_tokens = item.get("output_tokens")
                    if input_tokens is None:
                        input_tokens = len((message or "").split()) * 1.3
                    if output_tokens is None:
                        output_tokens = len((response_text or "").split()) * 1.3 if response_text else 0.0

                    try:
                        telemetry_service = ChatbotTelemetryService()
                        user_id = identity.user.id if identity.is_authenticated and identity.user else 0
                        session_id = request.headers.get("X-Session-ID") or str(uuid.uuid4())

                        estimated_cost = estimate_chat_cost(model_name or "", int(input_tokens or 0), int(output_tokens or 0)) if llm_provider == "openai" else 0.0

                        metrics = ChatbotMetrics(
                            user_id=user_id,
                            session_id=session_id,
                            timestamp=utcnow(),
                            message_length=len(message),
                            language=preferred_language,
                            page_context=json.dumps(page_context) if page_context else None,
                            llm_provider=llm_provider,
                            model_name=model_name,
                            function_calls_made=function_calls_used,
                            response_time_ms=response_time_ms,
                            success=success,
                            error_type=error_type,
                            input_tokens=int(input_tokens) if input_tokens else None,
                            output_tokens=int(output_tokens) if output_tokens else None,
                            estimated_cost_usd=estimated_cost if estimated_cost > 0 else None,
                            response_length=len(response_text or ""),
                            used_provenance=bool(function_calls_used),
                        )
                        telemetry_service.track_interaction(metrics)
                    except Exception as e:
                        current_app.logger.warning("Failed to track telemetry for v2 chat stream: %s", e, exc_info=True)

                    current_app.logger.info(
                        "AI SSE[%s] emitting done conversation_id=%s response_len=%s provider=%s model=%s",
                        request_id,
                        conversation_id,
                        len(response_text or ""),
                        llm_provider,
                        model_name,
                    )
                    if item.get("map_payload") or item.get("chart_payload") or item.get("table_payload"):
                        yield f"data: {json.dumps({'type': 'structured', 'map_payload': item.get('map_payload'), 'chart_payload': item.get('chart_payload'), 'table_payload': item.get('table_payload')})}\n\n"
                    done_payload = {'type': 'done', 'response': response_text, 'provider': llm_provider, 'model': model_name}
                    if conversation_id:
                        done_payload['conversation_id'] = conversation_id
                    if item.get("trace_id") is not None:
                        done_payload["trace_id"] = item.get("trace_id")
                    if item.get("map_payload"):
                        done_payload["map_payload"] = item.get("map_payload")
                    if item.get("chart_payload"):
                        done_payload["chart_payload"] = item.get("chart_payload")
                    if item.get("table_payload"):
                        done_payload["table_payload"] = item.get("table_payload")
                    if item.get("confidence") is not None:
                        done_payload["confidence"] = item.get("confidence")
                    if item.get("grounding_score") is not None:
                        done_payload["grounding_score"] = item.get("grounding_score")
                    yield f"data: {json.dumps(done_payload)}\n\n"
                    break

                yield f"data: {json.dumps(item)}\n\n"
        except GeneratorExit:
            client_disconnected.set()
            if not keep_running_on_disconnect:
                cancelled.set()
            with suppress(Exception):
                current_app.logger.info(
                    "AI SSE[%s] generator exit keep_running=%s conversation_id=%s",
                    request_id,
                    keep_running_on_disconnect,
                    conversation_id,
                )
            raise
        finally:
            if did_login:
                with suppress(Exception):
                    logout_user()
                with suppress(Exception):
                    session.modified = False

    resp = Response(stream_with_context(_sse_gen()), mimetype="text/event-stream")
    resp.headers["Cache-Control"] = "no-cache"
    resp.headers["X-Accel-Buffering"] = "no"
    return resp


@ai_bp.route("/feedback", methods=["POST"])
@limiter.limit(_ai_feedback_limit)
def submit_feedback():
    """
    Record like/dislike for an AI response. Used in traces to improve quality.
    Body: { "trace_id": int, "rating": "like" | "dislike" }.
    Trace must belong to the current user (user_id match).
    """
    identity = resolve_ai_identity()
    beta_denied = _ai_beta_denied_response(identity)
    if beta_denied is not None:
        return beta_denied
    if not identity.is_authenticated or not identity.user:
        return json_auth_required("Authentication required")

    data = get_json_safe()
    trace_id = data.get("trace_id")
    rating = (data.get("rating") or "").strip().lower()

    if trace_id is None or (isinstance(trace_id, str) and not trace_id.isdigit()):
        return json_bad_request("trace_id required (integer)")
    trace_id = int(trace_id)
    if rating not in ("like", "dislike"):
        return json_bad_request("rating must be 'like' or 'dislike'")

    trace = db.session.get(AIReasoningTrace, trace_id)
    if not trace:
        return json_not_found("Trace not found")
    if trace.user_id != identity.user.id:
        return json_forbidden("Not allowed")

    trace.user_rating = rating
    try:
        db.session.commit()
    except Exception as e:
        current_app.logger.exception("Failed to save feedback for trace %s: %s", trace_id, e)
        db.session.rollback()
        return json_server_error("Failed to save")

    return json_ok(trace_id=trace_id, rating=rating)


@ai_bp.route("/conversations", methods=["GET"])
def list_conversations():
    identity = resolve_ai_identity()
    beta_denied = _ai_beta_denied_response(identity)
    if beta_denied is not None:
        return beta_denied
    if not identity.is_authenticated or not identity.user:
        return json_auth_required("Authentication required")

    limit = min(int(request.args.get("limit", 50)), 200)
    convos = (
        AIConversation.query.filter_by(user_id=identity.user.id)
        .order_by(AIConversation.last_message_at.desc().nullslast(), AIConversation.updated_at.desc())
        .limit(limit)
        .all()
    )

    def _inflight_summary(meta: Any) -> Optional[Dict[str, Any]]:
        try:
            m = _normalize_conversation_meta(meta)
            inflight = m.get("inflight")
            if not isinstance(inflight, dict):
                return None
            if str(inflight.get("status") or "") != "in_progress":
                return None
            summary: Dict[str, Any] = {
                "status": "in_progress",
                "request_id": inflight.get("request_id"),
                "started_at": inflight.get("started_at"),
                "updated_at": inflight.get("updated_at"),
            }
            steps = inflight.get("steps")
            if isinstance(steps, list):
                summary["steps"] = steps
            return summary
        except Exception as e:
            logger.debug("_inflight_summary failed: %s", e)
            return None

    return json_ok(
        success=True,
        conversations=[
            {
                "id": c.id,
                "title": c.title,
                "updated_at": c.updated_at.isoformat() if c.updated_at else None,
                "last_message_at": c.last_message_at.isoformat() if c.last_message_at else None,
                "is_archived": bool(getattr(c, "is_archived", False)),
                "archived_at": c.archived_at.isoformat() if getattr(c, "archived_at", None) else None,
                "inflight": _inflight_summary(getattr(c, "meta", None)),
            }
            for c in convos
        ],
    )


@ai_bp.route("/conversations/<conversation_id>/clear-inflight", methods=["POST"])
@limiter.limit(_ai_clear_inflight_limit)
def clear_conversation_inflight(conversation_id: str):
    """
    Clear in-progress (inflight) state for a conversation.
    Used when the client shows 'service unavailable' or after a failed stream
    so the sidebar and reload don't show a stale loading state.
    """
    identity = resolve_ai_identity()
    beta_denied = _ai_beta_denied_response(identity)
    if beta_denied is not None:
        return beta_denied
    if not identity.is_authenticated or not identity.user:
        return json_auth_required("Authentication required")

    convo = AIConversation.query.filter_by(id=conversation_id, user_id=identity.user.id).first()
    if not convo:
        return json_not_found("Not found")

    try:
        meta = dict(convo.meta) if isinstance(convo.meta, dict) else {}
        meta.pop("inflight", None)
        convo.meta = meta if meta else None
        convo.updated_at = utcnow()
        db.session.commit()
        return json_ok()
    except Exception as e:
        logger.exception("Clear inflight failed: %s", e)
        with suppress(Exception):
            db.session.rollback()
        return json_server_error("Failed to clear")


@ai_bp.route("/conversations/<conversation_id>", methods=["GET", "DELETE"])
def get_or_delete_conversation(conversation_id: str):
    identity = resolve_ai_identity()
    beta_denied = _ai_beta_denied_response(identity)
    if beta_denied is not None:
        return beta_denied
    if not identity.is_authenticated or not identity.user:
        return json_auth_required("Authentication required")

    convo = AIConversation.query.filter_by(id=conversation_id, user_id=identity.user.id).first()
    if not convo:
        return json_not_found("Not found")

    if request.method == "DELETE":
        # Best-effort delete archived object as well
        with suppress(Exception):
            delete_archive_object(convo)
        # Delete all messages first (CASCADE should handle this, but explicit is safer)
        AIMessage.query.filter_by(conversation_id=conversation_id, user_id=identity.user.id).delete()
        # Delete the conversation
        db.session.delete(convo)
        db.session.commit()
        return json_ok()

    # GET method
    limit = min(int(request.args.get("limit", 200)), 500)
    msgs = (
        AIMessage.query.filter_by(conversation_id=conversation_id, user_id=identity.user.id)
        .order_by(AIMessage.created_at.asc())
        .limit(limit)
        .all()
    )

    # If messages were archived, load them from archive storage.
    if not msgs and getattr(convo, "is_archived", False) and getattr(convo, "archive_path", None):
        try:
            payload = load_archived_conversation(convo)
            archived_msgs = (payload or {}).get("messages") or []
            archived_msgs = archived_msgs[:limit]
            return json_ok(
                success=True,
                conversation={
                    "id": convo.id,
                    "title": convo.title,
                    "updated_at": convo.updated_at.isoformat() if convo.updated_at else None,
                    "last_message_at": convo.last_message_at.isoformat() if convo.last_message_at else None,
                    "is_archived": True,
                    "archived_at": convo.archived_at.isoformat() if getattr(convo, "archived_at", None) else None,
                },
                messages=archived_msgs,
                meta={"source": "archive"},
            )
        except Exception as e:
            current_app.logger.error(f"Failed to load archived conversation {conversation_id}: {e}", exc_info=True)
            return json_server_error("Failed to load archived conversation")

    return json_ok(
        success=True,
        conversation={
            "id": convo.id,
            "title": convo.title,
            "updated_at": convo.updated_at.isoformat() if convo.updated_at else None,
            "last_message_at": convo.last_message_at.isoformat() if convo.last_message_at else None,
            "is_archived": bool(getattr(convo, "is_archived", False)),
            "archived_at": convo.archived_at.isoformat() if getattr(convo, "archived_at", None) else None,
            "meta": convo.meta or None,
        },
        messages=[
            {
                "id": m.id,
                "role": m.role,
                "content": m.content,
                "created_at": m.created_at.isoformat() if m.created_at else None,
                "client_message_id": m.client_message_id,
                "meta": m.meta or None,
            }
            for m in msgs
        ],
        meta={"source": "db"},
    )


@ai_bp.route("/conversations/<conversation_id>/export", methods=["GET"])
@limiter.limit(_ai_export_limit)
def export_conversation(conversation_id: str):
    """
    Export a conversation as a downloadable JSON file.
    Includes archived conversations transparently.
    """
    identity = resolve_ai_identity()
    beta_denied = _ai_beta_denied_response(identity)
    if beta_denied is not None:
        return beta_denied
    if not identity.is_authenticated or not identity.user:
        return json_auth_required("Authentication required")

    convo = AIConversation.query.filter_by(id=conversation_id, user_id=identity.user.id).first()
    if not convo:
        return json_not_found("Not found")

    max_export = int(current_app.config.get("AI_CHAT_EXPORT_MAX_MESSAGES", 5000))
    try:
        limit = min(int(request.args.get("limit", max_export)), max_export)
    except (TypeError, ValueError):
        limit = max_export
    limit = max(1, min(limit, max_export))

    payload: Dict[str, Any]
    msgs = (
        AIMessage.query.filter_by(conversation_id=conversation_id, user_id=identity.user.id)
        .order_by(AIMessage.created_at.asc())
        .limit(limit)
        .all()
    )
    if msgs:
        payload = {
            "conversation": {
                "id": convo.id,
                "user_id": int(convo.user_id),
                "title": convo.title,
                "created_at": convo.created_at.isoformat() if convo.created_at else None,
                "updated_at": convo.updated_at.isoformat() if convo.updated_at else None,
                "last_message_at": convo.last_message_at.isoformat() if convo.last_message_at else None,
                "is_archived": bool(getattr(convo, "is_archived", False)),
                "archived_at": convo.archived_at.isoformat() if getattr(convo, "archived_at", None) else None,
                "meta": convo.meta or None,
            },
            "messages": [
                {
                    "id": int(m.id) if m.id is not None else None,
                    "role": m.role,
                    "content": m.content,
                    "created_at": m.created_at.isoformat() if m.created_at else None,
                    "client_message_id": m.client_message_id,
                    "meta": m.meta or None,
                }
                for m in msgs
            ],
            "exported_at": utcnow().isoformat(),
            "format_version": 1,
            "source": "db",
        }
    else:
        # fallback to archived payload if present
        if getattr(convo, "is_archived", False) and getattr(convo, "archive_path", None):
            payload = load_archived_conversation(convo)
            payload["exported_at"] = utcnow().isoformat()
            payload["source"] = "archive"
            # apply export limit defensively
            payload["messages"] = (payload.get("messages") or [])[:limit]
        else:
            payload = {
                "conversation": {
                    "id": convo.id,
                    "user_id": int(convo.user_id),
                    "title": convo.title,
                },
                "messages": [],
                "exported_at": utcnow().isoformat(),
                "format_version": 1,
                "source": "empty",
            }

    data = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    max_bytes = int(current_app.config.get("AI_CHAT_EXPORT_MAX_BYTES", 10 * 1024 * 1024))
    if len(data) > max_bytes:
        return json_error("Export too large", 413)

    filename = f"ai-conversation-{conversation_id}.json"
    return send_file(
        io.BytesIO(data),  # type: ignore[name-defined]
        mimetype="application/json",
        as_attachment=True,
        download_name=filename,
    )


@ai_bp.route("/table/export", methods=["POST"])
@limiter.limit(_ai_table_export_limit)
def export_table_as_excel():
    """
    Export a small table (from the chat UI) as a real .xlsx file.

    Payload:
      { "rows": [["Header1","Header2"], ["a","b"], ...] }
    """
    identity = resolve_ai_identity()
    beta_denied = _ai_beta_denied_response(identity)
    if beta_denied is not None:
        return beta_denied
    if not identity.is_authenticated or not identity.user:
        return json_auth_required("Authentication required")

    # Ensure current_user reflects Bearer auth for consistency with the rest of ai_v2.
    try:
        if identity.user and identity.auth_source == "bearer" and not current_user.is_authenticated:
            login_user(identity.user, remember=False)
    except Exception as e:
        logger.debug("login_user failed for export_table: %s", e, exc_info=True)

    data = get_json_safe()
    rows = data.get("rows") or []
    if not isinstance(rows, list) or not rows:
        return json_bad_request("rows must be a non-empty list")

    # Hard limits to avoid abuse / excessive memory
    max_rows = int(current_app.config.get("AI_TABLE_EXPORT_MAX_ROWS", 2000))
    max_cols = int(current_app.config.get("AI_TABLE_EXPORT_MAX_COLS", 100))
    rows = rows[:max_rows]

    # Normalize and bound columns
    normalized_rows = []
    computed_max_cols = 0
    for r in rows:
        if not isinstance(r, list):
            continue
        rr = [("" if v is None else str(v)) for v in r[:max_cols]]
        computed_max_cols = max(computed_max_cols, len(rr))
        normalized_rows.append(rr)

    if not normalized_rows:
        return json_bad_request("No valid rows")

    # Create workbook
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except Exception as e:
        current_app.logger.error("Excel export requires openpyxl: %s", e)
        return json_server_error("Excel export not available")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    for r_idx, row in enumerate(normalized_rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Keep long text readable
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Bold header row (first row)
    try:
        for cell in ws[1]:
            cell.font = Font(bold=True)
    except Exception as e:
        logger.debug("Failed to apply header font: %s", e)

    # Basic column width auto-fit (bounded)
    try:
        for col_idx in range(1, min(computed_max_cols, max_cols) + 1):
            max_len = 0
            for row in normalized_rows[: min(len(normalized_rows), 200)]:  # sample first 200 rows
                if col_idx - 1 < len(row):
                    max_len = max(max_len, len(row[col_idx - 1]))
            # Cap width to keep file usable
            width = max(8, min(50, max_len + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    except Exception as e:
        logger.debug("Column width auto-fit failed: %s", e)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "table-data.xlsx"
    resp = send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
    # Reuse existing frontend conventions
    resp.headers["X-NGO-Databank-Export-Completed"] = "1"
    resp.headers["X-NGO-Databank-Export-Filename"] = filename
    return resp


@ai_bp.route("/conversations", methods=["DELETE"])
@limiter.limit(_ai_delete_all_conversations_limit)
def delete_all_conversations():
    """
    Delete ALL AI conversations (and archives) for the current authenticated user.
    Requires explicit confirmation: ?confirm=true or JSON body {"confirm": true}
    """
    identity = resolve_ai_identity()
    beta_denied = _ai_beta_denied_response(identity)
    if beta_denied is not None:
        return beta_denied
    if not identity.is_authenticated or not identity.user:
        return json_auth_required("Authentication required")

    confirm = (request.args.get("confirm") or "").strip().lower() == "true"
    body = get_json_safe()
    if isinstance(body, dict) and bool(body.get("confirm")):
        confirm = True
    if not confirm:
        return json_bad_request("Confirmation required (confirm=true)")

    convos = AIConversation.query.filter_by(user_id=identity.user.id).all()
    deleted = 0
    deleted_archives = 0
    for c in convos:
        with suppress(Exception):
            if delete_archive_object(c):
                deleted_archives += 1
        AIMessage.query.filter_by(conversation_id=c.id, user_id=identity.user.id).delete(synchronize_session=False)
        db.session.delete(c)
        deleted += 1
    db.session.commit()
    return json_ok(deleted_conversations=deleted, deleted_archives=deleted_archives)


@ai_bp.route("/conversations/<conversation_id>/messages", methods=["POST"])
@limiter.limit(_ai_append_message_limit)
def append_conversation_message(conversation_id: str):
    """
    Append a single message to a conversation (e.g. client-side error message so it persists on refresh).
    Payload: { "role": "assistant", "content": "...", "meta": { "is_error": true, "retry_message": "..." } }
    """
    identity = resolve_ai_identity()
    beta_denied = _ai_beta_denied_response(identity)
    if beta_denied is not None:
        return beta_denied
    if not identity.is_authenticated or not identity.user:
        return json_auth_required("Authentication required")

    convo = AIConversation.query.filter_by(id=conversation_id, user_id=identity.user.id).first()
    if not convo:
        return json_not_found("Not found")

    data = get_json_safe()
    role = str(data.get("role") or "assistant").strip().lower()
    content = str(data.get("content") or "").strip()
    meta = data.get("meta")
    if not content:
        return json_bad_request("content required")
    if role not in ("user", "assistant", "system"):
        return json_bad_request("role must be user, assistant, or system")

    meta_dict = dict(meta) if isinstance(meta, dict) else {}
    meta_dict["client"] = "backoffice"

    try:
        db.session.add(
            AIMessage(
                conversation_id=conversation_id,
                user_id=identity.user.id,
                role=role,
                content=content,
                meta=meta_dict,
            )
        )
        convo.updated_at = utcnow()
        convo.last_message_at = utcnow()
        db.session.commit()
        return json_ok()
    except Exception as e:
        current_app.logger.exception("Failed to append AI conversation message")
        with suppress(Exception):
            db.session.rollback()
        return json_server_error("Failed to save message")


@ai_bp.route("/conversations/<conversation_id>/import", methods=["POST"])
@limiter.limit(_ai_import_conversation_limit)
def import_conversation_messages(conversation_id: str):
    """
    Import client-side messages into a server conversation (logged-in only).

    This supports an offline-first mobile UX: users can chat while logged out, then
    import that history to the server after login so conversations are consistent across devices.

    Payload:
      { "messages": [{"client_message_id": "...", "role": "user|assistant|system", "content": "...", "created_at": "iso"}], "client":"mobile" }
    """
    identity = resolve_ai_identity()
    beta_denied = _ai_beta_denied_response(identity)
    if beta_denied is not None:
        return beta_denied
    if not identity.is_authenticated or not identity.user:
        return json_auth_required("Authentication required")

    data = get_json_safe()
    raw_messages = data.get("messages") or []
    client = (data.get("client") or "unknown").strip()

    if not isinstance(raw_messages, list):
        return json_bad_request("messages must be a list")

    # Hard limits to protect DB and LLM costs
    max_import_messages = 5000
    max_import_messages = max(1, max_import_messages)
    if len(raw_messages) > max_import_messages:
        return json_bad_request(f"Too many messages (max {max_import_messages})")

    convo = AIConversation.query.filter_by(id=conversation_id, user_id=identity.user.id).first()
    if not convo:
        # Best-effort title from the first user message
        title = None
        for m in raw_messages:
            try:
                if (m or {}).get("role") == "user" and (m or {}).get("content"):
                    txt = str(m.get("content")).strip()
                    if txt:
                        title = _build_initial_conversation_title(txt)
                        break
            except Exception as e:
                logger.debug("Import message title parse failed: %s", e)
                continue
        convo = AIConversation(
            id=conversation_id,
            user_id=identity.user.id,
            title=title,
            created_at=utcnow(),
            updated_at=utcnow(),
            last_message_at=utcnow(),
            meta={"client": client, "imported": True, "title_source": "auto_initial"},
        )
        db.session.add(convo)

    inserted = 0
    skipped = 0

    def _parse_dt(s: Optional[str]):
        if not s:
            return None
        try:
            # Support trailing Z
            ss = s.strip().replace("Z", "+00:00")
            return datetime.fromisoformat(ss)
        except Exception as e:
            logger.debug("_parse_dt failed for %r: %s", s, e)
            return None

    allowed_roles = {"user", "assistant", "system"}

    for m in raw_messages:
        if not isinstance(m, dict):
            skipped += 1
            continue
        role = str(m.get("role") or "").strip().lower()
        content = str(m.get("content") or "").strip()
        client_message_id = (m.get("client_message_id") or None)
        if isinstance(client_message_id, str):
            client_message_id = client_message_id.strip() or None
        else:
            client_message_id = None

        if role not in allowed_roles or not content:
            skipped += 1
            continue

        # Idempotency: skip if client_message_id already exists for this conversation/user
        if client_message_id:
            exists = AIMessage.query.filter_by(
                conversation_id=conversation_id,
                user_id=identity.user.id,
                client_message_id=client_message_id,
            ).first()
            if exists:
                skipped += 1
                continue

        created_at = _parse_dt(m.get("created_at")) or utcnow()

        db.session.add(
            AIMessage(
                conversation_id=conversation_id,
                user_id=identity.user.id,
                role=role,
                content=content,
                created_at=created_at,
                client_message_id=client_message_id,
                meta={"imported": True, "client": client},
            )
        )
        inserted += 1

    convo.updated_at = utcnow()
    convo.last_message_at = utcnow()

    try:
        db.session.commit()
    except Exception as e:
        current_app.logger.exception("Failed to import AI conversation messages")
        with db.session.no_autoflush:
            db.session.rollback()
        return json_server_error("Failed to import messages")

    return json_ok(inserted=inserted, skipped=skipped)
